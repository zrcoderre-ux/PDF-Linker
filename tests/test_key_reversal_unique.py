"""
Exactly ONE key row reverses each fake.

`_PnFakeRegistry` is injective — no two real values are ever handed the same
fake — so two key rows sharing a Replacement are never two different parties.
They are always alternate SPELLINGS of one real value, registered on purpose
against the canonical value's fake so that every spelling scrubs:

  * a wrap-split hyphenated surname, where a line break opened the hyphen and
    extraction left "Ardeshirpour- Zartoshti"; and
  * a `_pn_name_variants` near-miss ("Sarra" for "Sara"), so the exhibit's
    misspelling is caught too.

Right for the forward pass, unusable in reverse: `DeAnonymize.bas` cannot tell
a synthetic spelling from the real one, so it treats a fake claimed by two Real
Values as ambiguous and retires the mapping — fail-safe, but the pseudonym is
then left standing in the tentative. Registering those spellings made every
hyphenated party and every OCR-matched name un-restorable.

So the non-canonical rows are MARKED `alt spelling` in Status, not dropped.
Dropping them was the first attempt and it broke the other half of the
contract: a re-run with only the key to go on (no party template to rebuild the
variants from) then scrubbed the wrap-split spelling half by half
("Sedgwick- Linford"), and the documents already sent back no longer matched.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_reversal_unique.py -v
"""
import json
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
ALT = P._PN_KEY_ALT_STATUS

HYPHEN_PARTY = "Sara Ardeshirpour-Zartoshti"
CASENO = "24STCV23198"
# The wrap-split spelling ("Ardeshirpour- Zartoshti") is what a line break
# leaves behind in the export, so it is the spelling that actually MATCHES —
# the canonical one may never appear in the batch at all.
WRAPPED = ("Declaration of Sara Ardeshirpour- Zartoshti in 24STCV23198. "
           "Dr. Ardeshirpour- Zartoshti examined the plaintiff.")


def _key_rows(path):
    # Both sheets: a binding no export carries is written to
    # `_PN_KEY_PINNED_SHEET`, out of the reversal macro's reach, but it is still
    # a binding and the one-row-per-fake invariant covers it.
    wb = openpyxl.load_workbook(path)
    head = [str(h).strip().lower()
            for h in next(wb.active.iter_rows(max_row=1, values_only=True))]
    return [dict(zip(head, r)) for ws in wb.worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]


def _fake_for(rows, real):
    """The Replacement a row carries for `real`. Read back rather than named:
    which pool word a value draws moves whenever the pools are resized, and
    what these tests are about is WHICH ROW owns a fake, not which fake."""
    return next(str(r["replacement"]) for r in rows
                if str(r["real value"]).lower() == str(real).lower())


def _reversible(rows):
    """The rows the macro will use fake->real: everything but `alt spelling`."""
    return [r for r in rows if str(r["status"]).strip().lower() != ALT]


def _run(tmp_path, names, text):
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(names, [], [], registry=reg),
                        {}, registry=reg)
    out = z.apply(text)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    return z, out, _key_rows(path)


@pytest.fixture
def template(tmp_path):
    """The E-Court party template, so a re-run can supplement from it."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant",
               "Other Names"])
    ws.append([CASENO, HYPHEN_PARTY, "Acme Corp", ""])
    wb.save(tmp_path / "Order_Template_Input.xlsx")
    return tmp_path


# ── the invariant ───────────────────────────────────────────────────────────

def test_one_reversible_row_per_fake(tmp_path):
    _z, _out, rows = _run(tmp_path, [HYPHEN_PARTY], WRAPPED)
    fakes = [str(r["replacement"]).lower() for r in _reversible(rows)]
    assert len(fakes) == len(set(fakes)), (
        "two reversible rows share a fake, so DeAnonymize retires both as "
        f"ambiguous and leaves the pseudonym standing: {sorted(fakes)}")


def test_wrap_split_spelling_does_not_own_the_reversal(tmp_path):
    """The row that reverses the compound fake must carry the CANONICAL
    surname; the wrap-split spelling is kept, marked forward-only."""
    _z, out, rows = _run(tmp_path, [HYPHEN_PARTY], WRAPPED)
    compound = _fake_for(rows, "Ardeshirpour-Zartoshti")
    group = [r for r in rows
             if str(r["replacement"]).lower() == compound.lower()]
    owner = _reversible(group)
    assert len(owner) == 1
    assert owner[0]["real value"] == "Ardeshirpour-Zartoshti"
    assert [r["real value"] for r in group if r not in owner] == [
        "Ardeshirpour- Zartoshti"]
    # And the fake it reverses really is the one in the export.
    assert compound in out
    assert "-" in compound            # each half is that half's own fake
    assert "Ardeshirpour" not in out


def test_owner_status_answers_for_the_whole_group(tmp_path):
    """The owner reads `replaced` even though the hits came in under the other
    spelling: the fake DID reach the export, and this is the row that reverses
    it. Reading `no match` would mark it a binding nothing ever used."""
    _z, _out, rows = _run(tmp_path, [HYPHEN_PARTY], WRAPPED)
    compound = _fake_for(rows, "Ardeshirpour-Zartoshti")
    owner = next(r for r in _reversible(rows)
                 if str(r["replacement"]).lower() == compound.lower())
    assert owner["status"] == "replaced"


def test_ocr_near_miss_spelling_does_not_own_the_reversal(tmp_path):
    """Same rule for a `_pn_name_variants` misspelling: "Sarra" is registered
    against Sara's fake so the exhibit's typo scrubs, but the key must reverse
    "Atwater" to "Sara"."""
    assert "Sarra" in P._pn_name_variants("Sara")   # guards the fixture
    _z, out, rows = _run(tmp_path, [HYPHEN_PARTY],
                         "Declaration of Sarra Ardeshirpour-Zartoshti.")
    sara = _fake_for(rows, "Sara")
    given = [r for r in _reversible(rows)
             if str(r["replacement"]).lower() == sara.lower()]
    assert len(given) == 1
    assert given[0]["real value"] == "Sara"
    assert sara in out


def test_a_variant_owns_the_reversal_when_the_canonical_earned_no_row(tmp_path):
    """A group that is ALL synthetic still gets an owner — otherwise the fake
    is in the export with nothing to reverse it. (Contrived: the canonical
    normally earns its row by being an authoritative party.)"""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms([HYPHEN_PARTY], [], [], registry=reg)
    # Strip the canonical whole-token term AND the full-name row it would be
    # re-derived from (`_pn_name_token_rows` zips the two word lists), leaving
    # only the synthetic spellings that point at its fake.
    terms = [t for t in terms
             if t.real not in ("Ardeshirpour-Zartoshti", HYPHEN_PARTY)]
    z = P.Pseudonymizer(terms, {}, registry=reg)
    z.apply(WRAPPED)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    rows = _key_rows(path)
    compound = _fake_for(rows, "Ardeshirpour- Zartoshti")
    owner = [r for r in _reversible(rows)
             if str(r["replacement"]).lower() == compound.lower()]
    assert len(owner) == 1
    assert owner[0]["real value"] == "Ardeshirpour- Zartoshti"


# ── the other half: the re-run must still reproduce the delivered exports ───

def test_marked_rows_still_pin_the_wrap_split_spelling(template):
    """Byte-identity on a re-run that has ONLY the key — no party template to
    rebuild the synthetic spellings from. Dropping the rows instead of marking
    them scrubbed the wrap-split form half by half ("Sedgwick- Linford")."""
    reg = P._PnFakeRegistry()
    names, casenos = P._pn_terms_from_xlsx(
        template / "Order_Template_Input.xlsx", None, log)
    z = P.Pseudonymizer(P._pn_build_terms(names, casenos, [], registry=reg),
                        {}, registry=reg)
    sent = z.apply(WRAPPED)
    z.write_key(template / "pseudonym_key.xlsx", log)
    compound = _fake_for(_key_rows(template / "pseudonym_key.xlsx"),
                         "Ardeshirpour-Zartoshti")
    assert compound in sent and "-" in compound

    (template / "Order_Template_Input.xlsx").unlink()
    reg2 = P._PnFakeRegistry()
    terms, _ = P._pn_load_key(template / "pseudonym_key.xlsx", reg2, log)
    assert P.Pseudonymizer(terms, {}, registry=reg2).apply(WRAPPED) == sent


def test_the_marker_survives_a_rewrite(template):
    """`_pn_load_key` must carry `alt spelling` back onto the term, or the
    re-run's write_key re-picks the owner by hit count — which the wrap-split
    spelling wins, since it is the form the text actually carries — and
    ownership flips to the synthetic spelling on the first re-run."""
    reg = P._PnFakeRegistry()
    names, casenos = P._pn_terms_from_xlsx(
        template / "Order_Template_Input.xlsx", None, log)
    z = P.Pseudonymizer(P._pn_build_terms(names, casenos, [], registry=reg),
                        {}, registry=reg)
    sent = z.apply(WRAPPED)
    z.write_key(template / "pseudonym_key.xlsx", log)

    for _ in range(3):
        reg2 = P._PnFakeRegistry()
        terms, _kd = P._pn_load_key(template / "pseudonym_key.xlsx", reg2, log)
        terms += P._pn_supplement_key_terms(terms, template, None, reg2, log)
        z2 = P.Pseudonymizer(terms, {}, registry=reg2)
        assert z2.apply(WRAPPED) == sent
        z2.write_key(template / "pseudonym_key.xlsx", log)
        rows = _key_rows(template / "pseudonym_key.xlsx")
        compound = _fake_for(rows, "Ardeshirpour-Zartoshti")
        owner = [r for r in _reversible(rows)
                 if str(r["replacement"]).lower() == compound.lower()]
        assert len(owner) == 1
        assert owner[0]["real value"] == "Ardeshirpour-Zartoshti"


def test_json_fallback_key_carries_the_marker(tmp_path, monkeypatch):
    """The openpyxl-missing path writes the same rows; it must not be the one
    place an unmarked duplicate survives."""
    real_import = __import__

    def no_openpyxl(name, *a, **kw):
        if name == "openpyxl":
            raise ImportError("no openpyxl")
        return real_import(name, *a, **kw)

    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms([HYPHEN_PARTY], [], [], registry=reg),
                        {}, registry=reg)
    z.apply(WRAPPED)
    monkeypatch.setattr("builtins.__import__", no_openpyxl)
    z.write_key(tmp_path / "pseudonym_key.xlsx", log)
    monkeypatch.undo()
    maps = json.loads((tmp_path / "pseudonym_key.json").read_text())["mappings"]
    fakes = [m["replacement"].lower() for m in maps if m["status"] != ALT]
    assert len(fakes) == len(set(fakes))
    assert any(m["status"] == ALT for m in maps)
