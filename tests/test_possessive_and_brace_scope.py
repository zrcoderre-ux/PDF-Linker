"""
Two ways one party came back as two names.

  * A POSSESSIVE drew its own unrelated fake — "Rasho -> ARCLIGHT" beside
    "RASHO'S -> BALFOUR" — because the registry memoizes on the string it is
    handed and "rasho's" is not "rasho".
  * A `{braced}` keep typed on ONE key row had to reach the registry before ANY
    row was read back, or every other row kept applying its stored composed fake
    for the same word and the operator had to run the folder twice.

Run:  cd PDF-Linker && python3 -m pytest tests/test_possessive_and_brace_scope.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}


# ── the possessive ──────────────────────────────────────────────────────────

def test_a_possessive_takes_the_partys_own_fake():
    reg = P._PnFakeRegistry()
    base = P._pn_fake_name_token("Rasho", reg)
    assert P._pn_fake_name_token("Rasho's", reg) == f"{base}'s"
    assert P._pn_fake_name_token("RASHO'S", reg) == f"{base.upper()}'S"
    assert P._pn_fake_name_token("Ross'", reg) == P._pn_fake_name_token("Ross", reg) + "'"


def test_one_party_one_name_across_both_forms():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Rasho Doe", "RASHO'S"], [], [], registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    surname = people["Rasho Doe"].split()[0]
    assert people["RASHO'S"] == f"{surname.upper()}'S"


def test_the_text_reads_as_one_person():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Rasho Doe", "RASHO'S"], [], [], registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    out = pz.apply("Rasho Doe filed. RASHO'S opposition. Rasho's brief.")
    assert "Rasho" not in out and "RASHO" not in out
    # Casing follows the document, so compare the names themselves.
    # removesuffix, not rstrip: a fake that itself ends in "s" ("Strangeways")
    # loses its own last letter to a character-set strip.
    stems = {w.removesuffix("'s").removesuffix("'S").lower()
             for w in out.split() if w.endswith(("'s", "'S"))}
    assert len(stems) == 1
    assert stems.pop() == out.split()[0].lower()


def test_an_apostrophe_inside_a_name_is_not_a_possessive():
    """O'Brien is a surname, not the possessive of "O"."""
    reg = P._PnFakeRegistry()
    fake = P._pn_fake_name_token("O'Brien", reg)
    assert "'" not in fake
    assert fake != P._pn_fake_name_token("O", reg)


def test_the_token_map_draws_on_the_same_base():
    """`_pn_person_token_map` fed the raw token to the registry while keying the
    result by base, so a possessive asked for a second fake under its own name."""
    reg = P._PnFakeRegistry()
    m = P._pn_person_token_map("Rasho's Deposition", reg)
    assert m["rasho"] == P._pn_fake_name_token("Rasho", reg)


def test_a_divergent_possessive_row_is_repaired_on_load(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person-token", "Rasho", "ARCLIGHT", "replaced", "spreadsheet", 12])
    ws.append(["person", "RASHO'S", "BALFOUR", "replaced", "--term", 4])
    wb.save(key)
    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    byreal = {t.real: t.fake for t in terms}
    assert byreal["RASHO'S"] == "ARCLIGHT'S"
    assert byreal["Rasho"] == "ARCLIGHT"


def test_a_possessive_row_with_no_base_row_is_left_alone(tmp_path):
    """Nothing authoritative to fold onto — the delivered export still says it."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person", "RASHO'S", "BALFOUR", "replaced", "--term", 4])
    wb.save(key)
    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    assert terms[0].fake == "BALFOUR"


# ── one brace, every row ────────────────────────────────────────────────────

def _key(tmp_path, rows):
    path = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return path


def test_one_brace_cleans_every_row_of_the_key(tmp_path):
    key = _key(tmp_path, [
        ("entity", "Alder Law, P.C.", "{Law}", "replaced", "spreadsheet", 21),
        ("person", "Law Offices of Scott C. Stratman",
         "Braxton Mansffield bancroft Merrick C. Whitlock", "replaced",
         "spreadsheet", 4),
        ("entity", "Mitilian Law Group", "Sackett Silvergate Group", "replaced",
         "spreadsheet", 7),
        ("person-token", "Law", "Braxton", "replaced", "spreadsheet", 19),
    ])
    reg = P._PnFakeRegistry()
    terms, decisions = P._pn_load_key(key, reg, log)
    assert "law" in reg.keep_words
    byreal = {t.real: t.fake for t in terms}
    assert byreal["Law Offices of Scott C. Stratman"] \
        == "Law Offices of Merrick C. Whitlock"
    assert byreal["Mitilian Law Group"] == "Sackett Law Group"
    assert "Law" not in byreal                      # the bare token row is dead
    assert decisions["alder law, p.c."]["type"] == P._PN_KEEP_NUCLEAR_TYPE


def test_a_brace_naming_nothing_in_its_row_does_not_nuke_the_word(tmp_path,
                                                                  caplog):
    """It falls through as a literal replacement, which is a mistake worth
    saying out loud rather than silently applying a word-wide decision."""
    key = _key(tmp_path, [
        ("person", "Steven Burt", "{Law}", "replaced", "spreadsheet", 4),
    ])
    reg = P._PnFakeRegistry()
    with caplog.at_level(logging.WARNING):
        terms, _dec = P._pn_load_key(key, reg, log)
    assert "law" not in reg.keep_words
    assert terms[0].fake == "{Law}"
    assert "does not name part of that value" in caplog.text


# ── a run-together name keeps its connector ─────────────────────────────────

def test_a_domain_core_folds_onto_the_partys_own_fake():
    """A domain core is the party with its spaces gone. Folded on the bound
    prefix alone, the connector's tail drew one unrelated pool word, so
    "cadillacofcalabasas" came back "eldridge" — nothing tying it to the party
    or to the "cadiilac" beside it."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Cadillac of Calabasas"], [],
                              ["cadillacofcalabasas"], registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    party = people["Cadillac of Calabasas"]
    assert people["cadillacofcalabasas"] == party.replace(" ", "").lower()
    assert " of " in party            # the connector is kept in both forms


def test_the_plain_two_part_weld_still_folds():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Adler Michael"], [], ["ADLERMICHAEL"], registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    assert people["ADLERMICHAEL"] == people["Adler Michael"].replace(" ", "").upper()


def test_a_name_that_merely_contains_a_connector_is_not_split():
    """"smiththeodore" must not lose its "the" and fold onto a mangled "odore" —
    the connector is only stripped when what remains is itself already bound."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Smith Theodore"], [], ["smiththeodore"],
                              registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    assert people["smiththeodore"] == people["Smith Theodore"].replace(" ", "").lower()


def test_an_ocr_typo_still_folds_onto_a_typo_of_the_same_fake():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Cadillac of Calabasas"], [], ["cadiilac"],
                              registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    bound = people["Cadillac of Calabasas"].split()[0]
    typo = people["cadiilac"]
    assert typo.lower() != bound.lower()                    # still distinct
    assert P._pn_osa_distance(typo.lower(), bound.lower()) <= 2


# ── the TYPOGRAPHIC apostrophe ──────────────────────────────────────────────
# Everything above is written with the straight `'`. A filing written in Word
# carries `’`, and the tool treated the two as unrelated characters — which
# broke the possessive three separate ways.

CURLY = "’"


def test_a_curly_possessive_is_not_an_initial():
    """The reported failure: "RACHEL GREEN’S" -> "RIDLEY YEARDLEY’H".

    `_PN_WORD_RE` kept a straight apostrophe inside a word but not a curly one,
    so "GREEN’S" read as "GREEN" plus a one-letter word "S". A single letter is
    kept verbatim by `_pn_fake_person`, which is exactly what a middle INITIAL
    looks like — so `_pn_align_initials` gave the possessive the first letter of
    the fake the middle name got. A possessive is not a name and has no fake to
    agree with."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(
        [f"Rachel Green{CURLY}s", "Rachel Susanna Green"], [], [], registry=reg)
    people = {t.real: t.fake for t in terms if t.category == "person"}
    given, _middle, surname = people["Rachel Susanna Green"].split()
    # The possessive keeps its own letters: the surname's fake, then "’s".
    assert people[f"Rachel Green{CURLY}s"] == f"{given} {surname}{CURLY}s"
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    out = pz.apply(f"RACHEL GREEN{CURLY}S OPPOSITION")
    assert out == f"{given.upper()} {surname.upper()}{CURLY}S OPPOSITION"


def test_both_apostrophes_screen_a_bare_token_the_same_way():
    """`_pn_is_name_token` stripped only a straight possessive, so "Green's"
    reduced to "green" and was refused a bare token (a common-word surname)
    while "Green’s" reduced to "green’s", matched no list, and became one — a
    token whose FAKE carries a possessive, then applied to every near-miss
    spelling of the surname ("Grreen" -> "Yeardley’s")."""
    for apos in ("'", CURLY):
        reg = P._PnFakeRegistry()
        _full, bare = P._pn_fake_person(f"Rachel Green{apos}s", reg)
        assert [b[0] for b in bare] == ["Rachel"], apos
    assert P._pn_word_base(f"Green{CURLY}s") == "green"
    assert P._pn_is_name_token(f"Green{CURLY}s") is False


def test_an_apostrophe_name_is_one_word():
    """The same split hit "O’Brien", where the leading "O" became a
    one-letter word — an initial to everything that reads one."""
    assert P._PN_WORD_RE.findall(f"O{CURLY}Brien") == [f"O{CURLY}Brien"]
    reg = P._PnFakeRegistry()
    full, bare = P._pn_fake_person(f"Sean O{CURLY}Brien", reg)
    assert CURLY not in full            # the surname is faked whole, not "O’x"
    assert len(full.split()) == 2


def test_a_term_matches_the_other_apostrophe():
    """The E-Court spreadsheet exports `'`; the filing carries `’`. Matched
    literally, the party was left standing whole — and `surviving_reals` scans
    with the same pattern, so nothing reported it either."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Rachel Green's Trust", "Sean O'Brien"], [], [],
                              registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    out = pz.apply(f"RACHEL GREEN{CURLY}S TRUST sued. Sean O{CURLY}Brien signed.")
    for real in ("GREEN", "RACHEL", "Brien", "Sean"):
        assert real not in out, out
    # …and the other direction: a term carrying the curly mark matches straight
    # text, so neither spelling is the privileged one.
    reg2 = P._PnFakeRegistry()
    t2 = P._pn_build_terms([f"Sean O{CURLY}Brien"], [], [], registry=reg2)
    assert "Brien" not in P.Pseudonymizer(t2, DET, registry=reg2).apply(
        "Sean O'Brien signed.")
