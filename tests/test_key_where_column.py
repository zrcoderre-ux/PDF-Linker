"""The pseudonym key says WHICH document each Context quote came from, and where.

The Context column carries a sentence of the real filing as the evidence for a
binding — and evidence the operator cannot go and check is worth much less than
it looks. A folder is a dozen filings, so "which one, and where in it" was
answerable only by searching every export for the quoted sentence.

File and Where (page:line) sit at E and F, following the quote they describe,
in the LEAKS worksheet's own format (`_pn_where_label`) because it is the same
measurement. The three cells are minted in one statement of `note_key_context`,
off the site the quote was cut at, so a row can never name a document the
sentence beside it did not come from — and they are carried forward from the
key on disk as a UNIT with that quote, since the key outlives the folder's
contents.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_where_column.py -v
"""
import logging

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")

# Two printed pages, the second carrying a printed page number of its own, and
# a sentence that wraps across two gutter lines.
SOURCE = "\n".join([
    "====== Page 1 ======",
    " 1  SUPERIOR COURT OF THE STATE OF CALIFORNIA",
    " 2  HELEN RASHO, an individual,",
    " 3          Plaintiff,",
    "====== Page 2 (printed p. 4) ======",
    " 6  Helen Rasho signed the lease on April 2, 2024 and paid the",
    " 7  deposit that day. Marcus Delacroix witnessed the signing at",
    " 8  the branch office in Pasadena.",
    "",
])


def _run(source=SOURCE, name="Opposition to MSJ.pdf",
         parties=("Helen Rasho", "Marcus Delacroix")):
    reg = P._PnFakeRegistry()
    pz = P.Pseudonymizer(
        P._pn_build_terms(list(parties), [], [], registry=reg), [], reg)
    pz.note_key_context(source, pz.apply(source), name)
    return pz


def _rows(path, sheet="Pseudonym Key"):
    ws = openpyxl.load_workbook(path)[sheet]
    rows = list(ws.iter_rows(values_only=True))
    return [dict(zip(rows[0], r)) for r in rows[1:]]


# ── the columns themselves ──────────────────────────────────────────────────

def test_file_and_where_sit_at_E_and_F():
    """After the Context they describe, before Status. Positional only here —
    every reader of the key resolves by header name."""
    assert P._PN_KEY_HEADERS[4] == "File"
    assert P._PN_KEY_HEADERS[5] == "Where (page:line)"
    assert P._PN_KEY_HEADERS[3] == "Context"
    assert P._PN_KEY_HEADERS[6] == "Status"


def test_the_where_header_is_the_leaks_wording():
    """One measurement, one name for it — a reader comparing the two sheets
    must not have to work out whether they mean the same thing."""
    assert P._PN_KEY_WHERE_HEADER in P._PN_LEAK_HEADERS
    assert P._PN_KEY_WHERE_HEADER in P._PN_KEY_HEADERS


def test_the_widths_stay_in_step():
    w = dict(zip(P._PN_KEY_HEADERS, P._PN_KEY_WIDTHS))
    assert len(P._PN_KEY_WIDTHS) == len(P._PN_KEY_HEADERS)
    assert w["File"] and w["Where (page:line)"]


# ── what a row says ─────────────────────────────────────────────────────────

def test_every_quoted_row_names_its_document_and_its_page(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    _run().write_key(key, log)
    rows = [r for r in _rows(key) if r["Context"]]
    assert rows
    for r in rows:
        assert r["File"] == "Opposition to MSJ.pdf", r
        assert str(r["Where (page:line)"]).startswith(("p.", "line ")), r


def test_the_location_is_the_pdf_page_the_quote_came_off(tmp_path):
    """The value's own page, taken from the site the quote was cut at — the PDF
    page, since that is the number a viewer's page box takes and the number the
    export's own header carries, with the PRINTED page beside it where the
    document numbers itself differently."""
    key = tmp_path / "pseudonym_key.xlsx"
    _run().write_key(key, log)
    by_real = {str(r["Real Value"]): r for r in _rows(key)}
    assert by_real["Helen Rasho"]["Where (page:line)"] == "p.1:2"
    # "Marcus Delacroix" is quoted from the sentence that wraps lines 7-8 of
    # the SECOND PDF page, which prints itself as 4 — one range, the page named
    # once, and both numbers said because they differ.
    assert (by_real["Marcus Delacroix"]["Where (page:line)"]
            == "p.2 (printed p.4):7-8")


def test_a_body_with_no_pages_is_located_by_line(tmp_path):
    """A Word export has no page headers, so 'p.?' would name a page the run
    does not know. It has a line number, and that is what it says."""
    key = tmp_path / "pseudonym_key.xlsx"
    _run("Susan Spellman signed the lease.\n\nThe deposit cleared Tuesday.",
         "Letter.docx", ("Susan Spellman",)).write_key(key, log)
    rows = [r for r in _rows(key) if r["Context"]]
    assert rows
    for r in rows:
        assert r["File"] == "Letter.docx"
        assert str(r["Where (page:line)"]).startswith("line "), r


def test_a_quote_crossing_a_page_spells_both_pages(tmp_path):
    src = "\n".join(["====== Page 7 ======",
                     " 27  Marcus Delacroix witnessed the",
                     "====== Page 8 ======",
                     " 1  signing of the lease agreement in full."])
    key = tmp_path / "pseudonym_key.xlsx"
    _run(src, "Ex A.pdf", ("Marcus Delacroix",)).write_key(key, log)
    where = {str(r["Real Value"]): r["Where (page:line)"] for r in _rows(key)}
    assert where["Marcus Delacroix"] == "p.7:27-p.8:1"


def test_a_row_with_no_quote_claims_no_location(tmp_path):
    """A binding this run never met has no sentence, so it has nothing to say
    about where one stands. An empty Context with a populated Where would read
    as a place the operator could go and look."""
    key = tmp_path / "pseudonym_key.xlsx"
    _run().write_key(key, log)
    for sheet in openpyxl.load_workbook(key).sheetnames:
        for r in _rows(key, sheet):
            if not r["Context"]:
                assert not r["File"] and not r["Where (page:line)"], r


# ── across runs ─────────────────────────────────────────────────────────────

def test_the_pair_is_carried_forward_with_its_quote(tmp_path):
    """The key outlives the folder's contents: a party named only in a filing
    delivered two runs ago keeps its binding, so it keeps its sentence — and a
    sentence whose document and page were dropped is evidence again turned
    uncheckable."""
    key = tmp_path / "pseudonym_key.xlsx"
    _run().write_key(key, log)
    before = {str(r["Real Value"]): (r["File"], r["Where (page:line)"])
              for r in _rows(key) if r["Context"]}

    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log)
    P.Pseudonymizer(terms, [], reg).write_key(key, log)   # no documents at all

    after = {str(r["Real Value"]): (r["File"], r["Where (page:line)"])
             for r in _rows(key) if r["Context"]}
    assert before and after == before


def test_this_runs_pair_wins_over_the_carried_one(tmp_path):
    """A re-derived quote arrives with a re-derived location; taking one from
    this run and the other from disk is exactly the mismatch these columns are
    written together to prevent."""
    key = tmp_path / "pseudonym_key.xlsx"
    _run().write_key(key, log)

    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log)
    pz = P.Pseudonymizer(terms, [], reg)
    src = "\n".join(["====== Page 3 ======",
                     " 9  Helen Rasho testified about the deposit in detail."])
    pz.note_key_context(src, pz.apply(src), "Reply.pdf")
    pz.write_key(key, log)

    row = {str(r["Real Value"]): r for r in _rows(key)}["Helen Rasho"]
    assert row["File"] == "Reply.pdf"
    assert row["Where (page:line)"] == "p.3:9"
    assert "testified" in str(row["Context"])


def test_an_older_key_reads_and_is_migrated(tmp_path):
    """A key written before these columns existed simply yields nothing for
    them, and the rewrite at the end of the run gives it the new layout without
    disturbing a binding or a quote."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement",
               "Context", "Status", "Source", "Occurrences"])
    ws.append(["person", "Helen Rasho", "Kelsall Strangeways",
               "Helen Rasho signed it.", "replaced", "spreadsheet", 3])
    wb.save(key)

    _orig, _scrub, where = P._pn_key_context_on_disk(key)
    assert where == {}

    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log)
    P.Pseudonymizer(terms, [], reg).write_key(key, log)

    rows = _rows(key)
    assert list(openpyxl.load_workbook(key)["Pseudonym Key"]
                .iter_rows(values_only=True))[0] == P._PN_KEY_HEADERS
    row = {str(r["Real Value"]): r for r in rows}["Helen Rasho"]
    assert row["Replacement"] == "Kelsall Strangeways"
    assert "Helen Rasho signed it." in str(row["Context"])
    assert not row["File"] and not row["Where (page:line)"]


# ── the location primitive ──────────────────────────────────────────────────

def test_the_line_label_never_invents_a_page():
    assert P._pn_where_label("8", "16", 40) == "p.8:16"
    assert P._pn_where_label("8", None, 40) == "p.8"
    assert P._pn_where_label("?", None, 40) == "line 40"
    assert P._pn_where_label("?", "3", 40) == "line 40"


def test_a_page_names_the_pdf_page_first_and_the_printed_one_beside_it():
    """A compiled filing RESTARTS its numbering at every sub-document, so the
    printed number names no page a reader can turn to: the 43rd page of an
    exhibit bundle prints "1", and so do a dozen others. The PDF page leads;
    the printed page is kept because it is the half a court cites, and is said
    only where it differs, so an ordinary filing's key does not move."""
    assert P._pn_page_label("43", "1") == "43 (printed p.1)"
    assert P._pn_where_label(P._pn_page_label("43", "1"), "16", 9) \
        == "p.43 (printed p.1):16"
    # The ordinary born-digital case: the two agree, so nothing is appended and
    # a delivered key's Where column reads exactly as it always did.
    assert P._pn_page_label("3", "3") == "3"
    assert P._pn_page_label("3", None) == "3"
    # Roman front matter always differs, and always says so.
    assert P._pn_page_label("5", "iv") == "5 (printed p.iv)"


def test_a_review_banner_does_not_cost_the_page_its_number():
    """A page whose text layer was rebuilt, whose images were read by OCR, or
    that was recognised below `_OCR_LOW_DPI` carries a " — REVIEW: …" clause in
    its header. The header pattern demanded the closing rule hard after the
    page number, so every such header failed to match — and a header that does
    not match does not merely lose its own page: the parser keeps the LAST page
    it matched, so a whole scanned exhibit set was reported at the number of
    the last clean page before it."""
    body = "\n".join([
        "====== Page 1 ======",
        " 1  Helen Rasho signed the lease.",
        "====== Page 43 — REVIEW: recognised at only 99 dpi, text is LOW "
        "CONFIDENCE ======",
        " 7  Helen Rasho signed it again.",
        "====== Page 44 (printed p. 2) — REVIEW: the text layer was unreadable "
        "and was REBUILT by OCR; spellings, numbers and citations on this page "
        "are GUESSES ======",
        " 3  Marcus Delacroix witnessed it.",
        "",
    ])
    parsed = P._pn_body_lines(body)
    assert P._pn_locate(parsed, "signed it again") == "p.43:7"
    assert (P._pn_locate(parsed, "Marcus Delacroix")
            == "p.44 (printed p.2):3")


def test_the_site_where_is_read_off_the_prep_table():
    """`_pn_context_prep` drops blank and gutter-only lines, so an index into
    the parsed body would be off by every one of them."""
    parsed = P._pn_body_lines(SOURCE)
    _q, site = P._pn_context_hit(parsed, "Marcus Delacroix")
    assert P._pn_site_where(parsed, site) == "p.2 (printed p.4):7-8"
    assert P._pn_site_where(parsed, None) == ""
