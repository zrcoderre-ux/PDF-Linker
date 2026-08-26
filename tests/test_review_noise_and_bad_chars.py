"""Two faults a landlord-tenant batch exposed, and one gazetteer hygiene rule.

1. KEPT WORDS ARE NOT OUR FAKES. `_pn_fake_person` preserves the furniture of a
   party name verbatim ("The Lovelace" -> "The Flintham"), and a `{braced}` keep
   too — but `name_fake_words()` split the composed fake and counted every word,
   so "the" became one of "our person fake words". `half_scrubbed_scan` then
   read EVERY title-case phrase in the document as a half-scrubbed pair: one
   filing reported 71, among them "The Bane Act", "The Unruh Civil Rights Act",
   "The Stanley Mosk Courthouse" and "TO THE HONORABLE COURT".

2. ONE INVISIBLE BYTE MUST NOT COST THE KEY. openpyxl refuses a C0 control
   character, and a scanned exhibit supplies them — OCR read a Bates stamp as
   "EQUITY-WALTON_0007 - HOIISING<BEL>COMMl,/Nl't'Y". Writing it threw, so the
   ENTIRE reversal key was lost while the exports were written: pseudonyms
   nobody can undo, reported as one "non-fatal" warning.

3. A WORD LIST IS DATA, NOT CODE. A `#` inside a triple-quoted word list is not
   a comment. Written that way first, the housing block put "#", "Act", "Bane,"
   and "Fair" into the gazetteer — swallowing the very surnames its own
   rationale said to protect.

Run:  cd PDF-Linker && python3 -m pytest tests/test_review_noise_and_bad_chars.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _pz(names):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    return P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)


# ── 1. a kept word is not one of our fakes ──────────────────────────────────

def test_kept_furniture_is_not_counted_as_a_person_fake():
    z = _pz(["The Lovelace", "Rolleston Devereaux"])
    fake = next(r["fake"] for (c, _r), r in z.records.items()
                if c == "person" and r["real"] == "The Lovelace")
    assert fake.startswith("The "), fake      # the furniture really is kept
    assert "the" not in z.name_fake_words()


def test_a_title_case_phrase_is_not_a_half_scrubbed_pair():
    z = _pz(["The Lovelace", "Rolleston Devereaux"])
    body = z.apply(
        "The Bane Act provides a private right of action. The Unruh Civil "
        "Rights Act says otherwise. The Stanley Mosk Courthouse is the proper "
        "venue. The Legislature commanded this. TO THE HONORABLE COURT: "
        "Rolleston Devereaux resided at The Lovelace.")
    assert z.half_scrubbed_scan(body) == []


def test_a_real_half_scrub_is_still_caught():
    # The scan's whole reason for existing must survive the fix.
    z = _pz(["The Lovelace", "Rolleston Devereaux"])
    body = z.apply("Contact Xiaoxia Devereaux about the lease.")
    assert [v for _c, v in z.half_scrubbed_scan(body)] == ["Xiaoxia"]


def test_a_braced_keep_is_not_counted_either(monkeypatch):
    # A nuclear keep rides on the registry and is kept verbatim while the fake
    # is composed, exactly like the built-in furniture — so it is the
    # document's word, not ours.
    reg = P._PnFakeRegistry()
    reg.keep_words = frozenset({"vanguard"})
    terms = P._pn_build_terms(["Quillmark Vanguard"], [], [], registry=reg)
    z = P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)
    assert "vanguard" not in z.name_fake_words()


# ── 2. an illegal character must not cost the key ───────────────────────────

BAD = "EQUITY-WALTON_0007 - HOIISING\x07COMMl,/Nl't'Y"


def test_xl_text_strips_what_excel_refuses():
    assert "\x07" not in P._pn_xl_text(BAD)
    assert P._pn_xl_text(BAD).startswith("EQUITY-WALTON_0007")
    assert P._pn_xl_text(7) == 7                 # non-strings pass through
    assert P._pn_xl_text(None) is None


def test_the_key_still_writes_when_a_value_carries_a_control_character(tmp_path):
    z = _pz(["Helen Rasho"])
    z.records[("identifier", BAD.lower())] = {
        "category": "identifier", "real": BAD, "fake": "DEAL-0001",
        "source": "prescan", "count": 3, "pattern": "x", "flags": 0}
    src = f"====== Page 1 ======\n 1  Helen Rasho signed {BAD} today.\n"
    z.apply(src)
    z.note_key_context(src)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)                        # used to raise
    rows = [r for ws in openpyxl.load_workbook(key).worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
    assert rows, "the key came out empty"
    # The BINDING is kept — stripping the invisible byte loses nothing, while
    # dropping the row would lose the reversal.
    kept = [r[1] for r in rows if "EQUITY-WALTON" in str(r[1])]
    assert kept and "\x07" not in kept[0], kept


def test_the_leaks_worksheet_survives_one_too(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "M.pdf", "type": "name?", "value": BAD, "where": "p.1:1",
         "context": f"The stamp reads {BAD} on every page.", "notes": ""}], log)
    ws = openpyxl.load_workbook(tmp_path / "LEAKS.xlsx").active
    vals = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert vals and "\x07" not in str(vals[0])


def test_the_master_sheets_survive_one_too(tmp_path, monkeypatch):
    master = tmp_path / "master.xlsx"
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    P._pn_update_master_keep(
        {}, {BAD.lower(): {"value": BAD, "type": "KEEP", "fix": "no",
                           "notes": "kept"}},
        "23STCV1 Test", "2026-08-06", log)
    assert master.exists()
    for ws in openpyxl.load_workbook(master).worksheets:
        for row in ws.iter_rows(values_only=True):
            assert not any("\x07" in str(c) for c in row if c is not None)


# ── 3. the gazetteer holds words, not prose ─────────────────────────────────

@pytest.mark.parametrize("name", [
    "_PN_COMMON_WORDS", "_PN_SERVICE_GENERIC_WORDS", "_PN_FORM_LABEL_WORDS",
    "_PN_REVIEW_NAME_STOP", "_PN_HOUSING_WORDS", "_PN_NARRATIVE_VERBS",
])
def test_a_word_list_contains_only_lower_case_words(name):
    """A `#` inside a triple-quoted list is data, not a comment.

    This is the assertion that would have caught the housing block's rationale
    being swallowed into the gazetteer along with "Act", "Bane," and "Fair".
    """
    bad = [w for w in getattr(P, name)
           if not w.replace("-", "").replace("'", "").isalpha()
           or w != w.lower()]
    assert not bad, f"{name} holds non-words: {sorted(bad)[:10]}"


@pytest.mark.parametrize("surname", ["fair", "bane", "prior", "person",
                                     "bureau", "act", "credit", "default",
                                     "rent", "senior", "public"])
def test_a_word_that_could_be_a_surname_stays_reportable(surname):
    # Vocabulary that DOUBLES as a real surname must not be swallowed: a word
    # withheld from the review gazetteer stops being reportable, and a party
    # really named Fair costs far more than one more triage row.
    assert surname not in P._PN_COMMON_WORDS


def test_the_housing_words_reached_the_gazetteer():
    assert P._PN_HOUSING_WORDS <= P._PN_COMMON_WORDS
    for w in ("premises", "utilities", "tenancy", "habitability", "writ"):
        assert w in P._PN_COMMON_WORDS


def test_the_housing_words_collide_with_no_fake_pool():
    pools = ({w.lower() for w in P._PN_NAME_WORDS}
             | {w.lower() for w in P._PN_ENTITY_WORDS}
             | {w.lower() for w in P._PN_STREET_NAMES})
    assert not (P._PN_HOUSING_WORDS & pools)
