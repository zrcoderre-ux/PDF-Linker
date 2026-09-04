"""`*CANONICAL` binds the canonical when this case has not.

The shape: the only spelling of a party ANY document in the folder carries is a
misspelling. "Vazqez" is flagged, the operator answers `*Vazquez`, and the
correct spelling appears nowhere — so there was nothing to mirror, the alias
was refused, and the value took an unrelated pool word. One party under a
stand-in that says nothing about the name it replaced, and the next document to
spell it RIGHT draws a second unrelated word.

Binding it costs one pool word and lands where a declared-but-absent value
belongs: `write_key` gives a binding no export carried Status `no match` and
puts it on `_PN_KEY_PINNED_SHEET`, which `DeAnonymize` cannot reach — forward
only, which is all this needs, while `_pn_load_key` reads both sheets so the
pin waits for the run where a document finally spells the name out.

What is given up is the refusal, which was the only screen on what was typed
after the star. Two things hold it: the canonical clears the same shape screens
a `--term` clears, and the binding is announced by name.

Run:  cd PDF-Linker && python3 -m pytest tests/test_alias_binds_unknown_canonical.py -v
"""
import logging

import pytest

import pdf_linker as P


class _Capture(logging.Handler):
    def __init__(self):
        super().__init__()
        self.lines = []

    def emit(self, record):
        self.lines.append((record.levelname, record.getMessage()))


def _run(canonical, value="Vazqez", already=()):
    """Apply `*canonical` to `value`; returns (values, log lines, registry)."""
    log = logging.getLogger(f"alias-{canonical}-{value}")
    log.handlers[:] = []
    cap = _Capture()
    log.addHandler(cap)
    log.setLevel(logging.INFO)
    log.propagate = False
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(already) or ["Manuel Sarkisyan"], [], [],
                              registry=reg)
    rows = [["Value", "Fix? (yes/no)", "Type", "Notes"],
            [value, canonical, "LEAK", ""]]
    d = P._pn_parse_decision_rows(rows)
    _terms, values = P._pn_apply_aliases(d, terms, reg, log)
    return values, cap.lines, reg


# ── the shape it exists for ────────────────────────────────────────────────

def test_an_unbound_canonical_is_bound_and_the_value_mirrors_it():
    values, lines, reg = _run("*Vazquez")
    assert set(values) == {"Vazqez", "Vazquez"}      # both need terms built
    tok = reg.tokens_for("nametok")
    assert tok["vazquez"] and tok["vazqez"]
    # one person spelled two ways, and two DISTINCT rows to reverse — a shared
    # Replacement is what `DeAnonymize.bas` calls ambiguous.
    assert tok["vazqez"] != tok["vazquez"]
    assert P._pn_osa_distance(tok["vazqez"].lower(),
                              tok["vazquez"].lower()) == 1


def test_the_binding_is_announced_by_name():
    """The refusal was the only screen on the spelling after the star, so the
    replacement for it has to be readable: a typo stays visible."""
    _values, lines, _reg = _run("*Vazquez")
    said = " ".join(m for _lvl, m in lines)
    assert "'Vazquez'" in said and "BOUND on your say-so" in said
    assert "check the spelling" in said.lower()


def test_a_canonical_this_case_already_binds_is_never_re_bound():
    """The ordinary path: mirror what is there, mint nothing, say nothing new."""
    values, lines, reg = _run("*Sarkisyan", already=["Manuel Sarkisyan"])
    assert "Sarkisyan" not in values          # not re-created
    assert not any("BOUND on your say-so" in m for _l, m in lines)


# ── the screens that replace the refusal ───────────────────────────────────

@pytest.mark.parametrize("canonical", [
    "*Doe",        # a common-word surname — `_pn_is_name_token` refuses it
    "*the",        # not a name at all
    "*Esq",        # a professional suffix, not a name
    "*Court",      # a role/procedural word
])
def test_a_canonical_that_is_not_name_shaped_still_refuses(canonical):
    """The same question the term builder asks before a bare token may exist,
    asked here because nothing else reads this value: it reaches the term list
    without having been read off any document."""
    values, lines, _reg = _run(canonical)
    assert values == ["Vazqez"]                       # value still faked
    assert any(lvl == "WARNING" and "nothing to mirror" in m
               for lvl, m in lines)


@pytest.mark.parametrize("canonical,value,bound", [
    ("*Vazquez", "Vazqez", True),      # one slip — a misspelling
    ("*Smythe", "Smyth", True),        # two, at a length that allows it
    ("*Nobody", "Antiono", False),     # five — not a spelling of anything
    ("*Vazquez", "Vatqual", False),    # a mangled scan, past the report reach
])
def test_the_pair_must_be_near_enough_to_be_ONE_misspelling(canonical, value,
                                                            bound):
    """`fold_onto`'s own `_PN_FOLD_MAX_REPS` is a LENGTH-delta bound, not a
    distance one, so it happily mirrors "ANTIONO" onto "NOBODY". That is fine
    when the operator names a value this case already holds and not fine when
    it INVENTS one, because a mistyped canonical would enter the key as a Real
    Value no document ever carried. The reach is the REPORT tier's own —
    "near enough to ask about"."""
    values, _lines, _reg = _run(canonical, value=value)
    assert (len(values) == 2) is bound


def test_the_reach_is_tighter_than_the_alias_onto_a_BOUND_canonical():
    """Deliberately asymmetric: folding onto a value the folder really contains
    is the operator settling which of two spellings is the person; inventing a
    third string is a bigger step. So the mangled scan refused above is honoured
    the moment the correct spelling is bound."""
    values, lines, reg = _run("*Vazquez", value="Vatqual",
                              already=["Manuel Vazquez"])
    assert reg.tokens_for("nametok")["vatqual"]        # honoured
    assert not any("BOUND on your say-so" in m for _l, m in lines)


def test_the_tool_never_takes_its_own_stand_in_as_a_canonical():
    """The `_pn_build_terms` gate, which matters more here than anywhere: this
    value reaches the term list without being read off a document, so a
    stand-in named after the star would enter the key as a Real Value."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Manuel Sarkisyan"], [], [], registry=reg)
    ours = next(str(t.fake) for t in terms if str(t.real) == "Sarkisyan")
    assert P._pn_alias_bind_canonical(reg, ours, ours.lower(), "vazqez",
                                      logging.getLogger("x")) == ""


def test_the_screen_reads_the_word_as_typed_not_its_folded_base():
    """`_pn_is_name_token`'s first question is whether the word is capitalised
    where it was written, so asking it about a lower-cased base answers False
    for every name there is — which silently disabled the whole feature."""
    reg = P._PnFakeRegistry()
    assert P._pn_is_name_token("Vazquez") and not P._pn_is_name_token("vazquez")
    assert P._pn_alias_bind_canonical(reg, "Vazquez", "vazquez", "vazqez",
                                      logging.getLogger("x"))


def test_the_stand_in_is_read_back_from_the_memo():
    """So the fake handed to `fold_onto` carries exactly the normalisation and
    canonical case it would have had if a document had spelled the name out."""
    reg = P._PnFakeRegistry()
    fake = P._pn_alias_bind_canonical(reg, "Vazquez", "vazquez", "vazqez",
                                      logging.getLogger("x"))
    assert fake == reg.tokens_for("nametok")["vazquez"]
    assert fake[:1].isupper()


# ── where the row goes ─────────────────────────────────────────────────────

def test_the_bound_canonical_is_pinned_and_the_misspelling_is_not(tmp_path):
    """A binding no export carried is right forward and a hazard in reverse, so
    it goes to the sheet `DeAnonymize` cannot reach — while the misspelling,
    which the export really does carry, stays on the main sheet."""
    import openpyxl
    log = logging.getLogger("test")
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Manuel Sarkisyan"], [], [], registry=reg)
    d = P._pn_parse_decision_rows(
        [["Value", "Fix? (yes/no)", "Type", "Notes"],
         ["Vazqez", "*Vazquez", "LEAK", ""]])
    terms, values = P._pn_apply_aliases(d, terms, reg, log)
    terms = P._pn_build_terms(["Manuel Sarkisyan"], [], values, registry=reg)
    pz = P.Pseudonymizer(terms, [], registry=reg)
    pz.apply("the Vazqez guaranty")            # only the misspelling appears
    kp = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(kp, log)

    wb = openpyxl.load_workbook(kp)
    where = {}
    for name in wb.sheetnames:
        ws = wb[name]
        hdr = [str(c.value or "").strip().lower() for c in ws[1]]
        rv = hdr.index("real value")
        for row in ws.iter_rows(min_row=2, values_only=True):
            where.setdefault(str(row[rv]), set()).add(name)
    assert where["Vazqez"] == {P._PN_KEY_MAIN_SHEET}
    assert where["Vazquez"] == {P._PN_KEY_PINNED_SHEET}
