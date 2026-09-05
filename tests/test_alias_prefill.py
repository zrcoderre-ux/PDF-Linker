"""A worksheet row that is a MISSPELLING of a tracked value arrives answered.

The fuzzy sweep already knows which tracked token a survivor is a slip of, and
the `~CANONICAL` alias control word is exactly the answer, so the Fix? cell is
PRE-FILLED with it — at the operator's direction: leave it if it is right,
change it if not. Confident cases only (the MINTING fold distance, or a clipped
lead), every word resolved, the canonical bound, never a LEAK row and never a
row that already carries a decision.

Run:  cd PDF-Linker && python3 -m pytest tests/test_alias_prefill.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _pz(*names):
    reg = P._PnFakeRegistry()
    return P.Pseudonymizer(P._pn_build_terms(list(names), [], [], registry=reg),
                           {}, registry=reg)


# ── the suggestion ──────────────────────────────────────────────────────────

@pytest.mark.parametrize("value,canon", [
    ("Miachael", "Michael"),                    # an adjacent transposition
    ("Miachael Rodgers", "Michael Rodgers"),    # …inside a two-word phrase
    ("Rodgerz", "Rodgers"),                     # one substitution
    ("idland", "Midland"),                      # a clipped lead
    ("Michal", "Michael"),                      # one deletion
])
def test_a_slip_of_a_tracked_token_is_suggested(value, canon):
    z = _pz("Michael Rodgers", "Midland States Bank")
    assert z.alias_suggestion(value) == canon


@pytest.mark.parametrize("value", [
    "Michael Rodgers",       # the tracked value itself — the survivor scan's row
    "Rodgers",               # a tracked token, nothing misspelled
    "Bob",                   # nothing near
    "M idland",              # a broken spelling: "M" resolves to nothing
    "Miachael Smithe",       # one word unresolved
    "Xiaoxia",               # no tracked token near it
])
def test_no_suggestion_where_the_answer_is_not_confident(value):
    z = _pz("Michael Rodgers", "Midland States Bank")
    assert z.alias_suggestion(value) == ""


def test_two_tokens_equally_near_is_ambiguity():
    z = _pz("Rachel Marlow", "Rachel Marlon")
    assert z.alias_suggestion("Marlox") == ""


@pytest.mark.parametrize("value,canon", [
    # The worksheet an operator filled by hand for one defendant, verbatim.
    ("Manual Vaiquez", "Manuel Vazquez"),
    ("Manue", "Manuel"),
    ("Manuel va2que1", "Manuel Vazquez"),
    ("Manuel Vaiquei", "Manuel Vazquez"),
    ("Manuel Varquez", "Manuel Vazquez"),
    ("Manuel vauiuez", "Manuel Vazquez"),
    ("Manuel Vazqoe", "Manuel Vazquez"),
    ("Manuel vazque", "Manuel Vazquez"),
    ("Manuel Vazqu~z", "Manuel Vazquez"),
    ("Manuel vazqvez", "Manuel Vazquez"),
    ("Manvel Vazquez", "Manuel Vazquez"),
    ("Va2quez", "Vazquez"),
    ("Vaiquel", "Vazquez"),
    ("Vaiquez", "Vazquez"),
    ("Vasquez", "Vazquez"),
    ("Vatquel", "Vazquez"),
    ("Vatquez", "Vazquez"),
    ("Vazquei", "Vazquez"),
    ("vizquez", "Vazquez"),
    ("zquez", "Vazquez"),
])
def test_the_operators_own_worksheet_is_answered(value, canon):
    z = _pz("Manuel Vazquez", "Midland States Bank")
    assert z.alias_suggestion(value) == canon


@pytest.mark.parametrize("value,canon", [
    ("Manuel Vazq~~1", "Manuel Vazquez"),    # three slips: the wide reach
    ("zquei", "Vazquez"),                    # a clipped lead AND a slip
])
def test_a_named_partys_net_is_cast_wide(value, canon):
    """Every row of that worksheet, then: a party the operator NAMED takes
    the sweep's widest reach on every page."""
    z = _pz("Manuel Vazquez", "Midland States Bank")
    assert z.alias_suggestion(value) == canon


@pytest.mark.parametrize("value", [
    "Manuel Vazq~~1", "zquei",
])
def test_a_harvested_name_keeps_the_clean_page_reach(value):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Midland States Bank"], [], [], registry=reg)
    P._pn_append_name_terms(terms, "Manuel Vazquez", "document", reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    assert z.alias_suggestion(value) == ""


def test_the_tracked_value_itself_is_never_an_alias():
    z = _pz("Manuel Vazquez")
    assert z.alias_suggestion("Manuel Vazquez") == ""


def test_the_wide_net_reaches_the_sweep_too():
    """The report casts the same net: a named party's token is asked about
    at the degraded reach on a clean page — once the document has shown a
    SECOND spelling of it. A lone three-slip spelling, however often it
    recurs, is as likely a different name and must be a close match
    (`test_variant_reach.py`); the pre-fill by value is unchanged."""
    z = _pz("Manuel Vazquez")
    out = z.apply("Name: Manuel Vazquez. Print Name: Manuel Vatqual, loan. "
                  "Vatqual again. Guarantor: Manuel Vazqoez.")
    got = [s for _c, s in z.fuzzy_survivor_scan(out)]
    assert "Vatqual" in got and "Vazqoez" in got            # 3 slips, 1 slip
    assert z.alias_suggestion("Vatqual") == "Vazquez"
    # A lone "Vatqual" is reported only while it is BARE — never beside a
    # name word nothing tracks. Behind a stranger's given name it is
    # somebody else.
    lone = _pz("Manuel Vazquez")
    out = lone.apply("Name: Manuel Vazquez. Print Name: Robert Vatqual, loan. "
                     "Robert Vatqual again.")
    assert "Vatqual" not in [s for _c, s in lone.fuzzy_survivor_scan(out)]


def test_an_unusual_letter_pair_is_what_licenses_the_wide_reach():
    """"zq" is in Vazquez and in almost nothing else a filing carries, so a
    survivor sharing it is reached at three edits; a six-letter party token
    with no such pair ("Manuel") stays at the ordinary reach, or "handle",
    "Model" and "Carmel" all arrive as the defendant."""
    assert P._pn_shares_rare_bigram("vatqual", "vazquez")
    assert not P._pn_shares_rare_bigram("manual", "manuel")
    assert P._pn_scan_fold_dist("vatqual", "vazquez", party=True) == 3
    assert P._pn_scan_fold_dist("vatqual", "vazquez", party=False) == 2
    assert P._pn_scan_fold_dist("handle", "manuel", party=True) == 2
    z = _pz("Manuel Vazquez")
    for word in ("handle", "Model", "Carmel"):
        assert z.alias_suggestion(word) == "", word


def test_a_word_the_original_writes_lower_case_is_vocabulary():
    """"Status" is one slip from a party token "States"; the document's own
    "the status of the loan" says it is a word."""
    z = _pz("Midland States Bank")
    z.note_original("The status of the loan was reported. Status: open.")
    assert z.alias_suggestion("Status") == ""
    z2 = _pz("Midland States Bank")
    assert z2.alias_suggestion("Status") == "States"     # no evidence, no screen


def test_the_prefill_reach_is_the_sweeps_clean_page_reach():
    """The sweep asks at the fold PLUS one on a clean page and plus two
    inside a degraded region; the pre-fill takes the first and not the
    second — it has no page in hand."""
    reg = P._PnFakeRegistry()
    terms = []
    P._pn_append_name_terms(terms, "Michael Rodgers", "document", reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    assert z.alias_suggestion("Mchaeil") == "Michael"      # two slips
    assert z.alias_suggestion("Mixhxxl") == ""              # three


def test_a_derived_spelling_never_names_the_canonical():
    """`_pn_name_variants` registers near-spellings that sit exactly as close
    as the real word; they must neither win nor make the answer ambiguous."""
    z = _pz("Michael Rodgers")
    assert z.alias_suggestion("Miachael") == "Michael"


# ── the worksheet ───────────────────────────────────────────────────────────

def _rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[P._PN_LEAK_SHEET]
    head = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    return [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def _write(folder, z, entries, decisions=None):
    P._pn_write_leak_report(folder, entries, log, decisions=decisions,
                            suggest_for=z.alias_suggestion,
                            note_for=z.triage_note)
    return {r["Value"]: r for r in _rows(folder / "LEAKS.xlsx")}


def test_the_fix_cell_arrives_with_the_alias(tmp_path):
    z = _pz("Michael Rodgers")
    rows = _write(tmp_path, z, [
        {"file": "Decl.txt", "type": "misspelled name?", "value": "Miachael",
         "where": "p.1:3", "context": "Miachael served it."},
        {"file": "Decl.txt", "type": "unscrubbed name?", "value": "Xiaoxia Deng",
         "where": "p.1:4", "context": "Xiaoxia Deng signed."},
    ])
    assert rows["Miachael"]["Fix? (yes/no)"] == "~Michael"
    assert "pre-filled" in rows["Miachael"]["Notes"]
    assert "Michael" in rows["Miachael"]["Notes"]
    assert not rows["Xiaoxia Deng"]["Fix? (yes/no)"]


def test_a_leak_row_is_never_prefilled(tmp_path):
    z = _pz("Michael Rodgers")
    rows = _write(tmp_path, z, [
        {"file": "Decl.txt", "type": "LEAK", "value": "Michal",
         "where": "p.1:3", "context": "Michal served it."}])
    assert not rows["Michal"]["Fix? (yes/no)"]


def _decision(value, fixcell, fix="yes", replacement=None):
    return {value.lower(): {"value": value, "type": "misspelled name?",
                            "fix": fix, "replacement": replacement,
                            "fake_values": None, "fixcell": fixcell,
                            "alias": None, "notes": "", "cases": "",
                            "origin": ""}}


def test_a_decision_already_typed_is_never_overwritten(tmp_path):
    """An operator's own answer — here a typed replacement — stands."""
    z = _pz("Michael Rodgers")
    rows = _write(tmp_path, z, [
        {"file": "Decl.txt", "type": "misspelled name?", "value": "Miachael",
         "where": "p.1:3", "context": "Miachael served it."}],
        _decision("Miachael", "Wemyss", replacement="Wemyss"))
    assert rows["Miachael"]["Fix? (yes/no)"] == "Wemyss"
    assert "pre-filled" not in str(rows["Miachael"]["Notes"])


def test_leaving_the_cell_reads_back_as_the_alias(tmp_path):
    """The operator opens nothing: the next pass reads the pre-filled cell as
    an ordinary alias decision, which is the whole point."""
    z = _pz("Michael Rodgers")
    _write(tmp_path, z, [
        {"file": "Decl.txt", "type": "misspelled name?", "value": "Miachael",
         "where": "p.1:3", "context": "Miachael served it."}])
    d = P._pn_read_leak_decisions(tmp_path)["miachael"]
    assert d["fix"] == "yes" and d["alias"] == "Michael"
    assert d["fixcell"] == "~Michael"


def test_a_prefilled_row_still_sorts_as_one_to_look_at(tmp_path):
    """`fix` stays "" on a pre-filled row, so it sorts with the undecided
    rows at the top rather than with the resolved ones at the bottom."""
    z = _pz("Michael Rodgers")
    P._pn_write_leak_report(tmp_path, [
        {"file": "Decl.txt", "type": "misspelled name?", "value": "Miachael",
         "where": "p.1:3", "context": "Miachael served it."}],
        log, decisions=_decision("Zed Quill", "yes"),      # resolved, absent
        suggest_for=z.alias_suggestion)
    ordered = [r["Value"] for r in _rows(tmp_path / "LEAKS.xlsx")]
    assert ordered.index("Miachael") < ordered.index("Zed Quill"), ordered


def test_the_alias_is_applied_the_way_a_typed_one_is():
    """End to end through the alias machinery: the misspelling is faked as
    the same slip of the canonical's stand-in."""
    z = _pz("Michael Rodgers")
    canon = z.alias_suggestion("Miachael")
    decisions = {"miachael": {"value": "Miachael", "alias": canon, "fix": "yes"}}
    terms, values = P._pn_apply_aliases(decisions, list(z.terms), z.registry,
                                        log)
    assert values == ["Miachael"]
    fake = z.registry.tokens_for("nametok")
    assert "miachael" in fake, fake
    real_fake = fake["michael"]
    assert fake["miachael"] != real_fake
    assert P._pn_osa_distance(fake["miachael"].lower(), real_fake.lower()) <= 2
