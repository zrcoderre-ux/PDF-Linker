"""
The NUCLEAR keep: `{braces}` in a Fix? / Replacement cell.

`[bracket]` says "this fragment is never a name HERE"; `{brace}` says "this can
never reveal anything, anywhere". So a braced word is never faked in any folder
— not even inside a party name, because the party's fake is COMPOSED with the
word left verbatim ("Alder Law, P.C." -> "Kaldor Law, P.C."). That is what makes
it safe to inherit globally: a keep the composing faker honours can never leave a
party in the clear.

Run:  cd PDF-Linker && python3 -m pytest tests/test_nuclear_keep.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}


def _decision(value, cell, origin="", cases=""):
    """A decision as `_pn_parse_decision_rows` would build it from a Fix? cell."""
    rows = [("Value", "Fix? (yes/no)", "Type", "Notes", "Cases", "Origin"),
            (value, cell, "", "", cases, origin)]
    return P._pn_parse_decision_rows(rows)


def _pz(names, decisions, extra=()):
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, decisions)
    terms = P._pn_build_terms(list(names), [], list(extra), registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    pz.keep_strict, pz.keep_soft, pz.keep_nuclear = P._pn_keep_values(decisions)
    pz._keep_decisions = {vl: d for vl, d in decisions.items()
                          if P._pn_decision_is_keep(d)}
    pz._keep_local = set()          # INHERITED everywhere below, on purpose
    return pz


# ── parsing ─────────────────────────────────────────────────────────────────

def test_braces_parse_like_brackets_and_cut_the_same_way():
    assert P._pn_bracket_keep("Mulliken Medical Center", "{Medical Center}") \
        == ["Mulliken"]
    assert P._pn_bracket_keep("Raytheon's Human Resources",
                              "{Human Resources}") == ["Raytheon's"]
    # Mixed in one cell, and a whole-value brace keeps everything.
    assert P._pn_bracket_keep("Alder Law Group", "{Law} [Group]") == ["Alder"]
    assert P._pn_bracket_keep("Attached", "{Attached}") == []
    # Not a keep-spec at all -> still an explicit replacement.
    assert P._pn_bracket_keep("Smith", "Jones") is None


def test_a_brace_is_a_nuclear_keep_a_bracket_is_not():
    braced = _decision("Mulliken Medical Center", "{Medical Center}")
    bracketed = _decision("Alder Law, P.C.", "[Law]")
    assert P._pn_keep_values(braced)[2] == {"Medical Center"}
    assert P._pn_keep_values(bracketed)[2] == set()
    assert P._pn_keep_values(bracketed)[0] == {"Law"}
    # A whole-value brace is nuclear, not an ordinary soft `no`.
    whole = _decision("Attached", "{Attached}")
    assert P._pn_keep_values(whole)[2] == {"Attached"}
    assert P._pn_keep_values(whole)[1] == set()


def test_nuclear_words_are_per_word():
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    assert P._pn_nuclear_words(d) == frozenset({"medical", "center"})


# ── the promise ─────────────────────────────────────────────────────────────

def test_kept_inside_a_party_name_and_the_party_still_scrubbed():
    """The whole point, and the rule every other keep tier has to yield on."""
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    pz = _pz(["Mulliken Medical Center"], d)
    out = pz.apply("Defendant Mulliken Medical Center answered.")
    assert "Medical Center" in out
    assert "Mulliken" not in out


def test_kept_standing_beside_a_name():
    d = _decision("Attached", "{Attached}")
    pz = _pz(["Jack Gerlach"], d)
    out = pz.apply("Jack Gerlach Attached is the declarant.")
    assert "Attached" in out and "Gerlach" not in out


def test_an_unbraced_party_is_untouched_by_the_rule():
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    pz = _pz(["Mulliken Medical Center", "CAL EQUIPMENT FE RANCH, LLC"], d)
    out = pz.apply("See CAL EQUIPMENT FE RANCH, LLC.")
    assert "EQUIPMENT" not in out and "RANCH" not in out


def test_a_braced_word_is_never_a_bare_token():
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, d)
    terms = P._pn_build_terms(["Mulliken Medical Center"], [], [], registry=reg)
    bare = {t.real.lower() for t in terms if len(t.real.split()) == 1}
    assert "medical" not in bare and "center" not in bare


def test_a_braced_value_is_not_a_leak():
    """It is present because the operator said it can never reveal anything —
    and unlike a soft keep, that holds in the party categories too."""
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    pz = _pz(["Mulliken Medical Center"], d)
    out = pz.apply("Mulliken Medical Center answered.")
    assert pz.surviving_reals(out) == []


def test_the_built_in_furniture_list_still_applies_on_its_own():
    """Braces ADD to `_PN_FIRM_WORDS`; they never replace it."""
    pz = _pz(["Law Offices of Scott C. Stratman"], {})
    out = pz.apply("Filed by Law Offices of Scott C. Stratman.")
    assert "Law Offices of" in out and "Stratman" not in out


# ── the key ─────────────────────────────────────────────────────────────────

def test_no_key_row_promises_a_fake_that_was_never_applied(tmp_path):
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    pz = _pz(["Mulliken Medical Center"], d)
    pz.apply("Mulliken Medical Center answered.")
    key = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(key, log)
    rows = list(openpyxl.load_workbook(key).active.iter_rows(values_only=True))[1:]
    reals = {str(r[1]).lower() for r in rows}
    assert "medical" not in reals and "center" not in reals
    full = next(r for r in rows if str(r[0]) in ("person", "entity"))
    _rp = P._PN_KEY_HEADERS.index("Replacement")
    assert str(full[_rp]).endswith("Medical Center")  # reversible as delivered


def test_a_brace_corrects_a_binding_an_older_key_baked_in(tmp_path):
    """A loaded key row is applied literally, so bracing a word has to reach the
    stored fake as well — or the folder keeps producing what was braced away."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person", "Mulliken Medical Center", "Wildmere Ashcroft Dunmore",
               "replaced", "spreadsheet", 4])
    ws.append(["person-token", "Medical", "Ashcroft", "replaced", "spreadsheet", 4])
    ws.append(["person-token", "Mulliken", "Wildmere", "replaced", "spreadsheet", 4])
    wb.save(key)

    d = _decision("Mulliken Medical Center", "{Medical Center}")
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, d)
    terms, _dec = P._pn_load_key(key, reg, log)
    assert "medical" not in {t.real.lower() for t in terms}
    full = next(t for t in terms if t.category == "person")
    assert full.fake == "Wildmere Medical Center"


def test_a_brace_typed_into_the_key_is_a_keep_spec(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["entity", "Mulliken Medical Center", "{Medical Center}",
               "replaced", "spreadsheet", 4])
    wb.save(key)
    terms, decisions = P._pn_load_key(key, P._PnFakeRegistry(), log)
    assert not terms                       # the row builds no faking term
    d = decisions["mulliken medical center"]
    assert d["type"] == P._PN_KEEP_NUCLEAR_TYPE
    assert d["fake_values"] == ["Mulliken"]
    assert P._pn_decision_nuclear_parts(d) == ["Medical Center"]


def test_the_master_sheet_labels_a_brace_row(tmp_path, monkeypatch):
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master_leaks.xlsx"))
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    P._pn_update_master_keep({}, d, "26STCV00001 Test", "2026-07-31", log)
    wb = openpyxl.load_workbook(tmp_path / "master_leaks.xlsx")
    row = [r for r in wb[P._PN_MASTER_KEEP_SHEET].iter_rows(values_only=True)
           if r[0] == "Mulliken Medical Center"][0]
    assert row[1] == "{Medical Center}"     # the instruction round-trips
    assert row[2] == P._PN_KEEP_NUCLEAR_TYPE
