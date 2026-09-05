"""
The OCR FIX: `*CORRECT TEXT` in a Fix? / Replacement cell, and the alias moved
to `~`.

`*Smith` says the value is a SCAN ERROR and the text it garbled is Smith, so
the value is replaced by what the correct text becomes: Smith's own stand-in
where this case binds Smith, and the correct text itself everywhere else — a
correction says what the page meant, not that the text is a name. Every OCR-fix row goes to the key's pinned sheet, forward-only,
under Status `ocr fix`, so the macro never sees two Real Values on one
Replacement and never un-fixes a corrected word. `~Smith` is the alias that
`*Smith` used to be: two spellings of one name, the second a slip of the
first's stand-in.

Run:  cd PDF-Linker && python3 -m pytest tests/test_ocr_fix_control.py -v
"""
import inspect
import logging

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
HDR = ("Value", "Fix? (yes/no)", "Type", "Notes", "Cases", "Origin")


def _decisions(*rows):
    return P._pn_parse_decision_rows(
        [HDR] + [(v, c, "", "", "", "") for v, c in rows])


def _run(names, decisions, text):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    terms = P._pn_apply_ocr_fixes(decisions, terms, reg, log)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    return pz, pz.apply(text)


# ── the two marks ───────────────────────────────────────────────────────────

def test_the_star_is_the_ocr_fix_and_the_tilde_the_alias():
    assert P._PN_OCR_MARK == "*" and P._PN_ALIAS_MARK == "~"
    d = _decisions(("Smlth", "*Smith"), ("ANTIONO", "~ANTIONIO"))
    assert d["smlth"]["fix"] == "yes" and d["smlth"]["ocr_fix"] == "Smith"
    assert d["smlth"]["alias"] is None and d["smlth"]["replacement"] is None
    assert d["antiono"]["alias"] == "ANTIONIO" and d["antiono"]["ocr_fix"] is None
    assert P._pn_alias_target("*Smith") is None
    assert P._pn_ocr_fix_target("~Smith") is None


def test_the_target_is_read_like_an_alias_target():
    assert P._pn_ocr_fix_target("  *  John Smith ") == "John Smith"
    assert P._pn_ocr_fix_target('*"John Smith"') == "John Smith"
    assert P._pn_ocr_fix_target("*") is None
    assert P._pn_ocr_fix_target("*!!") is None


def test_the_keep_spec_form_composes_with_the_fix():
    d = _decisions(("avidsaid", "*David {said}"))["avidsaid"]
    assert d["fix"] == "yes" and d["fake_values"] == ["avid"]
    assert d["ocr_fix"] == "David" and d["alias"] is None
    assert P._pn_decision_nuclear_parts(d) == ["said"]


def test_an_ocr_fix_is_never_a_keep():
    d = _decisions(("Smlth", "*Smith"))["smlth"]
    assert not P._pn_decision_is_keep(d)


# ── what the export reads ───────────────────────────────────────────────────

def test_a_scan_error_of_a_bound_name_takes_that_names_stand_in():
    d = _decisions(("Smlth", "*Smith"))
    pz, out = _run(["John Smith"], d, "John Smith met Smlth and Smith.")
    fake = next(r["fake"] for (c, rl), r in pz.records.items()
                if c == "person-token" and rl == "smith")
    assert out.count(fake) == 3, out           # the garble reads as Smith
    assert "Smlth" not in out


def test_a_multi_word_fix_reads_word_for_word():
    d = _decisions(("Jonh Smlth", "*John Smith"))
    pz, out = _run(["John Smith"], d, "John Smith and Jonh Smlth; Smlth alone.")
    full = next(r["fake"] for (c, rl), r in pz.records.items()
                if c == "person" and rl == "john smith")
    assert out.count(full) == 2, out
    surname = full.split()[-1]
    assert out.endswith(f"; {surname} alone."), out   # the garbled token too


def test_a_scan_error_of_ordinary_text_is_corrected_verbatim():
    d = _decisions(("cuve!nants", "*covenants"))
    _pz, out = _run([], d, "The cuve!nants were read.")
    assert out == "The covenants were read."


def test_a_correct_name_the_case_has_not_bound_is_written_verbatim(caplog):
    """A correction says what the page MEANT, not that the text is a name:
    nothing is bound or minted on its account. The corrected name stands in
    the export for the scans and the worksheet to ask about as they would of
    any other text."""
    d = _decisions(("Vazqez", "*Vazquez"))
    with caplog.at_level(logging.INFO):
        pz, out = _run([], d, "Vazqez signed.")
    assert out == "Vazquez signed."
    assert not any(rl == "vazquez" for (_c, rl) in pz.records)
    assert "corrected to 'Vazquez' verbatim" in caplog.text


def test_a_fix_naming_itself_does_nothing(caplog):
    d = _decisions(("Smith", "*Smith"))
    with caplog.at_level(logging.WARNING):
        reg = P._PnFakeRegistry()
        terms = P._pn_apply_ocr_fixes(d, [], reg, log)
    assert terms == [] and "names ITSELF" in caplog.text


# ── the key ─────────────────────────────────────────────────────────────────

def test_every_ocr_fix_row_is_pinned_under_its_own_status(tmp_path):
    d = _decisions(("Smlth", "*Smith"), ("cuve!nants", "*covenants"))
    pz, _out = _run(["John Smith"], d, "Smlth read the cuve!nants with Smith.")
    key = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(key, log)
    wb = openpyxl.load_workbook(key)
    main = {r[1]: r for r in wb[P._PN_KEY_MAIN_SHEET].iter_rows(
        min_row=2, values_only=True)}
    pinned = {r[1]: r for r in wb[P._PN_KEY_PINNED_SHEET].iter_rows(
        min_row=2, values_only=True)}
    hdr = [c.value for c in wb[P._PN_KEY_MAIN_SHEET][1]]
    st = hdr.index("Status")
    assert "Smlth" not in main and "cuve!nants" not in main
    assert pinned["Smlth"][st] == P._PN_OCR_STATUS
    assert pinned["cuve!nants"][st] == P._PN_OCR_STATUS
    assert pinned["Smlth"][2] == main["Smith"][2]      # Smith's own stand-in
    assert pinned["cuve!nants"][2] == "covenants"
    # …and the garble's words were never harvested as tokens of anything.
    assert "Smlth" not in {r[1] for r in main.values()}


def test_a_pinned_ocr_row_reloads_as_a_live_fix_and_stays_pinned(tmp_path):
    d = _decisions(("Jonh Smlth", "*John Smith"), ("cuve!nants", "*covenants"))
    pz, _out = _run(["John Smith"], d, "Jonh Smlth read the cuve!nants.")
    key = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(key, log)
    reg = P._PnFakeRegistry()
    terms, kd = P._pn_load_key(key, reg, log)
    fixes = {t.real: t for t in terms if getattr(t, "ocr_fix", False)}
    assert set(fixes) >= {"Jonh Smlth", "cuve!nants"}
    assert all(t.derived and t.loaded for t in fixes.values())
    assert kd == {}
    # The memo was not seeded under the garble: it is nobody's binding.
    assert ("name_or_entity", "jonh smlth") not in reg._memo
    pz2 = P.Pseudonymizer(terms, DET, registry=reg)
    out = pz2.apply("Jonh Smlth read the cuve!nants.")
    assert "Jonh Smlth" not in out and "covenants" in out
    key2 = tmp_path / "again.xlsx"
    pz2.write_key(key2, log)
    wb = openpyxl.load_workbook(key2)
    assert "Jonh Smlth" not in {r[1] for r in wb[P._PN_KEY_MAIN_SHEET]
                                .iter_rows(min_row=2, values_only=True)}
    assert "Jonh Smlth" in {r[1] for r in wb[P._PN_KEY_PINNED_SHEET]
                            .iter_rows(min_row=2, values_only=True)}


def test_a_star_in_the_key_hands_back_an_ocr_decision(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_KEY_MAIN_SHEET
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person-token", "Smlth", "*Smith", "replaced", "spreadsheet", 3])
    wb.save(key)
    terms, kd = P._pn_load_key(key, P._PnFakeRegistry(), log)
    assert all(t.real != "Smlth" for t in terms)
    d = kd["smlth"]
    assert d["fix"] == "yes" and d["ocr_fix"] == "Smith" and d["type"] == "OCR-FIX"


# ── the passes ──────────────────────────────────────────────────────────────

def test_fix_leaks_applies_a_worksheet_fix_and_refuses_a_key_one():
    src = inspect.getsource(P._fix_leaks_mode)
    assert 'if d.get("ocr_fix")]' in src
    assert "if key_aliases or key_phrases or key_ocr:" in src
    assert "_pn_apply_ocr_fixes(" in src
    assert 'if d.get("alias") or d.get("ocr_fix"):' in src


def test_a_bound_value_is_left_alone_by_the_text_only_pass(caplog):
    d = _decisions(("Smlth", "*Smith"))
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["John Smith", "Smlth"], [], [], registry=reg)
    with caplog.at_level(logging.WARNING):
        out = P._pn_apply_ocr_fixes(d, terms, reg, log, allow_rebind=False)
    assert not any(getattr(t, "ocr_fix", False) for t in out)
    assert "Re-run PDF-Linker" in caplog.text


def test_the_worksheet_names_both_marks():
    src = inspect.getsource(P._pn_write_leak_report)
    assert "~OTHER VALUE" in src and "*CORRECT TEXT" in src
    assert P._PN_ALIAS_MARKS[0] == "~"       # the pre-fill writes a tilde
