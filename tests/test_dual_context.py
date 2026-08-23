"""The pseudonym key and the LEAKS worksheet each carry ONE Context cell
holding TWO sentences, at the owner's direction: the ORIGINAL on top, a rule,
then the sentence the export now carries — what the document said above what
the deliverable says.

One CELL and not one column each: the pair is read as a single thing, and side
by side it costs two wide columns and a lot of sideways travel. One cell and
not two ROWS for a harder reason — the key is read by other programs a row at a
time, and a second row per binding would make every one of them read a phantom
party. Where the two sentences are the same the original stands alone, which is
most rows.

On the key each half is bolded on the word it was searched by (the real value
above, the fake below); on the worksheet the flagged value stands verbatim in
both bodies, being precisely what the scrub did not replace. The pair carries
forward from the key on disk — a merged cell splits back at its rule — and
every reader resolves the column by header NAME.

Run:  cd PDF-Linker && python3 -m pytest tests/test_dual_context.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _pz(names=("Roxane Estrada",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


ORIGINAL = ("====== Page 1 ======\n"
            " 1  The process server handed the summons to Roxane Estrada at\n"
            " 2  her residence in Van Nuys on the fourth of July.\n")


def _key_cells(path):
    # BOTH sheets: a binding no export carries lives on the pinned one, and a
    # re-run that matched nothing moves every row there.
    wb = openpyxl.load_workbook(path)
    hdr, out = None, []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = hdr or [str(h) for h in rows[0]]
        out.extend(rows[1:])
    return hdr, out


def _halves(cell):
    """(original, export) as the operator sees them in one cell."""
    return P._pn_context_split(str(cell or ""))


# ── the cell itself ──────────────────────────────────────────────────────────

class TestTheCell:
    def test_the_original_is_on_top(self):
        cell = str(P._pn_context_cell("Rasho performed the work.", "Rasho",
                                      "Yeardley performed the work.",
                                      "Yeardley"))
        top, bottom = _halves(cell)
        assert top == "Rasho performed the work."
        assert bottom == "Yeardley performed the work."
        assert cell.index("Rasho") < cell.index("Yeardley")

    def test_the_two_are_separated_by_a_rule_of_their_own(self):
        cell = str(P._pn_context_cell("a Rasho sentence", "Rasho",
                                      "a Yeardley sentence", "Yeardley"))
        # A line of its own, so the cell reads as two blocks once wrapped.
        assert f"\n{P._PN_CONTEXT_RULE}\n" in cell
        assert P._PN_CONTEXT_RULE in cell.splitlines()

    def test_an_unchanged_sentence_is_not_printed_twice(self):
        # Most rows are like this: nothing else in the sentence was faked, so
        # printing it twice would be a cell of noise that hides the rows where
        # the pair really does differ.
        same = "The parties met on the fourth of July."
        cell = str(P._pn_context_cell(same, "parties", same, "parties"))
        assert cell == same
        assert P._PN_CONTEXT_RULE not in cell

    def test_a_missing_export_quote_leaves_the_original_alone(self):
        cell = str(P._pn_context_cell("only this one", "this", "", "x"))
        assert cell == "only this one"

    def test_with_no_original_the_export_stands_alone(self):
        cell = str(P._pn_context_cell("", "x", "the export's own", "export"))
        assert cell == "the export's own"
        assert P._PN_CONTEXT_RULE not in cell

    def test_each_half_bolds_the_word_it_was_searched_by(self, tmp_path):
        cell = P._pn_context_cell("Rasho performed the work.", "Rasho",
                                  "Yeardley performed the work.", "Yeardley")
        bolded = "".join(str(part) for part in cell
                         if getattr(getattr(part, "font", None), "b", False))
        assert "Rasho" in bolded and "Yeardley" in bolded
        assert "performed" not in bolded

    def test_the_split_is_the_inverse_of_the_cell(self):
        # It is what a later run reads its own key back through, so it has to
        # return exactly what went in.
        for orig, exp in (("one", "two"),
                          ("a sentence with\ninternal newlines", "b"),
                          ("only", "")):
            cell = str(P._pn_context_cell(orig, "x", exp, "y"))
            top, bottom = _halves(cell)
            assert top == orig and bottom == exp

    def test_a_cell_from_before_the_merge_is_all_original(self):
        assert _halves("just the one sentence") == ("just the one sentence", "")


# ── the key ──────────────────────────────────────────────────────────────────

class TestKeyColumn:
    def _written(self, tmp_path):
        z = _pz()
        scrubbed = z.apply(ORIGINAL)
        z.note_key_context(ORIGINAL, scrubbed)
        key = tmp_path / "pseudonym_key.xlsx"
        z.write_key(key, log)
        return z, scrubbed, _key_cells(key)

    def test_the_replacement_then_one_context_column(self, tmp_path):
        _z, _s, (hdr, _rows) = self._written(tmp_path)
        assert hdr[2] == "Replacement"
        assert hdr[3] == "Context"
        assert "Scrubbed Context" not in hdr
        # The binding leads and its evidence follows: one Context cell holding
        # both sentences, not two columns wedged between value and replacement.
        assert hdr.index("Context") > hdr.index("Replacement")

    def test_the_one_cell_shows_the_two_texts(self, tmp_path):
        z, scrubbed, (hdr, rows) = self._written(tmp_path)
        cx, rp = hdr.index("Context"), hdr.index("Replacement")
        row = next(r for r in rows if str(r[1]) == "Roxane Estrada")
        fake = str(row[rp])
        top, bottom = _halves(row[cx])
        assert "Roxane Estrada" in top          # the document's own sentence
        assert fake in bottom                   # the export's
        assert "Roxane Estrada" not in bottom
        assert fake in scrubbed                 # and it really is there

    def test_the_export_half_is_carried_forward(self, tmp_path):
        _z, _s, _cells = self._written(tmp_path)
        z2 = _pz()
        nothing = "Nothing relevant here."
        z2.note_key_context(nothing, z2.apply(nothing))
        z2.write_key(tmp_path / "pseudonym_key.xlsx", log)
        hdr, rows = _key_cells(tmp_path / "pseudonym_key.xlsx")
        row = next(r for r in rows if str(r[1]) == "Roxane Estrada")
        top, bottom = _halves(row[hdr.index("Context")])
        assert top and bottom, "a half was lost on the re-run"

    def test_a_key_written_with_two_columns_still_reads(self, tmp_path):
        # The layout it replaced: the second column is taken as it stands.
        key = tmp_path / "pseudonym_key.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Category", "Real Value", "Context", "Scrubbed Context",
                   "Replacement", "Status", "Source", "Occurrences"])
        ws.append(["person", "Roxane Estrada", "the document said",
                   "the export says", "Marlow Deverell", "replaced",
                   "spreadsheet", 1])
        wb.save(key)
        orig, scrub = P._pn_key_context_on_disk(key)
        assert orig["roxane estrada"] == "the document said"
        assert scrub["roxane estrada"] == "the export says"

    def test_the_export_half_bolds_the_fake(self, tmp_path):
        z = _pz()
        scrubbed = z.apply(ORIGINAL)
        z.note_key_context(ORIGINAL, scrubbed)
        key = tmp_path / "pseudonym_key.xlsx"
        z.write_key(key, log)
        wb = openpyxl.load_workbook(key, rich_text=True)
        hdr = [str(c.value) for c in wb.active[1]]
        cx, rp = hdr.index("Context"), hdr.index("Replacement")
        for r in wb.active.iter_rows(min_row=2):
            if str(r[1].value) == "Roxane Estrada":
                bolded = "".join(
                    str(part) for part in r[cx].value
                    if getattr(getattr(part, "font", None), "b", False))
                assert "Roxane" in bolded                     # the real value
                assert str(r[rp].value).split()[0] in bolded  # and the fake
                # The rule survives the rich-text round trip: its run carries
                # the newlines that make the cell two blocks on the page, and a
                # run whose whitespace is dropped would run them together.
                whole = "".join(str(part) for part in r[cx].value)
                assert f"\n{P._PN_CONTEXT_RULE}\n" in whole
                return
        pytest.fail("no row for the party")


# ── the worksheet ────────────────────────────────────────────────────────────

class TestLeaksColumn:
    def test_worksheet_carries_both_quotes_in_one_cell(self, tmp_path,
                                                       monkeypatch):
        pz = _pz()
        src_text = ("Ashely attended the deposition of Roxane Estrada. "
                    "The parties met afterwards.")
        monkeypatch.setattr(pz, "surviving_reals", lambda body: {"Ashely"})
        src = tmp_path / "Letter.docx"
        src.write_bytes(b"")
        assert P._write_word_text_version(src, src_text, log, pz)
        fake = pz.apply("Roxane Estrada")
        row = next(r for r in pz.leak_report if r["value"] == "Ashely")
        # The row dict still carries the two separately — only the SHEET stacks
        # them, so nothing upstream had to learn about the layout.
        assert "Roxane Estrada" in row["context"]
        assert fake in row["scrubbed_context"]
        P._pn_write_leak_report(tmp_path, pz.leak_report, log)
        wb = openpyxl.load_workbook(P._pn_leak_xlsx_path(tmp_path))
        hdr = [str(c.value) for c in wb.active[1]]
        assert hdr[2] == "Context" and "Scrubbed Context" not in hdr
        body = next(r for r in wb.active.iter_rows(min_row=2, values_only=True)
                    if str(r[0]) == "Ashely")
        top, bottom = _halves(body[2])
        assert "Roxane Estrada" in top
        assert fake in bottom
        # The flagged value is what the scrub did NOT replace, so it stands in
        # both halves — that is what makes the pair worth reading.
        assert "Ashely" in top and "Ashely" in bottom


# ── the two halves describe ONE passage ──────────────────────────────────────
# They were independent searches of two different bodies, and came apart two
# ways: a different OCCURRENCE (the first prose occurrence of a value can sit
# inside a protected citation, where it is never replaced, so the export half
# went hunting) and a different amount of GROWTH (the span grows by whole
# sentences under _PN_CONTEXT_MIN, so a fake shorter than the real value drops
# its sentence under the floor and swallows the next one).

def _key_quotes(names, body):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    z.note_key_context(body, z.apply(body))
    return z


class TestOnePassage:
    def test_a_shorter_fake_does_not_pull_in_the_next_sentence(self):
        # The sentence clears the growth floor with the real value in it and
        # falls under it with the shorter fake — so the export half used to
        # grow and the original half did not.
        z = _key_quotes(["Konstantinopoulos"],
                        "====== Page 1 ======\n"
                        " 1  Konstantinopoulos signed the lease in March 2024"
                        " and moved in.\n"
                        " 2  The rent was never paid. A notice was served.\n")
        top = z._key_context["konstantinopoulos"]
        bottom = z._key_context_scrubbed["konstantinopoulos"]
        assert "The rent was never paid" not in bottom
        assert bottom.count(".") == top.count(".")     # the same sentence

    def test_an_occurrence_the_scrub_never_touched_has_no_second_half(self):
        # "Stockton" is a party AND the first word of a cited decision. The
        # citation is byte-preserved, so nothing in that passage was replaced:
        # the honest export half is none at all, and the cell shows the
        # document's sentence alone rather than an unrelated one from later.
        z = _key_quotes(["Stockton"],
                        "====== Page 1 ======\n"
                        " 1  The court relied on Stockton Theatres, Inc. v."
                        " Palermo (1956) 47\n"
                        " 2  Cal.2d 469, 476. That case is not on point.\n"
                        " 3  Defendant Stockton signed the lease in March.\n")
        top = z._key_context["stockton"]
        assert "Stockton Theatres" in top
        assert not z._key_context_scrubbed.get("stockton")
        cell = str(P._pn_context_cell(top, "Stockton",
                                      z._key_context_scrubbed.get("stockton", ""),
                                      "x"))
        assert cell == top and P._PN_CONTEXT_RULE not in cell

    def test_the_halves_come_from_one_document(self):
        # Two documents, and only the SECOND has the party in prose the scrub
        # replaced. The original half is claimed by the first document that
        # quotes the value; the export half must not then be taken from
        # another document's export, which is how one cell came to stack two
        # sentences from two filings.
        reg = P._PnFakeRegistry()
        terms = P._pn_build_terms(["Roxane Estrada"], [], [], registry=reg)
        z = P.Pseudonymizer(terms, {}, registry=reg)
        doc_a = ("====== Page 1 ======\n"
                 " 1  Estrada v. Roxane Estrada Holdings (2019) 40 Cal.App.5th"
                 " 1, 9.\n")
        doc_b = ("====== Page 1 ======\n"
                 " 1  Roxane Estrada served the summons on the fourth of"
                 " July at noon.\n")
        z.note_key_context(doc_a, z.apply(doc_a))
        z.note_key_context(doc_b, z.apply(doc_b))
        top = z._key_context["roxane estrada"]
        bottom = z._key_context_scrubbed.get("roxane estrada", "")
        # Whatever each half is, they cannot be sentences of different files.
        if bottom:
            assert ("served the summons" in top) == ("served the summons"
                                                     in bottom)

    def test_a_site_restricts_the_search_to_its_own_lines(self):
        parsed = P._pn_body_lines(
            "====== Page 1 ======\n"
            " 1  Alpha stood here. Beta stood there.\n"
            " 2  Alpha also stood on the second line of the page.\n")
        first, site = P._pn_context_hit(parsed, "Beta")
        assert "Beta stood there" in first
        # "Alpha" is on that line too, and on the next one: the site's lines
        # are the only ones this search may use.
        again, _ = P._pn_context_hit(parsed, "Alpha", within=site)
        assert "second line" not in again
        # ...and a value that is not in the passage at all yields nothing,
        # rather than the nearest sentence somewhere else.
        assert P._pn_context_hit(parsed, "Gamma", within=site)[0] == ""

    def test_the_shape_form_keeps_the_growth_and_drops_the_lines(self):
        # What --fix-leaks passes: its originals are ONE body while each export
        # is its own, so "the same lines" means nothing there — but "the same
        # number of sentences" still does.
        site = (4, 6, 2, 1)
        assert P._pn_quote_shape(site) == (None, None, 2, 1)
        assert P._pn_quote_shape(None) is None
