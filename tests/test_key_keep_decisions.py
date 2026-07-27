"""
KEEP / KEEP-PART edits typed into the pseudonym key's Replacement column.

Sometimes the pseudonym_key.xlsx itself carries a mistake that never surfaced as
a leak, so there is no LEAKS row to correct it on. The operator can fix it in
place, in the key's Replacement column, with the same vocabulary the LEAKS Fix?
column accepts:
  * 'no'          -> leave this Real Value verbatim (do not fake it), and
  * '[bracketed]' -> keep the bracketed part verbatim, auto-fake the rest.
Both build no faking term; the decision is applied (keep-protection / fragment
faking), persisted to LEAKS.xlsx, and recorded in the master leak log.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_keep_decisions.py -v
"""
import importlib.util
import logging
import os
from pathlib import Path

import openpyxl
import pytest

_ROOT = Path(__file__).resolve().parent.parent
_spec = importlib.util.spec_from_file_location("pdf_linker", _ROOT / "pdf_linker.py")
pl = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(pl)

log = logging.getLogger("test")
DET = {k: pl._PN_DETECTORS[k] for k in pl._PN_DEFAULT_DETECTORS}
_HDR = ["Category", "Real Value", "Replacement", "Status", "Source", "Occurrences"]


def _write_key(path, rows):
    """rows: list of (category, real, replacement)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HDR)
    for cat, real, repl in rows:
        ws.append([cat, real, repl, "replaced", "spreadsheet", 1])
    wb.save(path)
    return path


# ── _pn_load_key parses the control words ────────────────────────────────────

def test_load_key_parses_no(tmp_path):
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("person", "Acme Holdings", "no")])
    reg = pl._PnFakeRegistry()
    terms, decisions = pl._pn_load_key(kp, reg, log)
    assert terms == []                                   # no faking term built
    d = decisions["acme holdings"]
    assert d["fix"] == "no" and d["type"] == "KEEP"


def test_load_key_parses_bracket(tmp_path):
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("entity", "Raytheon Human Resources", "[Human Resources]")])
    reg = pl._PnFakeRegistry()
    terms, decisions = pl._pn_load_key(kp, reg, log)
    d = decisions["raytheon human resources"]
    assert d["type"] == "KEEP-PART" and d["fix"] == "yes"
    assert d["fake_values"] == ["Raytheon"]              # the non-kept fragment


def test_load_key_normal_replacement_unchanged(tmp_path):
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("person", "Jane Doe", "Keswick Bexley")])
    reg = pl._PnFakeRegistry()
    terms, decisions = pl._pn_load_key(kp, reg, log)
    assert decisions == {}                               # not a control word
    assert any(t.real == "Jane Doe" and t.fake == "Keswick Bexley" for t in terms)


# ── keep tiers: soft `no` vs strict [bracket], and what overrides them ───────

def _apply(text, term_srcs=(), soft=(), strict=()):
    reg = pl._PnFakeRegistry()
    pz = pl.Pseudonymizer(pl._pn_build_terms(list(term_srcs), [], [], registry=reg),
                          DET, registry=reg)
    pz.keep_soft = set(soft)
    pz.keep_strict = set(strict)
    return pz.apply(text)


def test_soft_no_protects_a_standalone_word():
    out = _apply("The lot was Attached. Ramona Vale signed it.",
                 term_srcs=["Ramona Vale"], soft={"Attached"})
    assert "Attached" in out                                # standalone kept
    assert "Ramona Vale" not in out                         # party faked


def test_keep_matches_on_word_boundaries_only():
    # "Cal" kept must NOT keep "Cal" inside a larger word.
    out = _apply("Cal went to California with Calvin.", soft={"Cal"})
    assert out.count("California") == 1 and "Calvin" in out


def test_full_party_match_overrides_a_soft_keep():
    out = _apply("Cal is fine. CAL EQUIPMENT FE RANCH, LLC sued here.",
                 term_srcs=["CAL EQUIPMENT FE RANCH, LLC"], soft={"Cal"})
    assert out.startswith("Cal is fine")                    # standalone kept
    assert "CAL EQUIPMENT FE RANCH" not in out              # full party faked


def test_soft_keep_released_inside_a_detected_name_run():
    # "Cal Equipment" is NOT a key term — just a capitalized name run. A soft
    # `no` on "Cal" must not SHIELD that phrase: the keep span is produced for
    # the standalone "Cal" but withheld inside the name run (so if a term ever
    # covers it, or the review scan flags it, the keep doesn't get in the way).
    reg = pl._PnFakeRegistry()
    pz = pl.Pseudonymizer([], DET, registry=reg)
    pz.keep_soft = {"Cal"}
    text = "Cal parked outside. Cal Equipment Leasing appeared."
    protected = [text[s:e] for s, e in pz._keep_spans(text)]
    starts = [s for s, e in pz._keep_spans(text)]
    assert text.index("Cal parked") in starts               # standalone protected
    assert text.index("Cal Equipment") not in starts        # name run NOT protected


def test_soft_keep_released_inside_John_Doe():
    out = _apply("John Doe testified. Later, Doe left the room.",
                 term_srcs=["John Doe"], soft={"Doe"})
    assert "John Doe" not in out                            # phrase faked
    assert "Doe left the room" in out                       # standalone kept


def test_strict_bracket_keep_survives_next_to_a_name():
    # "[Plaintiff]" on "Plaintiff John Doe": Plaintiff kept even beside the name;
    # the name fragment is faked.
    out = _apply("Plaintiff John Doe filed. The Plaintiff appeared.",
                 term_srcs=["John Doe"], strict={"Plaintiff"})
    assert out.count("Plaintiff") == 2                      # kept both times
    assert "John Doe" not in out                            # name faked


def test_soft_keep_blocks_a_detector():
    reg = pl._PnFakeRegistry()
    pz = pl.Pseudonymizer([], DET, registry=reg)
    pz.keep_soft = {"info@acme.com"}
    assert "info@acme.com" in pz.apply("Contact info@acme.com for details.")


# ── end-to-end through the reuse path (main) with a real PDF ─────────────────

def _make_pdf(path, lines):
    import fitz
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    y = 100
    for ln in lines:
        page.insert_text((72, y), ln, fontsize=11)
        y += 24
    doc.save(path)
    doc.close()


def _run(folder, monkeypatch, key):
    import sys
    monkeypatch.setattr(sys, "argv",
                        ["pdf_linker.py", str(folder), "--key", str(key)])
    pl.main()


def test_key_no_keeps_value_in_export(tmp_path, monkeypatch):
    _make_pdf(tmp_path / "Motion.pdf",
              ["Plaintiff Acme Holdings and Jane Roe appear.",
               "Acme Holdings is a party."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Acme Holdings", "no"),
                      ("person", "Jane Roe", "Keswick Bexley")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Acme Holdings" in txt          # marked 'no' — kept verbatim
    assert "Jane Roe" not in txt           # normal binding still fakes


def test_key_no_is_written_to_master_keep_sheet(tmp_path, monkeypatch):
    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action Acme Holdings appeared through counsel and filed",
               "an opposition to the plaintiff's motion, considered on the "
               "merits."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Acme Holdings", "no")])
    # The conftest points PDF_LINKER_MASTER at a per-test tmp file.
    master = Path(os.environ["PDF_LINKER_MASTER"])
    _run(tmp_path, monkeypatch, key)

    # The durable no/bracket decision lives in the cross-folder master KEEP
    # sheet (NOT the transient per-folder LEAKS triage).
    assert master.is_file()
    wb = openpyxl.load_workbook(master)
    keep = {ws.title: ws for ws in wb.worksheets}["KEEP"]
    rows = list(keep.iter_rows(values_only=True))
    hdr = [str(h).lower() if h else "" for h in rows[0]]
    vi, fi = hdr.index("value"), next(i for i, h in enumerate(hdr)
                                      if h.startswith("fix?"))
    kept = [r for r in rows[1:] if str(r[vi]) == "Acme Holdings"]
    assert kept and str(kept[0][fi]).strip().lower() == "no"

    # ...and NOT in the per-folder LEAKS worksheet (if one was written at all).
    leaks = tmp_path / "LEAKS.xlsx"
    if leaks.is_file():
        lrows = list(openpyxl.load_workbook(leaks).active.iter_rows(values_only=True))
        assert not any(str(r[0]) == "Acme Holdings" for r in lrows[1:])


def test_key_bracket_keeps_part_fakes_rest(tmp_path, monkeypatch):
    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action the Raytheon Human Resources department "
               "responded to",
               "the plaintiff's written discovery requests within the time "
               "allowed,",
               "and produced the documents that the parties had agreed were "
               "relevant."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Raytheon Human Resources", "[Human Resources]")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Human Resources" in txt        # bracketed part kept
    assert "Raytheon" not in txt           # the rest is faked


# ── cross-folder: the KEEP sheet is a single store that learns as it goes ─────

def _pdf_lines(*lines):
    return list(lines)


def test_global_keep_applies_in_a_later_folder(tmp_path, monkeypatch):
    # Folder A declares a KEEP in its key; folder B (a different case, no such
    # key row) inherits it from the single cross-folder master KEEP sheet.
    a, b = tmp_path / "A", tmp_path / "B"
    a.mkdir(); b.mkdir()
    body = ["In this matter Globex Industries appeared through its counsel of",
            "record and answered the complaint, and the parties then met and",
            "conferred about the schedule for the remainder of the litigation."]
    _make_pdf(a / "Motion.pdf", body)
    _make_pdf(b / "Brief.pdf", body)
    ka = _write_key(a / "pseudonym_key.xlsx",
                    [("entity", "Globex Industries", "no")])
    _run(a, monkeypatch, ka)
    # B's own key does not mention Globex at all.
    kb = _write_key(b / "pseudonym_key.xlsx",
                    [("person", "Jane Roe", "Keswick Bexley")])
    _run(b, monkeypatch, kb)

    btxt = "\n".join(p.read_text() for p in (b / "Text Files").glob("*.txt"))
    assert "Globex Industries" in btxt     # inherited global KEEP applied in B

    master = Path(os.environ["PDF_LINKER_MASTER"])
    keep = {ws.title: ws for ws in openpyxl.load_workbook(master).worksheets}["KEEP"]
    rows = [r for r in keep.iter_rows(values_only=True)][1:]
    g = next(r for r in rows if str(r[0]) == "Globex Industries")
    assert int(g[3]) >= 2                   # Times Seen accumulated across A and B


def test_global_keep_yields_to_a_current_case_party(tmp_path, monkeypatch):
    # Safety: a global KEEP must never leave a REAL party of the current case in
    # the clear. If this case binds the same value to a fake, faking wins.
    master = Path(os.environ["PDF_LINKER_MASTER"])
    pl._pn_update_master_keep(
        {}, {"vertex corp": {"value": "Vertex Corp", "type": "KEEP",
                             "fix": "no", "fake_values": None,
                             "fixcell": "no", "notes": ""}},
        "Some Other Case", "2026-01-01", logging.getLogger("t"))
    assert master.is_file()

    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action Vertex Corp appeared through its counsel and",
               "answered the complaint filed against it by the plaintiff here."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Vertex Corp", "Zenith Labs")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Vertex Corp" not in txt        # a current-case party is faked anyway


# ── the re-run leaves the key CLEAN: real values and applied fakes only ───────
#
# A key row is a FULL PARTY match, and the full-party override deliberately beats
# a keep — so without retiring the row an operator's own correction misfires: a
# `no` value is faked anyway (and its row dropped, leaving an unreversible fake),
# and a keep-spec fakes the whole flagged phrase while the dirty row survives.


def _write_leaks(path, rows):
    """rows: list of (value, Fix? cell)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LEAKS"
    ws.append(list(pl._PN_LEAK_HEADERS))
    hdr = [str(h).strip().lower() for h in pl._PN_LEAK_HEADERS]
    vi = hdr.index("value")
    fi = next(i for i, h in enumerate(hdr) if h.startswith("fix?"))
    for val, fix in rows:
        r = [""] * len(hdr)
        r[vi], r[fi] = val, fix
        ws.append(r)
    wb.save(path)
    return path


def _key_rows(folder):
    """{(category, real): replacement} from the key the run just wrote."""
    wb = openpyxl.load_workbook(folder / "pseudonym_key.xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    return {(str(r[0]), str(r[1])): str(r[2]) for r in rows[1:] if r[1]}


BODY = ["John Doe's Opposition was filed by Vertex Corp counsel of record in",
        "this action, and the court considered it on the merits after the",
        "hearing, together with the reply that the moving party filed."]


def test_leaks_bracket_rewrites_the_key_row_to_the_fragment(tmp_path, monkeypatch):
    # "John Doe's Opposition" leaked and the operator bracketed the non-name
    # part. The re-run key must hold the NAME and its fake — not the flagged
    # phrase, and never the bracket text.
    _make_pdf(tmp_path / "Motion.pdf", BODY)
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("person", "John Doe's Opposition", "Yorke Deverell")])
    _write_leaks(tmp_path / "LEAKS.xlsx",
                 [("John Doe's Opposition", "['s Opposition]")])
    _run(tmp_path, monkeypatch, key)

    rows = _key_rows(tmp_path)
    assert rows[("person", "John Doe")] == "Yorke Deverell"   # fake carried over
    assert not any(r.startswith("John Doe's") for _c, r in rows)
    assert not any("[" in v or v.strip().lower() == "no" for v in rows.values())

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Yorke Deverell's Opposition" in txt   # name faked, kept part verbatim


def test_leaks_no_retires_the_key_row_and_keeps_the_value(tmp_path, monkeypatch):
    # `no` on a value the key binds: kept verbatim AND its row is gone, so the
    # key never shows a fake that was not applied.
    _make_pdf(tmp_path / "Motion.pdf", BODY)
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Vertex Corp", "Zenith Labs")])
    _write_leaks(tmp_path / "LEAKS.xlsx", [("Vertex Corp", "no")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Vertex Corp" in txt and "Zenith Labs" not in txt
    assert not any(r == "Vertex Corp" for _c, r in _key_rows(tmp_path))


def test_no_also_retires_the_derived_token_rows(tmp_path, monkeypatch):
    # write_key emits a bare token row per name word; leaving them behind would
    # reassemble the kept name word by word.
    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action John Doe appeared through counsel of record and",
               "answered the complaint filed against him by the plaintiff here."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("person", "John Doe", "Yorke Deverell"),
                      ("person-token", "John", "Yorke"),
                      ("person-token", "Doe", "Deverell")])
    _write_leaks(tmp_path / "LEAKS.xlsx", [("John Doe", "no")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "John Doe" in txt and "Yorke" not in txt and "Deverell" not in txt
    assert not any(r in ("John Doe", "John", "Doe") for _c, r in _key_rows(tmp_path))


def test_a_token_another_party_still_uses_survives_the_keep(tmp_path, monkeypatch):
    # Safety on the token retirement: "Doe" is also Jane Doe's surname here, so
    # it must stay bound even though "John Doe" is kept.
    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action John Doe and Jane Doe appeared through counsel",
               "of record and answered the complaint filed by the plaintiff."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("person", "John Doe", "Yorke Deverell"),
                      ("person", "Jane Doe", "Marlow Deverell"),
                      ("person-token", "John", "Yorke"),
                      ("person-token", "Jane", "Marlow"),
                      ("person-token", "Doe", "Deverell")])
    _write_leaks(tmp_path / "LEAKS.xlsx", [("John Doe", "no")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "John Doe" in txt                  # kept
    assert "Jane Doe" not in txt              # the other party is still faked
    rows = _key_rows(tmp_path)
    assert rows[("person", "Jane Doe")] == "Marlow Deverell"
    assert rows[("person-token", "Doe")] == "Deverell"   # still reversible


def test_inherited_keep_does_not_retire_a_current_case_key_row(tmp_path, monkeypatch):
    # The mirror of the safety rule: a keep inherited from ANOTHER case fakes the
    # current party anyway, so its key row must stay — otherwise the export
    # carries a fake nothing can reverse.
    pl._pn_update_master_keep(
        {}, {"vertex corp": {"value": "Vertex Corp", "type": "KEEP",
                             "fix": "no", "fake_values": None,
                             "fixcell": "no", "notes": ""}},
        "Some Other Case", "2026-01-01", logging.getLogger("t"))
    _make_pdf(tmp_path / "Motion.pdf",
              ["In this action Vertex Corp appeared through its counsel and",
               "answered the complaint filed against it by the plaintiff here."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("entity", "Vertex Corp", "Zenith Labs")])
    _run(tmp_path, monkeypatch, key)

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Zenith Labs" in txt
    assert _key_rows(tmp_path)[("entity", "Vertex Corp")] == "Zenith Labs"


def test_key_control_words_leave_a_clean_idempotent_key(tmp_path, monkeypatch):
    # End to end through the key's own Replacement column, twice: run 1 cleans
    # the key, run 2 (key clean, master KEEP carrying the decisions) reproduces
    # it byte for byte.
    _make_pdf(tmp_path / "Motion.pdf", BODY)
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [("person", "John Doe's Opposition", "['s Opposition]"),
                      ("entity", "Vertex Corp", "no")])
    _run(tmp_path, monkeypatch, key)
    first = _key_rows(tmp_path)
    assert ("person", "John Doe") in first
    assert not any("[" in v or v.strip().lower() == "no" for v in first.values())
    assert not any(r in ("John Doe's Opposition", "Vertex Corp")
                   for _c, r in first)

    _run(tmp_path, monkeypatch, tmp_path / "pseudonym_key.xlsx")
    assert _key_rows(tmp_path) == first

    txt = "\n".join(p.read_text() for p in (tmp_path / "Text Files").glob("*.txt"))
    assert "Vertex Corp" in txt and "John Doe" not in txt
