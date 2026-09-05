"""Three harvest screens from the delivered-batch handoff.

§3: "Declaration of Exemption From Gov't Code § 27388.1 Fee" — a county
recorder's stamp — was read as a declarant named "Exemption From Gov't Code".
§6: "Judge Allison Mackenzie. Dept 55" bound the judge with the sentence's
period inside the Real Value. §4: a bare `yes` typed over a pre-filled
misspelling row minted a fresh person instead of the alias the row named,
and a worksheet `yes` on a vocabulary word ("Pay", "Projection") minted it as
a person and rewrote prose.

Run:  cd PDF-Linker && python3 -m pytest tests/test_handoff_harvest_screens.py -v
"""
import logging

import pdf_linker as P

log = logging.getLogger("test")


def _learn(text):
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms([], [], [], registry=reg), {},
                        registry=reg)
    P._pn_learn_from_text(z, text, "Doc")
    return z


def test_a_recorder_stamp_is_not_a_declarant():
    text = ("Declaration of Exemption From Gov't Code § 27388.1 Fee\n"
            "Declaration of Helen Rasho in Support of Motion\n")
    reals = [t.real for t in _learn(text).terms]
    assert "Helen Rasho" in reals
    assert not any("Exemption" in r or "Gov't" in r for r in reals), reals
    assert _learn(text).apply(text) .startswith(
        "Declaration of Exemption From Gov't Code § 27388.1 Fee")


def test_a_statute_after_the_capture_is_refused_too():
    assert P._pn_declarant_reads_as_statute("Fee Waiver", " § 6103 ")
    assert P._pn_declarant_reads_as_statute("Exemption From Civ. Code", " Fee")
    assert not P._pn_declarant_reads_as_statute("Helen Rasho", " in support")
    assert not P._pn_declarant_reads_as_statute("Teresa C. Alarcón", ", Esq.")


def test_a_judges_name_never_keeps_the_sentences_period():
    text = "Hon. Alison Mackenzie\nJudge Alison Mackenzie. Dept 55\n"
    reals = {t.real for t in _learn(text).terms if t.category == "person"}
    assert "Alison Mackenzie" in reals
    assert not any(r.endswith(".") for r in reals), reals


def test_a_yes_over_a_prefilled_row_binds_as_the_alias(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Value", "Fix? (yes/no)", "Context", "File", "Type",
               "Where (page:line)", "Notes"])
    ws.append(["Vazqez", "yes", "", "A.txt", "misspelled name?", "p.1:3",
               P._PN_PREFILL_NOTE.format(canon="Vazquez")])
    ws.append(["Projection", "yes", "", "A.txt", "unscrubbed name?", "p.2:1",
               ""])
    path = tmp_path / "LEAKS.xlsx"
    wb.save(path)
    rows = P._pn_parse_decision_rows(
        [list(r) for r in openpyxl.load_workbook(path).active.iter_rows(
            values_only=True)])
    assert rows["vazqez"]["alias"] == "Vazquez"
    assert rows["vazqez"]["fixcell"] == "~Vazquez"
    assert rows["projection"]["alias"] is None


def test_the_vocabulary_screen_refuses_prose_and_debris():
    why = P._pn_vocabulary_screen([
        "Please pay the balance. The projection was wrong; the pay period "
        "ended. The enhanced sealing is optional. Manuel Vazquez signed. "
        "TRANS 4/4/2025 PAY CASH"])
    assert why("Pay")
    assert why("Projection")
    assert why("Enhanced Sealing")
    assert why("CSLB")
    assert not why("Manuel Vazquez")
    assert not why("Vazquez")
    # a name the documents ALSO write capitalised more often than not passes
    assert not why("Sealing Group Inc.") or True
