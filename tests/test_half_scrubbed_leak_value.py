"""
A flagged leak is routinely HALF-SCRUBBED, and marking it yes must not feed the
run's own fake back in as a REAL value.

Only the bare surname was bound ("Penuela" -> "Sable"), so the export reads
"Melissa Sable" and the review scan rightly surfaces it — "Melissa" IS a real
name still standing. Taking that phrase at face value put a fake in the key's
Real Value column and faked it a second time ("Melissa Sable" -> "Ramsey
Ellery"), so the reversal macro walked a two-generation chain that never reached
"Penuela" — and every re-run grew another generation.

Only the real remainder is faked now: "Melissa" gets its own stand-in, "Sable"
is left alone, and the pair reverses to "Melissa Penuela".

Run:  cd PDF-Linker && python3 -m pytest tests/test_half_scrubbed_leak_value.py -v
"""
import logging
import sys
import types

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")
_HDR = ["Category", "Real Value", "Replacement", "Status", "Source",
        "Occurrences"]


def _write_key(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(_HDR)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


def _write_leaks(folder, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LEAKS"
    ws.append(["File", "Type", "Value", "Where (page:line)", "Fix? (yes/no)",
               "Notes"])
    for val, fix in rows:
        ws.append(["Motion.txt.LEAK", "LEAK", val, "p.1", fix, ""])
    wb.save(folder / "LEAKS.xlsx")


def _key_rows(folder):
    ws = openpyxl.load_workbook(folder / "pseudonym_key.xlsx",
                                data_only=True).active
    return [r for r in ws.iter_rows(values_only=True)][1:]


def _setup(folder, body, key_rows, leak_rows):
    tdir = folder / "Text Files"
    tdir.mkdir()
    (tdir / "Motion.txt.LEAK").write_text(f"====== Page 1 ======\n{body}\n",
                                          encoding="utf-8")
    _write_key(folder / "pseudonym_key.xlsx", key_rows)
    _write_leaks(folder, leak_rows)
    return tdir


def _args(folder):
    return types.SimpleNamespace(term=[],
                                 key=str(folder / "pseudonym_key.xlsx"))


# ── the unit: which words of a flagged value are already ours ────────────────

def test_strip_leaves_a_value_with_nothing_of_ours_alone():
    assert P._pn_strip_prior_fakes("Melissa Penuela", {"sable"}) == "Melissa Penuela"


def test_strip_drops_the_already_faked_word():
    assert P._pn_strip_prior_fakes("Melissa Sable", {"sable"}) == "Melissa"


def test_strip_handles_punctuation_and_possessives():
    assert P._pn_strip_prior_fakes("Melissa Sable's", {"sable"}) == "Melissa"
    assert P._pn_strip_prior_fakes("Sable, Melissa", {"sable"}) == "Melissa"


def test_strip_returns_none_when_all_of_it_is_ours():
    assert P._pn_strip_prior_fakes("Ramsey Sable", {"ramsey", "sable"}) is None


def test_prior_fake_words_covers_whole_and_token():
    t = P._PnTerm("person", "Melissa Penuela", "Ramsey Sable", whole_word=False,
                  case_sensitive=False, priority=2, source="key")
    words = P._pn_prior_fake_words([t])
    assert {"ramsey sable", "ramsey", "sable"} <= words


# ── end to end through --fix-leaks, where the text is already scrubbed ───────

def test_fix_leaks_never_puts_a_fake_in_the_real_value_column(tmp_path):
    tdir = _setup(tmp_path, "Declaration of Melissa Sable in support.",
                  [["person-token", "Penuela", "Sable", "replaced", "regex", 3]],
                  [("Melissa Sable", "yes")])
    assert P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log) == 0

    rows = _key_rows(tmp_path)
    reals = {str(r[1]) for r in rows}
    fakes = {str(r[2]) for r in rows}
    assert "Melissa Sable" not in reals      # the half-scrubbed phrase is gone
    assert "Sable" not in reals              # our own fake is never a Real Value
    assert not (reals & fakes)               # no row reverses into another row
    assert ("person-token", "Penuela", "Sable") == tuple(
        str(x) for x in next(r for r in rows if str(r[1]) == "Penuela")[:3])

    body = (tdir / "Motion.txt").read_text()
    assert "Melissa" not in body             # the real given name is scrubbed
    assert "Sable" in body                   # the established stand-in is kept


def test_fix_leaks_round_trips_the_pair_back_to_the_real_name(tmp_path):
    tdir = _setup(tmp_path, "Declaration of Melissa Sable in support.",
                  [["person-token", "Penuela", "Sable", "replaced", "regex", 3]],
                  [("Melissa Sable", "yes")])
    P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)

    # Reverse the export token by token the way the DeAnonymize macro does.
    body = (tdir / "Motion.txt").read_text()
    for r in sorted(_key_rows(tmp_path), key=lambda r: -len(str(r[2]))):
        body = body.replace(str(r[2]), str(r[1]))
    assert "Melissa Penuela" in body


def test_fix_leaks_skips_a_value_that_is_entirely_our_own_fakes(tmp_path):
    tdir = _setup(tmp_path, "Declaration of Ramsey Sable in support.",
                  [["person", "Melissa Penuela", "Ramsey Sable", "replaced",
                    "regex", 3]],
                  [("Ramsey Sable", "yes")])
    P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)

    reals = {str(r[1]) for r in _key_rows(tmp_path)}
    assert "Ramsey Sable" not in reals       # never re-faked into a new generation
    assert "Melissa Penuela" in reals        # the real binding is untouched
    body = (tdir / "Motion.txt").read_text() if (tdir / "Motion.txt").is_file() \
        else (tdir / "Motion.txt.LEAK").read_text()
    assert "Ramsey Sable" in body            # left exactly as the prior run wrote it


def test_a_bracket_fragment_is_stripped_too(tmp_path):
    # "[Declaration of]" keeps the boilerplate; the fragment left to fake is
    # "Melissa Sable", which must still shed our own "Sable".
    tdir = _setup(tmp_path, "Declaration of Melissa Sable in support.",
                  [["person-token", "Penuela", "Sable", "replaced", "regex", 3]],
                  [("Declaration of Melissa Sable", "[Declaration of]")])
    P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)

    reals = {str(r[1]) for r in _key_rows(tmp_path)}
    assert "Sable" not in reals and "Melissa Sable" not in reals
    body = (tdir / "Motion.txt").read_text()
    assert "Declaration of" in body and "Sable" in body and "Melissa" not in body


# ── the full run was already safe (the pre-scan re-reads the real name) ──────

def _make_pdf(path, lines):
    import fitz
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    y = 100
    for ln in lines:
        page.insert_text((72, y), ln, fontsize=11)
        y += 24
    doc.save(path)
    doc.close()


def test_full_run_keys_the_real_name_not_the_flagged_phrase(tmp_path, monkeypatch):
    _make_pdf(tmp_path / "Motion.pdf",
              ["Declaration of Melissa Penuela in support of the opposition",
               "filed in this action by the responding party, on the merits."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx",
                     [["person-token", "Penuela", "Sable", "replaced", "regex", 3]])
    _write_leaks(tmp_path, [("Melissa Sable", "yes")])
    monkeypatch.setattr(sys, "argv",
                        ["pdf_linker.py", str(tmp_path), "--key", str(key)])
    P.main()

    rows = _key_rows(tmp_path)
    reals = {str(r[1]) for r in rows}
    assert "Melissa Penuela" in reals
    assert "Melissa Sable" not in reals and "Sable" not in reals
    assert not (reals & {str(r[2]) for r in rows})
