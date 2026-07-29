"""
A value typed into the leak worksheet's Fix? column (anything other than
yes/no) is an explicit operator-provided replacement: it is applied verbatim,
bypassing the auto-faker. The fix persists into the reversal key and survives a
re-run, so --fix-leaks converges instead of re-deriving a different fake.

Run:  cd PDF-Linker && python3 -m pytest tests/test_typed_leak_fix.py -v
"""
import logging
import types
from pathlib import Path

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _key_with_self_map(folder):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status",
               "Source", "Occurrences"])
    ws.append(["person", "M & M", "M & M", "leaked", "--term", "21"])
    wb.save(folder / "pseudonym_key.xlsx")


def _worksheet(folder, fix_cell):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Potential Leaks"
    ws.append(["File", "Type", "Value", "Where (page:line)",
               "Fix? (yes/no)", "Notes"])
    ws.append(["Motion.txt.LEAK", "LEAK", "M & M", "p.1", fix_cell, ""])
    wb.save(folder / "LEAKS.xlsx")


def _setup(folder, fix_cell):
    tdir = folder / "Text Files"
    tdir.mkdir()
    (tdir / "Motion.txt.LEAK").write_text(
        "====== Page 1 ======\nPlaintiff sues M & M; M & M denies it.\n",
        encoding="utf-8")
    _key_with_self_map(folder)
    _worksheet(folder, fix_cell)
    return tdir


def _key_rows(folder):
    ws = openpyxl.load_workbook(folder / "pseudonym_key.xlsx",
                                data_only=True).active
    return [r for r in ws.iter_rows(values_only=True)
            if r and str(r[1]).strip() == "M & M"]


def _fix_cell(folder):
    ws = openpyxl.load_workbook(folder / "LEAKS.xlsx",
                                data_only=True).active
    for r in ws.iter_rows(values_only=True):
        if r and str(r[0]).strip() == "M & M":     # Value leads, Fix? is next
            return str(r[1] or "")
    return None


def _args(folder):
    return types.SimpleNamespace(term=[], key=str(folder / "pseudonym_key.xlsx"))


def test_typed_replacement_scrubs_and_unquarantines(tmp_path):
    tdir = _setup(tmp_path, "FOXGLEN & FOXGLEN")
    rc = P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)
    assert rc == 0                                   # loop cleared
    assert not (tdir / "Motion.txt.LEAK").exists()   # un-quarantined
    body = (tdir / "Motion.txt").read_text()
    assert "M & M" not in body
    assert "FOXGLEN & FOXGLEN" in body


def test_typed_replacement_persists_in_key(tmp_path):
    _setup(tmp_path, "FOXGLEN & FOXGLEN")
    P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)
    rows = _key_rows(tmp_path)
    assert rows and rows[0][2] == "FOXGLEN & FOXGLEN"   # not the self-map


def test_typed_replacement_is_idempotent(tmp_path):
    _setup(tmp_path, "FOXGLEN & FOXGLEN")
    assert P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log) == 0
    # The leak is resolved, so the worksheet is auto-removed (nothing left to
    # triage). The typed fake is persisted in the KEY, not re-derived, so a
    # second run still converges and the binding is unchanged.
    assert not (tmp_path / "LEAKS.xlsx").exists()
    rows = _key_rows(tmp_path)
    assert rows and rows[0][2] == "FOXGLEN & FOXGLEN"
    assert P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log) == 0
    rows = _key_rows(tmp_path)
    assert rows and rows[0][2] == "FOXGLEN & FOXGLEN"   # binding still present


def test_self_identical_typed_value_is_rejected(tmp_path, caplog):
    # Typing the value back in as its own replacement can never scrub it: it is
    # ignored (with a warning), leaving nothing to apply — the file stays
    # quarantined rather than shipping with the leak.
    tdir = _setup(tmp_path, "M & M")
    with caplog.at_level(logging.WARNING):
        rc = P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)
    assert rc == 0                                   # nothing actionable
    assert (tdir / "Motion.txt.LEAK").exists()       # stays quarantined
    assert not (tdir / "Motion.txt").exists()
    assert any("equals the value itself" in m for m in caplog.messages)


def test_plain_yes_still_auto_fakes(tmp_path):
    # A normal name the faker CAN handle still works via "yes".
    folder = tmp_path
    td = folder / "Text Files"
    td.mkdir()
    (td / "Opp.txt.LEAK").write_text(
        "====== Page 1 ======\nGregory Yu appeared.\n", encoding="utf-8")
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Gregory Yu"], [], [], registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    pz.apply("Gregory Yu")
    pz.write_key(folder / "pseudonym_key.xlsx", log)
    _worksheet2 = openpyxl.Workbook()
    ws = _worksheet2.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opp.txt.LEAK", "LEAK", "Gregory Yu", "p.1", "yes", ""])
    _worksheet2.save(folder / "LEAKS.xlsx")
    rc = P._fix_leaks_mode(folder, _args(folder), {}, log)
    assert rc == 0
    body = (td / "Opp.txt").read_text()
    assert "Gregory Yu" not in body


def test_no_leaves_the_value_but_releases_the_export(tmp_path):
    # "no" = leave the value verbatim, and it never gates delivery (parity with
    # the main leak gate, which filters a suppressed value out of `gating`). So
    # nothing is scrubbed, the export is released, and the folder's leak workflow
    # is cleared — the operator has said there is nothing left to triage.
    tdir = _setup(tmp_path, "no")
    rc = P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)
    assert rc == 0
    assert not (tdir / "Motion.txt.LEAK").exists()   # released
    assert "M & M" in (tdir / "Motion.txt").read_text()   # kept verbatim
    assert not (tmp_path / "LEAKS.xlsx").exists()    # nothing left to triage


def test_decisions_parse_typed_value(tmp_path):
    # Unit-level: the parser classifies a typed value as an explicit fix.
    folder = tmp_path
    _worksheet(folder, "FOXGLEN & FOXGLEN")
    d = P._pn_read_leak_decisions(folder)["m & m"]
    assert d["fix"] == "yes"
    assert d["replacement"] == "FOXGLEN & FOXGLEN"
    _worksheet(folder, "yes")
    assert P._pn_read_leak_decisions(folder)["m & m"]["replacement"] is None
    _worksheet(folder, "no")
    assert P._pn_read_leak_decisions(folder)["m & m"]["fix"] == "no"


# ── Bracket rule: keep the bracketed PART of a leak, auto-fake the rest ───────
# A Fix? cell that brackets part of the flagged value means "keep that fragment
# verbatim; fake the remainder" — e.g. "[Human Resources]" on the leak
# "Raytheon's Human Resources" keeps the department words and fakes the name.

def test_bracket_keep_parser():
    f = P._pn_bracket_keep
    assert f("Raytheon's Human Resources", "[Human Resources]") == ["Raytheon's"]
    assert f("RESPONSE TO RAYTEHEON", "[RESPONSE TO]") == ["RAYTEHEON"]
    assert f("Board of Raytheon Directors", "[Board of] [Directors]") == ["Raytheon"]
    assert f("Foo Corp", "[Foo Corp]") == []        # whole value bracketed
    assert f("Foo Corp", "[Acme]") is None          # not a substring of the value
    assert f("Foo Corp", "yes") is None             # no brackets -> not this rule


def test_bracket_cell_reads_as_partial_keep(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Potential Leaks"
    ws.append(["File", "Type", "Value", "Where (page:line)",
               "Fix? (yes/no)", "Notes"])
    ws.append(["A.pdf", "LEAK", "Raytheon Technologies", "p.1",
               "[Technologies]", ""])
    wb.save(tmp_path / "LEAKS.xlsx")
    d = P._pn_read_leak_decisions(tmp_path)["raytheon technologies"]
    assert d["fix"] == "yes" and d["replacement"] is None
    assert d["fake_values"] == ["Raytheon"]         # only the non-kept part
    assert d["fixcell"] == "[Technologies]"         # round-trips verbatim


def test_whole_value_bracketed_keeps_everything(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Potential Leaks"
    ws.append(["File", "Type", "Value", "Where (page:line)",
               "Fix? (yes/no)", "Notes"])
    ws.append(["A.pdf", "LEAK", "Superior Court", "p.1", "[Superior Court]", ""])
    wb.save(tmp_path / "LEAKS.xlsx")
    d = P._pn_read_leak_decisions(tmp_path)["superior court"]
    assert d["fix"] == "no" and d["fake_values"] is None


def _setup_partial(folder, value, fix_cell, line):
    tdir = folder / "Text Files"
    tdir.mkdir()
    (tdir / "Brief.txt.LEAK").write_text(
        f"====== Page 1 ======\n{line}\n", encoding="utf-8")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status",
               "Source", "Occurrences"])
    ws.append(["person", "Unrelated Party", "Fake Party", "replaced", "--term", "1"])
    wb.save(folder / "pseudonym_key.xlsx")
    wb2 = openpyxl.Workbook(); w2 = wb2.active; w2.title = "Potential Leaks"
    w2.append(["File", "Type", "Value", "Where (page:line)",
               "Fix? (yes/no)", "Notes"])
    w2.append(["Brief.txt.LEAK", "LEAK", value, "p.1", fix_cell, ""])
    wb2.save(folder / "LEAKS.xlsx")
    return tdir


def test_partial_keep_fakes_only_the_unbracketed_part(tmp_path):
    tdir = _setup_partial(tmp_path, "Raytheon Technologies", "[Technologies]",
                          "Raytheon Technologies opposed the motion.")
    rc = P._fix_leaks_mode(tmp_path, _args(tmp_path), {}, log)
    assert rc == 0                                   # leak cleared, un-quarantined
    body = (tdir / "Brief.txt").read_text()
    assert "Raytheon" not in body                    # the name was faked
    assert "Technologies" in body                    # the bracketed part kept


def test_partial_keep_roundtrips_the_bracket_spec_via_master(tmp_path):
    # A bracket keep-spec is a KEEP: it lives in the cross-folder master KEEP
    # sheet (not the transient per-folder LEAKS triage), and survives back
    # verbatim — NOT collapsed to "yes" (which would fake the whole phrase, kept
    # words and all). PDF_LINKER_MASTER (set by the conftest) isolates it here.
    cfg = {}
    d = {"value": "Raytheon Technologies", "type": "KEEP-PART", "fix": "yes",
         "replacement": None, "fake_values": ["Raytheon"],
         "fixcell": "[Technologies]", "notes": ""}
    P._pn_update_master_keep(cfg, {"raytheon technologies": d},
                             "Case A", "2026-01-01", log)
    back = P._pn_read_master_keep(cfg)
    got = back["raytheon technologies"]
    assert got["fixcell"] == "[Technologies]"      # round-trips verbatim
    assert got["fake_values"] == ["Raytheon"]      # keep "Technologies", fake rest
