"""
Key reuse: a follow-up single-file run that loads the key THIS tool wrote must
reproduce the original run's fakes verbatim (so a PDF forgotten from the batch
can be scrubbed separately and stay consistent), while a genuinely new value
gets a fresh non-colliding fake and the written-back key never shrinks.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_reuse.py -v
"""
import logging

import re
import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}


def _batch(names=("Ernest N Ramirez", "Ford Motor Company",
                  "BP Ford of Long Beach"), casenos=("24STCV23198",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), list(casenos), [], registry=reg)
    return P.Pseudonymizer(terms, DET, registry=reg)


def _reuse(key_path):
    reg = P._PnFakeRegistry()
    terms, _ = P._pn_load_key(key_path, reg, log)
    z = P.Pseudonymizer(terms, DET, registry=reg)
    for r in z.records.values():
        z._own_fakes.add(str(r["fake"]).lower().rstrip(" .,;:"))
    return z


SAMPLE = ("Plaintiff Ernest N Ramirez sued Ford Motor Company. Email "
          "rlally@mortensontaggart.com. Call (626) 292-0899. Offices at 1888 "
          "Century Park East, 19th Floor, Los Angeles, CA 90067. SBN 175977. "
          "VIN 1C4JJXP65PW699184.")


def test_reused_key_reproduces_batch_output_verbatim(tmp_path):
    zb = _batch()
    zb.register_identifiers(SAMPLE)
    batch_out = zb.apply(SAMPLE)
    key = tmp_path / "pseudonym_key.xlsx"
    zb.write_key(key, log)
    assert P._pn_key_looks_like_ours(key)

    zs = _reuse(key)
    zs.register_identifiers(SAMPLE)
    assert zs.apply(SAMPLE) == batch_out


def test_reused_key_new_value_gets_fresh_noncolliding_fake(tmp_path):
    zb = _batch()
    zb.apply(SAMPLE)
    zb.register_identifiers(SAMPLE)
    key = tmp_path / "pseudonym_key.xlsx"
    zb.write_key(key, log)
    loaded_fakes = {str(r[2]).lower()
                    for r in openpyxl.load_workbook(key).active.iter_rows(
                        min_row=2, values_only=True) if r[2]}

    zs = _reuse(key)
    # A brand-new e-mail (a detector value absent from the key) is faked …
    out = zs.apply("A new witness wrote from freshdomain@newfirm.example.")
    assert "freshdomain@newfirm.example" not in out
    new_email = re.search(r"\S+@\S+", out).group(0).rstrip(".")
    # … to a value that collides with NOTHING already handed out in the batch.
    assert new_email.lower() not in loaded_fakes


def test_reused_key_writeback_never_shrinks(tmp_path):
    zb = _batch()
    zb.apply("Ernest N Ramirez v. Ford Motor Company. Email a@b.com.")
    key = tmp_path / "pseudonym_key.xlsx"
    zb.write_key(key, log)
    before = {(r[0], r[1]) for r in openpyxl.load_workbook(key).active
              .iter_rows(min_row=2, values_only=True)}

    zs = _reuse(key)
    zs.register_declarant_names("DECLARATION OF QUENTIN ZABRISKIE")
    zs.apply("I, Quentin Zabriskie, declare. Contact c@d.com.")
    zs.write_key(key, log)
    after = {(r[0], r[1]) for r in openpyxl.load_workbook(key).active
             .iter_rows(min_row=2, values_only=True)}
    assert before <= after, f"key shrank, lost {before - after}"
    assert len(after) > len(before)          # the new declarant was added


def test_ecourt_namelist_is_not_treated_as_our_key(tmp_path):
    nk = tmp_path / "Order_Template_Input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant"])
    ws.append(["24STCV23198", "Ernest N Ramirez", "Ford Motor Company"])
    wb.save(nk)
    assert not P._pn_key_looks_like_ours(nk)


def test_reused_key_loaded_party_survives_citation_only_prune(tmp_path):
    # A loaded party that happens to appear ONLY inside a citation in this one
    # file is still a known party and must not be pruned.
    zb = _batch()
    zb.apply("Ford Motor Company is the defendant.")
    key = tmp_path / "pseudonym_key.xlsx"
    zb.write_key(key, log)
    zs = _reuse(key)
    pruned = zs.prune_citation_only_terms(
        "Only as authority: Donlen v. Ford Motor Co., 217 Cal.App.4th 138 (2013).")
    assert not any("Ford" in p for p in pruned)
    assert ("entity", "ford motor company") in zs.records


def _load(key_path):
    import logging
    reg = P._PnFakeRegistry()
    terms, _ = P._pn_load_key(key_path, reg, logging.getLogger("test"))
    return sorted((t.category, t.real, t.fake) for t in terms)


def test_key_is_read_by_sheet_name_not_by_the_tab_excel_left_selected(tmp_path):
    # `wb.active` is the tab selected at the last save, not a property of the
    # key. With the pinned tab selected the loader read the pinned sheet as
    # the main one and lost every applied binding.
    import openpyxl
    key = tmp_path / "pseudonym_key.xlsx"
    pz = _batch()
    pz.apply(SAMPLE)
    pz.write_key(key, log)
    wb = openpyxl.load_workbook(key)
    assert P._PN_KEY_PINNED_SHEET in wb.sheetnames
    expected = _load(key)
    assert any(cat == "entity" for cat, _r, _f in expected)

    wb.active = wb.sheetnames.index(P._PN_KEY_PINNED_SHEET)
    wb.save(key)
    assert openpyxl.load_workbook(key).active.title == P._PN_KEY_PINNED_SHEET
    assert P._pn_key_looks_like_ours(key)
    assert _load(key) == expected


def test_a_key_whose_main_sheet_carries_an_older_title_still_resolves(tmp_path):
    # No sheet named "Pseudonym Key": the one that is not the pinned sheet is
    # the main one, whichever tab is active.
    import openpyxl
    key = tmp_path / "pseudonym_key.xlsx"
    pz = _batch()
    pz.apply(SAMPLE)
    pz.write_key(key, log)
    expected = _load(key)
    wb = openpyxl.load_workbook(key)
    wb[P._PN_KEY_MAIN_SHEET].title = "Sheet1"
    wb.active = wb.sheetnames.index(P._PN_KEY_PINNED_SHEET)
    wb.save(key)
    assert P._pn_key_looks_like_ours(key)
    assert _load(key) == expected
