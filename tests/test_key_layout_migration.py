"""The key's column ORDER, and what happens to a key written in an older one.

The binding leads: Real Value then Replacement, side by side at B and C, which
is the pair the sheet exists to state. The Context column follows at D. The
earlier layout put the two stacked sentences BETWEEN the value and its
replacement — 120 characters of quote to travel across before the row says what
it maps to.

A key already on disk in that order is not left there. Every reader resolves by
header NAME (`_pn_load_key`, `_pn_key_context_on_disk`, and `DeAnonymize.bas`'s
`LoadKeyWorkbook`), and `write_key` re-emits `_PN_KEY_HEADERS` whole — so the
rewrite at the end of a run IS the migration, on the full path and the
`--fix-leaks` path alike, carrying every binding and both Context quotes across.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_layout_migration.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

# The order this project shipped before the Replacement moved to C.
_OLD_HEADERS = ("Category", "Real Value", "Context",
                "Replacement", "Status", "Source", "Occurrences")
# Older still: before a Context column existed at all.
_ANCIENT_HEADERS = ("Category", "Real Value", "Replacement", "Status",
                    "Source", "Occurrences")


def _write_old_key(path, headers=_OLD_HEADERS, rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(list(headers))
    for r in rows or _old_rows(headers):
        ws.append(list(r))
    wb.save(path)


def _old_rows(headers):
    cells = {"Category": "entity", "Real Value": "Ford Motor Company",
             "Replacement": "Halloran Trading Company",
             "Context": "Ford Motor Company built it.",
             "Status": "replaced", "Source": "template", "Occurrences": 4}
    tok = {**cells, "Category": "entity-token", "Real Value": "Ford",
           "Replacement": "Halloran"}
    return [[c[h] for h in headers] for c in (cells, tok)]


def _sheet(path, name="Pseudonym Key"):
    return list(openpyxl.load_workbook(path)[name].iter_rows(values_only=True))


# ── the order itself ────────────────────────────────────────────────────────

def test_the_binding_leads_and_context_follows_it():
    assert P._PN_KEY_HEADERS == ("Category", "Real Value", "Replacement",
                                 "Context", "File", "Where (page:line)",
                                 "Status", "Source", "Occurrences")


def test_the_widths_stay_in_step_with_the_headers():
    """`_PN_KEY_WIDTHS` is positional, so a reordered header with a stale width
    tuple silently gives Context's 120 characters to the Replacement."""
    assert len(P._PN_KEY_WIDTHS) == len(P._PN_KEY_HEADERS)
    w = dict(zip(P._PN_KEY_HEADERS, P._PN_KEY_WIDTHS))
    assert w["Context"] == 120                 # a whole sentence, twice over
    assert w["Real Value"] == w["Replacement"]  # the pair read across


def test_a_written_key_uses_that_order(tmp_path):
    reg = P._PnFakeRegistry()
    pz = P.Pseudonymizer(
        P._pn_build_terms(["Susan Spellman"], [], [], registry=reg), [], reg)
    pz.apply("Susan Spellman signed the lease.")
    p = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(p, log)
    assert _sheet(p)[0] == P._PN_KEY_HEADERS


# ── an older layout is READ, not misread ────────────────────────────────────

@pytest.mark.parametrize("headers", [_OLD_HEADERS, _ANCIENT_HEADERS])
def test_an_older_layout_still_reads_as_ours(tmp_path, headers):
    """The fingerprint is cut to the two headers every layout has led with,
    plus a by-NAME check that a Replacement column exists — so only a POSITIONAL
    fingerprint could have broken here, and there is not one."""
    p = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(p, headers)
    assert P._pn_key_looks_like_ours(p)


@pytest.mark.parametrize("headers", [_OLD_HEADERS, _ANCIENT_HEADERS])
def test_every_binding_loads_from_an_older_layout(tmp_path, headers):
    p = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(p, headers)
    terms, _decisions = P._pn_load_key(p, P._PnFakeRegistry(), log)[:2]
    assert {(t.real, str(t.fake)) for t in terms} == {
        ("Ford Motor Company", "Halloran Trading Company"),
        ("Ford", "Halloran")}


def test_the_context_loads_from_an_older_layout(tmp_path):
    p = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(p)
    ctx, _scrubbed, _where = P._pn_key_context_on_disk(p)
    assert ctx["ford motor company"] == "Ford Motor Company built it."


def test_the_older_order_is_announced(tmp_path, caplog):
    """A column order that changes with no line in the log reads as the tool
    having damaged the key — the one file this project treats as unlosable."""
    p = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(p)
    with caplog.at_level(logging.INFO):
        P._pn_load_key(p, P._PnFakeRegistry(), log)
    assert any("older column order" in r.message for r in caplog.records)


def test_a_current_key_is_NOT_announced(tmp_path, caplog):
    reg = P._PnFakeRegistry()
    pz = P.Pseudonymizer(
        P._pn_build_terms(["Susan Spellman"], [], [], registry=reg), [], reg)
    pz.apply("Susan Spellman signed.")
    p = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(p, log)
    caplog.clear()
    with caplog.at_level(logging.INFO):
        P._pn_load_key(p, P._PnFakeRegistry(), log)
    assert not any("older column order" in r.message for r in caplog.records)


# ── ...and REWRITTEN in the current one ─────────────────────────────────────

def test_a_full_run_rewrites_an_older_layout(tmp_path):
    """`write_key` re-emits `_PN_KEY_HEADERS` whole, so re-running the folder
    normalises the layout in place and loses no binding and no quote."""
    p = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(p)
    reg = P._PnFakeRegistry()
    terms, _d = P._pn_load_key(p, reg, log)[:2]
    pz = P.Pseudonymizer(terms, [], reg)
    pz.apply("Ford Motor Company built it.")
    pz.write_key(p, log)

    rows = _sheet(p)
    assert rows[0] == P._PN_KEY_HEADERS
    col = {h: i for i, h in enumerate(P._PN_KEY_HEADERS)}
    got = {r[col["Real Value"]]: r[col["Replacement"]] for r in rows[1:]}
    assert got["Ford Motor Company"] == "Halloran Trading Company"
    assert got["Ford"] == "Halloran"
    ctx = {r[col["Real Value"]]: r[col["Context"]] for r in rows[1:]}
    assert "Ford Motor Company built it." in str(ctx["Ford Motor Company"])


def test_fix_leaks_rewrites_an_older_layout(tmp_path):
    """The other path the operator has. `--fix-leaks` never reopens the PDFs
    but it does rewrite the same key it loaded, so clicking Apply Leak Fixes
    migrates the folder too."""
    tdir = tmp_path / "Text Files"
    tdir.mkdir()
    key = tmp_path / "pseudonym_key.xlsx"
    _write_old_key(key)
    (tdir / "Opposition.txt.LEAK").write_text(
        "====== Page 1 ======\n(Yu Decl.) Gregory Yu testified.\n",
        encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opposition.txt.LEAK", "LEAK", "Gregory Yu", "p.1:2", "yes", ""])
    wb.save(tmp_path / "LEAKS.xlsx")

    class _Args:
        key = None
        term = None
        provider = "lexis"
    args = _Args()
    args.key = str(key)
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0

    rows = _sheet(key)
    assert rows[0] == P._PN_KEY_HEADERS
    col = {h: i for i, h in enumerate(P._PN_KEY_HEADERS)}
    got = {r[col["Real Value"]]: r[col["Replacement"]] for r in rows[1:]}
    # the older key's bindings survived the migration...
    assert got["Ford Motor Company"] == "Halloran Trading Company"
    # ...its Context came with them...
    ctx = {r[col["Real Value"]]: r[col["Context"]] for r in rows[1:]}
    assert "Ford Motor Company built it." in str(ctx["Ford Motor Company"])
    # ...and the value this pass fixed earned its own row.
    assert "Gregory Yu" in got


def test_the_pinned_sheet_takes_the_same_order(tmp_path):
    """A binding no export carries lives on its own sheet, and it is written by
    the same `_sheet_row` — so the two sheets can never disagree about layout."""
    reg = P._PnFakeRegistry()
    # An authoritative term that matches nothing is pinned rather than dropped.
    pz = P.Pseudonymizer(
        P._pn_build_terms(["Never Mentioned"], [], [], registry=reg), [], reg)
    pz.apply("Nothing in this text names that party.")
    p = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(p, log)
    wb = openpyxl.load_workbook(p)
    if P._PN_KEY_PINNED_SHEET in wb.sheetnames:
        assert _sheet(p, P._PN_KEY_PINNED_SHEET)[0] == P._PN_KEY_HEADERS
