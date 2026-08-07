"""
The NEVER control word: "never" in a Fix? / Replacement cell.

It is the NUCLEAR keep of the WHOLE value — exactly what `{braces}` around the
entire value mean, without re-typing the value inside them. Braces stay for the
case they are the only way to say: nuclear treatment for PART of a value (and
they still work on a single word).

Two things have to hold, and they are tested here:
  * "never" and `{whole value}` are the SAME decision, everywhere the two are
    read (LEAKS worksheet, master KEEP sheet, pseudonym key); and
  * a value marked never is really never faked — including when it is one of
    this case's own parties, which is what the operator will type it against
    most. A party match normally RELEASES a nuclear keep, on the reasoning that
    it costs nothing (the party's fake is composed with the kept word left
    verbatim) — but that holds only while some word of the party is still left
    to fake. When the keep covers the whole party, releasing it fakes the very
    value the operator said to never fake.

Run:  cd PDF-Linker && python3 -m pytest tests/test_never_keep.py -v
"""
import logging

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}


def _decision(value, cell, origin="", cases=""):
    """A decision as `_pn_parse_decision_rows` would build it from a Fix? cell."""
    rows = [("Value", "Fix? (yes/no)", "Type", "Notes", "Cases", "Origin"),
            (value, cell, "", "", cases, origin)]
    return P._pn_parse_decision_rows(rows)


def _pz(names, decisions, extra=()):
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, decisions)
    terms = P._pn_build_terms(list(names), [], list(extra), registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    pz.keep_strict, pz.keep_soft, pz.keep_nuclear = P._pn_keep_values(decisions)
    pz._keep_decisions = {vl: d for vl, d in decisions.items()
                          if P._pn_decision_is_keep(d)}
    pz._keep_local = set()          # INHERITED everywhere below, on purpose
    return pz


def _key_row(path, category, real, replacement):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append([category, real, replacement, "replaced", "spreadsheet", 4])
    wb.save(path)


# ── parsing ─────────────────────────────────────────────────────────────────

def test_never_is_the_whole_value_brace_kept():
    d = _decision("Mulliken Medical Center", "never")["mulliken medical center"]
    assert d["fix"] == "no"                    # nothing is left to fake
    assert d["fake_values"] is None
    assert P._pn_decision_nuclear_parts(d) == ["Mulliken Medical Center"]
    assert P._pn_decision_is_keep(d)


def test_never_and_a_whole_value_brace_are_the_same_decision():
    val = "Mulliken Medical Center"
    never = _decision(val, "never")
    braced = _decision(val, "{Mulliken Medical Center}")
    assert P._pn_keep_values(never) == P._pn_keep_values(braced)
    assert P._pn_nuclear_words(never) == P._pn_nuclear_words(braced)
    assert (P._pn_decision_nuclear_parts(never[val.lower()])
            == P._pn_decision_nuclear_parts(braced[val.lower()]))


def test_the_control_word_is_case_and_space_insensitive():
    for cell in ("never", "NEVER", " Never "):
        d = _decision("Attached", cell)["attached"]
        assert P._pn_decision_nuclear_parts(d) == ["Attached"], cell
    # …and it is the BARE word only: anything else is still a typed replacement.
    d = _decision("Attached", "never fake this")["attached"]
    assert d["fix"] == "yes" and d["replacement"] == "never fake this"


def test_a_never_value_is_nuclear_not_a_soft_no():
    d = _decision("Mulliken Medical Center", "never")
    strict, soft, nuclear = P._pn_keep_values(d)
    assert nuclear == {"Mulliken Medical Center"}
    assert soft == set() and strict == set()
    assert P._pn_nuclear_words(d) == frozenset({"mulliken", "medical", "center"})


# ── the promise ─────────────────────────────────────────────────────────────

def test_a_never_party_is_left_standing_whole():
    """The headline case: the operator types it against a party name."""
    d = _decision("Mulliken Medical Center", "never")
    pz = _pz(["Mulliken Medical Center"], d)
    out = pz.apply("Defendant Mulliken Medical Center answered.")
    assert out == "Defendant Mulliken Medical Center answered."
    assert pz.surviving_reals(out) == []       # present because it was told to


def test_the_whole_value_brace_now_holds_against_a_party_too():
    """Same decision, so the same outcome — the equivalence has to be real."""
    d = _decision("Mulliken Medical Center", "{Mulliken Medical Center}")
    pz = _pz(["Mulliken Medical Center"], d)
    assert "Mulliken" in pz.apply("Defendant Mulliken Medical Center answered.")


def test_every_other_party_is_still_scrubbed():
    d = _decision("Mulliken Medical Center", "never")
    pz = _pz(["Mulliken Medical Center", "CAL EQUIPMENT FE RANCH, LLC"], d)
    out = pz.apply("See CAL EQUIPMENT FE RANCH, LLC and Mulliken Medical Center.")
    assert "EQUIPMENT" not in out and "RANCH" not in out
    assert "Mulliken Medical Center" in out


def test_a_wider_party_match_still_releases_the_keep():
    """The override earns its keep wherever the party reaches BEYOND the kept
    text: "Doe" survives, and "John Doe" is still faked — composed with the
    kept word left verbatim, which is what makes the release free."""
    d = _decision("Doe", "never")
    pz = _pz(["John Doe"], d)
    out = pz.apply("Plaintiff John Doe testified. Doe was there.")
    assert "John" not in out
    assert out.count("Doe") == 2               # kept in both forms
    assert " Doe was there." in out


def test_a_braced_part_of_a_value_is_untouched_by_the_new_word():
    d = _decision("Mulliken Medical Center", "{Medical Center}")
    pz = _pz(["Mulliken Medical Center"], d)
    out = pz.apply("Defendant Mulliken Medical Center answered.")
    assert "Medical Center" in out and "Mulliken" not in out


def test_a_brace_still_works_on_a_single_word():
    d = _decision("Alder Law, P.C.", "{Law}")
    pz = _pz(["Alder Law, P.C."], d)
    out = pz.apply("Filed by Alder Law, P.C. on Monday.")
    assert "Law, P.C." in out and "Alder" not in out


def test_a_never_word_is_never_a_bare_token():
    d = _decision("Mulliken Medical Center", "never")
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, d)
    terms = P._pn_build_terms(["Mulliken Medical Center"], [], [], registry=reg)
    bare = {t.real.lower() for t in terms if len(t.real.split()) == 1}
    assert not ({"mulliken", "medical", "center"} & bare)


def test_no_key_row_promises_a_fake_that_was_never_applied(tmp_path):
    d = _decision("Mulliken Medical Center", "never")
    pz = _pz(["Mulliken Medical Center"], d)
    pz.apply("Mulliken Medical Center answered.")
    key = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(key, log)
    rows = list(openpyxl.load_workbook(key).active.iter_rows(values_only=True))[1:]
    assert not [r for r in rows
                if "mulliken" in str(r[1]).lower() or "mulliken" in str(r[2]).lower()]


# ── the key ─────────────────────────────────────────────────────────────────

def test_never_typed_into_the_key_is_a_nuclear_keep(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    _key_row(key, "entity", "Mulliken Medical Center", "never")
    reg = P._PnFakeRegistry()
    terms, decisions = P._pn_load_key(key, reg, log)
    assert not terms                            # the row builds no faking term
    d = decisions["mulliken medical center"]
    assert d["type"] == P._PN_KEEP_NUCLEAR_TYPE
    assert P._pn_decision_nuclear_parts(d) == ["Mulliken Medical Center"]
    # …and the words reached the registry as the rows were read, so no fake
    # composed from this key can carry them.
    assert {"mulliken", "medical", "center"} <= set(reg.keep_words)


def test_a_never_row_keeps_the_word_out_of_every_OTHER_rows_fake(tmp_path):
    """The circular case a brace typed into the key has: the decision is inside
    the file being read, so it must land before any row is processed."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["entity", "Medical", "never", "replaced", "spreadsheet", 4])
    ws.append(["person", "Mulliken Medical Center", "Wildmere Ashcroft Dunmore",
               "replaced", "spreadsheet", 4])
    wb.save(key)
    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    full = next(t for t in terms if t.category == "person")
    assert full.fake == "Wildmere Medical Dunmore"


# ── persistence ─────────────────────────────────────────────────────────────

def test_the_master_sheet_stores_the_word_and_reads_it_back(tmp_path, monkeypatch):
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master_leaks.xlsx"))
    d = _decision("Mulliken Medical Center", "never")
    P._pn_update_master_keep({}, d, "26STCV00001 Test", "2026-08-02", log)
    wb = openpyxl.load_workbook(tmp_path / "master_leaks.xlsx")
    row = [r for r in wb[P._PN_MASTER_KEEP_SHEET].iter_rows(values_only=True)
           if r[0] == "Mulliken Medical Center"][0]
    assert row[1] == "never"                    # the operator's own word stands
    assert row[2] == P._PN_KEEP_NUCLEAR_TYPE
    # Re-read from the sheet, the decision is the one that was written.
    back = P._pn_read_master_keep({})
    assert (P._pn_decision_nuclear_parts(back["mulliken medical center"])
            == ["Mulliken Medical Center"])


class _Args:
    key = None
    term = None


def test_fix_leaks_honours_never_and_records_it_globally(tmp_path, monkeypatch):
    """End to end through the worksheet: the flagged value is left standing, the
    export is released, and the decision lands on the cross-folder KEEP sheet as
    a KEEP-ALWAYS so every future folder honours it too."""
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master_leaks.xlsx"))
    tdir = tmp_path / "Text Files"
    tdir.mkdir()
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Ford Motor Company"], [], [], registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    pz.apply("Ford Motor Company")
    pz.write_key(tmp_path / "pseudonym_key.xlsx", log)
    (tdir / "Opposition.txt.LEAK").write_text(
        "====== Page 1 ======\nGregory Yu testified.\n", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opposition.txt.LEAK", "LEAK", "Gregory Yu", "p.1:2", "never", ""])
    wb.save(tmp_path / "LEAKS.xlsx")

    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0
    assert not (tdir / "Opposition.txt.LEAK").exists()          # released
    assert "Gregory Yu" in (tdir / "Opposition.txt").read_text()

    keep = openpyxl.load_workbook(tmp_path / "master_leaks.xlsx")[
        P._PN_MASTER_KEEP_SHEET]
    row = [r for r in keep.iter_rows(values_only=True) if r[0] == "Gregory Yu"][0]
    assert row[1] == "never" and row[2] == P._PN_KEEP_NUCLEAR_TYPE


# ── the worksheet ───────────────────────────────────────────────────────────
# A worksheet row is a QUESTION, and `never` is the strongest answer the tool
# accepts. The exact value stops being a row the moment it is decided (that is
# `_pn_decision_is_keep`, and it always worked) — but the high-recall REVIEW
# scans go on flagging every PHRASE that contains it, and each phrasing is a
# distinct value with no decision of its own, so the same kept word came back
# under a new name on every re-run.

def _pz_with_keeps(cells, records=()):
    dec = {}
    for value, cell in cells:
        dec.update(_decision(value, cell))
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, dec)
    pz = P.Pseudonymizer(list(records), {}, registry=reg)
    pz.keep_strict, pz.keep_soft, pz.keep_nuclear = P._pn_keep_values(dec)
    pz._keep_decisions = {v: d for v, d in dec.items() if P._pn_decision_is_keep(d)}
    pz._keep_local = set(dec)
    return pz


def _triage(pz, values):
    pz.leak_report = [{"file": "A.txt", "type": "REVIEW", "value": v,
                       "where": "p.1:1"} for v in values]
    pz.confirm_findings(log)
    return [r["value"] for r in pz.leak_report]


def test_a_phrase_made_only_of_kept_words_stops_being_a_row():
    pz = _pz_with_keeps([("Labor", "never"), ("Business Manager", "never")])
    assert _triage(pz, ["Business Manager", "Labor"]) == []


def test_a_real_name_beside_a_kept_word_is_asked_about_AS_FOUND():
    """The row stays, and stays VERBATIM. A kept word is the document's own
    text, so striking it out leaves a value that matches nothing in the export
    — unlocatable, and on an opaque value plainly wrong ("553.com" -> "553").
    Marking it yes is safe anyway: the composing faker keeps the kept word."""
    pz = _pz_with_keeps([("Labor", "never")])
    assert _triage(pz, ["Labor Relations Leader", "Labor Rasho"]) == [
        "Labor Relations Leader", "Labor Rasho"]


def test_an_opaque_value_is_never_cut_up():
    """The nuclear set really does hold words like "com", "www", "no" and "n",
    harvested from braced URLs and addresses in a master KEEP sheet."""
    pz = _pz_with_keeps([("www.acme.com/path", "{www.acme.com/path}")])
    assert "com" in pz.registry.keep_words
    assert _triage(pz, ["553.com", "Tol. No.", "vrv.rtxpension.com"]) == [
        "553.com", "Tol. No.", "vrv.rtxpension.com"]


def test_no_original_text_is_needed():
    """Unlike a stand-in, a kept word is the operator's own declaration — there
    is nothing to check it against, so the reduction runs with no evidence
    recorded at all."""
    pz = _pz_with_keeps([("Labor", "never")])
    assert not pz._orig_words
    assert _triage(pz, ["Labor"]) == []


def test_a_braced_keep_drops_the_same_way():
    pz = _pz_with_keeps([("Mulliken Medical Center", "{Medical Center}")])
    assert _triage(pz, ["Medical Center", "Medical Center Rasho"]) == [
        "Medical Center Rasho"]


def test_a_soft_no_keep_does_not_strip_a_phrase():
    """A `no` is released inside a name run precisely because the word may be
    part of a party there, so a phrase carrying one can still be a real leak
    and must still be asked about."""
    pz = _pz_with_keeps([("Cal", "no")])
    assert _triage(pz, ["Cal Equipment"]) == ["Cal Equipment"]


def test_the_review_list_is_cleaned_too():
    pz = _pz_with_keeps([("Labor", "never"), ("Relations", "never")])
    pz.review = [("name", "Labor Relations"), ("name", "Marcus Ingram")]
    pz.confirm_findings(log)
    assert pz.review == [("name", "Marcus Ingram")]


def test_the_exact_value_never_becomes_a_row(tmp_path):
    """The half that always worked, pinned: a decided KEEP is excluded from the
    worksheet outright (it lives on the master KEEP sheet instead)."""
    dec = _decision("Gregory Yu", "never")
    P._pn_write_leak_report(
        tmp_path, [{"file": "A.txt", "type": "LEAK", "value": "Gregory Yu",
                    "where": "p.1:1"}], log, dec, cfg={})
    assert not (tmp_path / "LEAKS.xlsx").exists()


# ─── A value that GATES delivery must always be answerable ──────────────────

def test_a_keep_spec_leak_still_earns_a_row(tmp_path):
    """A `[bracketed]` keep-spec is a KEEP, but it does not suppress the gate.

    `_pn_decision_is_keep` is true of it, so the worksheet dropped its row —
    while its `fix` is "yes", so it never reaches the gate's `suppressed` set,
    and `surviving_reals` reports a party real regardless of a keep (the
    full-party match RELEASES the keep, so a real still standing was faked
    nowhere). The operator was left with quarantined exports, no worksheet to
    answer, and — the launcher being written beside the worksheet — no
    Apply-Leak-Fixes either, while the gate's own message pointed at both.
    """
    dec = P._pn_parse_decision_rows(
        [["Type", "Value", "Fix? (yes/no)", "Notes", "Cases", "Origin"],
         ["LEAK", "Lowther Rolleston EE", "[EE]", "", "", ""]])
    d = dec["lowther rolleston ee"]
    assert P._pn_decision_is_keep(d)          # it IS a keep…
    assert d["fix"] != "no"                   # …but it does not suppress

    P._pn_write_leak_report(
        tmp_path, [{"file": "A.txt", "type": "LEAK",
                    "value": "Lowther Rolleston EE", "where": "p.1:1"}],
        log, dec, cfg={})
    xlsx = tmp_path / "LEAKS.xlsx"
    assert xlsx.exists(), "a gating value must be answerable"
    ws = openpyxl.load_workbook(xlsx).active
    vals = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert "Lowther Rolleston EE" in vals, vals


def test_a_suppressed_keep_still_keeps_no_row(tmp_path):
    # The other half, unchanged: a `no`/`never` value is in `suppressed`, so it
    # gates nothing and a row would be a question no answer could clear.
    for word in ("no", "never"):
        out = tmp_path / word
        out.mkdir()
        dec = _decision("Gregory Yu", word)
        P._pn_write_leak_report(
            out, [{"file": "A.txt", "type": "LEAK", "value": "Gregory Yu",
                   "where": "p.1:1"}], log, dec, cfg={})
        assert not (out / "LEAKS.xlsx").exists(), word


def test_a_kept_review_finding_still_keeps_no_row(tmp_path):
    # Only a LEAK is gating; a kept REVIEW-class finding is the operator's own
    # answer and stays off the worksheet.
    dec = _decision("Premises", "never")
    P._pn_write_leak_report(
        tmp_path, [{"file": "A.txt", "type": "half-scrubbed name?",
                    "value": "Premises", "where": "p.2:3"}], log, dec, cfg={})
    assert not (tmp_path / "LEAKS.xlsx").exists()
