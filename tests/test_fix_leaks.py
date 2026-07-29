"""
--fix-leaks applies the worksheet Fix?=yes decisions to the .txt/.LEAK exports
directly (no PDFs), un-quarantines files that are now clean, and preserves +
extends the key. A companion 'Apply Leak Fixes' launcher runs it on the folder.

Run:  cd PDF-Linker && python3 -m pytest tests/test_fix_leaks.py -v
"""
import logging
from pathlib import Path

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _setup(folder):
    tdir = folder / "Text Files"
    tdir.mkdir()
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Ford Motor Company"], ["24STCV24253"], [],
                              registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    pz.apply("Ford Motor Company")
    pz.write_key(folder / "pseudonym_key.xlsx", log)
    (tdir / "Opposition.txt.LEAK").write_text(
        "====== Page 1 ======\n(Yu Decl.) Gregory Yu testified.\n",
        encoding="utf-8")
    (tdir / "Motion.txt").write_text(
        "====== Page 1 ======\nGregory Yu appeared.\n", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opposition.txt.LEAK", "LEAK", "Gregory Yu", "p.1:2", "yes", ""])
    wb.save(folder / "LEAKS.xlsx")
    return tdir


class _Args:
    key = None
    term = None


def test_fix_leaks_scrubs_unquarantines_and_extends_key(tmp_path):
    tdir = _setup(tmp_path)
    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    code = P._fix_leaks_mode(tmp_path, args, {}, log)
    assert code == 0                                   # no primary leaks left

    # .LEAK un-quarantined; the flagged name scrubbed in BOTH files
    assert (tdir / "Opposition.txt").exists()
    assert not (tdir / "Opposition.txt.LEAK").exists()
    opp = (tdir / "Opposition.txt").read_text()
    assert "Gregory Yu" not in opp and "Yu Decl." not in opp
    assert "Yu" not in (tdir / "Motion.txt").read_text()

    # key preserved the party AND added the fixed name
    reals = [r[1] for r in openpyxl.load_workbook(
        tmp_path / "pseudonym_key.xlsx")["Pseudonym Key"].iter_rows(
        min_row=2, values_only=True)]
    assert "Ford Motor Company" in reals and "Gregory Yu" in reals


def test_fix_leaks_needs_a_real_key(tmp_path):
    (tmp_path / "Text Files").mkdir()
    args = _Args()
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 1   # no key -> refuse


def test_no_decision_scrubs_nothing_but_still_releases(tmp_path):
    # "no" = leave the value verbatim. There is nothing to SCRUB, but a `no`
    # never gates delivery (the same rule the main leak gate applies), so the
    # pass still runs: the export is released and the flagged name is left
    # standing in it, exactly as the operator asked. Returning early here left
    # the folder quarantined forever for a leak already dismissed — while a full
    # re-run of the same folder delivered it.
    tdir = _setup(tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "LEAKS.xlsx")
    wb.active["E2"] = "no"
    wb.save(tmp_path / "LEAKS.xlsx")
    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0
    assert not (tdir / "Opposition.txt.LEAK").exists()       # released
    assert "Gregory Yu" in (tdir / "Opposition.txt").read_text()   # kept, as asked


def test_fix_launcher_spec_windows_and_frozen():
    n, c, _ = P._fix_launcher_spec(r"C:\Py\python.exe", r"C:\T\pdf_linker.py", True)
    assert n == "Apply Leak Fixes.bat"
    assert "--fix-leaks" in c and '"%~dp0."' in c and c.endswith("pause\r\n")
    _n, cf, _e = P._fix_launcher_spec(r"C:\App\app.exe", r"C:\x.py", True, frozen=True)
    assert "x.py" not in cf and "--fix-leaks" in cf       # no script arg when frozen


# ─── Bug fixes: tool artifacts in fallback layout; NFKC-only no-op ────────────

def test_fallback_layout_ignores_tool_artifacts(tmp_path):
    # Old single-folder layout: the worksheet's .txt companion legitimately
    # CONTAINS real leaked values and must not be scanned as an export (it made
    # the run report a phantom leak and exit 2); run markers stay untouched.
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Ford Motor Company"], ["24STCV24253"], [],
                              registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    pz.apply("Ford Motor Company")
    pz.write_key(tmp_path / "pseudonym_key.xlsx", log)
    (tmp_path / "Brief.txt").write_text("Gregory Yu appeared.", encoding="utf-8")
    (tmp_path / "LEAKS.txt").write_text(
        "LEAK  Ford Motor Company  p.1:2", encoding="utf-8")
    (tmp_path / "DONE 4.25PM.txt").write_text("", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Brief.txt", "LEAK", "Gregory Yu", "p.1:1", "yes", ""])
    wb.save(tmp_path / "LEAKS.xlsx")
    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0   # no phantom leak
    assert "Yu" not in (tmp_path / "Brief.txt").read_text()
    # The run stamps its own completion: the old DONE marker is replaced by a
    # fresh 0-byte one (a marker is never scanned or scrubbed as an export).
    dones = list(tmp_path.glob("DONE *.txt"))
    assert len(dones) == 1 and dones[0].stat().st_size == 0
    assert not list(tmp_path.glob("ETA *.txt"))


def test_nfkc_only_difference_is_not_rewritten(tmp_path):
    tdir = _setup(tmp_path)
    lig = tdir / "Clean.txt"
    lig.write_text("The ﬁnding was aﬃrmed.", encoding="utf-8")  # ligatures only
    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    P._fix_leaks_mode(tmp_path, args, {}, log)
    assert lig.read_text() == "The ﬁnding was aﬃrmed."          # untouched


# ── Auto-cleanup: once every LEAK file is fixed, remove the workflow files ────
def _fixable_folder(tmp_path):
    """A folder with one quarantined export and a key+worksheet that resolve it,
    plus a stand-in Apply-Leak-Fixes launcher."""
    td = tmp_path / "Text Files"
    td.mkdir()
    (td / "Brief.txt.LEAK").write_text(
        "====== Page 1 ======\nRaytheon Technologies opposed.\n", encoding="utf-8")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status",
               "Source", "Occurrences"])
    ws.append(["person", "Filler Party", "Fake Party", "replaced", "--term", "1"])
    wb.save(tmp_path / "pseudonym_key.xlsx")
    wb2 = openpyxl.Workbook(); w2 = wb2.active; w2.title = "Potential Leaks"
    w2.append(["File", "Type", "Value", "Where (page:line)", "Fix? (yes/no)", "Notes"])
    w2.append(["Brief.txt.LEAK", "LEAK", "Raytheon Technologies", "p.1", "yes", ""])
    wb2.save(tmp_path / "LEAKS.xlsx")
    (tmp_path / "Apply Leak Fixes.command").write_text("#!/bin/sh\n", encoding="utf-8")
    return td


def _fl_args(folder):
    import types
    return types.SimpleNamespace(term=[], key=str(folder / "pseudonym_key.xlsx"))


def test_resolved_run_deletes_worksheet_and_launcher(tmp_path):
    td = _fixable_folder(tmp_path)
    rc = P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log)
    assert rc == 0
    assert not (td / "Brief.txt.LEAK").exists()          # un-quarantined
    assert not (tmp_path / "LEAKS.xlsx").exists()         # worksheet removed
    assert not (tmp_path / "Apply Leak Fixes.command").exists()   # launcher removed


def test_unresolved_run_keeps_worksheet_and_launcher(tmp_path):
    # A leak the key can't resolve (no matching term) keeps the file
    # quarantined, so the worksheet and launcher must stay.
    td = tmp_path / "Text Files"; td.mkdir()
    (td / "Brief.txt.LEAK").write_text(
        "====== Page 1 ======\nOmega Dynamics opposed.\n", encoding="utf-8")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status",
               "Source", "Occurrences"])
    ws.append(["person", "Filler Party", "Fake Party", "replaced", "--term", "1"])
    wb.save(tmp_path / "pseudonym_key.xlsx")
    wb2 = openpyxl.Workbook(); w2 = wb2.active; w2.title = "Potential Leaks"
    w2.append(["File", "Type", "Value", "Where (page:line)", "Fix? (yes/no)", "Notes"])
    # A typed replacement equal to the value is rejected — nothing gets fixed.
    w2.append(["Brief.txt.LEAK", "LEAK", "Omega Dynamics", "p.1",
               "Omega Dynamics", ""])
    wb2.save(tmp_path / "LEAKS.xlsx")
    (tmp_path / "Apply Leak Fixes.command").write_text("#!/bin/sh\n", encoding="utf-8")
    P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log)
    assert (td / "Brief.txt.LEAK").exists()               # still quarantined
    assert (tmp_path / "LEAKS.xlsx").exists()             # worksheet kept
    assert (tmp_path / "Apply Leak Fixes.command").exists()


def test_fix_leaks_writes_eta_then_done_marker(tmp_path, monkeypatch):
    # Apply Leak Fixes projects a finish time (ETA marker) up front and replaces
    # it with a DONE stamp when the pass completes — no ETA marker lingers.
    seen = {}
    real_eta = P._write_eta_marker
    monkeypatch.setattr(P, "_write_eta_marker",
                        lambda folder, label: seen.setdefault("eta", label)
                        or real_eta(folder, label))
    td = _fixable_folder(tmp_path)
    rc = P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log)
    assert rc == 0
    assert "applying leak fixes" in seen.get("eta", "")   # ETA was projected
    markers = [p.name for p in tmp_path.iterdir()
               if p.name.startswith(("ETA ", "DONE "))]
    assert any(m.startswith("DONE ") for m in markers)     # finished stamp
    assert not any(m.startswith("ETA ") for m in markers)  # none left behind


def _markers(folder):
    return [p.name for p in folder.iterdir()
            if p.name.startswith(("ETA ", "DONE "))]


def test_nothing_to_apply_still_stamps_done(tmp_path):
    # A pass with nothing to scrub is still a pass that RAN: it must replace the
    # projected-finish marker with a DONE stamp. It used to return before the
    # marker was touched, so a folder the operator had just clicked kept reading
    # "ETA ... (applying leak fixes)" with nothing working on it.
    td = _fixable_folder(tmp_path)
    (td / "Brief.txt.LEAK").rename(td / "Brief.txt")       # already resolved
    (tmp_path / "ETA ~5.00PM (applying leak fixes).txt").write_text("")
    assert P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log) == 0
    assert any(m.startswith("DONE ") for m in _markers(tmp_path))
    assert not any(m.startswith("ETA ") for m in _markers(tmp_path))
    # ...and the resolved workflow's own files are gone
    assert not (tmp_path / "LEAKS.xlsx").exists()
    assert not (tmp_path / "Apply Leak Fixes.command").exists()


def test_rejected_fix_holds_its_own_file_but_not_the_batch(tmp_path):
    # Two rows: one applies, one is a self-identical typed replacement that has
    # to be dropped. The clean file is released; the file whose own fix was
    # dropped stays quarantined, and the worksheet stays so the cell can be
    # corrected — a typo in one cell must not ride out on the other row's
    # coat-tails.
    td = tmp_path / "Text Files"
    td.mkdir()
    reg = P._PnFakeRegistry()
    pz = P.Pseudonymizer(P._pn_build_terms(["Ford Motor Company"], [], [],
                                           registry=reg), {}, registry=reg)
    pz.apply("Ford Motor Company")
    pz.write_key(tmp_path / "pseudonym_key.xlsx", log)
    (td / "Opp.txt.LEAK").write_text(
        "====== Page 1 ======\nGregory Yu testified.\n", encoding="utf-8")
    (td / "Reply.txt.LEAK").write_text(
        "====== Page 1 ======\nM & M answered.\n", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opp.txt.LEAK", "LEAK", "Gregory Yu", "p.1", "yes", ""])
    ws.append(["Reply.txt.LEAK", "LEAK", "M & M", "p.1", "M & M", ""])
    wb.save(tmp_path / "LEAKS.xlsx")
    (tmp_path / "Apply Leak Fixes.command").write_text("#!/bin/sh\n")
    P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log)
    assert not (td / "Opp.txt.LEAK").exists()          # the applied row released
    assert (td / "Reply.txt.LEAK").exists()            # the dropped row held
    assert (tmp_path / "LEAKS.xlsx").exists()          # cell still to correct
    assert (tmp_path / "Apply Leak Fixes.command").exists()


def test_rejected_fix_alone_leaves_the_folder_untouched(tmp_path):
    # Nothing applied and a decision dropped: the folder is not resolved, so the
    # quarantine, the worksheet and the launcher all stand.
    td = tmp_path / "Text Files"
    td.mkdir()
    reg = P._PnFakeRegistry()
    pz = P.Pseudonymizer(P._pn_build_terms(["Ford Motor Company"], [], [],
                                           registry=reg), {}, registry=reg)
    pz.apply("Ford Motor Company")
    pz.write_key(tmp_path / "pseudonym_key.xlsx", log)
    (td / "Reply.txt.LEAK").write_text(
        "====== Page 1 ======\nM & M answered.\n", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Reply.txt.LEAK", "LEAK", "M & M", "p.1", "M & M", ""])
    wb.save(tmp_path / "LEAKS.xlsx")
    (tmp_path / "Apply Leak Fixes.command").write_text("#!/bin/sh\n")
    assert P._fix_leaks_mode(tmp_path, _fl_args(tmp_path), {}, log) == 0
    assert (td / "Reply.txt.LEAK").exists()
    assert (tmp_path / "LEAKS.xlsx").exists()
    assert (tmp_path / "Apply Leak Fixes.command").exists()
