"""
The PHRASE control word: "phrase" in a Fix? / Replacement cell, and its PART
form `(…)`.

A `never` keeps a word verbatim EVERYWHERE — inside a party name included,
because the composing fakers consult the operator's keeps while a fake is
composed. That is what a nuclear keep is for, and it has one wrong case: a
name whose words are fine alone and identifying together. With "River" on the
master sheet as `never`, "Cross River Bank" was matched, replaced, and came
out "<fake> River Bank" — the kept word riding through inside the fake.

`phrase` on the name says: fake it WHOLE, every word, an operator keep on one
of its words notwithstanding. It governs the INSIDE of the phrase and nothing
else — a bare "River" keeps whatever rule it already had — and it is not a
KEEP: a `yes` everywhere, flagged so the registry lifts the keeps inside it.
`(Cross River Bank)` on the row "Cross River Bank Tower" is the part form, as
braces are to `never`: the parenthesised run is the phrase and the rest of the
value keeps its own rules.

Run:  cd PDF-Linker && python3 -m pytest tests/test_phrase_control.py -v
"""
import inspect
import logging
import os
from pathlib import Path

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
HDR = ("Value", "Fix? (yes/no)", "Type", "Notes", "Cases", "Origin")


def _decisions(*rows):
    """Decisions as `_pn_parse_decision_rows` builds them from Fix? cells."""
    return P._pn_parse_decision_rows(
        [HDR] + [(v, c, "", "", "", "") for v, c in rows])


def _registry(decisions):
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, decisions)
    return reg


def _pz(names, decisions):
    reg = _registry(decisions)
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    pz.keep_strict, pz.keep_soft, pz.keep_nuclear = P._pn_keep_values(decisions)
    pz._keep_decisions = {vl: d for vl, d in decisions.items()
                          if P._pn_decision_is_keep(d)}
    pz._keep_local = set()
    return pz, reg


# ── parsing ─────────────────────────────────────────────────────────────────

def test_phrase_is_a_yes_flagged_as_a_phrase_and_never_a_keep():
    d = _decisions(("Cross River Bank", "phrase"))["cross river bank"]
    assert d["fix"] == "yes"
    assert d["phrase"] is True and d["phrase_parts"] is None
    assert d["fixcell"] == "phrase" and d["replacement"] is None
    assert not P._pn_decision_is_keep(d)
    assert P._pn_decision_nuclear_parts(d) == []
    assert P._pn_decision_keep_parts(d) == []


def test_the_control_word_is_read_in_any_case():
    for cell in ("phrase", "PHRASE", " Phrase "):
        assert P._pn_is_phrase_cell(cell), cell
        assert _decisions(("Cross River Bank", cell))["cross river bank"]["phrase"]


def test_a_parenthesised_part_names_the_run_inside_the_value():
    d = _decisions(("Cross River Bank Tower", "(Cross River Bank)"))[
        "cross river bank tower"]
    assert d["fix"] == "yes" and d["phrase"] is True
    assert d["phrase_parts"] == ["Cross River Bank"]
    assert d["replacement"] is None
    assert P._pn_fake_phrases({"x": d}) == {("cross", "river", "bank")}


def test_a_part_that_is_not_in_the_value_is_the_literal_replacement_a_brace_is():
    d = _decisions(("Cross River Bank Tower", "(Cross River Bancorp)"))[
        "cross river bank tower"]
    assert d["phrase"] is False and d["replacement"] == "(Cross River Bancorp)"


def test_a_replacement_carrying_a_parenthetical_is_still_a_replacement():
    d = _decisions(("Cross River Bank", "Acme (Holdings)"))["cross river bank"]
    assert d["phrase"] is False and d["replacement"] == "Acme (Holdings)"


def test_a_one_word_value_carries_no_phrase():
    assert P._pn_phrase_tuple("River") is None
    d = _decisions(("River", "phrase"))["river"]
    assert d["fix"] == "yes"              # an ordinary yes: fake it
    assert P._pn_fake_phrases({"x": d}) == frozenset()


# ── composition: the keep is lifted inside the phrase and nowhere else ──────

def test_a_never_holds_inside_a_party_name_without_the_phrase():
    reg = _registry(_decisions(("River", "never")))
    fake = P._pn_fake_entity("Cross River Bank", reg)
    assert fake.split()[1] == "River", fake
    assert fake.split()[0] != "Cross"


def test_phrase_fakes_the_kept_word_inside_the_name():
    reg = _registry(_decisions(("River", "never"), ("Cross River Bank", "phrase")))
    fake = P._pn_fake_entity("Cross River Bank", reg)
    words = fake.split()
    assert "River" not in words and "Cross" not in words, fake
    assert words[-1] == "Bank"            # built-in furniture is not an operator keep


def test_the_lift_reaches_neither_a_bare_word_nor_another_name():
    reg = _registry(_decisions(("River", "never"), ("Cross River Bank", "phrase")))
    # A bare "River" is still the operator's keep, and so is River inside a
    # name the phrase does not cover.
    assert reg.keeps_word("river")
    other = P._pn_fake_entity("River City Holdings", reg)
    assert other.split()[0] == "River", other


def test_the_person_path_and_its_token_map_lift_the_keep_together():
    reg = _registry(_decisions(("Smith", "never"), ("John Smith", "phrase")))
    fake, _bare = P._pn_fake_person("John Smith", reg)
    assert "Smith" not in fake.split(), fake
    assert set(P._pn_person_token_map("John Smith", reg)) == {"john", "smith"}
    # Without the phrase the surname is kept and never drawn.
    reg2 = _registry(_decisions(("Smith", "never")))
    assert P._pn_fake_person("John Smith", reg2)[0].split()[-1] == "Smith"
    assert set(P._pn_person_token_map("John Smith", reg2)) == {"john"}


def test_the_part_form_lifts_the_keep_on_its_run_only():
    reg = _registry(_decisions(("River", "never"), ("Tower", "never"),
                               ("Cross River Bank Tower", "(Cross River Bank)")))
    fake = P._pn_fake_entity("Cross River Bank Tower", reg)
    words = fake.split()
    assert "River" not in words, fake      # inside the parenthesised run: faked
    assert words[-1] == "Tower"            # outside it: the keep still holds


def test_a_kept_phrase_yields_to_a_faked_one_at_the_same_positions():
    reg = P._PnFakeRegistry()
    reg.keep_phrases = frozenset({("cross", "river")})
    reg.fake_phrases = frozenset({("cross", "river", "bank")})
    bases = ["cross", "river", "bank"]
    assert reg.kept_positions(bases) == set()
    assert reg.faked_positions(bases) == {0, 1, 2}
    assert reg.furniture_positions(bases) == set()
    reg.fake_phrases = frozenset()
    assert reg.kept_positions(bases) == {0, 1}


# ── the key: furniture repair, loader, re-fake ──────────────────────────────

def test_the_furniture_repair_does_not_undo_a_phrase():
    reg = _registry(_decisions(("River", "never"), ("Cross River Bank", "phrase")))
    fake = P._pn_fake_entity("Cross River Bank", reg)
    # The repair as an older key would have been read: it puts River back…
    repaired = P._pn_restore_furniture("Cross River Bank", fake,
                                       reg.keep_words, reg.keep_phrases)
    assert repaired and repaired.split()[1] == "River"
    # …and with the phrase in hand it leaves the stored fake alone.
    assert P._pn_restore_furniture("Cross River Bank", fake, reg.keep_words,
                                   reg.keep_phrases, reg.fake_phrases) is None


def _key(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_KEY_MAIN_SHEET
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def test_a_phrase_in_the_key_builds_no_term_and_hands_the_decision_back(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    _key(key, [("entity", "Cross River Bank", "phrase", "replaced",
                "spreadsheet", 4),
               ("entity-token", "Cross", "Riverton", "replaced", "spreadsheet", 4)])
    reg = P._PnFakeRegistry()
    reg.keep_words = frozenset({"river"})
    terms, kd = P._pn_load_key(key, reg, log)
    assert all(t.real != "Cross River Bank" for t in terms)
    d = kd["cross river bank"]
    assert d["fix"] == "yes" and d["phrase"] is True
    assert d["type"] == P._PN_PHRASE_TYPE
    # The pre-scan put the phrase on the registry before any row was read.
    assert ("cross", "river", "bank") in reg.fake_phrases
    # The run site rebuilds the binding from the value with every word faked,
    # the memo keeping the token fake the delivered exports already carry.
    rebuilt = P._pn_build_terms([], [], [d["value"]], reg)
    full = next(t for t in rebuilt if t.real == "Cross River Bank")
    assert full.fake.split()[0] == "Riverton"
    assert "River" not in full.fake.split(), full.fake


def test_the_part_form_in_the_key_is_read_the_same_way(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    _key(key, [("entity", "Cross River Bank Tower", "(Cross River Bank)",
                "replaced", "spreadsheet", 4)])
    reg = P._PnFakeRegistry()
    _terms, kd = P._pn_load_key(key, reg, log)
    d = kd["cross river bank tower"]
    assert d["phrase"] is True and d["phrase_parts"] == ["Cross River Bank"]
    assert reg.fake_phrases == {("cross", "river", "bank")}


def test_a_part_the_key_value_lacks_is_warned_and_taken_literally(tmp_path, caplog):
    key = tmp_path / "pseudonym_key.xlsx"
    _key(key, [("entity", "Cross River Bank", "(Cross River Bancorp)",
                "replaced", "spreadsheet", 4)])
    reg = P._PnFakeRegistry()
    with caplog.at_level(logging.WARNING):
        terms, kd = P._pn_load_key(key, reg, log)
    assert "does not name part of that value" in caplog.text
    assert kd == {} and reg.fake_phrases == frozenset()
    assert [t.fake for t in terms if t.real == "Cross River Bank"] \
        == ["(Cross River Bancorp)"]


# ── end to end: the export, and the bare word beside it ─────────────────────

def test_the_phrase_is_faked_whole_and_the_bare_word_is_still_kept():
    decisions = _decisions(("River", "never"), ("Cross River Bank", "phrase"))
    pz, _reg = _pz(["Cross River Bank"], decisions)
    out = pz.apply("Plaintiff Cross River Bank sued. The River is wide. "
                   "CROSS RIVER BANK answered.")
    assert "Cross River Bank" not in out and "CROSS RIVER BANK" not in out
    assert "River Bank" not in out, out           # the kept word went with it
    assert "The River is wide" in out             # …and stands alone untouched


def test_a_phrase_beats_a_keep_that_contains_it():
    """`never` on "Cross River Bank Tower" with `phrase` on "Cross River Bank":
    the phrase is protective only and beats every keep it meets, so the rest of
    the kept value stays and the phrase inside it is faked — in the composed
    fake and in the export alike."""
    decisions = _decisions(("Cross River Bank Tower", "never"),
                           ("Cross River Bank", "phrase"))
    pz, reg = _pz(["Cross River Bank"], decisions)
    out = pz.apply("The Cross River Bank Tower stands tall.")
    assert "Cross River Bank" not in out, out
    assert "Bank Tower stands tall" in out
    composed = P._pn_fake_entity("Cross River Bank Tower", reg)
    assert "River" not in composed.split() and composed.endswith("Tower")
    # Without the phrase the keep holds the whole value, as `never` promises.
    pz2, _ = _pz(["Cross River Bank"], _decisions(("Cross River Bank Tower", "never")))
    assert "Cross River Bank Tower stands tall" in pz2.apply(
        "The Cross River Bank Tower stands tall.")


def test_punching_a_hole_keeps_the_pieces_around_it():
    assert P._pn_punch_spans([(0, 10), (20, 30)], [(3, 5), (8, 25), (29, 40)]) \
        == [(0, 3), (5, 8), (25, 29)]
    assert P._pn_punch_spans([(0, 10)], []) == [(0, 10)]
    assert P._pn_punch_spans([(0, 10)], [(0, 10)]) == []


# ── persistence: the master sheet, and the text-only pass ───────────────────

def test_a_phrase_round_trips_through_the_master_keep_sheet(tmp_path, monkeypatch):
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master_leaks.xlsx"))
    decisions = _decisions(("River", "never"),
                           ("Cross River Bank Tower", "(Cross River Bank)"))
    P._pn_update_master_keep({}, decisions, "26STCV00001 Test", "2026-09-05", log)
    wb = openpyxl.load_workbook(Path(os.environ["PDF_LINKER_MASTER"]))
    rows = {r[0]: r for r in wb[P._PN_MASTER_KEEP_SHEET].iter_rows(
        min_row=2, values_only=True)}
    assert rows["Cross River Bank Tower"][1] == "(Cross River Bank)"
    assert rows["Cross River Bank Tower"][2] == P._PN_PHRASE_TYPE
    back = P._pn_read_master_keep({})
    d = back["cross river bank tower"]
    assert d["phrase"] is True and d["phrase_parts"] == ["Cross River Bank"]
    # Inherited, it shapes composition and mints nothing on its own.
    reg = P._PnFakeRegistry()
    P._pn_set_keep_words(reg, back)
    assert reg.fake_phrases == {("cross", "river", "bank")}
    assert "river" in reg.keep_words


def test_fix_leaks_refuses_a_phrase_typed_into_the_key():
    """Pinned on the SOURCE: the text-only pass never reopens the PDFs, so a
    phrase in the key (which re-composes a fake the exports already carry)
    is refused beside the key alias, and the folder is left as it stands."""
    src = inspect.getsource(P._fix_leaks_mode)
    assert 'if d.get("phrase")]' in src
    assert "if key_aliases or key_phrases or key_ocr:" in src
    assert "click 'Re-run PDF-Linker' instead" in src


def test_the_worksheet_offers_the_control_word():
    src = inspect.getsource(P._pn_write_leak_report)
    assert '"yes,no,never,phrase"' in src
    assert "phrase = fake it whole" in src
