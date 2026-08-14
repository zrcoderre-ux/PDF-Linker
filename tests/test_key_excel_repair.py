"""Why Excel kept offering to REPAIR `pseudonym_key.xlsx`, in three parts.

A repaired workbook is not a cosmetic complaint. Excel repairs by DROPPING what
it could not parse, and what it drops is reversal rows — so the symptom the
operator reports as "Excel had to fix the key" is the failure this project ranks
above a leak: a pseudonym standing in a delivered export with nothing left to
undo it.

1. THE CHARACTERS openpyxl LETS THROUGH. Its own filter covers the C0 controls
   and stops there, but XML 1.0 also forbids the surrogates and U+FFFE/U+FFFF.
   openpyxl writes those verbatim and the sheet XML that comes out is not
   well-formed at all — no reader can parse it, Excel included. The key's
   Context column quotes the UNSCRUBBED body, which carries the garbled text
   layer of every page with a broken encoding, so this is exactly the text that
   produces them.

2. THE LENGTH openpyxl STOPS CHECKING. It truncates an over-long cell for you in
   `Cell._bind_value` — and skips that step entirely for a `CellRichText`, which
   is precisely what the Context column is. Nothing else stood between a quote
   and Excel's 32,767-character cell.

3. THE WRITE THAT WAS NOT ATOMIC. `wb.save(path)` truncates the destination and
   then streams a zip into it. Anything that kills the process in that window —
   an out-of-memory kill on a big folder, a full disk, a sync client in a case
   folder that is by design synced — leaves a truncated zip WHERE THE KEY USED
   TO BE. `_pn_xl_save` writes beside it, reads it back, and only then replaces,
   so a key that cannot be opened is never the one on disk.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_excel_repair.py -v
"""
import logging
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

# A lone surrogate half, and the two non-characters. Each is legal in a Python
# str, illegal in XML, and invisible to openpyxl's own filter.
SURROGATE = "\ud800"
NONCHARS = "￾￿"


def _pz(names):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    return P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)


def _sheet_xml(path):
    """Every worksheet part of `path`, as raw bytes."""
    with zipfile.ZipFile(path) as z:
        return [z.read(n) for n in z.namelist()
                if n.startswith("xl/worksheets/") and n.endswith(".xml")]


# ── 1. the characters openpyxl lets through ─────────────────────────────────

def test_openpyxls_own_filter_does_not_cover_them():
    """The reason `_PN_XL_BAD_CHARS_RE` has to exist at all.

    If openpyxl ever widens its filter this test fails, which is the moment to
    reconsider ours — not a reason to drop it silently.
    """
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    for ch in SURROGATE + NONCHARS:
        assert ILLEGAL_CHARACTERS_RE.sub("", ch) == ch


@pytest.mark.parametrize("ch", list(SURROGATE + NONCHARS))
def test_xl_text_strips_what_xml_forbids(ch):
    assert P._pn_xl_text(f"a{ch}b") == "ab"


@pytest.mark.parametrize("ch", ["�", "\t", "\n", "﷐", "\U0001fffe"])
def test_xl_text_keeps_what_xml_allows(ch):
    """Stripping more than XML forbids would quietly edit the document's text.

    U+FFFD is what a mangled decode leaves behind, U+FDD0 and the plane-end
    non-characters are legal `Char`s, and a tab is real layout. All are kept.
    """
    assert P._pn_xl_text(f"a{ch}b") == f"a{ch}b"


def test_the_key_is_well_formed_xml_with_such_a_value(tmp_path):
    """The end-to-end shape: a garbled page puts one of these in a tracked value
    AND in the sentence the key quotes for it. Before the strip, the sheet part
    would not parse — which is Excel's repair prompt, from the other side."""
    bad = f"EQUITY-WALTON_0007{SURROGATE} - HOIISING{NONCHARS}COMM"
    z = _pz(["Helen Rasho"])
    z.records[("identifier", bad.lower())] = {
        "category": "identifier", "real": bad, "fake": "DEAL-0001",
        "source": "prescan", "count": 3, "pattern": "x", "flags": 0}
    src = f"====== Page 1 ======\n 1  Helen Rasho signed {bad} today.\n"
    z.apply(src)
    z.note_key_context(src)

    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)

    parts = _sheet_xml(key)
    assert parts
    for part in parts:                       # used to raise: not well-formed
        ET.fromstring(part)
    rows = [r for ws in openpyxl.load_workbook(key).worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
    # The BINDING survives — the strip loses invisible characters, not the row.
    kept = [r[1] for r in rows if "EQUITY-WALTON" in str(r[1])]
    assert kept, rows
    assert not (set(SURROGATE + NONCHARS) & set(kept[0]))


def test_the_leaks_worksheet_is_well_formed_too(tmp_path):
    # Same sanitizer, same quote-from-the-original column, so the same hazard.
    bad = f"HOIISING{SURROGATE}COMM"
    P._pn_write_leak_report(tmp_path, [
        {"file": "M.pdf", "type": "name?", "value": bad, "where": "p.1:1",
         "context": f"The stamp reads {bad} on every page.", "notes": ""}], log)
    for part in _sheet_xml(tmp_path / "LEAKS.xlsx"):
        ET.fromstring(part)


# ── 2. the length openpyxl stops checking ───────────────────────────────────

def test_a_rich_text_cell_is_cut_to_excels_limit():
    """openpyxl truncates a plain string and NOT a `CellRichText`, so the cap
    has to be applied where both kinds of cell pass."""
    quote = "Rasho " + "x" * 60000
    cell = P._pn_rich_context(quote, "Rasho")
    assert len("".join(str(part) for part in cell)) == P._PN_XL_CELL_MAX


def test_an_over_long_quote_still_writes_a_readable_key(tmp_path):
    z = _pz(["Helen Rasho"])
    z.apply("Helen Rasho signed it.")
    z._key_context = {"helen rasho": "Helen Rasho " + "x" * 60000}
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    for ws in openpyxl.load_workbook(key).worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                assert not isinstance(cell, str) or len(cell) <= P._PN_XL_CELL_MAX


# ── 2b. the cells openpyxl retypes out from under us ────────────────────────
#
# This tool writes no formulas. openpyxl writes them anyway, from the TEXT
# alone: any string beginning with "=" becomes a formula, and any string that is
# exactly one of the seven ERROR_CODES becomes a spreadsheet error. A flagged
# value is exactly where such text turns up, because the review scans read OCR'd
# exhibits and spreadsheet exports.
#
# `=Rasho v. Smith` is written as `<f>Rasho v. Smith</f>` — not a formula Excel
# can parse, so it repairs the workbook by dropping the cell, losing the value
# the operator opened the worksheet to decide. Neither shape is caught by the
# read-back: openpyxl reads both perfectly happily. They have to be fixed on the
# way out.

RETYPED = ["=Rasho v. Smith", "=SUM(A1)", "#N/A", "#REF!", "#NAME?"]


def _cell_types(path):
    """Every cell's data type, straight out of the written file."""
    return [c.data_type
            for ws in openpyxl.load_workbook(path, read_only=True).worksheets
            for row in ws.iter_rows() for c in row]


def test_openpyxl_really_does_retype_them():
    """The reason `_pn_xl_plain_cells` exists. If openpyxl ever stops, this
    fails and the pass can be reconsidered — not dropped unnoticed."""
    ws = openpyxl.Workbook().active
    for i, v in enumerate(RETYPED, start=1):
        ws.cell(row=i, column=1, value=v)
    assert {c.data_type for c in ws["A"]} == {"f", "e"}


def test_the_leaks_worksheet_writes_them_as_text(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "M.pdf", "type": "name?", "value": v, "where": "p.1:1",
         "context": f"The column reads {v} there.", "notes": ""}
        for v in RETYPED], log)
    xlsx = tmp_path / "LEAKS.xlsx"
    assert "f" not in _cell_types(xlsx) and "e" not in _cell_types(xlsx)
    # And the value survives whole — the leading "=" included, or the operator
    # is being asked about something the document does not say.
    vals = [r[0] for r in openpyxl.load_workbook(xlsx).active
            .iter_rows(min_row=2, values_only=True)]
    assert sorted(vals) == sorted(RETYPED), vals


def test_the_key_writes_them_as_text_too(tmp_path):
    # Same hazard on the Real Value column, and there a dropped cell is a
    # binding nothing can reverse.
    z = _pz(["Helen Rasho"])
    for v in RETYPED:
        z.records[("identifier", v.lower())] = {
            "category": "identifier", "real": v, "fake": f"X{len(v)}",
            "source": "prescan", "count": 1, "pattern": "x", "flags": 0}
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    assert "f" not in _cell_types(key) and "e" not in _cell_types(key)
    reals = [r[1] for ws in openpyxl.load_workbook(key).worksheets
             for r in ws.iter_rows(min_row=2, values_only=True)]
    for v in RETYPED:
        assert v in reals, (v, reals)


def test_a_retyped_value_round_trips_through_a_rewrite(tmp_path):
    """Reading one back gives the plain string, and writing it out again must
    not retype it a second time — every write goes through the same gate."""
    xlsx = tmp_path / "LEAKS.xlsx"
    rows = [{"file": "M.pdf", "type": "name?", "value": "=Rasho v. Smith",
             "where": "p.1:1", "context": "Stamp: =Rasho v. Smith.",
             "notes": ""}]
    P._pn_write_leak_report(tmp_path, rows, log)
    P._pn_write_leak_report(tmp_path, rows, log)   # the second run re-reads it
    assert "f" not in _cell_types(xlsx)
    vals = [r[0] for r in openpyxl.load_workbook(xlsx).active
            .iter_rows(min_row=2, values_only=True)]
    assert vals == ["=Rasho v. Smith"], vals


def test_the_master_sheets_are_covered_by_the_same_gate(tmp_path, monkeypatch):
    master = tmp_path / "master.xlsx"
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    P._pn_update_master_keep(
        {}, {"=rasho v. smith": {"value": "=Rasho v. Smith", "type": "KEEP",
                                 "fix": "no", "notes": "kept"}},
        "23STCV1 Test", "2026-08-13", log)
    assert "f" not in _cell_types(master) and "e" not in _cell_types(master)


# ── 3. the write that was not atomic ────────────────────────────────────────

def _one_row_workbook(text="ok"):
    wb = openpyxl.Workbook()
    wb.active.append([text])
    return wb


def test_a_save_that_fails_leaves_the_previous_key_standing(tmp_path, monkeypatch):
    key = tmp_path / "pseudonym_key.xlsx"
    _one_row_workbook("the good key").save(key)
    before = key.read_bytes()

    monkeypatch.setattr(P, "_pn_xl_verify",
                        lambda p: (_ for _ in ()).throw(ValueError("boom")))
    with pytest.raises(OSError):
        P._pn_xl_save(_one_row_workbook("the bad key"), key, "reversal key")

    assert key.read_bytes() == before
    assert not list(tmp_path.glob("*.tmp.xlsx")), "the temp file was left behind"


def test_write_key_leaves_the_old_key_and_raises(tmp_path, monkeypatch):
    """`write_key` does not swallow it: both call sites answer with
    `_PN_KEY_LOST_MSG`, which is the one message a run cannot afford to soften.
    """
    key = tmp_path / "pseudonym_key.xlsx"
    _one_row_workbook("the good key").save(key)
    before = key.read_bytes()

    monkeypatch.setattr(P, "_pn_xl_verify",
                        lambda p: (_ for _ in ()).throw(ValueError("boom")))
    z = _pz(["Helen Rasho"])
    z.apply("Helen Rasho signed it.")
    with pytest.raises(OSError):
        z.write_key(key, log)
    assert key.read_bytes() == before


def test_verify_rejects_a_truncated_workbook(tmp_path):
    """What an out-of-memory kill or a full disk leaves behind mid-`save`."""
    good = tmp_path / "good.xlsx"
    _one_row_workbook().save(good)
    cut = tmp_path / "cut.xlsx"
    cut.write_bytes(good.read_bytes()[:len(good.read_bytes()) // 2])
    with pytest.raises(Exception):
        P._pn_xl_verify(cut)


def test_verify_accepts_what_the_tool_writes(tmp_path):
    z = _pz(["Helen Rasho"])
    src = "====== Page 1 ======\n 1  Helen Rasho signed it today.\n"
    z.apply(src)
    z.note_key_context(src)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    P._pn_xl_verify(key)                     # the key's own sheets, walked


def test_the_temp_file_keeps_the_real_extension(tmp_path, monkeypatch):
    """A plain ".tmp" cannot be re-opened — openpyxl refuses the suffix — so
    every save would have reported itself broken on the name alone."""
    seen = []
    real_save = openpyxl.Workbook.save

    def spy(self, path):
        seen.append(str(path))
        return real_save(self, path)

    monkeypatch.setattr(openpyxl.Workbook, "save", spy)
    P._pn_xl_save(_one_row_workbook(), tmp_path / "pseudonym_key.xlsx", "key")
    assert seen and seen[0].endswith(".xlsx"), seen
    assert (tmp_path / "pseudonym_key.xlsx").exists()
    assert not list(tmp_path.glob("*.tmp.xlsx"))


# ── 4. the control characters that are legal XML and still not text ─────────
# DEL and the C1 block are inside XML 1.0's `Char` production, so neither
# openpyxl's filter nor the XML-legality strip above touches them — and the file
# they land in is well-formed, so `_pn_xl_verify` reads it back happily. Excel
# never writes one raw: it escapes a control character as `_xHHHH_`. They arrive
# the way the C0 controls did, off a page whose ToUnicode is broken.

C1 = "\x7f\x80\x9f"


@pytest.mark.parametrize("ch", list(C1))
def test_a_control_character_is_legal_xml_and_still_stripped(ch):
    doc = ET.fromstring(f"<t>a{ch}b</t>")        # well-formed: no reader objects
    assert doc.text == f"a{ch}b"
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    assert ILLEGAL_CHARACTERS_RE.sub("", ch) == ch      # openpyxl lets it past
    assert P._pn_xl_text(f"a{ch}b") == "ab"             # and we do not


def test_a_control_character_does_not_reach_the_key(tmp_path):
    z = _pz(["Helen Rasho"])
    src = f"====== Page 1 ======\n 1  Helen\x9f Rasho signed\x7f it today.\n"
    z.apply(src)
    z.note_key_context(src)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    for part in _sheet_xml(key):
        assert not P._PN_XL_AUDIT_BAD_RE.search(part.decode("utf-8"))


# ── 5. the rich-text runs Excel reads differently from openpyxl ─────────────

def test_a_whitespace_only_run_never_stands_alone():
    """openpyxl's `whitespace()` helper tests the STRIPPED text for truthiness,
    so a run that is all whitespace is written without `xml:space="preserve"`
    and Excel drops its text: a quote using the value twice in a row came back
    with the words run together."""
    rich = P._pn_rich_context("Rasho Rasho performed the work.", "Rasho")
    assert "".join(str(p) for p in rich) == "Rasho Rasho performed the work."
    assert not any(str(p) and not str(p).strip() for p in rich), list(rich)


@pytest.mark.parametrize("quote,value", [
    ("Rasho Rasho performed the work.", "Rasho"),
    ("  Rasho  ", "Rasho"),
    ("RASHO, an individual. Rasho signed.", "Rasho"),
    ("Rasho", "Rasho"),
])
def test_a_rich_quote_reaches_excel_whole(tmp_path, quote, value):
    wb = openpyxl.Workbook()
    wb.active["A1"] = P._pn_rich_context(quote, value)
    path = tmp_path / "k.xlsx"
    P._pn_xl_save(wb, path, "test")
    assert P._pn_xl_audit(path) == []


def test_the_bold_span_survives_a_length_changing_lower_case():
    """`str.lower()` is not length-preserving — "İ" lowers to two code points —
    so an index taken in the folded copy and used to slice the original walked
    off by one and bolded "asho " of "Rasho"."""
    rich = P._pn_rich_context("İSTANBUL CO. retained Rasho as its expert.",
                              "Rasho")
    bold = [str(p) for p in rich if not isinstance(p, str)]
    assert bold == ["Rasho"], list(rich)


# ── 6. the witness ─────────────────────────────────────────────────────────
# Four causes have now been diagnosed from a recovery log that names a PART and
# never a cell. `_pn_xl_audit` walks the saved file with Excel's own rules and
# names sheet, cell and reason in the log.

def _rich(*parts):
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    bold = InlineFont(b=True)
    return CellRichText([p if isinstance(p, str) else TextBlock(bold, p[0])
                         for p in parts])


@pytest.mark.parametrize("value,expected", [
    ("plain text", ""),
    ("carries \x7f a control", "control character U+007F"),
    (_rich(("Rasho",), " ", ("Rasho",)), "loses its whitespace"),
    (_rich(("Rasho",), ""), "empty text run"),
    (_rich(("Rasho",), "y" * 40000), "over Excel's limit"),
])
def test_the_audit_names_the_cell_and_the_reason(tmp_path, value, expected):
    wb = openpyxl.Workbook()
    wb.active.title = "Pseudonym Key"
    wb.active["A1"] = value
    path = tmp_path / "k.xlsx"
    wb.save(path)                       # not _pn_xl_save: the audit is the point
    found = P._pn_xl_audit(path)
    if not expected:
        assert found == []
    else:
        assert any(expected in f and "Pseudonym Key!A1" in f for f in found), found


def test_the_audit_never_raises_and_never_blocks_a_save(tmp_path):
    """It reports. A cell Excel would quietly fix is not worth discarding a key
    over, and the loud failure `_pn_xl_save` reserves for an unreadable file has
    to keep meaning that."""
    assert P._pn_xl_audit(tmp_path / "does-not-exist.xlsx")[0].startswith(
        "could not be audited")
    (tmp_path / "junk.xlsx").write_bytes(b"not a zip")
    assert P._pn_xl_audit(tmp_path / "junk.xlsx")


def test_what_the_tool_actually_writes_passes_the_audit(tmp_path):
    z = _pz(["Helen Rasho", "Southern Cal Construction, Inc."])
    src = ("====== Page 1 ======\n"
           " 1  HELEN RASHO v. SOUTHERN CAL CONSTRUCTION, INC.\n"
           " 2  Rasho Rasho, and Rasho again, performed the work.\n")
    z.apply(src)
    z.note_key_context(src)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    assert P._pn_xl_audit(key) == []
