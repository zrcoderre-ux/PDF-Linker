"""A name a printed word BREAK split in half is still the party's name.

Two authors, one shape. A born-digital pleading kerns a capital pair tightly
and the character joiner reads the gap as a space; a scan drops a speck in the
middle of the word. Either way the export carries a spelling no whole-word term
matches:

    VADIM SARKISYAN and SERGEY MIRZOYANS  ->  SHARNBROOK WRIGHTSON and
                                              RESTARICK GEDNEY
    V ADIM SARKISY AN / MIRZOY ANS        ->  unchanged
    SARKISYA.N / SARKISYA,N               ->  unchanged

One Song-Beverly fee batch shipped the first pair 77 times across four exports
— captions, attorney line, proof of service, service list — while the SAME
key's tokens scrubbed those two surnames 287 and 234 times everywhere the
spelling came out whole. The complaint's first page carries both forms four
lines apart, which is worse than a plain leak: a drafting pass read them as two
different sets of plaintiffs and reported a record inconsistency.

The reduced weld pass would match either and is right to refuse
(`_pn_span_is_unbroken`); the tolerance lives in the TOKEN'S OWN PATTERN
instead, so a broken spelling is matched by a real term — yielding to citation
protection and to an operator KEEP — and reported by the scan that shares that
pattern.

A pattern and not a term per spelling. That was the first shape of this fix and
it was wrong twice over: a real value carrying a phantom space makes half a
word look like a word to every pass that decomposes one. `_trusted_party_tokens`
— which decides when a cited authority LOSES its protection — took `sar`,
`sark`, `syan`, `yan`, `vad` and `dim` as party names, and `_weld_core` reduced
"Sarkisy an" to the prefix `sarkisy` and blindly rewrote `SARKISYA.N` as
`WRIGHTSONA.N`.

Run:  cd PDF-Linker && python3 -m pytest tests/test_broken_word_name.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

PARTIES = ["Vadim Sarkisyan", "Sergey Mirzoyans"]

# The shapes the delivered exports actually carry, verbatim.
CAPTION = (
    " 5  Attorneys for Plaintiffs, V ADIM SARKISY AN and SERGEY MIRZOY ANS\n"
    "11  V ADIM SARKISY AN, an individual; Case No. 25STCV63891\n"
    "12  SERGEY MIRZOY ANS, an individual,\n"
    " 4  Primary driver or drivers of the motor vehicle V adim Sarkisyan\n"
)
# …and what its scanned exhibits carry.
EXHIBIT = (
    "CASE NAME \\/4.DifVl SARKISYA.N, et al. vs GENERAL MOTORS,\n"
    "CASE NAME' VADlM SARKISYA,N, et al. vs HARBORVIEW QUILLMARK, LLC\n"
)


def _run(names, text, registry=None):
    reg = registry if registry is not None else P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(names, [], [], registry=reg),
                        {}, registry=reg)
    return z, z.apply(text)


# ── the leak ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("leaked", [
    "V ADIM SARKISY AN", "SARKISY AN", "MIRZOY ANS", "V adim",
    "SARKISY", "MIRZOY",
])
def test_a_space_break_does_not_reach_the_export(leaked):
    _z, out = _run(PARTIES, CAPTION)
    assert leaked not in out, out


@pytest.mark.parametrize("leaked", ["SARKISYA.N", "SARKISYA,N", "SARKISYA"])
def test_a_punctuation_break_does_not_reach_the_export(leaked):
    """The scanned exhibits of the same batch. `fuzzy_survivor_scan` flagged
    these for review, which is not the same as not shipping them."""
    _z, out = _run(PARTIES, EXHIBIT)
    assert leaked not in out, out


def test_no_fragment_of_either_party_survives():
    _z, out = _run(PARTIES, CAPTION + EXHIBIT)
    for real in ("vadim", "sarkisyan", "sergey", "mirzoyans",
                 "adim", "sarkisy", "mirzoy", "sarkisya"):
        assert real not in out.lower(), (real, out)


def test_the_broken_spelling_and_the_whole_one_are_ONE_party():
    """The half-scrubbed caption is what made a drafting pass report two
    different sets of plaintiffs."""
    _z, broken = _run(PARTIES, "V ADIM SARKISY AN and SERGEY MIRZOY ANS")
    _z2, whole = _run(PARTIES, "VADIM SARKISYAN and SERGEY MIRZOYANS")
    assert broken == whole


def test_a_spelling_that_never_broke_is_unchanged():
    reg = P._PnFakeRegistry()
    z, out = _run(PARTIES, "VADIM SARKISYAN, an individual;", registry=reg)
    fake = z.records[("person", "vadim sarkisyan")]["fake"]
    assert out == f"{fake.upper()}, an individual;"


def test_our_own_output_is_left_alone():
    z, _out = _run(PARTIES, CAPTION)
    fake = z.records[("person", "vadim sarkisyan")]["fake"]
    _z2, again = _run(PARTIES, f"{fake.upper()}, an individual;")
    assert again == f"{fake.upper()}, an individual;"


# ── what it costs ───────────────────────────────────────────────────────────

def _plain(monkeypatch):
    """The same build with the tolerance switched off."""
    monkeypatch.setattr(P, "_pn_term_is_breakable", lambda cat, src: False)


def test_it_mints_nothing_moves_nothing_and_adds_no_term(monkeypatch):
    """The tolerance is a pattern, so there is no new value: the term set, the
    used pool and every binding must come out identical. A fake that moved
    would re-scrub a folder already sent out under the old one."""
    def build():
        reg = P._PnFakeRegistry()
        terms = P._pn_build_terms(PARTIES, ["25STCV63891"], [], registry=reg)
        return (sorted(reg._used),
                sorted((t.category, t.real, t.fake) for t in terms))

    _plain(monkeypatch)
    used_before, terms_before = build()
    monkeypatch.undo()
    used_after, terms_after = build()
    assert used_before == used_after
    assert terms_before == terms_after


def test_ordinary_legal_prose_is_untouched():
    prose = (
        "Plaintiffs are informed and believe, and thereon allege, that in "
        "doing the things hereinafter alleged Defendants, and each of them, "
        "were acting in the course and scope of their employment as such "
        "agents, servants, and/or employees, and with the permission, "
        "consent, knowledge, and/or ratification of their Co-Defendants, "
        "principals, and/or employers. The Vehicle was/is a \"new motor "
        "vehicle\" under The Song-Beverly Warranty Act. As he was not a party "
        "to the within action, service was made in the manner ascribed.")
    _z, out = _run(PARTIES, prose)
    assert out == prose


def test_a_lower_case_occurrence_is_left_alone():
    """A bare token never matches a lower-case occurrence
    (`_pn_term_is_cap_only`) — the break tolerance does not change that."""
    _z, out = _run(PARTIES, "the sarkisy an filing")
    assert out == "the sarkisy an filing"


# ── the guards ──────────────────────────────────────────────────────────────

def test_a_middle_INITIAL_is_not_a_broken_word():
    """"Debora H. Smith" is how a filing writes a middle initial. A SPACE
    before a single letter is therefore never a break — only a mark is."""
    _z, out = _run(["Deborah Rasho", "Vadim Sarkisyan"],
                   "Served on Debora H. Smith and Sarkisya N. Doe at home.")
    assert out == "Served on Debora H. Smith and Sarkisya N. Doe at home."


@pytest.mark.parametrize("text", [
    # A mark followed by a SPACE is a boundary, not a break: a sentence end,
    # and — with a single letter after it — the surname-first roster shape a
    # docket really prints ("Sarkisya, N.").
    "The Sarkisya. N was never served.",
    "Sarkisya, N. appeared by counsel.",
    # …and the whole-word right edge still holds, so a break spelling cannot
    # run into the word after it.
    "The Sarkisya.Nothing clause is inapt.",
])
def test_a_boundary_is_never_read_as_a_break(text):
    _z, out = _run(PARTIES, text)
    assert out == text


def test_a_mark_with_no_space_around_it_IS_a_break():
    """Which is the whole difference: `SARKISYA.N` off a scanned exhibit."""
    _z, out = _run(PARTIES, "CASE NAME SARKISYA.N, et al.")
    assert "SARKISYA" not in out, out


def test_a_short_token_is_never_broken():
    for short in ("Kim", "Wu", "Ng", "Doe"):
        assert P._pn_word_breaks(short) == []


def test_a_one_letter_LEADING_half_is_kept():
    """It is where the observed breaks fell — "V ADIM", "V adim"."""
    assert ("V", "adim", P._PN_WORD_BREAK) in P._pn_word_breaks("Vadim")


def test_a_one_letter_TRAILING_half_takes_a_mark_only():
    breaks = {(l, r): b for l, r, b in P._pn_word_breaks("Sarkisyan")}
    assert breaks[("Sarkisya", "n")] == P._PN_WORD_BREAK_MARK
    assert breaks[("Sarkisy", "an")] == P._PN_WORD_BREAK


def test_two_ordinary_words_are_not_a_break():
    """"Nowhere" would offer "No where", which a sentence writes by accident.
    The concatenation is the corroboration everywhere else; where both halves
    are ordinary vocabulary it corroborates nothing."""
    assert not any(l == "No" for l, _r, _b in P._pn_word_breaks("Nowhere"))
    assert not any(l == "In" for l, _r, _b in P._pn_word_breaks("Inaction"))


def test_the_break_is_between_letters():
    """A hyphen or an inner dot is already a printed boundary — the wrap-split
    spelling of a compound surname has its own registration."""
    for left, right, _b in P._pn_word_breaks("Ardeshirpour-Zartoshti"):
        assert left[-1].isalpha() and right[0].isalpha(), (left, right)


@pytest.mark.parametrize("category,source,expected", [
    ("person-token", "spreadsheet", True),
    ("person-token", "--term", True),
    ("person-token", "document", False),
    ("person-token", "declarant", False),
    ("person-token", "judge", False),
    ("person", "spreadsheet", True),
    ("entity-token", "spreadsheet", True),
    ("entity", "spreadsheet", True),
    ("entity", "document", False),
    ("short-name", "spreadsheet", False),
    ("case_number", "spreadsheet", False),
    ("display-name", "spreadsheet", False),
])
def test_only_an_authoritative_name_term_is_breakable(category, source,
                                                      expected):
    """A break spelling is a guess about how a printed word came apart.
    Stacking it on a guess about who the party IS doubles the ways it can be
    wrong, so only the operator's own list qualifies — every NAME term on it,
    entity included (see `test_broken_entity_name.py` for why)."""
    assert P._pn_term_is_breakable(category, source) is expected


def test_a_harvested_name_is_not_broken():
    reg = P._PnFakeRegistry()
    terms = []
    P._pn_append_person_terms(terms, "Vadim Sarkisyan", "document", reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    assert z.apply("V ADIM SARKISY AN") == "V ADIM SARKISY AN"


# ── the two passes that decompose a real value into WORDS ───────────────────

def test_no_fragment_of_a_party_is_ever_a_trusted_token():
    """`_trusted_party_tokens` decides the caption exemption — when a cited
    authority loses its protection — and which short cores the blind reduced
    pass may rewrite. Registering a spelling per break position put `sar`,
    `sark`, `syan`, `yan`, `vad` and `dim` in it."""
    z, _out = _run(PARTIES, CAPTION)
    assert z._trusted_party_tokens() == {
        "vadim", "sarkisyan", "sergey", "mirzoyans"}


def test_a_synthetic_spelling_is_not_a_trusted_token():
    """Including one a key written by an older build hands back."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(PARTIES, [], [], registry=reg)
    stale = P._PnTerm("person-token", "Sarkisy an", "Wrightson",
                      whole_word=True, case_sensitive=False, priority=0,
                      source="spreadsheet", derived=True)
    z = P.Pseudonymizer(terms + [stale], {}, registry=reg)
    assert "sarkisy" not in z._trusted_party_tokens()


def test_a_person_weld_core_is_never_truncated():
    """`_pn_alnum_core` drops a trailing connector — right for
    `Schilleci & Tortorici, P.C.`, meaningless for a person. It cost the last
    word of a broken surname, and the seven-letter prefix left behind was
    hunted unanchored: `SARKISYA.N` came back `WRIGHTSONA.N`."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    rec = {"category": "person-token", "real": "Sarkisy an"}
    assert z._weld_core(rec)[0] == "sarkisyan"


def test_the_reduced_pass_does_not_mangle_a_broken_word():
    """It refuses a span holding a printed boundary and must go on refusing —
    the term above is what cures these, so nothing here needs to guess."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    for text in ("CASE NAME SARKISYA.N, et al.", "V ADIM SARKISY AN"):
        assert z.scrub_welded(text) == text
        assert z.surviving_reals_reduced(text) == []


# ── reversal, and the re-run ────────────────────────────────────────────────

def _key_rows(path):
    wb = openpyxl.load_workbook(path)
    head = [str(h).strip().lower()
            for h in next(wb.active.iter_rows(max_row=1, values_only=True))]
    return [dict(zip(head, r)) for ws in wb.worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]


def test_the_key_carries_no_spelling_the_document_never_used(tmp_path):
    """Nothing new is minted, so the key is exactly what it would be if no
    word had broken: one row per real value, and one reversible row per fake."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    z.apply(CAPTION + EXHIBIT)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    rows = _key_rows(path)
    assert sorted(str(r["real value"]) for r in rows) == [
        "Mirzoyans", "Sarkisyan", "Sergey", "Sergey Mirzoyans",
        "Vadim", "Vadim Sarkisyan"]
    # One reversible row per fake, with nothing marked forward-only: there is
    # no synthetic spelling for `DeAnonymize` to call ambiguous.
    fakes = [str(r["replacement"]).lower() for r in rows]
    assert len(fakes) == len(set(fakes)), fakes
    assert not any(str(r["status"]).strip().lower() == P._PN_KEY_ALT_STATUS
                   for r in rows), rows


def test_the_key_alone_reproduces_the_delivered_export(tmp_path):
    """A re-run has the key and may have no party template to rebuild from.
    The tolerance is asked of the row's own category and source, so the second
    export is the first one byte for byte."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    first = z.apply(CAPTION + EXHIBIT)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)

    reg2 = P._PnFakeRegistry()
    terms, *_ = P._pn_load_key(path, reg2, log)
    z2 = P.Pseudonymizer(terms, {}, registry=reg2)
    assert z2.apply(CAPTION + EXHIBIT) == first
