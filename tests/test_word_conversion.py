"""
Bulk Word-document conversion: when a case folder holds Word filings and NO
PDFs, each .docx/.docm is converted to the same scrubbed .txt export a PDF gets
— pseudonymized, leak-reported, name-scrubbed filename — but never hyperlinked.
When any PDF is present the Word docs are left alone so the usual PDF workflow is
unaffected. These tests need neither fitz nor openpyxl.

Run:  cd PDF-Linker && python3 -m pytest tests/test_word_conversion.py -v
"""
import importlib.util
import logging
import sys
import zipfile
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parent.parent
_spec = importlib.util.spec_from_file_location("pdf_linker", _ROOT / "pdf_linker.py")
pl = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(pl)

log = logging.getLogger("test")
DET = {k: pl._PN_DETECTORS[k] for k in pl._PN_DEFAULT_DETECTORS}
_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx(path, body_xml):
    """Write a minimal .docx whose word/document.xml body is `body_xml`."""
    doc = (f'<?xml version="1.0" encoding="UTF-8"?>'
           f'<w:document xmlns:w="{_W}"><w:body>{body_xml}</w:body></w:document>')
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("word/document.xml", doc)
    return path


def _para(*runs):
    inner = "".join(f'<w:r><w:t xml:space="preserve">{t}</w:t></w:r>' for t in runs)
    return f"<w:p>{inner}</w:p>"


def _pz():
    reg = pl._PnFakeRegistry()
    terms = pl._pn_build_terms(["Ernest N Ramirez", "Ford Motor Company"],
                               ["24STCV23198"], [], registry=reg)
    return pl.Pseudonymizer(terms, DET, registry=reg)


# ── extraction ───────────────────────────────────────────────────────────────

def test_extract_paragraphs(tmp_path):
    p = _docx(tmp_path / "a.docx", _para("Hello ", "world") + _para("Second line"))
    assert pl._extract_docx_text(p) == "Hello world\nSecond line"


def test_extract_tab_and_break(tmp_path):
    body = ("<w:p><w:r><w:t>Left</w:t><w:tab/><w:t>Right</w:t></w:r></w:p>"
            "<w:p><w:r><w:t>One</w:t><w:br/><w:t>Two</w:t></w:r></w:p>")
    p = _docx(tmp_path / "a.docx", body)
    assert pl._extract_docx_text(p) == "Left\tRight\nOne\nTwo"


def test_extract_table_cells_become_lines(tmp_path):
    body = ("<w:tbl><w:tr>"
            "<w:tc>" + _para("Cell A") + "</w:tc>"
            "<w:tc>" + _para("Cell B") + "</w:tc>"
            "</w:tr></w:tbl>")
    p = _docx(tmp_path / "a.docx", body)
    assert pl._extract_docx_text(p) == "Cell A\nCell B"


def test_extract_bad_file_raises(tmp_path):
    bad = tmp_path / "notzip.docx"
    bad.write_text("this is not a zip", encoding="utf-8")
    with pytest.raises(Exception):
        pl._extract_docx_text(bad)


# ── folder scan ──────────────────────────────────────────────────────────────

def test_word_docs_skips_lock_and_legacy(tmp_path):
    _docx(tmp_path / "Brief.docx", _para("x"))
    _docx(tmp_path / "Reply.docm", _para("x"))
    (tmp_path / "~$Brief.docx").write_text("lock", encoding="utf-8")  # Word lock file
    (tmp_path / "Old.doc").write_text("legacy binary", encoding="utf-8")
    names = [p.name for p in pl._word_docs_in_folder(tmp_path)]
    assert names == ["Brief.docx", "Reply.docm"]


# ── conversion + pseudonymization ────────────────────────────────────────────

def test_convert_scrubs_content_and_filename(tmp_path):
    src = _docx(tmp_path / "Ramirez Declaration.docx",
                _para("Plaintiff Ernest N Ramirez sued Ford Motor Company.")
                + _para("Case No. 24STCV23198."))
    z = _pz()
    text = pl._extract_docx_text(src)
    n = pl._convert_word_docs([(src, text)], log, z, "Text Files", None)
    assert n == 1

    out = list((tmp_path / "Text Files").glob("*.txt"))
    assert len(out) == 1
    body = out[0].read_text()
    # names scrubbed in the body
    assert "Ramirez" not in body and "Ford Motor Company" not in body
    assert "24STCV23198" not in body
    # filename scrubbed too — the real party name must not survive in the name
    assert "Ramirez" not in out[0].name
    # export is tracked for the leak gate
    assert z.written == out


def test_convert_off_writes_raw_text(tmp_path):
    src = _docx(tmp_path / "Memo.docx", _para("Ernest N Ramirez"))
    text = pl._extract_docx_text(src)
    n = pl._convert_word_docs([(src, text)], log, None, "Text Files", None)
    assert n == 1
    out = tmp_path / "Text Files" / "Memo.txt"
    assert out.is_file()
    # pseudonymization off → text is verbatim, filename unchanged
    assert out.read_text().strip() == "Ernest N Ramirez"


# ── the no-PDF gate, exercised through main() (pseudonymize off → no fitz) ────

def _run_main(folder, monkeypatch):
    monkeypatch.setattr(sys, "argv",
                        ["pdf_linker.py", str(folder), "--no-pseudonymize"])
    pl.main()


def test_gate_converts_when_no_pdfs(tmp_path, monkeypatch):
    for i in range(3):
        _docx(tmp_path / f"Filing{i}.docx", _para(f"Body {i}"))
    _run_main(tmp_path, monkeypatch)
    out = sorted(p.name for p in (tmp_path / "Text Files").glob("*.txt"))
    assert len(out) == 3


def test_gate_leaves_word_docs_when_a_pdf_is_present(tmp_path, monkeypatch):
    for i in range(3):
        _docx(tmp_path / f"Filing{i}.docx", _para(f"Body {i}"))
    # A single (even unreadable) PDF in the folder means the ordinary PDF
    # workflow is active, so the Word docs must be left untouched.
    (tmp_path / "Motion.pdf").write_bytes(b"%PDF-1.4 not a real pdf")
    _run_main(tmp_path, monkeypatch)
    text_dir = tmp_path / "Text Files"
    converted = list(text_dir.glob("Filing*.txt")) if text_dir.exists() else []
    assert converted == []


# ── an all-Word folder gets the same workflow launchers a PDF batch gets ─────
# A Word-only run produces the same scrubbed .txt exports, the same pseudonym
# key and the same LEAKS worksheet — so it needs the same two double-click
# launchers and the same finish stamp. All three used to be gated on the PDF
# list, so an all-Word folder finished with leaks to triage and nothing to
# click, and no marker to say the run had even completed.

def _launchers(folder):
    return {p.name for p in folder.iterdir()
            if p.suffix in (".bat", ".command")}


def test_word_only_folder_gets_the_rerun_launcher(tmp_path, monkeypatch):
    for i in range(2):
        _docx(tmp_path / f"Filing{i}.docx", _para(f"Body {i}"))
    _run_main(tmp_path, monkeypatch)
    assert any(n.startswith("Re-run PDF-Linker") for n in _launchers(tmp_path))


def test_word_only_folder_is_stamped_done(tmp_path, monkeypatch):
    _docx(tmp_path / "Filing.docx", _para("Body"))
    _run_main(tmp_path, monkeypatch)
    assert [p.name for p in tmp_path.glob("DONE *.txt")]


def test_word_only_folder_gets_the_fix_launcher_once_a_key_exists(
        tmp_path, monkeypatch):
    # The leak-fix launcher is the companion the triage loop needs, and it is
    # written only when the folder has a pseudonym key — that is what
    # --fix-leaks reads. A scrubbing run writes the key, so one run should
    # leave both launchers behind.
    pytest.importorskip("openpyxl")
    _docx(tmp_path / "Filing.docx", _para("Ernest N Ramirez appeared."))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(tmp_path),
                                      "--term", "Ernest N Ramirez"])
    try:
        pl.main()
    except SystemExit:
        pass
    assert (tmp_path / "pseudonym_key.xlsx").is_file()
    assert any(n.startswith("Apply Leak Fixes") for n in _launchers(tmp_path))
    # and the export really was scrubbed, so there is something to triage
    txt = (tmp_path / "Text Files" / "Filing.txt")
    assert "Ernest N Ramirez" not in txt.read_text(encoding="utf-8")


def test_empty_folder_still_gets_no_launchers(tmp_path, monkeypatch):
    # No PDFs and no Word docs is genuinely nothing to do — the folder must
    # not be littered with launchers for a batch that does not exist.
    _run_main(tmp_path, monkeypatch)
    assert _launchers(tmp_path) == set()


# ── an all-Word folder never borrows another case's party list ───────────────
# A Word batch is typically a one-off with no E-Court Order*.xlsx of its own.
# The Downloads fallback ("newest .xlsx there") then hands it whatever case was
# downloaded LAST: the run hunts for a stranger's parties, leaves this case's
# in the clear, and writes their names into this case's key as authoritative
# bindings. Folder-local inputs stay unambiguous and are still honoured.

def _order_sheet(path, plaintiff, defendant, caseno="24STCV00001"):
    openpyxl = pytest.importorskip("openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant"])
    ws.append([caseno, plaintiff, defendant])
    wb.save(path)
    return path


def test_is_word_only_folder(tmp_path):
    assert not pl._is_word_only_folder(tmp_path)          # empty
    _docx(tmp_path / "Filing.docx", _para("Body"))
    assert pl._is_word_only_folder(tmp_path)              # word, no pdf
    (tmp_path / "Motion.pdf").write_bytes(b"%PDF-1.4")
    assert not pl._is_word_only_folder(tmp_path)          # a pdf batch


def test_word_only_folder_ignores_a_downloads_template(tmp_path, monkeypatch):
    home = tmp_path / "home"
    (home / "Downloads").mkdir(parents=True)
    _order_sheet(home / "Downloads" / "Order_Other_Case.xlsx",
                 "Ernest N Ramirez", "Ford Motor Company")
    case = tmp_path / "Convert"
    case.mkdir()
    _docx(case / "Filing.docx", _para("Plaintiff Hollis Vantreight appeared."))
    monkeypatch.setenv("USERPROFILE", str(home))
    monkeypatch.setattr(Path, "home", staticmethod(lambda: home))
    assert pl._pn_find_downloads_key(log) is not None     # it IS findable …
    assert pl._pn_find_party_template(case, log) is None  # … but not used here


def test_pdf_folder_still_uses_the_downloads_template(tmp_path, monkeypatch):
    home = tmp_path / "home"
    (home / "Downloads").mkdir(parents=True)
    sheet = _order_sheet(home / "Downloads" / "Order_Case.xlsx", "A Party", "B Corp")
    case = tmp_path / "Case"
    case.mkdir()
    (case / "Motion.pdf").write_bytes(b"%PDF-1.4")
    monkeypatch.setenv("USERPROFILE", str(home))
    monkeypatch.setattr(Path, "home", staticmethod(lambda: home))
    # The E-Court export landing in Downloads is the designed PDF workflow.
    assert pl._pn_find_party_template(case, log) == sheet


def test_word_folder_still_uses_its_own_template(tmp_path, monkeypatch):
    home = tmp_path / "home"
    (home / "Downloads").mkdir(parents=True)
    _order_sheet(home / "Downloads" / "Order_Other.xlsx", "Stranger Party", "X Corp")
    case = tmp_path / "Convert"
    case.mkdir()
    mine = _order_sheet(case / "Order_Mine.xlsx", "Hollis Vantreight", "Cascadia Freight")
    _docx(case / "Filing.docx", _para("Body"))
    monkeypatch.setenv("USERPROFILE", str(home))
    monkeypatch.setattr(Path, "home", staticmethod(lambda: home))
    # A template the operator put IN the folder is unambiguous — use it.
    assert pl._pn_find_party_template(case, log) == mine


def test_word_run_scrubs_nobody_elses_parties(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    home = tmp_path / "home"
    (home / "Downloads").mkdir(parents=True)
    _order_sheet(home / "Downloads" / "Order_Other_Case.xlsx",
                 "Ernest N Ramirez", "Ford Motor Company")
    case = tmp_path / "Convert"
    case.mkdir()
    # A declaration anchor, so the pre-scan has a name to harvest with no
    # party list at all — that harvest is the only scrubbing left here.
    _docx(case / "Filing.docx",
          _para("Declaration of Hollis Vantreight in support.")
          + _para("Plaintiff Hollis Vantreight appeared."))
    monkeypatch.setenv("USERPROFILE", str(home))
    monkeypatch.setattr(Path, "home", staticmethod(lambda: home))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])
    try:
        pl.main()
    except SystemExit:
        pass
    import openpyxl
    reals = {str(r[1]) for r in
             openpyxl.load_workbook(case / "pseudonym_key.xlsx").active
             .iter_rows(min_row=2, values_only=True) if r[1]}
    assert "Ernest N Ramirez" not in reals      # the other case stays out …
    assert "Ford Motor Company" not in reals
    assert "Hollis Vantreight" in reals         # … the pre-scan still works
