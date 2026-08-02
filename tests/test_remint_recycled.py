"""
A re-run that finds leak triage still pending may re-assign a RECYCLED name.

A reused key exists to reproduce what was already delivered byte for byte, so a
binding is normally untouchable. But triage pending — a quarantined export, or a
worksheet row nobody has answered — means the folder was never handed on: no
draft has been written against these names, so a stand-in can still move for
free. That is the one moment the pool's exhausted-and-numbered "Deverell5" can
be repaired into a real name.

Two halves, both load-bearing:
  * `_pn_recycled_fake` recognises ONLY the shape an exhausted pool mints (a
    pool word hard against a number). A house number, a case number and a
    production stamp all carry digits and none of them is a recycled name.
  * `_pn_triage_pending` is the permission. Without it — a clean folder, or the
    text-only `--fix-leaks` pass, which never reopens the PDFs — every binding
    is pinned exactly as before.

Run:  cd PDF-Linker && python3 -m pytest tests/test_remint_recycled.py -v
"""
import logging

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")

POOL = P._PN_NAME_WORDS[0]              # a real pool word, whatever it is
RECYCLED = f"{POOL}4"                   # …and the shape a spent pool mints


def _key(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    for r in rows:
        ws.append(list(r) + ["replaced", "spreadsheet", 3])
    wb.save(path)
    return path


def _leaks(folder, value="Gregory Yu", fix=""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_LEAK_SHEET
    ws.append(["Value", "Fix? (yes/no)", "File", "Type", "Where (page:line)",
               "Notes"])
    ws.append([value, fix, "Motion.txt", "LEAK", "p.1:2", ""])
    wb.save(folder / "LEAKS.xlsx")


# ── what counts as recycled ─────────────────────────────────────────────────

def test_only_a_pool_word_against_a_number_is_recycled():
    assert P._pn_recycled_fake(RECYCLED) == RECYCLED
    assert P._pn_recycled_fake(f"{P._PN_NAME_WORDS[1]} {RECYCLED}") == RECYCLED
    assert P._pn_recycled_fake("quenby3@postbox9.org")           # local part
    assert P._pn_recycled_fake(f"x@{P._PN_EMAIL_DOMAINS[0].split('.')[0]}9.com")
    # …and the things that merely carry digits
    assert P._pn_recycled_fake(f"1450 {P._PN_STREET_NAMES[0]} Blvd") is None
    assert P._pn_recycled_fake("24STCV23198") is None
    assert P._pn_recycled_fake("DEAL# 23071") is None
    assert P._pn_recycled_fake("(555) 010-2929") is None
    assert P._pn_recycled_fake(f"{P._PN_NAME_WORDS[2]}") is None


# ── the permission ──────────────────────────────────────────────────────────

def test_a_quarantined_export_is_triage_pending(tmp_path):
    (tmp_path / "Text Files").mkdir()
    assert P._pn_triage_pending(tmp_path, "Text Files") is False
    (tmp_path / "Text Files" / "Motion.txt.LEAK").write_text("x")
    assert P._pn_triage_pending(tmp_path, "Text Files") is True


def test_an_unanswered_worksheet_row_is_triage_pending(tmp_path):
    (tmp_path / "Text Files").mkdir()
    _leaks(tmp_path, fix="")
    assert P._pn_triage_pending(tmp_path, "Text Files") is True
    # …and an answered one is not: the operator has dealt with the folder.
    _leaks(tmp_path, fix="no")
    assert P._pn_triage_pending(tmp_path, "Text Files") is False


# ── the effect ──────────────────────────────────────────────────────────────

def test_a_recycled_name_is_re_drawn_and_a_clean_one_is_not(tmp_path):
    key = _key(tmp_path / "pseudonym_key.xlsx",
               [("person", "Craig Garriott", f"{P._PN_NAME_WORDS[1]} {RECYCLED}"),
                ("person", "Dana Whitaker", f"{P._PN_NAME_WORDS[2]} "
                                            f"{P._PN_NAME_WORDS[3]}")])
    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log, remint_recycled=True)
    reals = {str(t.real) for t in terms}
    assert "Craig Garriott" not in reals      # dropped, to be drawn afresh
    assert "Dana Whitaker" in reals           # a clean binding is untouched
    # …and nothing of the recycled binding is left behind to pin it: neither the
    # memo nor the used-pool, or the re-draw would land on the same word again.
    assert RECYCLED.lower() not in reg._used
    assert not any(k[1] == "craig garriott" for k in reg._memo)


def test_without_the_permission_every_binding_is_pinned(tmp_path):
    key = _key(tmp_path / "pseudonym_key.xlsx",
               [("person", "Craig Garriott", f"{P._PN_NAME_WORDS[1]} {RECYCLED}")])
    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    assert [str(t.fake) for t in terms if t.category == "person"] == [
        f"{P._PN_NAME_WORDS[1]} {RECYCLED}"]


def test_the_re_drawn_name_is_a_clean_pool_word(tmp_path):
    """End to end: load with the permission, rebuild the term from the party
    template the way a full run does, and the party comes back with a name
    instead of a number."""
    key = _key(tmp_path / "pseudonym_key.xlsx",
               [("person", "Craig Garriott", f"{P._PN_NAME_WORDS[1]} {RECYCLED}")])
    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log, remint_recycled=True)
    terms += P._pn_build_terms(["Craig Garriott"], [], [], reg)
    fake = next(t.fake for t in terms
                if t.category == "person" and t.real == "Craig Garriott")
    assert P._pn_recycled_fake(fake) is None
    assert "Garriott" not in P.Pseudonymizer(terms, {}, reg).apply(
        "Declaration of Craig Garriott.")


def test_fix_leaks_never_re_assigns(tmp_path, monkeypatch):
    """The text-only pass works on exports that are ALREADY scrubbed and never
    reopens the PDFs, so a fake it moved would be left standing in the export
    with nothing in the key to reverse it — however much triage is pending."""
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master.xlsx"))
    tdir = tmp_path / "Text Files"
    tdir.mkdir()
    key = _key(tmp_path / "pseudonym_key.xlsx",
               [("person", "Craig Garriott", f"{P._PN_NAME_WORDS[1]} {RECYCLED}")])
    (tdir / "Motion.txt.LEAK").write_text(
        f"====== Page 1 ======\n{P._PN_NAME_WORDS[1]} {RECYCLED} testified.\n",
        encoding="utf-8")
    _leaks(tmp_path, value="Gregory Yu", fix="no")

    args = type("_Args", (), {"key": str(key), "term": None})()
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0
    # the stand-in already in the export is still the one the key reverses
    written = (tdir / "Motion.txt").read_text()
    assert RECYCLED in written
    rows = list(openpyxl.load_workbook(
        tmp_path / "pseudonym_key.xlsx").active.iter_rows(values_only=True))[1:]
    assert any(str(r[2]).endswith(RECYCLED) for r in rows)
