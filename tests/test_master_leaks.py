"""
A cross-case MASTER leak log: every flagged value accumulates into one
spreadsheet across all runs and case folders, sorted alphabetically, so a value
that keeps leaking over time is easy to spot. Opt-in (`master_leaks = on`),
relocatable (`master_leaks_path`), off by default.

Run:  cd PDF-Linker && python3 -m pytest tests/test_master_leaks.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _read(path):
    ws = openpyxl.load_workbook(path).active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def test_disabled_by_default_but_opt_in_and_relocatable():
    assert P._pn_master_leaks_path({}) is None
    assert P._pn_master_leaks_path({"master_leaks": "off"}) is None
    assert P._pn_master_leaks_path({"master_leaks": "on"}) is not None
    p = P._pn_master_leaks_path({"master_leaks_path": r"/tmp/Master Leaks.xlsx"})
    assert str(p).endswith("Master Leaks.xlsx")   # override works even w/o on


def test_accumulates_sorted_with_times_seen(tmp_path):
    mp = tmp_path / "master.xlsx"
    P._pn_update_master_leaks(
        mp, [("Zeta Corp", "LEAK"), ("Acme LLC", "LEAK")],
        "Case Alpha", "2026-01-01", log)
    P._pn_update_master_leaks(
        mp, [("Acme LLC", "LEAK"), ("Middle Co", "unknown name?")],
        "Case Beta", "2026-02-02", log)
    rows = _read(mp)
    assert rows[0] == list(P._PN_MASTER_HEADERS)
    values = [r[0] for r in rows[1:]]
    assert values == sorted(values, key=str.lower)          # alphabetical
    acme = next(r for r in rows[1:] if r[0] == "Acme LLC")
    assert acme[2] == 2                                       # Times Seen
    assert "Case Alpha" in acme[3] and "Case Beta" in acme[3]  # both cases
    assert acme[4] == "2026-01-01" and acme[5] == "2026-02-02"  # first/last


def test_same_value_same_case_not_double_counted_across_reruns(tmp_path):
    mp = tmp_path / "master.xlsx"
    P._pn_update_master_leaks(mp, [("Acme LLC", "LEAK")], "Case A", "2026-01-01", log)
    P._pn_update_master_leaks(mp, [("Acme LLC", "LEAK")], "Case A", "2026-01-05", log)
    acme = next(r for r in _read(mp)[1:] if r[0] == "Acme LLC")
    assert acme[2] == 2                     # times-seen counts runs
    assert acme[3] == "Case A"              # but the case is listed once
    assert acme[5] == "2026-01-05"          # last-seen advanced


def test_write_leak_report_updates_master_when_enabled(tmp_path):
    folder = tmp_path / "Case Gamma"
    folder.mkdir()
    entries = [{"file": "Brief.pdf", "type": "LEAK", "value": "Zeta Corp",
                "where": "p.1:1"},
               {"file": "Brief.pdf", "type": "LEAK", "value": "Acme LLC",
                "where": "p.1:2"}]
    mp = tmp_path / "Master Leaks.xlsx"
    cfg = {"master_leaks_path": str(mp)}
    P._pn_write_leak_report(folder, entries, log, cfg=cfg)
    rows = _read(mp)
    assert [r[0] for r in rows[1:]] == ["Acme LLC", "Zeta Corp"]   # sorted


def test_no_master_written_when_disabled(tmp_path):
    folder = tmp_path / "Case Delta"
    folder.mkdir()
    entries = [{"file": "B.pdf", "type": "LEAK", "value": "Zeta Corp",
                "where": "p.1"}]
    P._pn_write_leak_report(folder, entries, log, cfg={})     # disabled
    assert not list(tmp_path.glob("*.xlsx")) or \
        all("master" not in p.name.lower() for p in tmp_path.glob("*.xlsx"))


# ── an INHERITED decision contributes its KEEP and nothing else ──────────────
# The master KEEP sheet is read by every run in every folder. A KEEP-PART row
# ("Alder Law, P.C." -> "[Law]") carries one generalisable lesson — the
# BRACKETED fragment is never a name — and one case-local fact: "Alder" is that
# matter's law firm. Inheriting the keep is right; inheriting "fake the rest"
# is cross-case inference, and it put another case's firm in every folder's log
# and (once unmatched authoritative bindings were pinned) every folder's key.

def _decision(value, fixcell, fake_values):
    return {"value": value, "type": "KEEP-PART", "fix": "yes",
            "replacement": None, "fake_values": fake_values,
            "fixcell": fixcell, "notes": "pseudonym key"}


def test_inherited_keep_still_protects_its_bracketed_fragment():
    # The half that DOES generalise survives: "Law" is never a name, anywhere.
    strict, soft, _nuclear = P._pn_keep_values(
        {"alder law, p.c.": _decision("Alder Law, P.C.", "[Law]", ["Alder"])})
    assert "Law" in strict


def test_keep_protection_is_independent_of_the_faking_half():
    # _pn_keep_values reads the decisions directly, so withholding the
    # fragments as terms cannot weaken the keep.
    d = {"alder law, p.c.": _decision("Alder Law, P.C.", "[Law]", ["Alder"])}
    assert P._pn_keep_values(d)[0] == {"Law"}


def test_origin_marks_the_authoring_folder_not_every_folder_seen():
    # "Cases" accumulates every folder the keep protected text in — real
    # history, but not authorship. Only the AUTHOR fakes the remainder.
    d = {"value": "Alder Law, P.C.", "fix": "yes", "fake_values": ["Alder"],
         "fixcell": "[Law]", "cases": "Alder Matter; Other Matter",
         "origin": "Alder Matter"}
    assert P._pn_decision_is_ours(d, "Alder Matter")
    assert not P._pn_decision_is_ours(d, "Other Matter")


def test_legacy_row_without_origin_falls_back_to_a_lone_case():
    # A sheet written before the Origin column: a decision applied in exactly
    # one folder names it, so that folder is still recognised as the author.
    legacy = {"value": "Alder Law, P.C.", "fix": "yes", "fake_values": ["Alder"],
              "fixcell": "[Law]", "cases": "Alder Matter", "origin": ""}
    assert P._pn_decision_is_ours(legacy, "Alder Matter")
    legacy["cases"] = "Alder Matter; Other Matter"
    assert not P._pn_decision_is_ours(legacy, "Other Matter")


def test_local_decision_still_fakes_the_remainder():
    # In the folder that MADE the decision, the bracket means exactly what it
    # says: keep "Law", fake the rest.
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms([], [], ["Alder"], registry=reg)
    z = P.Pseudonymizer(terms, [], registry=reg)
    z.keep_strict = {"Law"}
    out = z.apply("Alder Law, P.C. appeared.")
    assert "Alder" not in out and "Law" in out


def test_typed_term_is_still_authoritative():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms([], [], ["Alder"], registry=reg)
    assert terms and all(t.source == "--term" for t in terms)
    assert "--term" in P._PN_KEY_UNMATCHED_SOURCES


def test_inherited_fragment_does_not_reach_an_unrelated_folder(tmp_path,
                                                               monkeypatch):
    """End to end: case B never mentions Alder, so nothing about Alder may
    appear in case B's key — while the inherited keep of "Law" still holds."""
    import sys
    import zipfile
    body = ("<w:p><w:r><w:t>Declaration of Hollis Vantreight in support."
            "</w:t></w:r></w:p>"
            "<w:p><w:r><w:t>Filed by Hollis Vantreight Law Group.</w:t></w:r></w:p>")
    case = tmp_path / "CaseB"
    case.mkdir()
    with zipfile.ZipFile(case / "Motion.docx", "w") as z:
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant"])
    ws.append(["24STCV00001", "Hollis Vantreight", "Cascadia Freight, Inc."])
    wb.save(case / "Order_Mine.xlsx")
    # the master KEEP sheet, carrying a bracket decision from ANOTHER matter
    master = tmp_path / "master_leaks.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_MASTER_KEEP_SHEET
    ws.append(list(P._PN_MASTER_KEEP_HEADERS))
    ws.append(["Alder Law, P.C.", "[Law]", "KEEP-PART", 3,
               "24STCV27134 Other Case", "2026-07-27", "2026-07-27", "key"])
    wb.save(master)

    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])
    try:
        P.main()
    except SystemExit:
        pass

    reals = {str(r[1]) for r in
             openpyxl.load_workbook(case / "pseudonym_key.xlsx").active
             .iter_rows(min_row=2, values_only=True) if r[1]}
    assert not any("alder" in r.lower() for r in reals), sorted(reals)
    # the inherited KEEP still did its job: "Law" survived beside a faked name
    txt = next((case / "Text Files").glob("*.txt")).read_text(encoding="utf-8")
    assert "Law Group" in txt
    assert "Hollis Vantreight" not in txt


def test_authoring_folder_keeps_the_bracket_and_fakes_the_rest(tmp_path,
                                                               monkeypatch):
    """The point of the bracket: "Alder Law, P.C." -> "<fake> Law, P.C." in the
    folder that made the decision — and still so on the NEXT run, when the
    local LEAKS.xlsx has been consumed and the decision lives only in the
    master sheet."""
    import sys
    import zipfile
    case = tmp_path / "Alder Matter"
    case.mkdir()
    body = ("<w:p><w:r><w:t>Declaration of Hollis Vantreight in support."
            "</w:t></w:r></w:p>"
            "<w:p><w:r><w:t>Counsel: Alder Law, P.C. appeared.</w:t></w:r></w:p>")
    with zipfile.ZipFile(case / "Motion.docx", "w") as z:
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant", "Other Names"])
    ws.append(["24STCV00001", "Hollis Vantreight", "Cascadia Freight, Inc.",
               "Alder Law, P.C."])
    wb.save(case / "Order_Mine.xlsx")
    # the operator's own triage decision
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LEAKS"
    ws.append(["Value", "Fix? (yes/no)", "File", "Type", "Where (page:line)",
               "Notes"])
    ws.append(["Alder Law, P.C.", "[Law]", "Motion.docx", "unscrubbed name?",
               "p.1", ""])
    wb.save(case / "LEAKS.xlsx")

    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master.xlsx"))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])

    def _run():
        try:
            P.main()
        except SystemExit:
            pass
        return next((case / "Text Files").glob("*.txt")).read_text(
            encoding="utf-8")

    first = _run()
    assert "Alder" not in first          # the identifying word is faked …
    assert "Law, P.C." in first          # … the bracketed part is kept
    # the decision is now recorded as OURS, authored by this folder — under
    # the folder's PSEUDONYM, never its real name: the party word that names
    # the folder is faked there exactly as it is in the export, and the id that
    # settles authorship rides in the same cell.
    keep = openpyxl.load_workbook(tmp_path / "master.xlsx")[
        P._PN_MASTER_KEEP_SHEET]
    row = [r for r in keep.iter_rows(min_row=2, values_only=True) if r[0]][0]
    assert row[8] == P._pn_case_origin(case.name, row[8].rsplit(" [", 1)[0])
    assert "Alder" not in row[8] and "Alder" not in str(row[4])   # Origin/Cases
    assert row[8].endswith(f"[{P._pn_case_id(case.name)}]")
    assert P._pn_decision_is_ours({"origin": row[8]}, case.name)

    import shutil
    shutil.rmtree(case / "Text Files")
    second = _run()                      # LEAKS.xlsx consumed — master only now
    assert "Alder" not in second and "Law, P.C." in second
