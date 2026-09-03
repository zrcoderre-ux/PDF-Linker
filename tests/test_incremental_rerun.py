"""
Adding a document to an already-pseudonymized folder.

The working pattern: the scrubbed .txt exports go out for a draft, the draft
comes back saying a filing is missing, the operator drops that PDF in and
re-runs. For that re-run to be usable, the documents already sent must come
back BYTE-IDENTICAL — otherwise every earlier draft is invalidated and the
work starts over.

So long as `pseudonym_key.xlsx` is beside the exports:

  * every value the key binds keeps its exact fake, whatever the new document
    adds (the key is loaded first and seeds the registry, so a later draw can
    neither move nor collide with an established fake);
  * a party the key never bound — one named ONLY in the missing document, so
    it matched nothing on the first run and `write_key` rightly omitted it —
    is still scrubbed, re-read from the E-Court party template; and
  * an operator KEEP still wins over that supplement: a value marked `no`
    does not come back to life just because it is also a template row.

Run:  cd PDF-Linker && python3 -m pytest tests/test_incremental_rerun.py -v
"""
import logging
import re

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}

PARTIES = ["Ernest N Ramirez", "Ford Motor Company", "Jane Roe",
           "Marcus Bellweather"]
CASENO = "24STCV23198"

# Docs 1-3, the batch that was sent out. Jane Roe is a listed party but is
# named only in the declaration that got left out.
SENT = ("Plaintiff Ernest N Ramirez sued Ford Motor Company in 24STCV23198. "
        "Marcus Bellweather declares. Offices at 414 S. Maple Ave., "
        "Montebello, CA 90640. Write rlally@mortensontaggart.com.")
# The document the draft flagged as missing.
ADDED = ("Declaration of Jane Roe. Jane Roe worked with Ernest N Ramirez at "
         "Ford Motor Company, 414-416 S Maple Avenue, in 24STCV23198.")


@pytest.fixture
def folder(tmp_path):
    """A case folder holding the E-Court party template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant",
               "Other Names"])
    ws.append([CASENO, PARTIES[0], PARTIES[1], "; ".join(PARTIES[2:])])
    wb.save(tmp_path / "Order_Template_Input.xlsx")
    return tmp_path


def _first_run(folder):
    """The original batch: mint fresh fakes from the template, write the key."""
    names, casenos = P._pn_terms_from_xlsx(
        folder / "Order_Template_Input.xlsx", None, log)
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(names, casenos, [], registry=reg),
                        DET, registry=reg)
    out = z.apply(SENT)
    z.write_key(folder / "pseudonym_key.xlsx", log)
    return z, out


def _rerun(folder, decisions=None):
    """The re-run with the missing document added: load the key, supplement
    from the template, honour any operator KEEP."""
    reg = P._PnFakeRegistry()
    terms, key_decisions = P._pn_load_key(folder / "pseudonym_key.xlsx", reg, log)
    terms += P._pn_supplement_key_terms(terms, folder, None, reg, log)
    terms, retired = P._pn_retire_kept_key_terms(
        terms, {**(decisions or {}), **key_decisions}, reg, log)
    return P.Pseudonymizer(terms, DET, registry=reg)


# ── the documents already sent must not move ────────────────────────────────

def test_already_sent_documents_are_byte_identical(folder):
    _z1, sent_out = _first_run(folder)
    assert _rerun(folder).apply(SENT) == sent_out


def test_every_keyed_fake_is_reused_exactly(folder):
    z1, _ = _first_run(folder)
    _rp = P._PN_KEY_HEADERS.index("Replacement")
    rows = {(str(r[0]), str(r[1])): str(r[_rp]) for ws in
            openpyxl.load_workbook(folder / "pseudonym_key.xlsx").worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r[1]}
    z2 = _rerun(folder)
    z2.apply(ADDED)                       # the new document draws its own fakes
    for (cat, real), fake in rows.items():
        # A bare token the BUILDER refuses a term ("Roe" is an ordinary word) is
        # exempt: its row stays in the key because the macro reverses a composed
        # fake word by word off it, but it builds no forward term — see
        # `test_a_withheld_token_is_reversible_but_matches_nothing`.
        if len(real.split()) == 1 and not P._pn_is_name_token(real):
            continue
        assert z2.apply(real) == fake, f"{real!r} moved: {z2.apply(real)!r}"


def test_a_withheld_token_is_reversible_but_matches_nothing(folder):
    """The two ends of the key must answer the same question the same way.

    `_pn_build_terms` refuses "Roe" a bare token — it is an ordinary English
    word, and the full name still scrubs. `write_key` harvests a row per word
    anyway, and `_pn_load_key` used to read every row back as a LIVE term, so
    the word the build had declined came back through the key: a first run left
    a bare "Roe" standing and the re-run scrubbed it, one folder answering one
    question two ways.

    The row must STAY, because the macro undoes "Jane Roe" -> "Widdecombe
    Lassiter" word by word off it. Only the forward term goes.
    """
    _first_run(folder)
    _rp = P._PN_KEY_HEADERS.index("Replacement")
    rows = {(str(r[0]), str(r[1])): str(r[_rp]) for ws in
            openpyxl.load_workbook(folder / "pseudonym_key.xlsx").worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r[1]}
    assert not P._pn_is_name_token("Roe")            # the builder's own verdict
    assert ("person-token", "Roe") in rows           # kept, for the reversal
    z2 = _rerun(folder)
    assert z2.apply("Roe declares.") == "Roe declares."      # no forward term
    assert "Jane Roe" not in z2.apply("Declaration of Jane Roe.")   # full name


def test_the_first_run_and_the_rerun_scrub_alike(folder):
    """The point of the rule: adding a document must not change what a value in
    the ALREADY-SENT batch would have been scrubbed to, and a re-run must not
    quietly scrub more than the run that produced the exports."""
    _first_run(folder)
    names, casenos = P._pn_terms_from_xlsx(
        folder / "Order_Template_Input.xlsx", None, log)
    reg = P._PnFakeRegistry()
    fresh = P.Pseudonymizer(P._pn_build_terms(names, casenos, [], registry=reg),
                            DET, registry=reg)
    z2 = _rerun(folder)
    for probe in ("Declaration of Jane Roe.", "Roe declares.",
                  "Ms. Roe's exhibit.", "Jane A. Roe signed.",
                  "JANE ROE testified.", "Roe, Jane, declarant."):
        assert fresh.apply(probe) == z2.apply(probe), probe


def _street(text):
    """The faked "<number> <name>" of the first address in `text`."""
    return re.search(r"\d+ [A-Z][a-z]+", text).group(0)


def test_case_number_and_address_stay_consistent(folder):
    _z1, sent_out = _first_run(folder)
    faked_caseno = sent_out.split(" in ")[1].split(".")[0]
    z2 = _rerun(folder)
    added_out = z2.apply(ADDED)
    assert faked_caseno in added_out
    # the same parcel written a second way folds onto the same faked street
    # NAME. The house number is kept verbatim now, so "414" and the range
    # "414-416" print their own real numbers — that is the parcel reading as
    # the document wrote it, not two different streets.
    name = lambda s: P._pn_addr_name_of(_street(s))
    assert name(added_out) == name(sent_out.split("Offices at ")[1])


# ── the party the key never bound is still scrubbed ─────────────────────────

def test_party_named_only_in_the_added_document_is_scrubbed(folder):
    z1, _ = _first_run(folder)
    # She matched nothing, but the party template names her, so the key pins
    # her binding ("no match") instead of discarding it.
    _st = list(P._PN_KEY_HEADERS).index("Status")        # not a fixed column
    _rp = P._PN_KEY_HEADERS.index("Replacement")
    rows = {str(r[1]): (str(r[_rp]), r[_st]) for ws in
            openpyxl.load_workbook(folder / "pseudonym_key.xlsx").worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r[1]}
    assert rows["Jane Roe"][1] == "no match"
    out = _rerun(folder).apply(ADDED)
    assert "Jane Roe" not in out
    # and she is faked to exactly the stand-in the first run had reserved
    assert rows["Jane Roe"][0] in out


def test_pinned_party_survives_even_without_the_template(folder):
    # The key alone now carries the binding, so the supplement is a fallback
    # for an older key or an amended template — not the only line of defence.
    _first_run(folder)
    (folder / "Order_Template_Input.xlsx").unlink()
    assert "Jane Roe" not in _rerun(folder).apply(ADDED)


def test_supplemented_party_does_not_disturb_the_others(folder):
    _z1, sent_out = _first_run(folder)
    z2 = _rerun(folder)
    z2.apply(ADDED)                       # supplement draws first, if it can
    assert z2.apply(SENT) == sent_out


def test_supplement_is_skipped_when_no_template_remains(folder):
    # The template may not have been kept beside the exports; the re-run must
    # still work off the key alone rather than failing.
    _z1, sent_out = _first_run(folder)
    (folder / "Order_Template_Input.xlsx").unlink()
    assert _rerun(folder).apply(SENT) == sent_out


# ── an operator KEEP still beats the supplement ─────────────────────────────

def test_keep_is_not_resurrected_by_the_party_template(folder):
    # The operator decided this listed value stays verbatim. Re-reading the
    # template must not quietly start faking it again.
    _first_run(folder)
    keep = {"marcus bellweather": {
        "value": "Marcus Bellweather", "type": "KEEP", "fix": "no",
        "replacement": None, "fake_values": None, "fixcell": None,
        "notes": "test"}}
    out = _rerun(folder, keep).apply("Marcus Bellweather declares.")
    assert "Marcus Bellweather" in out


def test_keep_is_not_reassembled_from_supplemented_tokens(folder):
    _first_run(folder)
    keep = {"marcus bellweather": {
        "value": "Marcus Bellweather", "type": "KEEP", "fix": "no",
        "replacement": None, "fake_values": None, "fixcell": None,
        "notes": "test"}}
    z = _rerun(folder, keep)
    assert z.apply("Marcus said so.") == "Marcus said so."
    assert z.apply("Bellweather said so.") == "Bellweather said so."
