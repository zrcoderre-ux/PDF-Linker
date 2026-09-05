"""Alternate spellings are written TOGETHER, under the value they are
spellings of.

Two key rows sharing a Replacement are never two parties — the registry is
injective, so they are one value written several ways: a wrap-split hyphen
("Ardeshirpour- Zartoshti"), a `_pn_name_variants` near-miss ("Sarra"), a
surname-first table spelling, an operator `*ANOTHER VALUE` alias. `write_key`
already marks the non-canonical ones `alt spelling` — but the sheet sorted
alphabetically, so they were scattered down the key and the one thing that
Status word says ("this is another spelling of some OTHER row") could only be
acted on by searching the sheet for the Replacement.

So the key is ordered as PARTY BLOCKS: the full name, its own alternate
spellings, then each of its bare tokens with that token's spellings under it.
And `LEAKS.xlsx` groups the same way, on the `*CANONICAL` cell — a badly
scanned party name arrives as a dozen pre-filled rows that are right together
or wrong together, and read scattered down an alphabet they cannot be.

Run:  cd PDF-Linker && python3 -m pytest tests/test_alt_spelling_grouping.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
ALT = P._PN_KEY_ALT_STATUS

PARTIES = ["Sara Ardeshirpour-Zartoshti", "Manuel Vazquez",
           "Midland States Bank"]
TEXT = ("Declaration of Sara Ardeshirpour- Zartoshti. Dr. Ardeshirpour "
        "examined the plaintiff. Sarra Ardeshirpour-Zartoshti signed it. "
        "MANUEL VAZQUEZ testified; Vazquez left. Midland States Bank is the "
        "plaintiff. Midland sued. Bank records were produced.")


def _run(tmp_path, names=PARTIES, text=TEXT):
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(list(names), [], [], registry=reg),
                        {}, registry=reg)
    out = z.apply(text)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    wb = openpyxl.load_workbook(path)
    head = [str(h).strip().lower() for h in
            next(wb.active.iter_rows(max_row=1, values_only=True))]
    rows = [dict(zip(head, r)) for ws in wb.worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    return out, rows


def _reals(rows):
    return [str(r["real value"]) for r in rows]


def _index(rows, real):
    return _reals(rows).index(real)


# ── the key: alternate spellings sit under their own row ────────────────────

def test_an_alt_spelling_follows_the_row_it_is_a_spelling_of(tmp_path):
    """Every `alt spelling` row is written immediately after a row carrying
    the same Replacement — the row that reverses it. Scattered, the Status
    word names no row at all."""
    _out, rows = _run(tmp_path)
    alts = [i for i, r in enumerate(rows)
            if str(r["status"]).strip().lower() == ALT]
    assert alts, "fixture stopped producing alternate spellings"
    for i in alts:
        assert i > 0, f"{rows[i]['real value']!r} leads the sheet"
        assert (str(rows[i - 1]["replacement"]).lower()
                == str(rows[i]["replacement"]).lower()), (
            f"{rows[i]['real value']!r} ({ALT}) does not sit under a row "
            f"carrying its Replacement; the row above is "
            f"{rows[i - 1]['real value']!r}")


def test_the_spellings_of_one_value_are_contiguous(tmp_path):
    """One Replacement, one run of rows — never two runs with unrelated
    parties between them."""
    _out, rows = _run(tmp_path)
    seen, last = set(), None
    for r in rows:
        fake = str(r["replacement"]).lower()
        if fake != last:
            assert fake not in seen, (
                f"the rows for {fake!r} are split across the sheet")
            seen.add(fake)
            last = fake


# ── the key: a party's tokens sit under the full name ───────────────────────

def test_both_names_tokens_are_grouped_under_the_full_name(tmp_path):
    """"If the first and last name are together in one real value cell, all of
    the variants of both first and last names should be grouped together below
    that." The full name leads; its own tokens follow it, before any other
    party's row."""
    _out, rows = _run(tmp_path)
    start = _index(rows, "Manuel Vazquez")
    assert _reals(rows)[start:start + 3] == ["Manuel Vazquez", "Manuel",
                                             "Vazquez"]


def test_a_tokens_own_alt_spelling_rides_with_it(tmp_path):
    """The near-miss "Sarra" is a spelling of the token "Sara", so it sits
    under the token, inside the party's own block."""
    _out, rows = _run(tmp_path)
    block = _reals(rows)[_index(rows, "Sara Ardeshirpour-Zartoshti"):][:7]
    assert block == ["Sara Ardeshirpour-Zartoshti",
                     "Sara", "Sarra",
                     "Ardeshirpour-Zartoshti", "Ardeshirpour- Zartoshti",
                     "Ardeshirpour", "Zartoshti"], block


def test_a_multi_word_short_form_is_part_of_its_party(tmp_path):
    """`_pn_entity_bare` registers "Midland States" off "Midland States Bank",
    and its fake is a RUN of the party's fake words rather than one of them.
    It belongs to the party — and taking it in is also what lets the block be
    ordered: it has to precede the "Midland" whose fake it begins with, while
    the full name has to precede them both."""
    _out, rows = _run(tmp_path)
    block = _reals(rows)[_index(rows, "Midland States Bank"):][:4]
    assert block == ["Midland States Bank", "Midland States", "Midland",
                     "States"], block


# ── the reversal order still holds ──────────────────────────────────────────

def test_no_replacement_is_written_before_a_longer_one_it_begins(tmp_path):
    """The rule `_pn_key_longer_first` exists for: a reader that searches
    substrings in row order would turn "Cranston" into "Kenston" on meeting
    "Cran" first. Grouping must not cost it."""
    _out, rows = _run(tmp_path)
    fakes = [str(r["replacement"]) for r in rows]
    for i, later in enumerate(fakes):
        for earlier in fakes[:i]:
            assert not (len(later) > len(earlier)
                        and later.lower().startswith(earlier.lower())), (
                f"{earlier!r} is written before the longer {later!r} "
                f"it begins")
    # …and the sheet really does hold such a pair, so the assertion above is
    # not vacuously true.
    assert any(a != b and len(a) > len(b) and a.lower().startswith(b.lower())
               for a in fakes for b in fakes)


def test_the_order_is_reproducible(tmp_path):
    """A key rewritten from the same folder must come back in the same order,
    or every re-run reads as the tool having shuffled the operator's file."""
    again = tmp_path / "again"
    again.mkdir()
    _out, first = _run(tmp_path)
    _out2, second = _run(again)
    assert _reals(first) == _reals(second)


def test_block_ordering_terminates_on_a_straddling_pair():
    """Two blocks can each hold a fake that is the front of one of the
    other's, which no ordering satisfies. The pass declines the second move
    rather than swapping forever."""
    blocks = [[{"fake": "ab"}, {"fake": "cdcd"}],
              [{"fake": "abc"}, {"fake": "cd"}]]
    P._pn_key_longer_first_blocks(blocks)          # must not hang
    assert len(blocks) == 2


# ── LEAKS: the pre-filled alias rows group too ──────────────────────────────

def _leaks(tmp_path, values, z, decisions=None):
    entries = [{"file": f"{chr(65 + i)}.txt", "type": "misspelled name?",
                "value": v, "where": "p.1:3", "context": f"{v} served it."}
               for i, v in enumerate(values)]
    P._pn_write_leak_report(tmp_path, entries, log, decisions=decisions,
                            suggest_for=z.alias_suggestion)
    wb = openpyxl.load_workbook(tmp_path / "LEAKS.xlsx")
    ws = wb[P._PN_LEAK_SHEET]
    head = [str(h).strip() for h in
            next(ws.iter_rows(max_row=1, values_only=True))]
    return [dict(zip(head, r)) for r in ws.iter_rows(min_row=2,
                                                     values_only=True)]


@pytest.fixture
def pz():
    reg = P._PnFakeRegistry()
    return P.Pseudonymizer(
        P._pn_build_terms(["Manuel Vazquez", "Rachel Ashworth"], [], [],
                          registry=reg), {}, registry=reg)


def test_the_prefilled_misspellings_of_one_value_are_contiguous(tmp_path, pz):
    """Twenty scanned spellings of one defendant are right together or wrong
    together; the file names sort them apart, so the alias does the grouping."""
    rows = _leaks(tmp_path, ["Vazqez", "Aardvark", "Vazqoe", "Bickerstaff",
                             "Vasquez", "Zenith"], pz)
    fix = "Fix? (yes/no)"
    hits = [i for i, r in enumerate(rows)
            if str(r[fix] or "").lower() == "*vazquez"]
    assert len(hits) == 3, [(r["Value"], r[fix]) for r in rows]
    assert hits == list(range(hits[0], hits[0] + len(hits))), (
        [(r["Value"], r[fix]) for r in rows])


def test_a_group_sorts_where_its_strongest_member_would_have(tmp_path, pz):
    """Grouping must not pull a family forward or bury it: the cluster lands
    at the place its best-sorting row already had."""
    rows = _leaks(tmp_path, ["Aardvark", "Vazqez", "Vazqoe"], pz)
    assert [r["Value"] for r in rows] == ["Aardvark", "Vazqez", "Vazqoe"]


def test_a_typed_alias_groups_with_the_prefilled_ones(tmp_path, pz):
    """The cell is read the same way whoever typed it — one reader, so the
    operator's own answer joins the family rather than sorting away from it."""
    rows = _leaks(tmp_path, ["Aardvark", "Vazqez", "Vazqoe"], pz,
                  decisions={"aardvark": {"value": "Aardvark", "fix": "yes",
                                          "fixcell": "*Vazquez",
                                          "alias": "Vazquez"}})
    assert [r["Value"] for r in rows] == ["Aardvark", "Vazqez", "Vazqoe"]


def test_an_undecided_row_never_sinks_to_sit_beside_a_resolved_sibling(
        tmp_path, pz):
    """The attention tier still leads the order: a row needing a look stays at
    the top even when a decided row names the same canonical."""
    rows = _leaks(tmp_path, ["Vazqez"], pz,
                  decisions={"vazqoe": {"value": "Vazqoe", "fix": "yes",
                                        "fixcell": "*Vazquez",
                                        "alias": "Vazquez"}})
    order = [r["Value"] for r in rows]
    assert order.index("Vazqez") < order.index("Vazqoe"), order


@pytest.mark.parametrize("cell,canon", [
    ("*Vazquez", "vazquez"),
    ("* Manuel Vazquez", "manuel vazquez"),
    ("=Vazquez", "vazquez"),                 # the older spelling still reads
    ("*David {said}", "david"),              # alias composed with a keep-spec
    ("no", ""),
    ("[Human Resources]", ""),
    ("", ""),
    (None, ""),
])
def test_the_grouping_reads_the_cell_the_way_every_other_pass_does(cell, canon):
    assert P._pn_leak_alias_canon(cell) == canon


# ── the key: spellings that do NOT share the Replacement ────────────────────

def _key_rows(z, tmp_path, name="pseudonym_key.xlsx"):
    path = tmp_path / name
    z.write_key(path, log)
    wb = openpyxl.load_workbook(path)
    head = [str(h).strip().lower() for h in
            next(wb[P._PN_KEY_MAIN_SHEET].iter_rows(max_row=1, values_only=True))]
    return path, [dict(zip(head, r)) for r in
                  wb[P._PN_KEY_MAIN_SHEET].iter_rows(min_row=2, values_only=True)
                  if r and r[0]]


def test_a_surname_first_table_spelling_sits_under_its_party(tmp_path):
    """"Vazquez Manuel" carries the party's words REVERSED, so it shares no
    Replacement with "Manuel Vazquez" — and as a `person` row of its own it
    sorted first, claimed the tokens in its own word order, and put the
    reversed spelling at the head of the party. It is a spelling of the
    party, marked as one, and written right after it; the tokens keep the
    party's own word order."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(["Manuel Vazquez", "Rachel Ashworth"],
                                          [], [], registry=reg), {}, registry=reg)
    z.apply("MANUEL VAZQUEZ testified. Vazquez Manuel signed the table row. "
            "Rachel Ashworth appeared. Ashworth Rachel is listed.")
    _path, rows = _key_rows(z, tmp_path)
    assert _reals(rows) == ["Manuel Vazquez", "Vazquez Manuel", "Manuel",
                            "Vazquez", "Rachel Ashworth", "Ashworth Rachel",
                            "Rachel", "Ashworth"], _reals(rows)
    status = {r["real value"]: str(r["status"]).lower() for r in rows}
    assert status["Vazquez Manuel"] == ALT
    assert status["Manuel Vazquez"] == "replaced"


def test_an_operator_alias_sits_under_the_token_it_misspells(tmp_path):
    """A `*` alias takes a mirrored SLIP of the canonical's fake and never the
    fake itself, so the twenty scanned spellings of one defendant shared
    nothing the same-Replacement grouping could see. Read off the values, they
    ride with the token they are slips of."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Manuel Vazquez", "Rachel Ashworth"], [], [],
                              registry=reg)
    decisions = {v.lower(): {"value": v, "fix": "yes", "fixcell": "*Vazquez",
                             "alias": "Vazquez"} for v in ("Vazqez", "Vatquel")}
    terms, vals = P._pn_apply_aliases(decisions, terms, reg, log)
    terms += P._pn_build_terms([], [], vals, reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    z.apply("MANUEL VAZQUEZ testified. Vazqez left. Vatquel signed. "
            "Rachel Ashworth appeared.")
    _path, rows = _key_rows(z, tmp_path)
    reals = _reals(rows)
    i = reals.index("Vazquez")
    assert reals[i + 1:i + 5] == ["Vatquel", "Vatquel", "Vazqez", "Vazqez"], reals
    assert reals.index("Rachel Ashworth") > i + 4


def test_an_inferred_typo_fold_sits_under_the_token_it_misspells(tmp_path):
    """The registry's own fold — "Palladina" beside "Palladino" — is the same
    shape as the operator's, and groups the same way."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(["Marco Palladino", "Marco Palladina"],
                                          [], [], registry=reg), {}, registry=reg)
    out = z.apply("Marco Palladino signed. Marco Palladina countersigned.")
    fakes = {r.real: r.fake for r in z.terms}
    assert fakes["Marco Palladina"] != fakes["Marco Palladino"]
    _path, rows = _key_rows(z, tmp_path)
    reals = _reals(rows)
    # Which spelling holds the POOL word is the registry's business (the
    # shortest-first pre-bind drew "Palladina" first here); the one that
    # folded onto it is written right under it, and their tokens likewise.
    assert set(reals[:2]) == {"Marco Palladino", "Marco Palladina"}, reals
    assert abs(reals.index("Palladina") - reals.index("Palladino")) == 1, reals


def test_a_fold_is_read_off_the_values_so_a_rerun_off_the_key_agrees(tmp_path):
    """A re-run off the delivered key re-folds nothing (every fake is pinned)
    and no longer holds the `*` cell, so the grouping must be decidable from
    the rows alone — or the sheet comes back shuffled on the first re-run."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Manuel Vazquez", "Marco Palladino",
                               "Marco Palladina"], [], [], registry=reg)
    decisions = {"vatquel": {"value": "Vatquel", "fix": "yes",
                             "fixcell": "*Vazquez", "alias": "Vazquez"}}
    terms, vals = P._pn_apply_aliases(decisions, terms, reg, log)
    terms += P._pn_build_terms([], [], vals, reg)
    text = ("MANUEL VAZQUEZ testified. Vazquez Manuel signed. Vatquel left. "
            "Marco Palladino signed. Marco Palladina countersigned.")
    z = P.Pseudonymizer(terms, {}, registry=reg)
    z.apply(text)
    path, first = _key_rows(z, tmp_path)

    reg2 = P._PnFakeRegistry()
    loaded, _decisions = P._pn_load_key(path, reg2, log)
    z2 = P.Pseudonymizer(loaded, {}, registry=reg2)
    again = tmp_path / "again"
    again.mkdir()
    assert z2.apply(text) == z.apply(text)
    _path2, second = _key_rows(z2, again)
    assert _reals(second) == _reals(first)
    assert [str(r["status"]).lower() for r in second] == [
        str(r["status"]).lower() for r in first]


@pytest.mark.parametrize("real,fake,preal,pfake,want", [
    ("Vazqez", "Inleby", "Vazquez", "Ingleby", True),     # a letter dropped
    ("Vatquel", "Inglebj", "Vazquez", "Ingleby", True),   # a confusable
    ("Palladina", "Keswicka", "Palladino", "Keswick", False),  # lengths do not track
    ("Palladinoo", "Keswickk", "Palladino", "Keswick", True),  # …and here they do
    ("Ken", "Cran", "Kenneth", "Cranston", True),         # a nickname's front
    ("Smiths", "Deverell5", "Jones", "Deverell", False),  # the recycled stand-in
    ("Radley", "Ridley", "Bradley", "Radley", False),     # a pool word is a draw
    ("Vazquez", "Ingleby", "Vazquez", "Ingleby", False),  # the same binding
])
def test_a_fold_is_read_off_the_four_words_alone(real, fake, preal, pfake, want):
    assert P._pn_key_word_fold(real, fake, preal, pfake) is want


def test_every_row_is_written_once_when_spellings_nest():
    """A fold of a fold that sorts AHEAD of the name it folds from is claimed
    as a spelling by the first fold, which is then itself claimed by the name.
    The walk is recursive, so the nested row is written under its family —
    and never dropped, since a row missing from the key is a fake nothing can
    reverse."""
    assert "ingleby" in P._PN_POOL_WORDS
    rows = [
        {"category": "person", "real": "Marco Zed", "fake": "Thorne Ingleby",
         "count": 3, "source": "spreadsheet"},
        {"category": "person", "real": "Marco Aed", "fake": "Thorne Inglebj",
         "count": 1, "source": "spreadsheet"},
        {"category": "person", "real": "Marco Bed", "fake": "Thorne Inglebi",
         "count": 1, "source": "spreadsheet"},
        {"category": "person-token", "real": "Zed", "fake": "Ingleby",
         "count": 0, "source": "spreadsheet"},
        {"category": "person-token", "real": "Marco", "fake": "Thorne",
         "count": 0, "source": "spreadsheet"},
        {"category": "address", "real": "12 Elm St", "fake": "12 Oak St",
         "count": 1, "source": "regex"},
    ]
    out = P._pn_key_party_order(list(rows), set())
    reals = [r["real"] for r in out]
    assert sorted(reals) == sorted(r["real"] for r in rows), reals
    assert reals[0] == "Marco Zed"
    assert reals.index("Marco Bed") < reals.index("12 Elm St")
    assert reals.index("Marco Aed") < reals.index("12 Elm St")


def _starred(tmp_path, text):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Manuel Vazqez"], [], [], registry=reg)
    decisions = {"vazqez": {"value": "Vazqez", "fix": "yes",
                            "fixcell": "*Vazquez", "alias": "Vazquez"}}
    terms, vals = P._pn_apply_aliases(decisions, terms, reg, log)
    terms += P._pn_build_terms([], [], vals, reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    z.apply(text)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    wb = openpyxl.load_workbook(path)
    return {ws.title: [(str(r[1]), str(r[2])) for r in
                       ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
            for ws in wb.worksheets}


def test_the_starred_spelling_takes_the_tokens_slot_in_its_party(tmp_path):
    """The template spelled the name wrong and a document spells it right.
    The value after the star is the correct spelling, so it takes the
    surname's slot in the party block and the misspelling is written under
    it — even though the misspelling is the one the party was built from."""
    sheets = _starred(tmp_path, "MANUEL VAZQEZ testified. Vazquez left.")
    reals = [r for r, _f in sheets[P._PN_KEY_MAIN_SHEET]]
    assert reals[:3] == ["Manuel Vazqez", "Manuel", "Vazquez"], reals
    assert reals[3:5] == ["Vazquez", "Vazqez"] or reals[3] == "Vazqez", reals
    fakes = dict(sheets[P._PN_KEY_MAIN_SHEET])
    assert fakes["Vazquez"].lower() in P._PN_POOL_WORDS
    assert fakes["Vazqez"].lower() not in P._PN_POOL_WORDS


def test_a_starred_spelling_no_document_carries_is_pinned_with_the_clean_word(
        tmp_path):
    """No export carries the correct spelling, so its row lives on the pinned
    sheet (`_PN_KEY_PINNED_SHEET`) as every unmatched authoritative binding
    does — holding the clean pool word the misspelling used to hold, while the
    misspelling on the main sheet carries the slip."""
    sheets = _starred(tmp_path, "MANUEL VAZQEZ testified. Vazqez left.")
    main, pinned = (dict(sheets[P._PN_KEY_MAIN_SHEET]),
                    dict(sheets[P._PN_KEY_PINNED_SHEET]))
    assert "Vazquez" in pinned and "Vazquez" not in main
    assert pinned["Vazquez"].lower() in P._PN_POOL_WORDS
    assert main["Vazqez"].lower() not in P._PN_POOL_WORDS
    assert P._pn_osa_distance(main["Vazqez"].lower(),
                              pinned["Vazquez"].lower()) == 1
