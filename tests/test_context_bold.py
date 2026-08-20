"""The flagged value is BOLD inside its own Context quote.

A Context cell is a whole sentence and the value is a word or two inside it, so
finding the value means reading the sentence. Bolding it makes the row
answerable at a glance, which is the entire job of the column.

Only the quote is styled. The Value column beside it stays in the ordinary font,
because there the value IS the cell and emphasis would say nothing.

The styling is DERIVED at write time from (quote, value), never stored — which
is what makes it survive the two round-trips that matter: a quote carried
forward from the key already on disk reads back as plain text and is re-bolded
on the next run, and `DeAnonymize.bas` (and every reader that does not ask for
rich text) sees the ordinary sentence.

Run:  cd PDF-Linker && python3 -m pytest tests/test_context_bold.py -v
"""
import logging

import openpyxl
import pytest
from openpyxl.cell.rich_text import CellRichText, TextBlock

import pdf_linker as P

log = logging.getLogger("test")

SOURCE = ("====== Page 1 ======\n"
          " 1  HELEN RASHO, an individual, brought this action. Counsel\n"
          " 2  served Marcus Delacroix at his residence, and Marcus\n"
          " 3  Delacroix accepted service on that date.\n")


def _pz():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Helen Rasho", "Marcus Delacroix"], [], [],
                              registry=reg)
    return P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)


def _key(tmp_path):
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    return path


def _bold_runs(cell_value):
    """The bolded substrings of a rich-text cell ([] when it is plain text)."""
    if not isinstance(cell_value, CellRichText):
        return []
    return [b.text for b in cell_value if isinstance(b, TextBlock)]


def _rows(path, rich=True):
    wb = openpyxl.load_workbook(path, rich_text=rich)
    return [(ws.title, r) for ws in wb.worksheets
            for r in ws.iter_rows(min_row=2) if r[0].value is not None]


# ── the key ─────────────────────────────────────────────────────────────────

def test_the_key_bolds_the_real_value_inside_its_quote(tmp_path):
    ctx = P._PN_KEY_HEADERS.index("Context")
    rv = P._PN_KEY_HEADERS.index("Real Value")
    seen = 0
    for _sheet, r in _rows(_key(tmp_path)):
        real, cell = str(r[rv].value), r[ctx].value
        if not cell:
            continue
        seen += 1
        runs = _bold_runs(cell)
        assert runs, f"{real!r} quote was not bolded: {cell!r}"
        assert all(b.lower() == real.lower() for b in runs), (real, runs)
    assert seen, "no quoted row to check"


def test_every_occurrence_is_bolded_not_just_the_first(tmp_path):
    # "Marcus Delacroix" is named twice in the quoted sentence run.
    ctx = P._PN_KEY_HEADERS.index("Context")
    rv = P._PN_KEY_HEADERS.index("Real Value")
    for _sheet, r in _rows(_key(tmp_path)):
        if str(r[rv].value) == "Marcus Delacroix":
            assert len(_bold_runs(r[ctx].value)) == 2, r[ctx].value
            return
    pytest.fail("Marcus Delacroix had no row")


def test_the_match_is_case_insensitive_and_keeps_the_documents_casing(tmp_path):
    # The caption shouts the name the key spells in title case; both are the
    # same value, and the quote must read as the document wrote it.
    ctx = P._PN_KEY_HEADERS.index("Context")
    rv = P._PN_KEY_HEADERS.index("Real Value")
    for _sheet, r in _rows(_key(tmp_path)):
        if str(r[rv].value) == "Helen Rasho":
            assert _bold_runs(r[ctx].value) == ["HELEN RASHO"], r[ctx].value
            return
    pytest.fail("Helen Rasho had no row")


def test_the_value_column_itself_stays_standard(tmp_path):
    path = _key(tmp_path)
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for r in ws.iter_rows(min_row=2):
            assert not r[P._PN_KEY_HEADERS.index("Real Value")].font.bold


# ── LEAKS ───────────────────────────────────────────────────────────────────

def _leaks(tmp_path, value="Rasho", context=None):
    P._pn_write_leak_report(tmp_path, [
        {"file": "M.pdf", "type": "name?", "value": value, "where": "p.1:1",
         "notes": "",
         "context": context if context is not None
         else "HELEN RASHO, an individual, brought this action."}], log)
    return tmp_path / "LEAKS.xlsx"


def _leak_col(name):
    return next(i for i, (h, _k, _w) in enumerate(P._PN_LEAK_COLUMNS)
                if h == name)


def test_the_leaks_worksheet_bolds_the_value_in_its_quote(tmp_path):
    path = _leaks(tmp_path)
    _sheet, r = _rows(path)[0]
    assert _bold_runs(r[_leak_col("Context")].value) == ["RASHO"]


def test_the_leaks_value_column_stays_standard(tmp_path):
    path = _leaks(tmp_path)
    wb = openpyxl.load_workbook(path)
    assert not wb.active.cell(row=2, column=_leak_col("Value") + 1).font.bold


# ── it must not disturb anything that reads these files ─────────────────────

def test_a_plain_reader_still_sees_the_sentence(tmp_path):
    """`DeAnonymize.bas` and `_pn_key_context_on_disk` do not ask for rich text.

    A rich-text cell read without `rich_text=True` yields the ordinary string,
    which is what keeps the macro and the carry-forward path unaffected.
    """
    path = _key(tmp_path)
    ctx = P._PN_KEY_HEADERS.index("Context")
    rv = P._PN_KEY_HEADERS.index("Real Value")
    for _sheet, r in _rows(path, rich=False):
        cell = r[ctx].value
        if cell:
            assert isinstance(cell, str), type(cell)
            assert str(r[rv].value).lower() in cell.lower()

    carried, _carried_scrubbed = P._pn_key_context_on_disk(path)
    assert carried, "the carry-forward read found nothing"
    assert all(isinstance(v, str) and v for v in carried.values())


def test_a_carried_forward_quote_is_re_bolded_next_run(tmp_path):
    # The bolding is derived, not stored, so a quote this run could not
    # re-derive still comes back bolded.
    path = _key(tmp_path)
    z2 = _pz()
    z2.apply("Nothing relevant here.")
    z2.note_key_context("Nothing relevant here.")
    assert not z2._key_context
    z2.write_key(path, log)

    ctx = P._PN_KEY_HEADERS.index("Context")
    rv = P._PN_KEY_HEADERS.index("Real Value")
    for _sheet, r in _rows(path):
        if str(r[rv].value) == "Marcus Delacroix":
            assert _bold_runs(r[ctx].value), r[ctx].value
            return
    pytest.fail("Marcus Delacroix had no row")


def test_wrapping_still_applies_to_a_rich_text_cell(tmp_path):
    path = _key(tmp_path)
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    assert c.alignment.wrap_text and c.alignment.vertical == "top"


# ── and it degrades rather than failing ─────────────────────────────────────

@pytest.mark.parametrize("quote,value", [
    ("a sentence with no match in it", "Rasho"),   # welded/reduced finding
    ("", "Rasho"),
    ("HELEN RASHO, an individual", ""),
    ("HELEN RASHO, an individual", None),
    (None, "Rasho"),
])
def test_an_unbolddable_cell_falls_back_to_plain_text(quote, value):
    got = P._pn_rich_context(quote, value)
    assert isinstance(got, str), got
    assert got == str(quote or "")
