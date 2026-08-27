"""
The `=ANOTHER VALUE` alias: telling the tool that one Real Value is a
MISSPELLING of another.

A filing that spells one party two ways ("ANTIONO" beside "ANTIONIO") gives the
tool two values. The automatic OCR/typo fold links them when it meets them
together and near enough; when it does not — a reused key pinned one of them, or
the two are further apart than `_pn_name_fold_dist` allows — each draws an
unrelated pool word and one person comes out under two names. The operator says
so by typing `=ANTIONIO` over the fake in the key's Replacement column, or into
the LEAKS worksheet's Fix? cell.

The fake is then the SAME misspelling of the canonical's fake — never the same
fake, because two Real Values sharing one Replacement is exactly what
DeAnonymize.bas calls ambiguous and refuses to reverse.

Run:  cd PDF-Linker && python3 -m pytest tests/test_misspelling_alias.py -v
"""
import logging
from pathlib import Path

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
_HDR = ["Category", "Real Value", "Replacement", "Status", "Source", "Occurrences"]


def _write_key(path, rows):
    """rows: list of (category, real, replacement)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(_HDR)
    for cat, real, repl in rows:
        ws.append([cat, real, repl, "replaced", "spreadsheet", 1])
    wb.save(path)
    return path


# ── parsing the cell ─────────────────────────────────────────────────────────

def test_alias_target_reads_the_bare_form():
    assert P._pn_alias_target("=ANTIONIO") == "ANTIONIO"
    assert P._pn_alias_target("  =  ANTIONIO  ") == "ANTIONIO"


def test_alias_target_reads_excels_quoted_form():
    # Excel refuses `=ANTIONIO SARKISYAN` as a malformed formula, so a
    # MULTI-WORD canonical has to be typed as a text formula. Same instruction.
    assert P._pn_alias_target('="ANTIONIO SARKISYAN"') == "ANTIONIO SARKISYAN"
    assert P._pn_alias_target("=“ANTIONIO SARKISYAN”") == "ANTIONIO SARKISYAN"


def test_alias_target_refuses_a_real_formula_and_a_plain_cell():
    for cell in ("=SUM(A1:A2)", "=B2+1", "=", "=123", "no", "never",
                 "[Human Resources]", "Keswick Bexley", "", None):
        assert P._pn_alias_target(cell) is None


def test_word_pairs_pair_by_position_then_by_nearest():
    assert P._pn_alias_word_pairs("ANTIONO SARKISYAN", "ANTIONIO SARKISYAN") == [
        ("ANTIONO", "ANTIONIO"), ("SARKISYAN", "SARKISYAN")]
    # Unequal counts: each word takes its nearest unclaimed counterpart.
    assert P._pn_alias_word_pairs("ANTIONO", "ANTIONIO SARKISYAN") == [
        ("ANTIONO", "ANTIONIO")]


def test_mirror_op_names_the_slip():
    assert P._pn_mirror_op("antionio", "antiono") == ("del", 1)
    assert P._pn_mirror_op("antiono", "antionio") == ("ins", 1)
    assert P._pn_mirror_op("adler", "alder") == ("trans", 1)
    assert P._pn_mirror_op("palladino", "pallodino") == ("sub", 1)


# ── the derivation ───────────────────────────────────────────────────────────

def test_fold_onto_mirrors_the_slip_and_stays_distinct():
    reg = P._PnFakeRegistry()
    canon = reg.token("ANTIONIO", P._PN_NAME_WORDS, "nametok")
    got = reg.fold_onto("antiono", "antionio", canon, "nametok")
    assert got and got != canon                    # never the SAME fake
    assert P._pn_osa_distance(got.lower(), canon.lower()) == 1
    assert len(got) == len(canon) - 1              # the real lost a letter, so
                                                   # the stand-in does too


def test_fold_onto_refuses_a_pair_that_is_not_a_misspelling():
    # Five letters apart is not a slip — `_PN_FOLD_MAX_REPS`.
    reg = P._PnFakeRegistry()
    canon = reg.token("ANTIONIO", P._PN_NAME_WORDS, "nametok")
    assert reg.fold_onto("ant", "antionio", canon, "nametok") is None


# ── the key's Replacement column ─────────────────────────────────────────────

def test_load_key_parses_the_alias(tmp_path):
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("person-token", "ANTIONIO", "Barlowe"),
                     ("person-token", "ANTIONO", "=ANTIONIO")])
    reg = P._PnFakeRegistry()
    terms, decisions = P._pn_load_key(kp, reg, log)
    assert [t.real for t in terms] == ["ANTIONIO"]      # no term from the alias
    d = decisions["antiono"]
    assert d["type"] == "ALIAS" and d["fix"] == "yes" and d["alias"] == "ANTIONIO"
    assert d["replacement"] is None and d["fake_values"] is None


def test_load_key_recovers_the_alias_excel_stored_as_a_formula(tmp_path):
    # Typed into Excel, `=ANTIONIO` is a FORMULA: an ordinary data_only read
    # hands back nothing at all (or "#NAME?"), so the instruction survives only
    # in the formula itself — which is what `_pn_xl_typed_text` goes and reads.
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("person-token", "ANTIONIO", "Barlowe"),
                     ("person-token", "ANTIONO", "PLACEHOLDER")])
    wb = openpyxl.load_workbook(kp)
    wb.active["C3"] = "=ANTIONIO"          # openpyxl stores this as a formula
    wb.save(kp)
    assert openpyxl.load_workbook(kp, data_only=True).active["C3"].value is None

    reg = P._PnFakeRegistry()
    _terms, decisions = P._pn_load_key(kp, reg, log)
    assert decisions["antiono"]["alias"] == "ANTIONIO"


def test_an_excel_error_cell_is_never_a_replacement(tmp_path):
    # "#NAME?" is what an alias looks like when the formula could not be read
    # back. Taking it as a typed replacement would put "#NAME?" in the export.
    kp = _write_key(tmp_path / "pseudonym_key.xlsx",
                    [("person", "Jane Doe", "#NAME?")])
    reg = P._PnFakeRegistry()
    terms, decisions = P._pn_load_key(kp, reg, log)
    assert terms == [] and decisions == {}


# ── the LEAKS worksheet's Fix? column ────────────────────────────────────────

def test_decision_rows_parse_the_alias():
    rows = [["Value", "Type", "Fix? (yes/no)", "Notes"],
            ["ANTIONO", "LEAK", "=ANTIONIO", ""]]
    d = P._pn_parse_decision_rows(rows)["antiono"]
    assert d["fix"] == "yes" and d["alias"] == "ANTIONIO"
    assert d["fixcell"] == "=ANTIONIO"        # round-trips back to the sheet
    # An alias is NOT a keep, so it never reaches the cross-folder KEEP sheet.
    assert not P._pn_decision_is_keep(d)


def test_decision_rows_refuse_an_excel_error_cell():
    rows = [["Value", "Type", "Fix? (yes/no)", "Notes"],
            ["ANTIONO", "LEAK", "#NAME?", ""]]
    d = P._pn_parse_decision_rows(rows)["antiono"]
    assert d["fix"] == "" and d["replacement"] is None      # left undecided


# ── applying it ──────────────────────────────────────────────────────────────

def _fake_of(terms, real):
    return next(t.fake for t in terms if t.real == real)


def test_alias_gives_one_person_two_spellings_of_one_fake():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["ANTIONIO SARKISYAN"], [], [], registry=reg)
    decisions = {"antiono sarkisyan": {
        "value": "ANTIONO SARKISYAN", "fix": "yes", "alias": "ANTIONIO SARKISYAN",
        "fake_values": None, "replacement": None}}
    terms, values = P._pn_apply_aliases(decisions, terms, reg, log)
    terms += P._pn_build_terms([], [], values, registry=reg)

    canon = _fake_of(terms, "ANTIONIO SARKISYAN")
    alias = _fake_of(terms, "ANTIONO SARKISYAN")
    assert canon != alias                              # two reversible rows
    # The SURNAME is untouched — the words the two spellings share are already
    # one binding, so an alias costs one pool word, not two.
    assert canon.split()[-1] == alias.split()[-1]
    # …and the given names are one slip apart, the same slip the reals are.
    assert P._pn_osa_distance(canon.split()[0].lower(),
                              alias.split()[0].lower()) == 1


def test_alias_re_derives_the_composed_row_the_moved_word_built():
    # The alias is typed on the TOKEN row; the composed full-name row in the
    # same key would otherwise go on applying its stored fake, which is the
    # half-applied fix that makes one party read as two.
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["ANTIONIO SARKISYAN", "ANTIONO SARKISYAN"],
                              [], [], registry=reg)
    stale = _fake_of(terms, "ANTIONO SARKISYAN")
    decisions = {"antiono": {"value": "ANTIONO", "fix": "yes",
                             "alias": "ANTIONIO", "fake_values": None,
                             "replacement": None}}
    terms, values = P._pn_apply_aliases(decisions, terms, reg, log)
    assert "ANTIONO SARKISYAN" in values            # handed back for rebuilding
    assert not any(t.real == "ANTIONO SARKISYAN" for t in terms)
    terms += P._pn_build_terms([], [], values, registry=reg)
    assert _fake_of(terms, "ANTIONO SARKISYAN") != stale
    assert (_fake_of(terms, "ANTIONO SARKISYAN").split()[-1]
            == _fake_of(terms, "ANTIONIO SARKISYAN").split()[-1])


def test_an_unresolvable_alias_still_fakes_the_value():
    # The canonical is not bound in this case, so there is nothing to mirror.
    # The value is a LEAK the operator answered, so it is still faked — refusing
    # it silently would leave the real name standing in the export.
    reg = P._PnFakeRegistry()
    decisions = {"antiono": {"value": "ANTIONO", "fix": "yes",
                             "alias": "NOBODY AT ALL", "fake_values": None,
                             "replacement": None}}
    terms, values = P._pn_apply_aliases(decisions, [], reg, log)
    assert values == ["ANTIONO"]
    terms += P._pn_build_terms([], [], values, registry=reg)
    assert _fake_of(terms, "ANTIONO") != "ANTIONO"


def test_fix_leaks_never_moves_a_binding_that_already_exists():
    # `allow_rebind=False`: the exports are already scrubbed under the stand-in
    # this would move, and that pass cannot rewrite them.
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["ANTIONIO", "ANTIONO"], [], [], registry=reg)
    before = _fake_of(terms, "ANTIONO")
    decisions = {"antiono": {"value": "ANTIONO", "fix": "yes",
                             "alias": "ANTIONIO", "fake_values": None,
                             "replacement": None}}
    kept, values = P._pn_apply_aliases(decisions, terms, reg, log,
                                       allow_rebind=False)
    assert kept == terms and values == ["ANTIONO"]
    assert _fake_of(kept, "ANTIONO") == before


# ── end to end through --fix-leaks ───────────────────────────────────────────

class _Args:
    key = None
    term = None


def _leaks_setup(folder, fix):
    tdir = folder / "Text Files"
    tdir.mkdir()
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["ANTIONIO SARKISYAN"], [], [], registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    pz.apply("ANTIONIO SARKISYAN filed suit.")
    pz.write_key(folder / "pseudonym_key.xlsx", log)
    (tdir / "Opposition.txt.LEAK").write_text(
        "====== Page 1 ======\nANTIONO SARKISYAN was served.\n", encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File", "Type", "Value", "Where", "Fix? (yes/no)", "Notes"])
    ws.append(["Opposition.txt.LEAK", "LEAK", "ANTIONO SARKISYAN", "p.1:2",
               fix, ""])
    wb.save(folder / "LEAKS.xlsx")
    return tdir


def _key_rows(folder):
    wb = openpyxl.load_workbook(folder / "pseudonym_key.xlsx")
    return {r[1]: r[2] for r in wb["Pseudonym Key"].iter_rows(
        min_row=2, values_only=True) if r[1]}


def test_fix_leaks_applies_a_worksheet_alias(tmp_path):
    tdir = _leaks_setup(tmp_path, "=ANTIONIO SARKISYAN")
    args = _Args()
    args.key = str(tmp_path / "pseudonym_key.xlsx")
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0

    out = (tdir / "Opposition.txt").read_text()
    assert "ANTIONO" not in out and "SARKISYAN" not in out
    rows = _key_rows(tmp_path)
    canon, alias = rows["ANTIONIO SARKISYAN"], rows["ANTIONO SARKISYAN"]
    assert canon != alias                              # each reverses on its own
    assert canon.split()[-1] == alias.split()[-1]      # one surname, one party
    assert P._pn_osa_distance(canon.split()[0].lower(),
                              alias.split()[0].lower()) == 1
    assert alias in out                                # and it is what shipped


def test_fix_leaks_refuses_an_alias_typed_into_the_key(tmp_path):
    # A key alias MOVES a binding the exports already carry, and this pass never
    # re-reads the PDFs — so it changes nothing at all and says to re-run.
    tdir = _leaks_setup(tmp_path, "no")
    kp = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.load_workbook(kp)
    ws = wb["Pseudonym Key"]
    ws.append(["person-token", "ANTIONO", "=ANTIONIO", "replaced",
               "spreadsheet", 0])
    wb.save(kp)
    before_key, before_txt = _key_rows(tmp_path), (
        tdir / "Opposition.txt.LEAK").read_text()

    args = _Args()
    args.key = str(kp)
    assert P._fix_leaks_mode(tmp_path, args, {}, log) == 0
    assert (tdir / "Opposition.txt.LEAK").exists()      # still quarantined
    assert (tdir / "Opposition.txt.LEAK").read_text() == before_txt
    assert _key_rows(tmp_path) == before_key            # key untouched
    assert (tmp_path / "LEAKS.xlsx").is_file()          # worksheet still there


# ── end to end through a full run (main) with a real PDF ─────────────────────

def _make_pdf(path, lines):
    import fitz
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    y = 100
    for ln in lines:
        page.insert_text((72, y), ln, fontsize=11)
        y += 24
    doc.save(path)
    doc.close()


def test_full_run_honours_an_alias_typed_into_the_key(tmp_path, monkeypatch):
    import sys
    _make_pdf(tmp_path / "Motion.pdf",
              ["Plaintiff ANTIONIO SARKISYAN filed this motion.",
               "The complaint names ANTIONO SARKISYAN as the same party."])
    key = _write_key(tmp_path / "pseudonym_key.xlsx", [
        ("person", "ANTIONIO SARKISYAN", "Cranfield Marlowe"),
        ("person-token", "ANTIONIO", "Cranfield"),
        ("person-token", "SARKISYAN", "Marlowe"),
        ("person", "ANTIONO SARKISYAN", "=ANTIONIO SARKISYAN")])
    monkeypatch.setattr(sys, "argv",
                        ["pdf_linker.py", str(tmp_path), "--key", str(key)])
    P.main()

    txt = "\n".join(p.read_text()
                    for p in (tmp_path / "Text Files").glob("*.txt")).lower()
    assert "antionio" not in txt and "antiono" not in txt   # both spellings gone
    rows = _key_rows(tmp_path)
    alias = rows["ANTIONO SARKISYAN"].lower()
    assert alias != "cranfield marlowe"                     # its own reversal row
    assert alias.split()[-1] == "marlowe"                   # one surname
    assert P._pn_osa_distance(alias.split()[0], "cranfield") == 1
    assert alias in txt                                     # and it is what shipped
