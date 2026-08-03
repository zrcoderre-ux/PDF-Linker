"""
Leak-triage worksheet: every potential leak the run flags is collected into
LEAKS.xlsx, each row located to the printed page and gutter line,
with a blank Fix? column for the reviewer. A clean run leaves no worksheet.

Column order is Value, Fix?, File, Type, Where, Notes — the flagged value and
its decision lead, the locating detail trails.

Run:  cd PDF-Linker && python3 -m pytest tests/test_leak_report.py -v
"""
import logging
from pathlib import Path

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

BODY = """====== Page 1 ======
 1  Attorneys for Defendant Travelers appeared.
 2  vs.
====== Page 2 (printed p. 5) ======
16  Defendant Travelers joined the motion. SBN 175977.
17  continuation line about Travelers again
====== Authorities cited (public verification links) ======
Donlen v. Ford Motor Co. -> http://scholar..."""


def _sheet(path):
    """[{header: value}] for the worksheet — read by header NAME, exactly as
    `_pn_read_leak_decisions` does, so adding a column never breaks a test on a
    positional index."""
    rows = list(openpyxl.load_workbook(path).active.iter_rows(values_only=True))
    return [dict(zip([str(h) for h in rows[0]], r)) for r in rows[1:]]


def test_locate_reports_printed_page_and_gutter_line():
    parsed = P._pn_body_lines(BODY)
    assert P._pn_locate(parsed, "Travelers") == "p.1:1, p.5:16, p.5:17"
    assert P._pn_locate(parsed, "175977") == "p.5:16"
    assert P._pn_locate(parsed, "absent-value") == "(not located)"


def test_worksheet_has_fix_column_and_severity_order(tmp_path):
    entries = [
        {"file": "Motion.txt", "type": "unscrubbed name?",
         "value": "Sunrise Motors", "where": "p.3:4"},
        {"file": "Motion.txt", "type": "LEAK", "value": "Travelers",
         "where": "p.5:16"},
        {"file": "Motion.txt", "type": "REID bar number", "value": "175977",
         "where": "p.5:16"},
    ]
    P._pn_write_leak_report(tmp_path, entries, log)
    xp = tmp_path / "LEAKS.xlsx"
    assert xp.exists()
    wb = openpyxl.load_workbook(xp)
    assert wb.active.title == "LEAKS"
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("Value", "Fix? (yes/no)", "Context", "File", "Type",
                       "Where (page:line)", "Notes")
    body = _sheet(xp)
    # real leaks first, then map-inverting REID, then ordinary review
    assert [r["Type"] for r in body] == ["LEAK", "REID bar number",
                                         "unscrubbed name?"]
    # each row is located, and the Fix? column is blank for the reviewer
    assert body[0]["Where (page:line)"] == "p.5:16"
    assert body[0]["Fix? (yes/no)"] in (None, "")


def test_clean_run_removes_stale_worksheet(tmp_path):
    xp = tmp_path / "LEAKS.xlsx"
    P._pn_write_leak_report(tmp_path, [{"file": "x", "type": "LEAK",
                                        "value": "v", "where": "p.1:1"}], log)
    assert xp.exists()
    P._pn_write_leak_report(tmp_path, [], log)      # a clean re-run
    assert not xp.exists()


def test_pseudonymizer_collects_located_findings():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Alejandro Orellana"], ["24STCV06764"], [],
                              registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    z = P.Pseudonymizer(terms, det, registry=reg)
    assert z.leak_report == []       # starts empty, accumulates during a run


def test_decisions_read_back_from_annotated_worksheet(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "M.txt", "type": "LEAK", "value": "Travelers", "where": "p.5:16"},
        {"file": "M.txt", "type": "possible business (purchase)",
         "value": "Worthington Motors", "where": "p.9:2"},
    ], log)
    xp = tmp_path / "LEAKS.xlsx"
    wb = openpyxl.load_workbook(xp)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Travelers":
            row[1].value = "yes"
        if row[0].value == "Worthington Motors":
            row[1].value = "No"          # case/space-insensitive
    wb.save(xp)

    dec = P._pn_read_leak_decisions(tmp_path)
    assert dec["travelers"]["fix"] == "yes"
    assert dec["worthington motors"]["fix"] == "no"


def test_roundtrip_persists_yes_suppresses_no_and_surfaces_new(tmp_path):
    decisions = {
        "travelers": {"value": "Travelers", "type": "LEAK", "fix": "yes",
                      "notes": ""},
        "worthington motors": {"value": "Worthington Motors",
                               "type": "possible business (purchase)",
                               "fix": "no", "notes": ""},
    }
    # This run: Travelers was scrubbed (gone), Worthington still present, and a
    # brand-new undecided finding appears.
    entries = [
        {"file": "M.txt", "type": "possible business (purchase)",
         "value": "Worthington Motors", "where": "p.9:2"},
        {"file": "M.txt", "type": "unscrubbed name?", "value": "New Corp",
         "where": "p.2:1"},
    ]
    P._pn_write_leak_report(tmp_path, entries, log, decisions)
    body = _sheet(tmp_path / "LEAKS.xlsx")
    by_val = {r["Value"]: r for r in body}
    # a marked-yes value that's now gone is RETAINED so the term keeps applying
    assert by_val["Travelers"]["Fix? (yes/no)"] == "yes"
    assert by_val["Travelers"]["Where (page:line)"] == P._PN_LEAK_ABSENT
    # a marked-NO value is a KEEP: it moves to the cross-folder master KEEP
    # sheet (handled by the run), so it no longer clutters this transient
    # per-folder triage.
    assert "Worthington Motors" not in by_val
    # a new undecided finding is blank and sorts ABOVE the resolved rows
    assert by_val["New Corp"]["Fix? (yes/no)"] in (None, "")
    order = [r["Value"] for r in body]
    assert order.index("New Corp") < order.index("Travelers")


def test_same_value_across_files_is_one_row(tmp_path):
    # A name that leaks in several files must appear ONCE — the operator
    # decides it a single time, not per file. Files and locations aggregate.
    entries = [
        {"file": "Complaint.pdf", "type": "LEAK", "value": "ca.gov",
         "where": "p.1, p.appendix"},
        {"file": "Motion.pdf", "type": "LEAK", "value": "ca.gov",
         "where": "p.appendix"},
        {"file": "Opposition.pdf", "type": "LEAK", "value": "ca.gov",
         "where": "p.C"},
        {"file": "RJN.pdf", "type": "LEAK", "value": "ca.gov", "where": "p.2"},
    ]
    P._pn_write_leak_report(tmp_path, entries, log)
    body = _sheet(tmp_path / "LEAKS.xlsx")
    assert len(body) == 1                       # one row for the four files
    row = body[0]
    assert row["Value"] == "ca.gov"
    assert row["File"] == "4 files"             # aggregated (>3 → a count)
    for loc in ("p.1", "p.appendix", "p.C", "p.2"):
        assert loc in row["Where (page:line)"]  # every location preserved


def test_merged_row_keeps_the_most_severe_type(tmp_path):
    entries = [
        {"file": "A.pdf", "type": "possible person name", "value": "M M",
         "where": "p.1"},
        {"file": "B.pdf", "type": "LEAK", "value": "M M", "where": "p.2"},
    ]
    P._pn_write_leak_report(tmp_path, entries, log)
    body = _sheet(tmp_path / "LEAKS.xlsx")
    assert len(body) == 1
    assert body[0]["Value"] == "M M" and body[0]["Type"] == "LEAK"
    assert body[0]["File"] == "A.pdf, B.pdf"    # <=3 files listed by name


def test_one_decision_covers_every_occurrence(tmp_path):
    # Mark the single merged row 'yes' once; on the next run the decision
    # applies to the value regardless of how many files carry it.
    entries = [{"file": f"F{i}.pdf", "type": "LEAK", "value": "ca.gov",
                "where": "p.1"} for i in range(5)]
    P._pn_write_leak_report(tmp_path, entries, log)
    xp = tmp_path / "LEAKS.xlsx"
    wb = openpyxl.load_workbook(xp)
    ws = wb.active
    body = list(ws.iter_rows(min_row=2))
    assert len(body) == 1
    body[0][1].value = "yes"                    # Fix? is column B
    wb.save(xp)
    dec = P._pn_read_leak_decisions(tmp_path)
    assert dec["ca.gov"]["fix"] == "yes"


def test_legacy_worksheet_name_is_still_read(tmp_path):
    # A folder triaged under the old name keeps its decisions: the reader
    # falls back to pdf_linker_leaks.xlsx when LEAKS.xlsx is absent.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(P._PN_LEAK_HEADERS))
    ws.append(["Travelers", "yes", "M.txt", "LEAK", "p.5:16", ""])
    wb.save(tmp_path / "pdf_linker_leaks.xlsx")
    dec = P._pn_read_leak_decisions(tmp_path)
    assert dec["travelers"]["fix"] == "yes"


def test_writing_migrates_off_the_legacy_name(tmp_path):
    # Once the report is rewritten, the stale old-named file is removed so the
    # operator's decisions don't split across two worksheets.
    (tmp_path / "pdf_linker_leaks.xlsx").write_bytes(b"stale")
    P._pn_write_leak_report(tmp_path, [{"file": "M.txt", "type": "LEAK",
                                        "value": "v", "where": "p.1:1"}], log)
    assert (tmp_path / "LEAKS.xlsx").exists()
    assert not (tmp_path / "pdf_linker_leaks.xlsx").exists()


def test_suppressed_value_excluded_from_gate():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["AHC Acquisition, LLC"], ["24STCV06764"], [],
                              registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    z = P.Pseudonymizer(terms, det, registry=reg)
    z.note_leaks({"ahc acquisition, llc"})
    z.suppressed = {"ahc acquisition, llc"}
    # the gate's expression must exclude a suppressed value
    gating = {v for v in z.primary_leaks()
              if str(v).lower() not in z.suppressed}
    assert gating == set()


# ── a resolved decision does not regenerate the worksheet ────────────────────
# A Fix?=yes mints a fake that the pseudonym KEY pins, and every later run
# re-applies it — so the value is legitimately absent from then on. Carrying
# the row forward preserved nothing and rebuilt LEAKS.xlsx on every clean run,
# a sheet whose only content was "(no longer present)".

def test_key_bound_decision_is_not_carried_forward(tmp_path):
    folder = tmp_path / "Case"
    folder.mkdir()
    decisions = {"travelers casualty": {
        "value": "TRAVELERS CASUALTY", "type": "unscrubbed name?", "fix": "yes",
        "replacement": None, "fake_values": None, "fixcell": "yes",
        "notes": ""}}
    # no findings this run, and the key already binds the value
    P._pn_write_leak_report(folder, [], log, decisions=decisions,
                            bound=["TRAVELERS CASUALTY"])
    assert not list(folder.glob("LEAKS.xlsx"))


def test_unbound_decision_is_still_carried_forward(tmp_path):
    folder = tmp_path / "Case"
    folder.mkdir()
    decisions = {"travelers casualty": {
        "value": "TRAVELERS CASUALTY", "type": "unscrubbed name?", "fix": "yes",
        "replacement": None, "fake_values": None, "fixcell": "yes",
        "notes": ""}}
    # nothing else holds it, so the instruction must survive in the worksheet
    P._pn_write_leak_report(folder, [], log, decisions=decisions, bound=[])
    sheet = folder / "LEAKS.xlsx"
    assert sheet.is_file()
    vals = [str(r[0]) for r in
            openpyxl.load_workbook(sheet).active.iter_rows(min_row=2,
                                                           values_only=True)]
    assert "TRAVELERS CASUALTY" in vals


# ───── a triage row names the authority it may have come from ───────────────

def _cited_pz():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Weishi Yang", "Ashley Liu"], ["26STCV08967"],
                              [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    z = P.Pseudonymizer(terms, det, registry=reg)
    z.note_authority_cites(
        "Defendant relies on Kremerman v. White (2021) 71 Cal.App.5th 358, 372, "
        "and on Pasadena Medi-Center Associates v. Superior Court (1973) 9 "
        "Cal.3d 773.")
    return z


def test_a_value_sharing_a_cited_partys_name_is_annotated():
    """"Angela White" in a worksheet is a question the operator can only answer
    by already knowing the authorities. The note IS the answer."""
    note = _cited_pz().authority_note("Angela White")
    assert note == ("cited authority: Kremerman v. White (2021) "
                    "71 Cal.App.5th 358"), note


def test_an_unrelated_value_gets_no_note():
    assert _cited_pz().authority_note("Bartholomew Quillfeather") == ""


def test_only_a_year_bearing_cite_contributes():
    # A bare "X v. Y" is as likely to be the document's own caption.
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer([], {}, registry=reg)
    z.note_authority_cites("HELEN RASHO v. GENERAL MOTORS, LLC\n  Defendant.")
    assert z.authority_note("General Motors") == ""


def test_the_note_reaches_the_worksheet(tmp_path):
    z = _cited_pz()
    P._pn_write_leak_report(tmp_path, [
        {"file": "Quash.pdf", "type": "unscrubbed name?",
         "value": "Angela White", "where": "p.2:20"},
        {"file": "Complaint.pdf", "type": "LEAK",
         "value": "Bartholomew Quillfeather", "where": "p.1:4"},
    ], log, note_for=z.authority_note)
    rows = {r[0]: r[6] for r in openpyxl.load_workbook(
        tmp_path / "LEAKS.xlsx").active.iter_rows(min_row=2, values_only=True)}
    assert "Kremerman v. White" in str(rows["Angela White"])
    assert not str(rows["Bartholomew Quillfeather"] or "")


def test_the_row_is_still_shown_not_suppressed(tmp_path):
    # Sharing a surname with a cited decision is not proof: a real witness can
    # be called White in a case that cites *Kremerman v. White*. The note
    # informs the operator; it does not decide for them.
    z = _cited_pz()
    P._pn_write_leak_report(tmp_path, [
        {"file": "Quash.pdf", "type": "unscrubbed name?",
         "value": "Angela White", "where": "p.2:20"},
    ], log, note_for=z.authority_note)
    values = [r[0] for r in openpyxl.load_workbook(
        tmp_path / "LEAKS.xlsx").active.iter_rows(min_row=2, values_only=True)]
    assert "Angela White" in values


def test_an_operators_own_note_survives(tmp_path):
    z = _cited_pz()
    P._pn_write_leak_report(tmp_path, [
        {"file": "Quash.pdf", "type": "unscrubbed name?",
         "value": "Angela White", "where": "p.2:20"},
    ], log, decisions={"angela white": {"value": "Angela White", "fix": "",
                                        "notes": "checked with counsel"}},
        note_for=z.authority_note)
    note = [r[6] for r in openpyxl.load_workbook(
        tmp_path / "LEAKS.xlsx").active.iter_rows(min_row=2, values_only=True)][0]
    assert "checked with counsel" in note and "Kremerman v. White" in note


# ── the Context column ──────────────────────────────────────────────────────
# A triage row asks "is this real?", and the value alone often cannot answer
# it. "Charge" is boilerplate in "CHARGE OF DISCRIMINATION" and a surname in
# "served on Charge at his residence", and the answer decides whether the word
# is faked in this and every future folder.

PLEADING = """====== Page 4 ======
 1  NOTICE OF MOTION AND MOTION TO QUASH SERVICE OF SUMMONS
 2  CHARGE OF DISCRIMINATION AND RIGHT TO SUE
 3  Plaintiff alleges that the summons was served on Charge at his
 4  residence in Montebello on March 3, 2024. (Rasho Decl. ¶ 7.)
"""


def _ctx(needle, body=PLEADING):
    return P._pn_context(P._pn_body_lines(body), needle)


def test_the_sentence_is_quoted_not_the_line():
    """A sentence on pleading paper is spread over numbered lines, so one line
    says almost nothing — and the gutter number is furniture."""
    got = _ctx("Montebello")
    assert got == ("Plaintiff alleges that the summons was served on Charge at "
                   "his residence in Montebello on March 3, 2024.")
    assert "  4  " not in got and "\n" not in got


def test_prose_beats_a_heading():
    """The first occurrence of "Charge" is a caption and proves nothing; the
    one that decides the row is the sentence."""
    assert _ctx("Charge").startswith("Plaintiff alleges")


def test_a_heading_only_value_is_quoted_as_its_heading():
    """Which is itself the answer: nothing but a title ever offered it."""
    assert _ctx("QUASH") == ("NOTICE OF MOTION AND MOTION TO QUASH SERVICE OF "
                             "SUMMONS CHARGE OF DISCRIMINATION AND RIGHT TO SUE")


def test_a_caption_is_not_swallowed_by_the_sentence_beside_it():
    assert "NOTICE OF MOTION" not in _ctx("Montebello")


def test_an_abbreviation_does_not_end_the_quote_early():
    """"(Rasho Decl." is not evidence of anything: legal prose is full of
    abbreviations, so a span too short to read grows into its neighbours."""
    got = _ctx("Rasho")
    assert "Rasho Decl." in got and len(got) > 60


def test_a_long_span_is_windowed_around_the_value():
    body = "====== Page 1 ======\n 1  " + ("filler words here " * 40) + "Wexford ok\n"
    got = _ctx("Wexford", body)
    assert "Wexford" in got and len(got) <= P._PN_CONTEXT_MAX + 2
    assert got.startswith("…")


def test_a_value_that_cannot_be_located_has_no_quote():
    """A welded or reduced finding has no clean line to quote, exactly as it
    has no location."""
    assert _ctx("Nowhere") == ""


def test_the_column_sits_between_the_decision_and_the_file(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "Motion.txt", "type": "LEAK", "value": "Charge",
         "where": "p.4:3", "context": "served on Charge at his residence"},
    ], log)
    rows = list(openpyxl.load_workbook(tmp_path / "LEAKS.xlsx")
                .active.iter_rows(values_only=True))
    assert rows[0][:4] == ("Value", "Fix? (yes/no)", "Context", "File")
    assert rows[1][2] == "served on Charge at his residence"


def test_one_sample_sentence_per_merged_row(tmp_path):
    """The row already merges every file and location; a cell holding nine
    sentences is one nobody reads."""
    P._pn_write_leak_report(tmp_path, [
        {"file": "A.txt", "type": "LEAK", "value": "Charge", "where": "p.1",
         "context": "first sentence"},
        {"file": "B.txt", "type": "LEAK", "value": "Charge", "where": "p.2",
         "context": "second sentence"},
    ], log)
    body = _sheet(tmp_path / "LEAKS.xlsx")
    assert len(body) == 1 and body[0]["Context"] == "first sentence"
