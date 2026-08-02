"""
Regression suite from the pseudonym-replacement review (2026-07-10).

Each test encodes ONE acceptance criterion for a defect found by comparing the
delivered pseudonymized .txt exports against the originals for
`25STCV37838 Estrada v. Azul Concreto`. Every test in the REVIEW section is
expected to FAIL against the pre-fix code and PASS once the corresponding fix
in HANDOFF.md lands. The GUARD section pins behavior that is already correct
and must not regress while the fixes are made.

Run:  cd PDF-Linker && python3 -m pytest tests/test_pseudonym_review.py -v
"""

import logging
import re
import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _pz(names=("Roxane Estrada", "Azul Concreto, Inc."),
        casenos=("25STCV37838",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), list(casenos), [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


# ─────────────────────────── REVIEW (red now) ───────────────────────────────

# P0-1  Defendant identity must not split into two unrelated fakes.
#   Delivered set called the same corporation both "Beacon Torchlight, Inc."
#   (entity pool) and "Nolan Isley, Inc." / "(Nolan)" (person pool), because a
#   word maps to a DIFFERENT fake depending on which pool asks for it.
#   Fix: a token already bound in the entity pool must reuse that fake when the
#   person path later meets the same word (cross-pool reconciliation), and a
#   suffix-less corporate name must not take the person path.
def test_token_fake_is_stable_across_person_and_entity_pools():
    reg = P._PnFakeRegistry()
    ent = reg.token("azul", P._PN_ENTITY_WORDS, "enttok")
    nam = reg.token("azul", P._PN_NAME_WORDS, "nametok")
    assert ent == nam, (
        f"'azul' got {ent!r} as an entity but {nam!r} as a person; the "
        "defendant renders under two identities in the same document set")


def test_defendant_bare_form_reuses_entity_fake():
    # "Azul Concreto, Inc." is the entity (from the spreadsheet). When a
    # document pass later meets the name WITHOUT its suffix and takes the person
    # path, it must reuse the entity's per-word fakes — never mint an unrelated
    # person ("Nolan Isley"). Every faked word of the bare person form must
    # therefore appear in the entity fake.
    reg = P._PnFakeRegistry()
    ent_fake, _map = P._pn_fake_entity_parts("Azul Concreto, Inc.", reg)
    person_fake, _bare = P._pn_fake_person("Azul Concreto", reg)
    for w in person_fake.split():
        assert w in ent_fake, (
            f"bare person form {person_fake!r} diverges from entity {ent_fake!r}"
            "; the defendant would appear under two identities")


# P0-2  A spelled-out state must never be corrupted.
#   "…, Montebello, California 90640" was emitted as "…, Clearwater, ia 30070":
#   the two-letter state group ate the tail of "California", leaving "ia".
def test_spelled_out_state_survives_addressing():
    z = _pz()
    out = z._fake_street("414 S. Maple Ave., Montebello, California 90640")
    assert "California" in out and " ia " not in f" {out} "
    out2 = z._fake_street("100 Centerview Drive, Suite 205, Birmingham, Alabama 35216")
    assert "Alabama" in out2 and " ma " not in f" {out2} "


# P1  House standard (and the repo owner): do NOT fake city, state, or ZIP.
#   Only the street number, street name, and unit change. Keeping the locality
#   also removes every state-corruption and city-drift failure at the source.
def test_city_state_zip_preserved_verbatim():
    z = _pz()
    for real, city, state, zipc in [
        ("414 S. Maple Ave., Montebello, California 90640", "Montebello", "California", "90640"),
        ("414 S. Maple Ave. Montebello, CA 90640", "Montebello", "CA", "90640"),
        ("100 Centerview Drive, Suite 205, Birmingham, Alabama 35216", "Birmingham", "Alabama", "35216"),
    ]:
        out = z._fake_street(real)
        assert city in out, f"city {city!r} was altered in {out!r}"
        assert re.search(rf"\b{state}\b", out), f"state {state!r} was altered in {out!r}"
        assert zipc in out, f"ZIP {zipc!r} was altered in {out!r}"


# P1  One parcel, however spelled, gets exactly one fake street.
#   Delivered key mapped 414 S. Maple to Cypress, Birch, Linden AND Tamarack.
def test_one_parcel_one_street_fake():
    z = _pz()
    spellings = [
        "414 S. Maple Ave.",
        "414 S. Maple Ave. Montebello, CA 90640",
        "414 S. Maple Ave., Montebello, California 90640",
    ]
    def street_of(s):
        # leading "<number> <name>" of the faked street, ignoring any tail
        m = re.match(r"\s*(\d[\d\-\u2013\u2014]*\s+[A-Za-z]+)", z._fake_street(s))
        return m.group(1) if m else z._fake_street(s)
    fakes = {street_of(s) for s in spellings}
    assert len(fakes) == 1, f"one parcel produced multiple street fakes: {fakes}"


# P1  A hyphenated unit range is the same parcel as its base number.
def test_unit_range_matches_base_parcel():
    z = _pz()
    base = re.split(r",", z._fake_street("414 S. Maple Ave."))[0]
    rng = re.split(r",", z._fake_street("414-416 S. Maple Ave."))[0]
    # same faked street NAME (the range may keep a range in the number)
    name = lambda s: re.sub(r"^[\d\-\u2013\u2014 ]+", "", s).strip()
    assert name(base) == name(rng), f"{base!r} vs {rng!r} disagree on street name"


# P2  Distinct real e-mail domains must map to distinct fakes (injective).
#   Delivered key collapsed gmail.com, yahoo.com and the firm domain all onto
#   letterbox.co, implying opposing parties shared an e-mail domain.
def test_email_domains_injective():
    hosts = ["gmail.com", "yahoo.com", "schillecitortoricilaw.com",
             "themillenniallawyer.com", "outlook.com"]
    fakes = [P._pn_fake_domain(h) for h in hosts]
    assert len(set(fakes)) == len(hosts), (
        f"domain map is not injective: {dict(zip(hosts, fakes))}")


# P2  Public consumer providers carry no identity and should pass through.
#   (Recommendation — delete this test if the firm prefers to fake them, but
#   then test_email_domains_injective still governs uniqueness.)
def test_public_email_providers_preserved():
    for h in ["gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com"]:
        assert P._pn_fake_domain(h) == h, f"public provider {h} was faked"


# P2  The macro-reversible key must not carry rows that matched nothing.
#   ReAnonymize runs the key in reverse and will replace a Real Value that was
#   never in the documents. Build terms, apply to text that hits ONLY some of
#   them, write the key, and assert no surviving "no match" rows.
def test_reversal_key_zero_occurrence_rows_are_authoritative_only(tmp_path):
    # A value the PARTY TEMPLATE names is authoritative: its row is written even
    # when this batch never mentioned it, so the fake is pinned for the run that
    # finally meets that party (the "add the document I forgot" path). What must
    # NOT appear is a SYNTHETIC spelling this tool invented to widen matching —
    # a near-miss variant is a real value only if a document contained it.
    z = _pz(names=["Roxane Estrada", "Azul Concreto, Inc.", "Someone Neverpresent"])
    z.apply("Plaintiff Roxane Estrada sued Azul Concreto, Inc. in 25STCV37838.")
    out = tmp_path / "pseudonym_key.xlsx"
    z.write_key(out, log)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    occ_i, real_i, stat_i = (header.index("Occurrences"),
                             header.index("Real Value"), header.index("Status"))
    rows = [r for s in wb.worksheets
            for r in s.iter_rows(min_row=2, values_only=True) if r and r[0]]
    zero = [r for r in rows if r[occ_i] in (0, None)]
    assert all(r[stat_i] == "no match" for r in zero)
    reals = {str(r[real_i]) for r in rows}
    assert "Someone Neverpresent" in reals          # pinned though never seen
    # …but OFF the sheet the reversal macro reads, or the reverse pass would
    # rewrite a Real Value that was never in the document.
    assert "Someone Neverpresent" not in {
        str(r[real_i]) for r in ws.iter_rows(min_row=2, values_only=True)
        if r and r[0]}
    # every variant spelling of a name that matched nothing stays out
    invented = {v for tok in ("Estrada", "Roxane", "Neverpresent")
                for v in P._pn_name_variants(tok)}
    assert not (reals & invented), sorted(reals & invented)
    # The separate audit spreadsheet was retired (repo owner: unnecessary for
    # a casual-recognition precaution) — the key is the ONLY file written, and
    # a stale audit from an earlier version is cleaned up.
    assert not (tmp_path / "pseudonym audit.xlsx").exists()
    assert not (tmp_path / "pseudonym_key audit.xlsx").exists()


def test_reversal_key_omits_an_inferred_value_that_matched_nothing(tmp_path):
    # A harvested guess is NOT authoritative: if it matched nothing it was
    # probably never a party, and ReAnonymize must not rewrite it.
    z = _pz(names=["Roxane Estrada"])
    z._add_terms([P._PnTerm("person", "Hollis Vantreight", "Ashby Renwick",
                           whole_word=False, case_sensitive=False,
                           priority=2, source="declarant")])
    z.apply("Plaintiff Roxane Estrada appeared.")
    out = tmp_path / "pseudonym_key.xlsx"
    z.write_key(out, log)
    reals = {str(r[1]) for r in
             openpyxl.load_workbook(out).active.iter_rows(min_row=2,
                                                          values_only=True)}
    assert "Hollis Vantreight" not in reals


def test_stale_audit_sheet_is_removed(tmp_path):
    # A leftover "pseudonym audit.xlsx" from an earlier tool version must not
    # sit beside a fresh key looking current.
    z = _pz(names=["Roxane Estrada"])
    z.apply("Plaintiff Roxane Estrada sued in 25STCV37838.")
    stale = tmp_path / "pseudonym audit.xlsx"
    stale.write_bytes(b"old")
    z.write_key(tmp_path / "pseudonym_key.xlsx", log)
    assert not stale.exists()


# P1  Every name in a phone-roster block must be harvested, not just the first.
#   Delivered FAC leaked `Juan Olivas` and `Viviana Gomez` (only `Jose Gomez`
#   was caught by the CONTRACTORS: label).
def test_contractor_roster_all_names_harvested():
    block = ("CONTRACTORS:\n"
             "Jose Gomez (323) 283-4603\n"
             "Juan Olivas (562) 239-8134\n"
             "Viviana Gomez (562) 456-6420 office\n")
    got = P._pn_label_names(block)
    for n in ("Jose Gomez", "Juan Olivas", "Viviana Gomez"):
        assert n in got, f"{n!r} not harvested from the roster: {got}"


# P1  Same-person aliases faked with unrelated surnames are surfaced (not
#   silently merged) so a reviewer can link them. Delivered set faked Roxane's
#   Estrada/Guzman/Purscelley names as three unrelated surnames.
def test_alias_candidates_surfaced():
    z = _pz(names=["Roxane Estrada", "Roxane Guzman", "Jason Tortorici"])
    flat = {n for grp in z.alias_candidates() for n in grp}
    assert {"Roxane Estrada", "Roxane Guzman"} <= flat
    assert "Jason Tortorici" not in flat   # unique given name — never flagged


# ─────────────────────────── GUARD (green now) ──────────────────────────────
# Behavior that is already correct; these must stay green through the fixes.

def test_guard_venue_city_not_scrubbed():
    # "Los Angeles" as a county/venue is locality, never a party.
    assert P._pn_is_protected_locality("Los Angeles")
    assert P._pn_is_protected_locality("Los Angeles County")


def test_guard_caseno_filing_year_kept():
    reg = P._PnFakeRegistry()
    fake = reg.digits("25STCV37838", "case_number", keep_prefix=2)
    assert fake.startswith("25") and fake != "25STCV37838"


def test_guard_reservation_id_is_faked_digitwise():
    reg = P._PnFakeRegistry()
    fake = reg.digits("935605884885", "reservation_id")
    assert fake != "935605884885" and fake.isdigit() and len(fake) == 12


# ─────────────────────── REVIEW (25STCV28413) ───────────────────────────────
# Second corpus (Orellana v. AHC Acquisition LLC). Each test encodes one
# acceptance criterion from the 2026-07-13 fix handoff; red against pre-fix
# code, green after its fix.

def _pz28413(names=("Alejandro Orellana", "AHC Acquisition LLC"),
             casenos=("25STCV28413",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), list(casenos), [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


# P0-A  A published decision's party names must never be rewritten. The address
#   detector rewrote "100 Oak Street" (a real party) inside a citation; the
#   entity harvester rewrote "DMS Services", "KPMG", "Valencia Holding Co.".
def test_citation_party_names_survive_replacement():
    z = _pz28413()
    for cite in [
        "Ericksen, Arbuthnot, McCarthy, Kearney & Walsh, Inc. v. 100 Oak Street "
        "(1983) 35 Cal.3d 312",
        "DMS Services, LLC v. Superior Court (2012) 205 Cal.App.4th 1346",
        "Goldman v. KPMG, LLP (2009) 173 Cal.App.4th 209",
        "Sanchez v. Valencia Holding Co., LLC (2015) 61 Cal.4th 899",
    ]:
        assert z.apply(cite) == cite, f"cited decision was corrupted: {cite!r}"


def test_citation_survives_even_when_defendant_is_a_harvested_entity():
    # "Valencia Holding Co., LLC" is ALSO in the term set as a harvested entity;
    # the citation span must still win and leave the authority byte-identical.
    z = _pz28413(names=["Alejandro Orellana", "AHC Acquisition LLC",
                        "Valencia Holding Co., LLC"])
    cite = "Sanchez v. Valencia Holding Co., LLC (2015) 61 Cal.4th 899"
    assert z.apply(cite) == cite


def test_caption_exemption_replaces_own_case_caption():
    # Both sides name a trusted party -> this is the ruling's OWN caption and
    # must still be replaced, not protected as if it were an authority.
    z = _pz28413()
    out = z.apply("Alejandro Orellana v. AHC Acquisition LLC")
    assert "Orellana" not in out and "AHC Acquisition" not in out


# P0-C  Attorney State Bar numbers unmask the pseudonyms in one public lookup.
def test_bar_number_scrubbed_label_kept():
    z = _pz28413()
    for real, num in [("Holloway Vance, Esq., SBN 175977", "175977"),
                      ("State Bar No. 291925", "291925"),
                      ("JENNINGS R. LANGLEY, S.B.# 108076", "108076")]:
        z.register_identifiers(real)
        out = z.apply(real)
        assert num not in out, f"bar number {num} survived in {out!r}"
        assert "SBN" in out or "Bar" in out or "S.B." in out


# P1-G  A 17-char VIN is a strong unique identifier and must be faked.
def test_vin_scrubbed():
    z = _pz28413()
    real = 'the VIN 1C4JJXP65PW699184 ("Vehicle")'
    z.register_identifiers(real)
    out = z.apply(real)
    assert "1C4JJXP65PW699184" not in out
    m = re.search(r"VIN\s+([A-HJ-NPR-Z0-9]{17})", out)
    assert m, f"replacement is not a 17-char VIN-alphabet run: {out!r}"


# P1-H  Directional + two-word street + OCR-split suffix must be detected.
def test_directional_two_word_street_detected():
    z = _pz28413()
    for real in ["817 N. La Brea Avenue, City of Inglewood",
                 "817 N. La Brea A venue, City of Inglewood"]:
        out = z.apply(real)
        # The street NAME is the identifying part and is faked; the house
        # NUMBER is kept verbatim, like the suffix and the City/ST/ZIP tail.
        assert "La Brea" not in out, f"street name survived: {out!r}"
        assert "817" in out, f"house number should be kept: {out!r}"
        assert "La Brea" not in out, f"street name survived: {out!r}"
        assert "Inglewood" in out, f"locality was altered: {out!r}"


# P2-E  Fake e-mail domains stay valid, and an OCR-typo host folds to the firm's.
def test_fake_domains_valid_shape():
    for h in ["schillecitortoricilaw.com", "themillenniallawyer.com",
              "autolegalgroup.com", "adr.org", "quirklawoffice.com"]:
        f = P._pn_fake_domain(h)
        assert re.match(r"^[a-z0-9-]+(\.[a-z]{2,})+$", f), \
            f"invalid fake domain {f!r} (trailing digit on TLD?)"


def test_ocr_typo_host_folds_to_firm_domain():
    assert (P._pn_fake_domain("autolegalgrouo.com")
            == P._pn_fake_domain("autolegalgroup.com"))


# P2-F  A 10-digit run that is really an OCR'd date is not faked as a phone.
def test_ocr_date_not_faked_as_phone():
    z = _pz28413()
    for run in ["0610612025", "0612112031", "0712112025"]:
        out = z.apply(f"Contract dated {run}.")
        assert run in out, f"OCR date {run} was rewritten as a phone: {out!r}"
    # a genuine bare NANP number is still faked
    assert "6262920899" not in z.apply("Call 6262920899.")


# P2-I  Label-anchored account and reservation identifiers must be scrubbed.
def test_account_and_reservation_ids_scrubbed():
    z = _pz28413()
    for real, val in [("DEAL# 23071", "23071"), ("CUST# 24248", "24248"),
                      ("CR-BFA76WFGYHSBGBCGZ", "CR-BFA76WFGYHSBGBCGZ")]:
        z.register_identifiers(real)
        out = z.apply(real)
        assert val not in out, f"identifier {val} survived in {out!r}"


# Pipeline leak gate: any surviving real value marks a `leaked` row and makes
# has_leaks() true, which drives the run's non-zero exit + export quarantine.
def test_leak_gate_flags_surviving_real():
    z = _pz28413()
    assert not z.has_leaks()
    z.note_leaks({"Travelers"})
    assert z.has_leaks()
    rec = {"real": "Travelers", "count": 0}
    assert z._status(rec) == "leaked"


# ─────────────────────── REVIEW (24STCV06764) ───────────────────────────────
# Third corpus (Tom of Finland Foundation v. CultureEdit). Acceptance criteria
# from the 2026 leak handoff: a document-defined acronym and a column-spliced
# caption both leaked past the scrub.

def _pz06764(names=("Tom of Finland Foundation, Inc.", "CultureEdit, LLC"),
             casenos=("24STCV06764",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), list(casenos), [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


# P0  A document-defined acronym short form must be registered and scrubbed.
def test_acronym_short_form_registered_and_scrubbed():
    z = _pz06764()
    z.register_short_names('Plaintiff Tom of Finland Foundation, Inc. ("ToFF") '
                           "is a foundation.")
    out = z.apply("ToFF promotes the artist. See Toff and TOFF too.")
    assert "toff" not in out.lower(), f"acronym survived: {out!r}"
    # a short-name row exists for the acronym
    assert any(cat == "short-name" and rl == "toff"
               for (cat, rl) in z.records), "no short-name row for ToFF"


def test_acronym_definition_with_descriptor_and_alternatives():
    # "…, INC., A PUBLIC BENEFIT CORPORATION ("TOFF" OR "TOM OF FINLAND …")"
    z = _pz06764(names=["Tom of Finland Foundation"])
    z.register_short_names(
        'TOM OF FINLAND FOUNDATION, INC., A PUBLIC BENEFIT CORPORATION '
        '("TOFF" OR "TOM OF FINLAND FOUNDATION") filed suit.')
    assert "toff" not in z.apply("TOFF did things.").lower()


def test_acronym_over_match_guard_rejects_non_initialism():
    # "Team" is not an initialism of Dream Team Real Estate; the initialism
    # path must reject it (it remains covered only by the parent-word path).
    assert P._pn_initialism_fake(
        "Team", ["Dream", "Team", "Real", "Estate"],
        ["Falcon", "Ridge", "Vale", "Court"]) is None


def test_acronym_review_backstop_reports_leftover():
    # Registration missed the definition shape; the leftover party acronym in the
    # output is caught by the review backstop.
    z = _pz06764(names=["Tom of Finland Foundation, Inc."])
    src = 'TOM OF FINLAND FOUNDATION, INC. ("TOFF") sued.'
    out = 'Kaldor of Ironbridge Foundation ("TOFF") sued. TOFF acted.'
    found = z.review_definition_survivors(src, out)
    assert ("party acronym", "TOFF") in found
    # ordinary defined terms are never surfaced
    z.review = []
    assert not z.review_definition_survivors(
        'the vehicle ("Vehicle") and ("ISO") and ("RJN")',
        "the Vehicle and ISO and RJN survive")


# P0  A column-spliced caption welds a party name to its neighbours; the
#   reduced-span replacement must remove it (write side mirrors detection).
def test_spliced_welded_entity_is_scrubbed():
    z = _pz06764()
    welded = z.apply("Defendant CULTUREEDITservice of process and cultureeditllc.")
    # the boundary-anchored pass can't remove the welded token …
    assert z.surviving_reals_reduced(welded), "fixture no longer welds"
    cured = z.scrub_welded(welded)
    assert "cultureedit" not in cured.lower().replace(" ", "")
    assert not z.surviving_reals_reduced(cured), "reduced survivor remains"


def test_spliced_scrub_respects_citation_spans():
    # The loose (reduced-substring) replacer must honor the same citation
    # protection as the precise path: a cited decision whose party name
    # contains a tracked core survives byte-for-byte even on a splice-flagged
    # page, while a welded occurrence of the same party elsewhere is cured.
    z = _pz06764(names=["Valencia Holding Co., LLC", "Alejandro Orellana"])
    cite = "Sanchez v. Valencia Holding Co., LLC (2015) 61 Cal.4th 899"
    out = z.scrub_welded(
        cite + ". Defendant VALENCIAHOLDINGservice of process appears.")
    assert cite in out, "scrub_welded rewrote a cited authority"
    assert "VALENCIAHOLDINGservice" not in out, "welded name survived"


def test_spliced_replacement_recorded_in_key():
    # A welded name cured by scrub_welded must be recorded (Status "replaced"),
    # so the reversal key carries the mapping instead of parking it in the audit.
    z = _pz06764()
    rec = z.records[("entity", "cultureedit, llc")]
    assert z._status(rec) == "no match"
    z.scrub_welded("Defendant CULTUREEDITservice appears.")
    assert rec["count"] > 0 and z._status(rec) == "replaced"


# P2  known_fake_words() contributes the bare host of a faked e-mail/URL, so the
#   open-world review scan no longer flags the tool's own numbered fake domain.
def test_review_scan_ignores_own_fake_domain():
    z = _pz06764(names=["Alpha Corp"])
    # One host more than the domain pool holds, so the pool is guaranteed to
    # run out and recycle — the numbered fake this test is about. Sized off the
    # pool rather than a fixed list, which stopped exhausting it when the pool
    # grew from 4 domains to 32.
    hosts = [f"firm{i:02d}legal{i:02d}.com"
             for i in range(len(P._PN_EMAIL_DOMAINS) + 1)]
    minted = None
    for h in hosts:
        out = z.apply(f"x@{h}")
        host = out.split("@")[-1]
        if any(c.isdigit() for c in host):
            minted = host
            break
    assert minted, "no numbered fake domain was minted"
    kf = z.known_fake_words()
    findings = P._pn_review_findings(f"see info@{minted} and www.{minted}", kf)
    assert not [f for f in findings if f[0] == "url/domain"], \
        f"tool fake domain {minted} was flagged: {findings}"
    # a real source domain is still surfaced
    assert any(f[0] == "url/domain"
               for f in P._pn_review_findings("visit www.realclientco.com", kf))


# P2  "Law Office(s) of <person>" must not cluster as a person alias on "Law".
def test_firm_wrappers_do_not_cluster_as_alias():
    z = _pz06764(names=["Law Office of Neal S. Zaslavsky",
                        "Law Offices of Rob Hennig"])
    pair = {"Law Office of Neal S. Zaslavsky", "Law Offices of Rob Hennig"}
    assert not any(pair <= set(c) for c in z.alias_candidates())


# ──────────────────── PROVEN APPROACHES (design adoption) ───────────────────
# Techniques adopted from the mature de-identification stack (Presidio's
# two-tier scored detection, checksum validators, entity resolution,
# normalize-before-detect). No dependencies added — the principles are
# implemented natively. Calibrated to the tool's actual purpose: a PRECAUTION
# against casual recognition of public filings, not defense against a
# motivated adversary — so fidelity outranks exhaustive recall, and only a
# party-naming survivor blocks delivery (see the tiered leak-gate tests).

# #1  Two-tier detection: absence from the term list raises a flag, never
#     grants a pass. Role-anchored unknown names surface as review findings.
def test_unknown_name_scan_flags_unlisted_party():
    z = _pz06764(names=["Alejandro Orellana", "AHC Acquisition LLC"])
    out = z.apply("Plaintiff Alejandro Orellana sued. Defendant Travelers "
                  "joined. Attorneys for Sunrise Motors Group filed.")
    flagged = {s for _c, s in z.unknown_name_scan(out)}
    assert any("Travelers" in s for s in flagged), flagged
    assert any("Sunrise Motors Group" in s for s in flagged), flagged


def test_unknown_name_scan_quiet_on_fakes_headings_and_public_entities():
    z = _pz06764(names=["Alejandro Orellana", "AHC Acquisition LLC"])
    out = z.apply("Defendant AHC Acquisition LLC moved. Defendants Oppose The "
                  "Motion To Compel Arbitration. Defendant City of Los "
                  "Angeles appeared.")
    flagged = {s for _c, s in z.unknown_name_scan(out)}
    fakes = z.known_fake_words()
    for s in flagged:
        assert "Oppose" not in s, f"heading false positive: {s}"
        assert "Los Angeles" not in s and s != "City", f"public entity: {s}"
        assert not all(w.lower() in fakes for w in s.split()), f"pure fake: {s}"


def test_unknown_name_scan_silent_on_fully_handled_prose():
    # Reviewer-fatigue guard: on a text where every party is tracked, the
    # scanner must emit NOTHING — title-case headings ("Defendants Are
    # Entitled To An Order…"), role phrases, and our own fakes are all
    # neutral. Review noise is how a real survivor gets waved past.
    z = _pz06764(names=["Alejandro Orellana", "AHC Acquisition, LLC",
                        "Rafael Quintero"])
    out = z.apply(
        "Defendant AHC Acquisition, LLC submits this Reply. Plaintiff "
        "Alejandro Orellana signed. Defendants Are Entitled To An Order "
        "Compelling Arbitration. Plaintiff May Not Avoid The Delegation "
        "Clause. Defendant Reserves All Rights. Plaintiff Cites No "
        "Authority. Attorneys for Defendant filed the Declaration of "
        "Rafael Quintero.")
    assert z.unknown_name_scan(out) == []


# #3  Validators: a fake VIN carries a CORRECT ISO 3779 check digit, the way a
#     fake phone is valid NANP and a fake SSN is valid SSA.
def test_fake_vin_is_facially_valid():
    # canonical ISO example: check digit (position 9) of this VIN is 'X'
    assert P._pn_vin_check_digit("1M8GDM9AXKP042788") == "X"
    reg = P._PnFakeRegistry()
    fake = P._pn_fake_vin("1C4JJXP65PW699184", reg)
    assert re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", fake)
    assert fake[8] == P._pn_vin_check_digit(fake)
    assert fake != "1C4JJXP65PW699184"


# #3  Re-identification scan: a lookup-key identifier in the OUTPUT that is not
#     one of our fakes is reported; our own fakes are not.
def test_reid_scan_flags_survivor_not_own_fake():
    z = _pz06764(names=["Alejandro Orellana"])
    src = "HOLLOWAY VANCE, SBN 175977."
    z.register_identifiers(src)
    out = z.apply(src)
    assert z.reid_scan(out) == [], "flagged our own fake"
    found = z.reid_scan("stray block SBN 332686 and VIN 1C4JJXP65PW699184")
    cats = {c for c, _v in found}
    assert "REID bar number" in cats and "REID vin" in cats, found


def test_reid_scan_ignores_short_bare_numbers_but_keeps_bar_numbers():
    # "When in doubt": a retail DEAL#/Account# under 6 bare digits is not a
    # re-identification key, so it is not reported. A 5-digit BAR number is a
    # definite State Bar lookup and stays reported.
    z = _pz06764()
    found = z.reid_scan("DEAL# 23071 and Account No. 512 today.")
    assert found == [], found
    bar = z.reid_scan("counsel of record, SBN 12345, appeared.")
    assert ("REID bar number", "12345") in bar


# #4  Entity resolution: a trailing corporate suffix (with period) is a
#     decisive entity signal; the suffix is never faked as a person surname;
#     the person-path form of the same party reuses the entity identity.
def test_trailing_corp_suffix_is_decisive_entity_signal():
    assert P._pn_looks_like_entity("AhC Acquisition, LLC.")
    assert P._pn_looks_like_entity("Smith & Co.")
    assert not P._pn_looks_like_entity("Li Na")       # ambiguous bare suffix
    assert not P._pn_looks_like_entity("Roxane Estrada")


def test_corp_suffix_never_a_person_surname_single_identity():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["AhC Acquisition, LLC."], ["24STCV06764"], [],
                              registry=reg)
    assert any(t.category == "entity" for t in terms)
    assert not any(t.category == "person" for t in terms)
    fake_full, bare = P._pn_fake_person("AhC Acquisition, LLC.", reg)
    assert fake_full.rstrip().endswith("LLC."), fake_full
    assert not any(rt.lower().startswith("llc") for rt, _f, _s in bare)
    ent_fake = next(t.fake for t in terms if t.category == "entity")
    for w in fake_full.replace(",", " ").split():
        if w != "LLC.":
            assert w in ent_fake, f"{w!r} diverges from {ent_fake!r}"


# ─────────────────────── REVIEW (24STCV23198, Ramirez) ──────────────────────
# Fifth corpus (Ramirez v. Ford Motor Company). Leaks that survive OUTSIDE the
# caption: name-bearing Bates stamps, an undefined firm acronym, a suffix-less
# street address, and label/anchor gaps.

def _pz23198(names=("Ernest Ramirez", "Ford Motor Company",
                    "BP Ford of Long Beach")):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), ["24STCV23198"], [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


# L1  A Bates/production stamp embeds the party's name in cleartext; word
#     boundaries can't see it, so the token is faked whole, and a production
#     series keeps ONE consistent fake prefix.
def test_bates_stamps_scrubbed_with_consistent_prefix():
    z = _pz23198()
    src = ("See FMC_RAMIREZ_ERNEST_000007 and RAM000013-RAM000018; "
           "also RAM000013 alone. Compare CACI No. 2320.")
    z.register_identifiers(src)
    out = z.apply(src)
    low = out.lower()
    for real in ("ramirez", "ernest", "ram00", "000007"):
        assert real not in low, f"{real!r} survived: {out}"
    assert "CACI No. 2320" in out            # citation-style refs untouched
    stamps = re.findall(r"[A-Z]{2,}\d{4,}", out)
    assert len({re.match(r"[A-Z]+", s).group(0) for s in stamps}) == 1, \
        f"inconsistent series prefixes: {stamps}"
    assert z.reid_scan(out) == []            # certified clean
    # a stamp that never got registered IS flagged on the way out
    assert any(c == "REID production number"
               for c, _v in z.reid_scan("produced as XYZ_CORP_004412"))


# L2  A firm acronym used with no defining parenthetical is derived from the
#     tracked entity's initials, occurrence-gated, and mapped to the initials
#     of the entity's fake. ISO/RJN and lowercase words are untouched.
def test_undefined_firm_acronym_scrubbed():
    z = _pz23198(names=["Ernest Ramirez", "Strategic Legal Practices, APC",
                        "Mortenson Taggart Adams LLP"])
    src = ("SLP's remote desktop failed. SLP does not employ in-house IT. "
           "Plaintiff filed an ISO and RJN. The slp lowercase word stays.")
    z.register_entity_acronyms(src)
    out = z.apply(src)
    assert not re.search(r"(?<!\w)SLP(?!\w)", out), out
    assert "ISO" in out and "RJN" in out
    assert "slp lowercase" in out
    # occurrence gate: MTA never used in the document, so never registered
    assert ("short-name", "mta") not in z.records
    # the acronym fake is the initials of the entity's fake
    ent = z.records[("entity", "strategic legal practices, apc")]["fake"]
    initials = "".join(
        w[0].upper() for w in P._pn_acronym_trim(ent.replace(",", " ").split())
        if P._pn_word_base(w) not in P._PN_ACRONYM_CONNECTIVES
        and len(P._pn_word_base(w)) >= 2)
    assert z.records[("short-name", "slp")]["fake"] == initials


# L3  A street with no street-type suffix ("Century Park East") is detected
#     when a floor/suite or City, ST ZIP tail follows; prose can't qualify.
def test_suffixless_street_detected_with_tail():
    z = _pz23198()
    out = z.apply("offices at 1888 Century Park East, 19th Floor, "
                  "Los Angeles, CA 90067.")
    assert "Century Park" not in out, out
    assert "1888" in out, f"house number should be kept: {out!r}"
    assert "19th Floor" in out and "Los Angeles, CA 90067" in out
    for prose in ("24 Hour Fitness Center opened a new gym",
                  "100 Years War East narrative"):
        assert P._PN_ADDR_RE.search(prose) is None, prose


# L4  A non-party dealer near a purchase cue is REPORTED, never replaced.
#     (A dealer sharing a tracked token — "Worthington FORD" with Ford a
#     party — is already partially handled by the token pass; the review net
#     is for the dealer no term touches at all.)
def test_dealer_purchase_cue_is_review_only():
    z = _pz23198(names=["Ernest Ramirez"])
    out = z.apply("Plaintiff purchased the vehicle from Worthington Motors "
                  "in Long Beach.")
    assert "Worthington Motors" in out       # no term -> not auto-replaced
    findings = P._pn_review_findings(out, z.known_fake_words())
    assert ("possible business (purchase)", "Worthington Motors") in findings


# L5  "RES. NO." is recognized directly, independent of any "Reservation ID".
def test_res_no_label_recognized():
    vals = dict(P._pn_identifier_values("RES. NO.: 134890490939"))
    assert vals.get("reservation id") == "134890490939"


# Item 7  The "Ford" short form must be replaced in prose while every
#     "... v. Ford Motor Co." authority survives byte-identical.
def test_ford_short_form_vs_ford_citations():
    z = _pz23198()
    z.register_short_names('Ford Motor Company ("Ford") makes vehicles.')
    cites = [
        "Donlen v. Ford Motor Co. (2013) 217 Cal.App.4th 138",
        "Valdez v. Ford Motor Co. (2022) 134 Cal.App.5th 1",
    ]
    body = ("Ford admitted the defect. " + cites[0] + ". Ford knew. "
            + cites[1] + ".")
    out = z.apply(body)
    for c in cites:
        assert c in out, f"authority corrupted: {out}"
    assert not re.search(r"(?<!\w)Ford(?!\w) (?:admitted|knew)", out), \
        f"prose short form survived: {out}"


# ────────────────── REVIEW (24STCV23198, delivered run) ─────────────────────
# Failures observed in the ACTUAL delivered run (quarantined *.LEAK exports +
# key): generic words registered as terms ("Name" x140, "warranty"->"langley",
# "Legal Standard"->"Granite Standard"); citation-harvested entities counted
# as leaks; federal dockets faked as VINs; authorities rewritten where the
# parser missed a shape; appendix queries faking real authorities.

def _pzR(names=("Ernest N Ramirez", "Ford Motor Company",
                "BP Ford of Long Beach")):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), ["24STCV23198"], [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


def test_generic_words_never_become_terms():
    z = _pzR(names=["Ernest N Ramirez", "Ford Motor Company",
                    "Strategic Legal Practices, APC", "BP Ford of Long Beach",
                    "Name"])
    for bad in [("person", "name"), ("entity", "name"),
                ("person-token", "legal"), ("person-token", "beach"),
                ("person-token", "long")]:
        assert bad not in z.records, f"{bad} was registered"
    out = z.apply("A. Legal Standard. The written warranty covers it. "
                  "Case Name: x. He lives in Long Beach. Their relationship "
                  "ended. Strategic Legal Practices appeared.")
    for kept in ("Legal Standard", "warranty", "Case Name", "Long Beach",
                 "relationship"):
        assert kept in out, f"{kept!r} corrupted: {out}"
    assert "Strategic Legal Practices" not in out   # the firm itself scrubs


def test_generic_label_harvest_rejected():
    z = _pzR(names=["Ernest N Ramirez"])
    z.register_label_names("Contractor: Customer Relations (323) 555-1212\n"
                           "Warranty Complaints (562) 555-3434")
    assert ("person", "customer relations") not in z.records
    assert ("person", "warranty complaints") not in z.records


def test_citation_only_harvest_is_pruned_and_stays_pruned():
    z = _pzR()
    corpus = ("McGee v. Mercedes-Benz USA, LLC, No. 19CV513-MMA (WVG), 2020 "
              "WL 1530921, at *5 (S.D. Cal. Mar. 30, 2020).")
    new = []
    P._pn_append_name_terms(new, "Mercedes-Benz USA, LLC", "document",
                            z.registry)
    z._add_terms(new)
    pruned = z.prune_citation_only_terms(corpus)
    assert any("Mercedes-Benz" in p for p in pruned)
    # a per-file re-harvest cannot resurrect a pruned value
    new2 = []
    P._pn_append_name_terms(new2, "Mercedes-Benz USA, LLC", "document",
                            z.registry)
    z._add_terms(new2)
    assert ("entity", "mercedes-benz usa, llc") not in z.records
    assert "Mercedes-Benz USA, LLC" in z.apply(corpus)


def test_leak_scan_ignores_protected_citations():
    # "Silvio v. Ford Motor Co." preserved in a citation is the protection
    # WORKING — counting it as a leak quarantined ten clean exports.
    z = _pzR()
    body = ("Prose about Zephyr here. Silvio v. Ford Motor Co., 109 "
            "Cal.App.4th 1205, 1207 (2003).")
    assert not any("ford" in s.lower() for s in z.surviving_reals(body))
    assert not any("ford" in s.lower()
                   for s in z.surviving_reals_reduced(body))
    # a genuine PROSE survivor still counts
    sv = z.surviving_reals("Defendant Ford Motor Company moved in prose.")
    assert "Ford Motor Company" in sv


def test_inre_nakedwl_and_yearless_wl_protected():
    z = _pzR()
    for cite in [
        'x."); In Re Ford Motor Co. DPS6 Powershift Transmission Products '
        "Liability Litigation, 689 F.Supp.3d 760, 776 (C.D. Cal. 2023)",
        "(x); Ford Motor Co., No. CV222111DSFAGRX, 2023 WL 3035369, at *3 "
        "(C.D. Cal. Feb. 21, 2023)",
        "see Noori v. Jaguar Land Rover N. Am., LLC, No. 220CV0813SVWJEMX, "
        "2021 WL 1232450, at *9 (C.D. Cal. Apr.);",
    ]:
        assert z.apply(cite) == cite, f"authority rewritten: {z.apply(cite)}"


def test_docket_numbers_are_not_vins():
    z = _pzR(names=["Ernest N Ramirez"])
    z.register_identifiers("No. 221CV01063RGKAFMX, 2022 WL 1591701 and "
                           "VIN 1C4JJXP65PW699184")
    assert ("vin", "221cv01063rgkafmx") not in z.records
    assert ("vin", "1c4jjxp65pw699184") in z.records   # real VIN still caught
    assert not any(c == "REID vin"
                   for c, _v in z.reid_scan("stray 20EDCV221896MWFJC here"))


def test_attorney_before_bar_label_registered():
    z = _pzR(names=["Ernest N Ramirez"])
    z.register_label_names("Michael D. Mortenson, State Bar No. 230831\n"
                           "Craig A. Taggart, State Bar No. 239168")
    out = z.apply("Michael D. Mortenson and Craig A. Taggart appeared.")
    assert "Mortenson" not in out and "Taggart" not in out, out


def test_appendix_query_keeps_authority_scrubs_own_caption():
    z = _pzR()
    authority = {"kind": "case", "plaintiff": "McGee",
                 "defendant": "Mercedes-Benz USA, LLC"}
    s = "McGee v. Mercedes-Benz USA, LLC 2020 WL 1530921"
    assert z.apply_to_citation(authority, s) == s
    own = {"kind": "case", "plaintiff": "Ernest N Ramirez",
           "defendant": "Ford Motor Company"}
    cap = "Ernest N Ramirez v. Ford Motor Company Case No. 24STCV23198"
    scrubbed = z.apply_to_citation(own, cap)
    assert "Ramirez" not in scrubbed and "Ford" not in scrubbed


def test_adjacency_warning_only_for_real_conflicts():
    recs = [
        {"category": "address", "real": "300 Spectrum Center Dr., Suite 1200",
         "fake": "7956 Foxglove Drive, Suite 1200"},
        {"category": "address",
         "real": "300 Spectrum Center Drive, Suite 1200, Irvine, CA 92618",
         "fake": "7956 Foxglove Drive, Suite 1200, Irvine, CA 92618"},
        {"category": "address", "real": "414 S. Maple Ave.",
         "fake": "1533 Cypress Ave."},
        {"category": "address", "real": "416 S. Maple Ave.",
         "fake": "2044 Birch Ave."},
    ]
    w = P._pn_address_adjacency(recs)
    assert len(w) == 1 and "Maple" in w[0], w


# ─────────────────────── CAPTION RECONSTRUCTION ─────────────────────────────
# The caption party block — where every hard extraction failure has lived —
# is cut out and re-rendered from the tracked parties' fakes, but ONLY when
# every word of it is attributable to a known party, role label, or caption
# boilerplate. Anything unattributed and the page falls back untouched.

_CAPTION_ROWS = [
    (1, [(110.0, "ALEJANDRO ORELLANA, an individual,"),
         (380.0, "Case No. 24STCV06764")]),
    (2, [(110.0, "Plaintiff,"), (380.0, "NOTICE OF MOTION")]),
    (3, [(110.0, "vs.")]),
    (4, [(110.0, "AHC ACQUISITION, LLC; and DOES 1-10,")]),
    (5, [(110.0, "inclusive,")]),
    (6, [(110.0, "Defendants.")]),
]


def _caption_pz():
    return _pz06764(names=["Alejandro Orellana", "AHC Acquisition, LLC"])


def test_caption_block_is_rebuilt_from_fakes():
    z = _caption_pz()
    out = P._pn_apply_page_rows(z, _CAPTION_ROWS)
    joined = " ".join(out).lower()
    for real in ("alejandro", "orellana", "ahc", "acquisition"):
        assert real not in joined, f"{real!r} survived: {out}"
    assert "does 1-10" in joined                     # boilerplate kept
    assert "Plaintiff," in out[1] and "vs." in out[2] and "Defendants." in out[5]
    assert "NOTICE OF MOTION" in out[1]              # right column untouched
    assert len(out) == len(_CAPTION_ROWS)            # row count preserved
    assert out[0].split(",")[0].isupper()            # caption casing followed
    assert z.records[("entity", "ahc acquisition, llc")]["count"] > 0
    # Line NUMBERING is untouched: reconstruction rewrites segment text in
    # place and never adds/removes rows, so every gutter number (including
    # None continuation rows) survives in order.
    rebuilt = P._pn_reconstruct_caption(_caption_pz(), _CAPTION_ROWS)
    assert [n for n, _ in rebuilt] == [n for n, _ in _CAPTION_ROWS]


def test_caption_failsafe_unknown_party_blocks_reconstruction():
    # An unattributed name in the block must abort reconstruction — silently
    # dropping a party the key doesn't know would misrepresent the case.
    z = _caption_pz()
    rows = list(_CAPTION_ROWS)
    rows[3] = (4, [(110.0, "AHC ACQUISITION, LLC; ZENITH PARTNERS; and "
                           "DOES 1-10,")])
    assert P._pn_reconstruct_caption(z, rows) is None


def test_caption_wrapped_name_is_attributed_and_rebuilt():
    # The DREAM TEAM failure shape: a defendant wrapped mid-name across rows
    # matches no whole-string pattern, but word-set attribution still binds it.
    z = _caption_pz()
    rows = [
        (1, [(110.0, "ALEJANDRO ORELLANA,"), (380.0, "Case No. 24STCV06764")]),
        (2, [(110.0, "Plaintiff,"), (380.0, "COMPLAINT")]),
        (3, [(110.0, "vs.")]),
        (4, [(110.0, "AHC")]),
        (5, [(110.0, "ACQUISITION, LLC,")]),
        (6, [(110.0, "Defendant.")]),
    ]
    out = P._pn_apply_page_rows(z, rows)
    joined = " ".join(out).lower()
    assert "ahc" not in joined and "acquisition" not in joined, out


def test_caption_tolerates_contact_furniture_and_descriptors():
    # A realistic first page: attorney block (SBN, address, phone), court
    # heading, and a multi-line class descriptor. None of it is a party, so
    # none of it may block reconstruction — only a distinctive capitalized
    # word unattributable to a tracked party does.
    z = _caption_pz()
    rows = [
        (1, [(110.0, "RAFAEL QUINTERO, ESQ., SBN 175977")]),
        (2, [(110.0, "817 N. La Brea Avenue, Suite 200")]),
        (3, [(110.0, "Telephone: (626) 292-0899")]),
        (4, [(110.0, "Attorneys for Plaintiff")]),
        (5, [(110.0, "SUPERIOR COURT OF THE STATE OF CALIFORNIA")]),
        (6, [(110.0, "COUNTY OF LOS ANGELES")]),
        (7, [(110.0, "ALEJANDRO ORELLANA, an individual, on behalf"),
             (380.0, "Case No. 24STCV06764")]),
        (8, [(110.0, "of himself and all others similarly situated,"),
             (380.0, "NOTICE OF MOTION")]),
        (9, [(110.0, "Plaintiff,")]),
        (10, [(110.0, "vs.")]),
        (11, [(110.0, "AHC ACQUISITION, LLC; and DOES 1-10, inclusive,")]),
        (12, [(110.0, "Defendants.")]),
    ]
    out = P._pn_apply_page_rows(z, rows)
    joined = " ".join(out).lower()
    for real in ("alejandro", "orellana", "ahc", "acquisition"):
        assert real not in joined, f"{real!r} survived: {out}"
    assert "COUNTY OF LOS ANGELES" in out[5]     # court heading untouched
    assert "does 1-10" in joined


def test_caption_not_reconstructed_without_structure():
    z = _caption_pz()
    # ordinary prose: no role/vs rows
    assert P._pn_reconstruct_caption(
        z, [(1, [(110.0, "The parties met and conferred.")])]) is None
    # single-column page: a caption block needs a second column beside it
    rows = [(n, [s for s in segs if s[0] < 300]) for n, segs in _CAPTION_ROWS]
    assert P._pn_reconstruct_caption(z, rows) is None


# Tiered leak gate: the pseudonymization is a precaution against casual
# recognition of a PUBLIC filing, so only a survivor that names a party
# outright justifies quarantining the exports. A bare token or identifier
# surviving is a warn-and-deliver, not a blocked run.
def test_leak_gate_tiers_primary_vs_lesser():
    z = _pz06764(names=["Alejandro Orellana", "AHC Acquisition, LLC"])
    # a bare surname token leaking is NOT primary
    z.note_leaks({"orellana"})
    assert z.has_leaks() and z.primary_leaks() == set()
    # a full entity name leaking IS primary
    z.note_leaks({"ahc acquisition, llc"})
    assert any("AHC Acquisition" in v for v in z.primary_leaks())


def test_leak_gate_short_name_is_primary():
    # A document-defined short form ("ToFF") names the party outright — its
    # survival makes the case recognizable, so it gates like a full name.
    z = _pz06764(names=["Tom of Finland Foundation, Inc."])
    z.register_short_names('Tom of Finland Foundation, Inc. ("ToFF") appeared.')
    z.note_leaks({"toff"})
    assert "ToFF" in z.primary_leaks()


def test_leak_gate_identifier_is_not_primary():
    # A surviving identifier is a review item, not a delivery blocker.
    z = _pz06764(names=["Alejandro Orellana"])
    z.register_identifiers("SBN 175977")
    z.note_leaks({"175977"})
    assert z.has_leaks() and z.primary_leaks() == set()


# ─────────────── REVIEW (25STCV24253, Emmett v. Ford, delivered run) ─────────
# Quarantined *.LEAK exports showed the defendant's name ("Ford Motor Company")
# rewritten INSIDE published authorities the strict parser couldn't read, and a
# Bates stamp whose interior apostrophe hid a party prefix. Renaming a cited
# decision is the method's cardinal failure, so each shape below must survive
# apply() byte-for-byte; the stamp must be faked WHOLE so no party token leaks.

def test_defendant_side_citations_survive_unreadable_shapes():
    z = _pzR()  # Ford Motor Company registered as a party
    cites = [
        # OCR split the docket number across spaces -> WL run stopped early
        "Hastings v. Ford Motor Co., No. 19-CV- 02217-BAS- MDD 2022 WL 848330",
        # two pin pages before "(year)" -> single-pin tail missed the close
        "Bowser v. Ford Motor Co., 78 Cal. App. 5th 587, 599, 617 (2022)",
        # short form cited by defendant name only, no "X v."
        "Ford Motor Co., No. CV222111DSFAGRX, 2023 WL 3035369, at *3 "
        "(C.D. Cal. Feb. 21, 2023)",
        # consolidated-proceeding title with no "v.", plus its short forms
        "Ford Motor Warranty Cases (2025) 17 Cal.5th 1122",
        "See Ford Motor Warranty, 333 17 Cal.5th at 1133-1134",
        "the Ford Motor Warranty Court highlighted",
        # In re short form, no reporter, "Litig." anchor (and welded "SeeIn re")
        "SeeIn re Ford Motor Co. DPS6 Powershift Transmission Prod. "
        "Liability Litig.",
        # In re with a naked WL number and no court paren
        "In re Ford Motor Co. DPS6 2019 WL 7185524",
    ]
    for c in cites:
        assert z.apply(c) == c, f"authority rewritten: {z.apply(c)!r}"


def test_warranty_prose_is_not_over_protected():
    # Lowercase "warranty" fills these very briefs; only the capital-C title and
    # the "Warranty <cite|Court>" short name are authorities. The party's own
    # name near ordinary "warranty" prose must still scrub.
    z = _pzR()
    out = z.apply("During the warranty coverage period, Ford Motor Company "
                  "warrants the vehicle under the New Vehicle Limited Warranty.")
    assert "Ford Motor Company" not in out, out
    assert "warranty coverage period" in out and "Limited Warranty" in out


def test_bates_stamp_with_apostrophe_scrubs_party_prefix():
    z = _pzR()  # Ford is a party base
    for stamp in ("FORD_O’WBYSOZ_91084", "FORD_O'WBYSOZ_44370"):
        z.register_identifiers(f"(See {stamp} for the calibration note.)")
        out = z.apply(f"(See {stamp}.)")
        assert stamp not in out, f"whole stamp survived: {out}"
        assert "FORD" not in out.upper(), f"party prefix leaked: {out}"


# ─── Accented declarant/party names (Spanish surnames, LA County corpus) ──────
# "Declaration of Teresa C. Alarcón" was captured as "Teresa C. Alarc" — the
# ASCII-only name class stopped at the "ó". The truncated stem went into the
# key (so the real name never round-tripped) and left the accented tail welded
# onto the fake in the body ("Isleyón Decl."). Names must be captured, keyed,
# and replaced WHOLE, accents included.

def test_accented_declarant_name_captured_and_replaced_whole():
    z = _pzR(names=["Ernest N Ramirez"])
    z.register_declarant_names("Declaration of Teresa C. Alarcón in support "
                               "of the motion.")
    # the WHOLE name is keyed, not a stem truncated at the accent
    assert ("person", "teresa c. alarcón") in z.records
    assert ("person", "teresa c. alarc") not in z.records
    out = z.apply("(Alarcón Decl., ¶ 4, Ex. 3 at 73:4-13.) Alarcón testified.")
    assert "Alarcón" not in out, f"real accented name survived: {out}"
    assert "Isleyón" not in out and "ón" not in out, f"welded fragment: {out}"
    # nothing real survives in the finished text
    assert z.surviving_reals(out) == [] and z.surviving_reals_reduced(out) == []


def test_word_tokenizer_keeps_accented_letters():
    assert P._PN_WORD_RE.findall("Alarcón Muñoz Hernández Peña") == \
        ["Alarcón", "Muñoz", "Hernández", "Peña"]


def test_reduced_scan_folds_accents_so_ocr_variants_match():
    # A spliced page often drops the accent ("ALARCON") or welds the name to a
    # neighbour; the accent-folded reduction still ties it to the real value.
    z = _pzR(names=["Ernest N Ramirez"])
    z.register_declarant_names("Declaration of Teresa C. Alarcón in support.")
    spliced = "TERESACALARCONDECLARATION"          # accent dropped + welded
    assert any("Alarcón" in s for s in z.surviving_reals_reduced(spliced))
    scrubbed = z.scrub_welded(spliced)
    assert "ALARCON" not in scrubbed.upper(), scrubbed


# ─── Short surnames + first/last token rows in the key (reversal-macro) ───────
# "Gregory Yu" -> "Ivers Kingsley": the surname "Yu" is two letters, so the old
# >=3 floor never registered it. A bare "Yu" leaked in the text, and the key had
# no standalone "Yu"->"Kingsley" row, so the token-level reversal macro could
# not undo a bare "Kingsley".

def test_two_letter_surname_is_scrubbed_and_keyed():
    z = _pzR(names=["Gregory Yu"])
    out = z.apply("(Yu Decl., ¶ 3.) Yu testified. Gregory Yu signed it.")
    assert "Yu" not in out, f"two-letter surname leaked: {out}"
    assert ("person-token", "yu") in z.records


def test_two_letter_english_word_surname_not_registered():
    # A surname spelled like a common word ("As", "Of") must NOT become a term,
    # or it would rewrite that word throughout the brief.
    z = _pzR(names=["Robert As", "Mary Of"])
    assert ("person-token", "as") not in z.records
    assert ("person-token", "of") not in z.records
    out = z.apply("This applies as of the date, as noted.")
    assert out == "This applies as of the date, as noted.", out


def _write_and_read_key(z, tmp_path):
    import openpyxl
    p = tmp_path / "pseudonym_key.xlsx"
    z.write_key(p, log)
    ws = openpyxl.load_workbook(p)["Pseudonym Key"]
    return list(ws.iter_rows(min_row=2, values_only=True))


def test_key_carries_first_and_last_name_rows(tmp_path):
    # Even when only the FULL name appears, the key gets a row per first/last
    # token so a token-level reversal macro can undo either alone. The exact
    # fake is pool-dependent (and so must not be hard-coded); the invariant is
    # that the full-name fake is exactly its two token fakes joined.
    z = _pzR(names=["Gregory Yu"])
    z.apply("Gregory Yu signed the declaration.")   # full name only
    rows = _write_and_read_key(z, tmp_path)
    full = {r[1]: r[2] for r in rows if r[0] == "person"}
    tok = {r[1]: r[2] for r in rows if r[0] == "person-token"}
    assert "Gregory Yu" in full
    assert "Gregory" in tok and "Yu" in tok
    assert full["Gregory Yu"] == f"{tok['Gregory']} {tok['Yu']}"


def test_key_orders_people_before_other_categories(tmp_path):
    z = _pzR(names=["Gregory Yu"])
    z.register_identifiers("VIN 1FA6P8TH5J5100001")
    z.apply("Gregory Yu drove VIN 1FA6P8TH5J5100001.")
    rows = _write_and_read_key(z, tmp_path)
    cats = [r[0] for r in rows]
    # every person/token row precedes the first non-people row
    last_person = max(i for i, c in enumerate(cats)
                      if c in ("person", "person-token"))
    first_other = min(i for i, c in enumerate(cats)
                      if c not in ("person", "person-token", "entity",
                                   "entity-token", "short-name", "display-name"))
    assert last_person < first_other, cats


# ─── Procedural phrases never flagged; declarant refs always flagged ──────────
# Document-type words ("Opposition", "Demurrer", "Reply") and any combination
# of them ("Plaintiff's Opposition to Mot.") name a pleading, not a person, and
# must never appear as an unscrubbed-name finding. Conversely, the name before a
# "Declaration"/"Decl."/"Dec." reference is the top place a witness name leaks
# and is always surfaced.

@pytest.mark.parametrize("phrase", [
    "Opposition", "Demurrer", "Reply", "Motion", "Complaint",
    "Plaintiff’s Opposition to Mot.", "Reply Declaration", "Notice of Motion",
    "Opposition to Motion to Compel", "Separate Statement",
    "Motion for Summary Judgment", "Defendant’s Reply Brief",
])
def test_procedural_phrases_are_never_names(phrase):
    assert P._pn_is_procedural_phrase(phrase) is True


@pytest.mark.parametrize("name", [
    "Smith", "Alarcón", "Gregory Yu", "Smith Opposition", "Worthington Ford",
])
def test_real_names_are_not_procedural(name):
    assert P._pn_is_procedural_phrase(name) is False


def test_procedural_phrase_suppressed_in_person_review():
    # Even reached through the "possible person name" scan, a doc phrase is
    # dropped. (Anchored via a "signed by:" cue so the person regex fires.)
    hits = P._pn_person_review_findings("Signed by: Reply Brief")
    assert hits == []
    hits = P._pn_person_review_findings("Signed by: Gregory Yu")
    assert any("Gregory Yu" in s for _c, s in hits)


@pytest.mark.parametrize("text,want", [
    ("(Smith Decl., ¶ 3.)", "Smith"),
    ("(Alarcón Declaration, ¶ 4.)", "Alarcón"),
    ("See Gregory Yu Decl.", "Gregory Yu"),
    ("cf. Nguyen Decl.", "Nguyen"),
    ("(Yu Dec.)", "Yu"),
])
def test_declarant_reference_is_always_flagged(text, want):
    hits = P._pn_declarant_ref_findings(text)
    assert hits and hits[0][1] == want, hits


@pytest.mark.parametrize("text", [
    "the Reply Declaration argues",       # procedural, not a person
    "In Dec. 2024 he signed",             # a date, not a declaration
    "Meeting on Dec. 5, 2024",            # a date
])
def test_declarant_scan_skips_nonnames(text):
    assert P._pn_declarant_ref_findings(text) == []


def test_declarant_scan_skips_our_own_fake():
    assert P._pn_declarant_ref_findings("(Kingsley Decl.)",
                                        known_fakes={"kingsley"}) == []


# ─── Declarant references are auto-PSEUDONYMIZED, not just flagged ────────────
# "Smith Decl.", "Alarcón Declaration", "Yu Dec." — the name is scrubbed even
# when that declaration isn't in the batch. The word "Declaration" and its
# variants ("Decl.", "Dec.") are NEVER faked.

def _pz_blank():
    return P.Pseudonymizer([], {}, registry=P._PnFakeRegistry())


def test_declarant_ref_name_is_registered_and_scrubbed():
    z = _pz_blank()
    doc = "(Smith Decl., ¶ 3.) Smith testified. See Alarcón Declaration, ¶ 4."
    z.register_declarant_refs(doc)
    out = z.apply(doc)
    assert "Smith" not in out and "Alarcón" not in out, out
    # the reference word itself is intact in both spellings
    assert "Decl." in out and "Declaration" in out


def test_two_letter_declarant_surname_is_scrubbed():
    z = _pz_blank()
    doc = "(Yu Dec.) Later Yu signed the declaration."
    z.register_declarant_refs(doc)
    out = z.apply(doc)
    assert "Yu" not in out, out
    assert "Dec." in out


def test_declaration_word_and_variants_are_never_faked():
    z = _pz_blank()
    doc = "Smith Decl.; Doe Declaration; Roe Dec.; the Declaration of Coe."
    z.register_declarant_refs(doc)
    # none of the reference words are registered as terms
    for w in ("declaration", "decl", "dec"):
        assert not any(P._pn_word_base(t) == w
                       for (_c, _rl), rec in z.records.items()
                       for t in str(rec["real"]).split())
    out = z.apply(doc)
    assert out.count("Decl") >= 2 and "Declaration" in out


def test_descriptor_and_date_declarations_are_left_alone():
    z = _pz_blank()
    doc = ("The Supporting Declaration and Expert Witness Declaration were "
           "filed. In Dec. 2024 they met.")
    z.register_declarant_refs(doc)
    out = z.apply(doc)
    assert out == doc, f"descriptor/date wrongly scrubbed: {out}"


def test_declarant_name_only_in_filename_is_scrubbed():
    # A declarant named only in the FILENAME ("Alarcón Declaration ISO Reply")
    # — never in the body — is registered from the stem so the output filename
    # scrubs it. (The prescan feeds the separator-normalized stem in.)
    import re
    z = _pz_blank()
    stem = re.sub(r"[_\-]+", " ", "Alarcón_Declaration_ISO_Reply")
    z.register_declarant_refs(stem)
    faked = z.apply(stem)
    assert "Alarcón" not in faked and "Declaration" in faked, faked


def test_procedural_filenames_are_not_renamed():
    import re
    for stem in ["Reply_Brief", "Opposition_to_Motion", "Expert_Declaration",
                 "Motion_for_Summary_Judgment"]:
        z = _pz_blank()
        sp = re.sub(r"[_\-]+", " ", stem)
        z.register_declarant_refs(sp)
        assert z.apply(sp) == sp, stem


# ─── Document terms survive a dropped space (OCR welds) ───────────────────────
# If OCR eats the space, a pleading phrase becomes one token
# ("OppositiontoMotion"); it must still read as procedural, not a name.

@pytest.mark.parametrize("welded", [
    "OppositiontoMotion", "ReplyBrief", "NoticeofMotion", "PointsandAuthorities",
    "MotionforSummaryJudgment", "ExParteApplication", "OppositiontoMotiontoCompel",
    "ReplyDeclaration", "Plaintiff’sOppositiontoMot.", "Defendant’sReplyBrief",
])
def test_welded_document_phrase_still_procedural(welded):
    assert P._pn_is_procedural_phrase(welded) is True


@pytest.mark.parametrize("name", [
    "Sotomayor", "Bennett", "Delacroix", "Motioned", "Reyes", "Ordonez",
    "Compton", "Ashford", "O’Brien", "OConnor", "Alarcón",
])
def test_welded_segmentation_does_not_swallow_real_names(name):
    assert P._pn_is_procedural_phrase(name) is False


# ─── Declaration references survive a dropped space too ───────────────────────
# OCR welding the space in "Smith Decl." -> "SmithDecl." must still scrub the
# name (the reference word "Decl."/"Declaration"/"Dec." is preserved).

@pytest.mark.parametrize("text,want", [
    ("(SmithDecl., ¶ 3.)", "Smith"),
    ("AlarcónDeclaration, ¶ 4", "Alarcón"),
    ("(YuDec.)", "Yu"),
    ("DeckerDecl.", "Decker"),
])
def test_welded_declaration_reference_name_is_found(text, want):
    names = P._pn_declarant_ref_names(text)
    assert names and names[0] == want, names


@pytest.mark.parametrize("text", [
    "SupportingDeclaration", "ExpertWitnessDeclaration",
    "CustodianofRecordsDeclaration", "ReplyDeclaration",
    "Redeclaration of rights", "Declan filed it", "In Dec. 2024 he signed",
])
def test_welded_declaration_nonnames_are_dropped(text):
    assert P._pn_declarant_ref_names(text) == []


def test_welded_declaration_is_auto_scrubbed_keeping_the_word():
    z = _pz_blank()
    doc = "(SmithDecl.) SmithDecl. again. YuDec. signed."
    z.register_declarant_refs(doc)
    out = z.apply(doc)
    assert "Smith" not in out and "Yu" not in out, out
    assert "Decl." in out and "Dec." in out          # reference word preserved


# ─── Quarantine only the files that carry a leak, not the whole batch ─────────

def test_only_leaking_files_are_quarantined():
    from pathlib import Path
    z = _pz_blank()
    a, b, c = Path("Motion.txt"), Path("Opposition.txt"), Path("Reply.txt")
    z.written = [a, b, c]
    # b carries a party-name leak; a carries a lesser token; c is clean
    z.leaked_by_file = {a: {"sometoken"}, b: {"ford motor company"}}
    # primary gate blocks only the party name -> only b is held
    assert z.files_to_quarantine({"Ford Motor Company"}) == [b]
    # strict gate (all leaks) -> a and b, never the clean c
    assert set(z.files_to_quarantine({"Ford Motor Company", "sometoken"})) == {a, b}
    assert c not in z.files_to_quarantine({"Ford Motor Company", "sometoken"})
    # nothing blocking -> nothing quarantined
    assert z.files_to_quarantine(set()) == []


# ─── Stray interior characters are ignored (OCR dropped a mark into a word) ────
# "Def.endant" reads as "Defendant"; comparison is on the letters only.

@pytest.mark.parametrize("mangled", [
    "Def.endant", "Opp.osition", "Mot.ion", "Decl.aration", "Plaint-iff",
    "Sep.arate State.ment", "Def.endant’s Opp.osition to Mot.ion",
])
def test_stray_char_document_words_still_procedural(mangled):
    assert P._pn_is_procedural_phrase(mangled) is True


@pytest.mark.parametrize("name", [
    "Sotomayor", "Bennett", "O’Brien", "D.E. Shaw", "Alarcón", "Ford",
    "Smith Opp.osition",   # a real surname next to a mangled doc word is kept
])
def test_stray_char_normalization_keeps_real_names(name):
    assert P._pn_is_procedural_phrase(name) is False


# ─── A stray char inside the DECLARATION reference word is tolerated ──────────
# "Smith Dec.laration" reads as a declaration cite; the name scrubs and the
# (mangled) reference word is preserved. The word is validated letters-only.

@pytest.mark.parametrize("text,want", [
    ("Alarcón Dec.laration", "Alarcón"),
    ("Smith Dec.laration, ¶ 3", "Smith"),
    ("Smith Decl.aration", "Smith"),
    ("Smith Declara.tion", "Smith"),
    ("(SmithDecl.)", "Smith"),
])
def test_mangled_reference_word_is_tolerated(text, want):
    names = P._pn_declarant_ref_names(text)
    assert names and names[0] == want, names


@pytest.mark.parametrize("text", [
    "In Dec. 2024 he signed", "Meeting on Dec. 5, 2024", "the Reply Declaration",
    "Declan filed it", "David Decker met", "See December filings", "Donald Trump",
])
def test_mangled_reference_word_guards(text):
    assert P._pn_declarant_ref_names(text) == []


def test_unknown_declarant_with_mangled_ref_is_autoscrubbed():
    z = _pz_blank()
    doc = "Smith Dec.laration, ¶ 3. Later Smith testified."
    z.register_declarant_refs(doc)
    out = z.apply(doc)
    assert "Smith" not in out, out
    assert "Dec.laration" in out          # the mangled reference word is kept


# ── 2026-07-21 review: false-positive classes from a delivered leak sheet ────
# Nine of fourteen rows were "ca.gov" LEAKs (a government host minted as a
# domain record, then "surviving" inside every kept citation link); the rest
# were a defined term sharing a tracked party's two initials, a court form
# stamp read as a person, and unregistered clerk signatures (see
# test_court_names for those).

def test_government_domain_is_never_minted_or_leak_flagged():
    z = _pz()
    out = z.apply("File at appellate.courts.ca.gov; forms wrapped to\n"
                  "ca.gov today.")
    assert "ca.gov" in out                       # kept verbatim
    assert not any(r["category"] == "url" for r in z.records.values())
    assert not z.surviving_reals(out)


def test_private_domain_is_still_faked():
    z = _pz()
    out = z.apply("Our site is estradalegal.com for intake.")
    assert "estradalegal.com" not in out


def test_gov_url_not_reported_for_review():
    findings = P._pn_review_findings("visit selfhelp.courts.ca.gov and ca.gov")
    assert not [f for f in findings if f[0] == "url/domain"]


def test_form_stamp_words_are_not_a_person():
    # "NEW QUALIFIER" is an e-filing form stamp: name-shaped to the scanner
    # (leading "New" + one distinctive word) but a form label, not a person.
    assert not P._pn_person_review_findings("By: NEW QUALIFIER")


# ── 2026-07-22 review: pool exhaustion, typo folding, bare-number noise ──────
# From a 6-file Palladino set whose key showed numbered-mint fakes
# ("Corwin Vance3", "HENDRY2 CORPORATIOLORNE10"), OCR name variants each given
# an unrelated fake ("Joe Palladina"->"Corwin Vance3" vs "Joe Palladino"->
# "Corwin Keswick"), and 300+ leak rows dominated by bare 3-digit account ids
# (101/110/120/130/214/342) recurring across every file.

def test_name_pools_are_large_and_valid():
    # The pool must stay well ahead of a real case's distinct name tokens
    # (parties + counsel + staff + declarants + e-mail display names).
    assert len(P._PN_NAME_WORDS) >= 150
    assert len(P._PN_ENTITY_WORDS) >= 90
    # No duplicates, and every surname is a valid bare name token.
    assert len({w.lower() for w in P._PN_NAME_WORDS}) == len(P._PN_NAME_WORDS)
    assert len({w.lower() for w in P._PN_ENTITY_WORDS}) == len(P._PN_ENTITY_WORDS)
    for w in P._PN_NAME_WORDS:
        assert P._pn_is_name_token(w), w


def test_ocr_name_variant_folds_onto_a_typo_of_the_base_fake():
    # "Palladino" and its OCR variants must read as typos of ONE fake, each
    # keeping its own distinct (reversible) stand-in.
    reg = P._PnFakeRegistry()
    base = reg.token("palladino", P._PN_NAME_WORDS, "nametok")
    v1 = reg.token("palladina", P._PN_NAME_WORDS, "nametok")   # substitution
    v2 = reg.token("pallading", P._PN_NAME_WORDS, "nametok")   # substitution
    v3 = reg.token("palladinos", P._PN_NAME_WORDS, "nametok")  # insertion
    fam = [base, v1, v2, v3]
    assert len(set(f.lower() for f in fam)) == 4               # all distinct
    for v in (v1, v2, v3):
        assert P._pn_edit_distance_le1(base.lower(), v.lower(), min_len=3), \
            f"{v} is not a typo-neighbour of {base}"


def test_length_tracks_the_real_deviation():
    reg = P._PnFakeRegistry()
    base = reg.token("manav", P._PN_NAME_WORDS, "nametok")
    longer = reg.token("manavi", P._PN_NAME_WORDS, "nametok")  # real gained a char
    assert len(longer) == len(base) + 1                        # so did the fake


def test_distinct_short_names_do_not_fold():
    # Below the fold floor, two genuinely different short surnames keep their
    # own independent fakes (no accidental typo-linking of "Vang"/"Vong").
    reg = P._PnFakeRegistry()
    a = reg.token("vang", P._PN_NAME_WORDS, "nametok")
    b = reg.token("vong", P._PN_NAME_WORDS, "nametok")
    assert not P._pn_edit_distance_le1(a.lower(), b.lower(), min_len=3)


# ── 2026-07-24 review: transposed and welded name variants ───────────────────
# From an "ADLER" set where the swapped "Alder" and the column-welded
# "ADLERMICHAEL" each minted an unrelated fake instead of tracking "ADLER".

def test_adjacent_transposition_is_one_edit():
    # A swapped pair reads as one OCR/keying slip (Damerau), not two subs.
    assert P._pn_edit_distance_le1("adler", "alder", min_len=5)
    assert P._pn_edit_distance_le1("gunderson", "gudnerson", min_len=5)
    # Two non-adjacent or non-swapped changes are still more than one edit.
    assert not P._pn_edit_distance_le1("adler", "obler", min_len=5)


def test_transposed_name_folds_onto_a_transposed_fake():
    # "Alder" (a transposition of "Adler") must read as a typo of the SAME fake
    # "Adler" got — a transposed stand-in — not an unrelated pool word.
    reg = P._PnFakeRegistry()
    base = reg.token("adler", P._PN_NAME_WORDS, "nametok")
    swap = reg.token("alder", P._PN_NAME_WORDS, "nametok")
    assert swap.lower() != base.lower()                        # its own stand-in
    assert P._pn_edit_distance_le1(base.lower(), swap.lower(), min_len=5), \
        f"{swap} is not a typo-neighbour of {base}"
    # It is specifically a transposition of the base fake, not a substitution.
    assert sorted(swap.lower()) == sorted(base.lower())


def test_welded_name_folds_onto_concatenated_fakes():
    # A column-splice glues two names ("ADLERMICHAEL"); the fake must show the
    # bound party ("Adler"->fake) with the other name's fake appended, and reuse
    # each part's own stand-in when it later appears alone.
    reg = P._PnFakeRegistry()
    adler = reg.token("adler", P._PN_NAME_WORDS, "nametok")
    weld = reg.token("adlermichael", P._PN_NAME_WORDS, "nametok")
    michael = reg.token("michael", P._PN_NAME_WORDS, "nametok")
    assert weld.lower().startswith(adler.lower())              # bound party first
    assert weld.lower() == (adler + michael).lower()           # + its neighbour
    assert len({adler.lower(), weld.lower(), michael.lower()}) == 3   # injective


def test_fold_distance_scales_with_length():
    # Short names stay strict (one slip); longer tokens, which plausibly carry
    # more independent typos, tolerate two or three.
    assert P._pn_name_fold_dist("adler", "alder") == 1
    assert P._pn_name_fold_dist("gunderson", "gundersen") == 1     # 9 chars
    assert P._pn_name_fold_dist("cunningham", "cunninghom") == 2   # 10 chars
    assert P._pn_name_fold_dist("schwarzeneger", "schwartzeneger") == 2  # 13-14
    assert P._pn_name_fold_dist("a" * 16, "b" * 16) == 3


def test_long_double_typo_folds_but_short_double_typo_does_not():
    # A long token two edits from a bound one still reads as the same name;
    # a short token two edits away is a genuinely different surname.
    reg = P._PnFakeRegistry()
    base = reg.token("cunningham", P._PN_NAME_WORDS, "nametok")
    dbl = reg.token("cunninghomm", P._PN_NAME_WORDS, "nametok")   # dist 2, 10+ chars
    assert P._pn_edit_distance_within(base.lower(), dbl.lower(), 2, min_len=5)

    reg2 = P._PnFakeRegistry()
    short = reg2.token("adler", P._PN_NAME_WORDS, "nametok")
    other = reg2.token("obler", P._PN_NAME_WORDS, "nametok")      # dist 2, only 5 chars
    assert not P._pn_edit_distance_within(short.lower(), other.lower(), 1, min_len=5)
    # It got an independent stand-in, not a corruption of the bound fake.
    assert sorted(other.lower()) != sorted(short.lower())


def test_multi_char_insertion_tracks_fake_length():
    # A real that gained two characters folds onto a fake that also grew by two,
    # so the stand-in's shape still mirrors the real's deviation.
    reg = P._PnFakeRegistry()
    base = reg.token("washington", P._PN_NAME_WORDS, "nametok")
    grown = reg.token("washingtonxy", P._PN_NAME_WORDS, "nametok")   # +2 chars
    assert len(grown) == len(base) + 2
    assert P._pn_edit_distance_within(base.lower(), grown.lower(), 2, min_len=5)


def test_osa_counts_a_transposition_as_one():
    assert P._pn_osa_distance("adler", "alder") == 1
    assert P._pn_osa_distance("abc", "abc") == 0
    assert P._pn_osa_distance("abc", "abcd") == 1
    assert P._pn_osa_distance("kitten", "sitting") == 3


def test_weld_folding_is_order_independent():
    # A weld is strictly longer than its base, so it can only fold once the base
    # is bound. The build pre-binds tokens shortest-first, so "ADLERMICHAEL"
    # folds onto "ADLER" no matter which order the document presents them in —
    # and the assignment is identical across orderings.
    def fakes_for(order):
        reg = P._PnFakeRegistry()
        terms = P._pn_build_terms(list(order), [], [], registry=reg)
        return {t.real.lower(): t.fake.lower()
                for t in terms if t.category in ("person", "entity")}

    orderings = [
        ["adlermichael", "adler", "michael"],   # weld first (the old failure)
        ["adler", "adlermichael", "michael"],
        ["michael", "adlermichael", "adler"],
    ]
    results = [fakes_for(o) for o in orderings]
    for f in results:
        assert f["adlermichael"].startswith(f["adler"])   # weld shows the base
        assert f["adlermichael"] == f["adler"] + f["michael"]
    # Deterministic: every ordering yields the very same three fakes.
    assert results[0] == results[1] == results[2]


def test_ordinary_long_surname_is_not_split_as_a_weld():
    # A single surname that merely starts with a bound short name must NOT be
    # torn into two fakes: too short to hold two full name tokens (< 2x the
    # fold floor), so it draws one ordinary pool word.
    reg = P._PnFakeRegistry()
    marsh = reg.token("marsh", P._PN_NAME_WORDS, "nametok")
    marshall = reg.token("marshall", P._PN_NAME_WORDS, "nametok")
    assert not marshall.lower().startswith(marsh.lower())      # not a weld
    pool = {w.lower() for w in P._PN_NAME_WORDS}
    assert marshall.lower().rstrip("0123456789") in pool       # one pool word


def test_bare_short_account_number_is_not_tracked():
    # A 1-3 digit "account number" is a document number, not a key — never
    # registered, so it never leaks as review noise.
    for label in ("Response No. 101", "Account No. 110", "DEAL# 214"):
        vals = {v for _c, v in P._pn_identifier_values(label)}
        assert "101" not in vals and "110" not in vals and "214" not in vals


def test_distinctive_account_number_is_still_tracked():
    got = {v for _c, v in P._pn_identifier_values(
        "DEAL# 23071 and Cust# 4433221 and Acct #: A203")}
    assert "23071" in got and "4433221" in got and "A203" in got


# ── 2026-07-22 review: REVIEW scans flagged already-faked content ────────────
# The delivered leak sheet flagged the run's OWN fakes as unscrubbed names
# whenever a fake sat next to a non-fake word: "ASIC Pruett Keswick",
# "Nolan Relations", "Operations Calder", the welded "HENDRY2 CORPORATIOLORNE10"
# and "POSTBOX4.ORGPANY". Known = the fake tokens those phrases are built from.
_KNOWN = {"keswick", "radley", "nolan", "pruett", "calder", "corwin",
          "hendry2", "lorne10", "postbox4", "colfax2"}


def test_welded_fake_is_recognised_as_own_fake():
    assert P._pn_word_is_own_fake("CORPORATIOLORNE10", _KNOWN)   # Lorne10 welded
    assert P._pn_word_is_own_fake("POSTBOX4.ORGPANY", _KNOWN)    # Postbox4 welded
    assert P._pn_word_is_own_fake("Keswick", _KNOWN)             # bare fake
    assert not P._pn_word_is_own_fake("Travelers", _KNOWN)       # a real survivor


def test_declarant_scan_ignores_a_fake_dragged_by_a_prefix():
    for ctx in ["ASIC Pruett Keswick Decl.", "CBA. Radley Decl.",
                "PDT Keswick Decl.", "Nolan Relations. Radley Decl.",
                "Keswick's January Radley Decl.", "See Radley Decl."]:
        assert P._pn_declarant_ref_findings(ctx, _KNOWN) == [], ctx


def test_declarant_scan_still_flags_a_real_declarant_leak():
    got = [s for _c, s in P._pn_declarant_ref_findings(
        "See Marisol Vandergriff Decl.", _KNOWN)]
    assert got and "Vandergriff" in got[0]


def test_person_scan_ignores_a_fake_beside_a_nonname_word():
    for ctx in ["By: Nolan Relations", "By: Operations Calder",
                "cc: Corwin Palladine"]:
        assert P._pn_person_review_findings(ctx, _KNOWN) == [], ctx


def test_person_scan_still_flags_a_real_person_leak():
    got = [s for _c, s in P._pn_person_review_findings("By: Aileen Neitzert", _KNOWN)]
    assert got == ["Aileen Neitzert"]


def test_unknown_name_scan_ignores_welded_fakes():
    for ctx in ["Against Defendant HENDRY2 Corporation",
                "Defendant HENDRY2 CORPORATIOLORNE10 moved",
                "Defendant POSTBOX4.ORGPANY responded"]:
        assert P._pn_unknown_name_findings(ctx, _KNOWN) == [], ctx


def test_unknown_name_scan_still_flags_a_real_entity():
    got = [s for _c, s in P._pn_unknown_name_findings(
        "Defendant Travelers Insurance moved", _KNOWN)]
    assert got == ["Travelers Insurance"]


# ── 2026-07-22 review: argument-heading phrases flagged as names ─────────────
# A separate statement / brief sets its argument headings in title case
# ("Defendant Cannot Establish...", "For Age Discrimination", "Voluntarily
# Resigned"), which the role-anchored unknown-name scan read as party names.
# The gazetteer (its stated design: high-frequency English + motion-practice
# vocabulary) now covers them. A real name in the same slot still flags.

_ARG_SALAD = [
    "Cannot Establish the Elements Of A", "Cannot Show He Experienced Intolerable",
    "Has No Direct Evidence of Discrimination", "Has Not Established Pretext",
    "For Age Discrimination", "For Failure To Prevent", "For Retaliation",
    "For Wrongful Termination In Violation", "Voluntarily Resigned",
    "Was Financially Motivated", "Requested and Was Approved", "Made A Rational",
    "Employees Grow Irate", "Board of Directors", "Employee Relations",
    "Disputed", "Undisputed", "Evidence", "Days", "Complains Again In January",
    "Admits He Was Accommodated", "For FEHA Retaliation",
]


def test_argument_headings_are_not_flagged_as_names():
    for phrase in _ARG_SALAD:
        got = P._pn_unknown_name_findings(
            f"Defendant {phrase} moved for judgment.", set())
        assert got == [], f"{phrase!r} flagged: {got}"


def test_real_names_after_a_role_anchor_still_flag():
    for name in ("Travelers Insurance", "Sunrise Motors Group", "Aileen Neitzert"):
        got = [s for _c, s in P._pn_unknown_name_findings(
            f"Defendant {name} moved for judgment.", set())]
        assert any(name.split()[0] in s for s in got), f"{name!r} not flagged"


def test_gazetteer_did_not_swallow_a_leaked_surname():
    # "Green", "Smart", "Raven", "Moore" leaked as real names in the case that
    # prompted this fix — none may become a neutral common word.
    for surname in ("green", "smart", "raven", "moore"):
        assert surname not in P._PN_COMMON_WORDS


def test_no_fake_pool_word_is_a_gazetteer_word():
    pool = {w.lower() for w in P._PN_NAME_WORDS} | {w.lower() for w in P._PN_ENTITY_WORDS}
    assert not (pool & P._PN_COMMON_WORDS)


def test_a_heading_built_on_our_own_fake_is_not_a_declarant_finding():
    """"F. Langley Submitted No Declaration and Relies on Inadmissible Hearsay"
    is an argument heading. "Langley" is this run's own stand-in; strip it and
    "Submitted No" is left, which cleared the old two-words test and put the
    tool's own fake in front of the operator as an unscrubbed declarant."""
    known = {"langley", "holloway"}
    assert P._pn_declarant_ref_findings(
        "F. Langley Submitted No Declaration and Relies on Hearsay", known) == []


def test_a_real_declarant_beside_our_fake_is_still_reported():
    known = {"langley", "holloway"}
    got = P._pn_declarant_ref_findings("Langley and Alarcon Decl. at 3", known)
    assert got and got[0][1] == "Alarcon", got
