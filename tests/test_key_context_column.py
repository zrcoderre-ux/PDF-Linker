"""The pseudonym key quotes the sentence each value stood in.

A key row says `Rasho -> Strangeways`, and whether that binding is right depends
on how the document used the word — the same question the LEAKS worksheet's
Context column exists to answer, asked of a decision already made.

It is read from the UNSCRUBBED body, because by the time the export exists the
real value has been replaced and the scrubbed copy cannot quote it. That is a
deliberate trade: the key now carries sentences of the real document, not merely
its real values. It was never a shareable file — it is the reversal map — so
this changes how revealing it is, not which file is safe to send.

The column sits beside the Replacement it explains -- column D, with the
Replacement at C -- because that is where it is read: the row states the
binding, and the sentence that follows is the evidence for it. Safe to insert
(and to reorder) rather than append, which is worth pinning because
the reverse is the obvious fear: `DeAnonymize.bas` scans the header row for
"real value" / "replacement" / "status" and uses whatever columns they land in,
and `_pn_load_key` resolves by header name too. The only thing a moved column
can break is a POSITIONAL fingerprint, so `_pn_key_looks_like_ours` is cut to
the three headers both layouts share and a six-column key from an older version
still reads as ours.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_context_column.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

SOURCE = "\n".join([
    "====== Page 1 ======",
    " 1  SUPERIOR COURT OF THE STATE OF CALIFORNIA",
    " 2  HELEN RASHO, an individual,",
    " 3  Plaintiff Helen Rasho alleges that on March 3, 2023, the",
    " 4  defendant refused to produce the billing records.",
    " 5  Counsel served the summons on Marcus Delacroix at his",
    " 6  residence in Pasadena on the following day.",
    "",
])


def _pz():
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["Helen Rasho", "Marcus Delacroix"], [], [],
                              registry=reg)
    return P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)


def _rows(path):
    wb = openpyxl.load_workbook(path)
    out = []
    for ws in wb.worksheets:
        hdr = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is not None:
                out.append((hdr, r))
    return out


# ── the column ──────────────────────────────────────────────────────────────

def test_the_key_carries_a_context_column_beside_the_replacement(tmp_path):
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)

    hdr, _r = _rows(key)[0]
    assert hdr[:2] == list(P._PN_KEY_FINGERPRINT), hdr
    # The binding leads — Real Value then Replacement, adjacent at B and C —
    # and the ONE Context column follows at D, holding both sentences stacked
    # (original on top) as the evidence for the mapping just read.
    assert hdr[2] == "Replacement", hdr
    assert hdr[3] == "Context", hdr
    # …and the File and Where the quote was read from follow it at E and F.
    assert hdr == ["Category", "Real Value", "Replacement",
                   "Context", "File", "Where (page:line)",
                   "Status", "Source", "Occurrences"], hdr


def test_every_quote_contains_its_own_real_value(tmp_path):
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)

    quoted = 0
    for hdr, r in _rows(key):
        real, ctx = str(r[1]), str(r[hdr.index("Context")] or "")
        if ctx:
            quoted += 1
            assert real.lower() in ctx.lower(), (real, ctx)
    assert quoted, "no row was quoted at all"


def test_the_quote_is_the_sentence_not_the_line(tmp_path):
    # A sentence on pleading paper is spread over several numbered lines, and
    # one of them alone says almost nothing.
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    ctx = z._key_context["marcus delacroix"]
    assert "served the summons on Marcus Delacroix" in ctx, ctx
    assert "residence in Pasadena" in ctx, ctx      # the wrapped line, joined
    assert "\n" not in ctx and "  " not in ctx      # rebuilt as prose
    assert not ctx.lstrip().startswith("5")         # the gutter number is gone


def test_a_key_written_before_the_column_existed_still_loads(tmp_path):
    # The fingerprint is the two headers every layout has led with plus a
    # by-name check for Replacement, so a key written before the Context
    # columns existed is still recognised as ours and still round-trips —
    # which is what stops a moved column stranding every key in circulation.
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])                       # the pre-Context layout
    ws.append(["person", "Helen Rasho", "Kelsall Strangeways", "replaced",
               "spreadsheet", 3])
    wb.save(key)

    assert P._pn_key_looks_like_ours(key)
    assert P._pn_key_context_on_disk(key) == ({}, {}, {})  # nothing to carry, no crash
    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    assert any(str(t.real) == "Helen Rasho" for t in terms)


# ── the quote survives a folder whose documents have moved on ───────────────

def test_a_quote_is_carried_forward_when_it_cannot_be_re_derived(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    z.write_key(key, log)
    first = {str(r[1]): str(r[h.index("Context")] or "")
             for h, r in _rows(key)}
    assert first["Marcus Delacroix"]

    # A later run that never sees that document: same bindings, no source to
    # quote from. The key must not lose what the last run learned.
    z2 = _pz()
    z2.apply("Nothing relevant here.")
    z2.note_key_context("Nothing relevant here.")
    assert not z2._key_context
    z2.write_key(key, log)
    second = {str(r[1]): str(r[h.index("Context")] or "")
              for h, r in _rows(key)}
    assert second["Marcus Delacroix"] == first["Marcus Delacroix"]


def test_this_runs_quote_beats_the_one_on_disk(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    z.write_key(key, log)

    z2 = _pz()
    other = ("====== Page 1 ======\n"
             " 1  The witness Marcus Delacroix testified for two hours about\n"
             " 2  the disputed invoices at the deposition.\n")
    z2.apply(other)
    z2.note_key_context(other)
    z2.write_key(key, log)
    got = {str(r[1]): str(r[h.index("Context")] or "")
           for h, r in _rows(key)}
    assert "testified for two hours" in got["Marcus Delacroix"], got


def test_the_first_document_to_use_a_value_owns_its_quote(tmp_path):
    # Stable across a re-run of the same folder: a value named in several
    # documents is quoted from the first that used it.
    z = _pz()
    z.note_key_context(SOURCE)
    firstq = z._key_context["marcus delacroix"]
    z.note_key_context("====== Page 1 ======\n 1  Marcus Delacroix again.\n")
    assert z._key_context["marcus delacroix"] == firstq


# ── and the rewrite that made it affordable ─────────────────────────────────

def test_context_prep_is_shared_across_values():
    parsed = P._pn_body_lines(SOURCE)
    a = P._pn_context_prep(parsed)
    b = P._pn_context_prep(parsed)
    assert a is b, "the per-body derivation must not be redone per value"
    # ...and a different body evicts it rather than answering from the old one.
    other = P._pn_body_lines("====== Page 1 ======\n 1  Something else.\n")
    assert P._pn_context_prep(other) is not a
    assert P._pn_context_prep(parsed)[2] != P._pn_context_prep(other)[2]


def test_line_offsets_match_the_joined_text():
    # The running offset replaced `len(" ".join(joined))` recomputed per line.
    # If the two ever disagree, every quote is cut in the wrong place.
    parsed = P._pn_body_lines(SOURCE)
    lines, lowers, text, _ends, _locs = P._pn_context_prep(parsed)
    for (start, end, _prose), low in zip(lines, lowers):
        assert text[start:end].lower() == low, (start, end, text[start:end], low)


@pytest.mark.parametrize("needle", ["Marcus Delacroix", "Rasho", "SUPERIOR",
                                    "billing", "Nonesuch", ""])
def test_context_still_answers_for_awkward_values(needle):
    parsed = P._pn_body_lines(SOURCE)
    got = P._pn_context(parsed, needle)
    assert isinstance(got, str)
    if needle and needle.lower() in SOURCE.lower():
        assert needle.lower() in got.lower(), (needle, got)
    if needle in ("Nonesuch", ""):
        assert got == ""
