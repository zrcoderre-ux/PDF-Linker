"""
D7 / D8 — what the delivered key and the OCR-mangled detectors got wrong.

D8. The delivered key had 335 rows and 133 of them were droppable or harmful.
`write_key` deliberately keeps a row the party template named even when the
batch never mentioned the party — the fake is already minted and the row is the
only durable record of a binding a later filing will need. Right forward, and a
hazard in reverse: `ReAnonymizeTentative` runs the map backwards and would
replace a Real Value that was never in the document. Status "no match"
documented the hazard without preventing it.

D7. Two OCR-mangled values the detectors did not recognise: a phone number whose
area-code brackets the scanner mangled, and an address whose "@" it read as
"(a)" — which left the local part (the attorney's own name) standing while the
url rule faked the host.

Run:  cd PDF-Linker && python3 -m pytest tests/test_key_hygiene.py -v
"""

import logging
import pathlib
import tempfile

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")


def _pz(names=(), casenos=("25STCV14710",)):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), list(casenos), [], registry=reg)
    det = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
    return P.Pseudonymizer(terms, det, registry=reg)


def _write(z, tmp_path):
    p = tmp_path / "pseudonym_key.xlsx"
    z.write_key(p, log)
    wb = openpyxl.load_workbook(p)
    macro = [r for r in wb.active.iter_rows(min_row=2, values_only=True) if r[1]]
    pinned = ([r for r in wb[P._PN_KEY_PINNED_SHEET]
               .iter_rows(min_row=2, values_only=True) if r[1]]
              if P._PN_KEY_PINNED_SHEET in wb.sheetnames else [])
    return p, macro, pinned


# ─────────────────── D8: the reverse pass sees only what shipped ────────────

def test_a_binding_no_export_carries_is_off_the_macros_sheet(tmp_path):
    z = _pz(names=["Helen Rasho", "Someone Neverpresent"])
    z.apply("Plaintiff Helen Rasho filed this action.")
    _p, macro, pinned = _write(z, tmp_path)
    assert "Someone Neverpresent" not in {r[1] for r in macro}, (
        "the reverse pass would rewrite a value that was never in the document")
    assert "Someone Neverpresent" in {r[1] for r in pinned}, (
        "the binding must still be pinned for the run that meets the party")


def test_a_pinned_binding_is_still_read_back(tmp_path):
    # The pin is the whole reason the row is written: a re-run must reuse the
    # exact stand-in the first run reserved.
    z = _pz(names=["Helen Rasho", "Someone Neverpresent"])
    z.apply("Plaintiff Helen Rasho filed this action.")
    reserved = z.records[("person", "someone neverpresent")]["fake"]
    p, _macro, _pinned = _write(z, tmp_path)

    reg = P._PnFakeRegistry()
    loaded, _dec = P._pn_load_key(p, reg, log)
    got = {t.real: t.fake for t in loaded}
    assert got.get("Someone Neverpresent") == reserved, got


def test_the_token_rows_of_a_shipped_name_stay_on_the_macros_sheet(tmp_path):
    # The macro reverses a composed fake WORD BY WORD, so the token rows of a
    # party whose full name is the only form the export used are load-bearing
    # even though they matched nothing themselves.
    z = _pz(names=["Gregory Yu"])
    z.apply("Gregory Yu signed the declaration.")
    _p, macro, _pinned = _write(z, tmp_path)
    reals = {r[1] for r in macro}
    assert {"Gregory Yu", "Gregory", "Yu"} <= reals, reals


def test_a_case_variant_is_one_row_not_two(tmp_path):
    # "Barry"/"BARRY" -> the same fake is ONE case-insensitive search; a second
    # row is dead weight.
    z = _pz(names=["Barry Abernathy"])
    z.apply("BARRY signed. Barry also signed.")
    _p, macro, _pinned = _write(z, tmp_path)
    barry = [r for r in macro if str(r[1]).lower() == "barry"]
    assert len(barry) == 1, barry
    assert barry[0][5] == 2, "both spellings must count toward the one row"


def test_two_spellings_of_one_address_each_reverse(tmp_path):
    """Two spellings of one parcel, and BOTH are fully reversible.

    They used to collide onto one fake — the whole "<number> <name> <suffix>"
    was memoized on the street identity — so one row had to be marked
    `alt spelling` and only the other could reverse. Now the house number and
    the suffix are kept verbatim and only the NAME is faked, so each spelling
    yields its own distinct fake, the key is bijective here, and nothing has to
    be retired. They still read as one street, which is the point."""
    z = _pz()
    z.apply("11845 W. Olympic Blvd., Suite 1270 and "
            "11845 W. Olympic Boulevard, Suite 1270.")
    _p, macro, _pinned = _write(z, tmp_path)
    addr = [r for r in macro if r[0] == "address"]
    assert len(addr) == 2, addr
    assert all(r[3] != P._PN_KEY_ALT_STATUS for r in addr), (
        f"a spelling was retired from the reverse pass: {addr}")
    assert len({str(r[2]) for r in addr}) == 2, f"two reals, one fake: {addr}"
    assert len({P._pn_addr_name_of(P._pn_addr_parts(str(r[2]))[0])
                for r in addr}) == 1, f"one parcel, two street names: {addr}"


# ──────────────────── D7: OCR-mangled contact details ───────────────────────

@pytest.mark.parametrize("text", [
    "Call (818) 953-0150 for details.",
    "Call 4(818) 953-0150 for details.",
    "Call 1{818) 953-0150 for details.",
    "Call [818] 953-0150 for details.",
])
def test_a_mangled_area_code_is_still_a_phone_number(text):
    out = _pz().apply(text)
    assert "953-0150" not in out and "818" not in out, out


@pytest.mark.parametrize("text", [
    "Bates RAM0001234567890 produced.",
    "Reservation 8113657265341 confirmed.",
])
def test_a_long_digit_run_is_not_a_phone_number(text):
    assert _pz().apply(text) == text


@pytest.mark.parametrize("addr", [
    "shabib@erskinelaw.com",
    "shabib(a)erskinelaw.com",
    "shabib[a]erskinelaw.com",
])
def test_an_ocr_at_sign_still_reads_as_an_address(addr):
    out = _pz().apply(f"Contact {addr} for service.")
    assert "shabib" not in out, f"the local part rode out in the clear: {out}"
    assert "erskinelaw" not in out, out


def test_the_local_part_is_rewritten_piece_by_piece():
    z = _pz(names=["Zachary Coderre"])
    out = z.apply("Contact zcoderre.enterprise1(a)erskinelaw.com today.")
    assert "coderre" not in out.lower(), out
    assert "enterprise1" not in out.lower(), (
        "a handle with no known name in it is a real, searchable handle")
