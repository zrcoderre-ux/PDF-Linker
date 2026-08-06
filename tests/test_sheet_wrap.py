"""Every spreadsheet this tool writes wraps its text.

All four of them now carry a cell that is a SENTENCE — the LEAKS Context
column, the key's, the Notes and the accumulating Cases list on the master KEEP
sheet. Excel's default spills a long cell across its empty neighbours and clips
it the moment the neighbour is occupied, so the column the operator is meant to
READ was the one they could not, without widening it by hand every time.

One helper (`_pn_wrap_sheet`) serves all of them, for the reason the two
launcher builders are shared: they should not be able to drift apart. This
covers each writer separately anyway, because "shared helper" is only true until
someone adds a fifth sheet.

Run:  cd PDF-Linker && python3 -m pytest tests/test_sheet_wrap.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")

SOURCE = ("====== Page 1 ======\n"
          " 1  Plaintiff Helen Rasho served Marcus Delacroix at his\n"
          " 2  residence in Pasadena on the following day.\n")


def _cells(path):
    """[(workbook, sheet, cell)] for every populated cell of every sheet."""
    wb = openpyxl.load_workbook(path)
    return [(path.name, ws.title, c)
            for ws in wb.worksheets
            for row in ws.iter_rows() for c in row if c.value is not None]


def _assert_wrapped(path):
    cells = _cells(path)
    assert cells, f"{path.name} has no populated cells to check"
    bad = [(s, c.coordinate) for _n, s, c in cells if not c.alignment.wrap_text]
    assert not bad, f"{path.name}: not wrapped at {bad[:6]}"
    # Vertical TOP as well: a four-line quote beside a one-line real value
    # centred against it makes the eye hunt for which row a value belongs to.
    bad = [(s, c.coordinate) for _n, s, c in cells
           if c.alignment.vertical != "top"]
    assert not bad, f"{path.name}: not top-aligned at {bad[:6]}"


def _pz(names=("Helen Rasho", "Marcus Delacroix")):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    return P.Pseudonymizer(terms, list(P._PN_DEFAULT_DETECTORS), registry=reg)


def test_the_pseudonym_key_wraps(tmp_path):
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    _assert_wrapped(key)


def test_the_keys_pinned_sheet_wraps_too(tmp_path):
    # A binding no export carried lives on its own sheet, and it is written by a
    # separate `append` loop — so it is a separate chance to forget.
    z = _pz(names=["Helen Rasho", "Someone Neverpresent"])
    z.apply(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    names = openpyxl.load_workbook(key).sheetnames
    assert P._PN_KEY_PINNED_SHEET in names, names
    _assert_wrapped(key)


def test_the_leaks_worksheet_wraps(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "Motion.pdf", "type": "name?", "value": "Ashely",
         "where": "p.3:7", "notes": "",
         "context": "The complaint spells the defendant's given name Ashely, "
                    "which is how it survived the scrub."}], log)
    _assert_wrapped(tmp_path / "LEAKS.xlsx")


def test_both_master_sheets_wrap(tmp_path, monkeypatch):
    master = tmp_path / "master.xlsx"
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    P._pn_update_master_keep(
        {}, {"cal": {"value": "Cal", "type": "KEEP", "fix": "no",
                     "notes": "kept in this folder"}},
        "23STCV14415 Test", "2026-08-06", log)
    P._pn_update_master_leaks(master, [("Ashely", "name?")],
                              "23STCV14415 Test", "2026-08-06", log)
    sheets = openpyxl.load_workbook(master).sheetnames
    assert P._PN_MASTER_KEEP_SHEET in sheets and P._PN_MASTER_LEAK_SHEET in sheets
    _assert_wrapped(master)


def test_the_header_row_wraps_with_the_data(tmp_path):
    # Applied to the header too, so a two-word header wraps instead of demanding
    # a column wider than its data.
    z = _pz()
    z.apply(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    ws = openpyxl.load_workbook(key).active
    assert all(c.alignment.wrap_text for c in ws[1] if c.value is not None)


def test_row_height_is_left_for_excel_to_fit(tmp_path):
    """No explicit height is stored, so Excel auto-fits the wrapped rows.

    A height written here would freeze the layout at whatever font metrics this
    machine happened to have, which is exactly the thing wrapping is supposed to
    stop mattering.
    """
    z = _pz()
    z.apply(SOURCE)
    z.note_key_context(SOURCE)
    key = tmp_path / "pseudonym_key.xlsx"
    z.write_key(key, log)
    ws = openpyxl.load_workbook(key).active
    heights = [d.height for d in ws.row_dimensions.values()]
    assert not any(h for h in heights), heights


@pytest.mark.parametrize("sheet_writer", ["key", "leaks"])
def test_wrapping_does_not_disturb_the_values(tmp_path, sheet_writer):
    # Formatting only: the cells must still say exactly what they said.
    if sheet_writer == "key":
        z = _pz()
        z.apply(SOURCE)
        z.note_key_context(SOURCE)
        path = tmp_path / "pseudonym_key.xlsx"
        z.write_key(path, log)
        ws = openpyxl.load_workbook(path).active
        reals = {r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[1]}
        assert {"Helen Rasho", "Marcus Delacroix"} <= reals, reals
    else:
        P._pn_write_leak_report(tmp_path, [
            {"file": "Motion.pdf", "type": "name?", "value": "Ashely",
             "where": "p.3:7", "context": "a sentence", "notes": ""}], log)
        ws = openpyxl.load_workbook(tmp_path / "LEAKS.xlsx").active
        vals = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
        assert "Ashely" in vals, vals
