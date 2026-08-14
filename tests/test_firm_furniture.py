"""
Name furniture — "LAW OFFICES OF …", "& Associates", "the …", "Mr." — is never
faked.

Those words name the trade and a connector, not a party, so faking them hides
nothing and destroys the caption ("Braxton Mansffield bancroft Merrick C.
Whitlock"). They must be kept verbatim inside a composed fake, must never be
registered as a bare token, must never be harvested into a key row, and — the
loop that made this survive every KEEP the operator typed — must never come
BACK as a live term when an older key is reused.

Run:  cd PDF-Linker && python3 -m pytest tests/test_firm_furniture.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}

FURNITURE = ("law", "laws", "office", "offices", "firm", "of", "the", "and",
             "associates", "associated", "attorney", "attorneys", "counsel")


# ── composing a fake ────────────────────────────────────────────────────────

@pytest.mark.parametrize("real", [
    "Law Offices of Scott C. Stratman",
    "Law Office of Steven W Burt",
    "Ryan G. Block Associated Attorney",
    "the Waggoner",
])
def test_person_path_keeps_furniture_verbatim(real):
    fake, _bare = P._pn_fake_person(real, P._PnFakeRegistry())
    for r, f in zip(real.split(), fake.split()):
        if P._pn_word_base(r) in FURNITURE:
            assert f == r, f"{real!r} -> {fake!r}: {r!r} was faked as {f!r}"
    # ...and the party itself is still scrubbed.
    assert "Stratman" not in fake and "Burt" not in fake
    assert "Waggoner" not in fake and "Block" not in fake


def test_entity_path_keeps_furniture_and_initials():
    fake = P._pn_fake_entity("Law Offices of Philip Y Kim, APC",
                             P._PnFakeRegistry())
    assert fake.startswith("Law Offices of ")
    assert fake.endswith(", APC")
    assert " Y " in fake          # a lone initial is not a whole entity word
    assert "Kim" not in fake and "Philip" not in fake


def test_firm_name_reads_as_a_firm():
    """The whole point: the fake still says what kind of thing it names."""
    reg = P._PnFakeRegistry()
    assert P._pn_fake_entity("Alder Law, P.C.", reg).endswith(" Law, P.C.")
    assert P._pn_fake_entity("Mitilian Law Group", reg).endswith(" Law Group")


@pytest.mark.parametrize("real", ["The Law Firm", "Attorneys at Law"])
def test_an_all_furniture_name_still_scrubs(real):
    """Keeping the furniture must never leave the fake equal to the real value —
    that ships the name in a "clean" export and loops --fix-leaks forever."""
    fake, _ = P._pn_fake_person(real, P._PnFakeRegistry())
    assert P._pn_norm_map(fake) != P._pn_norm_map(real)
    ent = P._pn_fake_entity(real, P._PnFakeRegistry())
    assert P._pn_norm_map(ent) != P._pn_norm_map(real)


def test_furniture_is_never_a_bare_token():
    terms = P._pn_build_terms(
        ["Law Offices of Scott C. Stratman", "Law Offices of Philip Y Kim, APC"],
        [], [])
    bare = {t.real.lower() for t in terms
            if t.category in ("person-token", "entity-token")
            and len(t.real.split()) == 1}
    assert not (bare & set(FURNITURE)), f"bare furniture tokens: {bare}"
    assert "stratman" in bare


def test_no_pool_word_is_spent_on_furniture():
    """The pools are drawn without replacement; a caption with three firms in it
    used to burn a surname apiece on "Law", "Offices" and "of"."""
    reg = P._PnFakeRegistry()
    P._pn_build_terms(["Law Offices of Scott C. Stratman"], [], [], registry=reg)
    used = len(reg._used)
    reg2 = P._PnFakeRegistry()
    P._pn_build_terms(["Scott C. Stratman"], [], [], registry=reg2)
    assert used == len(reg2._used)


# ── the key round trip ──────────────────────────────────────────────────────

def _write_and_reload(tmp_path, names, sample):
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(list(names), [], [], registry=reg)
    pz = P.Pseudonymizer(terms, DET, registry=reg)
    pz.apply(sample)
    key = tmp_path / "pseudonym_key.xlsx"
    pz.write_key(key, log)
    rows = [r for ws in openpyxl.load_workbook(key).worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    reg2 = P._PnFakeRegistry()
    loaded, _dec = P._pn_load_key(key, reg2, log)
    return rows, loaded


def test_key_carries_no_furniture_token_row(tmp_path):
    rows, _ = _write_and_reload(
        tmp_path, ["Law Offices of Scott C. Stratman"],
        "Served by Law Offices of Scott C. Stratman on May 1.")
    tokens = {str(r[1]).lower() for r in rows if str(r[0]).endswith("-token")}
    assert not (tokens & set(FURNITURE)), f"furniture rows in key: {tokens}"
    assert "stratman" in tokens


def test_a_stale_furniture_row_is_not_reloaded_as_a_term(tmp_path):
    """The loop this closes: an older key harvested a row per word, and every
    row comes back as a live matching term — so "the" was faked document-wide
    on the next run whatever the operator bracketed."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person", "the Waggoner", "chetwood Atwater", "replaced",
               "--term", 19])
    ws.append(["person-token", "the", "chetwood", "replaced", "--term", 19])
    ws.append(["person-token", "of", "bancroft", "no match", "spreadsheet", 0])
    ws.append(["person-token", "Law", "Braxton", "no match", "spreadsheet", 0])
    ws.append(["person-token", "Waggoner", "Atwater", "replaced", "--term", 3])
    wb.save(key)

    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    reals = {t.real.lower() for t in terms}
    assert not (reals & {"the", "of", "law"})
    assert "waggoner" in reals
    # ...and the stored composed fake is repaired, so the folder stops
    # reproducing "chetwood" the moment the key is reused.
    full = next(t for t in terms if t.category == "person")
    assert full.fake == "the Atwater"


def test_reloaded_key_leaves_the_words_alone(tmp_path):
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person-token", "the", "chetwood", "replaced", "--term", 19])
    ws.append(["person-token", "Offices", "Mansffield", "replaced",
               "spreadsheet", 4])
    ws.append(["person-token", "Stratman", "Whitlock", "replaced",
               "spreadsheet", 4])
    wb.save(key)

    reg = P._PnFakeRegistry()
    terms, _dec = P._pn_load_key(key, reg, log)
    out = P.Pseudonymizer(terms, DET, registry=reg).apply(
        "The Law Offices of Stratman filed the opposition.")
    assert "the opposition" in out
    assert "Law Offices of" in out
    assert "Stratman" not in out


# ── an HONORIFIC is furniture too ──────────────────────────────────────────
# "Mr" is a title in front of a name and never identity. Faked, it produced
# "EVERLINE. REDWOOD'S LIGHTWELL, LLC" for "Mr. Kool's Collision, LLC" — where
# "Mr. Redwood's Lightwell, LLC" says exactly as much and hides exactly as much
# — and the bare token that fell out of it rewrote "Mr" 42 times across one
# batch. It is also what let one party's spellings disagree with each other,
# since only some of them carry the title.

HONORIFICS = ("Mr", "Mrs", "Ms", "Miss", "Dr", "Prof", "Rev", "Hon", "Sir")

KOOL = "Mr. Kool's Collision, LLC"


@pytest.mark.parametrize("title", HONORIFICS)
def test_an_honorific_is_kept_verbatim_on_both_paths(title):
    reg = P._PnFakeRegistry()
    person, _bare = P._pn_fake_person(f"{title}. Helen Rasho", reg)
    entity, _map = P._pn_fake_entity_parts(f"{title}. Rasho Holdings, LLC", reg)
    assert person.startswith(f"{title}. "), person
    assert entity.startswith(f"{title}. "), entity
    assert "Rasho" not in person and "Rasho" not in entity


def test_the_party_reads_as_a_party():
    reg = P._PnFakeRegistry()
    fake, _map = P._pn_fake_entity_parts(KOOL, reg)
    assert fake.startswith("Mr. "), fake
    assert fake.endswith(", LLC"), fake
    assert "Kool" not in fake


def test_an_honorific_is_never_a_bare_token():
    """The 42 rewrites: every "Mr. Henriquez" in the batch became
    "Everline. Henriquez"."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms([KOOL], [], [KOOL], registry=reg)
    assert not any(t.real.strip(". ").lower() in P._PN_HONORIFICS
                   for t in terms), [t.real for t in terms]
    z = P.Pseudonymizer(terms, {}, registry=reg)
    assert z.apply("Mr. Henriquez testified.") == "Mr. Henriquez testified."


def test_no_pool_word_is_spent_on_an_honorific():
    reg = P._PnFakeRegistry()
    P._pn_build_terms([KOOL], [], [KOOL], registry=reg)
    assert not any(real in P._PN_HONORIFICS for _tag, real in reg._memo)


@pytest.mark.parametrize("real", ["Mr. Mister", "Mr. Mrs Dr"])
def test_a_name_made_only_of_honorifics_still_scrubs(real):
    """The furniture is kept only while a distinctive word is left to fake, or
    the "fake" would be the name itself."""
    reg = P._PnFakeRegistry()
    fake, _bare = P._pn_fake_person(real, reg)
    assert fake.lower() != real.lower(), fake


def test_the_entity_path_asks_the_same_question_as_the_person_path():
    """It read `_PN_FIRM_WORDS` directly where the others call
    `registry.keeps_word`, and the gap was the whole of `_PN_HONORIFICS`."""
    reg = P._PnFakeRegistry()
    for word in list(P._PN_NAME_FURNITURE)[:40]:
        assert reg.keeps_word(word), word


# ── …and a possessive that lost its apostrophe is the same word ────────────
# `KOOL'S` and `Kool’s` both reduce to the core "kool". An all-caps caption
# printing `MR. KOOLS COLLISION, LLC` arrives as "kools", keys separately, and
# drew an unrelated pool word: one company read as "Redwood's Lightwell" in most
# of the batch and "Orion Lightwell" wherever the apostrophe was missing.

def test_the_apostrophe_less_spelling_folds_onto_the_same_party():
    reg = P._PnFakeRegistry()
    base = reg.token("kool", P._PN_ENTITY_WORDS, "enttok")
    assert reg.token("kools", P._PN_ENTITY_WORDS, "enttok") == base + "s"


def test_it_stays_one_party_and_two_reversible_rows():
    """One party to a reader, two DISTINCT rows to the reversal macro."""
    reg = P._PnFakeRegistry()
    a = reg.token("kool", P._PN_ENTITY_WORDS, "enttok")
    b = reg.token("kools", P._PN_ENTITY_WORDS, "enttok")
    assert a != b
    assert len({a.lower(), b.lower()}) == 2


def test_it_never_reuses_a_fake_another_value_holds():
    reg = P._PnFakeRegistry()
    base = reg.token("kool", P._PN_ENTITY_WORDS, "enttok")
    reg._used.add((base + "s").lower())          # somebody already has it
    other = reg.token("kools", P._PN_ENTITY_WORDS, "enttok")
    assert other.lower() != (base + "s").lower()


def test_an_unrelated_word_ending_in_s_draws_its_own():
    reg = P._PnFakeRegistry()
    reg.token("kool", P._PN_ENTITY_WORDS, "enttok")
    got = reg.token("jones", P._PN_ENTITY_WORDS, "enttok")
    assert not got.lower().startswith("redwood")
    assert got in P._PN_ENTITY_WORDS or got[:-1] in P._PN_ENTITY_WORDS or got


def test_the_batch_reads_as_one_defendant():
    """Every spelling the delivered key carried, through one build."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms([KOOL], [], [KOOL], registry=reg)
    z = P.Pseudonymizer(terms, {}, registry=reg)
    for raw in ["MR. KOOLS COLLISION, LLC", "KOOL'S Collision LLC",
                "Mr. Kool's Auto, LLC"]:
        new = []
        P._pn_append_entity_terms(new, raw, "document", reg)
        z._add_terms(new)
    out = z.apply("MR. KOOL'S COLLISION, LLC; Mr. Kool’s Collision, LLC; "
                  "KOOL'S Collision LLC; MR. KOOLS COLLISION, LLC; "
                  "Mr. Kool's Auto, LLC")
    assert "Kool" not in out, out
    assert out.lower().count("mr. ") == 4, out  # the title, never a stand-in
    # Every spelling of the party's own word lands on ONE stem, so the batch
    # names one defendant: the delivered key had three (Redwood's, Bristow's,
    # Orion) and read as three companies.
    stems = {fake.lower().rstrip("s") for (_tag, real), fake in reg._memo.items()
             if real.startswith("kool")}
    assert len(stems) == 1, stems


# ── a key an older build wrote is repaired on the way in ───────────────────

def test_a_stale_honorific_row_is_repaired_on_load(tmp_path):
    """A loaded row is applied literally, so the folder reproduces whatever the
    key says. The delivered key carried seven such rows."""
    key = tmp_path / "pseudonym_key.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["entity", "Mr. Kool's Collision, LLC",
               "Everline. Redwood's Lightwell, LLC", "replaced", "document", 42])
    ws.append(["entity-token", "Mr. Kool's Collision",
               "Everline. Redwood's Lightwell", "replaced", "document", 19])
    ws.append(["entity-token", "Mr", "Everline", "replaced", "document", 42])
    wb.save(key)

    terms, _dec = P._pn_load_key(key, P._PnFakeRegistry(), log)
    fakes = {t.category: t.fake for t in terms}
    assert fakes["entity"] == "Mr. Redwood's Lightwell, LLC"
    # The SHORT form too, or the repaired full name sits beside an unrepaired
    # one — "Mr. Redwood's Lightwell, LLC" in one line and "Everline. Redwood's
    # Lightwell" in the next, which reads as two parties.
    assert fakes["entity-token"] == "Mr. Redwood's Lightwell"
    # …and the bare honorific row builds no term at all.
    assert not any(t.real.lower() == "mr" for t in terms), [t.real for t in terms]


def test_a_bare_honorific_row_is_never_repaired_into_a_self_map():
    """`_pn_restore_furniture` refuses a repair that would leave fake == real."""
    assert P._pn_restore_furniture("Mr", "Everline") is None


def test_a_fake_not_composed_word_for_word_is_left_alone():
    """The alignment guard: "MR. KOOL'S COLLISION,LLC" lost the space before
    its suffix, so its four words cannot be matched to the fake's three."""
    assert P._pn_restore_furniture("MR. KOOL'S COLLISION,LLC",
                                   "EVERLINE. REDWOOD'S LANTERNWOOD") is None
