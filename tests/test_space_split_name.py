"""A name EXTRACTION broke in half is still the party's name.

A born-digital pleading kerns a capital pair tightly, and the gap the kern
leaves is wide enough that the character joiner reads it as a space. The
printed page is perfectly normal; the export is not. One Song-Beverly fee
batch shipped

    VADIM SARKISYAN and SERGEY MIRZOYANS   ->  SHARNBROOK WRIGHTSON and
                                               RESTARICK GEDNEY
    V ADIM SARKISY AN  and  MIRZOY ANS     ->  unchanged

in its captions, its attorney line, its proof of service and its service list,
across four exports — while the same key's tokens scrubbed those two surnames
287 and 234 times everywhere the spelling came out whole. The complaint's first
page carries both forms four lines apart: line 10 reads `SHARNBROOK WRIGHTSON`
and line 11 `RESTARICK MIRZOY ANS`. That is worse than a plain leak — a
drafting pass read the two spellings as two different sets of plaintiffs and
reported a record inconsistency.

The reduced weld pass WOULD match it (`SARKISY AN` folds to `sarkisyan`) and
deliberately refuses: `_pn_span_is_unbroken` rejects a match holding a printed
word boundary, which is the rule that stopped it deleting the text between two
real words ("Further, a substantial" -> "Furtthorpe substantial"). That rule
stands. The split spelling is registered as a real TERM instead — the way the
wrap-split spelling of a compound surname already is — so it still yields to
citation protection and to an operator KEEP, which a blind reduced substitution
cannot.

Run:  cd PDF-Linker && python3 -m pytest tests/test_space_split_name.py -v
"""
import logging

import openpyxl
import pytest

import pdf_linker as P

log = logging.getLogger("test")
ALT = P._PN_KEY_ALT_STATUS

PARTIES = ["Vadim Sarkisyan", "Sergey Mirzoyans"]

# The shapes the delivered exports actually carry, verbatim.
CAPTION = (
    " 5  Attorneys for Plaintiffs, V ADIM SARKISY AN and SERGEY MIRZOY ANS\n"
    "11  V ADIM SARKISY AN, an individual; Case No. 25STCV63891\n"
    "12  SERGEY MIRZOY ANS, an individual,\n"
    " 4  Primary driver or drivers of the motor vehicle V adim Sarkisyan\n"
)


def _run(names, text, registry=None):
    reg = registry if registry is not None else P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(names, [], [], registry=reg),
                        {}, registry=reg)
    return z, z.apply(text)


# ── the leak ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("leaked", [
    "V ADIM SARKISY AN",
    "SARKISY AN",
    "MIRZOY ANS",
    "V adim",
    "SARKISY",
    "MIRZOY",
])
def test_the_split_spelling_does_not_reach_the_export(leaked):
    _z, out = _run(PARTIES, CAPTION)
    assert leaked not in out, out


def test_no_fragment_of_either_party_survives():
    _z, out = _run(PARTIES, CAPTION)
    for real in ("vadim", "sarkisyan", "sergey", "mirzoyans",
                 "adim", "sarkisy", "mirzoy"):
        assert real not in out.lower(), (real, out)


def test_the_split_spelling_and_the_whole_one_are_ONE_party():
    """The half-scrubbed caption is what made a drafting pass report two
    different sets of plaintiffs. Both spellings must land on the same fake."""
    _z, split = _run(PARTIES, "V ADIM SARKISY AN and SERGEY MIRZOY ANS")
    _z2, whole = _run(PARTIES, "VADIM SARKISYAN and SERGEY MIRZOYANS")
    assert split == whole


def test_a_spelling_that_never_broke_is_unchanged():
    """The whole spelling scrubbed correctly before this and must still."""
    reg = P._PnFakeRegistry()
    z, out = _run(PARTIES, "VADIM SARKISYAN, an individual;", registry=reg)
    fake = z.records[("person", "vadim sarkisyan")]["fake"]
    assert out == f"{fake.upper()}, an individual;"


def test_our_own_output_is_left_alone():
    """`SHARNBROOK WRIGHTSON` — the fake already standing in the export beside
    the leaked spelling — is not touched by any variant."""
    z, _out = _run(PARTIES, CAPTION)
    fake = z.records[("person", "vadim sarkisyan")]["fake"]
    _z2, again = _run(PARTIES, f"{fake.upper()}, an individual;")
    assert again == f"{fake.upper()}, an individual;"


# ── what it costs ───────────────────────────────────────────────────────────

def _without_split(monkeypatch):
    monkeypatch.setattr(P, "_pn_space_split_spellings", lambda w: set())


def test_no_pool_word_is_drawn_and_no_fake_moves(monkeypatch):
    """A split spelling shares the token's own fake, so it draws nothing: the
    used-pool and every binding must come out identical to a build without it.
    A fake that moved would re-scrub a folder already sent out under the old
    one."""
    def build():
        reg = P._PnFakeRegistry()
        terms = P._pn_build_terms(PARTIES, ["25STCV63891"], [], registry=reg)
        bindings = {(t.category, t.real): t.fake for t in terms
                    if " " not in t.real or t.category != "person-token"}
        return sorted(reg._used), bindings

    _without_split(monkeypatch)
    used_before, bind_before = build()
    monkeypatch.undo()
    used_after, bind_after = build()
    assert used_before == used_after
    assert bind_before == bind_after


def test_ordinary_legal_prose_is_untouched():
    """Every printed word pair in a paragraph of a real complaint, against two
    parties whose tokens split seven ways each."""
    prose = (
        "Plaintiffs are informed and believe, and thereon allege, that in "
        "doing the things hereinafter alleged Defendants, and each of them, "
        "were acting in the course and scope of their employment as such "
        "agents, servants, and/or employees, and with the permission, "
        "consent, knowledge, and/or ratification of their Co-Defendants, "
        "principals, and/or employers. The Vehicle was/is a \"new motor "
        "vehicle\" under The Song-Beverly Warranty Act. As he was not a party "
        "to the within action, service was made in the manner ascribed.")
    _z, out = _run(PARTIES, prose)
    assert out == prose


def test_a_lower_case_occurrence_is_left_alone():
    """A bare token never matches a lower-case occurrence
    (`_pn_term_is_cap_only`), and a split spelling of one is still a bare
    token. Residual and accepted, exactly as it is for the whole token."""
    _z, out = _run(PARTIES, "the sarkisy an filing")
    assert out == "the sarkisy an filing"


# ── the guards ──────────────────────────────────────────────────────────────

def test_a_short_token_is_never_split():
    """Under `_PN_SPACE_SPLIT_MIN` the halves carry too little for the
    concatenation to be evidence of anything."""
    for short in ("Kim", "Wu", "Ng", "Doe"):
        assert P._pn_space_split_spellings(short) == set()


def test_a_trailing_single_letter_is_never_a_half():
    """"Sarkisya N" is the shape of a middle INITIAL, which a filing really
    writes; the leading half has no such twin."""
    out = P._pn_space_split_spellings("Sarkisyan")
    assert not any(v.split()[-1] and len(v.split()[-1]) < 2 for v in out), out
    assert "Sarkisya n" not in out


def test_a_one_letter_LEADING_half_is_kept():
    """It is where the observed breaks fell — "V ADIM", "V adim"."""
    assert "V adim" in P._pn_space_split_spellings("Vadim")
    assert "S arkisyan" in P._pn_space_split_spellings("Sarkisyan")


def test_two_ordinary_words_are_not_a_split_spelling():
    """"Nowhere" would offer "No where", which a sentence writes by accident.
    The concatenation is the corroboration everywhere else; where both halves
    are ordinary vocabulary it corroborates nothing."""
    assert "No where" not in P._pn_space_split_spellings("Nowhere")
    assert "In action" not in P._pn_space_split_spellings("Inaction")


def test_the_split_is_between_letters():
    """A hyphen or an inner dot is already a printed boundary — the wrap-split
    spelling of a compound surname has its own registration."""
    for v in P._pn_space_split_spellings("Ardeshirpour-Zartoshti"):
        left, right = v.split(" ")
        assert left[-1].isalpha() and right[0].isalpha(), v


def test_only_an_authoritative_source_is_split():
    """A split spelling is a guess about how a printed word came apart.
    Stacking it on a guess about who the party IS doubles the way it can be
    wrong, so a document harvest never earns one."""
    def variants(source):
        reg = P._PnFakeRegistry()
        terms = []
        P._pn_append_person_terms(terms, "Vadim Sarkisyan", source, reg)
        return {t.real for t in terms if " " in t.real}

    assert "Sarkisy an" in variants("spreadsheet")
    assert "Sarkisy an" in variants("--term")
    for guess in ("document", "declarant", "judge", "court-staff"):
        assert not any(" " in v and v != "Vadim Sarkisyan"
                       for v in variants(guess)), guess


# ── reversal, and the re-run ────────────────────────────────────────────────

def _key_rows(path):
    wb = openpyxl.load_workbook(path)
    head = [str(h).strip().lower()
            for h in next(wb.active.iter_rows(max_row=1, values_only=True))]
    return [dict(zip(head, r)) for ws in wb.worksheets
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]


def test_the_split_spelling_does_not_own_the_reversal(tmp_path):
    """It is a spelling this tool invented, so it earns a key row by MATCHING
    and is marked forward-only: `DeAnonymize` retires a fake two Real Values
    claim, which would leave the pseudonym standing in the tentative."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    z.apply(CAPTION)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)
    rows = _key_rows(path)

    surname_fake = z.records[("person-token", "sarkisyan")]["fake"]
    group = [r for r in rows
             if str(r["replacement"]).lower() == surname_fake.lower()]
    owners = [r for r in group
              if str(r["status"]).strip().lower() != ALT]
    assert len(owners) == 1, group
    assert str(owners[0]["real value"]) == "Sarkisyan"
    assert "Sarkisy an" in {str(r["real value"]) for r in group}


def test_the_key_alone_reproduces_the_delivered_export(tmp_path):
    """A re-run has the key and may have no party template to rebuild the
    variants from. The split spelling is pinned, so the second export is the
    first one byte for byte."""
    reg = P._PnFakeRegistry()
    z = P.Pseudonymizer(P._pn_build_terms(PARTIES, [], [], registry=reg),
                        {}, registry=reg)
    first = z.apply(CAPTION)
    path = tmp_path / "pseudonym_key.xlsx"
    z.write_key(path, log)

    reg2 = P._PnFakeRegistry()
    terms, *_ = P._pn_load_key(path, reg2, log)
    z2 = P.Pseudonymizer(terms, {}, registry=reg2)
    assert z2.apply(CAPTION) == first
