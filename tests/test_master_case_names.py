"""
The cross-case master workbook names a case folder by its PSEUDONYM.

That workbook is permanent, lives outside every case folder (next to the
config, routinely on a synced drive) and is never pruned, so the real folder
name in its Cases and Origin columns made the one file whose purpose is
cross-matter history a standing list of every matter's parties. It carries
this case's own stand-ins where the run can prove that is safe, and an opaque
per-folder id where it cannot — never half of each.

Run:  cd PDF-Linker && python3 -m pytest tests/test_master_case_names.py -v
"""
import logging

import openpyxl

import pdf_linker as P

log = logging.getLogger("test")


def _pz(*terms):
    reg = P._PnFakeRegistry()
    built = P._pn_build_terms([], [], list(terms), registry=reg)
    return P.Pseudonymizer(built, {}, registry=reg)


def _fake_for(pz, real):
    return next(r["fake"] for r in pz.records.values()
                if r["real"].lower() == real.lower())


# ── the label ───────────────────────────────────────────────────────────────

def test_a_bound_folder_name_is_written_as_this_case_s_own_fakes():
    """The point of preferring the readable form: the row is traceable through
    `pseudonym_key.xlsx`, because the stand-in is the one the exports use."""
    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark - MTC", pz)
    assert "Rasho" not in label and "Quillmark" not in label
    assert label == (f"{_fake_for(pz, 'Rasho')} v "
                     f"{_fake_for(pz, 'Quillmark')} - MTC")


def test_the_label_is_stable_across_runs():
    """A fake is seeded on its real value, so two independent runs over the
    same folder and party list agree — which is what lets the Cases column be a
    set of MATTERS rather than a list of run outputs."""
    assert (P._pn_case_label("Rasho v Quillmark", _pz("Rasho", "Quillmark"))
            == P._pn_case_label("Rasho v Quillmark", _pz("Rasho", "Quillmark")))


def test_a_half_bound_name_falls_back_whole_rather_than_shipping_half_real():
    """A party the folder names that this key never bound is the case the id
    exists for — and `surviving_reals` cannot see it, since nothing TRACKS it.
    The leftover screen is what catches it, and it takes the WHOLE name with
    it: a partly real name is the failure being replaced."""
    pz = _pz("Rasho")                      # Quillmark is not a term here
    label = P._pn_case_label("Rasho v Quillmark", pz)
    assert label == P._pn_case_opaque("Rasho v Quillmark")
    assert "Quillmark" not in label and "Rasho" not in label


def test_docket_shorthand_does_not_cost_a_folder_its_readable_label():
    """The screen has to leave the abbreviations folders are really named with
    alone, or every label would be an id."""
    pz = _pz("Rasho", "Quillmark")
    assert P._pn_case_label("Rasho v Quillmark - MTC", pz).endswith(" - MTC")
    assert P._pn_case_label("Rasho MSJ 25STCV14710", pz).endswith(
        " MSJ 25STCV14710")                 # a case number is not name-shaped


def test_the_screen_is_strict_because_the_fallback_is_cheap():
    """An unbound name-shaped word forces the id even where it is ordinary
    docket vocabulary ("Ex Parte") — losing a readable label costs one folder's
    rows their readability, where guessing wrong costs a real name in a
    permanent cross-case file."""
    pz = _pz("Rasho", "Quillmark")
    assert P._pn_case_label("Rasho v Quillmark Ex Parte", pz) == \
        P._pn_case_opaque("Rasho v Quillmark Ex Parte")


def test_an_untouched_name_is_not_evidence_of_a_safe_one():
    """Nothing bound in the name means nothing established about it, so it
    takes the id too — the same guard `_real_remainder` states."""
    pz = _pz("Rasho")
    assert (P._pn_case_label("Dept 55 Tentatives", pz)
            == P._pn_case_opaque("Dept 55 Tentatives"))


def test_no_pseudonymizer_still_never_yields_the_real_name():
    assert P._pn_case_label("Rasho v Quillmark") == \
        P._pn_case_opaque("Rasho v Quillmark")


def test_the_label_never_mints_a_fake_or_counts_an_occurrence():
    """A folder name is not a document: it must not inflate the counts
    `write_key` reports, nor add a binding to the key."""
    pz = _pz("Rasho")
    before = {k: r["count"] for k, r in pz.records.items()}
    P._pn_case_label("Rasho v Quillmark", pz)
    assert {k: r["count"] for k, r in pz.records.items()} == before
    assert len(pz.records) == len(before)


# ── the id ──────────────────────────────────────────────────────────────────

def test_the_id_is_derived_from_the_name_alone_and_is_stable():
    assert P._pn_case_id("Rasho v Quillmark") == P._pn_case_id(
        "  rasho   V QUILLMARK ")               # normalized, case-folded
    assert P._pn_case_id("Rasho v Quillmark") != P._pn_case_id("Other Matter")
    assert len(P._pn_case_id("x")) == P._PN_CASE_ID_LEN
    assert P._pn_case_id("") == ""


def test_origin_carries_the_id_beside_the_readable_label():
    """`_pn_decision_is_ours` runs before any key is loaded or term built, so
    the readable half cannot be re-derived there — the id rides along."""
    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark", pz)
    origin = P._pn_case_origin("Rasho v Quillmark", label)
    assert origin == f"{label} [{P._pn_case_id('Rasho v Quillmark')}]"
    # the opaque form already IS the id and takes no suffix
    assert P._pn_case_origin("Dept 55") == P._pn_case_opaque("Dept 55")


# ── ownership ───────────────────────────────────────────────────────────────

def test_ownership_is_settled_on_the_id_without_any_bindings_in_hand():
    pz = _pz("Rasho", "Quillmark")
    origin = P._pn_case_origin(
        "Rasho v Quillmark", P._pn_case_label("Rasho v Quillmark", pz))
    assert P._pn_decision_is_ours({"origin": origin}, "Rasho v Quillmark")
    assert not P._pn_decision_is_ours({"origin": origin}, "Another Matter")


def test_ownership_still_recognises_a_workbook_written_by_an_older_version():
    """An Origin holding the REAL folder name, or a lone Cases entry, still
    identifies its author — the next write migrates it."""
    assert P._pn_decision_is_ours({"origin": "Rasho v Quillmark"},
                                  "Rasho v Quillmark")
    assert P._pn_decision_is_ours({"cases": "Rasho v Quillmark"},
                                  "Rasho v Quillmark")
    assert not P._pn_decision_is_ours({"cases": "Some Other Matter"},
                                      "Rasho v Quillmark")


def test_the_opaque_form_is_recognised_too_so_a_label_may_flip():
    """An early run with nothing bound writes the id; a later one with a key
    writes the readable label. Both are this folder."""
    assert P._pn_decision_is_ours(
        {"origin": P._pn_case_opaque("Rasho v Quillmark")}, "Rasho v Quillmark")


# ── migration of a workbook already carrying real names ─────────────────────

def test_a_stored_real_case_name_is_rewritten_not_accumulated(tmp_path):
    mp = tmp_path / "master.xlsx"
    P._pn_update_master_leaks(mp, [("Acme LLC", "LEAK")], "Rasho v Quillmark",
                              "2026-01-01", log)          # an older version
    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark", pz)
    P._pn_update_master_leaks(
        mp, [("Acme LLC", "LEAK")], label, "2026-02-02", log,
        aliases=P._pn_case_aliases("Rasho v Quillmark", label))
    row = next(r for r in openpyxl.load_workbook(mp).active.iter_rows(
        min_row=2, values_only=True) if r[0] == "Acme LLC")
    assert row[3] == label          # one entry, and it is the pseudonym
    assert "Rasho" not in row[3] and "Quillmark" not in row[3]
    assert row[2] == 2              # still the same matter, seen twice


def test_migration_leaves_another_matter_s_entries_alone(tmp_path):
    mp = tmp_path / "master.xlsx"
    P._pn_update_master_leaks(mp, [("Acme LLC", "LEAK")], "Other Matter",
                              "2026-01-01", log)
    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark", pz)
    P._pn_update_master_leaks(
        mp, [("Acme LLC", "LEAK")], label, "2026-02-02", log,
        aliases=P._pn_case_aliases("Rasho v Quillmark", label))
    row = next(r for r in openpyxl.load_workbook(mp).active.iter_rows(
        min_row=2, values_only=True) if r[0] == "Acme LLC")
    assert "Other Matter" in row[3] and label in row[3]


def test_a_keep_row_s_origin_migrates_and_keeps_its_authorship(tmp_path,
                                                               monkeypatch):
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master.xlsx"))
    cfg, dec = {}, {"alder law, p.c.": {"value": "Alder Law, P.C.",
                                        "fix": "no", "fixcell": "no",
                                        "type": "KEEP", "notes": ""}}
    P._pn_update_master_keep(cfg, dec, "Rasho v Quillmark", "2026-01-01", log)
    stored = P._pn_read_master_keep(cfg)["alder law, p.c."]
    assert stored["origin"] == "Rasho v Quillmark"     # the older version's row
    assert P._pn_decision_is_ours(stored, "Rasho v Quillmark")

    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark", pz)
    P._pn_update_master_keep(
        cfg, dec, label, "2026-02-02", log,
        aliases=P._pn_case_aliases("Rasho v Quillmark", label),
        origin=P._pn_case_origin("Rasho v Quillmark", label))
    stored = P._pn_read_master_keep(cfg)["alder law, p.c."]
    assert "Rasho" not in stored["origin"] and "Rasho" not in stored["cases"]
    assert stored["cases"] == label                    # not two entries
    assert P._pn_decision_is_ours(stored, "Rasho v Quillmark")


def test_another_matter_s_origin_is_never_restated_as_ours(tmp_path,
                                                           monkeypatch):
    """Rewriting it would hand this folder another matter's keeps — the
    faking half of a keep-spec is the author's alone."""
    monkeypatch.setenv("PDF_LINKER_MASTER", str(tmp_path / "master.xlsx"))
    cfg, dec = {}, {"alder law, p.c.": {"value": "Alder Law, P.C.",
                                        "fix": "no", "fixcell": "no",
                                        "type": "KEEP", "notes": ""}}
    P._pn_update_master_keep(cfg, dec, "Other Matter", "2026-01-01", log)
    pz = _pz("Rasho", "Quillmark")
    label = P._pn_case_label("Rasho v Quillmark", pz)
    P._pn_update_master_keep(
        cfg, dec, label, "2026-02-02", log,
        aliases=P._pn_case_aliases("Rasho v Quillmark", label),
        origin=P._pn_case_origin("Rasho v Quillmark", label))
    stored = P._pn_read_master_keep(cfg)["alder law, p.c."]
    assert stored["origin"] == "Other Matter"
    assert not P._pn_decision_is_ours(stored, "Rasho v Quillmark")


def test_a_label_can_never_split_its_own_cases_cell():
    """`Cases` is a semicolon-joined LIST: a folder named "Smith; Jones v Acme"
    would come back as two matters and never match itself again."""
    pz = _pz("Smith", "Jones", "Rasho")
    label = P._pn_case_label("Smith; Jones v Rasho", pz)
    assert ";" not in label
    assert P._pn_decision_is_ours(
        {"origin": P._pn_case_origin("Smith; Jones v Rasho", label)},
        "Smith; Jones v Rasho")


def test_a_label_never_carries_a_bracket_the_id_reader_would_claim():
    pz = _pz("Rasho")
    label = P._pn_case_label("Rasho [abcdef12]", pz)
    assert P._PN_CASE_ID_RE.search(label) is None
    origin = P._pn_case_origin("Rasho [abcdef12]", label)
    assert P._PN_CASE_ID_RE.search(origin).group(1) == \
        P._pn_case_id("Rasho [abcdef12]")


def test_a_citation_shaped_folder_name_cannot_ship_a_real_party():
    """A folder named like an authority puts the substitution against its own
    citation protection. Whichever way that falls, the leftover screen is what
    settles it — it asks nothing of any parse."""
    pz = _pz("Rasho")                     # Quillmark unbound
    label = P._pn_case_label("Rasho v. Quillmark (2021) 71 Cal.App.5th 358", pz)
    assert "Quillmark" not in label and "Rasho" not in label


def test_a_recycled_stand_in_is_still_recognised_as_ours():
    """A pool that ran out mints "<word><n>" (`_pn_recycled_fake`), and a word
    regex that stops at the first digit asks `known_fake_words` for a word it
    does not hold — so the screen would call this run's own output an unknown
    name and every such folder would lose its readable label."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(
        [], [], [f"Party{i} Surname{i}" for i in range(520)], registry=reg)
    pz = P.Pseudonymizer(terms, {}, registry=reg)
    label = P._pn_case_label("Party7 Surname7 v Party9 Surname9 - MTC", pz)
    assert label != P._pn_case_opaque("Party7 Surname7 v Party9 Surname9 - MTC")
    assert "Party7" not in label and "Surname9" not in label
    assert label.endswith(" - MTC")
