"""
`**CORRECT TEXT` is the DURABLE OCR fix: the same correction `*CORRECT TEXT`
makes, remembered on the master workbook's KEEP sheet and applied in every
folder from then on. It is for a scan's habitual misreading of a GENERIC term
— "SanDiega" for San Diego, a courthouse, a code name — that would otherwise
be answered again in every folder it turns up in.

A single star is a statement about one scan of one document and is never
persisted. An inherited fix touches nothing where its garble is absent: the
term matches nothing, and an unmatched OCR-fix row is never written to the key.

Run:  cd PDF-Linker && python3 -m pytest tests/test_durable_ocr_fix.py -v
"""
import logging
import sys
import types
import warnings
import zipfile

import fitz
import openpyxl

import pdf_linker as P

warnings.filterwarnings("ignore", category=DeprecationWarning)
log = logging.getLogger("test")
DET = {k: P._PN_DETECTORS[k] for k in P._PN_DEFAULT_DETECTORS}
HDR = ("Value", "Fix? (yes/no)", "Type", "Notes", "Cases", "Origin")
LEAK_HDR = ("Value", "Fix? (yes/no)", "File", "Type", "Where (page:line)",
            "Notes")


def _decisions(*rows):
    return P._pn_parse_decision_rows(
        [HDR] + [(v, c, "", "", "", "") for v, c in rows])


# ── 1. parsing ───────────────────────────────────────────────────────────────

def test_two_stars_strip_to_the_correct_text():
    assert P._pn_ocr_fix_target("**San Diego") == "San Diego"
    assert P._pn_ocr_fix_target("** San Diego ") == "San Diego"
    assert P._pn_ocr_fix_target('**"San Diego"') == "San Diego"
    assert P._pn_ocr_fix_target("*San Diego") == "San Diego"
    assert P._pn_ocr_fix_target("**") is None


def test_durability_is_read_off_the_second_star_only():
    assert P._pn_ocr_fix_durable("**San Diego")
    assert P._pn_ocr_fix_durable("**David {said}")
    assert not P._pn_ocr_fix_durable("*San Diego")
    assert not P._pn_ocr_fix_durable("*David {said}")
    assert not P._pn_ocr_fix_durable("~San Diego")
    assert not P._pn_ocr_fix_durable("San Diego")
    assert not P._pn_ocr_fix_durable("**")


def test_the_decision_carries_the_flag_from_both_readers():
    d = _decisions(("SanDiega", "**San Diego"), ("Smlth", "*Smith"),
                   ("avidsaid", "**David {said}"))
    assert d["sandiega"]["fix"] == "yes"
    assert d["sandiega"]["ocr_fix"] == "San Diego"
    assert d["sandiega"]["ocr_durable"] is True
    assert d["smlth"]["ocr_fix"] == "Smith"
    assert d["smlth"]["ocr_durable"] is False
    assert d["avidsaid"]["ocr_durable"] is True
    assert d["avidsaid"]["fake_values"] == ["avid"]
    # Neither is a KEEP, so neither retires a row or protects a span.
    assert not P._pn_decision_is_keep(d["sandiega"])
    assert P._pn_ocr_whole_text(d["sandiega"]) == "San Diego"
    assert P._pn_ocr_whole_text(d["avidsaid"]) == "David said"


def test_the_key_loader_reads_the_two_stars_too(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_KEY_MAIN_SHEET
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person", "John Smith", "Fake Party", "replaced", "--term", "1"])
    ws.append(["entity", "SanDiega", "**San Diego", "replaced", "prescan", "2"])
    ws.append(["entity", "LosAngelas", "*Los Angeles", "replaced", "prescan", "2"])
    wb.save(tmp_path / "pseudonym_key.xlsx")
    reg = P._PnFakeRegistry()
    out = P._pn_load_key(tmp_path / "pseudonym_key.xlsx", reg, log)
    decisions = out[-1] if isinstance(out, tuple) else out
    if not isinstance(decisions, dict) or "sandiega" not in decisions:
        decisions = next(x for x in out if isinstance(x, dict)
                         and "sandiega" in x)
    assert decisions["sandiega"]["ocr_fix"] == "San Diego"
    assert decisions["sandiega"]["ocr_durable"] is True
    assert decisions["sandiega"]["type"] == P._PN_OCR_FIX_TYPE
    assert decisions["losangelas"]["ocr_durable"] is False


# ── 2. persistence ───────────────────────────────────────────────────────────

def _master_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[P._PN_MASTER_KEEP_SHEET]
    return {str(r[0]): r for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0]}


def test_a_durable_fix_lands_on_the_master_sheet_under_its_own_type(tmp_path):
    cfg = {"master_leaks_path": str(tmp_path / "master.xlsx")}
    d = _decisions(("SanDiega", "**San Diego"))
    P._pn_update_master_keep(cfg, d, "Case A", "2026-09-05", log)
    rows = _master_rows(tmp_path / "master.xlsx")
    assert rows["SanDiega"][1] == "**San Diego"
    assert rows["SanDiega"][2] == P._PN_OCR_FIX_TYPE
    # …and reads back as the same decision from the master sheet.
    back = P._pn_read_master_keep(cfg)
    assert back["sandiega"]["ocr_fix"] == "San Diego"
    assert back["sandiega"]["ocr_durable"] is True


def test_the_full_run_persists_two_stars_and_not_one(tmp_path, monkeypatch):
    """A worksheet `**` reaches the master KEEP sheet; a worksheet `*` beside
    it does not — one scan of one document is nobody else's business."""
    case = _word_case(tmp_path / "Case A",
                      "Plaintiff Hollis Vantreight lives in SanDiega, near the "
                      "Superior Court. Defendant Marcus Smlth denies it.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LEAKS"
    ws.append(list(LEAK_HDR))
    ws.append(["SanDiega", "**San Diego", "Motion.docx", "name?", "p.1", ""])
    ws.append(["Smlth", "*Smith", "Motion.docx", "name?", "p.1", ""])
    wb.save(case / "LEAKS.xlsx")
    master = tmp_path / "master.xlsx"
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])
    try:
        P.main()
    except SystemExit:
        pass
    txt = next((case / "Text Files").glob("*.txt")).read_text(encoding="utf-8")
    assert "San Diego" in txt and "SanDiega" not in txt
    assert "Smith" in txt and "Smlth" not in txt
    rows = _master_rows(master)
    assert "SanDiega" in rows and rows["SanDiega"][2] == P._PN_OCR_FIX_TYPE
    assert "Smlth" not in rows


# ── 3. inheritance ───────────────────────────────────────────────────────────

def _word_case(case, body_text):
    case.mkdir()
    body = "".join(f"<w:p><w:r><w:t>{t}</w:t></w:r></w:p>"
                   for t in body_text.split("\n"))
    with zipfile.ZipFile(case / "Motion.docx", "w") as z:
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Case Number", "Title Plaintiff", "Title Defendant"])
    ws.append(["24STCV00001", "Hollis Vantreight", "Cascadia Freight, Inc."])
    wb.save(case / "Order_Mine.xlsx")
    return case


def _master_with(path, *rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = P._PN_MASTER_KEEP_SHEET
    ws.append(list(P._PN_MASTER_KEEP_HEADERS))
    for value, cell, vtype in rows:
        ws.append([value, cell, vtype, 1, "Case 0f0f0f0f", "2026-09-01",
                   "2026-09-01", "", "Case 0f0f0f0f"])
    wb.save(path)


def test_a_fresh_folder_inherits_the_correction(tmp_path, monkeypatch):
    """Case B never typed anything: the master sheet carries `**San Diego`
    from another matter, and Case B's export reads San Diego where its
    document said SanDiega — with no worksheet row asking about it."""
    master = tmp_path / "master.xlsx"
    _master_with(master, ("SanDiega", "**San Diego", P._PN_OCR_FIX_TYPE))
    case = _word_case(tmp_path / "Case B",
                      "Plaintiff Hollis Vantreight lives in SanDiega, near the "
                      "Superior Court.")
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])
    try:
        P.main()
    except SystemExit:
        pass
    txt = next((case / "Text Files").glob("*.txt")).read_text(encoding="utf-8")
    assert "lives in San Diego" in txt and "SanDiega" not in txt
    assert "Hollis Vantreight" not in txt
    # The inherited fix that corrected text here is re-affirmed: Case B joins
    # the row's Cases so the sheet says where the garble has turned up.
    rows = _master_rows(master)
    assert rows["SanDiega"][2] == P._PN_OCR_FIX_TYPE
    assert int(rows["SanDiega"][3]) == 2
    # The correction is never a REVERSAL: the body was corrected before it
    # was scrubbed, so nothing about the garble reaches the key's main sheet
    # (the fix lives on the master sheet, which is where a `**` belongs).
    kb = openpyxl.load_workbook(case / "pseudonym_key.xlsx")
    main_reals = {str(r[1]).lower() for r in kb[P._PN_KEY_MAIN_SHEET]
                  .iter_rows(min_row=2, values_only=True) if r[1]}
    assert "sandiega" not in main_reals
    leaks = case / "LEAKS.xlsx"
    if leaks.exists():
        vals = {str(r[0]).lower() for r in openpyxl.load_workbook(leaks).active
                .iter_rows(min_row=2, values_only=True) if r and r[0]}
        assert "sandiega" not in vals


def test_a_folder_without_the_garble_is_untouched(tmp_path, monkeypatch):
    """The inherited fix mints nothing where the garble is absent: no key
    row, no worksheet row, and the master row is not re-affirmed."""
    master = tmp_path / "master.xlsx"
    _master_with(master, ("SanDiega", "**San Diego", P._PN_OCR_FIX_TYPE))
    case = _word_case(tmp_path / "Case C",
                      "Plaintiff Hollis Vantreight lives in Fresno, near the "
                      "Superior Court.")
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    monkeypatch.setattr(sys, "argv", ["pdf_linker.py", str(case)])
    try:
        P.main()
    except SystemExit:
        pass
    kb = openpyxl.load_workbook(case / "pseudonym_key.xlsx")
    for name in kb.sheetnames:
        reals = {str(r[1]).lower() for r in kb[name]
                 .iter_rows(min_row=2, values_only=True) if r and r[1]}
        assert "sandiega" not in reals, (name, reals)
    rows = _master_rows(master)
    assert int(rows["SanDiega"][3]) == 1
    leaks = case / "LEAKS.xlsx"
    if leaks.exists():
        vals = {str(r[0]).lower() for r in openpyxl.load_workbook(leaks).active
                .iter_rows(min_row=2, values_only=True) if r and r[0]}
        assert "sandiega" not in vals


def test_an_inherited_fix_takes_the_bound_names_stand_in(tmp_path):
    """Where the correct text IS a party this case binds, the garble reads as
    that party's stand-in — the `*` rule, inherited unchanged."""
    reg = P._PnFakeRegistry()
    terms = P._pn_build_terms(["John Smith"], [], [], registry=reg)
    d = _decisions(("Smlth", "**Smith"))
    terms = P._pn_apply_ocr_fixes(d, terms, reg, log)
    smith = next(t for t in terms if t.real == "Smith")
    fix = next(t for t in terms if t.real == "Smlth")
    assert fix.fake == smith.fake and fix.ocr_fix


def test_fix_leaks_applies_an_inherited_correction(tmp_path, monkeypatch):
    td = tmp_path / "Text Files"
    td.mkdir()
    (td / "Brief.txt.LEAK").write_text(
        "====== Page 1 ======\nThe hearing was set in SanDiega County.\n",
        encoding="utf-8")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Pseudonym Key"
    ws.append(["Category", "Real Value", "Replacement", "Status", "Source",
               "Occurrences"])
    ws.append(["person", "Filler Party", "Fake Party", "replaced", "--term", "1"])
    wb.save(tmp_path / "pseudonym_key.xlsx")
    wb2 = openpyxl.Workbook(); w2 = wb2.active; w2.title = "LEAKS"
    w2.append(["File", "Type", "Value", "Where (page:line)", "Fix? (yes/no)",
               "Notes"])
    w2.append(["Brief.txt.LEAK", "REVIEW", "SanDiega", "p.1:1", "", ""])
    wb2.save(tmp_path / "LEAKS.xlsx")
    master = tmp_path / "master.xlsx"
    _master_with(master, ("SanDiega", "**San Diego", P._PN_OCR_FIX_TYPE))
    monkeypatch.setenv("PDF_LINKER_MASTER", str(master))
    args = types.SimpleNamespace(term=[], key=str(tmp_path / "pseudonym_key.xlsx"))
    P._fix_leaks_mode(tmp_path, args, {}, log)
    body = next(td.glob("Brief.txt*")).read_text(encoding="utf-8")
    assert "San Diego County" in body and "SanDiega" not in body


def test_the_prompt_names_the_two_star_form(tmp_path):
    P._pn_write_leak_report(tmp_path, [
        {"file": "Motion.pdf", "type": "name?", "value": "SanDiega",
         "where": "p.1:14", "notes": "", "context": "lives in SanDiega."}], log)
    wb = openpyxl.load_workbook(tmp_path / "LEAKS.xlsx")
    dv = wb["LEAKS"].data_validations.dataValidation[0]
    assert "**" in dv.prompt and len(dv.prompt) <= P._PN_XL_DV_TEXT_MAX


def test_an_undecided_worksheet_row_does_not_shadow_the_inherited_fix():
    """A folder run BEFORE the `**` reached the master sheet carries the
    garble as an undecided row; on the re-run the inherited fix answers it."""
    master = _decisions(("SanDiega", "**San Diego"))
    folder = P._pn_parse_decision_rows(
        [HDR, ("SanDiega", "", "", "", "", ""), ("Other", "no", "", "", "", "")])
    merged = P._pn_layer_decisions(master, folder)
    assert merged["sandiega"]["ocr_fix"] == "San Diego"
    assert merged["other"]["fix"] == "no"
    # …while a row the operator DID answer still wins locally.
    folder2 = P._pn_parse_decision_rows([HDR, ("SanDiega", "no", "", "", "", "")])
    assert P._pn_layer_decisions(master, folder2)["sandiega"]["fix"] == "no"
