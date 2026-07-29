"""
pdf_linker.py - Add citation hyperlinks to PDFs in a case folder, build
                navigation bookmarks, and export a plain-text companion for
                each text-based PDF.

Usage:
    pythonw pdf_linker.py "C:\\path\\to\\case\\folder"
    pythonw pdf_linker.py "C:\\path\\to\\case\\folder" --provider westlaw

Pseudonymization can be turned on/off without editing the code: set
`pseudonymize = on` (or `off`) in the `pdf_linker.config` file next to this
script (auto-created on first run). A --pseudonymize / --no-pseudonymize flag
overrides the config for that run.

For each *.pdf in the folder (processed from shortest to longest by file size):
  1. If the PDF has no text layer, runs Tesseract OCR via OCRmyPDF-style flow
     using pytesseract page-by-page to add a text layer. Pages whose existing
     text layer looks obviously garbled (broken font encoding, unmapped
     glyphs) are rebuilt the same way: the bad text is removed and a fresh OCR
     layer is laid down over the (correctly-rendering) page.
  1b. Pseudonymizes the .txt export (only) — ON by default (--no-pseudonymize
     to disable): party/case/attorney names, a declaration's own declarant
     name ("DECLARATION OF X" / "I, X, declare", even when not in the key),
     and detected PII are swapped for stable, deterministic fakes, and a real->fake key file is written for the
     folder. Party/case names come from a spreadsheet key — --key, or by
     default the most recent Order*.xlsx in the user's Downloads folder (the
     E-Court order-template export) — plus any --term values. The PDFs are
     never modified. The .txt FILENAME is pseudonymized too (a party/attorney
     name in the source PDF's name is scrubbed from the .txt's name, falling
     back to a neutral name if anything real would survive). A surviving real
     value (a per-file leak, or a spreadsheet key that matched nothing) is
     reported in the log for review — the .txt is still written, not withheld.
     A mistake baked into the pseudonym_key.xlsx (a value it should not have
     faked) can be corrected in place in the key's Replacement column, using the
     same words the LEAKS worksheet's Fix? column accepts: "no" leaves that Real
     Value verbatim, and a [bracketed] keep-spec keeps the bracketed part and
     fakes the rest. These no/bracket KEEP decisions are recorded to a SINGLE
     cross-folder sheet — the KEEP tab of master_leaks.xlsx — so they are applied
     on every future run in EVERY folder (accumulating Times Seen / Cases /
     dates), letting the screening learn from real history. A KEEP matches on
     WORD BOUNDARIES and loses to a full party name: "no" on "Cal" keeps the
     standalone word yet still fakes "CAL EQUIPMENT FE RANCH, LLC" whole and is
     released inside a detected name run like "Cal Equipment", while a
     [bracketed] keep ("this fragment is never a name") stays even beside a name
     ("[Plaintiff]" in "Plaintiff John Doe"). A keep never leaves a party in the
     clear.
  1a. Writes a plain-text companion (.txt) for each PDF that has a real text
     layer into a "Text Files" subfolder of the case folder (name overridable
     via text_subfolder in pdf_linker.config), so a text-only copy can be
     uploaded instead of page images. On pleading-paper pages each line is
     prefixed with its printed
     line number (and the page header carries the printed page number), so a
     pinpoint "p.X:Y" cite lands on the same text as in the PDF. Extracts whenever the document opens with native text (even if
     scanned exhibits follow); if the head isn't native it falls back to a
     random sample of the rest. Scanned-only PDFs are skipped. Decided from
     the native text, before OCR. Disable with --no-text. The .txt ends with
     an "Authorities cited" appendix linking each citation to a FREE public
     database (California leginfo / Cornell LII / Google Scholar) instead of
     the paid Lexis/Westlaw link the PDF uses, for no-login verification.
  2. For files whose name contains "Declaration", "Decl.", "Separate Statement",
     "Compl.", "Complaint", "FAC", "SAC", "TAC", or "Proof of Service", link
     insertion is skipped — these documents rarely have citation-worthy material
     in the working judge's chambers. Bookmark/structure passes still run on
     these files.
  3. Extracts case, statute, and rule citations using the same patterns the
     workup search tool uses (extended to recognise CSM, Bluebook, the
     "flat" no-comma practitioner form, and Westlaw-only WL cites).
  4. Tracks first-seen full case citations and resolves later "supra" cites
     within the same PDF; also runs a second pass that links bare short-form
     "X v. Y" references (no reporter or year) to the URL of the matching
     full citation.
  5. Resolves each citation to a search URL on the active provider
     (--provider lexis or westlaw). Westlaw-only WL cites (e.g.
     "2023 WL 3035369") always resolve to a Westlaw URL even when the
     active provider is Lexis — Lexis doesn't carry those reporters.
  6. Adds a blue-and-underlined link annotation over each citation. Multi-
     line citations are matched safely: if the full text doesn't appear on
     a single line, line-fragments are used only when they are distinctive
     enough (no bare "Cal." or "Cal. App. 4th" links across the document).
  7. Replaces the original PDF with the linked version (original is deleted).
  8. Builds a PDF outline (sidebar bookmarks) from the detected structure —
     Contents (TOC), Exhibits, Paragraphs, Causes of Action, and combined-
     filing sub-documents.

If the folder holds Word documents (.docx/.docm) and NO PDFs, each is instead
bulk-converted to a scrubbed .txt export exactly as a PDF is (same
pseudonymization, leak report and gate, name-scrubbed filename), but Word docs
are never hyperlinked. When any PDF is present the Word docs are left untouched,
so the usual PDF workflow is unaffected. The legacy binary .doc format can't be
read (re-save it as .docx).

The Westlaw and Lexis search prefix tables, statute name variants, and URL
forms here are kept in sync with the cross-opener extension's content.js
and workups.html, which are validated against live Westlaw and Lexis+
pages. Those files are authoritative whenever they disagree with this one.

Logs to <folder>/pdf_linker.log.
"""

import datetime
import logging
import os
import re
import statistics
import sys
import time
import traceback
from pathlib import Path

# ────────────────────────────────────────────────────────────────────────────
# Configuration
# ────────────────────────────────────────────────────────────────────────────
SHAREPOINT_BASE = (
    Path(os.environ.get("USERPROFILE", "C:/Users/ZCoderre"))
    / "Los Angeles Superior Court"
    / "Research Attorney and Law Clerk Unit - Zachary Coderre"
)

# Standard Tesseract install paths to try if not on PATH
TESSERACT_CANDIDATES = [
    Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tesseract.exe",
    Path(os.environ.get("LOCALAPPDATA", "")) / "Tesseract-OCR" / "tesseract.exe",
    Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
    Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
    Path(os.environ.get("USERPROFILE", "")) / "AppData" / "Local" / "Tesseract-OCR" / "tesseract.exe",
]

# Visual link styling
LINK_COLOUR = (0.0, 0.0, 0.85)  # blue (RGB 0..1)

# ────────────────────────────────────────────────────────────────────────────
# Code conversion (matches workups.html / cross-opener / harvester)
# ────────────────────────────────────────────────────────────────────────────
WL_SEARCH_PREFIX = {
    "BPC": "CA BUS & PROF", "COM": "CA COML", "CIV": "CA CIVIL",
    "CCP": "CA CIV PRO", "CORP": "CA CORP", "EDC": "CA EDUC",
    "ELEC": "CA ELEC", "EVID": "CA EVID", "FAM": "CA FAM",
    "FIN": "CA FIN", "FGC": "CA FISH & G", "FAC": "CA FOOD & AG",
    "GOV": "CA GOVT", "HNC": "CA HARB & NAV", "HSC": "CA HLTH & S",
    "INS": "CA INS", "LAB": "CA LABOR", "MVC": "CA MIL & VET",
    "PEN": "CA PENAL", "PROB": "CA PROBATE", "PCC": "CA PUB CONT",
    "PRC": "CA PUB RES", "PUC": "CA PUB UTIL", "RTC": "CA REV & TAX",
    "SHC": "CA STR & HWY", "UIC": "CA UNEMP INS", "VEH": "CA VEHICLE",
    "WAT": "CA WATER", "WIC": "CA WELF & INST",
}

LEXIS_SEARCH_PREFIX = {
    "BPC": "Cal Bus & Prof Code", "COM": "Cal U Com Code", "CIV": "Cal Civ Code",
    "CCP": "Cal Code Civ Proc", "CORP": "Cal Corp Code", "EDC": "Cal Ed Code",
    "ELEC": "Cal Elec Code", "EVID": "Cal Evid Code", "FAM": "Cal Fam Code",
    "FIN": "Cal Fin Code", "FGC": "Cal Fish & G Code", "FAC": "Cal Food & Agr Code",
    "GOV": "Cal Gov Code", "HNC": "Cal Harb & Nav Code", "HSC": "Cal Health & Saf Code",
    "INS": "Cal Ins Code", "LAB": "Cal Lab Code", "MVC": "Cal Mil & Vet Code",
    "PEN": "Cal Pen Code", "PROB": "Cal Prob Code", "PCC": "Cal Pub Contract Code",
    "PRC": "Cal Pub Resources Code", "PUC": "Cal Pub Util Code", "RTC": "Cal Rev & Tax Code",
    "SHC": "Cal Sts & Hy Code", "UIC": "Cal Unemp Ins Code", "VEH": "Cal Veh Code",
    "WAT": "Cal Wat Code", "WIC": "Cal Welf & Inst Code",
}

# ────────────────────────────────────────────────────────────────────────────
# Reporter list - both CSM (compact) and Bluebook (spaced) forms
# ────────────────────────────────────────────────────────────────────────────
_REPORTERS_RAW = [
    # California
    "Cal.5th", "Cal. 5th", "Cal.4th", "Cal. 4th", "Cal.3d", "Cal. 3d",
    "Cal.2d", "Cal. 2d", "Cal.",
    "Cal.App.5th Supp.", "Cal. App. 5th Supp.",
    "Cal.App.4th Supp.", "Cal. App. 4th Supp.",
    "Cal.App.3d Supp.", "Cal. App. 3d Supp.",
    "Cal.App.2d Supp.", "Cal. App. 2d Supp.",
    "Cal.App.5th", "Cal. App. 5th", "Cal.App.4th", "Cal. App. 4th",
    "Cal.App.3d", "Cal. App. 3d", "Cal.App.2d", "Cal. App. 2d",
    "Cal.App.", "Cal. App.",
    "Cal.Rptr.3d", "Cal. Rptr. 3d", "Cal.Rptr.2d", "Cal. Rptr. 2d",
    "Cal.Rptr.", "Cal. Rptr.",
    # California Style Manual ultra-compact forms, ubiquitous in practitioner
    # briefs and CEB/Witkin (e.g. "39 C3d 311", "116 CA4th 968", "216 CR 718",
    # "11 CR3d 45"). C=Cal., CA=Cal.App., CR=Cal.Rptr. These are normalised to
    # the canonical reporter for keys and search URLs (see _normalize_reporter)
    # so a cite written only in compact form still resolves correctly.
    "C5th", "C4th", "C3d", "C2d",
    "CA5th", "CA4th", "CA3d", "CA2d",
    "CR3d", "CR2d", "CR",
    # Federal
    "U.S.", "S.Ct.", "S. Ct.", "L.Ed.2d", "L. Ed. 2d", "L.Ed.", "L. Ed.",
    "F.4th", "F. 4th", "F.3d", "F. 3d", "F.2d", "F. 2d", "F.",
    "F.Supp.3d", "F. Supp. 3d", "F.Supp.2d", "F. Supp. 2d", "F.Supp.", "F. Supp.",
    # F. App'x (Federal Appendix) — both straight apostrophe and curly U+2019.
    # PDF text extraction commonly produces the curly form; we accept either.
    "F. App'x", "F. App\u2019x", "F.App'x", "F.App\u2019x",
    # Common out-of-state
    "N.Y.2d", "N.Y. 2d", "N.Y.3d", "N.Y. 3d",
    "P.3d", "P. 3d", "P.2d", "P. 2d", "P.",
    "A.3d", "A. 3d", "A.2d", "A. 2d",
    "N.E.3d", "N.E. 3d", "N.E.2d", "N.E. 2d",
    "N.W.2d", "N.W. 2d",
    "S.E.2d", "S.E. 2d",
    "S.W.3d", "S.W. 3d", "S.W.2d", "S.W. 2d",
    "So.3d", "So. 3d", "So.2d", "So. 2d",
]
REPORTERS_SORTED = sorted(_REPORTERS_RAW, key=len, reverse=True)


def _flex_ws_reporter(reporter: str) -> str:
    # Escape the reporter, then let any internal whitespace match one-or-more
    # whitespace chars (spaces, tabs, or a normalized line-wrap). PyMuPDF emits
    # a trailing space before a wrapped line ("Cal. App. \n4th"), and
    # _normalize_for_detection turns the newline into a second space, yielding
    # a DOUBLE space inside the reporter ("Cal. App.  4th"). A literal-single-
    # space pattern silently misses those cites (e.g. a reporter split as
    # "134 Cal. App." / "4th 365"). \s+ absorbs the extra whitespace.
    return re.sub(r"\\?\s+", r"\\s+", re.escape(reporter))


REPORTER_PATTERN = "|".join(_flex_ws_reporter(r) for r in REPORTERS_SORTED)

# California Style Manual compact reporter abbreviations → canonical compact
# form (matching the rest of pdf_linker's keys). Keyed by the whitespace-
# stripped matched text so both "CA4th" and a stray-spaced "CA 4th" normalise.
_REPORTER_NORMALIZE = {
    "C5th": "Cal.5th", "C4th": "Cal.4th", "C3d": "Cal.3d", "C2d": "Cal.2d",
    "CA5th": "Cal.App.5th", "CA4th": "Cal.App.4th",
    "CA3d": "Cal.App.3d", "CA2d": "Cal.App.2d",
    "CR3d": "Cal.Rptr.3d", "CR2d": "Cal.Rptr.2d", "CR": "Cal.Rptr.",
}


def _normalize_reporter(reporter: str) -> str:
    """Strip internal whitespace from a matched reporter and map California
    Style Manual compact forms (C3d, CA4th, CR3d, …) to the canonical
    reporter so citation keys and provider search URLs are consistent
    regardless of which form the brief used. The appellate-department
    "Supp." suffix keeps its separating space — "Cal.App.4th Supp." glued
    to "Cal.App.4thSupp." is a citation form no provider recognizes."""
    compact = re.sub(r"\s+", "", reporter)
    compact = _REPORTER_NORMALIZE.get(compact, compact)
    return re.sub(r"(?<=[^\s])Supp\.", " Supp.", compact)

# ────────────────────────────────────────────────────────────────────────────
# Statute code recognition
# ────────────────────────────────────────────────────────────────────────────
STATUTE_CODES = [
    # Long forms. Internal spaces are `\s+` (not literal " ") so the
    # patterns survive line-wrapped citations like "Code of Civil
    # \nProcedure section 430.10(e)" — common in body text where the
    # code name straddles a line break. A literal-space pattern would
    # silently fail to match such wraps, leaving the citation unlinked.
    (r"Code\s+of\s+Civil\s+Procedure", "CCP"),
    (r"Civil\s+Code", "CIV"),
    (r"Penal\s+Code", "PEN"),
    (r"Evidence\s+Code", "EVID"),
    (r"Business\s+(?:and|&)\s+Professions\s+Code", "BPC"),
    (r"Family\s+Code", "FAM"),
    (r"Government\s+Code", "GOV"),
    (r"Health\s+(?:and|&)\s+Safety\s+Code", "HSC"),
    (r"Labor\s+Code", "LAB"),
    (r"Probate\s+Code", "PROB"),
    (r"Vehicle\s+Code", "VEH"),
    (r"Welfare\s+(?:and|&)\s+Institutions\s+Code", "WIC"),
    (r"Corporations\s+Code", "CORP"),
    (r"Insurance\s+Code", "INS"),
    (r"Revenue\s+(?:and|&)\s+Taxation\s+Code", "RTC"),
    (r"Education\s+Code", "EDC"),
    (r"Elections\s+Code", "ELEC"),
    (r"Financial\s+Code", "FIN"),
    (r"Fish\s+(?:and|&)\s+Game\s+Code", "FGC"),
    (r"Food\s+(?:and|&)\s+Agricultural\s+Code", "FAC"),
    (r"Harbors\s+(?:and|&)\s+Navigation\s+Code", "HNC"),
    (r"Military\s+(?:and|&)\s+Veterans\s+Code", "MVC"),
    (r"Public\s+Contract\s+Code", "PCC"),
    (r"Public\s+Resources\s+Code", "PRC"),
    (r"Public\s+Utilities\s+Code", "PUC"),
    (r"Streets\s+(?:and|&)\s+Highways\s+Code", "SHC"),
    (r"Unemployment\s+Insurance\s+Code", "UIC"),
    (r"Water\s+Code", "WAT"),
    (r"Commercial\s+Code", "COM"),

    # CSM short forms
    (r"Code Civ\.\s*Proc\.", "CCP"),
    (r"Civ\.\s*Code", "CIV"),
    (r"Pen\.\s*Code", "PEN"),
    (r"Evid\.\s*Code", "EVID"),
    (r"Bus\.\s*(?:&|and)\s*Prof\.\s*Code", "BPC"),
    (r"Fam\.\s*Code", "FAM"),
    (r"Gov\.\s*Code", "GOV"),
    (r"Health\s*(?:&|and)\s*Saf\.\s*Code", "HSC"),
    (r"Lab\.\s*Code", "LAB"),
    (r"Prob\.\s*Code", "PROB"),
    (r"Veh\.\s*Code", "VEH"),
    (r"Welf\.\s*(?:&|and)\s*Inst\.\s*Code", "WIC"),
    (r"Corp\.\s*Code", "CORP"),
    (r"Ins\.\s*Code", "INS"),
    (r"Rev\.\s*(?:&|and)\s*Tax\.\s*Code", "RTC"),
    (r"Educ\.\s*Code", "EDC"),
    (r"Elec\.\s*Code", "ELEC"),
    (r"Fin\.\s*Code", "FIN"),
    (r"Fish\s*(?:&|and)\s*Game Code", "FGC"),
    (r"Food\s*(?:&|and)\s*Agric\.\s*Code", "FAC"),
    (r"Harb\.\s*(?:&|and)\s*Nav\.\s*Code", "HNC"),
    (r"Mil\.\s*(?:&|and)\s*Vet\.\s*Code", "MVC"),
    (r"Pub\.\s*Cont(?:ract)?\.?\s*Code", "PCC"),
    (r"Pub\.\s*Res(?:ources)?\.?\s*Code", "PRC"),
    (r"Pub\.\s*Util(?:ities)?\.?\s*Code", "PUC"),
    (r"Sts\.\s*(?:&|and)\s*Hy\.\s*Code", "SHC"),
    (r"Unemp\.\s*Ins\.\s*Code", "UIC"),
    (r"Wat\.\s*Code", "WAT"),
    (r"Com\.\s*Code", "COM"),

    # Extra variants validated by the cross-opener extension's content.js
    (r"Govt\.\s*Code", "GOV"),
    (r"Fish\s*(?:&|and)\s*G\.\s*Code", "FGC"),
    (r"Food\s*(?:&|and)\s*Agr\.\s*Code", "FAC"),

    # Practitioner-style "X Code" reorderings of multi-word codes. The CSM
    # short forms above use "Code Civ. Proc." (Code first), but California
    # briefs frequently write "Cal. Civ. Proc. Code" (Code last). Add the
    # reversed orderings so both word orders match.
    (r"Civ\.\s*Proc\.\s*Code", "CCP"),
    (r"Civil\s+Procedure\s+Code", "CCP"),

    # Bare short form without "Code". Some briefs write "Civ. Proc.
    # § 430.10(e)" with no preceding "Code". STATUTE_RE requires a "§"
    # (or "section") + section number right after the code name, so a
    # standalone "Civ. Proc." in body prose can't accidentally trigger
    # this — only the citation form will. Listed last so the longer
    # "Code Civ. Proc." and "Civ. Proc. Code" patterns win when present
    # (alternation tries patterns in length-descending order).
    (r"Civ\.\s*Proc\.", "CCP"),

    # Bare initialism. "CCP" for Code of Civil Procedure is ubiquitous in
    # California unlawful-detainer and law-and-motion practice ("CCP section
    # 1170", "CCP § 1005"). Dots and spacing are optional so "C.C.P." and
    # "C. C. P." also match. Like the other bare forms it only fires in
    # citation position — a following section number is required.
    # (?<![A-Za-z]\.) keeps the "C.C.P." inside "J.C.C.P. 5237" (a Judicial
    # Council coordination number) from reading as a CCP statute cite.
    (r"(?<![A-Za-z]\.)C\.?\s*C\.?\s*P\.?", "CCP"),
]
STATUTE_CODES_SORTED = sorted(STATUTE_CODES, key=lambda x: len(x[0]), reverse=True)


def _build_statute_re():
    parts = [f"(?P<c{i}>{pat})" for i, (pat, _) in enumerate(STATUTE_CODES_SORTED)]
    code_alt = "|".join(parts)
    # Section number: digits, optional .digits, optional letter suffix (e.g.
    # "437c"), optional subsection like (b)(1)
    full = (
        r"\b(?:Cal\.\s*|California\s+)?"
        rf"(?:{code_alt})"
        r",?\s*"
        # Section marker (§, "section", or "sec.") is OPTIONAL: practitioners
        # routinely drop it, writing "Code of Civil Procedure 430.30(a)" or
        # "Penal Code 187". We capture whether it was present in `mk` so the
        # caller can apply a stricter section shape when it's absent (a bare
        # code name followed by a list counter like "2." at a paragraph break
        # must not be mistaken for a citation — see find_statute_citations).
        r"(?:(?P<mk>§§?|sections?|secs?\.?)\s*)?"
        r"(?P<sec>\d+(?:\.\d+)?[a-z]?(?:\([a-z0-9]+\))*)"
    )
    # IGNORECASE so all-caps practitioner forms like "CAL. CIV. PROC. CODE
    # § 1281.2" match alongside the conventional title-case forms. False
    # positives are held down by requiring the full code name AND, when no
    # section marker is present, a section number distinctive enough not to
    # be a list/paragraph counter (enforced in find_statute_citations).
    return re.compile(full, re.DOTALL | re.IGNORECASE)


STATUTE_RE = _build_statute_re()


# Federal statutes: "9 U.S.C. § 1", "42 U.S.C. § 1983", etc. These have a
# title number preceding the code abbreviation rather than the optional
# "Cal."/"California" prefix that STATUTE_RE handles. We detect them with
# a parallel regex and emit keys in their natural form: "9 U.S.C. § 1".
# URL building (see _build_usc_term) reads the title and section back out.
USC_RE = re.compile(
    # Word-boundary, then title number (1-3 digits), space-separated
    # "U.S.C." (with optional intervening spaces between letters as PDFs
    # sometimes render it), then section marker and section identifier.
    r"\b(?P<title>\d{1,3})\s+U\.\s*S\.\s*C\."
    r"(?:\s*App\.)?"                              # optional ", App."
    r"\s*"
    r"(?:§§?|sections?|secs?\.?)\s*"
    r"(?P<sec>\d+(?:\.\d+)?[a-z]?(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)


def _statute_abbrev(match):
    for i, (_, abbrev) in enumerate(STATUTE_CODES_SORTED):
        if match.group(f"c{i}"):
            return abbrev
    return None


# After a primary statute citation matches, additional sections may follow
# in chained form like:
#   "Code of Civil Procedure sections 598 and 1048(b)"
#   "Civ. Code §§ 1542, 1543, and 1544"
#   "Pen. Code §§ 187, 189"
# This pattern matches a single continuation: a connector (",", " and ",
# ", and ") followed by another section number. The caller applies it
# repeatedly starting at the end of the previous match to extract every
# section in the chain. The continuation must immediately follow the
# previous match (anchored via re.match at scan_pos) — any intervening
# non-connector text breaks the chain.
ADDL_SEC_RE = re.compile(
    r"\s*(?:,\s*and|,|\s+and)\s+"
    r"(?P<sec>\d+(?:\.\d+)?[a-z]?(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)


# ────────────────────────────────────────────────────────────────────────────
# Bare ("naked") section citations
# ────────────────────────────────────────────────────────────────────────────
# A section number with a marker (§ / "section" / "sec.") but NO code name —
# "Section 1150", "§ 1177", "Section 1150 through 1179a". These are linked to
# the code that is unambiguously referenced *around* them: if exactly one
# California code is cited or named within _BARE_CONTEXT_WINDOW characters on
# either side, that code is assumed (the common brief pattern where a writer
# names "Code of Civil Procedure"/"CCP" once and then refers to bare sections
# of it). When two different codes appear nearby (ambiguous), or none does,
# the bare section is left unlinked. A distinctive section shape is required
# so prose like "section 5 of the lease" or "Part 2" is not picked up.
BARE_SECTION_RE = re.compile(
    r"(?:§§?|sections?|secs?\.?)\s*"
    r"(?P<sec>\d+(?:\.\d+)?[a-z]?(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)

# Continuation of a bare-section chain or range that inherits the inferred
# code: "Section 1150 through 1179a", "§§ 1150 to 1179a", "Section 1150,
# 1177, and 1179a". Same numeric shape as a primary section.
BARE_CHAIN_RE = re.compile(
    r"\s*(?:through|thru|to|[-‒–—−]|,\s*and|,|\s+and)\s+"
    r"(?P<sec>\d+(?:\.\d+)?[a-z]?(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)

_BARE_CONTEXT_WINDOW = 1200  # characters on each side of a bare section


def _build_code_name_re():
    parts = [f"(?P<c{i}>{pat})" for i, (pat, _) in enumerate(STATUTE_CODES_SORTED)]
    return re.compile(
        r"\b(?:Cal\.\s*|California\s+)?(?:" + "|".join(parts) + r")",
        re.DOTALL | re.IGNORECASE,
    )


# Matches any California code NAME (with or without a trailing section), used
# only to locate code mentions for bare-section context inference.
CODE_NAME_RE = _build_code_name_re()


def _code_context_positions(text):
    """Return [(center_index, abbrev)] for every California code name mentioned
    in `text`, whether or not a section number follows."""
    out = []
    for m in CODE_NAME_RE.finditer(text):
        for i, (_, abbrev) in enumerate(STATUTE_CODES_SORTED):
            if m.group(f"c{i}"):
                out.append(((m.start() + m.end()) // 2, abbrev))
                break
    return out


def _section_is_distinctive(sec: str) -> bool:
    """True if `sec` looks like a real statute section rather than a list or
    paragraph counter: it has a decimal part, a letter suffix, a subdivision,
    or is at least three digits."""
    if re.search(r"\.\d|[a-z]|\(", sec, re.IGNORECASE):
        return True
    return len(re.sub(r"\D", "", sec)) >= 3


def find_bare_section_citations(text, occupied_spans, code_positions):
    """Link bare section citations to the single code referenced nearby.

    `occupied_spans` are (start, end) ranges already claimed by code-named
    statute citations, so a section that is part of one is not re-matched.
    `code_positions` comes from `_code_context_positions`.
    """
    def _overlaps(a, b):
        return a[0] < b[1] and b[0] < a[1]

    def _infer(pos):
        near = {ab for (cpos, ab) in code_positions
                if abs(cpos - pos) <= _BARE_CONTEXT_WINDOW}
        return next(iter(near)) if len(near) == 1 else None

    results = []
    for m in BARE_SECTION_RE.finditer(text):
        if any(_overlaps(m.span(), sp) for sp in occupied_spans):
            continue
        if not _section_is_distinctive(m.group("sec")):
            continue
        abbrev = _infer((m.start() + m.end()) // 2)
        if not abbrev:
            continue
        results.append({
            "kind": "statute",
            "key": f"{abbrev} § {m.group('sec')}",
            "span": m.span(),
            "match_text": m.group(0),
        })
        # Chain/range continuations ("through 1179a", ", 1177", "and 1179a")
        # inherit the same inferred code.
        scan = m.end()
        while True:
            cont = BARE_CHAIN_RE.match(text, scan)
            if not cont:
                break
            if _section_is_distinctive(cont.group("sec")):
                results.append({
                    "kind": "statute",
                    "key": f"{abbrev} § {cont.group('sec')}",
                    "span": cont.span(),
                    "match_text": text[cont.start():cont.end()].lstrip(),
                })
            scan = cont.end()
    return results


# ────────────────────────────────────────────────────────────────────────────
# Rule pattern
# ────────────────────────────────────────────────────────────────────────────
RULE_RE = re.compile(
    r"\b(?:Cal\.\s*Rules?\s*of\s*Court|California\s*Rules?\s*of\s*Court),?\s*"
    r"rules?\s+(\d+(?:\.\d+)*(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)

# Rules of Professional Conduct.
RPC_RE = re.compile(
    r"\b(?:Cal(?:ifornia)?\.?\s+)?Rules?\s+of\s+(?:Prof(?:essional)?\.?\s+)?Conduct\s+"
    r"(\d+(?:\.\d+)*(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)

# ────────────────────────────────────────────────────────────────────────────
# Case patterns - v.-anchor approach for accuracy across CSM/Bluebook
# ────────────────────────────────────────────────────────────────────────────
ANCHOR_RE = re.compile(r"(?<=\s)v\.(?=\s)")
# vol REPORTER page (with optional pin or pin range like 247-48)
REPORTER_PART = (
    rf"(\d{{1,4}})\s+({REPORTER_PATTERN})\s+(\d{{1,5}})"
    # Optional pin or pin-range. The dash character class accepts all common
    # PDF-extracted dashes: ASCII hyphen-minus, en dash (U+2013), em dash
    # (U+2014), figure dash (U+2012), and minus sign (U+2212). PDFs from
    # different layout tools render the same Bluebook pin range with
    # different code points, and missing one breaks otherwise-valid cites
    # (this caused "Santana v. FCA US, LLC, 56 Cal.App.5th 324, 345‒46
    # (2020)" to be undetected — the figure dash U+2012 wasn't in the class).
    # Repeated (`*`, not `?`) so a cite with more than one pin page still
    # reaches the "(year)" close: "Bowser v. Ford Motor Co., 78 Cal.App.5th
    # 587, 599, 617 (2022)" pins to two pages, and a single-pin tail stopped
    # before the year \u2014 leaving the whole span undetected and the party token
    # inside it ("Ford Motor Co.") exposed to rewriting.
    rf"(?:,\s*\d{{1,5}}(?:[-\u2012\u2013\u2014\u2212]\d{{1,5}})?)*"
)
CSM_TAIL = re.compile(rf"\s*\((?:[^)]*?\b)?(\d{{4}})\)\s+{REPORTER_PART}")
# Bluebook tail: ", REPORTER (court year)" — allow optional court abbreviation
# in parens like (9th Cir. 2015), (S.D.N.Y. 2009), (Cal. 2004), or just (2001).
# The non-greedy [^)]*? before the year matches optional court text without
# backtracking issues.
#
# An optional comma is allowed AFTER the pin and before "(year)" — some
# briefs write "108 Cal. App. 4th 773, 780, (2003)" with a trailing comma
# after the pin. The trailing comma adds no information but is a real
# practitioner habit; we accept it rather than miss the cite.
BB_TAIL = re.compile(
    rf",\s+{REPORTER_PART}\s*,?\s*\((?:[^)]*?\b)?(\d{{4}})\)"
)

# Flat tail: " REPORTER (court year)" — same as Bluebook but WITHOUT the comma
# between the defendant and the volume. Common in California practitioner
# briefs and in tables of authorities:
#   "Donlen v. Ford Motor Co. 217 Cal. App. 4th 138 (2013)"
#   "Rattagan v. Uber Technologies, Inc. 17 Cal. 5th 1 (2024)"
#   "LiMandri v. Judkins 52 Cal.App.4th 326 (1997)"
# Required: a space (not comma) before the volume. Group order matches BB_TAIL
# so downstream code in find_case_citations can treat them identically.
# Some defendant names end with a period (e.g. "Inc.", "Corp.", "Am.", "Co.")
# — we accept that, otherwise we'd require the period to be optional and
# end up matching mid-sentence noise. The walk-back already validates that
# what precedes "v." is a real party name.
FLAT_TAIL = re.compile(
    rf"\s+{REPORTER_PART}\s*,?\s*\((?:[^)]*?\b)?(\d{{4}})\)"
)

# Westlaw-only citation tail: ", YYYY WL NNNNNN, at *N (court date)"
# (or without the pin "at *N"). Used for unpublished decisions that exist
# only on Westlaw; the year inside the parens may include a court abbreviation
# and a date like (C.D. Cal. Nov. 2, 2021).
# Group layout: (1) year-of-cite (in WL key), (2) WL number, (3) decision year
WL_TAIL = re.compile(
    r",\s+(\d{4})\s+WL\s+(\d{4,8})"          # YYYY WL NNNNNN
    # Optional ", at *N" pin, optionally followed by " n.M" footnote pin
    # ("at *5 n.7" is common in S.D.N.Y. citations).
    r"(?:,\s*at\s+\*?\d+(?:\s+n\.\d+)?)?"
    r"\s*\((?:[^)]*?\b)?(\d{4})\)"            # (court date)
)

# Lexis-only citation tail: ", YYYY U.S. Dist. LEXIS NNNNN (court date)".
# Structural parallel to WL_TAIL — both encode an online-database number
# that isn't a real page in any printed reporter, so the trailing number
# needs broader tolerance than the \d{1,5} page constraint used in the
# CSM/BB/FLAT reporter patterns. Same optional pin/footnote shape as WL.
# Group layout: (1) year-of-cite, (2) LEXIS number, (3) decision year.
# URL routing: cites matched here get lexis_only=True (mirrors wl_only),
# forcing resolution through Lexis regardless of the active provider —
# Westlaw doesn't carry LEXIS database numbers as a lookup key.
LEXIS_TAIL = re.compile(
    r",\s+(\d{4})\s+U\.S\.\s*Dist\.\s*LEXIS\s+(\d{4,8})"
    r"(?:,\s*at\s+\*?\d+(?:\s+n\.\d+)?)?"
    r"\s*\((?:[^)]*?\b)?(\d{4})\)"
)

# Slip-cite tail: "[, ]?Case No. <docket-id> (<court> [date])". These are
# decisions that haven't been published in a reporter and don't have a
# WL/LEXIS number assigned yet — the brief identifies them by their
# trial-court docket number and the court parenthetical. Because there's
# no reporter cite to look up, URL resolution falls back to a case-name
# search (handled in resolve_url via the slip_only branch).
#
# Match requirements:
#   * Comma is optional (the spreadsheet has "Lee v. Creditco Info. Sols.,
#     Inc. Case No. 30STCV12347" with no comma after "Inc.").
#   * Docket-id is letters, digits, and common docket punctuation
#     (BCV-30-123456, 30STCV12347, 1:16-cv-12653-ADB, 19-cv-01080).
#   * The court parenthetical may or may not contain a date — we don't
#     require a year because slip cites without dates do exist ("(Sup. Ct.
#     Cal. Los Angeles Cnty.)"). The parenthetical itself is required.
# Group layout: (1) docket id (for diagnostics), (2) full parenthetical
# contents (court + optional date), (3) year if present, else empty.
SLIP_TAIL = re.compile(
    r",?\s+Case\s+No\.\s+([A-Z0-9][A-Z0-9:\-]{3,30})"
    r"\s*\(([^)]{3,80})\)",
    re.IGNORECASE,
)

# Probate, conservatorship, and family-law cases use "in the matter of"
# conventions instead of "X v. Y": "In re Marriage of Smith", "Estate of
# Bowles", "Guardianship of Doe", "Conservatorship of Roe", "Adoption of
# Jones". All share the same citation-tail structure as In re, so we use
# one regex over a prefix alternation. The prefix isn't captured as its
# own group — that would shift the numeric group indices the consumer
# code relies on — so the consumer re-extracts it from `match.group(0)`
# via _NONV_PREFIX_RE below.
_NONV_PREFIX = (
    # "In re" in any casing an OCR layer produces ("In Re", "IN RE").
    r"In [Rr][Ee]|IN RE|"
    r"Estate of|"
    r"Guardianship of|"
    r"Conservatorship of|"
    r"Adoption of|"
    r"Marriage of"
)

# Used by the consumer to recover which prefix matched, since the prefix
# itself isn't a numbered capture group. Anchored at start-of-match.
_NONV_PREFIX_RE = re.compile(rf"^\s*({_NONV_PREFIX})\b")

# In re / Estate of / Guardianship of / etc. cases: support CSM "(year)
# vol REPORTER page", Bluebook ", vol REPORTER page (court year)", and
# flat " vol REPORTER page (court year)" (no-comma) forms. The
# Bluebook/flat forms also accept court abbreviations in parens (e.g.
# "In re Doe, 555 F.3d 100 (9th Cir. 2009)").
INRE_RE = re.compile(
    rf"\b(?:{_NONV_PREFIX})\s+([A-Z][A-Za-z0-9.\-'&, ]+?)\s*"
    rf"(?:"
    rf"\((\d{{4}})\)\s+(\d{{1,4}})\s+({REPORTER_PATTERN})\s+(\d{{1,5}})"
    rf"|"
    rf",?\s+(\d{{1,4}})\s+({REPORTER_PATTERN})\s+(\d{{1,5}})"
    rf"(?:,\s*\d{{1,5}}(?:[–\-]\d{{1,5}})?)?\s*"
    rf"\((?:[^)]*?\b)?(\d{{4}})\)"
    rf"|"
    # WL alternative: "In re X, [docket text], YYYY WL NNNNNN (court year)".
    # Federal district court WL cites commonly carry docket info between the
    # case name and the WL number ("In re Intuniv Antitrust Litig., Civil
    # Action No. 1:16-cv-12653-ADB, 2021 WL 517386 (D. Mass. Feb. 11, 2021)").
    # We allow up to ~80 characters of non-newline filler before the WL
    # number. The filler may include colons (docket numbers) which aren't
    # part of the case-name character class.
    rf"[,\s][^\n]{{0,80}}?,\s+(\d{{4}})\s+WL\s+(\d{{4,8}})"
    rf"(?:,\s*at\s+\*?\d+(?:\s+n\.\d+)?)?"
    rf"\s*\((?:[^)]*?\b)?(\d{{4}})\)"
    rf")"
)

SUPRA_RE = re.compile(
    rf"\b((?:(?:{_NONV_PREFIX})\s+)?[A-Z][A-Za-z0-9.\-'&]+(?:\s+v\.\s+[A-Z][A-Za-z0-9.\-'&]+)?)"
    r",\s*supra\b"
)

# Consolidated-litigation cases ending with the literal word "Cases", with no
# "v." or "In re": "Ford Motor Warranty Cases (2025) 17 Cal.5th 1122",
# "Gilead Tenofovir Cases (2024) 98 Cal.App.5th 911". Each prefix word must
# be Title Case (capital letter followed by at least one lowercase letter)
# so we don't match "TABLE OF AUTHORITIES Cases" headings as a citation.
_TITLE_WORD = r"[A-Z][a-z][A-Za-z]*"
# First word is constrained to NOT be one of the common connectors that
# could precede a case-name reference in body text ("The Ford Motor Warranty
# Cases held...", "In Ford Motor Warranty Cases the court..."). Without this
# guard the leading "The"/"In"/"See" gets glued onto the case name.
_CASES_FIRST = r"(?!The\b|In\b|See\b|Cf\b|But\b)" + _TITLE_WORD
CASES_RE = re.compile(
    rf"\b({_CASES_FIRST}(?:\s+{_TITLE_WORD}){{1,5}}\s+Cases)\s*"
    rf"(?:"
    rf"\((\d{{4}})\)\s+(\d{{1,4}})\s+({REPORTER_PATTERN})\s+(\d{{1,5}})"
    rf"|"
    rf",?\s+(\d{{1,4}})\s+({REPORTER_PATTERN})\s+(\d{{1,5}})\s*"
    rf"\((?:[^)]*?\b)?(\d{{4}})\)"
    rf")"
)

SIGNAL_PREFIXES = {
    "see", "cf", "cf.", "per", "in", "but", "compare", "accord", "e.g.",
    "also", "n", "of", "the", "and", "to", "by", "for", "with", "from",
    "as", "if", "when", "while", "since", "because", "though", "although",
    "court", "supreme", "federal", "state", "california",
}

# Tokens that are TOA section headers — when newline normalization runs, lines
# like "Cases\nSmith v. Jones..." become "Cases Smith v. Jones...", and the
# walk-back from "v." would otherwise pull "Cases" into the plaintiff name.
# We stop walk-back when we hit any of these as a standalone token. The match
# is case-insensitive and on the cleaned token (no surrounding punctuation).
_TOA_HEADERS = {
    "cases", "statutes", "rules", "authorities", "treatises",
    "regulations", "constitutional", "miscellaneous",
}

# Sentence-internal signal words that look like corporate abbreviations
# ("E.g.", "I.e.", "Cf.") and would otherwise pass the cap-then-short
# heuristic in walk-back. Listed here so `_walk_back_for_name` can refuse
# to collect them. Without this guard, "Song Beverly Act. E.g., Noori v.
# Jaguar..." gets collected as "Song Beverly Act. E.g., Noori".
# Stored without trailing punctuation because the matcher strips it before
# checking — entries are matched against `lower(strip("(.,;:\"\'", ",.;:"))`.
_STOPPER_ABBREVS = {
    "e.g", "i.e", "cf", "etc", "viz", "supra",
    "eg", "ie", "see", "accord", "compare",
}

# Corporate-suffix tokens that mark the end of a party name. Used by the
# digit-token rule in walk-back: a digit is only kept as part of the
# plaintiff if at least one of these has already been collected (otherwise
# the digit is almost certainly a page number bleeding in from a TOA layout).
# Stored without trailing punctuation — comparisons strip ".,;:" first.
_CORP_SUFFIX_LOWER = {
    "inc", "co", "corp", "ltd", "grp", "ass'n", "assn", "lp",
}
_CORP_SUFFIX_UPPER = {"LLC", "LLP", "LP", "LLLP", "PLLC", "PC", "PLC"}


def _short_name(plaintiff: str) -> str:
    p = plaintiff.strip()
    # Strip non-v. case-name prefixes ("In re", "Estate of",
    # "Guardianship of", etc.) plus "Ex parte" and "People v." so the
    # short name is the distinguishing subject ("Bowles", not "Estate").
    # The prefix group REPEATS ("+") because prefixes nest: "In re Marriage
    # of Smith" is "In re" + "Marriage of" + "Smith" — a single strip left
    # "Marriage of Smith", whose first word made every Marriage-of case share
    # the short name "Marriage" and never match its supra cites.
    p = re.sub(rf"^(?:(?:{_NONV_PREFIX})\s+|Ex parte\s+|People v\.\s+)+",
               "", p, flags=re.IGNORECASE)
    parts = p.split()
    return parts[0].rstrip(",.;:") if parts else p


def _walk_back_for_name(text: str, v_pos: int):
    """Return start index of plaintiff name, or None."""
    pos = v_pos - 1
    while pos > 0 and text[pos] == " ":
        pos -= 1

    # Skip a trailing ", et al." if present immediately before v.
    # Real text is "Juan Carlos Meneses, et al. v. FCA US LLC" — the walk-back
    # would otherwise hit "al." (lowercase) and abandon. We treat ", et al."
    # (or ", et al" without trailing period) as if it weren't there.
    et_al_match = re.search(r",\s*et\s+al\.?\s*$", text[: pos + 1])
    if et_al_match:
        pos = et_al_match.start() - 1
        while pos > 0 and text[pos] == " ":
            pos -= 1

    tokens = []  # (start, end, text) closest-to-v.-first
    while pos >= 0:
        while pos >= 0 and text[pos] in " \t":
            pos -= 1
        if pos < 0:
            break
        # Newline = stop (citation can't span paragraph)
        if text[pos] == "\n":
            break
        tok_end = pos + 1
        while pos >= 0 and text[pos] not in " \n\t":
            pos -= 1
        tok_start = pos + 1
        tok = text[tok_start:tok_end]
        if not tok:
            break

        # Hard sentence boundary: token ends with : ; ! ?  (the case name
        # cannot include these as the final character of a preceding token)
        if tok and tok[-1] in ":;!?":
            break

        # Stopper-abbreviations: "E.g.,", "I.e.,", "Cf.", "Etc." etc. look
        # superficially like the corporate-abbreviation pattern (capital-
        # then-lowercase-with-dots), but they're actually sentence-internal
        # signal words that mark the END of any case-name we should still be
        # collecting. Reject them explicitly so the walk-back doesn't slurp
        # text like "Song Beverly Act. E.g., Noori v. ..." into a plaintiff
        # name. The check is on the cleaned, lowercased form.
        _tok_clean_low = tok.lstrip("(.,;:\"'").rstrip(",.;:").lower()
        if _tok_clean_low in _STOPPER_ABBREVS:
            if tokens:
                break
            return None

        # End-of-sentence: ends with "." preceded by lowercase. We allow
        # any capitalized-then-short-lowercase token like "Co.", "Inc.",
        # "Ref.", "Mfg.", "Sav.", "Bldg." as part of a corporate name —
        # these appear constantly inside party names. The cap-then-short
        # heuristic matches abbreviations without false-positives on real
        # sentence ends ("approved.", "court.").
        if tok.endswith(".") and len(tok) > 1 and tok[-2].islower():
            inner = tok.rstrip(".")
            is_short_cap_abbrev = (
                inner and inner[0].isupper() and 1 <= len(inner) <= 6
            )
            if not is_short_cap_abbrev and tok.lower() not in {
                "co.", "inc.", "corp.", "ltd.", "ass'n.", "ass'n.",
            }:
                break

        connectors = {"of", "the", "and", "&", "de", "la", "du", "von", "van", "re"}
        # Strip leading punctuation. We also strip a leading hyphen because
        # PDFs sometimes render hyphenated party names like "Bigler-Engler"
        # with a stray space-hyphen-space sequence ("Bigler -Engler"), which
        # the tokenizer splits into two tokens "Bigler" and "-Engler". We
        # want "-Engler" to clean to "Engler" so the walk-back keeps going.
        clean = tok.lstrip("(.,;:\"'\u2010\u2011\u2012\u2013\u2014\u2212-").rstrip(",.;:")

        if not clean:
            break

        # Citation-join boundary: two citations strung together with "and"
        # (or "&") \u2014 "Orozco v. Casimiro (2004) 121 Cal.App.4th Supp. 7 and
        # Del Monte ... v. Dolan ...". Walking back from the second "v." we
        # must not cross into the first citation. If the connector's left
        # neighbour is a citation tail (a page number, or a reporter
        # abbreviation), the connector separates two citations rather than
        # two words of one party name, so stop and keep the name gathered so
        # far. A name-internal "&"/"and" ("Properties & Investments") has a
        # capitalized word \u2014 not a number/reporter \u2014 to its left and is kept.
        if clean.lower() in {"and", "&"} and tokens:
            _left = re.search(r"(\S+)\s*$", text[:tok_start])
            if _left:
                _lt = _left.group(1).rstrip(",.;:")
                if _lt.isdigit() or re.search(REPORTER_PATTERN, _lt):
                    break

        # Pure-digit tokens. These appear in real party names ("Studio
        # 1220, Inc.", "Advanced Grp. 400" — though that one's on the
        # defendant side) but ALSO appear as page numbers in TOAs that
        # bleed into the walk-back after newline normalization (e.g.
        # "...14, 16 McGee v. Mercedes-Benz..."). To distinguish, we
        # accept the digit if EITHER:
        #   (a) we've already collected a corporate-suffix token closer
        #       to v. — "Studio 1220, Inc." reaches "Inc." first, then
        #       "1220", at which point _has_corp_marker is true; OR
        #   (b) the digit is immediately preceded (in source order, i.e.
        #       the next leftward token we haven't tokenized yet) by a
        #       "local number" introducer like "Local", "Loc.", "No.",
        #       or "Chapter" — these unambiguously mark the digit as
        #       part of a party name ("Service Employees Local 660",
        #       "L.A. Coll. Fac. Guild Loc. 1521") rather than a page
        #       reference. Page references after newline normalization
        #       look like ", 16, 14, ..." with no such introducer.
        if clean[0].isdigit():
            # Comma-suffixed digit (e.g. "16,") is a page-reference list item
            # after TOA newline normalisation — never a company number.
            if tok.rstrip().endswith(","):
                if tokens:
                    break
                return None
            _has_corp_marker = any(
                t[2].rstrip(",.;:").lower() in _CORP_SUFFIX_LOWER
                or t[2].upper() in _CORP_SUFFIX_UPPER
                for t in tokens
            )
            # Peek at the leftward source text for a local-number introducer.
            # tok_start is the start index of the digit token in `text`.
            _local_intro = False
            _peek_left = text[:tok_start].rstrip()
            _last_tok_match = re.search(r"(\S+)$", _peek_left)
            if _last_tok_match:
                _prev = _last_tok_match.group(1).rstrip(",.;:").lower()
                if _prev in {"local", "loc", "no", "chapter", "ch"}:
                    _local_intro = True
            if not (_has_corp_marker or _local_intro):
                if tokens:
                    break
                return None
            tokens.append((tok_start, tok_end, tok))
            continue

        if clean[0].islower() and clean.lower() not in connectors:
            if tokens:
                break
            return None
        if not clean[0].isupper() and clean.lower() not in connectors:
            if tokens:
                break
            return None

        # Reject ALLCAPS tokens that are clearly heading text. We use a
        # length threshold of 5+ because:
        #   - Real party names sometimes have ALLCAPS abbreviations of 2-4
        #     chars: "OCM Principal", "FCA US LLC", "B.B.", "L.A. Times".
        #   - Heading words ("TABLE", "AUTHORITIES", "DEFENDANT", "MOTION",
        #     "SUMMARY", "JUDGMENT") are almost always 5+ chars.
        # We also reject law-firm suffixes that commonly appear in page
        # footers immediately above the body. "EXAMPLE COUNSEL GROUP LLP"
        # in a footer + "Santa Clara Valley Water Dist. v. ..." on the next
        # line yielded "LLP Santa Clara Valley..." as the captured plaintiff
        # before this guard.
        alpha_chars = [c for c in clean if c.isalpha()]
        if (len(alpha_chars) >= 5
                and all(c.isupper() for c in alpha_chars)
                and clean.lower() not in connectors):
            if tokens:
                break
            return None
        if clean.upper() in {"LLP", "LLC", "LLLP", "PLLC", "PC", "PLC"}:
            # Law-firm-suffix token. Two scenarios:
            #   (a) Part of the plaintiff name — "Smith LLC v. Jones". When
            #       walked backward from v., LLC is the FIRST token collected,
            #       directly attached to the rest of the party name.
            #   (b) Page-footer artifact — "EXAMPLE COUNSEL GROUP LLP\n
            #       Santa Clara Valley Water Dist. v. ...". Here LLC/LLP
            #       appears AFTER several plaintiff tokens have been
            #       collected, because the real plaintiff is "Santa Clara
            #       Valley Water Dist." and LLP is part of an upstream
            #       law-firm letterhead in the same flow.
            # We allow (a) and reject (b): break if we've already collected
            # tokens, accept otherwise.
            if tokens:
                break
            # Fall through to normal handling for first-token case.

        # Stop at TOA section-header words ("Cases", "Statutes", "Rules"...).
        # After newline normalization, a TOA layout that puts "Cases" on its
        # own line directly above a citation looks like "Cases Smith v. Jones".
        # Without this guard, "Cases" would get pulled into the plaintiff name.
        if clean.lower() in _TOA_HEADERS:
            if tokens:
                break
            return None

        tokens.append((tok_start, tok_end, tok))

    if not tokens:
        return None

    # tokens are reverse-order; reverse to forward
    tokens.reverse()

    # Strip leading signal words
    while tokens:
        first = tokens[0][2].lower().rstrip(",.;:").lstrip("(.,;:\"'")
        if first in SIGNAL_PREFIXES:
            # Preserve "In re"
            if first == "in" and len(tokens) > 1:
                second = tokens[1][2].lower().rstrip(",.;:")
                if second == "re":
                    break
            tokens.pop(0)
        else:
            break

    if not tokens:
        return None

    # Advance start past leading non-letter punctuation like "(" or quotation
    start = tokens[0][0]
    end = tokens[0][1]
    while start < end and not text[start].isalpha():
        start += 1
    return start


def find_case_citations(text: str):
    """Return list of citation dicts for full case cites in `text`."""
    results = []

    # v.-anchored cases
    for m in ANCHOR_RE.finditer(text):
        v_start = m.start()
        v_end = m.end()
        plaintiff_start = _walk_back_for_name(text, v_start)
        if plaintiff_start is None:
            continue
        plaintiff = text[plaintiff_start:v_start].strip()

        # Restrict tail to within ~80 chars of v. - cite shouldn't span far.
        # With newline normalization, a single citation comfortably fits in
        # this window; anything farther is almost certainly a different case.
        max_dist = 80
        # Search a BOUNDED window, not the whole remaining document: a tail
        # must START within max_dist and a real tail is well under 400 chars,
        # but searching all of `rest` made every "v." anchor pay a full-
        # document scan per tail pattern — O(n²) overall, seconds on a TOA-
        # heavy brief and hang-grade on pathological input.
        rest = text[v_end: v_end + max_dist + 400]

        # Strategy: find first occurrence of any tail in `rest`, then pick
        # whichever appears earliest. This handles CSM, Bluebook, Westlaw-only,
        # Lexis-only, and flat (no-comma) citation forms uniformly.
        csm_search   = CSM_TAIL.search(rest)
        bb_search    = BB_TAIL.search(rest)
        wl_search    = WL_TAIL.search(rest)
        lexis_search = LEXIS_TAIL.search(rest)
        flat_search  = FLAT_TAIL.search(rest)

        candidates = []
        if csm_search and csm_search.start() <= max_dist:
            candidates.append(("csm", csm_search))
        if bb_search and bb_search.start() <= max_dist:
            candidates.append(("bb", bb_search))
        if wl_search and wl_search.start() <= max_dist:
            candidates.append(("wl", wl_search))
        if lexis_search and lexis_search.start() <= max_dist:
            candidates.append(("lexis", lexis_search))
        if flat_search and flat_search.start() <= max_dist:
            candidates.append(("flat", flat_search))

        # Slip cite is a *fallback*: only consider it if no reporter-shaped
        # tail matched. Slip cites have no reporter cite to anchor a strong
        # match, so they're vulnerable to misreading "Case No." references
        # in body text that AREN'T citations. Requiring no other tail first
        # avoids picking the slip pattern over a real reporter cite when
        # both appear in the same sentence.
        if not candidates:
            slip_search = SLIP_TAIL.search(rest)
            if slip_search and slip_search.start() <= max_dist:
                candidates.append(("slip", slip_search))
        if not candidates:
            continue
        # Earliest tail wins. When two tails start at the same position
        # (FLAT_TAIL is a strict superset of BB_TAIL minus the comma —
        # they never tie because BB's comma takes a position FLAT can't),
        # the earliest-start rule resolves cleanly. CSM also can't tie BB/FLAT
        # because CSM's "(year)" lead disambiguates.
        chosen = min(candidates, key=lambda c: c[1].start())

        kind, mm = chosen
        defendant_text = rest[: mm.start()].rstrip(", ").strip()
        # Defendant normally begins with a capital, but a real party can be an
        # address or numbered entity ("100 Oak Street", "1200 Fifth Ave.
        # Partners") — anchored by the reporter tail it is still a citation, and
        # protecting that span keeps the address detector from rewriting a cited
        # decision's name. Allow a leading digit; the reporter cite is the guard.
        if not defendant_text or not (defendant_text[0].isupper()
                                      or defendant_text[0].isdigit()):
            continue
        if len(defendant_text) > 200:
            continue

        if kind == "csm":
            year, vol, reporter, page = mm.group(1), mm.group(2), mm.group(3), mm.group(4)
            rep_compact = _normalize_reporter(reporter)
            tail_for_key = f"({year}) {vol} {rep_compact} {page}"
        elif kind in ("bb", "flat"):
            # Group layout is identical: (vol, reporter, page, year).
            # The only structural difference is the comma at the start of BB,
            # which both patterns absorb internally before the captured groups.
            vol, reporter, page, year = mm.group(1), mm.group(2), mm.group(3), mm.group(4)
            rep_compact = _normalize_reporter(reporter)
            tail_for_key = f"({year}) {vol} {rep_compact} {page}"
        elif kind == "wl":
            wl_year, wl_num, decision_year = mm.group(1), mm.group(2), mm.group(3)
            tail_for_key = f"{wl_year} WL {wl_num}"
        elif kind == "lexis":
            lx_year, lx_num, decision_year = mm.group(1), mm.group(2), mm.group(3)
            tail_for_key = f"{lx_year} U.S. Dist. LEXIS {lx_num}"
        else:  # slip
            # Slip cites have no reporter cite — the docket id (group 1)
            # and court parenthetical (group 2) ARE the citation. Encode
            # both in the key so duplicate detection still works; URL
            # resolution falls back to a case-name search.
            docket = mm.group(1)
            court_paren = mm.group(2).strip()
            tail_for_key = f"Case No. {docket} ({court_paren})"

        plaintiff_clean = re.sub(r"\s+", " ", plaintiff).strip()
        defendant_clean = re.sub(r"\s+", " ", defendant_text).strip()
        key = f"{plaintiff_clean} v. {defendant_clean} {tail_for_key}"
        full_span = (plaintiff_start, v_end + mm.end())

        results.append({
            "kind": "case",
            "key": key,
            "span": full_span,
            "match_text": text[full_span[0]: full_span[1]],
            "short": _short_name(plaintiff_clean),
            # Party sides kept separately so the pseudonymizer's citation-span
            # protection can apply the caption exemption: a "X v. Y" span is
            # replaced (not protected) only when BOTH sides name a trusted party.
            "plaintiff": plaintiff_clean,
            "defendant": defendant_clean,
            # Westlaw-only unpublished decisions can't be served by Lexis;
            # always resolve these to a Westlaw URL even when the user's
            # default provider is Lexis.
            "wl_only": (kind == "wl"),
            # Lexis-only mirror: U.S. Dist. LEXIS database numbers can't be
            # served by Westlaw, so always resolve those to a Lexis URL.
            "lexis_only": (kind == "lexis"),
            # Slip cites have no reporter cite. URL resolution falls back
            # to a case-name search against the active provider.
            "slip_only": (kind == "slip"),
        })

    # In re / Estate of / Guardianship of / Conservatorship of /
    # Adoption of / Marriage of cases (no v.)
    for m in INRE_RE.finditer(text):
        name = re.sub(r"\s+", " ", m.group(1)).strip()
        # Recover which prefix matched. INRE_RE doesn't capture the
        # prefix as a numbered group (doing so would shift the numbered
        # captures the branch logic below depends on), so we re-extract
        # it from the matched text.
        prefix_m = _NONV_PREFIX_RE.match(m.group(0))
        prefix = prefix_m.group(1) if prefix_m else "In re"
        # Group layout: 1=name, then EITHER (2,3,4,5) for CSM form,
        # (6,7,8,9) for Bluebook form, or (10,11,12) for the WL form
        # (year-in-cite, WL-number, decision-year). Whichever branch
        # matched is non-None.
        wl_only = False
        if m.group(2):
            year, vol, reporter, page = m.group(2), m.group(3), m.group(4), m.group(5)
        elif m.group(6):
            vol, reporter, page, year = m.group(6), m.group(7), m.group(8), m.group(9)
        else:
            wl_year, wl_num, year = m.group(10), m.group(11), m.group(12)
            wl_only = True
        full_name = f"{prefix} {name}"
        if wl_only:
            key = f"{full_name} {wl_year} WL {wl_num}"
        else:
            rep_compact = _normalize_reporter(reporter)
            key = f"{full_name} ({year}) {vol} {rep_compact} {page}"
        results.append({
            "kind": "case",
            "key": key,
            "span": m.span(),
            "match_text": m.group(0),
            "short": _short_name(full_name),
            "wl_only": wl_only,
        })

    # "[Subject] Cases" — consolidated-litigation case names with no v./In re.
    for m in CASES_RE.finditer(text):
        name = re.sub(r"\s+", " ", m.group(1)).strip()
        if m.group(2):
            year, vol, reporter, page = m.group(2), m.group(3), m.group(4), m.group(5)
        else:
            vol, reporter, page, year = m.group(6), m.group(7), m.group(8), m.group(9)
        rep_compact = _normalize_reporter(reporter)
        key = f"{name} ({year}) {vol} {rep_compact} {page}"
        results.append({
            "kind": "case",
            "key": key,
            "span": m.span(),
            "match_text": m.group(0),
            "short": name.split()[0],
        })

    return results


def _skip_to_name_end(s: str) -> str:
    return s


def find_statute_citations(text: str):
    results = []
    for m in STATUTE_RE.finditer(text):
        abbrev = _statute_abbrev(m)
        if not abbrev:
            continue
        # First section: matched by the primary regex (anchored on a
        # code name).
        section = m.group("sec")
        # When the citation had no explicit "§"/"section"/"sec." marker, the
        # code name was directly followed by a number. Accept that only when
        # the number is clearly a statute section — it has a decimal part
        # ("430.30", "1010.6"), a letter suffix ("437c"), a subdivision
        # ("(a)"), or is at least three digits ("187"). This rejects the
        # paragraph/list counters ("1.", "2.") that frequently follow a code
        # name at a numbered-list break and would otherwise be linked as a
        # spurious "§ 2".
        if not m.group("mk"):
            distinctive = re.search(r"\.\d|[a-z]|\(", section, re.IGNORECASE)
            if not distinctive and len(re.sub(r"\D", "", section)) < 3:
                continue
        results.append({
            "kind": "statute",
            "key": f"{abbrev} § {section}",
            "span": m.span(),
            "match_text": m.group(0),
        })
        # Chained additional sections: "§§ A, B, and C" or
        # "sections A and B". The first section grabs the code-name
        # context; subsequent sections inherit the same abbreviation
        # without needing to repeat the code name.
        #
        # Guarded — the continuation used to accept ANY following number, so
        # ordinary prose after a cite became a statute link ("section 1005 and
        # 16 court days" -> a spurious CCP § 16 over "and 16"; "section 998
        # and 1998 amendments" -> CCP § 1998). A continuation now needs:
        #   * a PLURAL marker on the cite ("§§"/"sections"/"secs.") plus a
        #     section passing the usual distinctiveness test; or
        #   * under a singular/absent marker, a STRICTLY distinctive section
        #     (decimal part, letter suffix, or subdivision — "128.7", "437c",
        #     "1005(b)"), so "section 128.5 and 128.7" still chains while a
        #     bare prose number never does.
        marker = (m.group("mk") or "").lower().rstrip(".")
        plural = marker in ("§§", "sections", "secs")
        scan_pos = m.end()
        while True:
            cont = ADDL_SEC_RE.match(text, scan_pos)
            if not cont:
                break
            sec = cont.group("sec")
            if plural:
                ok = _section_is_distinctive(sec)
            else:
                ok = bool(re.search(r"\.\d|[a-z]|\(", sec, re.IGNORECASE))
            if not ok:
                break
            results.append({
                "kind": "statute",
                "key": f"{abbrev} § {sec}",
                "span": cont.span(),
                "match_text": text[cont.start():cont.end()].lstrip(),
            })
            scan_pos = cont.end()
    # Federal statutes: "9 U.S.C. § 1", "42 U.S.C. § 1983", etc.
    # Key form preserves the title number explicitly: "9 U.S.C. § 1".
    # These are detected by their own regex (USC_RE) rather than via
    # STATUTE_RE's California-code machinery; URL building keys off the
    # "U.S.C." literal in the key.
    for m in USC_RE.finditer(text):
        title = m.group("title")
        section = m.group("sec")
        results.append({
            "kind": "statute",
            "key": f"{title} U.S.C. \u00a7 {section}",
            "span": m.span(),
            "match_text": m.group(0),
        })
    return results


def find_rule_citations(text: str):
    results = []
    for m in RULE_RE.finditer(text):
        rule_num = m.group(1)
        key = f"Cal. Rules of Court, rule {rule_num}"
        results.append({
            "kind": "rule",
            "key": key,
            "span": m.span(),
            "match_text": m.group(0),
        })
    for m in RPC_RE.finditer(text):
        rule_num = m.group(1)
        key = f"Cal. Rules of Prof. Conduct, rule {rule_num}"
        results.append({
            "kind": "rule",
            "key": key,
            "span": m.span(),
            "match_text": m.group(0),
        })
    return results


def find_supra_citations(text: str, full_cites_in_order):
    """Find supra cites and resolve to first matching short_name."""
    seen = {}
    for c in full_cites_in_order:
        if c["kind"] == "case" and c.get("short"):
            seen.setdefault(c["short"], c)

    results = []
    for m in SUPRA_RE.finditer(text):
        cite_text = m.group(1)
        sname = _short_name(cite_text)
        if sname in seen:
            full = seen[sname]
            results.append({
                "kind": "case",
                "key": full["key"],
                "span": m.span(),
                "match_text": m.group(0),
                "short": sname,
                "is_supra": True,
                # Provider routing must FOLLOW the full cite: a WL-only key
                # resolved through the default provider built a Lexis URL for
                # a database number Lexis doesn't carry (and vice versa).
                "wl_only": full.get("wl_only", False),
                "lexis_only": full.get("lexis_only", False),
                "slip_only": full.get("slip_only", False),
            })
    return results


def _normalize_for_detection(text: str) -> str:
    """Normalize text for citation detection while preserving offsets.

    Replaces single newlines (bare line wraps) with a single space, while
    preserving paragraph breaks. PyMuPDF often emits paragraph breaks as
    \\n followed by space-runs and another \\n (e.g. "...end.\\n \\nNext..."),
    so we look through intervening whitespace when deciding whether a given
    newline is part of a paragraph break.

    Output length matches input length so spans returned by detection still
    index into the original text and PyMuPDF's `page.search_for(match_text)`
    still finds the original glyphs.
    """
    out = list(text)
    n = len(out)

    def _has_newline_within(i: int, direction: int, max_ws: int = 3) -> bool:
        # Look up to max_ws chars in `direction` (-1 or +1) past whitespace.
        # If we hit another \n, it's part of a paragraph break (keep it).
        j = i + direction
        steps = 0
        while 0 <= j < n and steps <= max_ws:
            ch = out[j]
            if ch == "\n":
                return True
            if ch == "\f":
                return True
            if not ch.isspace():
                return False
            j += direction
            steps += 1
        return False

    for i, ch in enumerate(out):
        if ch != "\n":
            continue
        if _has_newline_within(i, -1) or _has_newline_within(i, +1):
            continue  # part of a paragraph break, keep
        out[i] = " "
    return "".join(out)


def find_all_citations(text: str):
    """Find all citations in text, ordered by position."""
    # Normalize line wraps so cites split across lines still match. Spans
    # remain valid against the original text because length is preserved.
    norm = _normalize_for_detection(text)
    full_cases = find_case_citations(norm)
    statutes = find_statute_citations(norm)
    # Bare section citations ("Section 1150 through 1179a") linked to the code
    # referenced around them. Code-named statute citations already found take
    # priority and define both the occupied spans and the surrounding context.
    bare_sections = find_bare_section_citations(
        norm,
        [c["span"] for c in statutes],
        _code_context_positions(norm),
    )
    statutes = statutes + bare_sections
    rules = find_rule_citations(norm)

    # Update match_text to use the original text (preserves original
    # whitespace for downstream page.search_for calls).
    for c in full_cases + statutes + rules:
        s, e = c["span"]
        c["match_text"] = text[s:e]

    # Order full cases by position for supra resolution
    full_ordered = sorted(full_cases, key=lambda c: c["span"][0])
    supras = find_supra_citations(norm, full_ordered)
    for c in supras:
        s, e = c["span"]
        c["match_text"] = text[s:e]

    all_cites = full_cases + statutes + rules + supras
    all_cites.sort(key=lambda c: c["span"][0])

    # Deduplicate overlapping spans (e.g. "Smith, supra" inside a longer match)
    dedup = []
    last_end = -1
    for c in all_cites:
        if c["span"][0] >= last_end:
            dedup.append(c)
            last_end = c["span"][1]
    return dedup


# ────────────────────────────────────────────────────────────────────────────
# URL resolution
# ────────────────────────────────────────────────────────────────────────────
def _bare_section(sec: str) -> str:
    """Strip parenthetical subdivisions from a section identifier so it
    can be used as a search term.

    Search engines on Lexis and Westlaw don't separately index
    subdivisions — there's no page for "§ 430.10(e)", only "§ 430.10".
    A search term that includes the subdivision returns no results and
    the user lands on an empty search page instead of the statute.
    The citation key keeps the subdivision (it drives the underline
    text on the link), but URL building uses just the bare section.

    Examples:
      "430.10(e)"   → "430.10"
      "430.10(e)(1)"→ "430.10"
      "1714.10"     → "1714.10"   (decimal IS the section)
      "437c"        → "437c"      (letter suffix is part of section)
      "15200"       → "15200"
    """
    return re.sub(r"\([^)]*\).*$", "", sec)


def _wl_search_term(key: str):
    # Federal U.S.C.: key is "9 U.S.C. § 1" — Westlaw accepts that form
    # directly as a search term.
    if re.match(r"^\d+\s+U\.S\.C\.\s*\u00a7", key):
        return key
    m = re.match(r"^([A-Z]+)\s*§\s*(.+)$", key)
    if not m:
        return None
    p = WL_SEARCH_PREFIX.get(m.group(1))
    return f"{p} § {_bare_section(m.group(2))}" if p else None


def _lexis_search_term(key: str):
    # Federal U.S.C.: same direct form works for Lexis.
    if re.match(r"^\d+\s+U\.S\.C\.\s*\u00a7", key):
        return key
    m = re.match(r"^([A-Z]+)\s*§\s*(.+)$", key)
    if not m:
        return None
    p = LEXIS_SEARCH_PREFIX.get(m.group(1))
    return f"{p} § {_bare_section(m.group(2))}" if p else None


# Validated against the cross-opener extension's injectFloatingButton — these
# are the URL forms that have been confirmed to land on the right pages.
LEXIS_PDMFID = "1530671"


# Westlaw's findType=Y&cite= and Lexis's pdsearchterms= both expect a bare
# reporter citation like "13 Cal.App.5th 1152" or "2021 WL 1234567" - NOT the
# full key with case name and year. The extraction below pulls the reporter
# portion out of any of pdf_linker's case-key forms:
#   "Smith v. Jones (2017) 13 Cal.App.5th 1152"      (CSM)
#   "Anderson v. Liberty Lobby, Inc. (1986) 477 U.S. 242"  (Bluebook)
#   "In re Doe (2009) 555 F.3d 100"                  (In re)
#   "Ford Motor Warranty Cases (2025) 17 Cal.5th 1122"  (Cases)
#   "Smith v. Jones 2021 WL 1234567"                 (Westlaw-only)
_CASE_TAIL_RE = re.compile(r"\((\d{4})\)\s+(\d{1,4})\s+(\S+?)\s+(\d{1,5})\s*$")
_WL_TAIL_RE   = re.compile(r"(\d{4})\s+WL\s+(\d{4,8})\s*$")
# Lexis database tail: "YYYY U.S. Dist. LEXIS NNNNN" at end of a case key.
# Parallel to _WL_TAIL_RE for use in URL building.
_LEXIS_TAIL_RE = re.compile(r"(\d{4})\s+U\.S\.\s*Dist\.\s*LEXIS\s+(\d{4,8})\s*$")


def _case_reporter_cite(case_key: str):
    """Return the reporter portion of a case key, e.g. '13 Cal.App.5th 1152'.
    Returns None if the key doesn't end with a parseable reporter cite."""
    m = _CASE_TAIL_RE.search(case_key)
    if m:
        _year, vol, reporter, page = m.groups()
        return f"{vol} {reporter} {page}"
    m = _WL_TAIL_RE.search(case_key)
    if m:
        year, num = m.groups()
        return f"{year} WL {num}"
    m = _LEXIS_TAIL_RE.search(case_key)
    if m:
        year, num = m.groups()
        return f"{year} U.S. Dist. LEXIS {num}"
    return None


def _slip_search_term(case_key: str) -> str:
    """Build a search term for a slip-cite key.

    Slip-cite keys are shaped:
      "<plaintiff> v. <defendant>[, ?]Case No. <docket> (<court>)"
    Both Westlaw and Lexis return useful results when searched by case
    name alone (the docket id is too narrow and depends on the
    provider's database indexing). We strip the "Case No. ... (...)"
    tail and return just the case name.
    """
    # Strip the slip tail.
    m = re.search(r",?\s*Case\s+No\.\s+", case_key, re.IGNORECASE)
    if m:
        return case_key[: m.start()].strip()
    return case_key


def _disambiguated_lexis_term(case_key: str) -> str:
    """Lexis search term: the FULL case name (both parties) + reporter cite,
    e.g. "People v. Smith 13 Cal.App.5th 1152".

    Lexis resolves cases through a search query, and the reporter cite
    ("13 Cal.App.5th 1152") is the unique anchor. Including both party names
    (rather than just the lead plaintiff word) disambiguates when the lead
    party is generic — "People v.", "City of …", "In re …" — and tolerates a
    stray leading token from imperfect name extraction without hurting the
    match, since the reporter cite still pins the result. (Westlaw's direct
    findType=Y deep-link is the opposite case and must stay name-free — it
    uses _case_reporter_cite, not this function.)

    The full name also subsumes the "In re X" / "Estate of X" / "X Cases"
    forms, so no special-casing is needed: whatever precedes the reporter
    tail in the key is the case name we search on.

    Kept byte-for-byte in step with pdf-viewer's disambiguatedLexisTerm
    (viewer/code-tables.js), which the cross-opener tooling also mirrors.
    """
    m = _CASE_TAIL_RE.search(case_key)
    if not m:
        return case_key
    _year, vol, reporter, page = m.groups()
    reporter_cite = f"{vol} {reporter} {page}"
    # Trim a trailing comma OR semicolon left on the name by the tail split.
    name_part = case_key[: m.start()].strip().rstrip(",;").strip()
    return f"{name_part} {reporter_cite}" if name_part else reporter_cite


def _build_westlaw_case_url(cite_text: str) -> str:
    """Westlaw direct-link form. cite_text MUST be just the reporter cite
    (vol REPORTER page) - including the case name confuses findType=Y and
    yields 'page not found'.
    
    WL citations (e.g., '2015 WL 13626022') use a search-based URL instead of
    findType=Y, which doesn't work reliably for unpublished WL reporters."""
    from urllib.parse import quote
    
    # Check if this is a WL citation (format: YEAR WL NUMBER)
    if " WL " in cite_text:
        # WL citations need search URLs, not direct link format
        return (
            f"https://1.next.westlaw.com/Search/Results.html"
            f"?query={quote(cite_text)}&jurisdiction=CA&contentType=CASE"
        )
    
    # Published reporters use direct link format
    return (
        f"https://1.next.westlaw.com/Link/Document/FullText"
        f"?findType=Y&cite={quote(cite_text)}"
    )


def _build_westlaw_search_url(query: str, statute: bool) -> str:
    from urllib.parse import quote
    base = (
        f"https://1.next.westlaw.com/Search/Results.html"
        f"?query={quote(query)}&jurisdiction=CA"
    )
    return base + "&contentType=STATUTE" if statute else base


def _build_lexis_search_url(term: str) -> str:
    from urllib.parse import quote
    return (
        f"https://plus.lexis.com/search/"
        f"?pdmfid={LEXIS_PDMFID}&pdsearchterms={quote(term)}"
    )


def resolve_url(cite, provider: str = "lexis") -> str:
    """Resolve citation to a search URL on the active provider.

    `provider` is "lexis" (default) or "westlaw".

    Westlaw-only unpublished decisions (`wl_only=True` on the cite) override
    the provider and always resolve to a Westlaw URL — Lexis doesn't carry
    those reporters, so a Lexis search would fail. Lexis-only unpublished
    decisions (`lexis_only=True`, used for U.S. Dist. LEXIS database
    numbers) are the symmetric case: always resolve to Lexis, regardless
    of the active provider.

    Earlier versions consulted a curated citation_repo.json that mapped
    keys to direct-link URLs (Lexis `lni=` document URLs and Westlaw
    direct-cite URLs). Those direct links rot — Lexis re-indexes content
    periodically, and stale `lni`s land on broken pages. Search URLs
    don't rot: the search engines tolerate variations in formatting and
    consistently land on the right document. The repo lookup was
    removed in favor of always building a search URL.
    """
    # Override for unpublished WL / LEXIS cites: each provider can't serve
    # the other's database numbers.
    if cite.get("wl_only"):
        effective_provider = "westlaw"
    elif cite.get("lexis_only"):
        effective_provider = "lexis"
    else:
        effective_provider = provider

    # Build a search URL using the active provider's form.
    if cite["kind"] == "case":
        if cite.get("slip_only"):
            # Slip cites have no reporter to anchor a direct-link URL.
            # Use a case-name search against the active provider. The full
            # key includes case name + "Case No. X (court)" which is enough
            # for the search engines to find the case (and short enough not
            # to overflow the URL).
            search_term = _slip_search_term(cite["key"])
            if effective_provider == "lexis":
                return _build_lexis_search_url(search_term)
            return _build_westlaw_search_url(search_term, statute=False)
        if effective_provider == "lexis":
            return _build_lexis_search_url(_disambiguated_lexis_term(cite["key"]))
        reporter_cite = _case_reporter_cite(cite["key"]) or cite["key"]
        return _build_westlaw_case_url(reporter_cite)

    if cite["kind"] == "statute":
        if effective_provider == "lexis":
            term = _lexis_search_term(cite["key"]) or cite["key"]
            return _build_lexis_search_url(term)
        term = _wl_search_term(cite["key"]) or cite["key"]
        return _build_westlaw_search_url(term, statute=True)

    # Rules (and any unrecognised statute)
    if effective_provider == "lexis":
        return _build_lexis_search_url(cite["key"])
    return _build_westlaw_search_url(cite["key"], statute=False)


# ────────────────────────────────────────────────────────────────────────────
# OCR helpers
# ────────────────────────────────────────────────────────────────────────────
def _find_tesseract():
    """Locate tesseract.exe; return path string or None."""
    # Check PATH first
    import shutil
    path_hit = shutil.which("tesseract")
    if path_hit:
        return path_hit
    for cand in TESSERACT_CANDIDATES:
        if cand and cand.exists():
            return str(cand)
    return None


def _suppress_child_error_dialogs():
    """Windows: stop child processes from showing the system error dialog when
    they fail to start (e.g. a broken tesseract.exe dying with 0xc0000142).
    The error mode is inherited by children, so a per-page OCR loop over a
    corrupted install degrades to log lines instead of one popup PER PAGE.
    No-op elsewhere; best-effort."""
    if os.name != "nt":
        return
    try:
        import ctypes
        SEM = 0x0001 | 0x0002 | 0x8000   # FAILCRITICALERRORS | NOGPFAULTERRORBOX
        ctypes.windll.kernel32.SetErrorMode(         # | NOOPENFILEERRORBOX
            ctypes.windll.kernel32.SetErrorMode(0) | SEM)
    except Exception:
        pass


# Probe result per tesseract path, so a BROKEN install (present on disk but
# unable to start — DLL init failure 0xc0000142, wrong bitness, missing VC++
# runtime) is detected ONCE per run instead of failing per page.
_TESS_PROBE = {}


def _tesseract_usable(tess, log):
    """True when `tess` actually RUNS (`tesseract --version` exits 0). A
    corrupted install used to be discovered one page at a time — every page
    spawned a tesseract.exe that died with a Windows error popup. Probed once
    and cached for the run; on failure OCR is disabled with one log line."""
    cached = _TESS_PROBE.get(tess)
    if cached is not None:
        return cached
    _suppress_child_error_dialogs()
    import subprocess
    try:
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
        r = subprocess.run([tess, "--version"], capture_output=True,
                           timeout=20, creationflags=flags)
        ok = r.returncode == 0
    except Exception:
        ok = False
    _TESS_PROBE[tess] = ok
    if not ok:
        log.warning(
            f"Tesseract is installed but cannot start ({tess}) — skipping OCR "
            f"for this run. This is usually a corrupted install or a missing "
            f"Visual C++ runtime (Windows error 0xc0000142); reinstall from "
            f"https://github.com/UB-Mannheim/tesseract/wiki")
    return ok


# A really large scanned page can make Tesseract stall indefinitely. pytesseract
# defaults to timeout=0 (wait forever), so a single stuck page hangs the WHOLE
# run — the process sits in subprocess.communicate() at ~0% CPU and never
# finishes. We bound each ATTEMPT (so a stall is escapable) but never skip the
# page: on timeout we re-render it at a lower resolution and try again, grinding
# down until Tesseract completes, so every page still gets a text layer. The
# per-attempt budget (10 min) comfortably covers a genuinely slow page at full
# resolution while a real stall gives up and grinds down without pausing its
# batch too long; override with PDF_LINKER_OCR_TIMEOUT (seconds).
_OCR_PAGE_TIMEOUT = None


def _ocr_page_timeout():
    global _OCR_PAGE_TIMEOUT
    if _OCR_PAGE_TIMEOUT is None:
        try:
            _OCR_PAGE_TIMEOUT = max(
                30, int(os.environ.get("PDF_LINKER_OCR_TIMEOUT", "600")))
        except (ValueError, TypeError):
            _OCR_PAGE_TIMEOUT = 600
    return _OCR_PAGE_TIMEOUT


# OCR is the dominant cost on a scanned set, and Tesseract runs as a subprocess
# (the calling thread blocks in a C wait that releases the GIL), so N worker
# threads run N Tesseract processes across N cores for a near-linear speedup at
# NO quality cost — the alternative, lowering dpi, would degrade the exhibit
# image itself. Rendering stays on the main thread (PyMuPDF is not thread-safe);
# only the Tesseract call is parallel. Default to cores-1 (leaving one core for
# the main-thread render/overlay and the OS), capped at 10 so RAM (one rendered
# page per worker) and process oversubscription stay sane on a many-core box;
# override with PDF_LINKER_OCR_WORKERS. 1 disables parallelism entirely.
_OCR_WORKERS = None
_OCR_WORKER_CAP = 10


def _ocr_workers():
    global _OCR_WORKERS
    if _OCR_WORKERS is None:
        env = os.environ.get("PDF_LINKER_OCR_WORKERS")
        if env:
            try:
                _OCR_WORKERS = max(1, int(env))
            except (ValueError, TypeError):
                _OCR_WORKERS = 1
        else:
            _OCR_WORKERS = min(_OCR_WORKER_CAP, max(1, (os.cpu_count() or 2) - 1))
    return _OCR_WORKERS


def _chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


# Cap the rasterised image so a pathologically large page (a big scanned exhibit,
# an engineering drawing, a mis-sized page) doesn't produce a 100+ megapixel
# bitmap that makes Tesseract crawl or stall. 30 MP is ~ a US-letter page at
# 470 dpi — well above the 300 dpi we target, so a normal page is unaffected.
_OCR_MAX_PIXELS = 30_000_000


def _ocr_base_dpi(page):
    """The starting OCR resolution for `page`: 300 dpi, reduced only when that
    would blow past the pixel cap so a very large page still starts within a size
    Tesseract can handle (and within memory)."""
    dpi = 300
    rect = page.rect
    scale = dpi / 72.0
    px = (rect.width * scale) * (rect.height * scale)
    if px > _OCR_MAX_PIXELS:
        dpi = max(72, int(dpi * (_OCR_MAX_PIXELS / px) ** 0.5))
    return dpi


def _ocr_pixmap(page):
    """Render `page` for OCR at its base resolution (see `_ocr_base_dpi`)."""
    return page.get_pixmap(dpi=_ocr_base_dpi(page))


# Resolution ladder for the grind-don't-skip retry: each attempt renders the page
# at this fraction of its base dpi. A stall usually tracks image size/complexity,
# so a smaller render escapes it while still producing a text layer; the floor
# keeps even the last attempt legible enough to be useful.
_OCR_RETRY_FACTORS = (1.0, 0.66, 0.5, 0.33, 0.24)
_OCR_MIN_DPI = 60


def _ocr_page_to_pdf(page, pytesseract, Image, io, log, label, start=0):
    """OCR one page to overlay-PDF bytes, GRINDING rather than skipping. Try at
    full resolution; if Tesseract times out (a stall or a pathological page),
    re-render at successively lower resolution and try again. Only the render
    resolution drops — the page always gets a text layer. Returns the overlay
    PDF bytes, or None only if even the lowest resolution fails (effectively
    never, and logged loudly if so). `start` skips the first rungs (a page that
    already stalled at full resolution in the parallel pass resumes lower)."""
    base = _ocr_base_dpi(page)
    timeout = _ocr_page_timeout()
    last = len(_OCR_RETRY_FACTORS) - 1
    for i, factor in enumerate(_OCR_RETRY_FACTORS):
        if i < start:
            continue
        dpi = max(_OCR_MIN_DPI, int(base * factor))
        try:
            pix = page.get_pixmap(dpi=dpi)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            return pytesseract.image_to_pdf_or_hocr(
                img, extension="pdf", timeout=timeout)
        except Exception as e:
            if i < last:
                nxt = max(_OCR_MIN_DPI, int(base * _OCR_RETRY_FACTORS[i + 1]))
                log.warning(
                    f"  {label} page {page.number} did not finish within "
                    f"{timeout}s at {dpi} dpi ({e}); retrying at {nxt} dpi")
                continue
            log.warning(f"  {label} ultimately failed on page {page.number}: {e}")
            return None


def _ocr_pdf(doc, log):
    """OCR pages of a PyMuPDF doc that have no text. Adds an invisible text
    layer using the recognised text. Modifies doc in place."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        log.warning("pytesseract or Pillow not installed - skipping OCR")
        return False

    tess = _find_tesseract()
    if not tess:
        log.warning("Tesseract not found - skipping OCR. Install from "
                    "https://github.com/UB-Mannheim/tesseract/wiki")
        return False
    if not _tesseract_usable(tess, log):
        return False
    pytesseract.pytesseract.tesseract_cmd = tess

    import io
    import fitz

    pages = [p for p in doc if not p.get_text("text").strip()]
    if not pages:
        return False

    ocr_count = 0

    def _overlay(page, pdf_bytes):
        nonlocal ocr_count
        try:
            ocr_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            page.show_pdf_page(page.rect, ocr_doc, 0, overlay=True)
            ocr_doc.close()
            ocr_count += 1
        except Exception as e:
            log.warning(f"  Could not overlay OCR text on page {page.number}: {e}")

    workers = _ocr_workers()
    if workers <= 1 or len(pages) <= 1:
        # Sequential: render + OCR + grind, one page at a time.
        for page in pages:
            hocr = _ocr_page_to_pdf(page, pytesseract, Image, io, log, "OCR")
            if hocr is not None:
                _overlay(page, hocr)
    else:
        # Parallel: RENDER on the main thread (PyMuPDF is not thread-safe), OCR
        # in a thread pool (each Tesseract subprocess runs on its own core, GIL
        # released while it works), OVERLAY back on the main thread. Pages that
        # stall at full resolution are collected and ground down sequentially
        # afterwards, so grind-don't-skip is preserved.
        from concurrent.futures import ThreadPoolExecutor
        # One thread per worker keeps each Tesseract single-threaded, so N
        # workers cleanly use N cores instead of oversubscribing.
        os.environ.setdefault("OMP_THREAD_LIMIT", "1")
        timeout = _ocr_page_timeout()
        log.info(f"  OCR: {len(pages)} page(s) across {workers} workers")
        stalled = []
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for chunk in _chunked(pages, workers):
                rendered = []
                for page in chunk:
                    try:
                        pix = _ocr_pixmap(page)
                        img = Image.open(io.BytesIO(pix.tobytes("png")))
                    except Exception as e:
                        log.warning(f"  OCR render failed on page {page.number}: {e}")
                        continue
                    rendered.append((page, img))
                futs = [(page, ex.submit(
                            pytesseract.image_to_pdf_or_hocr, img,
                            extension="pdf", timeout=timeout))
                        for page, img in rendered]
                for page, fut in futs:
                    try:
                        _overlay(page, fut.result())
                    except Exception as e:
                        log.warning(f"  OCR did not finish on page {page.number} at "
                                    f"full resolution ({e}); will grind it down")
                        stalled.append(page)
        # Grind the stragglers on the main thread, resuming below the resolution
        # that already stalled.
        for page in stalled:
            hocr = _ocr_page_to_pdf(page, pytesseract, Image, io, log, "OCR", start=1)
            if hocr is not None:
                _overlay(page, hocr)

    if ocr_count:
        log.info(f"  OCR'd {ocr_count} page(s)")
    return ocr_count > 0


_GARBLE_VOWELS = set("aeiouy")


def _text_looks_garbled(text: str) -> bool:
    """Heuristic: does this page's *extracted* text look like broken output
    (bad font/ToUnicode encoding, unmapped glyphs, junk) rather than real
    prose? Deliberately conservative — it only fires on obviously bad text, so
    a good text layer is never needlessly re-OCR'd.

    Signals, in order:
      * "(cid:NN)" tokens — PDFMiner-style unmapped glyphs.
      * A real density of U+FFFD replacement characters.
      * Mostly non-letters (symbol soup from a broken encoding).
      * A low fraction of word-shaped alphabetic tokens. On real legal prose
        this fraction sits at ~0.99 (measured), so a 0.50 cut has a wide
        safety margin while still catching gibberish.
    """
    if not text:
        return False
    # Ignore any legacy right-margin markers a prior version may have stamped.
    text = _MARKER_DETECT_RE.sub("", text)
    non_space = sum(1 for c in text if not c.isspace())
    if non_space < 200:
        return False  # too little content to judge reliably
    if "(cid:" in text:
        return True
    if text.count("�") >= max(8, 0.03 * non_space):
        return True
    # Letters AND digits both count as meaningful extraction: a damages table
    # or repair invoice is digit-dominated yet perfectly well extracted, and
    # the letters-only ratio sent those pages into destructive re-OCR (all
    # real text redacted, replaced by 300-dpi OCR guesses) and gave them the
    # 40x OCR ETA weight. Symbol soup from a broken encoding is low on BOTH.
    meaningful = sum(1 for c in text if c.isalpha() or c.isdigit())
    if meaningful / non_space < 0.35:
        return True  # dominated by symbols/punctuation
    tokens = re.findall(r"[A-Za-z]{2,}", text)
    if len(tokens) < 30:
        return False

    def _wordish(t: str) -> bool:
        tl = t.lower()
        if len(t) > 24:
            return False                              # words don't run this long
        if len(t) > 3 and not any(c in _GARBLE_VOWELS for c in tl):
            return False                              # no vowel in a long token
        if re.search(r"[bcdfghjklmnpqrstvwxz]{5,}", tl):
            return False                              # 5+ consonants in a row
        return True

    good = sum(1 for t in tokens if _wordish(t))
    return good / len(tokens) < 0.50


def _reocr_garbled_pages(doc, log):
    """Rebuild the text layer of any page whose extracted text looks garbled.

    A page's glyphs still *render* correctly even when a broken encoding makes
    them *extract* as gibberish, so we: render the page (correct visual),
    remove the bad text, and overlay a fresh OCR text layer. Only pages that
    look clearly garbled (see `_text_looks_garbled`) are touched, so a good
    text layer is never rasterised or degraded. No-op with a warning if OCR
    tooling is unavailable. Returns the number of pages re-OCR'd.
    """
    garbled = [p.number for p in doc if _text_looks_garbled(p.get_text("text"))]
    if not garbled:
        return 0
    try:
        import io
        import pytesseract
        import fitz
        from PIL import Image
    except ImportError:
        log.warning(f"  {len(garbled)} page(s) look garbled but pytesseract/"
                    f"Pillow isn't installed - leaving text as-is")
        return 0
    tess = _find_tesseract()
    if not tess:
        log.warning(f"  {len(garbled)} page(s) look garbled but Tesseract "
                    f"isn't installed - leaving text as-is")
        return 0
    if not _tesseract_usable(tess, log):
        return 0
    pytesseract.pytesseract.tesseract_cmd = tess
    keep_img = getattr(fitz, "PDF_REDACT_IMAGE_NONE", 0)
    keep_art = getattr(fitz, "PDF_REDACT_LINE_ART_NONE", 0)

    # Phase 1 — render + OCR every garbled page, in parallel across workers. This
    # is the heavy step (448 pages was ~15 min single-threaded). Rendering stays
    # on the main thread (PyMuPDF is not thread-safe); only Tesseract runs in the
    # pool. Rendering BEFORE any redaction is required — the visible glyphs are
    # correct until we strip the bad text in phase 2, and a redaction on one page
    # never affects another page's already-captured image. Stalled pages grind
    # down afterwards, so nothing is skipped.
    pages = [doc[n] for n in garbled]
    workers = _ocr_workers()
    ocr_by_page = {}
    if workers <= 1 or len(pages) <= 1:
        for page in pages:
            b = _ocr_page_to_pdf(page, pytesseract, Image, io, log, "Re-OCR")
            if b is not None:
                ocr_by_page[page.number] = b
    else:
        from concurrent.futures import ThreadPoolExecutor
        os.environ.setdefault("OMP_THREAD_LIMIT", "1")
        timeout = _ocr_page_timeout()
        log.info(f"  Re-OCR: {len(pages)} page(s) across {workers} workers")
        stalled = []
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for chunk in _chunked(pages, workers):
                rendered = []
                for page in chunk:
                    try:
                        pix = _ocr_pixmap(page)
                        img = Image.open(io.BytesIO(pix.tobytes("png")))
                    except Exception as e:
                        log.warning(f"  Re-OCR render failed on page {page.number}: {e}")
                        continue
                    rendered.append((page, img))
                futs = [(page, ex.submit(
                            pytesseract.image_to_pdf_or_hocr, img,
                            extension="pdf", timeout=timeout))
                        for page, img in rendered]
                for page, fut in futs:
                    try:
                        ocr_by_page[page.number] = fut.result()
                    except Exception as e:
                        log.warning(f"  Re-OCR did not finish on page {page.number} "
                                    f"at full resolution ({e}); will grind it down")
                        stalled.append(page)
        for page in stalled:
            b = _ocr_page_to_pdf(page, pytesseract, Image, io, log, "Re-OCR", start=1)
            if b is not None:
                ocr_by_page[page.number] = b

    # Phase 2 — strip the garbled text and overlay the fresh OCR (serial; mutates
    # the doc, which PyMuPDF requires single-threaded).
    done = 0
    for pno in garbled:
        ocr_bytes = ocr_by_page.get(pno)
        if ocr_bytes is None:
            continue
        page = doc[pno]
        # Strip the garbled text (text only; keep images and line art, no fill)
        # so get_text no longer returns the gibberish.
        try:
            page.add_redact_annot(page.rect, fill=False)
            try:
                page.apply_redactions(images=keep_img, graphics=keep_art)
            except TypeError:
                page.apply_redactions(images=keep_img)
        except Exception as e:
            log.warning(f"  Could not remove garbled text on page {pno}: {e}")
            continue
        # Overlay the fresh OCR render (rendered image + clean invisible text).
        try:
            ocr_doc = fitz.open(stream=ocr_bytes, filetype="pdf")
            page.show_pdf_page(page.rect, ocr_doc, 0, overlay=True)
            ocr_doc.close()
            done += 1
        except Exception as e:
            log.warning(f"  Could not overlay re-OCR on page {pno}: {e}")
    if done:
        log.info(f"  Re-OCR'd {done} page(s) with garbled text layers")
    return done


# ────────────────────────────────────────────────────────────────────────────
# Files that should be skipped for hyperlink insertion
# ────────────────────────────────────────────────────────────────────────────
# Declarations and separate statements rarely contain citations worth linking
# (they're factual recitations or ruling-by-ruling responses), and Zachary's
# workflow doesn't benefit from links there. The right-margin marker
# injection still runs on these files — that's the citation-paste aid the
# Word macro relies on, separate from the link annotations — but the
# case/statute hyperlink insertion is skipped entirely.
#
# Match is case-insensitive against the filename stem (no extension), with
# spaces and underscores treated equivalently. Patterns must appear as a
# whole token or sequence so partial words ("Declaratory", "Decline") don't
# fire.
_SKIP_LINK_PATTERNS = [
    re.compile(r"(?:^|[\s_-])Declaration(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Decl\.?(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Separate[\s_-]+Statement(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Compl\.?(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Complaint(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])FAC(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])SAC(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])TAC(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Proof[\s_-]+of[\s_-]+Service(?:[\s_-]|$)", re.IGNORECASE),
]


def should_skip_linking(filename: str) -> bool:
    """Return True if this filename indicates a doc type we don't link."""
    name = filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    stem = name.rsplit(".", 1)[0]
    for pat in _SKIP_LINK_PATTERNS:
        if pat.search(stem):
            return True
    return False


def _join_spans_spaced(spans):
    """Concatenate a row's spans (already x-sorted) into one string, inserting
    a single space at a span boundary when there's a real horizontal gap and
    neither side already has whitespace. PyMuPDF sometimes splits a line into
    spans at a style change and drops the space between them (e.g. "1." + "I
    am" -> "1.I am"); this restores it without gluing within-word pieces like
    the superscript "th" in "4th"."""
    out = ""
    prev_x1 = None
    for s in spans:
        t = s.get("text", "")
        if not t:
            continue
        x0, _y0, x1, y1 = s["bbox"]
        if (out and prev_x1 is not None and not out[-1].isspace()
                and not t[:1].isspace() and (x0 - prev_x1) > 0.2 * (y1 - s["bbox"][1])):
            out += " "
        out += t
        prev_x1 = x1
    return out


# A physical row of text is identified by its BASELINE, not by its bbox
# mid-point. Glyph boxes of different point sizes on one row have different
# mid-points but share a baseline, so bucketing by a rounded mid-point splits
# one row into two buckets — and the old code emitted at most one bucket per
# gutter line, silently discarding the other. That is how caption text went
# missing from the .txt.
_ROW_BASELINE_TOL = 3.0     # pt; spans this close share a row
_GUTTER_MATCH_TOL = 24.0    # pt; hard ceiling on how far a row may sit from its
                            # line number (the effective tolerance is half the
                            # page's own median lead — see _detect_line_anchors)
_COLUMN_GAP_MIN = 20.0      # pt; a horizontal gap this wide separates columns
_COLUMN_BAND_MIN_ROWS = 3   # distinct physical rows that must break at an x
                            # before it counts as a page column (fewer is a
                            # within-column tab/leader gap, not a column). On a
                            # sparse page the floor scales down (rows // 4) so a
                            # short two-column block is still two columns.
_FOOTER_MASK_PT = 50.0      # pt; bottom band holding running footers, printed
                            # page numbers and Bates stamps. Rows whose baseline
                            # falls below it are page furniture and are never
                            # adopted onto a numbered line.

# A caption's vertical brace column (")" repeated down the page) is typography,
# not content: it carries no text and its rows sit at their own baselines, so it
# only ever reorders the segments around it. Dropped from the numbered-line
# export.
_CAPTION_DIVIDER_RE = re.compile(r"^[)(|\]\[]+$")
# The same brace, but close enough to the right-hand column that no gutter
# separates them (") Case No.: 25STCV37838"). Only a CLOSING brace is stripped,
# and only with whitespace after it — "(FAC, p. 8:17-18.)" must survive intact.
_CAPTION_DIVIDER_LEAD_RE = re.compile(r"^[)\]|]+\s+")


def _span_baseline(sp):
    """Baseline y of a span. PyMuPDF gives `origin`; fall back to the bbox
    mid-point for span dicts that lack it."""
    origin = sp.get("origin")
    if origin:
        return origin[1]
    return (sp["bbox"][1] + sp["bbox"][3]) / 2


def _cluster_rows(spans, tol=_ROW_BASELINE_TOL):
    """Group spans into physical rows by baseline. Returns [{y, spans}] sorted
    top to bottom, with every input span in exactly one row (nothing dropped).

    Superscripts and subscripts ("5th", a footnote mark) are set well below the
    body size and their RAISED baseline pulls them onto the row above, where
    they surface as an orphan segment ("...LAW OFFICE OF PAUL GREEN  th") and,
    worse, skew that row toward the wrong gutter number. So body-size spans are
    clustered first; each small span is then attached to the row whose baseline
    its glyph-box BOTTOM sits on (that bottom rests on the base line of the word
    it belongs to). The row's baseline `y` is always taken from its body spans,
    so an attached superscript can't move it."""
    sizes = [sp.get("size") for sp in spans if sp.get("size")]
    med = statistics.median(sizes) if sizes else 0.0

    def _is_small(sp):
        return bool(med) and sp.get("size", med) < 0.72 * med

    rows = []
    for sp in sorted((s for s in spans if not _is_small(s)), key=_span_baseline):
        y = _span_baseline(sp)
        if rows and abs(y - rows[-1]["y"]) <= tol:
            rows[-1]["spans"].append(sp)
        else:
            rows.append({"y": y, "spans": [sp]})

    for sp in (s for s in spans if _is_small(s)):
        bottom = sp["bbox"][3]
        best = min(rows, key=lambda r: abs(r["y"] - bottom)) if rows else None
        if best is not None and abs(best["y"] - bottom) <= tol + 2.0:
            best["spans"].append(sp)
        else:
            rows.append({"y": _span_baseline(sp), "spans": [sp]})

    rows.sort(key=lambda r: r["y"])
    return rows


def _split_char_runs(chars, gap_min=_COLUMN_GAP_MIN, back_tol=5.0):
    """Split one span's characters into contiguous runs wherever the x-advance
    JUMPS — forward by more than `gap_min` (the glyphs hopped to another page
    column) or backward past `back_tol` (two columns interleaved, so the pen
    snapped back left mid-span). This is the character-geometry cure for a
    welded caption: the text layer hands back ONE span reading
    "CULTUREEDIT, LLCAttorneys for Plaintiff", but its glyph boxes still sit in
    two distinct columns, and splitting at the jump restores what the page
    actually shows. Ordinary kerning overlaps by well under `back_tol`, so
    clean spans come back as a single run.

    `chars` is [(x0, x1, ch, origin_y), ...] in stream order; returns
    [(x0, x1, origin_y, text), ...] with whitespace-only runs dropped."""
    runs, cur, prev_x1 = [], None, None
    for x0, x1, ch, oy in chars:
        if cur is not None and prev_x1 is not None and (
                x0 - prev_x1 > gap_min or x0 < prev_x1 - back_tol):
            runs.append(cur)
            cur = None
        if cur is None:
            cur = [x0, x1, oy, []]
        cur[1] = max(cur[1], x1)
        cur[3].append(ch)
        prev_x1 = x1
    if cur is not None:
        runs.append(cur)
    out = []
    for x0, x1, oy, chs in runs:
        text = "".join(chs)
        if text.strip():
            out.append((x0, x1, oy, text))
    return out


def _despliced_body_spans(page, body_x_min, footer_top):
    """Body-text spans rebuilt from CHARACTER geometry (rawdict), each welded
    span split at its column jumps — the normalize-before-detect stage for a
    page whose ordinary extraction shows splice corruption. Returns span dicts
    shaped like the "dict" spans downstream code expects, or None when the
    rawdict layer is unavailable (caller keeps the ordinary extraction)."""
    try:
        blocks = page.get_text("rawdict")["blocks"]
    except Exception:
        return None
    spans = []
    for b in blocks:
        if "lines" not in b:
            continue
        for ln in b["lines"]:
            for sp in ln.get("spans", []):
                chars = [(c["bbox"][0], c["bbox"][2], c.get("c", ""),
                          c.get("origin", (0, c["bbox"][3]))[1])
                         for c in sp.get("chars", [])]
                if not chars:
                    continue
                y0, y1 = sp["bbox"][1], sp["bbox"][3]
                for x0, x1, oy, text in _split_char_runs(chars):
                    syn = {"text": text, "bbox": (x0, y0, x1, y1),
                           "origin": (x0, oy), "size": sp.get("size")}
                    if (text.strip() and x0 >= body_x_min
                            and _span_baseline(syn) <= footer_top):
                        spans.append(syn)
    return spans


def _split_row_columns(spans, gap_min=_COLUMN_GAP_MIN):
    """Split one row's spans into column segments at wide horizontal gaps.

    A pleading caption is two columns — party names left, document title right —
    and both sit on the same printed line. Joining them left-to-right into one
    string welds the two into a phrase that exists nowhere on the page
    ("DREAM TEAM REAL ESTATE" + "BAY EQUITY LENDING, INC."), which no party-name
    term can match. Returns [(x0, text), ...] in left-to-right order."""
    segments, current, prev_x1 = [], [], None
    for sp in sorted(spans, key=lambda s: s["bbox"][0]):
        x0, x1 = sp["bbox"][0], sp["bbox"][2]
        if current and prev_x1 is not None and (x0 - prev_x1) > gap_min:
            segments.append(current)
            current = []
        current.append(sp)
        prev_x1 = max(prev_x1 or x1, x1)
    if current:
        segments.append(current)
    out = []
    for seg in segments:
        text = _join_spans_spaced(seg).strip()
        if text:
            out.append((seg[0]["bbox"][0], text))
    return out


def _detect_line_anchors(page, desplice=False):
    """Per-page: find pleading-paper line numbers and gather body text on
    each numbered row.

    With `desplice=True` the body spans are rebuilt from character-level
    geometry (`_despliced_body_spans`), splitting any span whose glyphs jump
    between columns — the retry path for a page whose ordinary extraction
    shows splice corruption.

    Returns list of {line_num, body_text, segments, y_mid} for each row that has
    both a line number in the gutter and body text in the same horizontal band.
    `segments` is [(x0, text), ...] — the row split at column gutters, so a
    caller can pseudonymize each column as its own text stream. `body_text` is
    the segments joined left-to-right, i.e. the page as it reads on paper.
    Returns [] if the page lacks a recognisable pleading-paper line-number
    column (e.g. exhibits, signature pages, cover sheets without lines).

    One gutter number routinely owns MORE THAN ONE physical row: a caption's
    right column is set at ~11 pt while the gutter runs at 24 pt. Those rows are
    merged column band by column band, each band's rows joined top to bottom.
    Flattening every row's spans into one pool and re-sorting that pool by x —
    what this function used to do — interleaved row 2 into row 1 and produced
    "SCHILLECIJASON P. TORTORICI& TORTORICI,, STATE BARP.C. NO. 207972". A name
    spliced like that matches no pseudonymizer term, so the extraction bug was
    also a disclosure bug.
    """
    from collections import Counter, defaultdict
    blocks = page.get_text("dict")["blocks"]

    # Step 1: find spans whose text is a small integer (1-30) on the left
    # third of the page. These are line-number candidates.
    line_spans = []
    for b in blocks:
        if "lines" not in b:
            continue
        for ln in b["lines"]:
            for sp in ln["spans"]:
                t = sp["text"].strip()
                if (re.fullmatch(r"\d{1,2}", t) and 1 <= int(t) <= 30
                        and sp["bbox"][0] < page.rect.width / 3):
                    line_spans.append({
                        "num": int(t),
                        "y_mid": _span_baseline(sp),
                        "x0": sp["bbox"][0],
                    })
    if not line_spans:
        return []

    # Step 2: cluster by x to find the dominant line-number column.
    x_buckets = Counter(round(s["x0"] / 5) * 5 for s in line_spans)
    dominant_x = x_buckets.most_common(1)[0][0]
    line_col = [s for s in line_spans if abs(s["x0"] - dominant_x) < 8]
    if len(line_col) < 5:
        # Fewer than 5 line numbers: probably a caption page or similar
        # - not a real pleading-paper page.
        return []

    # Step 2a: the page's own line lead. A body row belongs to a line number
    # only if it sits within HALF a lead of it; anything further is furniture.
    # A flat 24 pt tolerance was a whole lead, so the running footer and the
    # printed page number (7-18 pt below line 28) were adopted onto line 28:
    #   "28  (FAC, p. 8:17-18.) 1"
    #   "28  ...a flood on OPPOSITION TO DEFENDANT'S MOTION TO STRIKE..."
    ys = sorted(s["y_mid"] for s in line_col)
    leads = [b - a for a, b in zip(ys, ys[1:]) if b - a > 1]
    tol = min(_GUTTER_MATCH_TOL,
              (statistics.median(leads) if leads else _GUTTER_MATCH_TOL) * 0.5)

    # Step 3: collect body-text spans (right of the line-number column, above
    # the footer band) and cluster them into physical rows by baseline.
    body_x_min = dominant_x + 20
    footer_top = page.rect.height - _FOOTER_MASK_PT
    body_spans = [sp
                  for b in blocks if "lines" in b
                  for ln in b["lines"]
                  for sp in ln["spans"]
                  if sp["text"].strip()
                  and sp["bbox"][0] >= body_x_min
                  and _span_baseline(sp) <= footer_top]
    if desplice:
        cured = _despliced_body_spans(page, body_x_min, footer_top)
        if cured is not None:
            body_spans = cured
    rows = _cluster_rows(body_spans)
    if not rows:
        return []

    # Step 4: assign every row to its nearest gutter line number.
    rows_by_num = defaultdict(list)
    for row in rows:
        best, best_diff = None, float("inf")
        for lnd in line_col:
            diff = abs(row["y"] - lnd["y_mid"])
            if diff < best_diff:
                best, best_diff = lnd, diff
        if best is not None and best_diff <= tol:
            rows_by_num[best["num"]].append(row)

    # Step 5: page-level column bands, seeded from every row's own column split.
    # A band start is a segment x0 further than _COLUMN_GAP_MIN from any other —
    # but a seed only becomes a band when enough DISTINCT rows break at it
    # (_COLUMN_BAND_MIN_ROWS). A caption's true columns run down many rows; a
    # label-value tab gap breaks only its own row or two ("Trial Date:   Nov-
    # ember 16, 2026", "Department:   55"). Treating every wide gap as a page
    # column — the old behaviour — cut those pairs into separate column
    # streams, so the label's stream read "Department:" while the number sat
    # in a phantom band of its own, and the label-anchored department detector
    # never saw the pair in one piece: the REAL department number rode along
    # into the export. An under-supported seed folds into the band on its
    # left and its text re-joins its row with an ordinary space.
    seed_rows = defaultdict(set)          # band-start x -> ids of rows breaking there
    band_starts = []
    for num in rows_by_num:
        for row in rows_by_num[num]:
            for x, _t in _split_row_columns(row["spans"]):
                for bx in band_starts:
                    if abs(x - bx) <= _COLUMN_GAP_MIN:
                        seed_rows[bx].add(id(row))
                        break
                else:
                    band_starts.append(x)
                    seed_rows[x].add(id(row))
    # The floor scales with the page: a wild caption page carries 15-25 body
    # rows (floor 3); a sparse page — a near-empty cover, a signature block —
    # keeps every seed, since "several rows break here" is not evidence a
    # 3-row page can offer.
    n_rows = sum(len(v) for v in rows_by_num.values())
    min_support = min(_COLUMN_BAND_MIN_ROWS, max(1, n_rows // 4))
    band_starts = sorted(bx for bx in band_starts
                         if len(seed_rows[bx]) >= min_support)

    def _band_of(x):
        idx = 0
        for k, bx in enumerate(band_starts):
            if x >= bx - 1.0:
                idx = k
        return idx

    # Step 6: reconstruct each gutter number's lines from its column stacks.
    #
    # A dense letterhead or caption sets several physical rows under ONE gutter
    # number (the block runs at ~half the numbered-body lead). Two shapes hide
    # in that: a caption's two COLUMNS (party name left, case caption right) are
    # one logical line and must read left to right; a letterhead's stacked ROWS
    # in a SINGLE column are sequential lines and must stay on their own lines.
    # Welding everything under the number into one line (the old behaviour)
    # served the first shape but mangled the second — "Paul Green, Esq. (SBN
    # 237707) LAW OFFICE OF PAUL GREEN" on one line.
    #
    # So: bucket the number's row segments into page COLUMNS; within a column
    # keep an ordered stack of its distinct physical rows (same-row pieces
    # joined). The number then emits `depth` lines, where depth is the tallest
    # column's stack and line k pairs the k-th row of every column, left to
    # right. The first line carries the gutter number; deeper lines are
    # continuations (line_num=None) so a pinpoint "p.X:Y" never lands on one.
    results = []
    for num in sorted(rows_by_num):
        stacks = {}                              # band -> {row_y: [x0, text]}
        for row in sorted(rows_by_num[num], key=lambda r: r["y"]):
            for x, text in _split_row_columns(row["spans"]):
                if _CAPTION_DIVIDER_RE.match(text):
                    continue
                text = _CAPTION_DIVIDER_LEAD_RE.sub("", text)
                if not text:
                    continue
                col = stacks.setdefault(_band_of(x), {})
                if row["y"] in col:              # another piece of the same row
                    col[row["y"]][0] = min(col[row["y"]][0], x)
                    col[row["y"]][1] += " " + text
                else:
                    col[row["y"]] = [x, text]
        if not stacks:
            continue
        bands = sorted(stacks)
        depth = max(len(col) for col in stacks.values())
        for k in range(depth):
            segs = []
            for b in bands:
                items = sorted(stacks[b].items())     # by row_y, top to bottom
                if k < len(items):
                    row_y, (x, text) = items[k]
                    segs.append((x, row_y, text))
            if not segs:
                continue
            segs.sort()                               # left to right by x0
            results.append({
                "line_num": num if k == 0 else None,
                "body_text": " ".join(t for _x, _y, t in segs),
                "segments": [(x, t) for x, _y, t in segs],
                "y_mid": min(y for _x, y, _t in segs),
            })
    return results


def _annotate_paragraphs(anchors):
    """For each line, detect numbered-paragraph starts (1., 2., 3., ...) and
    forward-fill the active paragraph number to subsequent lines on the page.

    A page is treated as declaration-style only if it contains AT LEAST 3
    sequential paragraph starts (e.g. 1., 2., 3., or 4., 5., 6.). Pages with
    fewer paragraph starts (or non-sequential ones) get paragraph_num = None
    on every line — this avoids tagging an isolated "7." inside body text on
    a page that isn't actually a declaration body page.

    For declaration-style pages, every line within paragraph N carries
    paragraph_num = N (not just the line where N starts), so a paste from
    the middle of a paragraph still surfaces the paragraph reference for
    citation purposes.

    The ``starts_paragraph`` flag is also added: True on the line where the
    paragraph starts, False on continuation lines.
    """
    # First pass: find candidate paragraph-start lines and strip the leading
    # "N. " prefix into a separate field. Heuristic guard: only treat N. as
    # a paragraph start if 1 <= N <= 999 AND the next char is uppercase
    # (filters out "7." appearing mid-sentence). A long complaint runs well
    # past 99 paragraphs, so the cap must reach triple digits — the earlier
    # 99 ceiling silently dropped every paragraph from 100 on, and a page of
    # ¶¶ 100-103 then had zero candidates and produced no bookmarks.
    candidates = []  # list of (anchor_index, paragraph_num)
    for i, a in enumerate(anchors):
        m = re.match(r"^(\d{1,3})\.\s*(.*)$", a["body_text"])
        if m:
            num = int(m.group(1))
            rest = m.group(2)
            if 1 <= num <= 999 and rest and rest[0].isupper():
                candidates.append((i, num, rest))

    # Determine whether this page qualifies as declaration-style: need 3+
    # paragraph starts in strictly increasing sequence (need not be
    # contiguous integers — page might start mid-doc at 4., 5., 6. — but
    # must be monotonically increasing).
    qualifies = False
    if len(candidates) >= 3:
        nums = [c[1] for c in candidates]
        if all(nums[k] < nums[k + 1] for k in range(len(nums) - 1)):
            qualifies = True

    # Initialise all anchors with no paragraph context.
    for a in anchors:
        a["paragraph_num"] = None
        a["starts_paragraph"] = False

    if not qualifies:
        return anchors

    # Apply paragraph numbers: at each candidate index, strip the leading
    # "N." from body_text and set starts_paragraph; then forward-fill
    # paragraph_num to every subsequent anchor until the next candidate.
    candidate_indices = {c[0]: (c[1], c[2]) for c in candidates}
    active = None
    for i, a in enumerate(anchors):
        if i in candidate_indices:
            num, rest = candidate_indices[i]
            a["paragraph_num"] = num
            a["body_text"] = rest
            a["starts_paragraph"] = True
            active = num
        elif active is not None:
            a["paragraph_num"] = active
    return anchors


# ────────────────────────────────────────────────────────────────────────────
# Legacy right-margin marker detection
# ────────────────────────────────────────────────────────────────────────────
# pdf_linker used to stamp invisible white "[ShortName|p3:7¶4]" markers into
# the right margin of pleading pages (to feed a Word paste macro). That
# watermarking has been removed. This pattern is kept only so that
# re-processing a PDF stamped by an older version still recognises and strips
# those markers - see _collect_rows (row detection) and _write_text_version
# (clean text export). It matches the full form "[ShortName|p3:7¶4]" and the
# compact form "[p3:7¶4]"; the paragraph suffix is optional.
_MARKER_DETECT_RE = re.compile(
    r"\[(?:[A-Za-z0-9_.&'\- ]+\|)?p\d+:\d+(?:\u00b6\d+)?\]"
)


def _detect_paragraph_anchors(doc):
    """Return [(page_index, paragraph_num), ...] - one entry per page on which
    at least one *new* numbered paragraph starts, recording the lowest such
    number. Feeds the bookmark builder's "Paragraphs" branch. Documents
    without pleading-paper line numbers (most briefs) yield no anchors.

    This used to also inject invisible right-margin citation markers; that
    watermarking has been removed. Only the paragraph-anchor detection, which
    drives bookmarks and never modifies the page, remains.
    """
    paragraph_anchors = []
    for page in doc:
        anchors = _detect_line_anchors(page)
        if not anchors:
            continue  # caption / exhibit / signature page
        anchors = _annotate_paragraphs(anchors)
        # Lowest paragraph number that *starts* on this page, if any.
        # Continuation pages (every line mid-paragraph) produce no entry, so a
        # paragraph spanning three pages adds one bookmark, not three.
        starts = [a.get("paragraph_num") for a in anchors
                  if a.get("starts_paragraph") and a.get("paragraph_num")]
        if starts:
            paragraph_anchors.append((page.number, min(starts)))
    return paragraph_anchors


# ────────────────────────────────────────────────────────────────────────────
# Safe citation-text search (multi-line aware, no spurious wide matches)
# ────────────────────────────────────────────────────────────────────────────
# Minimum length for a search-for fragment to be considered "safe" — anything
# shorter risks matching tons of unrelated locations. "Cal." (4 chars) is the
# canonical example of an unsafe short fragment.
_MIN_FRAG_LEN = 12

# Generic substrings that, even if longer than _MIN_FRAG_LEN, appear too often
# in legal documents to be safe link targets on their own. We require a
# fragment to contain something *more* than these to be linkable.
# The patterns capture: California reporter abbreviations alone or with an
# ordinal ("Cal. App. 4th"), plus generic connector words. A safe fragment
# must contain content beyond these.
_GENERIC_FRAG_RE = re.compile(
    r"^(?:"
    r"Cal\.(?:\s*(?:App|Rptr))?\.?(?:\s*\d?(?:st|nd|rd|th|d))?"
    r"|U\.S\."
    r"|F\.\s*Supp\.(?:\s*\d?d)?"
    r"|F\.\s*\d?d"
    r"|the(?:\s+\w+)?"
    r"|see(?:\s+\w+)?"
    r"|in\s+re"
    r")\s*\d*\s*$",
    re.IGNORECASE,
)


def _is_safe_fragment(s: str) -> bool:
    """Return True if `s` is distinctive enough to use as a search target.

    A fragment is unsafe if it's too short to be unique, or if it consists
    only of reporter abbreviations and generic words. Both conditions
    matter: "Cal. App. 4th" is 13 chars (long enough) but matches every
    case in a TOA. "Smith" is distinctive but only 5 chars — likely too
    common as a name. We pair the length check with a generic-pattern check.
    """
    s = s.strip()
    if len(s) < _MIN_FRAG_LEN:
        return False
    if _GENERIC_FRAG_RE.match(s):
        return False
    return True


def _merge_quads_by_line(quads, *, require_x_adjacent: bool = False):
    """Merge a list of fitz.Quad into one quad per text-line.

    PyMuPDF's `Page.search_for` can return multiple per-word quads for a
    single match:
      - On phrases containing `\\n`, justified-text spacing sometimes
        causes per-word splitting on a line (e.g., "Intengan v. BAC
        Home Loans..." → one quad per word).
      - Even on single-line phrases, superscript glyphs like the "th"
        in "Cal.4th" come back as a separate quad with a slightly
        different y-range and font.

    Drawing an underline per quad makes one citation look like several
    disjoint underlined fragments. This helper groups quads by y-band
    (same logical line) and emits one merged quad per line, spanning
    min(x0) to max(x1).

    Args:
      quads: input quads from `page.search_for`.
      require_x_adjacent: when True, only merge quads on the same line
        if they are horizontally contiguous (x-gap < ~5pt). This is the
        safer mode for single-line searches, where multiple separate
        occurrences of a short phrase on one line must remain distinct.
        For multi-line searches, the same-line word-splits are always
        contiguous parts of one match, so this flag can be left False.
    """
    if not quads:
        return quads
    import fitz
    LINE_TOL = 4.0    # pt — quads on the same line are within ~4pt vertically
    X_GAP_TOL = 6.0   # pt — horizontally adjacent if x-gap is at most this

    # Sort by y-center then x to make line-grouping a linear scan.
    sorted_q = sorted(quads, key=lambda q: ((q.rect.y0 + q.rect.y1) / 2,
                                            q.rect.x0))
    groups = []  # list of list[Quad]
    for q in sorted_q:
        y_mid = (q.rect.y0 + q.rect.y1) / 2
        if groups:
            last_grp = groups[-1]
            last_y = (last_grp[0].rect.y0 + last_grp[0].rect.y1) / 2
            same_line = abs(y_mid - last_y) <= LINE_TOL
            if same_line and require_x_adjacent:
                # Also require horizontal adjacency to the previous quad
                # in this group — guards against two separate occurrences
                # of a short phrase on the same physical line.
                prev_x1 = max(qq.rect.x1 for qq in last_grp)
                if q.rect.x0 - prev_x1 > X_GAP_TOL:
                    same_line = False
            if same_line:
                last_grp.append(q)
                continue
        groups.append([q])

    merged = []
    for grp in groups:
        x0 = min(q.rect.x0 for q in grp)
        y0 = min(q.rect.y0 for q in grp)
        x1 = max(q.rect.x1 for q in grp)
        y1 = max(q.rect.y1 for q in grp)
        merged.append(fitz.Rect(x0, y0, x1, y1).quad)
    return merged


def _statute_section_token(cite):
    """Return the disambiguating section token for a statute cite, else None.

    For a statute the code name ("Code of Civil Procedure") is shared by
    every section of that code, so it is NOT distinctive — the section
    number is. Callers use this token to refuse anchoring a link on a bare
    code-name fragment. Returns the leading numeric part of the section
    (e.g. "430.10" from key "CCP § 430.10(a)", "1010.6" from
    "CCP § 1010.6", "1983" from "42 U.S.C. § 1983"); None for non-statutes.
    """
    if not cite or cite.get("kind") != "statute":
        return None
    m = re.search(r"§\s*(\d+(?:\.\d+)?)", cite.get("key", ""))
    return m.group(1) if m else None


def _safe_search_for_citation(page, match_text: str, cite=None):
    """Locate the citation `match_text` on `page`, handling multi-line wraps
    safely. Returns a list of fitz.Quad covering the citation's glyphs, or
    an empty list if not safely locatable.

    Strategy:
      1. Try the full match_text. If found, return its quads.
      2. If match_text spans multiple lines (contains \\n), split into
         fragments at line breaks. For each fragment that passes
         `_is_safe_fragment`, search the page. We require fragments to
         appear in vertically-adjacent positions on the page (within ~3
         lines of each other) so a TOA "Smith" plus an unrelated body
         "Smith" don't get linked together.
      3. If no fragment is safe, return [] — better to miss a link than to
         spray hyperlinks across unrelated text.

    `cite` (optional) lets statute citations be matched safely: the code
    name alone is not distinctive (every section shares it), so when a
    wrapped statute citation falls back to fragments, only a fragment that
    carries the section number may anchor the link.
    """
    section_token = _statute_section_token(cite)
    # Step 1: full text search
    quads = page.search_for(match_text, quads=True)
    if quads:
        # PyMuPDF may return multiple per-word/per-glyph quads for a
        # single match — driven by justified-text spacing (multi-line)
        # or superscripts like the "th" in "Cal.4th" (single-line).
        # Merging them produces a single rect per text-line so the
        # underline draws as one continuous span. For single-line
        # searches we additionally require x-adjacency so two distinct
        # occurrences of the same short phrase on one line don't get
        # accidentally fused into one mega-rect.
        is_multiline = "\n" in match_text
        quads = _merge_quads_by_line(quads,
                                     require_x_adjacent=not is_multiline)
        return quads

    # Step 2: multi-line splitting
    if "\n" not in match_text:
        return []
    fragments = [f.strip() for f in match_text.splitlines() if f.strip()]
    safe_frags = [f for f in fragments if _is_safe_fragment(f)]
    if not safe_frags:
        return []

    # Search each safe fragment, then group adjacent results.
    fragment_hits = []  # list of (fragment, quads)
    for frag in safe_frags:
        fq = page.search_for(frag, quads=True)
        if fq:
            fragment_hits.append((frag, fq))
    if not fragment_hits:
        return []

    # Statute guard: a section-bearing fragment must actually be present on
    # this page. If only code-name fragments were found, the citation isn't
    # really here (just a mention of the same code for a different section),
    # so linking would point at the wrong section. Covers both the single-
    # and multi-fragment branches below (e.g. a code name that wraps across
    # several lines while the section sits on its own short line).
    if section_token is not None and not any(
        section_token in f for f, _ in fragment_hits
    ):
        return []

    # If only one fragment is safely findable, decide whether it is
    # distinctive enough to anchor the link by itself.
    #
    # For a STATUTE the code name is never distinctive on its own. A
    # citation that wraps a line — e.g. "Code of Civil Procedure\nsection
    # 430.10" — splits into a code-name fragment ("Code of Civil Procedure")
    # and a section fragment ("section 430.10"). On a page that cites a
    # *different* section of the same code ("Code of Civil Procedure
    # § 1161(2)", or "...§ 1010.6"), only the generic code-name fragment is
    # present. Anchoring on it would attach THIS citation's URL (§ 430.10) to
    # a span that is really a different citation, so the reader clicks
    # "...§ 1161(2)" and lands on § 430.10. We therefore require the lone
    # fragment to carry the section number; a bare code name links nothing,
    # and the citation is picked up wherever its section actually appears
    # (via the full-text search in step 1 or the multi-fragment branch).
    #
    # For non-statutes the original rule stands: the fragment must be
    # >= 20 chars and identify exactly one location on the page.
    if len(fragment_hits) == 1:
        frag, fq = fragment_hits[0]
        if len(page.search_for(frag)) != 1:
            return []
        if section_token is not None:
            return fq if section_token in frag else []
        if len(frag) >= 20:
            return fq
        return []

    # Multiple fragments found: only keep quads from each fragment that
    # have a peer fragment-quad within 3 line-heights. This rejects cases
    # where one fragment matches in an unrelated place on the page.
    LINE_H_TOL = 60  # ~3 lines at 20pt
    out_quads = []
    for i, (_, fq) in enumerate(fragment_hits):
        for q in fq:
            y_mid = (q.rect.y0 + q.rect.y1) / 2
            paired = False
            for j, (_, fq2) in enumerate(fragment_hits):
                if j == i:
                    continue
                for q2 in fq2:
                    y2 = (q2.rect.y0 + q2.rect.y1) / 2
                    if abs(y_mid - y2) <= LINE_H_TOL:
                        paired = True
                        break
                if paired:
                    break
            if paired:
                out_quads.append(q)
    return out_quads


# ────────────────────────────────────────────────────────────────────────────
# Short-form case linking (second pass)
# ────────────────────────────────────────────────────────────────────────────
# Match a bare "X v. Y" (or "X v. Y, Inc." etc.) — no reporter, no year.
# Used in the second pass after full citations are linked, to also link
# subsequent shorter mentions of cases already cited in long form.
# Plaintiff: must start uppercase, allow internal letters/digits/'-./&.
# Defendant: same, plus permit a comma and a corporate suffix like ", Inc."
# We deliberately keep this conservative — only sequences that look like
# party-v-party, with no surrounding reporter/year, qualify. Limited to
# 1-4 tokens on each side to avoid runaway matches.
_PARTY_TOKEN = r"[A-Z][A-Za-z0-9.\-'&]*"
_SHORT_FORM_RE = re.compile(
    rf"\b({_PARTY_TOKEN}(?:\s+{_PARTY_TOKEN}){{0,3}})\s+v\.\s+"
    rf"({_PARTY_TOKEN}(?:\s+{_PARTY_TOKEN}){{0,4}}(?:,\s*(?:Inc|LLC|LLP|Ltd|Corp|Co)\.?)?)"
)

# Leading words to strip from a short-form plaintiff capture. These are
# common when a brief introduces a case — "In Smith v. Jones, ..." or
# "See Smith v. Jones, ..." — and would otherwise pollute the registry
# lookup. Mirrors the SIGNAL_PREFIXES set used by `_walk_back_for_name`.
_SHORTFORM_LEAD_RE = re.compile(
    r"^(?:In|See|Cf|Cf\.|Compare|Accord|But|Following|"
    r"Per|Under|Like|Citing|Quoting)\s+",
    re.IGNORECASE,
)


def _normalize_party(s: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    s = re.sub(r"[.,;:'\"]", "", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def _link_short_form_cases(doc, full_cites,
                           provider: str, log: logging.Logger) -> int:
    """For each full case citation already matched, also link bare
    "X v. Y" mentions of the same case elsewhere in the document.

    Returns the number of additional citation occurrences linked.
    """
    try:
        import fitz
    except ImportError:
        return 0

    # Build a registry of (plaintiff_norm, defendant_norm) -> URL from
    # full citations. We extract plaintiff/defendant from the case key
    # ("Plaintiff v. Defendant (year) ..." or "... 2023 WL ...").
    registry = {}  # (plaintiff_norm, defendant_norm) -> (key, url, full_match)
    # NB: the word boundary belongs INSIDE the WL branch only. Written after
    # the alternation ("…|\d{4}\s+WL)\b"), it also applied to the "(year)"
    # branch — and ")" followed by a space has no \b, so no published-reporter
    # key ever matched and short-form linking silently ran for WL cites alone.
    case_key_re = re.compile(r"^(.+?)\s+v\.\s+(.+?)\s+(?:\(\d{4}\)|\d{4}\s+WL\b)")
    for c in full_cites:
        if c.get("kind") != "case":
            continue
        url = resolve_url(c, provider)
        if not url:
            continue
        m = case_key_re.match(c["key"])
        if not m:
            continue
        p_norm = _normalize_party(m.group(1))
        d_norm = _normalize_party(m.group(2))
        if not p_norm or not d_norm:
            continue
        # First-seen wins (later identical references resolve to same URL anyway)
        registry.setdefault((p_norm, d_norm), (c["key"], url, c["match_text"]))

    if not registry:
        return 0

    extra_links = 0
    # Track spans we've already annotated (per page) to avoid double-linking
    # a span that was already linked in the main pass.
    for page in doc:
        text = page.get_text("text")
        # Find all bare X v. Y patterns on this page
        for m in _SHORT_FORM_RE.finditer(text):
            plaintiff_raw = m.group(1).strip()
            defendant = m.group(2).strip()
            # Strip leading signal/connector words ("In", "See", "Following"
            # etc.) so the normalized plaintiff matches the registry.
            plaintiff = _SHORTFORM_LEAD_RE.sub("", plaintiff_raw).strip()
            if not plaintiff:
                continue
            p_norm = _normalize_party(plaintiff)
            d_norm = _normalize_party(defendant)

            # Try exact (plaintiff, defendant) match first; then a relaxed
            # match where the short defendant is a prefix of the registered
            # one (e.g. short "Ford" matches registered "Ford Motor Co.").
            url = None
            if (p_norm, d_norm) in registry:
                _, url, _full = registry[(p_norm, d_norm)]
            else:
                # Prefix match on WORD boundaries: short "ford" matches
                # registered "ford motor co", but "jones" must not match
                # "jonesboro" (a char-level prefix did).
                for (rp, rd), (_k, ru, _ft) in registry.items():
                    if rp == p_norm and (
                            rd == d_norm
                            or rd.startswith(d_norm + " ")
                            or d_norm.startswith(rd + " ")):
                        url = ru
                        break
            if not url:
                continue

            # Locate this bare reference on the page. We use the cleaned
            # case-name portion (without any leading "In"/"See"/etc.) so
            # the highlight covers only "Plaintiff v. Defendant", not the
            # connecting word that precedes it.
            snippet = f"{plaintiff} v. {defendant}"
            quads = page.search_for(snippet, quads=True)
            if not quads:
                continue

            # If the matched span overlaps a span already covered by a full
            # citation's annotations, skip — the full citation's link is
            # already there. We approximate "already covered" by checking
            # whether the rect overlaps any existing link annotation.
            existing_links = page.get_links()
            for q in quads:
                rect = q.rect
                already = False
                for el in existing_links:
                    er = el.get("from")
                    if er and rect.intersects(er):
                        # Substantial overlap means the span is already linked
                        # (probably as part of the full citation).
                        already = True
                        break
                if already:
                    continue
                if rect.x0 > page.rect.width - 90:
                    continue
                page.insert_link({
                    "kind": fitz.LINK_URI,
                    "from": rect,
                    "uri": url,
                })
                underline_y = rect.y1 - 0.5
                page.draw_line(
                    fitz.Point(rect.x0, underline_y),
                    fitz.Point(rect.x1, underline_y),
                    color=LINK_COLOUR,
                    width=1.0,
                )
                extra_links += 1

    if extra_links:
        log.info(f"  Linked {extra_links} short-form case reference(s)")
    return extra_links


# ────────────────────────────────────────────────────────────────────────────
# Table-of-Contents internal linking
# ────────────────────────────────────────────────────────────────────────────
# Detect a "Table of Contents" heading in the front matter and turn each
# entry into an internal jump to the page bearing that printed page number.
#
# Scope:
#   * Only the first TOC_SCAN_PAGES pages are searched for the *heading*.
#     This is a search budget to keep us out of the body of long briefs that
#     might mention "table of contents" in passing. Once the heading is
#     found, the TOC itself is walked to its natural end (terminator
#     heading or empty page) regardless of how many pages it spans.
#   * The "Table of Authorities" page is intentionally NOT processed here —
#     its entries are case/statute citations and the main citation-linking
#     pass already turns those into external Westlaw/Lexis links. A TOA
#     heading appearing after the TOC also acts as a TOC terminator.
#
# Page-number mapping:
#   The number printed on a TOC line ("3", "iii") is the page label the
#   author chose, not the PDF's underlying page index. A 2-page TOC after
#   a cover sheet means TOC's "1" might be PDF page index 3. We resolve
#   this by scanning each PDF page for a stamped page number in the top
#   or bottom margin band, building a {printed -> pdf_index} map, and
#   resolving each TOC entry through that map. Entries whose printed
#   number can't be found are left unlinked (better no link than wrong).
#
# TOC_SCAN_PAGES is the budget for *finding the heading*. Pleading front
# matter (multi-page Notice of Motion/Demurrer + caption) routinely pushes
# the TOC to page 7 or later, so the budget is generous; _find_toc_heading's
# standalone-row requirement is what keeps the wider scan from false-matching
# body prose.
TOC_SCAN_PAGES = 20

_TOC_HEADING_RE = re.compile(
    r"\bT\s*A\s*B\s*L\s*E\s+O\s*F\s+C\s*O\s*N\s*T\s*E\s*N\s*T\s*S\b",
    re.IGNORECASE,
)

# Headings that end the TOC. Matching any of these on a line stops further
# entry collection on that page and prevents continuation onto later pages.
_TOC_END_RE = re.compile(
    r"\b("
    r"T\s*A\s*B\s*L\s*E\s+O\s*F\s+A\s*U\s*T\s*H\s*O\s*R\s*I\s*T\s*I\s*E\s*S"
    r"|INTRODUCTION"
    r"|STATEMENT\s+OF\s+(?:FACTS|THE\s+CASE)"
    r"|MEMORANDUM\s+OF\s+POINTS"
    r"|ARGUMENT(?!\s+(?:I|II|III|IV|V|\d))"   # bare "ARGUMENT" as a top-level section,
    r")\b",                                    # not a numbered subheading inside the TOC
    re.IGNORECASE,
)

# A TOC entry line: some label text, then dot leaders or wide whitespace,
# then a trailing page number (arabic or roman). The label is kept
# non-greedy so the longest trailing token is the page number.
#
# The leader class accepts ASCII periods AND the Unicode dot-leader / ellipsis
# characters Word and many PDF generators emit instead of literal periods:
# U+2024 ONE DOT LEADER, U+2025 TWO DOT LEADER, U+2026 HORIZONTAL ELLIPSIS.
# Without these, a TOC whose leaders render as "……………" (the common case)
# never matches and no entry is linked.
#
# Two leader shapes are allowed: 2+ characters of dots/ellipses/whitespace
# (the usual long leader run, or wide whitespace alignment), OR a SINGLE
# dot-leader/ellipsis character. The single-character case matters when a
# long label nearly fills the line and only one "…" fits before the page
# number ("…Personal Characteristics…8"); a lone ellipsis is still an
# unambiguous leader (one "…" renders as three dots). Whitespace still needs
# 2+ — a single space would match ordinary "word number" prose.
_TOC_ENTRY_RE = re.compile(
    r"""
    ^
    (?P<label>\S.*?\S)
    (?: [\s.․‥…]{2,} | [.․‥…]\s* )
    (?P<page>
        \d{1,4}
      | [ivxlcdm]{1,6}
      | [IVXLCDM]{1,6}
    )
    \s*$
    """,
    re.VERBOSE,
)

# A page-number stamp: a line that is *only* a page number (with optional
# decorations). Used to find the printed page label of each PDF page.
_PAGE_STAMP_RE = re.compile(
    r"""
    ^
    \s*
    (?:Page\s+|-\s*|—\s*)?
    # Roman labels must be GRAMMATICALLY valid numerals, not just roman
    # letters: "CIVIL" is five roman letters, so a court form's footer word
    # was read as a page-number stamp ("p.CIVIL" in the leak report's Where
    # column, and a wrong printed-page map for TOC links).
    (?P<num>\d{1,4}
        | (?=[ivxlcdm]) m{0,3}(?:cm|cd|d?c{0,3})(?:xc|xl|l?x{0,3})(?:ix|iv|v?i{0,3})
        | (?=[IVXLCDM]) M{0,3}(?:CM|CD|D?C{0,3})(?:XC|XL|L?X{0,3})(?:IX|IV|V?I{0,3})
    )
    \s*
    (?:\s*-|\s*—|\.|\s+of\s+\d+)?
    \s*
    $
    """,
    re.VERBOSE,
)

# Margin band (in PDF points) within which a line qualifies as a page-number
# stamp. ~100pt covers the typical footer zone on pleading paper, where the
# printed page number often sits about an inch above the page edge — outside
# a strict 1-inch band but well above the bottom margin. The band is just a
# coarse filter; the candidate-ranking logic below disambiguates among
# multiple hits within the band.
_STAMP_BAND_PT = 100.0


def _collect_rows(page, y_tol: float = 4.0, drop_gutter_numbers: bool = False):
    """Return [(text, bbox)] for the page, grouping text runs that share a
    baseline. ReportLab and most pleading-paper layouts draw a TOC entry's
    label, dot leaders, and page number as separate text runs, which
    PyMuPDF returns as separate `lines` despite their visual co-location
    on one row. We merge runs whose y-center differs by no more than
    `y_tol` points and concatenate left-to-right with single spaces.

    The returned bbox is the union of the merged runs, which is what the
    TOC linker uses as the clickable rectangle.

    Marker runs (the invisible right-margin `[Demurrer|p1:2]`-style
    tags pdf_linker writes into its own output) are filtered out before
    grouping. Without this filter, a TOC entry like
    "TABLE OF AUTHORITIES ... ii" merges with the adjacent marker into
    "TABLE OF AUTHORITIES ... ii [Demurrer|p5:3]", which then fails
    `_TOC_ENTRY_RE`'s end-of-line page-number requirement — silently
    dropping every TOC entry on re-runs of an already-linked PDF.

    `drop_gutter_numbers` removes pleading-paper gutter line-numbers (the
    "1..28" column down the left edge) before grouping. On a TOC page these
    share a baseline with each entry and would otherwise merge into the row
    ("2 I. INTRODUCTION … 1"), polluting the entry label and stretching the
    clickable rectangle across the left margin. Used by the TOC linker;
    left off elsewhere so no other caller's behaviour changes.
    """
    try:
        d = page.get_text("dict")
    except Exception:
        return []
    gutter_x_max = page.rect.width * 0.18
    runs = []  # (y_center, bbox, text)
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            bbox = line.get("bbox")
            if not bbox:
                continue
            text = "".join(
                span.get("text", "") for span in line.get("spans", [])
            ).strip()
            if not text:
                continue
            # Skip pdf_linker's own right-margin markers; otherwise they
            # contaminate merged rows on re-runs of an already-linked PDF.
            if _MARKER_DETECT_RE.fullmatch(text):
                continue
            # Skip a left-margin gutter line-number when requested.
            if (drop_gutter_numbers and bbox[0] < gutter_x_max
                    and re.fullmatch(r"\d{1,2}", text) and 1 <= int(text) <= 40):
                continue
            y_center = (bbox[1] + bbox[3]) / 2
            runs.append((y_center, bbox, text))
    if not runs:
        return []
    runs.sort(key=lambda r: (r[0], r[1][0]))
    rows = []
    current = [runs[0]]
    for r in runs[1:]:
        if abs(r[0] - current[0][0]) <= y_tol:
            current.append(r)
        else:
            rows.append(current)
            current = [r]
    rows.append(current)
    merged = []
    for g in rows:
        g.sort(key=lambda r: r[1][0])
        text = " ".join(r[2] for r in g)
        x0 = min(r[1][0] for r in g)
        y0 = min(r[1][1] for r in g)
        x1 = max(r[1][2] for r in g)
        y1 = max(r[1][3] for r in g)
        merged.append((text, (x0, y0, x1, y1)))
    return merged


def _find_toc_heading(doc, max_pages: int):
    """Return the PDF page index whose front matter carries a standalone
    'Table of Contents' heading, or None.

    The heading is required to be (essentially) a row of its own — after
    stripping a leading pleading-paper gutter line number it must match the
    heading pattern with nothing else on the line. That standalone
    requirement is what makes it safe to scan well past the cover/notice
    pages (a long Notice of Demurrer can push the TOC to page 7+), without
    false-matching a body sentence that merely mentions a "table of
    contents". `max_pages` bounds the scan so we never walk an entire
    compendium looking for front matter.
    """
    limit = min(max_pages, len(doc))
    for i in range(limit):
        for text, _ in _collect_rows(doc[i], drop_gutter_numbers=True):
            m = _TOC_HEADING_RE.match(text.strip())
            if m and not text.strip()[m.end():].strip():
                return i
    return None


def _build_printed_page_map(doc) -> dict:
    """Scan every page for a printed page-number stamp in the top or bottom
    margin band. Returns {printed_label_lower: [pdf_page_index, ...]}.

    The value is a list (not a single index) because reset numbering is
    common in briefs: the cover and notice section often carries arabic
    numbers 1-4, then TOA/TOC switches to roman i-iii, then the MPA
    restarts arabic from 1. A TOC entry "Introduction ... 1" targets the
    MPA's page 1, not the cover's page 1. Callers resolve the ambiguity
    with `_resolve_target`, which picks the first occurrence strictly
    after the TOC entry's own page (TOC entries always point forward).

    Algorithm per page:
      1. Walk merged rows from `_collect_rows`. A row whose entire text
         matches `_PAGE_STAMP_RE` is a candidate.
      2. Additionally scan per-line spans from `get_text("dict")` for
         span-level matches. This handles pleading paper where the
         left-margin line-number column (e.g., "28") shares a baseline
         with the centered page-number stamp (e.g., "i"); the row merger
         fuses them into "28 i", which no longer matches the stamp
         regex. The per-span pass recovers the legitimate stamp.
      3. Discard candidates whose horizontal center is far from page
         center (line-number-column matches). The legitimate stamp is
         essentially always centered or near-centered.
      4. Among remaining candidates, prefer (a) the bottom-most span
         (typical pleading footer), then (b) the span closest to
         horizontal page-center.
    """
    # Maximum horizontal distance from page center for a stamp candidate
    # to be considered legitimate. A real centered page-number stamp on
    # a letter-sized page sits within ~10pt of center; allow generous
    # slack for slightly off-center stamps and right-aligned footer
    # numbering. The pleading-paper line-number column sits 200+ pt
    # off-center and is reliably excluded.
    CENTER_TOL_PT = 120.0

    mapping: dict = {}  # label_lower -> [pdf_page_idx, ...] in document order
    for i, page in enumerate(doc):
        page_height = page.rect.height
        page_center_x = page.rect.width / 2
        candidates = []  # (y0, num_str, x_center)

        def _consider(text: str, bbox):
            y0, y1 = bbox[1], bbox[3]
            in_top = y1 <= _STAMP_BAND_PT
            in_bot = y0 >= page_height - _STAMP_BAND_PT
            if not (in_top or in_bot):
                return
            m = _PAGE_STAMP_RE.match(text)
            if not m:
                return
            x_center = (bbox[0] + bbox[2]) / 2
            if abs(x_center - page_center_x) > CENTER_TOL_PT:
                # Far from center: almost certainly a left-margin
                # line-number column entry, not a page stamp.
                return
            candidates.append((y0, m.group("num"), x_center))

        # Pass 1: whole-row matches via _collect_rows.
        for line_text, bbox in _collect_rows(page):
            _consider(line_text, bbox)

        # Pass 2: per-span fallback for column-collision rows. When the
        # line-number column shares a baseline with the centered page
        # stamp, _collect_rows merges them ("28 i") and Pass 1 misses
        # the stamp. Spans are individual atoms, so the stamp survives.
        try:
            raw = page.get_text("dict")
        except Exception:
            raw = {"blocks": []}
        for block in raw.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span.get("text", "").strip()
                    if not txt:
                        continue
                    bbox = span.get("bbox")
                    if not bbox:
                        continue
                    _consider(txt, bbox)

        if not candidates:
            continue

        # Rank: bottom-most first (footers are the typical stamp
        # location), then closest to page center.
        candidates.sort(key=lambda c: (-c[0], abs(c[2] - page_center_x)))
        num = candidates[0][1]
        key = num.lower()
        mapping.setdefault(key, []).append(i)
    return mapping


def _footer_page_label(page, center_tol: float = 120.0):
    """Return the page number *printed* on this page (the "footer page
    number"), as a raw label string ("3", "ii"), or None if no centred
    page-number stamp is found.

    This is the per-page inverse of what `_build_printed_page_map`
    collects across the whole document: rather than {label -> [pages]},
    it answers "what page number does a reader see on THIS page?".
    Right-margin citation markers use it so a pasted pinpoint reference
    carries the page number a court actually cites — the printed page —
    instead of the PDF's physical page index.

    Detection mirrors `_build_printed_page_map` exactly (kept as a small
    standalone copy so the proven TOC-linking path is left untouched):
    scan merged rows and raw spans in the top/bottom margin band for a
    token that is *only* a page number (`_PAGE_STAMP_RE`); discard
    candidates far from horizontal centre (pleading-paper line-number
    column digits, right-margin Bates stamps); then prefer the
    bottom-most, most-centred candidate — i.e. the footer.

    Why per-page (not a running offset): in compiled filings — compendia,
    declaration bundles — page numbering restarts at every sub-document.
    Each declaration begins again at its own page 1/2/3, and a new
    exhibit typically switches scheme entirely (Bates numbers, "Page X of
    Y") or restarts. Reading each page's own stamp absorbs those resets
    for free: no counter is carried across pages, so a declaration or
    exhibit boundary can never desynchronise the count.

    The raw label is returned verbatim; the caller decides whether to
    accept non-arabic labels (the marker format and the Word macro expect
    arabic digits, so roman front-matter labels are declined upstream).
    """
    page_height = page.rect.height
    page_center_x = page.rect.width / 2
    candidates = []  # (y0, num_str, x_center)

    def _consider(text: str, bbox):
        y0, y1 = bbox[1], bbox[3]
        in_top = y1 <= _STAMP_BAND_PT
        in_bot = y0 >= page_height - _STAMP_BAND_PT
        if not (in_top or in_bot):
            return
        m = _PAGE_STAMP_RE.match(text)
        if not m:
            return
        x_center = (bbox[0] + bbox[2]) / 2
        if abs(x_center - page_center_x) > center_tol:
            return
        candidates.append((y0, m.group("num"), x_center))

    # Pass 1: whole-row matches (label + leaders fused rows are harmless
    # here since a pure stamp row is just the number).
    for line_text, bbox in _collect_rows(page):
        _consider(line_text, bbox)

    # Pass 2: per-span fallback for column-collision rows, where the
    # left-margin line-number digit shares a baseline with the centred
    # stamp and the row merger fuses them ("28 3").
    try:
        raw = page.get_text("dict")
    except Exception:
        raw = {"blocks": []}
    for block in raw.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                txt = span.get("text", "").strip()
                bbox = span.get("bbox")
                if txt and bbox:
                    _consider(txt, bbox)

    if not candidates:
        return None
    # Bottom-most first (footers), then closest to horizontal centre.
    candidates.sort(key=lambda c: (-c[0], abs(c[2] - page_center_x)))
    return candidates[0][1]


def _resolve_target(printed_map: dict, page_label: str, after_idx: int):
    """Resolve a printed page label (e.g. "1", "ii") to a PDF page index.

    `after_idx` is the index of the page containing the reference (a TOC
    entry's own page). TOC entries always point forward, so we return
    the first occurrence strictly after `after_idx`. If the label has no
    forward occurrence, we fall back to the first occurrence anywhere —
    a reasonable last resort, though this case is unusual and may
    indicate a malformed TOC.

    Returns None if the label is unknown.
    """
    occurrences = printed_map.get(page_label.lower())
    if not occurrences:
        return None
    for idx in occurrences:
        if idx > after_idx:
            return idx
    # No forward occurrence — return the first occurrence as a fallback.
    return occurrences[0]


def _link_toc_entries(doc, log: logging.Logger):
    """Find the Table of Contents (if any) within the first TOC_SCAN_PAGES
    pages and add internal links from each entry to the PDF page bearing
    the entry's printed page number.

    Returns (linked_count, entries, toc_page_range) where:
      * entries is a list of (label, target_page_index) tuples in
        document order — one per successfully-resolved TOC entry,
        regardless of whether a new link was inserted for it this run
        (entries are produced even when the overlap-check skips
        re-linking, so re-runs surface the same set of bookmark
        candidates as first runs).
      * toc_page_range is (start_page_idx, end_page_idx_inclusive) for
        the pages that contain the TOC itself (the heading page plus
        any continuation pages with TOC entries). Used by the bookmark
        builder to point the "Contents" branch header at the TOC's own
        page, and by the section-heading detector to skip TOC pages.
        Returns None if no TOC heading was found at all.
    """
    try:
        import fitz
    except ImportError:
        return 0, [], None

    toc_start = _find_toc_heading(doc, TOC_SCAN_PAGES)
    if toc_start is None:
        return 0, [], None

    printed_map = _build_printed_page_map(doc)
    if not printed_map:
        log.info("  TOC heading found but no printed page numbers detected; "
                 "skipping TOC linking")
        # We still know where the TOC page itself is — return it so the
        # bookmark builder can point Contents at it, and so the section
        # detector knows to skip it.
        return 0, [], (toc_start, toc_start)

    linked = 0
    entries = []   # [(label, target_page_index), ...]
    last_toc_page = toc_start  # tracks the furthest TOC page walked
    # Walk forward from the TOC heading page. Stop on the first page that
    # contains a TOC-terminating heading (TOA, Introduction, Argument, etc.)
    # or that yields no entries — whichever comes first. The TOC_SCAN_PAGES
    # budget applied only to *finding* the heading; once found, the TOC
    # itself is walked to its natural end regardless of length.
    for page_idx in range(toc_start, len(doc)):
        last_toc_page = page_idx
        page = doc[page_idx]
        # Snapshot existing link annotations so a re-run on an already-
        # linked PDF doesn't stack a second GOTO on every TOC entry. New
        # links added during this page's processing are appended to the
        # snapshot list so we also avoid double-linking within one run.
        existing_links = page.get_links()
        # Group text runs by baseline so a TOC entry's label, dot leaders,
        # and right-aligned page number — which PyMuPDF often returns as
        # three separate "lines" — combine into one logical row before
        # entry matching. Drop gutter line-numbers so they don't prepend to
        # the entry label or widen the link rect. See _collect_rows.
        page_lines = _collect_rows(page, drop_gutter_numbers=True)

        # Stop the TOC if any line on this page is a terminating heading.
        # A line that parses as a TOC entry (label + leaders + page number)
        # is NEVER treated as a terminator, even if its label happens to be
        # "Introduction" or "Argument" — those are legitimate first entries
        # in most briefs. Terminators are bare standalone headings.
        end_hit = False
        for line_text, _ in page_lines:
            if _TOC_HEADING_RE.search(line_text):
                continue
            if _TOC_ENTRY_RE.match(line_text):
                continue
            if _TOC_END_RE.search(line_text):
                end_hit = True
                break

        entries_on_this_page = 0
        # A TOC entry whose label is too long for one line wraps across
        # several rows, with the dot leaders and page number only on the
        # LAST row. We accumulate the leading rows in `pending` and attach
        # them to the row that finally carries the page number, so the whole
        # entry becomes one logical entry (clean merged label) and every one
        # of its rows is made clickable — not just the final line.
        pending = []           # [(text, bbox)] continuation rows
        _PENDING_MAX = 5       # guard against runaway accumulation
        top_band = _STAMP_BAND_PT
        bot_band = page.rect.height - _STAMP_BAND_PT
        for line_text, bbox in page_lines:
            if _TOC_HEADING_RE.search(line_text):
                pending = []
                continue
            m = _TOC_ENTRY_RE.match(line_text)
            if not m:
                # No trailing page number. Either a terminator (stop) or a
                # continuation line of a wrapped entry. Only accumulate rows
                # that sit in the page body and carry text — this excludes
                # the running header/footer title and page-number stamps in
                # the margin bands, which would otherwise prepend to the
                # next entry's label.
                if _TOC_END_RE.search(line_text):
                    break
                y0, y1 = bbox[1], bbox[3]
                in_margin = (y1 <= top_band) or (y0 >= bot_band)
                if (not in_margin and not _PAGE_STAMP_RE.match(line_text)
                        and any(ch.isalpha() for ch in line_text)):
                    pending.append((line_text, bbox))
                    if len(pending) > _PENDING_MAX:
                        pending.pop(0)
                continue
            # The line LOOKS like a TOC entry. Count it toward the "is
            # this page still part of the TOC?" signal whether or not we
            # can actually resolve its page target — a TOC that runs
            # several pages can have entries with targets that we failed
            # to detect (e.g., a missing printed page number), and that
            # shouldn't abort the walk through subsequent TOC pages.
            entries_on_this_page += 1

            # Continuation rows gathered since the previous entry belong to
            # THIS entry; consume them whether or not the target resolves.
            group_rows = pending + [(line_text, bbox)]
            pending = []

            page_label = m.group("page").lower()
            target_idx = _resolve_target(printed_map, page_label, page_idx)
            if target_idx is None:
                continue
            # Don't link an entry to the very page it sits on.
            if target_idx == page_idx:
                continue

            # Merge the wrapped label: continuation text + this row's label.
            label_parts = [t.strip() for t, _ in group_rows[:-1]]
            label_parts.append(m.group("label").strip())
            label = " ".join(p for p in label_parts if p)

            # Record this entry for the bookmark builder regardless of
            # whether we end up inserting a new link below — re-runs hit
            # the overlap-check skip path but should still surface the
            # same bookmark set as a first run.
            entries.append((label, target_idx))

            # Link EVERY row of the entry so the whole (possibly multi-line)
            # entry is clickable, not just the final page-numbered line.
            for _, gbbox in group_rows:
                rect = fitz.Rect(gbbox)
                # Skip if this rect overlaps an existing link annotation —
                # primarily a re-run guard, but also catches the unlikely
                # case of a TOC line that overlaps a citation rect.
                already = False
                for el in existing_links:
                    er = el.get("from")
                    if er and rect.intersects(er):
                        already = True
                        break
                if already:
                    continue
                page.insert_link({
                    "kind": fitz.LINK_GOTO,
                    "from": rect,
                    "page": target_idx,
                    "to": fitz.Point(0, 0),
                })
                existing_links.append({"from": rect, "kind": fitz.LINK_GOTO})

            # Underline the trailing page number only, not the whole row —
            # underlining a full TOC line (which is mostly dot leaders)
            # looks like a typesetting error. We locate just the number
            # via search_for, restricted to the final row's bbox.
            final_rect = fitz.Rect(bbox)
            num_quads = page.search_for(m.group("page"), clip=final_rect, quads=True)
            for q in num_quads:
                qr = q.rect
                # Only the right-most occurrence on this line — the leader
                # number, not a digit that happens to appear in the label.
                if qr.x1 < final_rect.x1 - 5:
                    continue
                underline_y = qr.y1 - 0.5
                page.draw_line(
                    fitz.Point(qr.x0, underline_y),
                    fitz.Point(qr.x1, underline_y),
                    color=LINK_COLOUR,
                    width=1.0,
                )
                break
            linked += 1

        if end_hit or entries_on_this_page == 0:
            # This page was either past the TOC (terminator heading or no
            # entries). Roll back last_toc_page to the previous page since
            # this one isn't TOC content. If this is the first page (no
            # prior TOC content), last_toc_page stays at toc_start since
            # that page at least had the heading.
            if entries_on_this_page == 0 and page_idx > toc_start:
                last_toc_page = page_idx - 1
            break

    if linked:
        log.info(f"  Linked {linked} TOC entry(ies) to internal pages")
    return linked, entries, (toc_start, last_toc_page)


# ────────────────────────────────────────────────────────────────────────────
# Section-heading detection (TOC fallback for the bookmark builder)
# ────────────────────────────────────────────────────────────────────────────
# When a document has no table of contents, the bookmark builder's
# `Contents` branch would be empty — we'd offer the user no high-level
# navigation. This block detects section headings heuristically so we
# can populate `Contents` from the document's structure even without a
# parsed TOC.
#
# Signal layering: a row is treated as a heading only if it combines a
# *visual* signal (larger font than body, or bold formatting) with a
# *structural* signal (an outline label like "I.", "A.", "1.", or a
# short ALL-CAPS line). Either signal alone produces too many false
# positives — body text is often bold for defined terms, and outline
# labels appear in body sentences ("Part I argues..."). Together they
# concentrate the matches on real headings.
#
# Scope: this fallback runs ONLY when TOC parsing produced no entries.
# Documents that have a TOC use the parsed entries — the heading
# detector is never invoked for them, so we don't risk regressing
# TOC-bearing documents with a noisier source.

# Outline labels: Roman numerals (capped at X to limit false positives
# from body text starting with "I."), single uppercase letter, 1-2
# digit number, and single lowercase letter. Each is followed by a
# period, then either end-of-line (label alone on the row, common when
# the label and heading text are typeset as separate runs) or
# whitespace + heading text.
_HEADING_LABEL_RE = re.compile(
    r"""
    ^
    \s*
    (?P<label>
        (?:I|II|III|IV|V|VI|VII|VIII|IX|X)\.        # Roman I-X
      | [A-Z]\.                                       # single uppercase letter
      | \d{1,2}\.                                     # 1-2 digit number
      | [a-z]\.                                       # single lowercase letter
    )
    (?:\s+(?P<text>\S.*?))?                           # optional heading text
    \s*$
    """,
    re.VERBOSE,
)

# Bare ALL-CAPS heading without outline label: "INTRODUCTION",
# "STATEMENT OF FACTS", "ARGUMENT", "CONCLUSION", "FIRST CAUSE OF
# ACTION", "SECTION 2", "COUNT 1". Require 2+ words OR a single word
# of 6+ chars (rules out short noise like "FILED" stamps or "EXHIBIT"
# which are caught elsewhere). The character class allows digits so
# numbered headings ("SECTION 2", "COUNT 1") match — these are common
# in complaints and longer briefs.
_HEADING_ALLCAPS_RE = re.compile(
    r"^\s*[A-Z][A-Z0-9\s'\-,&]{4,}[A-Z0-9]\s*$"
)

# A line that looks like a complete sentence (ends in period and has
# enough internal whitespace to be a sentence, not just a label
# followed by period). Used as a negative filter — sentences are not
# headings. Threshold: a line ending in ".", "!" or "?" with more than
# 8 whitespace-separated tokens is a sentence and not a heading.
_HEADING_MAX_TOKENS = 15

# Pages to skip at the start of every document when running the
# heading-detection pass. Pages 0-1 are reserved for caption layouts
# whose ALLCAPS court-name lines ("SUPERIOR COURT OF THE STATE OF
# CALIFORNIA", "COUNTY OF LOS ANGELES") and bold-centered case
# captions would otherwise be misread as section headings. By page 2
# the caption block has ended and either TOC or body content begins.
_HEADING_SKIP_LEAD_PAGES = 2

# Hard cap on how far into a document the heading detector scans.
# Briefs are limited to 20-25 pages by most court rules; 30 is a
# comfortable ceiling that covers oversized briefs without venturing
# into exhibit territory in documents that lack cover-page markers.
# The exhibit-cover stop (see _detect_section_headings) usually
# triggers first when exhibits exist — this cap is the fallback for
# memos whose exhibits are unmarked or absent but have a long
# appendix-style tail (e.g. proofs of service, declarations stapled
# without cover sheets).
_HEADING_MAX_SCAN_PAGES = 30


def _document_body_font_size(doc) -> float:
    """Return the modal font size across the document (the size that
    appears on the most text spans). This is the document's body font
    — any heading must be larger than this OR bold to qualify.

    Sampling: scan every page, count span sizes weighted by character
    length so a long body paragraph counts more than a short bold
    label. Returns 12.0 as a safe default if the document has no
    text spans at all.
    """
    from collections import Counter
    counts: Counter = Counter()
    for page in doc:
        try:
            d = page.get_text("dict")
        except Exception:
            continue
        for block in d.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for sp in line.get("spans", []):
                    t = sp.get("text", "")
                    if not t.strip():
                        continue
                    size = sp.get("size", 0)
                    if not size:
                        continue
                    # Round to 0.5pt to merge near-identical sizes
                    bucket = round(size * 2) / 2
                    counts[bucket] += len(t.strip())
    if not counts:
        return 12.0
    return counts.most_common(1)[0][0]


def _row_visual_signals(page, bbox, line_text):
    """Return (max_size, is_bold, is_centered) for the row at `bbox` on `page`.

    Walks every span whose vertical center lies inside the row's bbox
    and reports the largest font size seen and whether any span is
    bold (either flagged bold or with a font name containing "Bold").
    Centering is computed geometrically: the row's horizontal midpoint
    is within ~10pt of the page's horizontal midpoint, AND the row is
    narrower than ~70% of the page width (so a long body paragraph
    that happens to span the full page isn't treated as centered).
    """
    max_size = 0.0
    is_bold = False
    try:
        d = page.get_text("dict")
    except Exception:
        return max_size, is_bold, False
    y0, y1 = bbox[1], bbox[3]
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            lb = line.get("bbox")
            if not lb:
                continue
            line_y_center = (lb[1] + lb[3]) / 2
            if not (y0 - 1 <= line_y_center <= y1 + 1):
                continue
            for sp in line.get("spans", []):
                if not sp.get("text", "").strip():
                    continue
                size = sp.get("size", 0)
                if size > max_size:
                    max_size = size
                flags = sp.get("flags", 0)
                font = sp.get("font", "") or ""
                if (flags & 16) or "Bold" in font or "bold" in font:
                    is_bold = True
    # Centering: compare row x-midpoint to page x-midpoint.
    # Use the page's full width as the reference. Pleading paper has
    # a gutter on the left, but a centered heading on pleading paper
    # is centered relative to the whole page anyway (publishers center
    # over the printable area, not the gutter-adjusted body column),
    # so this works in both layouts. The narrow-line guard
    # (width < 70% of page width) filters out body paragraphs that
    # happen to have a midpoint near the page center.
    is_centered = False
    try:
        x0, _, x1, _ = bbox
        row_mid = (x0 + x1) / 2
        page_mid = page.rect.width / 2
        row_width = x1 - x0
        if abs(row_mid - page_mid) <= 10 and row_width < page.rect.width * 0.70:
            is_centered = True
    except Exception:
        pass
    return max_size, is_bold, is_centered


def _is_heading_row(line_text: str, font_size: float, is_bold: bool,
                    is_centered: bool, body_size: float) -> bool:
    """Decide whether a row is a section heading.

    Combines visual signals (font size > body, bold, centered) with
    structural signals (outline label, ALL-CAPS pattern). Rejects rows
    that look like full sentences. Conservative by design — false
    positives in the bookmark tree are more annoying than false
    negatives.

    Two paths to qualify:
      * Outline-label heading ("I. INTRODUCTION", "A. The Court..."):
        passes if EITHER larger-than-body OR bold. These headings
        are visually marked by their label structure, so a single
        visual cue is enough corroboration.
      * Bare ALL-CAPS heading ("SUMMARY OF ARGUMENT", "INTRODUCTION"):
        passes only if AT LEAST 2 of {bold, larger-than-body,
        centered} are true. The all-caps text alone could occur in
        body text (court names in captions, statute references), so
        we require two corroborating visual cues to commit to a
        bookmark.
    """
    text = line_text.strip()
    if not text:
        return False

    # Length cap — headings are never long sentences
    tokens = text.split()
    if len(tokens) > _HEADING_MAX_TOKENS:
        return False

    larger = font_size >= body_size + 1.0

    # Sentence rejection: ends in period with enough words to be a
    # sentence. Outline-label headings ("I. INTRODUCTION") end at
    # "INTRODUCTION" without a period, so a trailing period on a long
    # line is a sentence-shape signal. "I. Introduction" (no period
    # at end) passes; "We address each claim in turn." rejects.
    if text[-1] in ".!?" and len(tokens) > 5:
        label_only = _HEADING_LABEL_RE.match(text)
        if not (label_only and not label_only.group("text")):
            return False

    # Outline-label path: structural cue is strong, single visual cue
    # is enough.
    if _HEADING_LABEL_RE.match(text):
        if larger or is_bold:
            return True
        return False

    # ALL-CAPS path: structural cue is weaker (body text can be ALL-CAPS
    # for proper nouns or stylistic emphasis), so require 2 of 3 visual
    # cues. This is the user spec — "bold, large, underlined, and
    # centered" relaxed to a 2-of-3 majority (with underline omitted
    # because PyMuPDF doesn't natively expose underline as a span flag
    # and the styling rarely appears alone).
    if _HEADING_ALLCAPS_RE.match(text) and len(text) >= 6:
        visual_score = (1 if is_bold else 0) + (1 if larger else 0) \
                       + (1 if is_centered else 0)
        if visual_score >= 2:
            return True
        return False

    return False


def _detect_section_headings(doc, log: logging.Logger,
                              toc_page_range=None):
    """Return [(label, page_index), ...] for every section heading
    found in the document, in document order.

    Runs unconditionally on briefs (gated upstream by filename via
    should_skip_linking). Pages identified as TOC pages by
    _link_toc_entries are skipped — those headings are already
    covered by the Contents branch and would otherwise double-up
    in the outline.

    Skips the first _HEADING_SKIP_LEAD_PAGES PDF pages to avoid
    misreading caption ALL-CAPS text ("SUPERIOR COURT OF...") as
    headings.

    Scan range: stops at whichever comes first —
      * _HEADING_MAX_SCAN_PAGES (30) — hard cap covering even
        oversized briefs without venturing into exhibits.
      * The earliest exhibit cover page detected by
        _find_exhibit_cover_pages — once the exhibits start, headings
        inside attached documents are false positives. The detector
        is reused here without modification; it returns the same
        cover map _link_exhibit_references uses later. Calling it
        twice is cheap (no link insertion) and avoids reordering the
        process_pdf pipeline.

    toc_page_range: (start_idx, end_idx_inclusive) for pages occupied
    by the TOC itself, or None. Those pages are skipped during the
    section scan.

    Deduplicates against repeated identical headings appearing on
    consecutive pages (page headers/footers that read e.g. "ARGUMENT"
    on every page of the argument section). Only the first occurrence
    is kept as a bookmark target — subsequent identical hits are
    page-header artefacts.
    """
    body_size = _document_body_font_size(doc)

    # Build the set of pages to skip: TOC pages, computed once upfront.
    toc_skip_pages = set()
    if toc_page_range is not None:
        t_start, t_end = toc_page_range
        toc_skip_pages.update(range(t_start, t_end + 1))

    # Determine the scan stop. The page cap and the earliest exhibit
    # cover both lower-bound the inclusive end-of-scan index.
    scan_end = min(_HEADING_MAX_SCAN_PAGES, len(doc))
    try:
        cover_map, _ = _find_exhibit_cover_pages(doc)
        if cover_map:
            earliest_cover = min(min(pages) for pages in cover_map.values() if pages)
            # Only honor the exhibit-cover stop if the earliest cover
            # is at PDF index 3 or later (page 4+). Earlier "covers"
            # are almost certainly false positives — a TOC entry, a
            # caption block, or a heading that happens to mention an
            # exhibit. A real brief reserves its first 3 pages for
            # caption + TOC at minimum.
            if 3 <= earliest_cover < scan_end:
                scan_end = earliest_cover
    except Exception:
        # Cover detection failures are non-fatal here — fall back to
        # the page cap alone. The bookmark feature isn't worth
        # aborting over a cover-scan glitch.
        pass

    entries = []
    seen_labels = set()

    for page_idx in range(_HEADING_SKIP_LEAD_PAGES, scan_end):
        if page_idx in toc_skip_pages:
            continue
        page = doc[page_idx]
        rows = _collect_rows(page)
        for line_text, bbox in rows:
            max_size, is_bold, is_centered = _row_visual_signals(
                page, bbox, line_text)
            if not _is_heading_row(line_text, max_size, is_bold,
                                    is_centered, body_size):
                continue
            label = line_text.strip()
            # Dedup: only the first sighting of a given label string
            # becomes a bookmark. Repeated identical strings on later
            # pages are usually running headers.
            if label in seen_labels:
                continue
            seen_labels.add(label)
            entries.append((label, page_idx))

    if entries:
        log.info(f"  Detected {len(entries)} section heading(s) "
                 f"(scanned pages {_HEADING_SKIP_LEAD_PAGES+1}-"
                 f"{scan_end}; body font ~{body_size:.1f}pt)")
    return entries


# ────────────────────────────────────────────────────────────────────────────
# Cause-of-action detection and linking (Complaint-only)
# ────────────────────────────────────────────────────────────────────────────
# Complaints (and amended complaints — FAC/SAC/TAC) carry a distinctive
# structure: each claim is introduced by a heading like "FIRST CAUSE OF
# ACTION", "SECOND CAUSE OF ACTION", etc. We bookmark every one of these
# and, when the cover page also lists the causes in the same ALL-CAPS
# form, we add hyperlinks from each cover-page listing to its body
# occurrence.
#
# This block runs only for files whose name matches `is_complaint()`.
# Briefs, declarations, and other doc types pass through untouched —
# their bookmarks come from the TOC parser or the heading detector
# instead.
#
# Identifying a cause-of-action heading:
#   * "FIRST CAUSE OF ACTION", "SECOND CAUSE OF ACTION", … (English ordinals)
#   * "COUNT I", "COUNT II", … (Roman numerals)
#   * "COUNT ONE", "COUNT TWO", … (English cardinals)
#   * "FIRST CLAIM FOR RELIEF", "SECOND CLAIM FOR RELIEF", … (federal style)
# All in ALL-CAPS. Mixed-case body text mentioning these phrases ("the
# first cause of action…") never matches.
#
# Picking the bookmark target page when an ordinal appears multiple
# times: complaints commonly list every cause on the caption/cover page
# AND then have a section heading deeper in the document. We want the
# bookmark to jump to the body heading, not the cover listing. Rule:
# when an ordinal appears ≥2 times, the first occurrence (within the
# first 3 PDF pages) is treated as the cover listing and the next
# occurrence is the bookmark target. Single-occurrence ordinals
# bookmark wherever they appear.

# English ordinals up to 30 (rare to see complaints with more)
_CAUSE_ORDINALS_ENGLISH = [
    "FIRST", "SECOND", "THIRD", "FOURTH", "FIFTH", "SIXTH", "SEVENTH",
    "EIGHTH", "NINTH", "TENTH", "ELEVENTH", "TWELFTH", "THIRTEENTH",
    "FOURTEENTH", "FIFTEENTH", "SIXTEENTH", "SEVENTEENTH", "EIGHTEENTH",
    "NINETEENTH", "TWENTIETH", "TWENTY-FIRST", "TWENTY-SECOND",
    "TWENTY-THIRD", "TWENTY-FOURTH", "TWENTY-FIFTH", "TWENTY-SIXTH",
    "TWENTY-SEVENTH", "TWENTY-EIGHTH", "TWENTY-NINTH", "THIRTIETH",
]
_ORDINAL_TO_NUMBER = {w: i + 1 for i, w in enumerate(_CAUSE_ORDINALS_ENGLISH)}

# Roman numerals I-XXX for "COUNT I", "COUNT II", etc.
_CAUSE_ROMANS = [
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X",
    "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX",
    "XXI", "XXII", "XXIII", "XXIV", "XXV", "XXVI", "XXVII", "XXVIII",
    "XXIX", "XXX",
]
_ROMAN_TO_NUMBER = {r: i + 1 for i, r in enumerate(_CAUSE_ROMANS)}

# English cardinals (ONE, TWO, ...) for "COUNT ONE" style.
_CAUSE_CARDINALS = [
    "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT",
    "NINE", "TEN", "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN",
    "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN", "TWENTY",
]
_CARDINAL_TO_NUMBER = {w: i + 1 for i, w in enumerate(_CAUSE_CARDINALS)}

# Matches: "FIRST CAUSE OF ACTION", "FIRST CLAIM FOR RELIEF",
# "SECOND CAUSE OF ACTION (Breach of Contract)", and "TWENTY-FIRST CAUSE OF
# ACTION". The ordinal is the only capture group used; downstream we
# normalize to an integer.
_CAUSE_ORDINAL_RE = re.compile(
    r"^\s*(?P<ord>" + "|".join(_CAUSE_ORDINALS_ENGLISH) + r")"
    r"\s+(?:CAUSE\s+OF\s+ACTION|CLAIM\s+FOR\s+RELIEF)"
    r"(?:\s|,|$|\(|—|-|:)",
    re.IGNORECASE,
)

# Matches "COUNT I", "COUNT II", … (Roman) and "COUNT ONE", "COUNT TWO", …
# (English cardinal). Case-insensitive for safety though most occur in
# ALL-CAPS.
_COUNT_ROMAN_RE = re.compile(
    r"^\s*COUNT\s+(?P<rom>" + "|".join(_CAUSE_ROMANS) + r")"
    r"(?:\s|,|$|\(|—|-|:|\.)",
    re.IGNORECASE,
)
_COUNT_CARDINAL_RE = re.compile(
    r"^\s*COUNT\s+(?P<card>" + "|".join(_CAUSE_CARDINALS) + r")"
    r"(?:\s|,|$|\(|—|-|:|\.)",
    re.IGNORECASE,
)

# Pages on which a cause-of-action ordinal is treated as a cover-page
# *listing* rather than the body target. Complaints have caption + cover
# listing within the first 3 PDF pages essentially always; once we're
# past page 3 we're in the body.
_COVER_LISTING_LAST_PAGE = 3   # 0-indexed exclusive bound — pages 1-3

# Document-name patterns that identify a complaint or amended complaint.
# Independent of `_SKIP_LINK_PATTERNS` so we can run cause-of-action
# linking on these documents even though TOC and citation linking are
# disabled for them. Declarations and separate statements are
# intentionally NOT here.
_COMPLAINT_PATTERNS = [
    re.compile(r"(?:^|[\s_-])Complaint(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])Compl\.?(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])FAC(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])SAC(?:[\s_-]|$)", re.IGNORECASE),
    re.compile(r"(?:^|[\s_-])TAC(?:[\s_-]|$)", re.IGNORECASE),
]


def is_complaint(filename: str) -> bool:
    """Return True if filename suggests a complaint or amended complaint."""
    name = filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    stem = name.rsplit(".", 1)[0]
    for pat in _COMPLAINT_PATTERNS:
        if pat.search(stem):
            return True
    return False


def _cause_number_from_row(text: str):
    """Return the integer ordinal (1-30) if `text` is a cause-of-action
    heading, else None. Recognises English ordinals, Roman COUNTs, and
    English cardinal COUNTs."""
    t = text.strip()
    m = _CAUSE_ORDINAL_RE.match(t)
    if m:
        return _ORDINAL_TO_NUMBER.get(m.group("ord").upper().replace(" ", "-"))
    m = _COUNT_ROMAN_RE.match(t)
    if m:
        return _ROMAN_TO_NUMBER.get(m.group("rom").upper())
    m = _COUNT_CARDINAL_RE.match(t)
    if m:
        return _CARDINAL_TO_NUMBER.get(m.group("card").upper())
    return None


def _detect_causes_of_action(doc, log: logging.Logger):
    """Find every cause-of-action heading and return:
      bookmark_entries: [(label, target_page_index), ...] ordered by
                        ordinal number, one entry per distinct cause.
      cover_occurrences: [(ordinal_n, page_index, label, bbox), ...]
                        every match found on PDF pages 1-3 (the cover
                        listing source for the link pass).
      body_occurrences: {ordinal_n: (page_index, label, bbox)}
                        the chosen body target per ordinal.

    Selection rule: for each ordinal, if it appears ≥2 times and at
    least one appearance is on PDF pages 1-3, the FIRST such
    appearance is the cover listing and the NEXT appearance is the
    body target. If only one appearance exists (cover only or body
    only), that one is the body target.

    No scan range cap — complaints can be 80+ pages. The structural
    constraint of ALL-CAPS-with-ordinal pattern keeps false positives
    extremely low without needing a page limit.
    """
    # all_hits[ordinal_n] = list of (page_idx, label, bbox) in document order
    all_hits = {}
    for page_idx, page in enumerate(doc):
        for line_text, bbox in _collect_rows(page):
            n = _cause_number_from_row(line_text)
            if n is None:
                continue
            all_hits.setdefault(n, []).append(
                (page_idx, line_text.strip(), bbox)
            )

    if not all_hits:
        return [], [], {}

    cover_occurrences = []
    body_occurrences = {}
    for n, hits in sorted(all_hits.items()):
        # Partition rule: for each ordinal, the FIRST occurrence (in
        # document order) is treated as the cover listing IF it sits
        # on PDF pages 1-3; subsequent occurrences are body
        # candidates. Single-occurrence ordinals bookmark wherever they
        # appear. This handles all common layouts:
        #   * Cover listing + body heading           → cover/body split
        #   * Body heading only (no cover listing)   → bookmark to body
        #   * Cover listing only (no body heading)   → bookmark to cover
        #     (it's the only target we have)
        first_page, _, _ = hits[0]
        if len(hits) >= 2 and first_page < _COVER_LISTING_LAST_PAGE:
            cover_occurrences.append(
                (n, hits[0][0], hits[0][1], hits[0][2])
            )
            body_occurrences[n] = hits[1]
        else:
            body_occurrences[n] = hits[0]

    # Bookmark entries: one per ordinal, ordered by ordinal number.
    bookmark_entries = [
        (body_occurrences[n][1], body_occurrences[n][0])
        for n in sorted(body_occurrences)
    ]
    return bookmark_entries, cover_occurrences, body_occurrences


def _link_cover_to_causes(doc, cover_occurrences, body_occurrences,
                           log: logging.Logger) -> int:
    """Insert GOTO links from each cover-page cause-of-action listing
    to the body occurrence. Returns count of links inserted.

    Skipped per-ordinal when (a) there's no separate body target for
    that ordinal (cover only), or (b) the cover and body occurrences
    are on the same page (rare but possible if a complaint puts both
    a cover-style listing and the actual heading on the same page).
    Idempotent: existing link annotations on the cover rect are
    detected and the new link is suppressed.
    """
    try:
        import fitz
    except ImportError:
        return 0

    linked = 0
    for n, cover_page_idx, label, bbox in cover_occurrences:
        body = body_occurrences.get(n)
        if not body:
            continue
        body_page_idx = body[0]
        if body_page_idx == cover_page_idx:
            # Same page — there's nothing to link to. Common when the
            # ordinal only appears once on the cover and the "body
            # target" we recorded is actually that same cover listing.
            continue
        page = doc[cover_page_idx]
        rect = fitz.Rect(bbox)

        # Idempotency: skip if an existing link already covers this rect
        existing_links = page.get_links()
        already = False
        for el in existing_links:
            er = el.get("from")
            if er and rect.intersects(er):
                already = True
                break
        if already:
            continue

        page.insert_link({
            "kind": fitz.LINK_GOTO,
            "from": rect,
            "page": body_page_idx,
            "to": fitz.Point(0, 0),
        })
        underline_y = rect.y1 - 0.5
        page.draw_line(
            fitz.Point(rect.x0, underline_y),
            fitz.Point(rect.x1, underline_y),
            color=LINK_COLOUR,
            width=1.0,
        )
        linked += 1

    if linked:
        log.info(f"  Linked {linked} cover-page cause-of-action reference(s)")
    return linked


# ────────────────────────────────────────────────────────────────────────────
# Exhibit cross-reference linking
# ────────────────────────────────────────────────────────────────────────────
# Turn body references like "Exhibit 5", "Ex. 12", "Exh. 3", "Exhibit A",
# "Ex. AA" into internal jumps to the corresponding exhibit page. Runs
# whenever 2+ distinct exhibit cover pages are detected, regardless of
# the document's total page count — the cover-count gate is sufficient
# to avoid false positives in short briefs that mention "Exhibit A" in
# passing (those won't have multiple labelled cover rows).
#
# Identifier forms:
#   Numeric  -- Exhibit 1, Ex. 12, Exh. 3   (1-3 digits)
#   Letter   -- Exhibit A, Ex. AA           (uppercase, 1-2 letters)
#   The same document can use both formats; we link whichever appears.
#   Each form is also matched in quoted variants — Exhibit "3" (straight)
#   and Exhibit \u201c3\u201d (curly). California pleading templates often
#   wrap exhibit numbers in quotes, and a single document frequently mixes
#   bare and quoted forms across paragraphs. Cover-page detection and
#   body-reference matching both handle all three spellings.
#   Plurals ("Exhibits"), ranges ("Exhibit 3-5"), and lowercase letter
#   references ("exhibit a", which is almost always natural English, not
#   a cite) are intentionally left alone.
#
# Letter-specific safeguards:
#   Letter references have a much higher false-positive rate than numeric
#   ones because most English sentences begin with a capital letter. To
#   keep this safe:
#     * Letters must be UPPERCASE. "Exhibit a" never matches.
#     * Letter cover rows must either be the label alone or be followed
#       by a separator (— – - :) before a descriptor. "Exhibit A copy of
#       the contract" is NOT treated as a cover row.
#     * Body-text linking uses a rect-validation probe to reject any quad
#       that's a fragment of a longer token. PyMuPDF's search_for is
#       glyph-based and would otherwise latch onto the "Exhibit A" part
#       of "Exhibit Apple" or "Exhibits A-C". The probe (see
#       _is_complete_phrase_rect) checks whether the next ~3pt to the
#       right of a match contains an alphanumeric glyph — if so, the
#       reference is a prefix of a longer word and is discarded. The
#       same logic protects numeric matches: "Exhibit 1" won't latch onto
#       "Exhibit 12" because "2" is the next glyph.
#
# Target selection:
#   For each exhibit identifier, we find every page whose layout shows the
#   label as a row of its own. The *nearest* such page AFTER the body
#   reference is the link target. This matters in combined filings that
#   contain two declarations each numbering their own exhibits 1, 2, 3 —
#   a reference in declaration A's body should jump to declaration A's
#   Exhibit 1, not declaration B's. If no cover page exists after the
#   reference, the nearest one before is used as a fallback.
# EXHIBIT_LINK_MIN_PAGES was a 30-page minimum gate retired when an
# 18-page declaration with Exhibits A/B/C failed to link them. The
# cover-count >= 2 gate inside _link_exhibit_references is now the
# sole guard. If you want to bring back a page-count minimum, restore
# the `if len(doc) <= N: return 0, {}` check at the top of that
# function — but be aware short multi-exhibit declarations exist.

# Quote characters that may surround an exhibit identifier. Both straight
# ("3") and the curly opener/closer (\u201c3\u201d) appear in real-world
# filings — the curly form is what Word produces with smart-quotes on,
# and many caption-page templates wrap the number that way.
_EXHIBIT_QUOTE_OPEN = "[\"'\u201c\u2018]"
_EXHIBIT_QUOTE_CLOSE = "[\"'\u201d\u2019]"

_EXHIBIT_COVER_RE = re.compile(
    r"""
    ^\s*
    (?:EXHIBIT|Exhibit|EX\.|Ex\.|EXH\.|Exh\.)
    \s+
    """ + _EXHIBIT_QUOTE_OPEN + r"""?                  # optional opening quote
    (?:
        (?P<num>\d{1,3}) \b
        """ + _EXHIBIT_QUOTE_CLOSE + r"""?             # optional closing quote
        .{0,40}                                        # numeric: any short trailing descriptor
      |
        (?P<letter>[A-Z]{1,2}) \b
        """ + _EXHIBIT_QUOTE_CLOSE + r"""?             # optional closing quote
        (?:\s*[\u2014\u2013\-:]\s*.{1,40})?            # letter: optional separator + descriptor
    )
    \s*$
    """,
    re.VERBOSE,
)


def _find_exhibit_cover_pages(doc):
    """Scan the document for pages bearing an exhibit label as a standalone
    row, and return (targets, label_pages):

      targets     -- {exhibit_id: [page_idx, ...]} with contiguous runs
                     collapsed to their first page. Exhibit IDs are
                     normalized strings: "1", "12", "A", "AA". These are
                     the candidate link targets for body references.
      label_pages -- set of every page index whose layout contains an
                     exhibit label row. These are excluded from the body-
                     reference scan so we don't link an exhibit content
                     page's own banner ("Exhibit 5 — page 2 of 9") back
                     to the cover sheet — that would be a self-link from
                     within the exhibit itself.

    A "standalone row" is what _collect_rows returns when the label is the
    whole text run on its baseline — this catches both true cover sheets
    (one big "EXHIBIT 5" on an otherwise blank page) and content pages
    whose top banner reads "Exhibit 5" or "Exhibit 5 — Description".
    """
    found: dict = {}
    label_pages: set = set()
    for i, page in enumerate(doc):
        # Drop pleading-paper gutter line-numbers: on a pleading-paper
        # exhibit cover the centered "EXHIBIT C" can share a baseline with a
        # left-margin line number, which _collect_rows would otherwise fuse
        # into "14 EXHIBIT C" — breaking _EXHIBIT_COVER_RE's ^EXHIBIT anchor
        # and leaving that exhibit (and every body reference to it) unlinked.
        for line_text, _bbox in _collect_rows(page, drop_gutter_numbers=True):
            m = _EXHIBIT_COVER_RE.match(line_text)
            if not m:
                continue
            # Exactly one of `num` or `letter` matches; normalize both to
            # strings so the cover map can mix numbered and lettered IDs.
            ident = m.group("num") or m.group("letter")
            if not ident:
                continue
            found.setdefault(ident, []).append(i)
            label_pages.add(i)
            break  # one label-row per page is enough to register the page
    # Collapse contiguous runs to their first page: if Exhibit 5 has a
    # cover sheet at page 42 and content pages 43-50 each carry a banner
    # "Exhibit 5 — page N of 9", we want page 42 as the single link
    # target and don't need pages 43-50 as additional candidates. The
    # comparison tracks the previous *input* page, not the previous kept
    # page, so a full unbroken run collapses correctly. The `label_pages`
    # set keeps all of 42-50 — content pages still count as "inside the
    # exhibit" for self-link suppression.
    for ident, pages in found.items():
        pages.sort()
        deduped = [pages[0]]
        prev = pages[0]
        for p in pages[1:]:
            if p != prev + 1:
                deduped.append(p)
            prev = p
        found[ident] = deduped
    return found, label_pages


def _exhibit_target_for(ref_page: int, candidates) -> int:
    """Pick the nearest exhibit cover page for a body reference.

    Preference: the smallest candidate index that is > ref_page. If none
    exist after the reference, fall back to the largest candidate index
    that is < ref_page. Returns None if `candidates` is empty.
    """
    if not candidates:
        return None
    after = [p for p in candidates if p > ref_page]
    if after:
        return after[0]
    before = [p for p in candidates if p < ref_page]
    if before:
        return before[-1]
    return None


def _is_complete_phrase_rect(page, rect, probe_pt: float = 3.0) -> bool:
    """Return True if `rect` is the visual extent of a complete reference,
    i.e. not a prefix of a longer word/number. We check by probing the
    next ~3pt to the right of the rect: if a glyph there is alphanumeric
    AND not separated from the rect by whitespace, the rect is a fragment
    of a longer token (e.g. the "Exhibit 1" part of "Exhibit 12", or the
    "Exhibit A" part of "Exhibit Apple") and we should refuse to link it.

    PyMuPDF's search_for matches phrases case-insensitively at the glyph
    level and doesn't honour word boundaries. We use this probe to recover
    the boundary check our regex applies to the text stream.
    """
    try:
        import fitz
    except ImportError:
        return True
    probe = fitz.Rect(rect.x1, rect.y0, rect.x1 + probe_pt, rect.y1)
    after = page.get_text("text", clip=probe)
    for ch in after:
        if ch.isspace():
            return True  # whitespace before any other char → complete reference
        if ch.isalnum():
            return False  # word continues immediately → fragment
        # Punctuation (period, comma, paren) is fine — that's a real boundary
        return True
    return True  # nothing visible in the probe → complete reference


# Common identifier-form prefixes used in exhibit references. Used by the
# body-link pass to drive page.search_for for every cover-map identifier.
# (We don't loop over the regex directly because PyMuPDF's search_for is
# glyph-based and would miss case-folded variants if we passed only one
# spelling.)
_EXHIBIT_PREFIXES = ("Exhibit", "EXHIBIT", "Ex.", "EX.", "Exh.", "EXH.")


def _link_exhibit_references(doc, log: logging.Logger):
    """Find body references to exhibits and link them to the matching
    exhibit page.

    Returns (linked_count, cover_map) where cover_map is the
    {ident: [page_indices]} dict from _find_exhibit_cover_pages. The
    cover_map is empty (and linked_count is 0) when the document is
    too short, has fewer than 2 exhibits, or PyMuPDF isn't available;
    in those cases no linking happens and no bookmarks should be made.
    """
    try:
        import fitz
    except ImportError:
        return 0, {}

    cover_map, label_pages = _find_exhibit_cover_pages(doc)
    if len(cover_map) < 2:
        # Requires multiple detected exhibit covers — a single Exhibit 1
        # alone isn't worth a pass (and an empty map definitely isn't).
        # This gate is sufficient on its own: short documents without
        # real exhibits won't have 2+ covers detected. The previous
        # belt-and-suspenders page-count gate (EXHIBIT_LINK_MIN_PAGES)
        # was dropped because it excluded short declarations with
        # multiple legitimate exhibits (an 18-page declaration with
        # Exhibits A, B, C would never reach this code).
        return 0, {}

    linked = 0
    for page_idx, page in enumerate(doc):
        if page_idx in label_pages:
            # Don't link from within an exhibit's own pages — that would
            # turn the exhibit's header banner into a self-referential link.
            continue
        existing_links = page.get_links()
        # For each known exhibit identifier, look for every spelling of
        # "<prefix> <ident>" on this page. PyMuPDF's search_for is
        # glyph-based and doesn't honour word boundaries, so each quad
        # gets validated by _is_complete_phrase_rect before linking —
        # this is what prevents "Exhibit 1" from latching onto "Exhibit
        # 12", or "Exhibit A" from latching onto "Exhibit Apple".
        for ident, candidates in cover_map.items():
            target = _exhibit_target_for(page_idx, candidates)
            if target is None or target == page_idx:
                continue
            # Build every spelling we want to match for this identifier.
            # The bare form covers "Exhibit 5"; the quoted forms cover
            # "Exhibit \"5\"" (straight) and "Exhibit \u201c5\u201d" (curly).
            # All three appear in real filings, sometimes in the same
            # document — the body text of a declaration commonly switches
            # between forms paragraph to paragraph. Cover-page detection
            # is already quote-tolerant via _EXHIBIT_COVER_RE; this loop
            # is the body-reference counterpart.
            ident_forms = (
                ident,
                f"\"{ident}\"",
                f"\u201c{ident}\u201d",
            )
            for prefix in _EXHIBIT_PREFIXES:
                for ident_form in ident_forms:
                    phrase = f"{prefix} {ident_form}"
                    quads = page.search_for(phrase, quads=True)
                    for q in quads:
                        rect = q.rect
                        # PyMuPDF's search_for is case-insensitive, so a search
                        # for "Exhibit A" also matches lowercase "exhibit a"
                        # which is almost always natural English text rather
                        # than a cite. Clip the rect and confirm the glyph
                        # text matches the requested case exactly.
                        clipped = page.get_text("text", clip=rect).strip()
                        if clipped != phrase:
                            continue
                        # Reject quads that are fragments of longer tokens
                        # (e.g. "Exhibit 1" inside "Exhibit 12"). See the
                        # docstring of _is_complete_phrase_rect for the rule.
                        if not _is_complete_phrase_rect(page, rect):
                            continue
                        # Skip if an existing link annotation already covers
                        # this span — handles dedup across the multiple
                        # spellings ("Exhibit 1" / "EXHIBIT 1" both find the
                        # same glyph rect on a case-insensitive search) and
                        # also avoids stomping a TOC entry that happens to
                        # read "Exhibit 5 — Police Report".
                        already = False
                        for el in existing_links:
                            er = el.get("from")
                            if er and rect.intersects(er):
                                already = True
                                break
                        if already:
                            continue
                        page.insert_link({
                            "kind": fitz.LINK_GOTO,
                            "from": rect,
                            "page": target,
                            "to": fitz.Point(0, 0),
                        })
                        underline_y = rect.y1 - 0.5
                        page.draw_line(
                            fitz.Point(rect.x0, underline_y),
                            fitz.Point(rect.x1, underline_y),
                            color=LINK_COLOUR,
                            width=1.0,
                        )
                        linked += 1
                        existing_links.append({"from": rect,
                                               "kind": fitz.LINK_GOTO})

    if linked:
        log.info(f"  Linked {linked} exhibit reference(s) to exhibit pages "
                 f"({len(cover_map)} exhibits detected)")
    return linked, cover_map


# ────────────────────────────────────────────────────────────────────────────
# Bookmark (outline) builder
# ────────────────────────────────────────────────────────────────────────────
# Produce a PDF outline (the sidebar bookmarks Acrobat shows) from the same
# data the link passes already collected:
#   * Contents   <- TOC entries detected by _link_toc_entries
#   * Exhibits   <- exhibit cover pages from _link_exhibit_references
#   * Paragraphs <- per-page first-new-paragraph anchors from
#                   _detect_paragraph_anchors
#
# The outline is set via doc.set_toc(), which *replaces* any pre-existing
# outline rather than appending. Two consequences:
#   * Idempotency is free — re-runs produce the same tree, no stacking.
#   * Bookmarks the source PDF came with are intentionally discarded, in
#     line with the policy decision to use a fresh, predictable outline
#     derived from the detected structure.
#
# Only the branches with content appear. A simple brief with no exhibits
# and no pleading-paper paragraphs gets just "Contents"; a declaration
# gets "Exhibits" and "Paragraphs"; a document with none of the three
# yields no outline change at all (set_toc is not called).


def _exhibit_sort_key(ident: str):
    """Order exhibits as a human would expect: numerics ascending first,
    then letters alphabetically. ("1", "2", "12", "A", "B", "AA")."""
    if ident.isdigit():
        return (0, int(ident), "")
    return (1, 0, ident.upper())


# ────────────────────────────────────────────────────────────────────────────
# Document-footer detection (for combined-filing splitter bookmarks)
# ────────────────────────────────────────────────────────────────────────────
# Many filings in Zachary's workflow are combined PDFs: a Motion followed by
# its Memorandum, a Declaration, and exhibits, all concatenated into one
# document. Each sub-document typically has its own running footer in the
# bottom margin (the document title, sometimes with a page number). The
# transition between sub-documents shows up as a footer change.
#
# This pass finds the *first* page of each unique footer and produces one
# bookmark per sub-document. Single-document PDFs (only one footer text
# across the entire file) get nothing — there's no value in a "Documents"
# branch with one entry; the user can navigate to page 1 trivially.
#
# Normalization for comparison:
#   * Bottom 72pt band only.
#   * Pleading-paper line-number gutter digits (lone "27", "28") are
#     filtered out before joining the band text.
#   * Trailing page-number tokens are stripped (" — 5", "- 12", "5.",
#     bare " 5" at end). This handles "OPPOSITION TO MSJ — 5" → "OPPOSITION
#     TO MSJ" across pages with different page numbers.
#   * Case-folded + whitespace-collapsed for the comparison key. The
#     bookmark label uses the cleaned text in its original case (the first
#     occurrence's casing).
#
# Behavioural rules (from user spec):
#   * Same footer appearing later in the file does NOT create a second
#     bookmark — first-appearance only.
#   * 1-page sub-documents (a unique footer that appears on exactly one
#     page) ARE bookmarked.
#   * 1-page footer "flicker" (footer A → footer B → footer A on
#     consecutive pages) is suppressed: the lone B page is treated as
#     noise and does NOT start a new document. This handles OCR/scan
#     artifacts on a single page without losing genuine 1-page sub-docs
#     (a real 1-page sub-doc doesn't revert to the prior footer
#     immediately after).
#   * Pages with no detectable footer inherit the previous page's footer
#     for grouping purposes.
#   * The whole pass is gated on ≥2 distinct footers in the document.

_FOOTER_BAND_PT = 50.0

# Minimum x-coordinate at which footer content can begin. Anything whose
# bbox sits entirely to the left of this threshold is considered to be in
# the left margin (pleading-paper line-number gutter or vertical firm
# letterhead) and excluded from the footer text. 90pt clears both:
#   * The line-number gutter (typically x ≈ 50-85pt on standard
#     pleading paper).
#   * Vertical firm letterheads ("OGLETREE, DEAKINS, NASH, SMOAK &
#     STEWART, P.C.") which run in the left margin and are read by
#     PyMuPDF as ordinary text spans with x_right ≈ 80pt or less.
# Standard 8.5x11 briefs and declarations put their body content (and
# any centered/right-aligned footer text) well to the right of 90pt.
_FOOTER_LEFT_MARGIN_PT = 90.0


def _normalize_footer_for_key(s: str) -> str:
    """Produce the comparison key for a footer string: strip every
    standalone digit-shaped token (page numbers, gutter digits, "of"-
    counters), normalize quote and dash variants, collapse whitespace,
    uppercase. Returns empty string if nothing meaningful remains.

    Standalone digit tokens are stripped wherever they occur — not
    just at the trailing edge — because real-world footers can have
    page numbers leading ("7 EXHIBIT A"), trailing ("OPPOSITION — 5"),
    or both ("Page 1 of 2 9 EXHIBIT A"). Any of these must normalize
    to the same key for grouping to work.

    Quote and dash normalization: page-to-page PDF text extraction
    can render the same source string with different code points
    depending on the font (curly U+2019 apostrophe on one page vs.
    straight U+0027 on another, en-dash vs. em-dash vs. hyphen).
    Without folding these to a common form, "DEFENDANTS' MOTION" on
    page 1 (curly) and "DEFENDANTS' MOTION" on page 5 (straight) would
    produce different keys and be detected as two separate documents.
    Labels (in _clean_footer_label) keep the original characters; only
    the comparison key is folded.

    What counts as a "digit token" here:
      * A bare 1-4 digit run, optionally preceded by a dash, "Page",
        "PAGE", or "p." and optionally followed by ".".
      * The literal word "of" between two digit tokens ("5 of 12").
    """
    if not s:
        return ""
    # Fold quote variants to ASCII apostrophe, dash variants to ASCII
    # hyphen-minus. This only affects the key, not the displayed label.
    cleaned = s.translate(_FOOTER_KEY_FOLD_MAP)
    # Strip "N of M" forms first ("5 of 12", "Page 1 of 2").
    cleaned = re.sub(
        r"(?:Page|PAGE|p\.)?\s*\d{1,4}\s+of\s+\d{1,4}",
        " ", cleaned, flags=re.IGNORECASE,
    )
    # Strip remaining standalone digit-shaped tokens with optional
    # leading dash/Page/p. and optional trailing period or dash. The
    # word boundary on both sides keeps "1281.2" (statute section) and
    # "12,345" (an unlikely but conceivable footer figure) intact —
    # we only match isolated 1-4 digit numbers, not numbers embedded
    # in a larger word. The trailing-dash allowance handles stylized
    # page indicators like "- 4 -" where the digit is bracketed by
    # dashes; without it, the leading dash gets consumed but the
    # trailing dash survives and corrupts the key.
    cleaned = re.sub(
        r"(?:^|\s)"
        r"(?:-\s*|Page\s+|PAGE\s+|p\.\s*)?"
        r"\d{1,4}"
        r"\s*[-.]?"
        r"(?=\s|$)",
        " ", cleaned, flags=re.IGNORECASE,
    )
    cleaned = cleaned.strip()
    if not cleaned:
        return ""
    return re.sub(r"\s+", " ", cleaned).upper()


# Maps used by _normalize_footer_for_key to fold typographic variants
# (curly quotes, fancy dashes) to ASCII equivalents BEFORE digit-token
# stripping. PDFs from different page sources can render the same
# source text with different Unicode code points; folding ensures the
# comparison key is stable across pages even when the underlying glyph
# code points differ. The label-cleanup pass uses no fold so the
# bookmark text preserves what the PDF actually contains.
_FOOTER_KEY_FOLD_MAP = str.maketrans({
    # Apostrophes / single quotes
    "\u2018": "'", "\u2019": "'",
    "\u201A": "'", "\u201B": "'",
    # Double quotes
    "\u201C": '"', "\u201D": '"',
    "\u201E": '"', "\u201F": '"',
    # Dash variants
    "\u2010": "-", "\u2011": "-",
    "\u2012": "-", "\u2013": "-",
    "\u2014": "-", "\u2015": "-",
    "\u2212": "-",
})


def _clean_footer_label(s: str) -> str:
    """Produce the bookmark label: same digit-stripping as the key,
    but preserve the original case and apply minimal cleanup.
    """
    if not s:
        return ""
    cleaned = s
    cleaned = re.sub(
        r"(?:Page|PAGE|p\.)?\s*\d{1,4}\s+of\s+\d{1,4}",
        " ", cleaned, flags=re.IGNORECASE,
    )
    cleaned = re.sub(
        r"(?:^|\s)"
        r"(?:[-\u2010\u2011\u2012\u2013\u2014\u2015]\s*|Page\s+|PAGE\s+|p\.\s*)?"
        r"\d{1,4}"
        r"\s*[-\u2010\u2011\u2012\u2013\u2014\u2015.]?"
        r"(?=\s|$)",
        " ", cleaned, flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _extract_footer_text(page) -> str:
    """Return the joined footer text from the bottom _FOOTER_BAND_PT of
    the page, or empty string if no footer text is present.

    Works at the SPAN level rather than the row level. PyMuPDF's row
    merger ('_collect_rows') joins all spans on a shared baseline,
    which on pleading-paper layouts splices vertical-letterhead text
    in the left margin together with the real centered footer text
    on the same y. To avoid that, we walk individual spans inside
    the footer y-band and drop any span whose bbox sits entirely
    inside the left margin (x_right < _FOOTER_LEFT_MARGIN_PT). That
    cleanly excludes:
      * Pleading-paper gutter line-number digits ("28").
      * Vertical firm letterhead text ("OGLETREE, DEAKINS, NASH,
        SMOAK & STEWART, P.C.").

    Remaining spans are sorted by visual reading order (top-to-bottom,
    then left-to-right) and joined with a single space. The page-
    number stripping in _normalize_footer_for_key handles any page
    labels (leading or trailing) baked into the joined text.
    """
    try:
        d = page.get_text("dict")
    except Exception:
        return ""
    page_height = page.rect.height
    band_top = page_height - _FOOTER_BAND_PT

    # Collect surviving spans as (y_mid, x0, text) for sorting.
    spans = []
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for sp in line.get("spans", []):
                t = sp.get("text", "").strip()
                if not t:
                    continue
                bbox = sp.get("bbox")
                if not bbox:
                    continue
                x0, y0, x1, y1 = bbox
                y_mid = (y0 + y1) / 2
                if y_mid < band_top:
                    continue
                # Drop spans entirely inside the left margin
                # (firm letterheads, gutter line numbers).
                if x1 < _FOOTER_LEFT_MARGIN_PT:
                    continue
                # Drop lone digit tokens (lone bare page numbers) outright
                # — these otherwise survive the span filter and get joined
                # mid-string into the comparison key.
                if re.fullmatch(r"\d{1,4}", t):
                    continue
                spans.append((y_mid, x0, t))
    if not spans:
        return ""
    # Sort by y first (top-to-bottom of the footer band), then x.
    spans.sort(key=lambda s: (s[0], s[1]))
    return " ".join(s[2] for s in spans)


# ────────────────────────────────────────────────────────────────────────────
# Fuzzy footer-key matching
# ────────────────────────────────────────────────────────────────────────────
# Real-world PDF text extraction produces footer text that varies slightly
# from page to page within the same document, even after the
# digit-stripping / quote-folding normalization in _normalize_footer_for_key.
# Sources of variation:
#   * Court reporters / e-discovery vendors stamp watermark text into the
#     bottom margin ("STENO.COM (888) 707-8366", "Page 234", "YVer1f"
#     anti-piracy fingerprints). These tokens are unique per page and
#     would split a single exhibit into N sub-documents in strict matching.
#   * Stylized page indicators like "- 4 -" leave punctuation residue
#     even after digit stripping.
#   * Optional appendix labels appear only on some pages ("Exhibit A
#     Page 1 of 6" on the first page, "Page · Exhibit A Page 5 of 6" on
#     a later page).
# To handle these, we compute token-set Jaccard similarity between keys
# and treat any pair above _FOOTER_KEY_SIMILARITY_THRESHOLD as the same
# document.
#
# Safeguards against over-merging:
#   * Short footers (< _FOOTER_MIN_TOKENS_FOR_FUZZY significant tokens)
#     are compared by exact equality only — fuzzy matching on tiny token
#     sets has too much false-positive risk ("EXHIBIT A" vs "EXHIBIT B"
#     would share 1 of 2 tokens, 0.33 Jaccard, so they wouldn't merge
#     anyway, but the policy makes the intent explicit and protects
#     against future regressions).
#   * Stopwords and short tokens are dropped before comparison, so a
#     footer of mostly noise can't accidentally match a real title
#     just because both contain "for the of".
#   * First-seen-key wins: when a new page key fuzzy-matches a key we've
#     already seen, the new page inherits the prior key's identity. It
#     does NOT update the canonical key text — the canonical key stays
#     as whatever was first detected for that document.

_FOOTER_KEY_SIMILARITY_THRESHOLD = 0.7
_FOOTER_MIN_TOKENS_FOR_FUZZY = 3

# Stopwords excluded from token-set comparisons. Lowercase. The intent is
# to eliminate function words that contribute nothing to identity, so a
# single substantive token disagreement weighs more in the similarity
# score. Conservative list — only words that appear in essentially every
# title/footer in California legal practice. We intentionally do NOT
# include single-letter articles "a"/"an": exhibit identifiers like
# "Exhibit A" rely on the letter to distinguish from "Exhibit B", and
# isolated articles are rare enough in footer text that keeping them
# doesn't hurt Jaccard discrimination.
_FOOTER_FUZZY_STOPWORDS = frozenset({
    "of", "the", "and", "in", "to", "for", "by",
    "on", "at", "or", "with", "from",
})


def _fuzzy_token_set(key: str) -> frozenset:
    """Return the set of significant tokens for fuzzy comparison.

    Lowercases, strips surrounding punctuation, then drops:
      * Empty tokens (length 0 after punctuation stripping).
      * Pure-digit tokens (page numbers, phone-number fragments, year
        stamps) — these carry no identity and would inflate Jaccard
        similarity between unrelated footers that happen to share
        watermark numerics.
      * Stopwords (see _FOOTER_FUZZY_STOPWORDS).
    Single-letter tokens ARE kept: they're how exhibit identifiers
    ("Exhibit A", "Exhibit B") stay distinguishable, and the Jaccard
    threshold of 0.7 prevents incidental shared single letters from
    over-merging unrelated footers.
    Returns a frozenset for hashability so callers can cache.
    """
    if not key:
        return frozenset()
    raw_tokens = key.lower().split()
    out = set()
    for tok in raw_tokens:
        # Strip surrounding punctuation. Keep internal apostrophes
        # and dots so "bernards'" or "steno.com" stay distinguishable.
        stripped = tok.strip(".,;:!?()[]{}\u2018\u2019\u201C\u201D\"'·")
        if not stripped:
            continue
        if stripped.isdigit():
            continue
        if stripped in _FOOTER_FUZZY_STOPWORDS:
            continue
        out.add(stripped)
    return frozenset(out)


def _fuzzy_keys_equivalent(key_a: str, key_b: str,
                           tokens_a=None, tokens_b=None) -> bool:
    """Return True if two normalized footer keys should be treated as
    the same document under the fuzzy-match policy.

    Two paths to "equivalent":
      * Identical keys (cheap fast path).
      * Token-set Jaccard ≥ _FOOTER_KEY_SIMILARITY_THRESHOLD, when both
        keys have ≥ _FOOTER_MIN_TOKENS_FOR_FUZZY significant tokens.
      * SUBSET COVERAGE: the short key's tokens are a non-empty
        subset of the long key's tokens, AND the short key has ≥ 2
        significant tokens, AND the long key has ≥
        _FOOTER_MIN_TOKENS_FOR_FUZZY significant tokens. This handles
        the case where one page renders only the core title
        ("EXHIBIT A") and another page adds watermark noise around
        the same title ("STENO.COM 707-8366 EXHIBIT A"). Without
        subset coverage, the short version would never reach the
        ≥3-token gate and the two would be split into separate
        documents.

    Pre-computed token sets may be passed via tokens_a/tokens_b to
    avoid recomputation in inner loops.
    """
    if key_a == key_b:
        return True
    if tokens_a is None:
        tokens_a = _fuzzy_token_set(key_a)
    if tokens_b is None:
        tokens_b = _fuzzy_token_set(key_b)

    # Subset-coverage path. Identify short/long by token count.
    if len(tokens_a) <= len(tokens_b):
        short, long_ = tokens_a, tokens_b
    else:
        short, long_ = tokens_b, tokens_a
    if (len(short) >= 2
            and len(long_) >= _FOOTER_MIN_TOKENS_FOR_FUZZY
            and short
            and short <= long_):
        return True

    # Strict Jaccard path — both keys must have enough tokens.
    if (len(tokens_a) < _FOOTER_MIN_TOKENS_FOR_FUZZY
            or len(tokens_b) < _FOOTER_MIN_TOKENS_FOR_FUZZY):
        return False
    union = tokens_a | tokens_b
    if not union:
        return False
    inter = tokens_a & tokens_b
    return len(inter) / len(union) >= _FOOTER_KEY_SIMILARITY_THRESHOLD


def _canonicalize_key(key: str, canon_index: list) -> str:
    """Map `key` to a canonical key from `canon_index` (a list of
    previously-seen keys with their pre-computed token sets) using the
    fuzzy-equivalence rule. If no canonical match exists, append the
    new key to canon_index and return it as-is. First-seen-key wins.

    canon_index is a list of (canonical_key, token_set) tuples and is
    mutated in place.
    """
    new_tokens = _fuzzy_token_set(key)
    for canon_key, canon_tokens in canon_index:
        if _fuzzy_keys_equivalent(key, canon_key,
                                   tokens_a=new_tokens,
                                   tokens_b=canon_tokens):
            return canon_key
    canon_index.append((key, new_tokens))
    return key


def _detect_document_footers(doc, log: logging.Logger):
    """Walk the document; for each transition to a new unique footer,
    record the page where the new footer first appears. Returns a list
    of (label, page_idx) tuples, or [] if fewer than 2 distinct footers
    are detected (single-document file — no value in a sub-document
    bookmark branch).

    Implementation notes:
      * Pages with no detectable footer inherit the previous page's
        footer key for grouping (they're "inside" the prior sub-doc).
      * 1-page flicker (A → B → A) is suppressed by buffering one page
        of look-ahead: a new key only commits if the next page either
        also has the new key, has no footer (still inside the new
        sub-doc), or has a *different* new key (i.e. moves forward to
        another sub-doc rather than reverting).
      * Same-key reappearance later in the file is ignored — we only
        bookmark the *first* occurrence of each key. Tracked via
        `seen_keys`.
    """
    if len(doc) < 2:
        # Single-page documents don't need a Documents branch.
        return []

    # First pass: per-page (raw_text, key). Empty key means "inherit
    # from previous page". Keys are canonicalized via fuzzy matching:
    # when a page's normalized key is "close enough" (≥70% Jaccard
    # similarity on significant tokens) to a previously-seen canonical
    # key, it inherits that canonical key. This handles per-page
    # variation introduced by court-reporter watermarks, stylized page
    # indicators, and other low-information noise that survives the
    # strict-normalization pass. See the comment block above
    # _fuzzy_token_set for the rationale.
    canon_index: list = []  # list of (canonical_key, token_set)
    per_page = []
    for page in doc:
        raw = _extract_footer_text(page)
        if raw:
            strict_key = _normalize_footer_for_key(raw)
            key = _canonicalize_key(strict_key, canon_index) if strict_key else ""
        else:
            key = ""
        per_page.append((raw, key))

    # Forward-fill empty keys from the previous non-empty key. Pages
    # before the first detected footer keep an empty key.
    filled = []
    prev_key = ""
    prev_raw = ""
    for raw, key in per_page:
        if key:
            filled.append((raw, key))
            prev_key = key
            prev_raw = raw
        else:
            # Inherit prior key (and reuse its raw text for label
            # purposes — won't be used unless this page becomes a
            # boundary, which it won't, since key matches prior).
            filled.append((prev_raw, prev_key))

    # Walk filled keys; collect first-appearance boundaries with the
    # 1-page-flicker guard. A boundary is committed when:
    #   * The page's key differs from the previously-committed key
    #     (or this is the first non-empty key in the file), AND
    #   * The new key has not been seen before, AND
    #   * Either: this is the last page, OR the next page's key is
    #     not equal to the previously-committed key (i.e. we don't
    #     immediately revert — that would be flicker).
    entries = []
    seen_keys = set()
    last_committed_key = ""
    n = len(filled)
    for i, (raw, key) in enumerate(filled):
        if not key:
            # Pages before any footer ever appears.
            continue
        if key == last_committed_key:
            continue
        if key in seen_keys:
            # Reappearance later in the file: per user spec, do NOT
            # create a second bookmark. We also DON'T update
            # last_committed_key — if the document flips back to an
            # earlier footer and then to yet another new one, the next
            # new one is still a fresh sub-document.
            last_committed_key = key
            continue
        # Flicker check: if the *next* page reverts to last_committed_key,
        # this page is a 1-page noise blip, not a real sub-document.
        # 1-page sub-docs are allowed only when they don't revert.
        next_key = filled[i + 1][1] if i + 1 < n else None
        if (last_committed_key and next_key == last_committed_key
                and key != last_committed_key):
            # Flicker: skip this page, don't commit, don't mark seen.
            # The next iteration will pick up the reverted key (but
            # last_committed_key already equals it, so it'll be a no-op).
            continue
        # Commit a new boundary.
        label = _clean_footer_label(raw)
        if not label:
            # Shouldn't happen — non-empty key implies non-empty raw —
            # but guard anyway.
            continue
        entries.append((label, i))
        seen_keys.add(key)
        last_committed_key = key

    # Gate: need at least 2 distinct footers for the branch to be useful.
    if len(entries) < 2:
        return []

    log.info(f"  Detected {len(entries)} sub-document(s) by unique footer")
    return entries


def _build_bookmark_tree(doc, toc_entries, exhibit_cover_map,
                        paragraph_anchors, cause_entries=None,
                        document_entries=None, section_entries=None,
                        toc_page_range=None):
    """Assemble the [level, title, page_1_based] list that doc.set_toc
    expects. Returns the list, or None if there's nothing worth adding.

    toc_entries:        [(label, target_page_index), ...]
    exhibit_cover_map:  {ident: [page_indices]}
    paragraph_anchors:  [(page_index, paragraph_num), ...]
    cause_entries:      [(label, target_page_index), ...] — ordered by
                        ordinal number, one entry per distinct cause of
                        action. Only populated for complaint files.
    document_entries:   [(label, target_page_index), ...] — one entry per
                        sub-document detected by unique footer text, in
                        page order. Populated for combined filings where
                        multiple sub-docs share one PDF (e.g. Motion +
                        Declaration + Exhibits concatenated together).
    section_entries:    [(label, page_index), ...] — section headings
                        detected by _detect_section_headings, in
                        document order. Only populated for briefs (the
                        filename-skip gate excludes declarations and
                        complaints).
    toc_page_range:     (start, end_inclusive) page-index range of the
                        TOC itself, or None. The Contents branch header
                        is pointed at this range's start page so
                        clicking "Contents" navigates to the actual
                        TOC page rather than the first entry's target.

    Hierarchy rules (from user spec):
      * Contents and Causes of Action are flat top-level branches.
      * Exhibits and Documents are top-level branches whose children are
        the individual exhibit / sub-document entries.
      * Sections (when present) nest under whichever Document contains
        their page. If no Documents exist, sections become a top-level
        "Sections" branch. Sections inside an exhibit's page range are
        dropped — exhibits don't get section bookmarks.
      * Paragraphs nest under whichever Section, Document, or Exhibit
        owns their page, in that priority order. A paragraph at page P
        nests under the Section that contains P if any; else under the
        Document; else under the Exhibit; else (no parent) appears in
        the top-level fallback Paragraphs branch.
      * Exhibits take precedence over Documents on contested pages:
        once the document enters its first exhibit's page range, all
        subsequent paragraphs nest under exhibits regardless of any
        footer-detected sub-document that may continue.
      * A sub-document whose start page falls inside an exhibit's page
        range becomes a CHILD of that exhibit, not a top-level Documents
        entry. Its paragraphs then nest under it (one level deeper).
      * Inside an exhibit with both direct paragraphs (no inner sub-doc)
        AND a child sub-document, the direct paragraphs come first and
        the sub-document(s) follow. (Within each group, page order.)
      * Redundant nested labels suppressed: a sub-document inside
        Exhibit N whose cleaned label is just "Exhibit N" / "EXHIBIT N" /
        "Exhibit N — something-trivial" is dropped — the exhibit
        bookmark already covers it.
      * If paragraphs exist but no Documents/Exhibits/Sections do, the
        Paragraphs branch falls back to a top-level flat list so
        standalone declarations still get paragraph navigation.
    """
    n_pages = len(doc)
    tree = []

    # ── Contents (flat, top-level) ──────────────────────────────────────
    valid_toc = [(lbl, p) for (lbl, p) in toc_entries if 0 <= p < n_pages]
    if valid_toc:
        # Branch header points at the TOC's own page (where "TABLE OF
        # CONTENTS" appears), not at the first entry's target. So
        # clicking "Contents" in the bookmark sidebar navigates to the
        # TOC page itself. If toc_page_range is unavailable for any
        # reason, fall back to the first entry's target page.
        contents_header_page = (toc_page_range[0] + 1
                                if toc_page_range
                                else valid_toc[0][1] + 1)
        tree.append([1, "Contents", contents_header_page])
        for label, page_idx in valid_toc:
            tree.append([2, label, page_idx + 1])

    # ── Causes of Action (flat, top-level) ──────────────────────────────
    valid_causes = [(lbl, p) for (lbl, p) in (cause_entries or [])
                    if 0 <= p < n_pages]
    if valid_causes:
        tree.append([1, "Causes of Action", valid_causes[0][1] + 1])
        for label, page_idx in valid_causes:
            tree.append([2, label, page_idx + 1])

    # ── Compute page-range ownership for Documents and Exhibits ─────────
    # Each document/exhibit owns a contiguous range [start, end_inclusive].
    # Documents end where the next document starts OR where the first
    # exhibit cover begins (exhibits take precedence).
    # Exhibits end where the next exhibit cover begins.
    valid_docs = sorted(
        [(lbl, p) for (lbl, p) in (document_entries or [])
         if 0 <= p < n_pages],
        key=lambda x: x[1],
    )
    exhibits_sorted = sorted(
        ((ident, pages[0]) for ident, pages in (exhibit_cover_map or {}).items()
         if pages and 0 <= pages[0] < n_pages),
        key=lambda x: (x[1], _exhibit_sort_key(x[0])),
    )

    # The page where exhibits start (exhibits take precedence past here).
    first_exhibit_page = exhibits_sorted[0][1] if exhibits_sorted else n_pages

    # Build exhibit ranges: each exhibit owns from its cover page to one
    # before the next exhibit's cover page. The LAST exhibit normally
    # runs to the end of the document, but there's one important case
    # where it must end sooner: a TRAILING POST-EXHIBIT DOCUMENT, like
    # a Proof of Service appearing after Exhibit C. Without capping,
    # such a document would be nested under the last exhibit (just
    # because nothing else followed it) instead of being a top-level
    # peer.
    #
    # To distinguish a trailing post-exhibit document from a genuine
    # sub-document INSIDE an exhibit (e.g. Exhibit 5 is itself a
    # declaration with its own "Declaration of Smith" footer), we
    # require a confirmation signal that the exhibit had its own
    # running footer text: at least one footer-detected document inside
    # the tentative exhibit range whose cleaned label is redundant
    # with the exhibit's identifier (e.g. footer "EXHIBIT C" detected
    # on pages within Exhibit C's range). The presence of this
    # redundant entry means there's a clear "this is still the
    # exhibit" footer marker, so when the footer changes to a non-
    # redundant label, we know the exhibit's content has ended.
    #
    # Without this signal — e.g. exhibits whose content pages have no
    # footer at all — we don't cap, and any sub-document detected
    # inside the tentative range gets nested as before.
    exhibit_ranges = []  # list of (ident, start_idx, end_inclusive)
    for i, (ident, start) in enumerate(exhibits_sorted):
        if i + 1 < len(exhibits_sorted):
            end = exhibits_sorted[i + 1][1] - 1
        else:
            end = n_pages - 1
            # Tentative range is [start, n_pages-1]. Check whether any
            # footer-detected document inside this range has a redundant
            # label — i.e. the exhibit has its own running footer.
            has_redundant_marker = any(
                start <= d_start <= end
                and _label_is_just_exhibit_id(d_lbl, ident)
                for d_lbl, d_start in valid_docs
            )
            if has_redundant_marker:
                # Look for a post-exhibit, non-redundant Document to
                # cap at.
                for d_lbl, d_start in valid_docs:
                    if d_start <= start:
                        continue
                    if _label_is_just_exhibit_id(d_lbl, ident):
                        continue
                    end = d_start - 1
                    break
        exhibit_ranges.append((ident, start, end))

    # Partition documents: top-level vs. nested-under-exhibit. A document
    # whose start page is within some exhibit's range becomes that
    # exhibit's child; otherwise it's top-level.
    def _containing_exhibit(page_idx):
        for ident, start, end in exhibit_ranges:
            if start <= page_idx <= end:
                return ident
        return None

    top_level_docs = []           # [(label, start_idx, end_inclusive), ...]
    nested_docs_by_exhibit = {}   # {ident: [(label, start, end_inclusive)]}
    for i, (lbl, start) in enumerate(valid_docs):
        # Range end of this document candidate: bounded by the next
        # document's start AND by the first exhibit page (documents
        # cede to exhibits at that boundary).
        next_doc_start = (valid_docs[i + 1][1]
                          if i + 1 < len(valid_docs) else n_pages)
        end = min(next_doc_start, first_exhibit_page, n_pages) - 1
        containing = _containing_exhibit(start)
        if containing is not None:
            # Nested under the exhibit. The end is bounded by the
            # exhibit's own end too — a sub-document can't extend past
            # its parent exhibit.
            for ex_ident, ex_start, ex_end in exhibit_ranges:
                if ex_ident == containing:
                    # The nested doc's end is bounded by the NEXT nested
                    # doc in the same exhibit (if any) OR the exhibit's
                    # own end, whichever comes first.
                    next_in_same_ex = None
                    for j in range(i + 1, len(valid_docs)):
                        _, p_next = valid_docs[j]
                        if ex_start <= p_next <= ex_end:
                            next_in_same_ex = p_next
                            break
                    nested_end = (next_in_same_ex - 1
                                  if next_in_same_ex is not None
                                  else ex_end)
                    # Redundant-label suppression: if the cleaned label
                    # is essentially just the exhibit identifier
                    # ("Exhibit N", "EXHIBIT N", "Ex. N"), drop it.
                    if _label_is_just_exhibit_id(lbl, ex_ident):
                        break
                    nested_docs_by_exhibit.setdefault(ex_ident, []).append(
                        (lbl, start, nested_end)
                    )
                    break
        else:
            # Top-level document. Its end is bounded by the first
            # exhibit page (exhibits take precedence past that point).
            top_level_docs.append((lbl, start, end))

    # ── Section partitioning ────────────────────────────────────────────
    # Each section owns a contiguous range from its start page to one
    # before the next section's start (or to the end of its containing
    # Document, whichever comes first). Sections inside exhibit ranges
    # are dropped — exhibits don't carry sections. Sections that fall
    # outside any Document range, when no Documents exist at all,
    # become top-level under a "Sections" branch.
    valid_sections = sorted(
        [(lbl, p) for (lbl, p) in (section_entries or [])
         if 0 <= p < n_pages],
        key=lambda x: x[1],
    )
    # Drop sections inside any exhibit range.
    valid_sections = [(lbl, p) for (lbl, p) in valid_sections
                      if _containing_exhibit(p) is None]

    # Map each section to its containing Document (or None if it falls
    # outside all Document ranges — possible when no Documents exist).
    # Compute each section's page-range end: bounded by the next section
    # in the same parent (or globally if top-level), and by the parent
    # Document's end if any.
    sections_in_doc = {}  # doc_index_in_top_level_docs -> [(lbl, start, end)]
    top_level_sections = []  # [(lbl, start, end), ...] when no Documents

    def _containing_doc_idx(page_idx):
        for di, (_, dstart, dend) in enumerate(top_level_docs):
            if dstart <= page_idx <= dend:
                return di
        return None

    for si, (lbl, start) in enumerate(valid_sections):
        # End of section: next section's start - 1, capped by containing
        # parent's end.
        next_section_start = (valid_sections[si + 1][1]
                              if si + 1 < len(valid_sections)
                              else n_pages)
        di = _containing_doc_idx(start)
        if di is not None:
            parent_end = top_level_docs[di][2]
            # Only allow next-section bound if that section is in the
            # same parent Document. Otherwise this section runs to
            # parent_end.
            if next_section_start - 1 > parent_end:
                end = parent_end
            else:
                # Check that the next section is still in this Document
                # (if not, this section runs to parent_end).
                next_di = _containing_doc_idx(next_section_start)
                if next_di == di:
                    end = next_section_start - 1
                else:
                    end = parent_end
            sections_in_doc.setdefault(di, []).append((lbl, start, end))
        else:
            # No containing Document — only relevant when there are no
            # Documents at all. If Documents exist but this section
            # falls outside them all, drop the section.
            if not top_level_docs:
                end = min(next_section_start - 1, n_pages - 1)
                # Cap by first exhibit if any (sections don't extend
                # into exhibit territory).
                if exhibits_sorted:
                    end = min(end, first_exhibit_page - 1)
                if end >= start:
                    top_level_sections.append((lbl, start, end))

    # ── Paragraphs: filter to valid page range; dedup happens per-range ─
    # The source list records the lowest paragraph number that *starts*
    # on each page, so duplicates within one sub-document shouldn't
    # occur. Global dedup would be incorrect because paragraph numbers
    # reset at each sub-document boundary (every declaration starts at
    # ¶ 1). We dedup inside _paragraphs_in_range so each range gets a
    # fresh dedup window.
    valid_paras = [(p_idx, num) for (p_idx, num) in (paragraph_anchors or [])
                   if 0 <= p_idx < n_pages]

    # Helper: emit the paragraphs that fall within [start, end_inclusive]
    # at a given indent level. Paragraphs are filtered by page range and
    # deduplicated against consecutive identical numbers WITHIN this range.
    def _paragraphs_in_range(start, end_inclusive):
        out = []
        last_num = None
        for (p_idx, num) in valid_paras:
            if not (start <= p_idx <= end_inclusive):
                continue
            if num == last_num:
                continue
            out.append((p_idx, num))
            last_num = num
        return out

    # Helper: emit paragraphs in [start, end_inclusive] that do NOT fall
    # inside any of the given sub-ranges (which represent child
    # sub-document or section ranges that will get their paragraphs
    # under their own bookmarks).
    def _paragraphs_in_range_excluding(start, end_inclusive, sub_ranges):
        def _in_any(p_idx):
            return any(s <= p_idx <= e for s, e in sub_ranges)
        return [(p_idx, num) for (p_idx, num)
                in _paragraphs_in_range(start, end_inclusive)
                if not _in_any(p_idx)]

    # ── Documents branch (top-level) ────────────────────────────────────
    if top_level_docs:
        tree.append([1, "Documents", top_level_docs[0][1] + 1])
        for di, (lbl, start, end) in enumerate(top_level_docs):
            tree.append([2, lbl, start + 1])
            sections_here = sections_in_doc.get(di, [])
            section_ranges = [(s, e) for (_, s, e) in sections_here]
            # Direct paragraphs of the Document: those not under any
            # section in this Document.
            for p_idx, num in _paragraphs_in_range_excluding(
                    start, end, section_ranges):
                tree.append([3, f"\u00b6 {num}", p_idx + 1])
            # Sections under this Document, each with its paragraphs.
            for slbl, s_start, s_end in sections_here:
                tree.append([3, slbl, s_start + 1])
                for p_idx, num in _paragraphs_in_range(s_start, s_end):
                    tree.append([4, f"\u00b6 {num}", p_idx + 1])

    # ── Sections branch (top-level, when no Documents) ──────────────────
    # Sections become a top-level branch only when there are no
    # Documents to nest them under. This is the standalone-brief case
    # — single document with detected section headings.
    if not top_level_docs and top_level_sections:
        tree.append([1, "Sections", top_level_sections[0][1] + 1])
        for slbl, s_start, s_end in top_level_sections:
            tree.append([2, slbl, s_start + 1])
            for p_idx, num in _paragraphs_in_range(s_start, s_end):
                tree.append([3, f"\u00b6 {num}", p_idx + 1])

    # ── Exhibits branch (top-level, with optional nested docs) ──────────
    if exhibits_sorted:
        tree.append([1, "Exhibits", exhibits_sorted[0][1] + 1])
        for ident, ex_start, ex_end in exhibit_ranges:
            tree.append([2, f"Exhibit {ident}", ex_start + 1])
            # Direct paragraphs of this exhibit: paragraphs whose page is
            # within the exhibit's range but NOT within any nested doc's
            # range. Per user spec, these come BEFORE the nested
            # sub-documents in the bookmark tree.
            nested = nested_docs_by_exhibit.get(ident, [])
            nested_page_ranges = [(n_start, n_end)
                                  for (_, n_start, n_end) in nested]

            direct_paras = _paragraphs_in_range_excluding(
                ex_start, ex_end, nested_page_ranges)
            for p_idx, num in direct_paras:
                tree.append([3, f"\u00b6 {num}", p_idx + 1])
            # Nested sub-documents and their paragraphs.
            for nlbl, n_start, n_end in nested:
                tree.append([3, nlbl, n_start + 1])
                for p_idx, num in _paragraphs_in_range(n_start, n_end):
                    tree.append([4, f"\u00b6 {num}", p_idx + 1])

    # ── Paragraphs top-level fallback ───────────────────────────────────
    # If neither Documents nor Exhibits nor Sections exist but there
    # ARE paragraphs, show them as a flat top-level branch — preserves
    # the standalone-declaration use case.
    if (valid_paras and not top_level_docs and not exhibits_sorted
            and not top_level_sections):
        all_paras = _paragraphs_in_range(0, n_pages - 1)
        if all_paras:
            tree.append([1, "Paragraphs", all_paras[0][0] + 1])
            for page_idx, num in all_paras:
                tree.append([2, f"\u00b6 {num}", page_idx + 1])

    return tree if tree else None


# Match labels that are essentially just an exhibit identifier — used to
# suppress redundant nested bookmarks inside the exhibit they belong to.
# "Exhibit 5", "EXHIBIT 5", "Ex. 5", "Exhibit B", "EXHIBIT AA". Allows
# optional trailing punctuation. The label has ALREADY been stripped of
# its page-number tail by _clean_footer_label upstream, so this is
# purely a label-shape match.
_LABEL_JUST_EXHIBIT_RE = re.compile(
    r"^\s*(?:Exhibit|EXHIBIT|Ex\.?|EX\.?|Exh\.?|EXH\.?)\s+([A-Za-z0-9]+)\s*[.:]?\s*$"
)


def _label_is_just_exhibit_id(label: str, ident: str) -> bool:
    """Return True if `label` (a sub-document label) is essentially just
    the identifier of `ident` (an exhibit identifier). Used to suppress
    redundant nested bookmarks.

    Numeric idents are compared numerically (so "Exhibit 5" matches
    ident "5"); letter idents are compared case-insensitively.
    """
    m = _LABEL_JUST_EXHIBIT_RE.match(label or "")
    if not m:
        return False
    label_id = m.group(1)
    if ident.isdigit() and label_id.isdigit():
        return int(label_id) == int(ident)
    return label_id.upper() == ident.upper()


def _set_bookmarks(doc, toc_entries, exhibit_cover_map, paragraph_anchors,
                   log: logging.Logger, cause_entries=None,
                   document_entries=None, section_entries=None,
                   toc_page_range=None):
    """Build and apply the bookmark outline. No-op if all inputs are
    empty. Failure is non-fatal — the function logs and returns."""
    try:
        tree = _build_bookmark_tree(doc, toc_entries, exhibit_cover_map,
                                    paragraph_anchors, cause_entries,
                                    document_entries, section_entries,
                                    toc_page_range)
        if not tree:
            return
        doc.set_toc(tree)
        # Count second-level entries (the actual bookmarks; level 1 are
        # the branch headers) for the log line.
        n_leaves = sum(1 for entry in tree if entry[0] == 2)
        branches = [entry[1] for entry in tree if entry[0] == 1]
        log.info(f"  Bookmarks: {n_leaves} entry(ies) across "
                 f"{', '.join(branches)}")
    except Exception as e:
        log.warning(f"  Bookmark generation failed (non-fatal): {e}")


# ────────────────────────────────────────────────────────────────────────────
# Pseudonymization (applied to the .txt export only, never the PDF)
# ────────────────────────────────────────────────────────────────────────────
# Ported from the standalone PDF-Redactor's pseudonymize.py, but retargeted at
# plain text: every party name, case number, attorney name, and regex-detected
# identifier (e-mail, phone, SSN, street address) is swapped for a STABLE,
# realistic fake so the exported text stays readable and coherent while never
# naming the real parties. "Stable" = deterministic: the same real value always
# maps to the same fake (seed-per-value), so roles stay consistent across pages
# and files, and a key spreadsheet lets a downstream step map a draft back to
# reality. Opt-in (--pseudonymize / --key / --term); the PDF is untouched.
import hashlib as _pn_hashlib
import random as _pn_random
import unicodedata as _pn_unicodedata

_NFKC = lambda s: _pn_unicodedata.normalize("NFKC", s)


def _pn_rng(*parts):
    key = "\x1f".join(str(p) for p in parts).encode("utf-8")
    return _pn_random.Random(int(_pn_hashlib.sha256(key).hexdigest(), 16))


# Neutral name-word pool, shared by given names and surnames so a token maps to
# one fake regardless of role ("Smith" in "John Smith" and a bare "Smith" both
# resolve to the same stand-in).
_PN_NAME_WORDS = [
    "Ashford", "Bennett", "Calder", "Danforth", "Ellery", "Fenwick", "Garrick",
    "Halloran", "Ingram", "Jarrett", "Keswick", "Langley", "Marlowe", "Nash",
    "Orwell", "Prescott", "Quill", "Radley", "Sable", "Thorne", "Underwood",
    "Vance", "Whitlock", "Yardley", "Ashby", "Brandt", "Corwin", "Delacroix",
    "Everts", "Fairfax", "Grantham", "Holloway", "Isley", "Jennings", "Kingsley",
    "Lathrop", "Merrick", "Norwood", "Ackerly", "Bramble", "Colfax", "Denning",
    "Emmett", "Forsythe", "Gable", "Hendry", "Ivers", "Joplin", "Kessler",
    "Lorne", "Mabry", "Nolan", "Ondine", "Pruett", "Renwick", "Sterling",
    "Tolliver", "Ursin", "Verity", "Waverly", "Alden", "Beaumont", "Carrow",
    "Delane",
    # The pool is drawn without replacement per case; a real filing carries far
    # more than the original 64 distinct name tokens (parties, counsel, staff,
    # every declarant and e-mail display name), so it was exhausting and minting
    # ugly numbered stand-ins ("Ashford2", "Corwin Vance3", and in body text
    # "HENDRY2 CORPORATIOLORNE10"). These keep the pool ahead of a large case.
    "Abernathy", "Alcott", "Amberly", "Ashcombe", "Atwater", "Balfour",
    "Bancroft", "Barlowe", "Bexley", "Blackwood", "Braddock", "Brimley",
    "Cadwell", "Calloway", "Carden", "Cartwright", "Chadwick", "Chamberlin",
    "Chetwood", "Clarendon", "Cleary", "Cranston", "Cresswell", "Darrow",
    "Davenport", "Deverell", "Doran", "Dunhill", "Eastwick", "Edgerton",
    "Ellsworth", "Fairbank", "Fallon", "Farraday", "Fenmore", "Finnegan",
    "Gaskell", "Gearhart", "Goddard", "Hadley", "Halstead", "Hargrove",
    "Hartwell", "Hawkridge", "Hemsley", "Hollis", "Huxley", "Ingersoll",
    "Jarrow", "Kimball", "Kinsley", "Larkin", "Ledbetter", "Linford",
    "Lockwood", "Ludlow", "Mallory", "Mansfield", "Marsden", "Mayhew",
    "Milburn", "Montrose", "Mowbray", "Oakley", "Ormsby", "Paget",
    "Parrish", "Pemberton", "Penrose", "Prentiss", "Quenby", "Ramsey",
    "Rathbone", "Redmond", "Ridley", "Rockwell", "Rutherford", "Sackett",
    "Selwyn", "Sheridan", "Sinclair", "Stanhope", "Stockton", "Swinton",
    "Thorpe", "Trafford", "Tremaine", "Underhill", "Vickers", "Wadsworth",
    "Waldron", "Warwick", "Wescott", "Wexford", "Whitby", "Winslow",
    "Wolcott", "Wycliffe", "Yates", "Yorke", "Ainsworth", "Braxton",
    "Denholm", "Harrell", "Kimbrell", "Lassiter", "Mercer", "Northcott",
    "Ravenscar", "Stroud", "Thackeray", "Weatherby", "Aldous", "Birkett",
    "Crandall", "Eldridge", "Fanshawe", "Grimwood", "Harkness", "Loxley",
    "Merton", "Pennington", "Rushton", "Sedgwick", "Tennant", "Waverley",
    "Wharton", "Yeardley",
]
_PN_ENTITY_WORDS = [
    "Aldrin", "Brightwater", "Cascadia", "Dunmore", "Everline", "Foxglen",
    "Granite", "Havenwood", "Ironbridge", "Juniper", "Kestrel", "Lumen",
    "Meridian", "Northgate", "Oakmont", "Pinnacle", "Quarry", "Redwood",
    "Silverpeak", "Torchlight", "Umbra", "Vantage", "Westmark", "Zephyr",
    "Ambrose", "Beacon", "Cobalt", "Drayton", "Emberly", "Falcon", "Gladstone",
    "Harborview", "Ivory", "Jetstream", "Kaldor", "Larkspur", "Monarch",
    "Nimbus", "Orion", "Pembroke",
    # Enlarged for the same reason as _PN_NAME_WORDS: a case with many entity
    # parties, affiliates and firm names outran 40 words and minted
    # "POSTBOX4.ORGPANY"-class numbered fakes.
    "Arclight", "Brookstone", "Cairnwood", "Clearspring", "Crestline",
    "Dovewood", "Eastmark", "Eldergrove", "Ferncliff", "Fieldstone",
    "Foxbridge", "Glenrock", "Goldcrest", "Graystone", "Highpoint",
    "Hollowmere", "Ironwood", "Kirkwall", "Lakemont", "Lanternwood",
    "Ledgewood", "Marbury", "Millbrook", "Moorland", "Oakspire", "Overland",
    "Parkhurst", "Pinehurst", "Ravenwood", "Riverton", "Rockhaven",
    "Sablewood", "Sandpiper", "Shorewood", "Silvergate", "Solstice",
    "Springvale", "Starling", "Stonehaven", "Thornfield", "Timberline",
    "Wexmoor", "Whitfield", "Wildmere", "Windermere", "Ashcroft",
    "Brightmoor", "Coppervale", "Dawnfield", "Emberton", "Frostgate",
    "Greenhollow", "Hartland", "Ironclad", "Keystone", "Lightwell",
    "Meadowgate", "Northwind", "Opalridge", "Pinecrest", "Quillmark",
    "Rosemont", "Stormont", "Truenorth", "Umberwood", "Vanguard",
    "Wellspring", "Yarrowvale",
]
_PN_ENTITY_SUFFIXES = {
    "llc", "l.l.c.", "inc", "inc.", "incorporated", "corp", "corp.",
    "corporation", "co", "co.", "company", "lp", "l.p.", "llp", "ltd", "ltd.",
    "limited", "pc", "p.c.", "plc", "na", "n.a.", "trust", "bank", "assn",
    "association", "partners", "partnership", "group", "fund", "holdings",
    "enterprises", "systems", "services", "solutions", "foundation",
    "apc", "a.p.c.", "pllc", "p.l.l.c.",
}
_PN_ENTITY_KEEP = {"the", "of", "and", "for", "a", "an", "&", "de", "la"} | _PN_ENTITY_SUFFIXES
# Suffix set with dots/spaces removed, so "P.C." / "L.L.C." normalize to the
# same key as "pc" / "llc" when deciding whether a token is a corporate suffix.
_PN_ENTITY_SUFFIXES_NORM = {re.sub(r"[.\s]", "", s) for s in _PN_ENTITY_SUFFIXES}

# ── "dba" (fictitious business name) markers ────────────────────────────────
# A caption cell routinely names one party twice: the legal entity and the
# fictitious name it trades under ("SMITH AUTO GROUP, INC. dba TOYOTA OF
# DOWNTOWN"). The brief then calls the party by ONE of those names, almost never
# by the joined string, so the joined string must be split and each side
# registered as its own term. The marker itself is public boilerplate: it is
# kept verbatim, never faked. (To recognize "aka"/"fka" the same way, add them
# to _PN_DBA_ALTS and _PN_DBA_NORM.)
_PN_DBA_ALTS = (r"doing\s+business\s+as", r"d[ \t]*/[ \t]*b[ \t]*/[ \t]*a\.?",
                r"d\.[ \t]*b\.[ \t]*a\.?", r"dba\.?")
# "aka"/"fka" mark another name for the SAME kind of thing: a person's alias is a
# person, whereas a dba is always a business. They are split like a dba but the
# tail inherits the head's person/entity classification.
_PN_AKA_ALTS = (r"a[ \t]*/[ \t]*k[ \t]*/[ \t]*a\.?", r"a\.[ \t]*k\.[ \t]*a\.?",
                r"aka\.?", r"f[ \t]*/[ \t]*k[ \t]*/[ \t]*a\.?",
                r"f\.[ \t]*k\.[ \t]*a\.?", r"fka\.?",
                r"also\s+known\s+as", r"formerly\s+known\s+as",
                r"erroneously\s+(?:sued|named)\s+(?:herein\s+)?as",
                r"sued\s+herein\s+as", r"improperly\s+(?:sued|named)\s+as")
_PN_DBA_NORM = {"dba", "doingbusinessas"}
# Letter guards keep the punctuated forms from straddling unrelated text: the
# sentence "...IS INITIALED. \nB. ADDITIONAL MEDIATION TERMS" otherwise reads as
# "d. b. a" spanning the end of one word and a subsection heading. The dotted
# and slashed forms may not cross a line break either (only "doing business as"
# may wrap).
_PN_DBA_MARK = r"(?<![A-Za-z])(?i:" + "|".join(_PN_DBA_ALTS) + r")(?![A-Za-z])"
_PN_AKA_MARK = r"(?<![A-Za-z])(?i:" + "|".join(_PN_AKA_ALTS) + r")(?![A-Za-z])"
_PN_DBA_SPLIT_RE = re.compile(
    r"(?:^|(?<=[\s,;(]))\(?" + _PN_DBA_MARK + r"\)?(?=$|[\s,;)])")
_PN_AKA_SPLIT_RE = re.compile(
    r"(?:^|(?<=[\s,;(]))\(?" + _PN_AKA_MARK + r"\)?(?=$|[\s,;)])")
# Two-letter tokens that are ordinary English words. A surname of the same
# letters is rare, but registering one as a case-insensitive whole-word term
# would rewrite the word throughout a brief. Ambiguous real surnames that are
# NOT everyday formal-writing words (Yu, Ng, Le, Wu, Vo, Xu, Ho, Oh) are left
# OUT so they still register and scrub.
_PN_SHORT_TOKEN_STOP = frozenset({
    "as", "at", "am", "an", "be", "by", "do", "go", "he", "if", "in", "is",
    "it", "me", "my", "no", "of", "on", "or", "so", "to", "up", "us", "we",
})
_PN_COMMON_WORD_SURNAMES = {
    "green", "brown", "white", "black", "gray", "grey", "young", "long",
    "short", "small", "rich", "best", "price", "king", "bell", "cook", "wood",
    "park", "day", "may", "will", "mark", "grace", "hope", "rose", "ray",
    "dawn", "hall", "moore", "west", "east", "north", "south", "case", "law",
    "field", "flowers", "banks", "waters", "rivers", "stone", "reed", "fields",
    "love", "just", "good", "power", "peace", "sharp", "swift", "noble",
}
# Fake street-name pool. Kept deliberately arboreal-but-uncommon; a name that
# turns up as a REAL street corrupts the record (an earlier build minted "Maple",
# which WAS the plaintiff's street). The registry rejects a fake that collides
# with another real, but cannot know a name is real prose, so the pool avoids
# the common ones.
_PN_STREET_NAMES = [
    "Cedar", "Birch", "Willow", "Aspen", "Juniper", "Laurel",
    "Poplar", "Hawthorn", "Linden", "Chestnut",
    "Sequoia", "Cypress", "Alder", "Dogwood", "Hickory", "Rosewood",
    "Foxglove", "Larkspur", "Tamarack", "Sorrel",
]
# Fake localities. The STATE is kept (venue is jurisdictionally meaningful and a
# two-letter code identifies nobody); the city name and the ZIP are faked, and a
# fake city must look like a real one to `_PN_ADDR_RE` or re-scrubbing an already
# scrubbed export would stop being a fixed point.
# Kept disjoint from _PN_NAME_WORDS, _PN_ENTITY_WORDS and _PN_STREET_NAMES (see
# TestPoolsAreDisjoint): a fake surname that is also a fake city reads as a
# collision to anyone auditing the key, even though the map stays a bijection.
_PN_CITY_NAMES = [
    "Fairview", "Brookfield", "Rosedale", "Elmwood", "Kingsbury",
    "Northvale", "Westbrook", "Clearwater", "Havenport", "Stonebridge",
    "Marlow", "Redhill", "Glenmore", "Oakhurst", "Bridgeton",
]
# NB: there is no fake street-TYPE pool. A road type (Street/Road/Court/Way…) is
# generic, not identifying, so the real one is always kept — see
# `_pn_addr_suffix_of`. Only the number and the street NAME are faked.
_PN_EMAIL_DOMAINS = ["example.com", "mailhaven.net", "postbox.org", "letterbox.co"]
# Generational and professional/degree suffixes that trail a name — never a name
# themselves, always kept verbatim ("Jane Smith, M.D." -> "<fake> <fake>, M.D.").
_PN_SUFFIX_TOKENS = {"jr", "jr.", "sr", "sr.", "ii", "iii", "iv", "v", "esq",
                     "esq.", "md", "m.d.", "phd", "ph.d.", "do", "d.o.",
                     "dds", "d.d.s.", "rn", "r.n.", "cpa", "c.p.a.", "jd",
                     "j.d.", "llm", "ll.m.", "np", "pa", "dvm", "psyd", "psy.d."}
# Matched dot/space-insensitively so "M.D.", "MD", "M.D" and "m.d." all count.
_PN_SUFFIX_NODOT = frozenset(re.sub(r"[.\s]", "", s) for s in _PN_SUFFIX_TOKENS)


def _pn_is_suffix_token(word):
    """True for a generational/degree suffix in any punctuation ("M.D.", "MD",
    "Ph.D.", "Esq", "Jr.") — never pseudonymized, always kept as written."""
    return re.sub(r"[.\s]", "", word).lower().removesuffix("'s") in _PN_SUFFIX_NODOT


# Latin letters INCLUDING the accented forms that appear in California party
# and declarant names — Spanish ("Alarcón", "Muñoz", "Peña", "Hernández"),
# Portuguese, French. An ASCII-only [A-Za-z] class split "Alarcón" into "Alarc"
# + "n" at the "ó": the truncated stem was stored in the key (so the real name
# no longer round-trips) and, on a spliced page, matched the prefix of the real
# word and left the accented tail welded onto the fake ("Isleyón"). The range
# covers Latin-1 Supplement letters (U+00C0–U+00FF, skipping the × and ÷ math
# signs) plus Latin Extended-A (U+0100–U+017F); digits and underscores stay out.
_PN_LAT = "A-Za-zÀ-ÖØ-öø-ÿĀ-ſ"
_PN_LAT_UPPER = "A-ZÀ-ÖØ-Þ"
_PN_WORD_RE = re.compile(rf"[{_PN_LAT}][{_PN_LAT}'\-]*")


def _pn_build_accent_fold():
    """Char->char map folding an accented Latin letter to its ASCII base
    (ó->o, ñ->n, é->e). ONE char in, ONE char out, so it can be applied inside
    a reduced-index map without shifting offsets — only letters whose NFKD base
    is a single ASCII letter are mapped; ligatures (æ->ae) and ß are left alone
    (they drop out of the [a-z0-9] reduction anyway)."""
    fold = {}
    for cp in range(0xC0, 0x180):
        ch = chr(cp)
        base = "".join(c for c in _pn_unicodedata.normalize("NFKD", ch)
                       if not _pn_unicodedata.combining(c))
        if len(base) == 1 and ("A" <= base <= "Z" or "a" <= base <= "z"):
            fold[cp] = ord(base)
    return fold


_PN_ACCENT_FOLD = _pn_build_accent_fold()


def _pn_ascii_fold(s):
    """`s` with accented Latin letters folded to their ASCII base. Used by the
    reduced-alphabet leak scan / welded scrub so "Alarcón" and an un-accented
    "Alarcon" (a common OCR/typing variant, and the only form an e-mail carries)
    share one reduced core instead of missing each other."""
    return s.translate(_PN_ACCENT_FOLD)


def _pn_titlecase_like(fake, original):
    if original.isupper():
        return fake.upper()
    if original.islower():
        return fake.lower()
    return fake


# The shortest name token that folds an OCR/typo near-variant (see
# `_PnFakeRegistry.token`). Shorter tokens (Ng, Yu, Vang, Choi) are left to
# mint their own fakes, since at 4 chars two genuinely different surnames sit
# an edit apart too often to fold safely.
_PN_NAME_FOLD_MIN = 5

# Letter-only visual confusables for deriving a substitution typo of a fake —
# the result must stay a name-shaped token, so a glyph never maps to a digit.
_PN_CONFUSABLES = {
    "i": "l", "l": "i", "o": "a", "a": "o", "m": "n", "n": "m", "u": "v",
    "v": "u", "e": "c", "c": "e", "s": "z", "z": "s", "b": "h", "h": "b",
    "g": "q", "q": "g", "t": "f", "f": "t", "r": "n", "d": "b", "w": "v",
    "y": "v",
}


def _pn_typo_variants(fake, op, seed, reps=1):
    """Yield typo'd forms of `fake` — the fake a near-variant real folds onto, so
    a real OCR typo reads as a typo of the fake rather than a brand-new name.
    `op` mirrors the real's own deviation so lengths track: 'ins' duplicates a
    letter (real gained one), 'del' drops one, 'sub' swaps one for a visual
    confusable, 'trans' swaps two adjacent letters (real transposed a pair, e.g.
    "Adler"/"Alder" -> "Darrow"/"Drarow"). `reps` applies a length-changing op
    ('ins'/'del') that many times, so a long token with several typos gets a
    fake whose length delta matches (real +2 chars -> fake +2). Interior
    positions are tried in a per-variant deterministic order; the caller takes
    the first form no other real has been given."""
    if op in ("ins", "del") and reps > 1:
        # Compose `reps` single-letter edits so the net length change tracks the
        # real's. Each level branches over positions; dedup and keep only forms
        # that actually reached the target length delta.
        seen, frontier = set(), [fake]
        for _ in range(reps):
            nxt = []
            for f in frontier:
                nxt.extend(_pn_typo_variants(f, op, seed))
            frontier = nxt
        want = len(fake) + (reps if op == "ins" else -reps)
        for v in frontier:
            if len(v) == want and v != fake and v not in seen:
                seen.add(v)
                yield v
        return
    rng = _pn_rng("nametypo", op, seed)
    idxs = list(range(1, len(fake)))          # never touch the leading capital
    rng.shuffle(idxs)
    for i in idxs:
        ch = fake[i]
        if op == "ins":
            yield fake[:i] + ch + fake[i:]                 # duplicate a letter
        elif op == "del":
            if len(fake) > 3:
                yield fake[:i] + fake[i + 1:]              # drop a letter
        elif op == "trans":
            # Swap this letter with the next, mirroring the real's transposition.
            # Skip a doubled pair (swapping "rr" changes nothing).
            if i + 1 < len(fake) and fake[i] != fake[i + 1]:
                yield fake[:i] + fake[i + 1] + fake[i] + fake[i + 2:]
        else:  # sub
            repl = _PN_CONFUSABLES.get(ch.lower())
            if repl is None:
                continue
            repl = repl.upper() if ch.isupper() else repl
            if repl != ch:
                yield fake[:i] + repl + fake[i + 1:]


def _pn_case_word_like(fake, original):
    """`fake` re-cased to match how `original` was written.

    Unlike `_pn_titlecase_like`, this also handles a MIXED-case original against
    a fake that was minted in a single case. A fake is stored in whatever casing
    the spreadsheet cell used, so an ALL-CAPS "Title Defendant" yields an
    all-caps fake, which then reads as a shout inside ordinary prose. When the
    original is neither all-upper nor all-lower, an all-one-case fake is folded
    to Title Case per letter-run — which leaves "Inc." and "L.L.C." intact."""
    if original.isupper() and any(c.isalpha() for c in original):
        return fake.upper()
    if original.islower():
        return fake.lower()
    if fake.isupper() or fake.islower():
        # Fold each letter-run to Title Case. A possessive is part of its word
        # ("BENNETT'S" -> "Bennett's", not "Bennett'S"), while the dot-separated
        # runs of "L.L.C." stay separate letters.
        return re.sub(r"[A-Za-z]+(?:['’][A-Za-z]+)*",
                      lambda m: m.group(0).capitalize(), fake)
    return fake


def _pn_case_like(original, fake):
    """`fake` re-cased word by word against `original`. Name fakes are composed
    token-for-token with the real name, so the counts line up; when they don't
    (a phone number, an address), the whole string is cased as one."""
    otoks, ftoks = original.split(), fake.split()
    if len(otoks) == len(ftoks):
        return " ".join(_pn_case_word_like(f, o) for o, f in zip(otoks, ftoks))
    return _pn_case_word_like(fake, original)


class _PnFakeRegistry:
    """Hands out fakes so the real->fake map stays a bijection for one case.

    Two guarantees, both required for the pseudonym key to round-trip:
      * stable  — the same real value always gets the same fake (across pages,
        files, and full-name vs. bare-token forms);
      * unique  — two DIFFERENT real values never get the same fake.

    The fake-name pool is small, so without this the per-token hash would
    occasionally map two unrelated people to one fake (e.g. a party's surname
    and an attorney's middle name both becoming "Forsythe"), which made the
    fake ambiguous and broke mapping the .txt back to the real names."""

    # A word must map to ONE fake regardless of whether the person path or the
    # entity path asks for it, or the same party splits into two identities
    # (the delivered set called the defendant both "Beacon Torchlight, Inc."
    # and "Nolan Isley, Inc."). These two token seed-tags therefore share a
    # single memo slot; whichever path binds the word first wins, and the other
    # reuses it — so a bare "Azul Concreto" met on the person path renders with
    # the same words as the entity "Azul Concreto, Inc."
    _NAME_ENTITY_TAGS = frozenset({"nametok", "enttok"})

    @classmethod
    def _memo_tag(cls, seed_tag):
        return "name_or_entity" if seed_tag in cls._NAME_ENTITY_TAGS else seed_tag

    def __init__(self):
        self._memo = {}     # (memo_tag, real_lower) -> fake
        self._used = set()  # every fake handed out, lower-cased (global)
        self._domain_reals = {}  # real host -> fake, for OCR-typo folding

    def _take(self, key, fake):
        self._memo[key] = fake
        self._used.add(fake.lower())
        return fake

    def tokens_for(self, seed_tag):
        """{real_lower: fake} for every value assigned under `seed_tag` — used to
        reuse a person's name-token fake elsewhere (e.g. inside an e-mail).
        Person and entity tokens share a slot, so either resolves the other."""
        want = self._memo_tag(seed_tag)
        return {real: fake for (tag, real), fake in self._memo.items()
                if tag == want}

    def token(self, real, words, seed_tag):
        """A single canonical-case stand-in word for `real`, never reused for a
        different real value. Draws from `words` in a per-value deterministic
        order, skipping any already taken; if the whole pool is exhausted a
        numeric suffix keeps it unique. Person and entity token pools share one
        memo slot per word (see `_NAME_ENTITY_TAGS`)."""
        tag = self._memo_tag(seed_tag)
        key = (tag, real.lower())
        if key in self._memo:
            return self._memo[key]
        low = real.lower()
        # Fold an OCR/typo near-variant (edit distance 1) onto a TYPO of the
        # base token's fake, so "Palladina"/"Pallading" read as typos of the
        # same "Keswick" the canonical "Palladino" got, instead of three
        # unrelated stand-ins. Each still keeps its OWN fake (a distinct typo),
        # so the key round-trips one-to-one.
        if len(low) >= _PN_NAME_FOLD_MIN:
            for (t, prev), prev_fake in self._memo.items():
                if (t != tag or prev == low or len(prev) < _PN_NAME_FOLD_MIN
                        or not _pn_edit_distance_within(
                            prev, low, _pn_name_fold_dist(prev, low),
                            min_len=_PN_NAME_FOLD_MIN)):
                    continue
                dl = len(low) - len(prev)
                if dl > 0:
                    op, reps = "ins", dl        # real gained dl chars -> so does fake
                elif dl < 0:
                    op, reps = "del", -dl        # real lost chars -> fake drops them
                else:
                    # Equal length: an adjacent transposition (two swapped,
                    # side-by-side mismatches) mirrors as a swap; anything else
                    # as a confusable substitution.
                    mism = [i for i in range(len(low)) if low[i] != prev[i]]
                    op = ("trans" if len(mism) == 2 and mism[1] == mism[0] + 1
                          and low[mism[0]] == prev[mism[1]]
                          and low[mism[1]] == prev[mism[0]] else "sub")
                    reps = 1
                for cand in _pn_typo_variants(prev_fake, op, low, reps):
                    if cand.lower() not in self._used:
                        return self._take(key, cand)
                break     # a near-variant was found but yielded no free typo
        # A HYPHENATED compound surname is two name components, not one word,
        # and its fake must show that: "Ardeshirpour-Zartoshti" -> "Sedgwick-
        # Linford", each half the SAME fake that half gets standing alone. A
        # single pool word ("Gaskell") made the compound and its shorthand
        # ("Dr. Ardeshirpour" -> "Dr. Sedgwick") read as two unrelated people,
        # and it hid the compound shape from anyone reading the export. It also
        # makes the reversal ROBUST: because the compound fake is exactly its
        # parts' fakes joined, the macro's token-by-token pass rebuilds it
        # correctly whichever row it applies first.
        parts = low.split("-")
        if (len(parts) > 1 and all(len(p) >= 2 and p.isalpha() for p in parts)
                and any(len(p) >= _PN_NAME_FOLD_MIN for p in parts)):
            cand = "-".join(self.token(p, words, seed_tag) for p in parts)
            if cand.lower() not in self._used:
                return self._take(key, cand)
        # Fold a WELDED token — a column-splice glued two names with no space
        # ("ADLERMICHAEL" = "ADLER" + "MICHAEL") — onto the CONCATENATION of the
        # two fakes, so the stand-in still shows the bound party ("Darrow" +
        # whatever "Michael" faked to) instead of an unrelated word. Guarded so
        # an ordinary long surname is not split: the whole token must be letters
        # and long enough to hold two full name tokens, and BOTH the bound
        # prefix/suffix and the leftover must clear the fold floor. The leftover
        # is faked through the registry too, so it stays unique and reversible
        # and reuses that name's fake if it appears on its own elsewhere.
        if len(low) >= 2 * _PN_NAME_FOLD_MIN and low.isalpha():
            for (t, prev), prev_fake in list(self._memo.items()):
                if t != tag or len(prev) < _PN_NAME_FOLD_MIN or prev == low:
                    continue
                if low.startswith(prev):
                    rest, is_prefix = low[len(prev):], True
                elif low.endswith(prev):
                    rest, is_prefix = low[:-len(prev)], False
                else:
                    continue
                if len(rest) < _PN_NAME_FOLD_MIN or not rest.isalpha():
                    continue
                rest_fake = self.token(rest, words, seed_tag)
                cand = (prev_fake + rest_fake if is_prefix
                        else rest_fake + prev_fake)
                if cand.lower() not in self._used:
                    return self._take(key, cand)
                break     # a weld was found but the joined fake was taken
        rng = _pn_rng(seed_tag, real.lower())
        order = list(words)
        rng.shuffle(order)
        for cand in order:
            if cand.lower() not in self._used:
                return self._take(key, cand)
        base, n = order[0], 2
        while f"{base}{n}".lower() in self._used:
            n += 1
        return self._take(key, f"{base}{n}")

    def digits(self, real, seed_tag, keep_prefix=0):
        """A digit-for-digit fake of `real` (e.g. a case number), unique per
        case. Re-seeds on the rare collision so two real numbers never share
        one fake. The first `keep_prefix` characters are copied verbatim — a
        case number's two-digit filing year is not an identifier, and randomising
        it produced numbers that are facially impossible for the filing date
        printed beside them ("25STCV37838" -> "48STCV51378")."""
        key = (seed_tag, real.lower())
        if key in self._memo:
            return self._memo[key]
        head, tail = real[:keep_prefix], real[keep_prefix:]
        for attempt in range(10000):
            r = _pn_rng(seed_tag, real, attempt)
            cand = head + re.sub(r"\d", lambda m: str(r.randrange(10)), tail)
            if cand.lower() not in self._used:
                return self._take(key, cand)
        return self._take(key, cand)  # give up after 10k tries; keep it stable

    def alnum(self, real, seed_tag, letters="ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
        """A char-for-char fake of an alphanumeric identifier (a VIN, a
        confirmation code): each digit -> a random digit, each ASCII letter -> a
        random letter drawn from `letters` (the VIN alphabet excludes I/O/Q),
        every other character kept. Unique per case."""
        key = (seed_tag, real.lower())
        if key in self._memo:
            return self._memo[key]
        low = letters.lower()

        def sub(c, r):
            if c.isdigit():
                return str(r.randrange(10))
            if c.isascii() and c.isalpha():
                pool = letters if c.isupper() else low
                return pool[r.randrange(len(pool))]
            return c
        cand = real
        for attempt in range(10000):
            r = _pn_rng(seed_tag, real, attempt)
            cand = "".join(sub(c, r) for c in real)
            if cand.lower() not in self._used:
                return self._take(key, cand)
        return self._take(key, cand)

    def domain(self, host, pool):
        """An injective fake host. Folds a near-duplicate real host (an OCR typo,
        edit distance 1) onto the same fake, and mints uniqueness by mutating the
        registrable LABEL ("letterbox2.co"), never the TLD ("letterbox.co2",
        which isn't a valid domain)."""
        host = host.lower()
        key = ("domain", host)
        if key in self._memo:
            return self._memo[key]
        for prev, fake in self._domain_reals.items():
            if _pn_edit_distance_le1(prev, host):
                self._memo[key] = fake
                return fake
        rng = _pn_rng("emaildom", host)
        order = list(pool)
        rng.shuffle(order)
        fake = next((c for c in order if c.lower() not in self._used), None)
        if fake is None:
            label, _dot, tld = order[0].partition(".")
            n = 2
            while f"{label}{n}.{tld}".lower() in self._used:
                n += 1
            fake = f"{label}{n}.{tld}"
        self._used.add(fake.lower())
        self._memo[key] = fake
        self._domain_reals[host] = fake
        return fake

    def unique(self, real, seed_tag, make):
        """A fake produced by `make(rng)`, retried with a fresh seed until it is
        one no other real value has been given. `real` should be the CANONICAL
        form of the value (e.g. an address folded to one spelling) so its
        spelling variants all resolve to the same fake. Guarantees injectivity —
        two different reals never collide onto one fake — which the address key
        needs so one parcel is not written under two fakes."""
        key = (seed_tag, real.lower())
        if key in self._memo:
            return self._memo[key]
        cand = None
        for attempt in range(10000):
            cand = make(_pn_rng(seed_tag, real.lower(), attempt))
            if cand.lower() not in self._used:
                return self._take(key, cand)
        return self._take(key, cand)  # give up after 10k tries; keep it stable


def _pn_fake_name_token(word, registry):
    return _pn_titlecase_like(registry.token(word, _PN_NAME_WORDS, "nametok"), word)


# Words that mark a business even with no corporate suffix attached. Without
# these, "South Bay Equity" and "Mortgage 2000" take the PERSON path, and the
# bare token "Equity" is registered and rewritten wherever the word appears.
_PN_ENTITY_HINT_WORDS = {
    "equity", "mortgage", "lending", "loans", "realty", "capital", "financial",
    "finance", "properties", "property", "brokerage", "escrow", "title",
    "insurance", "ventures", "management", "consultants", "consulting",
    "motors", "automotive", "industries", "technologies", "laboratories",
    "medical", "hospital", "university", "church", "credit", "investments",
    "development", "construction", "builders", "realtors", "escrows",
}


# Corporate suffixes that are decisive on their own ("LLC" is never a surname)
# vs. ones that need punctuation context ("Co."/"Na" could be name syllables:
# "Li Na" is a person; "Smith & Co." is not).
_PN_CORP_SUFFIX_UNAMBIG = {
    "llc", "llp", "inc", "corp", "ltd", "plc", "incorporated", "corporation",
    "company", "lp", "pc",
}


def _pn_trailing_corp_suffix(name):
    """The name's final token when it is a corporate suffix — with or without a
    trailing period, with or without a preceding comma ("AhC Acquisition, LLC.")
    — else None. Classify a name by the entity it belongs to, not by its own
    shape: a trailing corporate suffix is a decisive entity signal, and the
    token itself must never be faked as a person surname. An ambiguous suffix
    word ("Co", "Na") qualifies only with punctuation context (its own period or
    a preceding comma), so a person name like "Li Na" is untouched."""
    toks = name.split()
    if len(toks) < 2:
        return None
    last = toks[-1]
    norm = re.sub(r"[.\s]", "", _pn_word_base(last))
    if norm not in _PN_ENTITY_SUFFIXES_NORM:
        return None
    if (norm in _PN_CORP_SUFFIX_UNAMBIG or "." in last
            or toks[-2].endswith(",")):
        return last
    return None


# Bare (unpunctuated) suffix words that are also plausible name syllables. As
# naked tokens they must not classify a name as an entity ("Li Na", "Tyler Co"
# the surname): they count only with punctuation context — "Co.", "N.A.", a
# preceding comma — via _pn_trailing_corp_suffix or their dotted set entries.
_PN_CORP_SUFFIX_AMBIG = {"na", "co", "lp", "pc"}


def _pn_looks_like_entity(name):
    if _pn_trailing_corp_suffix(name):
        return True
    low = " " + re.sub(r"[^\w&. ]", " ", name).lower() + " "
    if any(f" {s} " in low
           for s in _PN_ENTITY_SUFFIXES if s not in _PN_CORP_SUFFIX_AMBIG):
        return True
    if any(f" {w} " in low for w in _PN_ENTITY_HINT_WORDS):
        return True
    return bool(re.search(r"\b(city|county|people|state) of\b", low))


# The state-of-organization descriptor a caption appends to an entity party:
# "Acme Widgets, Inc., a Delaware corporation", "Sunrise Motors, LLC, a
# California limited liability company", "Doe Holdings, an Illinois limited
# partnership". It names the STATE of incorporation and the LEGAL FORM — stock
# boilerplate that identifies no one and reads identically for thousands of
# companies, so it must stay verbatim and NEVER be faked, even though it sits
# inside the party's name. Only the distinctive name in front of it is a term.
_PN_JURIS_STATES = (
    "alabama", "alaska", "arizona", "arkansas", "california", "colorado",
    "connecticut", "delaware", "florida", "georgia", "hawaii", "idaho",
    "illinois", "indiana", "iowa", "kansas", "kentucky", "louisiana", "maine",
    "maryland", "massachusetts", "michigan", "minnesota", "mississippi",
    "missouri", "montana", "nebraska", "nevada", "ohio", "oklahoma", "oregon",
    "pennsylvania", "tennessee", "texas", "utah", "vermont", "virginia",
    "washington", "wisconsin", "wyoming",
    "new hampshire", "new jersey", "new mexico", "new york", "north carolina",
    "north dakota", "rhode island", "south carolina", "south dakota",
    "west virginia", "district of columbia",
)
# Adjectives that qualify the legal form ("a California PROFESSIONAL corporation",
# "a NONPROFIT PUBLIC BENEFIT corporation", "a FOREIGN limited partnership").
_PN_JURIS_QUALIFIERS = (
    "professional", "nonprofit", "non-profit", "public", "benefit", "mutual",
    "general", "limited", "domestic", "foreign", "close", "closely", "held",
    "unincorporated", "incorporated", "registered", "statutory", "charitable",
    "religious", "family", "business",
    # "…, an unincorporated INTERINDEMNITY arrangement" — the physician
    # mutual-protection form. Unrecognised, the whole clause was harvested as
    # part of the entity name and each of its ordinary words became a bare
    # pseudonym token the operator had to retire by hand.
    "interindemnity", "reciprocal", "inter-insurance",
)
_PN_ENTITY_FORM_NOUN = (
    r"(?:corporations?|corp|compan(?:y|ies)|"
    r"limited\s+liability\s+(?:compan(?:y|ies)|partnerships?)|"
    r"limited\s+partnerships?|general\s+partnerships?|partnerships?|"
    r"joint\s+venture|unincorporated\s+association|associations?|"
    r"business\s+trust|trust|cooperative|"
    r"arrangements?|organizations?|exchange|"
    r"l\.?\s*l\.?\s*c|l\.?\s*l\.?\s*p|l\.?\s*p|p\.?\s*l\.?\s*l\.?\s*c|"
    r"p\.?\s*c|p\.?\s*a)")
_PN_DESCRIPTOR_BODY = (
    r"(?:(?:" + "|".join(_PN_JURIS_STATES) + r"|"
    + "|".join(_PN_JURIS_QUALIFIERS) + r")\s+){0,4}?"  # state / qualifier words
    + _PN_ENTITY_FORM_NOUN + r"\.?")
_PN_ENTITY_DESCRIPTOR_RE = re.compile(
    r",?\s+an?\s+" + _PN_DESCRIPTOR_BODY + r"\s*$", re.IGNORECASE)
# A value that is NOTHING but a descriptor ("a California corporation") names no
# party at all — it must not become a term (and so must never be faked).
_PN_ONLY_DESCRIPTOR_RE = re.compile(
    r"^\s*an?\s+" + _PN_DESCRIPTOR_BODY + r"\s*$", re.IGNORECASE)


def _pn_strip_entity_descriptor(name):
    """(name_without_descriptor, had_descriptor). Removes a trailing
    "..., a <State> <legal-form>" clause so the registered term is just the
    distinctive name — the descriptor then stays verbatim in the document and is
    never faked. Only strips when a distinctive name remains in front of it, so a
    value that is nothing but the descriptor is left as-is for the caller to
    reject elsewhere."""
    m = _PN_ENTITY_DESCRIPTOR_RE.search(name)
    if not m or m.start() == 0:
        return name, False
    head = name[:m.start()].rstrip(" ,")
    if not re.search(r"[A-Za-z]", head):
        return name, False
    return head, True


# A state (or DC) name is geography, not identity: it names no one and reads the
# same for every business incorporated or located there, so it is kept VERBATIM
# inside a company/entity party name and only the other words are faked
# ("California Pizza Kitchen" -> keep "California", fake the rest). Scoped to the
# ENTITY path on purpose: a state word that is part of a PERSON's name — the
# surname "Washington", the first name "Georgia" — is still an identifier and is
# faked, so it never leaks. Multi-word states ("New York") are kept as a unit.
_PN_STATE_SINGLE = frozenset(s for s in _PN_JURIS_STATES if " " not in s)
_PN_STATE_MULTI = tuple(tuple(s.split()) for s in _PN_JURIS_STATES if " " in s)


def _pn_state_keep_flags(bases):
    """Per-token bool over lower-cased word bases: True where the token is part
    of a US state / DC name (single- or multi-word), so the entity faker keeps
    it verbatim. Multi-word matches guard the ambiguous leading word ("new" is
    kept only inside "new york"/"new jersey"/..., never on its own)."""
    n = len(bases)
    keep = [b in _PN_STATE_SINGLE for b in bases]
    for phrase in _PN_STATE_MULTI:
        L = len(phrase)
        for i in range(n - L + 1):
            if tuple(bases[i:i + L]) == phrase:
                for j in range(i, i + L):
                    keep[j] = True
    return keep


# ── Party-role labels (never pseudonymized) ─────────────────────────────────
# "Plaintiff", "Defendant", "Cross-Complainant", etc. are procedural roles, not
# identities: they occur on nearly every line of a pleading and are public, so
# replacing them would corrupt the text while protecting nothing. A spreadsheet
# name cell (especially the E-Court "Other Names" column) or a --term is
# sometimes JUST a role word; such a value must never become a pseudonym term,
# and no bare name-token equal to a role word may be registered either.
_PN_PARTY_ROLE_WORDS = frozenset({
    "plaintiff", "plaintiffs", "defendant", "defendants",
    "petitioner", "petitioners", "respondent", "respondents",
    "complainant", "complainants",
    "cross-complainant", "cross-complainants",
    "crosscomplainant", "crosscomplainants",
    "cross-defendant", "cross-defendants",
    "crossdefendant", "crossdefendants",
    "cross-claimant", "cross-claimants",
    "counter-claimant", "counter-claimants",
    "counterclaimant", "counterclaimants",
    "counter-defendant", "counter-defendants",
    "counterdefendant", "counterdefendants",
    "claimant", "claimants", "movant", "movants",
    "appellant", "appellants", "appellee", "appellees",
    "intervenor", "intervenors", "intervener", "interveners",
    "creditor", "creditors", "debtor", "debtors",
    "garnishee", "garnishees", "party", "parties",
    "third party", "third-party", "real party in interest",
    "aggrieved employee", "aggrieved employees",
})
# Separators that join role words into a compound role phrase
# ("Plaintiff and Cross-Complainant", "Plaintiffs/Cross-Defendants").
_PN_ROLE_SPLIT_RE = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.IGNORECASE)


def _pn_is_party_role(value):
    """True when `value` is nothing but one or more common party-role labels
    (e.g. "Plaintiff", "Defendants", "Plaintiff and Cross-Complainant"). Such a
    value names a role rather than a person or entity, so it is dropped before
    it can ever become a pseudonym term."""
    if not value:
        return False
    norm = re.sub(r"\s+", " ", str(value)).strip().strip('.,:;"’\'').lower()
    if not norm:
        return False
    if norm in _PN_PARTY_ROLE_WORDS:
        return True
    parts = [p for p in _PN_ROLE_SPLIT_RE.split(norm) if p]
    return len(parts) > 1 and all(p in _PN_PARTY_ROLE_WORDS for p in parts)


def _pn_is_role_token(token):
    """True when a single bare name-token is itself a party-role word."""
    return token.strip('.,:;"’\'').lower() in _PN_PARTY_ROLE_WORDS


# ── Capacity phrases and other words that are never a person's name ──────────
# An E-Court name cell routinely carries the capacity a party sues in:
# "Kristopher Edwards as successor in interest to Erika Edwards". Faked whole,
# that cell registers "successor" and "interest" as bare name-tokens, and every
# later "successor in interest" in the briefs is silently rewritten. The cell is
# split at the capacity phrase instead, so both real names are registered and
# the connective tissue is left alone. The word list is a second, independent
# guard: no bare token drawn from any of these words ever becomes a term.
_PN_NON_NAME_WORDS = frozenset({
    "successor", "successors", "interest", "trustee", "trustees", "executor",
    "executrix", "administrator", "administratrix", "guardian", "conservator",
    "representative", "individually", "individual", "capacity", "behalf",
    "deceased", "decedent", "minor", "estate", "heir", "heirs", "beneficiary",
    "regarding", "concerning", "compliance", "support", "opposition",
    "response", "reply", "declaration", "motion", "notice", "exhibit",
    "county", "state", "city", "court", "doe", "does", "roe", "roes",
    # document/caption abbreviations, never a person's name:
    "iso", "rjn",   # "In Support Of", "Request for Judicial Notice"
})
# Abbreviations kept VERBATIM even when they land inside a captured name, so
# "Declaration of X ISO Motion" never fakes the "ISO".
_PN_DOC_ABBREV = frozenset({"iso", "rjn"})
_PN_CAPACITY_RE = re.compile(
    r"[\s,;]*\b(?:"
    r"(?:as\s+)?(?:an?\s+)?(?:the\s+)?successors?[-\s]in[-\s]interest"
    r"|as\s+(?:an?\s+|the\s+)?(?:successor|trustee|executor|executrix|"
    r"administrator|administratrix|guardian(?:\s+ad\s+litem)?|conservator|"
    r"personal\s+representative|individual)"
    r"|individually(?:\s+and\s+as)?"
    r"|an?\s+individual"
    r"|on\s+behalf\s+of"
    r"|in\s+(?:his|her|its|their)\s+capacity"
    r"|as\s+trustee\s+of"
    r")\b(?:\s+(?:of|to|for))?[\s,;]*", re.IGNORECASE)


def _pn_split_capacities(raw):
    """['Kristopher Edwards', 'Erika Edwards'] for a name cell that states the
    capacity a party sues in; a single-element list otherwise. Each piece is a
    name in its own right and is registered separately."""
    parts = [re.sub(r"\s+", " ", p).strip().strip(",;()\"' ")
             for p in _PN_CAPACITY_RE.split(raw)]
    return [p for p in parts if p and re.search(r"[A-Za-z]", p)] or [raw]


def _pn_is_name_token(word):
    """Whether a bare token of a person's name may become a term on its own.
    It must be capitalised where it was written (so the lower-case "interest"
    of a capacity phrase can never qualify), must not be a role label, and must
    not be one of the procedural words above."""
    if not word or not word[0].isupper():
        return False
    if _pn_is_suffix_token(word):  # "M.D.", "Ph.D.", "Esq" — a suffix, not a name
        return False
    base = word.strip('.,:;"’\'').lower().removesuffix("'s")
    return (base not in _PN_NON_NAME_WORDS and base not in _PN_PARTY_ROLE_WORDS
            and base not in _PN_COMMON_WORD_SURNAMES)


_PN_VOWEL_CONS = "bcdfghjklmnpqrstvwxz"


def _pn_name_variants(word):
    """Near-spelling variants of a surname/given name, so the key's "Roxane"
    still scrubs the exhibit's "Roxanne". Covers a doubled/collapsed internal
    consonant, a y<->ie ending, and an e<->ee ending. Each variant is registered
    at LOW priority pointing at the SAME fake, so the exact spelling always wins
    and only a genuine near-miss falls through to a variant. Kept conservative
    on purpose — an over-eager variant rewrites an unrelated word silently."""
    core = _pn_word_affixes(word)[1]
    if len(core) < 4 or not core[0].isupper():
        return set()
    low = core.lower()
    out = set()
    # collapse a doubled consonant ("Purscelley" has none; "Bennett"->"Benet")
    for i in range(1, len(core)):
        if core[i].lower() == core[i - 1].lower() and core[i].lower() in _PN_VOWEL_CONS:
            out.add(core[:i] + core[i + 1:])
    # double a single internal consonant ("Roxane"->"Roxanne")
    for i in range(1, len(core) - 1):
        c = core[i].lower()
        if (c in _PN_VOWEL_CONS and core[i].lower() != core[i - 1].lower()
                and core[i].lower() != core[i + 1].lower()):
            out.add(core[:i + 1] + core[i] + core[i + 1:])
    # y <-> ie
    if low.endswith("ie"):
        out.add(core[:-2] + "y")
    elif low.endswith("y"):
        out.add(core[:-1] + "ie")
    # e <-> ee
    if low.endswith("ee"):
        out.add(core[:-1])
    elif low.endswith("e") and not low.endswith("ee"):
        out.add(core + "e")
    # Only variants that are themselves plausible name tokens, and never the
    # word itself.
    return {v for v in out
            if v.lower() != low and len(v) >= 4 and _pn_is_name_token(v)}


_PN_INITIAL_POOL = tuple("ABCDEFGHIJKLMNOPQRSTUVWXYZ")


def _pn_norm_map(s):
    """Comparison form for a real/fake pair: NFKC-folded, whitespace-collapsed,
    trailing punctuation stripped, case-folded — so a value and a would-be fake
    that differ only in spacing, a trailing dot or case are recognised as the
    SAME string (a self-map that never scrubs)."""
    return re.sub(r"\s+", " ", _NFKC(str(s))).strip().strip(" .,;:").casefold()


def _pn_fake_initials_name(real, registry):
    """A distinct, reversible fake for a name made ENTIRELY of single-letter
    tokens plus punctuation — "M & M", "J&J", "A. B." — which the per-token
    person faker leaves byte-for-byte identical (every initial is kept verbatim),
    a self-map that both ships the real value and makes --fix-leaks loop forever.

    Map each real letter to a fake letter (repeats and separators preserved, so
    "M & M" -> "Q & Q" and "A & B" -> "Q & R"), each drawn through the registry
    so it is deterministic, globally unique and reversible. The pool excludes the
    real letter, so the result can never equal the input."""
    out = []
    for ch in real:
        if ch.isalpha():
            pool = [c for c in _PN_INITIAL_POOL if c != ch.upper()]
            fake = registry.token(ch.lower(), pool, "initial")
            out.append(fake.upper() if ch.isupper() else fake.lower())
        else:
            out.append(ch)
    return "".join(out)


def _pn_guard_distinct_fake(real, fake, registry, log=None):
    """Guarantee a term's fake differs from its real value under normalization.
    A self-identical mapping is a no-op scrub: it ships the real name in a
    "clean" export AND makes the post-scrub survivor scan re-flag the value on
    every --fix-leaks pass (a non-terminating loop). Returns the given fake when
    already distinct, else a re-mint; None only if no distinct fake can be made
    (the caller drops the term — a drop flags the value once, an installed no-op
    never resolves)."""
    if _pn_norm_map(fake) != _pn_norm_map(real):
        return fake
    remint = _pn_fake_initials_name(real, registry)
    if _pn_norm_map(remint) != _pn_norm_map(real):
        if log:
            log.info(f"  Pseudonymize: re-minted self-mapped value {real!r} -> "
                     f"{remint!r} (a fake equal to the real value never scrubs)")
        return remint
    if log:
        log.warning(f"  Pseudonymize: dropping {real!r} — no distinct fake could "
                    f"be minted; add an explicit replacement in the key.")
    return None


def _pn_fake_person(name, registry):
    """(fake_full_name, [(bare_real, bare_fake, is_surname), ...]) — composed
    token-by-token so a bare surname resolves to the same fake used here. A
    single-letter initial is kept verbatim: faking "J." to a whole surname makes
    "J. Brett Griffin" render as "TOLLIVER. Forsythe Ivers" in one place and
    "J. Forsythe Ivers" in another, since the bare tokens leave the initial
    alone."""
    words = list(_PN_WORD_RE.finditer(name))
    # A trailing corporate suffix ("…, LLC.") that slipped down the person path
    # must never be faked as a surname — it is structure, not identity. Kept
    # verbatim and excluded from surname selection (single-pool classification:
    # the entity path is where such a name belongs; this is the safety net).
    trail = _pn_trailing_corp_suffix(name)
    trail_at = (words[-1].start()
                if (trail and words
                    and _pn_word_base(words[-1].group(0)) == _pn_word_base(trail))
                else None)
    def _keep(w):
        return (len(w) == 1 or _pn_is_suffix_token(w)
                or w.strip(".,").lower() in _PN_DOC_ABBREV)
    mappable = [m for m in words
                if not _keep(m.group(0)) and m.start() != trail_at]
    # An all-initials name ("M & M", "J&J", "A. B.") has no mappable token, so
    # the loop below would keep every character verbatim and return the input —
    # a self-map that never scrubs. Fake the initials distinctly instead.
    if not mappable and re.search(r"[A-Za-z]", name):
        return _pn_fake_initials_name(name, registry), []
    surname_at = None
    if mappable:
        if "," in name:
            comma = name.index(",")
            before = [m for m in mappable if m.start() < comma]
            surname_at = (before[0] if before else mappable[0]).start()
        else:
            surname_at = mappable[-1].start()
    parts, bare, cursor = [], [], 0
    for m in words:
        w = m.group(0)
        parts.append(name[cursor:m.start()])
        if _keep(w) or m.start() == trail_at:
            parts.append(w)
        else:
            fake = _pn_fake_name_token(w, registry)
            parts.append(fake)
            is_surname = (m.start() == surname_at)
            # A surname as short as two letters (Yu, Ng, Le, Wu, Vo) is real and
            # identifying — the old `len >= 3` floor dropped it, so a bare "Yu"
            # both leaked in the text and had no token row for the reversal
            # macro. Single letters are already held out by `_keep`. A non-
            # surname (first/middle) still needs four letters to earn its own
            # token; the common-word guard in `_pn_append_person_terms` keeps a
            # two-letter word ("As", "Of") from becoming a term.
            if _pn_is_name_token(w) and (is_surname or len(w) >= 4):
                bare.append((w, fake, is_surname))
        cursor = m.end()
    parts.append(name[cursor:])
    return "".join(parts), bare


def _pn_name_token_rows(real, fake):
    """Aligned (real_word, fake_word) pairs for a composed full name.

    The reversal macro replaces token by token, so it needs a row per first and
    last name — "Kingsley" -> "Yu" on its own, not only inside the full-name row
    "Ivers Kingsley" -> "Gregory Yu". A person/entity fake is built word-for-word
    from the real, so zipping the two word lists recovers each token's fake.
    Yields nothing when the sides don't tokenize to the same length (a fake that
    was not composed word-for-word), so a mis-aligned pair is never emitted.
    Identical words (kept initials, shared verbatim tokens) and lone initials are
    skipped — there is nothing to reverse there."""
    rw = _PN_WORD_RE.findall(str(real))
    fw = _PN_WORD_RE.findall(str(fake))
    if not rw or len(rw) != len(fw):
        return
    for rtok, ftok in zip(rw, fw):
        if rtok == ftok:
            continue
        if len(rtok.strip(".'’-")) < 2:
            continue
        yield rtok, ftok


def _pn_word_affixes(tok):
    """(prefix, core, suffix) — strips surrounding punctuation and a trailing
    possessive so the core word is what gets faked: "(Smith's," -> ("(",
    "Smith", "'s,"). Keeps "L.L.C." intact as a core (inner dots are kept)."""
    i, j = 0, len(tok)
    while i < j and not tok[i].isalnum():
        i += 1
    while j > i and not tok[j - 1].isalnum():
        j -= 1
    pre, core, post = tok[:i], tok[i:j], tok[j:]
    if len(core) > 2 and core[-2] in "'\u2019" and core[-1] in "sS":
        core, post = core[:-2], core[-2:] + post
    return pre, core, post


def _pn_word_base(word):
    """Canonical lookup key for a name word: lower-cased, outer punctuation and
    any possessive removed. "SMITH'S" and "Smith" share the key "smith"."""
    return _pn_word_affixes(word)[1].lower()


def _pn_is_entity_keep(base):
    """True for a word that survives entity faking verbatim: a connector, a
    corporate suffix in ANY punctuation style ("llc" / "L.L.C." / "l.l.c"), or a
    "dba" marker. The suffix set is matched on its punctuation-stripped form —
    a plain `strip(".,")` leaves "l.l.c", which is not in the literal set, so
    "L.L.C." used to be replaced by a fake word."""
    if base in _PN_ENTITY_KEEP:
        return True
    if re.sub(r"[.\s]", "", base) in _PN_ENTITY_SUFFIXES_NORM:
        return True
    return re.sub(r"[./\s]", "", base) in _PN_DBA_NORM


# Connective words an initialism may skip or fold in ("Tom OF Finland" -> the
# "o" in ToFF, or dropped entirely in TFF), and the narrow set of purely legal
# corporate suffixes that a trailing acronym letter never stands for ("Inc.").
# Deliberately excludes name-bearing suffix words like "Foundation"/"Group"
# that DO contribute a letter (ToFF's second F is Foundation).
_PN_ACRONYM_CONNECTIVES = {"of", "and", "the", "for", "a", "an", "de", "la", "&"}
_PN_ACRONYM_DROP_SUFFIX = {
    "inc", "incorporated", "llc", "llp", "lp", "corp", "corporation", "co",
    "company", "ltd", "pc", "plc", "na", "apc", "pllc",
}


def _pn_acronym_trim(words):
    """Drop a trailing purely-legal corporate suffix ("Inc.", "LLC") so it does
    not have to be represented by an acronym letter. Name-bearing words are
    kept."""
    out = list(words)
    while out and re.sub(r"[.\s]", "", _pn_word_base(out[-1])) in _PN_ACRONYM_DROP_SUFFIX:
        out.pop()
    return out


def _pn_initialism_fake(short, real_words, fake_words):
    """If single-token `short` is an initialism of `real_words` (initials in
    order; a connective word's initial may be included or skipped; a trailing
    legal suffix is ignored), return the matching acronym built from the aligned
    `fake_words`, re-cased like `short`. Otherwise None.

    "ToFF" over ["Tom","of","Finland","Foundation","Inc."] with the fake
    ["Kaldor","of","Ironbridge","Foundation","Inc."] -> "KoIF": every name word
    contributes its fake's initial, so the short form stays the same party and
    the reversal key round-trips. Returns None (caller falls back) when the real
    and fake word lists don't align 1:1."""
    letters = re.sub(r"[^A-Za-z]", "", short)
    if len(letters) < 2:
        return None
    rw = _pn_acronym_trim(real_words)
    fw = _pn_acronym_trim(fake_words)
    if not rw or len(rw) != len(fw):
        return None
    low = letters.lower()
    out, i = [], 0
    for r_word, f_word in zip(rw, fw):
        base = _pn_word_base(r_word)
        if i < len(low) and base[:1] == low[i]:
            f_init = (re.sub(r"[^A-Za-z]", "", f_word)[:1] or base[:1])
            out.append(f_init.upper() if letters[i].isupper() else f_init.lower())
            i += 1
        elif base in _PN_ACRONYM_CONNECTIVES:
            continue          # an unused connective ("of" in TFF) is allowed
        else:
            return None        # a name word with no acronym letter -> not one
    return "".join(out) if i == len(low) else None


def _pn_paren_short_forms(body):
    """Candidate short forms inside a parenthetical definition body. Every
    double-quoted alternative is returned (`"TOFF" OR "TOM OF FINLAND
    FOUNDATION"` -> two), so a multi-alternative definition is fully harvested.
    When nothing is quoted, only a leading capitalised short phrase is taken
    (mirroring the original pattern) so a descriptive parenthetical isn't
    swept in. Double-quote delimiters only, so an apostrophe inside a name
    ("O'Brien") is never mistaken for a quote."""
    quoted = re.findall(r'["“”]\s*([^"“”]{1,60}?)\s*["“”]', body)
    forms = []
    for s in (quoted or [body]):
        s = re.sub(r"\s+", " ", s).strip(" .,;")
        if not quoted:
            mm = re.match(r"[A-Z][\w'’.\-]*(?:\s+[A-Z0-9][\w'’.\-]*){0,3}", s)
            if not mm:
                continue
            s = mm.group(0).strip(" .,")
        if s:
            forms.append(s)
    return forms


def _pn_fake_entity_parts(name, registry, prefer=None):
    """(fake_name, {word_base: canonical_fake}) — the entity faked word by word.

    `prefer` maps a word base to the fake already minted for that same word in
    this party's legal name (possibly from the person pool). A preferred fake
    always wins, including over an entity-pool binding some UNRELATED party's
    name made for the same word — otherwise "John Smith dba Smith's Auto Body"
    would render its dba with the fake belonging to a different Smith."""
    out, mapping = [], {}
    prefer = prefer or {}
    toks = name.split()
    # A state / DC name is geography, not identity — keep it verbatim and fake
    # only the other words ("California Pizza Kitchen" -> "California <fake>").
    state_keep = _pn_state_keep_flags([_pn_word_affixes(t)[1].lower() for t in toks])
    for idx, tok in enumerate(toks):
        pre, core, post = _pn_word_affixes(tok)
        base = core.lower()
        if (not core or _pn_is_entity_keep(base) or state_keep[idx]
                or not re.search(r"[A-Za-z]", core)):
            out.append(tok)
            continue
        fake = prefer.get(base) or registry.token(base, _PN_ENTITY_WORDS, "enttok")
        mapping[base] = fake
        out.append(pre + _pn_titlecase_like(fake, core) + post)
    return " ".join(out), mapping


def _pn_fake_entity(name, registry, prefer=None):
    return _pn_fake_entity_parts(name, registry, prefer)[0]


def _pn_person_token_map(name, registry):
    """{word_base: canonical_fake} for every mappable word of a person name,
    using the same per-word fakes `_pn_fake_person` assigns."""
    out = {}
    for m in _PN_WORD_RE.finditer(name):
        w = m.group(0)
        if _pn_is_suffix_token(w):
            continue
        base = _pn_word_base(w)
        if base:
            out[base] = registry.token(w, _PN_NAME_WORDS, "nametok")
    return out


def _pn_entity_bare(name, registry, prefer=None):
    """Return (bare_real, bare_fake) for the entity name with its trailing
    corporate suffix/connector(s) removed — e.g. "Air Tutors LLC" -> "Air
    Tutors" — so the entity is still pseudonymized when it's written WITHOUT
    the suffix. Only when the remaining core is a distinctive MULTI-word phrase;
    a single leftover word (often an ordinary English word like "Redwood" or
    "Tutors") is skipped to avoid corrupting unrelated prose. Returns None when
    there's nothing safe to add. The bare fake is computed the same per-word
    way as the full name, so both forms stay consistent."""
    toks = name.split()
    while toks:
        base = _pn_word_base(toks[-1])
        if _pn_is_entity_keep(base) or not re.search(r"[A-Za-z0-9]", toks[-1]):
            toks.pop()
        else:
            break
    bare = re.sub(r"[.,;]+$", "", " ".join(toks)).strip()
    # A digit token is as distinctive as a name word ("Mortgage 2000"), so it
    # counts toward the two-word minimum and is never popped as a suffix.
    core = [t for t in toks
            if re.search(r"[A-Za-z0-9]", t) and not _pn_is_entity_keep(_pn_word_base(t))]
    if len(core) < 2 or not bare:
        return None
    if bare.lower() == re.sub(r"[.,;]+$", "", name.strip()).lower():
        return None  # nothing was stripped; the full-name term already covers it
    return bare, _pn_fake_entity(bare, registry, prefer)


def _pn_append_entity_terms(terms, raw, source, registry, prefer=None):
    """Register the full entity term plus, when safe, a suffix-stripped
    'bare' variant that catches the name written without its corporate
    suffix. Returns the {word_base: fake} map so a dba sharing words with this
    name can reuse the same fakes."""
    fake, mapping = _pn_fake_entity_parts(raw, registry, prefer)
    terms.append(_PnTerm("entity", raw, fake,
                         whole_word=False, case_sensitive=False, priority=2,
                         source=source))
    bare = _pn_entity_bare(raw, registry, prefer)
    if (bare and not _pn_is_party_role(bare[0])
            and not (len(bare[0].split()) <= 2
                     and all(_pn_is_generic_token(_pn_word_base(w))
                             for w in bare[0].split()))):
        terms.append(_PnTerm("entity-token", bare[0], bare[1], whole_word=True,
                             case_sensitive=False, priority=1, source=source))
    return mapping


def _pn_split_dba(raw):
    """['SMITH AUTO GROUP, INC.', 'TOYOTA OF DOWNTOWN'] for a name written with
    a dba marker; a single-element list otherwise. The first element is the
    legal name, the rest are fictitious business names."""
    parts = [re.sub(r"\s+", " ", p).strip().strip(",;()\"' ")
             for p in _PN_DBA_SPLIT_RE.split(raw)]
    return [p for p in parts if p]


def _pn_split_aka(raw):
    """['Tanooa Sparks', 'Tanooa Sherrell Sparks'] for a name written with an
    aka/fka marker. Every element is the same kind of thing as the first."""
    parts = [re.sub(r"\s+", " ", p).strip().strip(",;()\"' ")
             for p in _PN_AKA_SPLIT_RE.split(raw)]
    return [p for p in parts if p]


def _pn_append_person_terms(terms, raw, source, registry):
    """Register a person's full name plus its safe bare tokens, and — at low
    priority, sharing each token's fake — its near-spelling variants, so a
    surname the key spells one way is still scrubbed where a document spells it
    another ("Roxane" vs "Roxanne")."""
    fake_full, bare = _pn_fake_person(raw, registry)
    terms.append(_PnTerm("person", raw, fake_full, whole_word=False,
                         case_sensitive=False, priority=2, source=source))
    for real_tok, fake_tok, _is_surname in bare:
        # A generic word ("Legal" of a firm name, "Warranty", "Beach" of
        # "Long Beach") must never be a free-standing token — it rewrites
        # ordinary prose and the venue everywhere. The full name still scrubs.
        base = _pn_word_base(real_tok)
        if _pn_is_generic_token(base):
            continue
        # A two-letter surname that spells an ordinary English word ("As", "Of",
        # "In") would rewrite that word throughout the brief; skip it. A short
        # surname that is NOT a word (Yu, Ng, Le, Wu) still registers.
        if len(base) <= 2 and base in _PN_SHORT_TOKEN_STOP:
            continue
        terms.append(_PnTerm("person-token", real_tok, fake_tok,
                             whole_word=True, case_sensitive=False,
                             priority=1, source=source))
        for var in _pn_name_variants(real_tok):
            terms.append(_PnTerm("person-token", var, fake_tok,
                                 whole_word=True, case_sensitive=False,
                                 priority=0, source=source, derived=True))
        # A HYPHENATED surname is one token to the tokenizer, so nothing else
        # can see its halves — and a brief uses them freely: a line wrap opens
        # the hyphen ("Ardeshirpour- Zartoshti"), a shorthand reference drops
        # one half ("Dr. Ardeshirpour"). The whole-token term matches neither,
        # and the surviving half stood beside the faked given name as a half-
        # scrubbed pair. Register the wrap-split spelling (same fake — it IS
        # the same token) and each half as its own low-priority token with its
        # own registry fake, so every form is scrubbed and reversible.
        core = _pn_word_affixes(real_tok)[1]
        halves = [p for p in core.split("-") if p]
        if len(halves) >= 2 and all(
                len(p) >= 4 and _pn_is_name_token(p)
                and not _pn_is_generic_token(_pn_word_base(p))
                for p in halves):
            for k in range(1, len(halves)):
                split_var = ("-".join(halves[:k]) + "- "
                             + "-".join(halves[k:]))
                terms.append(_PnTerm("person-token", split_var, fake_tok,
                                     whole_word=True, case_sensitive=False,
                                     priority=0, source=source, derived=True))
            for part in halves:
                terms.append(_PnTerm("person-token", part,
                                     _pn_fake_name_token(part, registry),
                                     whole_word=True, case_sensitive=False,
                                     priority=0, source=source))
    return _pn_person_token_map(raw, registry)


def _pn_append_shortname_term(terms, raw, source, registry):
    """Register the court's own short form for a party ("(Dream Team)") as a
    whole-phrase term and nothing more.

    No bare tokens are expanded from it. A short form is chosen to be brief, not
    distinctive — run through the person path, "Dream Team" would register
    "Team" and rewrite the word wherever it appeared. The fake is composed from
    the entity word pool, so it reuses the fakes already minted for the full
    name and the two forms stay recognisably the same party."""
    if _pn_is_party_role(raw) or len(raw.split()) < 2:
        return
    terms.append(_PnTerm("short-name", raw, _pn_fake_entity(raw, registry),
                         whole_word=True, case_sensitive=False, priority=1,
                         source=source))


_PN_TITLE_CONNECTORS = frozenset({"to", "of", "for", "re", "on", "against"})


def _pn_strip_pleading_title_prefix(name):
    """A harvested name cell sometimes carries a DOCUMENT TITLE, not a bare
    party: "DEMURRER TO RESCORE HOLLYWOOD, LLC". Registered whole, the
    pleading words become part of the "party" (and their bare tokens rewrite
    every later "Demurrer" in the briefs). Strip a leading run of pleading/
    procedural words ending in a connector so only the party remains — and
    only that shape: "Motion Picture Industry Pension Plan" has no connector
    after its pleading word, so a real name that merely STARTS with one is
    never truncated. Returns the party remainder, or `name` unchanged."""
    words = name.split()
    i = 0
    while True:
        j = i
        while j < len(words) and _pn_word_base(words[j]) in _PN_PLEADING_WORDS:
            j += 1
        if j == i or j >= len(words):
            break
        if _pn_word_base(words[j]) not in _PN_TITLE_CONNECTORS:
            break
        i = j + 1
    if not i:
        return name
    rest = " ".join(words[i:])
    # "MOTION TO STRIKE": the "remainder" is itself procedural — the whole
    # cell is a title with no party in it, so there is nothing to register.
    if (not re.search("[" + _PN_LAT_UPPER + "]", rest)
            or _pn_is_procedural_phrase(rest)):
        return ""
    return rest


def _pn_append_name_terms(terms, raw, source, registry):
    """Register every term implied by one raw name cell.

    Three things can be packed into that cell, and each is unpacked here:

      * a CAPACITY phrase, which names a second real party and a run of ordinary
        words ("Kristopher Edwards as successor in interest to Erika Edwards");
      * an ALIAS (aka/fka), which is another name for the same kind of thing;
      * a DBA, which is always a business name whatever the head is.

    A dba is registered as its OWN entity term, because a brief refers to the
    party by the legal name or by the fictitious name, hardly ever by the joined
    caption string. When two names share words, the shared words keep one fake
    apiece: the first name is faked and its word map is preferred for the rest,
    so "Culver City Toyota, Inc. dba Toyota of Culver City" yields fakes that
    share the same three words in the same order."""
    for chunk in _pn_split_capacities(raw):
        if _pn_is_party_role(chunk):  # a bare role label is never a name
            continue
        aliases = _pn_split_aka(chunk)
        prefer, head_is_entity = {}, None
        for alias in aliases:
            parts = _pn_split_dba(alias)
            head, dbas = parts[0], parts[1:]
            # A cell carrying a document TITLE names its target party after
            # the pleading words ("DEMURRER TO RESCORE HOLLYWOOD, LLC") — or
            # nobody at all ("MOTION TO STRIKE"). Register the party, not the
            # pleading vocabulary.
            head = _pn_strip_pleading_title_prefix(head)
            if (not head or _pn_is_party_role(head)
                    or not re.search(r"[A-Za-z]", head)
                    or _PN_ONLY_DESCRIPTOR_RE.match(head)):
                continue
            # Drop a trailing state-of-organization descriptor ("Acme, Inc., a
            # Delaware corporation" -> "Acme, Inc."). The descriptor names no
            # one, so it stays verbatim in the document; only the distinctive
            # name in front of it becomes a term. A descriptor is a decisive
            # entity signal, so its presence forces the entity path.
            head, had_descriptor = _pn_strip_entity_descriptor(head)
            # A SHORT candidate made entirely of generic words is a header
            # cell or a harvested label ("Name", "Customer Relations",
            # "Warranty Complaints"), never a party — registering one rewrites
            # ordinary prose document-wide (the run that replaced all 140
            # occurrences of the word "Name" with a surname). A LONGER
            # all-generic name ("Strategic Legal Practices") is a plausible
            # firm and stays: its whole-phrase match is safe, and its generic
            # bare tokens are suppressed separately.
            head_words = [w for w in head.split() if re.search(r"[A-Za-z]", w)]
            if 1 <= len(head_words) <= 2 and all(
                    _pn_is_generic_token(_pn_word_base(w)) or
                    _pn_is_entity_keep(_pn_word_base(w))
                    for w in head_words):
                continue
            # A government/public-entity party is public record and shares its
            # text with the venue line — keep it verbatim, don't build a term.
            if _pn_is_public_entity(head):
                continue
            if head_is_entity is None:
                head_is_entity = had_descriptor or _pn_looks_like_entity(head)
            if head_is_entity:
                prefer.update(_pn_append_entity_terms(terms, head, source,
                                                      registry, prefer))
            else:
                prefer.update(_pn_append_person_terms(terms, head, source,
                                                      registry))
            # A fictitious business name is a business name whatever its head
            # is, so it always takes the entity path (never the person pool).
            for dba in dbas:
                if _pn_is_party_role(dba) or _pn_is_public_entity(dba):
                    continue
                prefer.update(_pn_append_entity_terms(terms, dba, source,
                                                      registry, prefer))


# A California case number opens with the two-digit filing year ("25STCV37838").
# That year is printed beside the number in every caption ("Complaint Filed: Dec
# 29, 2025"), so randomising it does not hide anything — it only makes the fake
# internally inconsistent. Keep it; fake the rest.
_PN_CASENO_YEAR_RE = re.compile(r"^\d{2}(?=[A-Za-z])")


def _pn_fake_caseno(real, registry):
    keep = 2 if _PN_CASENO_YEAR_RE.match(real) else 0
    return registry.digits(real, "caseno", keep_prefix=keep)


# Court FORM identifiers and field labels that are public boilerplate, never a
# party or a real case number: a Judicial Council form number ("CIV-100"), the
# "CASE NUMBER" field label, the "Default Only" checkbox. Faking one corrupts
# the form (and a form number faked as a case number is a nonsense stamp), so
# these are never registered as terms and never flagged for review. Compared on
# an alphanumeric-only, case-folded reduction, so spacing/dash variants
# ("CIV 100", "Case Number") all match. Extend as more form boilerplate turns up.
_PN_NEVER_FAKE = frozenset({"casenumber", "civ100", "defaultonly"})


def _pn_is_never_fake(value):
    return re.sub(r"[^a-z0-9]", "", str(value).lower()) in _PN_NEVER_FAKE


def _pn_reapply_digits(fake_digits, real):
    """Lay `fake_digits` back into `real`'s exact template: each real DIGIT
    position takes the next fake digit; every separator, bracket and letter is
    kept verbatim. "(626) 381-9893" -> "(NNN) NNN-NNNN"; "626.381.9893" keeps
    its dots; a run-together "123456789" stays run-together. Format in == format
    out, so a re-scan of the export is a fixed point."""
    out, i = [], 0
    for ch in real:
        if ch.isdigit() and i < len(fake_digits):
            out.append(fake_digits[i])
            i += 1
        else:
            out.append(ch)
    return "".join(out)


# NANP validity: an area code and a central-office (exchange) code are each
# [2-9] then any two digits, minus the N11 service codes (211, 911, ...). A fake
# phone that randomises every digit comes out facially bogus ~40% of the time
# (area/exchange leading 0 or 1), which reads as an error in a filing and can
# even land on a real, in-service number; a valid NANP shape reads naturally and
# still identifies no one.
def _pn_nanp_nxx(rng):
    a = rng.randrange(2, 10)
    while True:
        b, c = rng.randrange(10), rng.randrange(10)
        if not (b == 1 and c == 1):        # reject N11 service codes
            return f"{a}{b}{c}"


def _pn_phone_digits(rng, n):
    """`n` fake digits shaped like a valid NANP number: 7 == exchange+line,
    10 == area+exchange+line, 11 keeps a leading country-code 1. Any other
    length falls back to per-digit random (nothing to validate)."""
    if n == 11:
        return "1" + _pn_nanp_nxx(rng) + _pn_nanp_nxx(rng) + f"{rng.randrange(10000):04d}"
    if n == 10:
        return _pn_nanp_nxx(rng) + _pn_nanp_nxx(rng) + f"{rng.randrange(10000):04d}"
    if n == 7:
        return _pn_nanp_nxx(rng) + f"{rng.randrange(10000):04d}"
    return "".join(str(rng.randrange(10)) for _ in range(n))


def _pn_fake_phone(real):
    digits = re.sub(r"\D", "", real)
    return _pn_reapply_digits(_pn_phone_digits(_pn_rng("phone", digits), len(digits)),
                              real)


# A scanned RISC/contract date whose slashes the OCR dropped or read as "1"
# collapses "06/06/2025" into the 10-digit run "0610612025", which the phone
# detector's bare-digit branch then matches and fakes — silently rewriting a
# load-bearing date. A run whose leading pair is a month (01-12) can never be a
# valid NANP number anyway (an area code never starts with 0 or 1), so rejecting
# these date shapes drops no real phone. Only the separator-free form is
# ambiguous; a run carrying any phone punctuation is left to the detector.
def _pn_phone_is_ocr_date(match):
    if re.search(r"\D", match):          # has a separator -> a real phone
        return False
    d = match
    if len(d) != 10 or d[2] != "1" or d[5] != "1":
        return False
    mm, dd, yyyy = int(d[0:2]), int(d[3:5]), int(d[6:10])
    return 1 <= mm <= 12 and 1 <= dd <= 31 and 1900 <= yyyy <= 2099


# SSN validity: the SSA never issues area 000, 666 or 900-999, group 00, or
# serial 0000. A fake in those ranges is recognisably impossible (~10% of a
# blind randomisation lands there); a valid-shaped fake reads naturally.
def _pn_ssn_digits(rng):
    area = rng.randrange(1, 900)
    if area == 666:
        area = 665
    return f"{area:03d}{rng.randrange(1, 100):02d}{rng.randrange(1, 10000):04d}"


def _pn_fake_ssn(real):
    # Seed on digits only so every separator spelling of one SSN
    # ("123-45-6789", "123 45 6789", "123456789") maps to the same fake, then
    # re-apply the original punctuation.
    return _pn_reapply_digits(_pn_ssn_digits(_pn_rng("ssn", re.sub(r"\D", "", real))),
                              real)


def _pn_fake_email(real):
    r = _pn_rng("email", real.lower())
    return (f"{r.choice(_PN_NAME_WORDS).lower()}.{r.choice(_PN_NAME_WORDS).lower()}"
            f"@{r.choice(_PN_EMAIL_DOMAINS)}")


# ── Address handling ────────────────────────────────────────────────────────
# Street-type suffixes the detector recognises, in every spelling seen in the
# corpus. Court/Way/Place/Terrace/Square are back (an earlier build dropped them
# for "Superior Court" / "by way of" false positives, yet _pn_fake_street EMITS
# Court/Way/Place — so the tool produced addresses it could not re-recognise,
# breaking idempotency). They are safe here because the whole pattern is
# anchored on a leading street NUMBER on the same row: "\d+ ... Court" never
# matches "Superior Court".
_PN_ADDR_SUFFIX = (
    r"(?i:Street|St|A[ \t]?venue|Ave|Boulevard|Blvd|Road|Rd|Lane|Ln|Drive|Dr|"
    r"Circle|Cir|Highway|Hwy|Parkway|Pkwy|Trail|Trl|Court|Ct|Place|Pl|Way|"
    r"Terrace|Ter|Square|Sq)")   # "A[ ]?venue" tolerates the OCR split "A venue"
# A number component: a single number, or a hyphenated RANGE captured as ONE
# span ("414–416"), so the leading half is never left stranded next to a fake.
_PN_ADDR_NUM = r"\d{1,6}(?:[ \t]*[-–—][ \t]*\d{1,6})?"
# A name word: a capitalised word (no IGNORECASE, so "the court" can't qualify),
# an ordinal ("5th", "101st"), or a directional ("N", "S.", "NW").
_PN_ADDR_WORD = r"(?:[NSEW]{1,3}\.?|\d{1,3}(?:st|nd|rd|th)|[A-Z][A-Za-z0-9.'&-]*)"
# Optional trailing suite/floor and City, ST ZIP, so the locality is kept too.
# The floor form is "<ordinal> Floor" — value BEFORE the label, unlike Suite.
_PN_ADDR_SUITE = (
    r"(?:[ \t]*,?[ \t]*(?:(?i:Suite|Ste|Unit|Apt|Bldg|#)[ \t]*[\w-]+"
    r"|\d{1,3}(?:st|nd|rd|th)[ \t]+(?i:Floor|Fl)\b\.?))?")
_PN_ADDR_CITYZIP = (
    r"(?:[ \t]*,?[ \t]*(?:[A-Z][A-Za-z]+[ \t]*,?[ \t]*){1,3}[A-Za-z]{2}\.?[ \t]+"
    r"\d{5}(?:-\d{4})?)?")
# A street with no street-TYPE suffix ("1888 Century Park East") may close on a
# trailing directional or plaza-head word instead — but ONLY when a suite/floor
# token or a City, ST ZIP tail actually follows, so prose that merely starts
# with a number ("24 Hour Fitness Center opened") can never qualify by shape.
_PN_ADDR_CLOSE = r"(?i:East|West|North|South|Park|Center|Centre|Plaza)"
_PN_ADDR_TAIL_CUE = (
    r"[ \t]*,?[ \t]*(?:(?i:Suite|Ste|Unit|Apt|Bldg|#)[ \t]*[\w-]"
    r"|\d{1,3}(?:st|nd|rd|th)[ \t]+(?i:Floor|Fl)\b"
    r"|(?:[A-Z][A-Za-z]+[ \t]*,?[ \t]*){1,3}[A-Za-z]{2}\.?[ \t]+\d{5})")
_PN_ADDR_STREET_PAT = (
    rf"{_PN_ADDR_NUM}[ \t]+"
    rf"(?:(?:{_PN_ADDR_WORD}[ \t]+){{1,4}}{_PN_ADDR_SUFFIX}\b\.?"
    rf"|(?:{_PN_ADDR_WORD}[ \t]+){{1,3}}{_PN_ADDR_CLOSE}\b(?={_PN_ADDR_TAIL_CUE}))")
_PN_ADDR_RE = re.compile(
    rf"\b{_PN_ADDR_STREET_PAT}{_PN_ADDR_SUITE}{_PN_ADDR_CITYZIP}")

_PN_ADDR_ABBR = {
    "st": "Street", "ave": "Avenue", "av": "Avenue", "blvd": "Boulevard",
    "rd": "Road", "dr": "Drive", "ln": "Lane", "hwy": "Highway",
    "pkwy": "Parkway", "cir": "Circle", "ct": "Court", "pl": "Place",
    "ter": "Terrace", "sq": "Square", "trl": "Trail",
}
_PN_ADDR_DIRS = {"n": "N", "s": "S", "e": "E", "w": "W", "ne": "NE",
                 "nw": "NW", "se": "SE", "sw": "SW"}


def _pn_addr_canon(real):
    """Fold an address to one canonical spelling so `21225 Pacific Coast Hwy`,
    `21225 Pacific Coast Highway`, and `21225 PACIFIC COAST HWY` seed ONE fake —
    the audited run gave one office three fakes. Expands the suffix and
    directional abbreviations, lower-cases, and collapses punctuation/space."""
    toks = re.split(r"[ \t]+", real.strip())
    out = []
    for t in toks:
        base = t.strip(".,").lower()
        if base in _PN_ADDR_ABBR:
            out.append(_PN_ADDR_ABBR[base].lower())
        elif base in _PN_ADDR_DIRS:
            out.append(_PN_ADDR_DIRS[base].lower())
        else:
            out.append(base)
    return re.sub(r"[^a-z0-9 ]", "", " ".join(out)).strip()


def _pn_addr_street_key(real):
    """(street-identity, [house-numbers]) — the directional + name + suffix
    core, with the house number(s) AND any City/ST/ZIP tail removed, so every
    spelling of one parcel folds to a single key. Keying the fake on THIS
    (not on the whole address string) is what makes `414 S. Maple Ave.`,
    `414 S. Maple Ave. Montebello, CA 90640`, and `414-416 S. Maple Ave.` share
    one faked street instead of the four the audited run produced.
    `414–416 S Maple Ave, Montebello, CA 90640` -> ("s maple avenue", [414, 416]).
    """
    street, _suite, _cz = _pn_addr_parts(real)
    if not street:
        street = real
    nums = [int(n) for n in re.findall(r"\d{1,6}", street)]
    # canon of the STREET part only (no locality), then drop the leading
    # number(s) / range so only the directional+name+suffix identity remains.
    core = re.sub(r"^[\d\- \u2013\u2014]+", "", _pn_addr_canon(street)).strip()
    return core, nums


# Capture the street-type suffix as it was actually written (with any trailing
# period), so the fake can reuse it. The suffix is a generic road TYPE, not an
# identifier — turning a real "Road" into a fake "Court" changes the record for
# nothing — so only the number and name words are faked; Street/Road/Court/Way…
# are always left as they stand.
_PN_ADDR_SUFFIX_OF_RE = re.compile(
    rf"{_PN_ADDR_NUM}[ \t]+(?:{_PN_ADDR_WORD}[ \t]+){{1,4}}({_PN_ADDR_SUFFIX}\b\.?)")


def _pn_addr_suffix_of(real):
    """The street-type suffix of `real`, verbatim ("Ave.", "Court", "Highway"),
    or "Street" if none is found."""
    m = _PN_ADDR_SUFFIX_OF_RE.search(real)
    return m.group(1) if m else "Street"


_PN_ADDR_STRIP_SUFFIX_RE = re.compile(rf"[ \t]+{_PN_ADDR_SUFFIX}\b\.?\s*$")


def _pn_addr_strip_suffix(street):
    """A street with its trailing street-type suffix removed — the house number
    + name only: "1055 E Colorado Blvd." -> "1055 E Colorado", "4372 Cedar Ave."
    -> "4372 Cedar". Used to register a suffix-less address FRAGMENT so a
    corrupted scan that wraps the suffix off ("1055 E Colorado" | "Blvd.") is
    still scrubbed."""
    return _PN_ADDR_STRIP_SUFFIX_RE.sub("", street.strip()).strip()


# `_PN_ADDR_RE` deliberately consumes an optional suite and City, ST ZIP so the
# locality is scrubbed too — but the fakers used to emit only "<num> <name>
# <suffix>", so the tail was DELETED rather than replaced:
#     "414 S. Maple Ave. Montebello, CA 90640"  ->  "1533 Cypress Ave."
# The export lost the fact that there was a city at all, while "residing in
# Montebello, California" three lines below survived untouched. These parse the
# match back into its parts so each is replaced in kind.
_PN_ADDR_PARSE_RE = re.compile(
    rf"^(?P<street>{_PN_ADDR_STREET_PAT})"
    rf"(?P<suite>{_PN_ADDR_SUITE})(?P<cityzip>{_PN_ADDR_CITYZIP})$")
_PN_CITYZIP_PARSE_RE = re.compile(
    r"^(?P<lead>[ \t]*,?[ \t]*)(?P<city>.+?)[ \t]*,?[ \t]*"
    r"(?P<state>[A-Za-z]{2}\.?)[ \t]+(?P<zip>\d{5}(?:-\d{4})?)$")


def _pn_addr_parts(real):
    """(street, suite, cityzip) for an address matched by `_PN_ADDR_RE`.
    Missing parts come back as empty strings."""
    m = _PN_ADDR_PARSE_RE.match(real.strip())
    if not m:
        return real.strip(), "", ""
    return m.group("street"), m.group("suite") or "", m.group("cityzip") or ""


def _pn_fake_cityzip(cityzip, city_word, zip_digits):
    """Rebuild a "…, City, ST 90640" tail with a fake city and ZIP, keeping the
    separators and the STATE exactly as written. Returns "" for an empty tail."""
    if not cityzip:
        return ""
    m = _PN_CITYZIP_PARSE_RE.match(cityzip)
    if not m:
        return cityzip
    return (f"{m.group('lead')}{city_word}, {m.group('state')} {zip_digits}")


# A "City, ST 90640" standing on its own line. A pleading's service list prints
# the street and the locality on separate lines, and `_PN_ADDR_RE` may not cross
# a newline (a ZIP at a line end must not weld to a "Street" three lines below),
# so the locality half is invisible to the address detector. `register_localities`
# uses this to learn it anyway.
_PN_CITYSTZIP_LINE_RE = re.compile(
    r"(?m)^[ \t]*(?P<city>[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z.'’-]+){0,2})"
    r"[ \t]*,[ \t]*(?P<state>[A-Z]{2})\.?[ \t]+(?P<zip>\d{5}(?:-\d{4})?)[ \t]*$")

# A city that names the venue is not an identifier — it is the court. Scrubbing
# "Los Angeles" out of "COUNTY OF LOS ANGELES" would make the ruling nonsense.
# Only the LABEL is case-insensitive: an `(?i)` covering the name too would let
# `[A-Z]` match a lower-case word, and "in Los Angeles County" captured as
# "in Los Angeles".
_PN_VENUE_NAME = r"[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z.'’-]+){0,2}"
_PN_VENUE_CITY_RES = (
    re.compile(r"(?i:count(?:y|ies)\s+of)\s+(?P<c>" + _PN_VENUE_NAME + r")"),
    re.compile(r"\b(?P<c>" + _PN_VENUE_NAME + r")\s+(?i:county)\b"),
    re.compile(r"(?i:(?:superior\s+)?court\s+of\s+(?:the\s+)?(?:state\s+of\s+)?)"
               r"(?P<c>" + _PN_VENUE_NAME + r")"),
)


def _pn_venue_cities(text):
    """Lower-cased city/county names that `text` uses to name the venue."""
    out = set()
    for rx in _PN_VENUE_CITY_RES:
        for m in rx.finditer(_NFKC(text)):
            out.add(re.sub(r"\s+", " ", m.group("c")).strip().lower())
    return out


# Localities never anonymized on their own — every California county (which is
# the venue/jurisdiction, not a party identifier; many share their name with a
# city, so faking it makes the record read oddly and helps no one). A county is
# still scrubbed when it is PART OF a party name (e.g. an entity literally named
# "County of Los Angeles"), which the party's own term handles — this set only
# stops a STANDALONE mention from being learned as a locality or faked inside an
# address tail.
_PN_CA_COUNTIES = frozenset({
    "alameda", "alpine", "amador", "butte", "calaveras", "colusa",
    "contra costa", "del norte", "el dorado", "fresno", "glenn", "humboldt",
    "imperial", "inyo", "kern", "kings", "lake", "lassen", "los angeles",
    "madera", "marin", "mariposa", "mendocino", "merced", "modoc", "mono",
    "monterey", "napa", "nevada", "orange", "placer", "plumas", "riverside",
    "sacramento", "san benito", "san bernardino", "san diego", "san francisco",
    "san joaquin", "san luis obispo", "san mateo", "santa barbara",
    "santa clara", "santa cruz", "shasta", "sierra", "siskiyou", "solano",
    "sonoma", "stanislaus", "sutter", "tehama", "trinity", "tulare",
    "tuolumne", "ventura", "yolo", "yuba",
})
_PN_NEVER_SCRUB_LOCALITIES = _PN_CA_COUNTIES
# The individual WORDS of the protected localities ("long", "beach", "los",
# "angeles"): a bare pseudonym token equal to one of these corrupts the venue
# everywhere it appears ("Long Beach" -> "Long Waverly"), so no bare token may
# be one.
_PN_LOCALITY_WORDS = frozenset(
    w for loc in _PN_NEVER_SCRUB_LOCALITIES for w in loc.split()) | frozenset("""
beach hills valley springs grove heights gardens lake canyon ranch palms
shores vista mesa park city long
""".split())


def _pn_is_generic_token(base):
    """True for a word that must never become a BARE pseudonym token, however
    it was harvested: ordinary English / procedural vocabulary ("Warranty",
    "Legal", "Name" — the run that faked every 'warranty' as 'langley' and
    'Legal Standard' as 'Granite Standard' came from exactly these), an
    institution word, or a word of a protected locality ("Beach"). The FULL
    name a token came from is still registered; only the free-standing token
    is withheld, so the cost is a missed bare occurrence (caught by the
    unknown-name net), never a corrupted document."""
    return (base in _PN_COMMON_WORDS or base in _PN_REVIEW_NAME_STOP
            or base in _PN_LOCALITY_WORDS)


def _pn_is_protected_locality(city):
    norm = re.sub(r"\s+", " ", city).strip().lower()
    if norm in _PN_NEVER_SCRUB_LOCALITIES:
        return True
    # "Orange County" / "Los Angeles County" -> the county itself.
    return norm.removesuffix(" county").strip() in _PN_NEVER_SCRUB_LOCALITIES


# A government / public-entity party ("County of Los Angeles", "Los Angeles
# County", "City of Montebello", "People of the State of California", "State of
# California", "United States"). Its involvement is public record, so
# pseudonymizing it protects nothing — and because the same string names the
# VENUE ("Superior Court of California, County of Los Angeles"), scrubbing it as
# a party would corrupt the court reference. Such a party is left as written.
# A PRIVATE business whose name merely CONTAINS a place ("Los Angeles Widgets,
# Inc.") is not matched here and is still scrubbed.
_PN_PUBLIC_ENTITY_RE = re.compile(
    r"(?ix)^\s*(?:the\s+)?(?:"
    r"count(?:y|ies)\s+of\s+[A-Za-z]"        # County of X
    r"|city\s+of\s+[A-Za-z]"                  # City of X
    r"|town\s+of\s+[A-Za-z]"                  # Town of X
    r"|people(?:\s+of\b|\s*$)"               # (The) People / People of the State…
    r"|state\s+of\s+[A-Za-z]"                 # State of X
    r"|united\s+states(?:\s+of\s+america)?"   # United States [of America]
    r")")


def _pn_is_public_entity(name):
    """True for a government / public-entity party that is kept verbatim (see
    above). Covers the "<protected county> County" form too."""
    n = re.sub(r"\s+", " ", name or "").strip().strip('.,')
    if _PN_PUBLIC_ENTITY_RE.match(n):
        return True
    m = re.match(r"(?i)(.+?)\s+county\b", n)
    return bool(m and _pn_is_protected_locality(m.group(1)))


def _pn_locality_pairs(text):
    """[(city, state, zip), ...] every locality `text` states, whether it sits at
    the tail of a full street address or alone on its own line."""
    text = _NFKC(text)
    out, seen = [], set()

    def add(city, state, zipd):
        city = city.strip(" ,")
        key = (city.lower(), zipd)
        if city and key not in seen:
            seen.add(key)
            out.append((city, state, zipd))

    for m in _PN_ADDR_RE.finditer(text):
        _s, _su, cityzip = _pn_addr_parts(m.group(0))
        cm = _PN_CITYZIP_PARSE_RE.match(cityzip) if cityzip else None
        if cm:
            add(cm.group("city"), cm.group("state"), cm.group("zip"))
    for m in _PN_CITYSTZIP_LINE_RE.finditer(text):
        add(m.group("city"), m.group("state"), m.group("zip"))
    return out


def _pn_fake_street(real):
    """Module-level fallback (no registry): a stable but NOT guaranteed-unique
    fake. The Pseudonymizer routes through `_fake_street` for injectivity. Only
    the house number and street NAME are faked; the real street-type suffix, the
    unit/suite, and the whole City/State/ZIP tail are kept verbatim. Keyed on the
    street identity so spelling variants of one parcel share a fake."""
    _street, suite, cityzip = _pn_addr_parts(real)
    core, _nums = _pn_addr_street_key(real)
    r = _pn_rng("street", core)
    fake_street = (f"{r.randrange(100, 9999)} {r.choice(_PN_STREET_NAMES)} "
                   f"{_pn_addr_suffix_of(real)}")
    return fake_street + suite + cityzip


# ── URL / bare-domain handling ──────────────────────────────────────────────
# Citation hosts the authorities appendix emits on purpose — never rewrite them,
# or the public verification links break. The second group is e-filing /
# court-services infrastructure: a One Legal or usLegalPro status page in a
# proof of service names the FILING CHANNEL, not a party — like a `.gov` host
# it identifies no one, so faking it protects nothing and flagging it put a
# worksheet row (and an operator "no") in front of every filing-receipt page.
_PN_URL_WHITELIST = (
    "leginfo.legislature.ca.gov", "law.cornell.edu", "scholar.google.com",
    "courts.ca.gov", "google.com", "casetext.com", "justia.com",
    "onelegal.com", "uslegalpro.com", "lawhelpcalifornia.org",
)


def _pn_url_fragmentary(real):
    """True when a URL match carries no registrable host — a line-wrap split
    the URL and the detector saw only the lead fragment ("http://www",
    "https://", "http://]", a bare "www"). Such a fragment identifies no one,
    so faking it corrupts the wrapped URL's visible text for nothing — and the
    minted record then "survived" as a LEAK row wherever an intact kept URL
    (every whitelisted court link starts with the same "https://www") contained
    the fragment. Not a detector miss to hide: the tail half of the wrap is a
    bare domain and still matches on its own line."""
    host = _pn_url_host(real)
    if "." not in host:
        return True                     # "www", "http", "]" — no domain at all
    stem, _, tld = host.rpartition(".")
    # A registrable host needs an alphabetic TLD and a non-empty label before
    # it ("law.com" qualifies; ".com" and "www." do not).
    return not (stem.strip(".") and re.fullmatch(r"[a-z]{2,}", tld))


def _pn_url_host(real):
    """The bare host of a URL match: scheme, leading www., and any path removed,
    lower-cased. `https://www.Firm.com/bio` -> `firm.com`."""
    m = re.match(r"(?:https?://)?(?:www\.)?([^/\s]+)", real.strip(), re.IGNORECASE)
    host = (m.group(1) if m else real).lower()
    return host.rstrip("/").rstrip(".")


def _pn_url_whitelisted(real):
    host = _pn_url_host(real)
    # A government host names an agency, not a party — faking it protects no
    # one, and flagging it buries the real findings. This must cover the
    # whitelisted hosts' PUBLIC PARENTS too: a line-wrapped court URL leaves a
    # bare "ca.gov" tail on its own line, which the child-only whitelist
    # missed — the detector minted a domain record for it, and that record
    # then "survived" inside every deliberately-kept courts.ca.gov citation
    # link, a LEAK row per file across the whole folder.
    if host in ("gov", "mil") or host.endswith((".gov", ".mil")):
        return True
    return any(host == w or host.endswith("." + w) for w in _PN_URL_WHITELIST)


# Consumer mail providers shared by millions: the host identifies no one, and
# faking it both protects nothing and, once two distinct real hosts collapse
# onto one fake, misleads a reader into thinking opposing parties shared an
# inbox. These pass through unchanged; every other host is faked injectively.
_PN_PUBLIC_EMAIL_DOMAINS = frozenset({
    "gmail.com", "googlemail.com", "yahoo.com", "ymail.com", "yahoo.co.uk",
    "outlook.com", "hotmail.com", "hotmail.co.uk", "live.com", "msn.com",
    "icloud.com", "me.com", "mac.com", "aol.com", "proton.me", "protonmail.com",
    "gmx.com", "gmx.net", "mail.com", "zoho.com", "yandex.com",
})

# Default registry for bare `_pn_fake_domain` calls (module fallbacks / tests).
# The Pseudonymizer passes its own per-case registry so domain fakes reset
# between documents; either way the map stays injective (one real host -> one
# fake, distinct hosts never collide).
_PN_DOMAIN_REGISTRY = _PnFakeRegistry()


def _pn_osa_distance(a, b):
    """Optimal-string-alignment (restricted Damerau-Levenshtein) distance: the
    fewest insertions, deletions, substitutions, and ADJACENT TRANSPOSITIONS
    that turn `a` into `b`. A swapped pair ("adler"/"alder") costs one, not two,
    so an OCR/keying transposition reads as a single slip. Strings are short
    (name tokens, hosts), so the full DP is cheap."""
    la, lb = len(a), len(b)
    if not la:
        return lb
    if not lb:
        return la
    prev2 = None
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        cur = [i] + [0] * lb
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
            if (prev2 is not None and i > 1 and j > 1
                    and a[i - 1] == b[j - 2] and a[i - 2] == b[j - 1]):
                cur[j] = min(cur[j], prev2[j - 2] + 1)   # adjacent transposition
        prev2, prev = prev, cur
    return prev[lb]


def _pn_edit_distance_within(a, b, k, min_len=8):
    """True when `a` and `b` are within OSA edit distance `k` (see
    `_pn_osa_distance`). Requires both strings be at least `min_len` long so two
    genuinely different SHORT values aren't folded, and short-circuits on the
    length gap before running the DP."""
    if a == b:
        return True
    if min(len(a), len(b)) < min_len or abs(len(a) - len(b)) > k:
        return False
    return _pn_osa_distance(a, b) <= k


def _pn_edit_distance_le1(a, b, min_len=8):
    """True when `a` and `b` differ by at most one insertion, deletion,
    substitution, or adjacent transposition — a single OCR/keying slip (e.g.
    autolegalgrouo vs autolegalgroup, or the swapped "ADLER"/"ALDER"). Requires
    both strings be at least `min_len` long. Thin `k == 1` wrapper over
    `_pn_edit_distance_within`; the domain/host fold and single-typo callers use
    it, while the name fold scales `k` up with token length."""
    return _pn_edit_distance_within(a, b, 1, min_len)


def _pn_name_fold_dist(a, b):
    """The edit distance at which two NAME tokens still fold onto one fake, by
    length: the longer a token, the more independent typos it plausibly carries,
    and the smaller the chance two genuinely different surnames sit that close.
    Short names stay strict (one slip) so distinct four/five-letter surnames are
    never linked; long tokens tolerate two or three."""
    L = min(len(a), len(b))
    if L >= 16:
        return 3
    if L >= 10:
        return 2
    return 1


def _pn_fake_domain(domain, registry=None):
    """A stable, injective fake host for a real domain. Common public providers
    (gmail, yahoo, outlook, …) pass through unchanged — they carry no identity.
    Every other distinct real host gets its OWN fake from the neutral pool, so
    two parties never collapse onto one domain (the earlier `_pn_rng(...).choice`
    put gmail, yahoo and the firm domain all on `letterbox.co`). Keyed on the
    registrable host, so `paula@themillenniallawyer.com` and
    `www.TheMillennialLawyer.com` land on the SAME fake."""
    host = domain.lower()
    if host.startswith("www."):
        host = host[4:]
    if host in _PN_PUBLIC_EMAIL_DOMAINS:
        return host
    reg = registry if registry is not None else _PN_DOMAIN_REGISTRY
    return reg.domain(host, _PN_EMAIL_DOMAINS)


def _pn_fake_url(real):
    """Module-level fallback URL faker (no registry)."""
    m = re.match(r"(?P<scheme>https?://)?(?P<www>www\.)?(?P<host>[^/\s]+)",
                 real.strip(), re.IGNORECASE)
    host = m.group("host") if m else real
    return f"{m.group('scheme') or ''}{m.group('www') or ''}{_pn_fake_domain(host)}"


# Regex PII detectors found in the document body (not the spreadsheet).
_PN_DETECTORS = {
    # An SSN in its 3-2-4 grouped shape with a dash, dot or SPACE separator. The
    # old dash-only rule let "123 45 6789" ride straight into a shareable export.
    # The 3-2-4 grouping is distinctive: a 3-3-4 phone or a 2-2-4 date can't
    # match. The run-together "123456789" form is deliberately NOT here (it is
    # indistinguishable from a Bates/account/reservation number) — it is caught
    # only behind an SSN label, in _PN_ID_RES below.
    "ssn": (re.compile(r"(?<!\d)\d{3}[-. ]\d{2}[-. ]\d{4}(?!\d)"), _pn_fake_ssn),
    "email": (re.compile(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"), _pn_fake_email),
    "phone": (re.compile(
        r"(?<!\d)(?:\+?1[\s.\-]?)?(?:\(\d{3}\)|\d{3})[\s.\-]?\d{3}[\s.\-]?\d{4}(?!\d)"),
        _pn_fake_phone),
    # Street address — anchored on a leading street number on the same row, then
    # 1-4 name words (capitalised, or an ordinal/directional), then a street-type
    # suffix; extended through an optional suite and City, ST ZIP. Separators are
    # spaces/tabs, never a newline (a zip at a line end must not weld to the word
    # "Street" three lines below). See _PN_ADDR_RE for the components.
    "address": (_PN_ADDR_RE, _pn_fake_street),
    # Bare URL or domain OUTSIDE an @-address (an @-address is the email
    # detector's job and wins overlap resolution). Citation hosts are filtered
    # out in _detector_cands so the authorities appendix survives.
    # Curly quotes are sentence punctuation, not URL characters: a citation
    # ending “…onelegal.com/.” swallowed the closing `.”` into the match, so
    # the same URL was tracked twice (with and without the tail) and the
    # duplicate row survived every dedup keyed on the exact value.
    "url": (re.compile(
        r"(?:https?://|www\.)[^\s<>()\"'‘’“”]*[^\s<>()\"'‘’“”.,;:!?]"
        r"|(?<!@)(?<![\w.])[A-Za-z0-9][\w-]*(?:\.[\w-]+)*"
        r"\.(?:com|net|org|law|gov|edu|us|biz|info)\b"
        r"(?:/[^\s<>()\"'‘’“”.,;:!?]*)?"),
        _pn_fake_url),
}
_PN_DEFAULT_DETECTORS = ["ssn", "email", "phone", "address", "url"]


# ── Label-anchored identifiers (auto-faked) ─────────────────────────────────
# A bar number, a contractor licence, a court-reservation ID and an attorney file
# number each resolve to a real name in a single public lookup, so none of them
# may ride along in a shareable export. Each is anchored on its own printed label
# (there is no free-floating "\d{6}" rule), which keeps the false-positive rate
# at zero on the corpus, so they are REPLACED rather than merely reported.
#
# The captured group is the identifier; the label is left as written. Once found,
# the value is registered as a whole-word term, so bare repeats elsewhere in the
# document (a caption's "Res. I.D." echoed on an exhibit) are scrubbed too — and
# `surviving_reals` can then report it if one gets through.
_PN_ID_RES = {
    "license number": re.compile(
        r"(?i)\blicen[cs]e\s*(?:no\.?|number|#)?\s*:?\s*(\d{6,})"),
    # State Bar number, in every printed form seen in the corpus — "SBN 175977",
    # "State Bar No. 207972", "S.B.# 108076". The old rule required the literal
    # word "bar", so "SBN"/"S.B.#" (no "bar") rode straight through beside the
    # fake attorney name — one public lookup then inverts the whole pseudonym map.
    "bar number": re.compile(
        r"(?i)\b(?:(?:state\s+)?bar|sbn|s\.\s*b\.?)"
        r"\s*(?:n\.?|no\.?|number|#)?\s*:?\s*(\d{5,7})\b"),
    # "RES. I.D.", "RES. NO.", "Reservation Number" — the caption block and the
    # portal page label the same value differently, and requiring "I.D." made
    # the "RES. NO." spelling invisible (it was saved only transitively, when a
    # "Reservation ID" occurrence elsewhere registered the value).
    "reservation id": re.compile(
        r"(?i)\bres(?:ervation)?\.?\s*(?:i\.?\s*d\.?|no\.?|num(?:ber)?\.?|#)"
        r"\s*:?\s*(\d{8,})"),
    "file number": re.compile(
        r"(?i)\b(?:my\s+)?file\s+no\.?\s*:?\s*([\w\-]{3,})"),
    # A run-together 9-digit SSN, caught ONLY when an SSN label anchors it, so an
    # ordinary 9-digit Bates/account number is never mistaken for one. Grouped
    # SSNs (with dashes/dots/spaces) are handled by the "ssn" detector instead.
    "ssn": re.compile(
        r"(?i)\b(?:ssn|social\s+security)"
        r"(?:\s*(?:no\.?|number|account|acct\.?|#))?\s*:?\s*(\d{9})\b"),
    # Retail-installment account identifiers printed on a RISC/contract stamp:
    # "DEAL# 23071", "CUST# 24248", "Account No. 55512".
    "account id": re.compile(
        r"(?i)\b(?:deal|cust(?:omer)?|acct|account|stock|inventory|inv)"
        r"\s*(?:no\.?|number|#)?\s*:?\s*#?\s*([A-Z]*\d[\w-]{2,})"),
    # A court-reservation CONFIRMATION code is alphanumeric ("CR-BFA76WFGYHSBGBCGZ")
    # — its digits alone (registry.digits) leave the distinctive letters intact,
    # so it is faked char-wise instead (see register_identifiers).
    "confirmation code": re.compile(
        r"(?i)\b(?:confirmation\s+(?:code|no\.?|number)\s*:?\s*|reservation\s+code\s*:?\s*)?"
        r"(CR-[A-Z0-9]{6,})"),
    # A Bates / document-production stamp: an uppercase alpha prefix welded (or
    # underscore-joined) to a zero-padded digit run, optionally a range —
    # "FMC_RAMIREZ_ERNEST_000007", "RAM000013-RAM000018". These frequently
    # EMBED a party's name, and the whole-word term patterns can't see it:
    # `_` and digits are \w, so there is no word boundary at the name inside
    # the token. Shape-anchored (no label exists in print); the space-separated
    # "SBN 175977" style stays with its own labelled class.
    # The alpha/underscore segments tolerate an interior apostrophe: a Bates
    # stamp built from a name like O'Neill welds it in as "FORD_O'WBYSOZ_91084",
    # and an apostrophe-free class split the token at the "'", matched only the
    # "WBYSOZ_91084" tail, and left the party prefix "FORD" in cleartext. Both
    # straight (') and curly (’) marks are allowed.
    "production number": re.compile(
        r"(?<![A-Za-z0-9'’])([A-Z]{2,}[A-Z0-9'’]*(?:_[A-Z0-9'’]+)*_?\d{4,}"
        r"(?:[ \t]*[-–][ \t]*[A-Z0-9_'’]*\d{2,})?)(?![A-Za-z0-9])"),
}
# Identifier classes whose value is ALPHANUMERIC and must be faked char-wise
# (letters too), not just digit-wise — a production stamp's letters ARE the
# party-name abbreviation, so they must change along with the digits.
_PN_ALNUM_IDS = {"confirmation code", "production number"}

# Identifier classes that are RE-IDENTIFICATION KEYS: each resolves to a real
# name/asset in one public lookup (State Bar search, DMV/title records, court
# reservation system), so one surviving in an export inverts the pseudonym map
# no matter how clean the names are. The reid_scan below re-checks the FINISHED
# output for these shapes as its own adversarial pass — an export is not
# certified clean until the scan comes back empty.
_PN_REID_CLASSES = frozenset({
    "bar number", "license number", "reservation id", "account id",
    "confirmation code", "ssn", "production number",
})

# A Vehicle Identification Number: 17 chars in the VIN alphabet (no I/O/Q). A
# strong unique key — decodes to make/model/year/plant and links to title and
# owner. Cued by "VIN" OR the 17-char VIN-alphabet shape with at least one digit
# (so a 17-letter word can't match). Faked char-wise in the VIN alphabet.
_PN_VIN_RE = re.compile(
    r"(?i)\bVIN\b[:\s#]*([A-HJ-NPR-Z0-9]{17})\b"
    r"|(?<![A-Z0-9])((?=[A-HJ-NPR-Z0-9]{17}(?![A-Z0-9]))(?=[A-Z0-9]*\d)[A-HJ-NPR-Z0-9]{17})")

# ISO 3779 VIN check digit (position 9): transliterate letters to values, weight
# each position, mod 11 ('X' encodes 10). Paired with the shape detector the way
# NANP ranges pair with the phone faker and SSA ranges with the SSN faker: the
# validator lets recall stay high while the FAKE stays facially valid — a fake
# VIN with a bad check digit is recognisably bogus to anyone who runs it.
_PN_VIN_VALUES = {c: v for c, v in zip("ABCDEFGH", range(1, 9))}
_PN_VIN_VALUES.update(zip("JKLMNPR", (1, 2, 3, 4, 5, 7, 9)))
_PN_VIN_VALUES.update(zip("STUVWXYZ", (2, 3, 4, 5, 6, 7, 8, 9)))
_PN_VIN_VALUES.update((str(d), d) for d in range(10))
_PN_VIN_WEIGHTS = (8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2)


def _pn_vin_check_digit(vin):
    """The ISO 3779 check character for a 17-char VIN (digit or 'X')."""
    total = sum(_PN_VIN_VALUES.get(c.upper(), 0) * w
                for c, w in zip(vin, _PN_VIN_WEIGHTS))
    r = total % 11
    return "X" if r == 10 else str(r)


def _pn_fake_vin(vin, registry):
    """A char-wise fake VIN with a CORRECT check digit, so the stand-in is
    facially valid (same principle as the valid-NANP phone and valid-SSA SSN
    fakes: an impossible fake reads as an error and invites 'correction')."""
    fake = registry.alnum(vin, "vin", "ABCDEFGHJKLMNPRSTUVWXYZ")
    if len(fake) == 17:
        fake = fake[:8] + _pn_vin_check_digit(fake) + fake[9:]
    return fake


# A Westlaw-cite run, tolerant of the shapes the full citation parser can't
# read: an optional capitalized case-name lead, "No. <docket>," then
# "<year> WL <number>", pin cite and court paren optional (the year inside the
# paren may be OCR-damaged or missing). Used ONLY to extend citation
# PROTECTION, never for linking.
#
# The docket segment allows interior whitespace: OCR routinely splits a docket
# number ("No. 19-CV- 02217-BAS- MDD") mid-token, which stopped the old
# single-run docket class before it reached "WL" and left the party name in
# "Hastings v. Ford Motor Co., No. 19-CV- 02217-BAS- MDD 2022 WL 848330"
# rewritable. Over-consuming a couple of words before "WL" is harmless here —
# this run only PROTECTS a span, it never links it.
_PN_WL_RUN_RE = re.compile(
    r"(?:[A-Z][\w.,&'’\- ]{0,60}?,?\s+)?"
    r"(?:No\.\s*[\w:().\-]+(?:[ \t]+[\w:().\-]+){0,5}?,?\s+)?"
    r"\d{4}\s+WL\s+"
    r"\d{4,}(?:,?\s*at\s+\*\d+)?(?:\s*\([^)\n]{0,50}\))?")

# ── Protection-only authority-name shapes the full parser can't read ─────────
# These extend citation PROTECTION (never linking) to two families the strict
# parsers in find_case_citations miss, both of which recur in warranty/lemon-law
# briefs and both of which weld a defendant's name INTO a published authority:
#
#   * "In re <Name> ... Litig." — a coordinated/MDL proceeding cited by name,
#     often as a bare short form with no reporter tail ("In re Ford Motor Co.
#     DPS6 Powershift Transmission Prod. Liability Litig."), and sometimes with
#     the leading signal welded on by OCR ("SeeIn re ..."). Anchored on the
#     distinctive "Litig."/"Litigation" suffix so it can't run away into prose.
_PN_INRE_LITIG_RE = re.compile(
    r"(?<![A-Za-z])(?:See\s*)?In\s*[Rr]e\s+"
    r"[A-Z][\w.'’&\-]*(?:\s+[A-Za-z0-9][\w.'’&\-]*){0,12}?\s+Litig(?:ation)?\.?"
    r"(?![A-Za-z])")

# A Title-Case run that does not open with a sentence connector, used to anchor
# the two "Cases"/"Warranty" authority shapes below.
_PN_TITLE_RUN = (
    r"(?!The\b|In\b|See\b|Cf\b|But\b|These\b|Those\b|Such\b|All\b|This\b|"
    r"That\b|Its\b|Under\b|When\b|Where\b|Here\b|A\b|An\b)"
    r"[A-Z][a-z][A-Za-z]*(?:\s+[A-Z][a-z][A-Za-z]*){0,5}")
#   * "<Name> Warranty Cases" — a consolidated-proceeding title cited by name
#     with no "v.", plus its later short forms ("Ford Motor Warranty Cases",
#     then "Ford Motor Warranty, 17 Cal.5th at 1133", "the Ford Motor Warranty
#     Court"). Anchored on the capital-C "Cases" or the "Warranty" short-name
#     immediately followed by a cite/Court signal, so lowercase "warranty" in
#     ordinary prose (which fills these very briefs) is never touched.
_PN_CASES_RUN_RE = re.compile(
    rf"(?<![A-Za-z])(?:{_PN_TITLE_RUN}\s+Warranty\s+Cases\b"
    rf"|{_PN_TITLE_RUN}\s+Warranty(?=\s+Court\b|,?\s+\d|,?\s*supra\b|\s+at\s)"
    rf"|{_PN_TITLE_RUN}\s+Cases\b(?=[\s,]+(?:\(?\d|supra\b|at\s)))")


def _pn_fake_production(val, registry):
    """Fake a production/Bates stamp with a CONSISTENT head: the alpha prefix
    maps as its own unit and the digit run as its own, so "RAM000013" and
    "RAM000018" — and both endpoints of "RAM000013-RAM000018" — come out under
    one fake prefix, the way a real production series reads."""
    def one(tok):
        m = re.match(r"^([A-Za-z][A-Za-z0-9_]*?)(\d{2,})$", tok)
        if not m:
            return registry.alnum(tok, "production_number")
        head, digits = m.groups()
        return (registry.alnum(head, "prod_prefix")
                + registry.digits(digits, "prod_digits"))
    parts = re.split(r"([ \t]*[-–][ \t]*)", val)
    return "".join(p if i % 2 or not p else one(p)
                   for i, p in enumerate(parts))


def _pn_identifier_values(text):
    """[(class, value), ...] — label-anchored identifiers found in `text`.
    A "file no." value must contain a digit, so a stray word cannot qualify."""
    text = _NFKC(text)
    out, seen = [], set()
    for cls, rx in _PN_ID_RES.items():
        for m in rx.finditer(text):
            val = m.group(1).strip()
            if not re.search(r"\d", val):
                continue
            # A bare 1-3 digit "account number" carries no distinguishing
            # entropy and collides with ubiquitous document numbers: a separate
            # statement's "Response No. 101" / "Material Fact No. 110", a page,
            # an exhibit, a statute section. Registering it faked "101"
            # everywhere it could — corrupting those references — and left it in
            # cleartext everywhere it couldn't (citation-protected spans), so
            # the real number "leaked" on page after page as pure review noise.
            # Don't track it. A 4+ digit deal/customer stamp ("DEAL# 23071") is
            # distinctive enough to keep scrubbing, and an alphanumeric id keeps
            # its letters, so both still register.
            if (cls == "account id" and not re.search(r"[A-Za-z]", val)
                    and len(re.sub(r"\D", "", val)) < 4):
                continue
            if val.lower() in seen:
                continue
            seen.add(val.lower())
            out.append((cls, val))
    return out


# ── Open-world REVIEW scan ──────────────────────────────────────────────────
# `surviving_reals` only re-checks values the tool already tracks, so it cannot
# see a real identifier that was never a term. These shape detectors flag what
# must be looked at before an export is shared. They are REVIEW findings, NOT
# auto-replaced: a statute URL is legitimate and only a human can tell the
# difference.
_PN_REVIEW_RES = {
    "url/domain": _PN_DETECTORS["url"][0],
}

# A person name that no key, declarant anchor or label anchor caught. Two shapes
# carry nearly all of them in this corpus:
#   * a capitalised bigram/trigram immediately before a phone number
#     ("Jose Gomez   (323) 283-4603" under a "CONTRACTORS:" heading);
#   * a capitalised bigram/trigram after a signature or courtesy anchor.
# Reported, never replaced: only a human can tell "Superior Court" from a name.
_PN_REVIEW_NAME = r"[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z.'’-]+){1,2}"
_PN_REVIEW_PERSON_RES = (
    # A contractor list extracts as "Jose Gomez\n(323) 283-4603" — the name and
    # the number it belongs to land on separate lines, so the gap must admit one
    # newline or the shape is never seen.
    re.compile(r"(?P<n>" + _PN_REVIEW_NAME + r")[ \t]*\n?[ \t]*"
               r"(?=\(?\d{3}\)?[\s.\-]\d{3}[\s.\-]\d{4})"),
    re.compile(r"(?i:by|signed|executed\s+by|prepared\s+by)[ \t]*:[ \t]*"
               r"(?P<n>" + _PN_REVIEW_NAME + r")"),
    re.compile(r"(?i:c\.?c\.?|cc)[ \t]*:[ \t]*(?P<n>" + _PN_REVIEW_NAME + r")"),
)
# Words that make a capitalised bigram an institution, not a person.
_PN_REVIEW_NAME_STOP = frozenset({
    "city", "court", "courthouse", "county", "state", "superior", "department",
    "dept", "office", "law", "firm", "inc", "llc", "llp", "company", "corp",
    "corporation", "avenue", "street", "boulevard", "road", "drive", "suite",
    "california", "angeles", "america", "united", "district", "division",
    "clerk", "judge", "honorable", "hon", "plaintiff", "defendant", "attorney",
    # E-filing form-field vocabulary: an LASC-generated notice prints stamps
    # like "NEW QUALIFIER" in the caption area, which is name-shaped to the
    # scanner ("New" + one distinctive word) but is a form label, not a person.
    "qualifier",
})

# Document-type / procedural vocabulary — the canonical pleading names and their
# abbreviations (mirrors the pdf-viewer footer-naming rules the same author
# maintains). A phrase built ENTIRELY from these plus role words, connectors,
# and corporate suffixes names a DOCUMENT or an argument, never a person or an
# entity, so it must never be surfaced as an unscrubbed-name finding
# ("Opposition", "Demurrer", "Plaintiff's Opposition to Mot.").
_PN_PLEADING_WORDS = frozenset({
    "motion", "motions", "mot", "mtn", "opposition", "oppositions", "opp",
    "reply", "replies", "demurrer", "demurrers", "complaint", "complaints",
    "petition", "petitions", "pet", "answer", "answers", "cross-complaint",
    "declaration", "declarations", "decl", "dec", "order", "orders", "notice",
    "notices", "memorandum", "memo", "points", "authorities", "brief", "briefs",
    "stipulation", "stipulations", "objection", "objections", "application",
    "applications", "app", "ex", "parte", "exparte", "statement", "statements",
    "separate", "umf", "aumf", "rjn", "iso", "fac", "sac", "tac", "amendment",
    "amended", "proof", "service", "dismissal", "request", "requests",
    "judgment", "judgement", "adjudication", "compel", "strike", "quash",
    "sanction", "sanctions", "reconsideration", "continuance", "arbitration",
    "arbitrate", "summary", "support", "opposition", "opposing", "response",
    "responses", "responsive", "supplemental", "amend", "trial", "hearing",
})


# Vocabulary used to re-split a token whose spaces OCR dropped ("OppositiontoMot",
# "ReplyBrief", "ExpertWitness") back into document words. Substantive document /
# role / declaration-descriptor words plus the few connectors that join them in a
# pleading title ("Notice OF Motion", "Points AND Authorities"). Kept deliberately
# small so an ordinary surname doesn't happen to segment into it. Defined lazily
# (after _PN_DECL_DESCRIPTOR exists) via _pn_proc_seg_vocab().
_PN_PROC_SEG_CONNECTORS = frozenset({"of", "to", "the", "and", "for", "in", "on"})
_PN_PROC_SEG_VOCAB = None


def _pn_proc_seg_vocab():
    global _PN_PROC_SEG_VOCAB
    if _PN_PROC_SEG_VOCAB is None:
        # "attorney"/"counsel" live in the review stop-list, not the role
        # words, so a welded "DefendantAttorneys" failed segmentation and was
        # flagged as an unscrubbed name (a KEEP `no` per spelling, forever).
        substantive = (_PN_PLEADING_WORDS | _PN_DECL_DESCRIPTOR
                       | {w for w in _PN_PARTY_ROLE_WORDS if " " not in w}
                       | {"attorney", "attorneys", "counsel", "counsels"})
        _PN_PROC_SEG_VOCAB = (substantive, substantive | _PN_PROC_SEG_CONNECTORS)
    return _PN_PROC_SEG_VOCAB


def _pn_token_is_procedural(base):
    """True when `base` (lower-cased) is a document/role/descriptor/boilerplate
    word OR an OCR-welded run of them ("oppositiontomotion", "expertwitness").
    Segmentation requires every piece to be >=2 chars and at least one to be a
    substantive document word, so a surname doesn't fragment into connectors.

    Comparison is on the LETTERS ONLY, so a stray character OCR dropped into a
    word ("Def.endant", "Opp-osition", "Mot1on") still reads as the document
    word. Accented letters are kept (str.isalpha covers them)."""
    substantive, seg_vocab = _pn_proc_seg_vocab()
    letters = "".join(c for c in base if c.isalpha()).lower()
    if (letters in _PN_PLEADING_WORDS or letters in _PN_COMMON_WORDS
            or letters in _PN_REVIEW_NAME_STOP or letters in _PN_PARTY_ROLE_WORDS
            or letters in _PN_DECL_DESCRIPTOR or _pn_is_entity_keep(base)):
        return True
    n = len(letters)
    if n < 6 or n > 40:
        return False
    # DP: can `letters` be split entirely into vocab words, using >=1 doc word?
    reach = [False] * (n + 1)
    doc = [False] * (n + 1)
    reach[0] = True
    for i in range(n):
        if not reach[i]:
            continue
        for j in range(i + 2, min(n, i + 16) + 1):
            piece = letters[i:j]
            if piece in seg_vocab:
                reach[j] = True
                if doc[i] or piece in substantive:
                    doc[j] = True
    return reach[n] and doc[n]


def _pn_is_email_value(value):
    """True when `value` is an e-mail address — a local part, an `@`, and a
    dotted host. The `@` is what settles it (a bare `courts.ca.gov` is a
    government WEBSITE and stays; `clerk@courts.ca.gov` names a person).

    An e-mail is always faked, so an e-mail row in the triage worksheet is never
    a decision the operator can make differently — it is a row to clear. Such
    values are kept OUT of `LEAKS.xlsx` (the console log and `pdf_linker.log`
    still report them, so a detector miss is not hidden), and `scrub_emails`
    cures the export so the miss does not survive either."""
    v = str(value).strip()
    if v.count("@") != 1:
        return False
    local, _, host = v.partition("@")
    return bool(local.strip()
                and re.fullmatch(r"[A-Za-z0-9.\-_]+\.[A-Za-z]{2,}", host.strip()))


def _pn_is_procedural_phrase(value):
    """True when EVERY word in `value` is a document type, a party role, a
    connector, a corporate suffix, or common/legal boilerplate — so the phrase
    names a pleading or an argument, not a person or entity. Such a value is
    never a genuine unscrubbed name and is not surfaced for review ("Opposition",
    "Plaintiff's Opposition to Mot.", "Reply Declaration"). Robust to a dropped
    space: a welded run of document words ("OppositiontoMotion") still counts."""
    words = re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-ſ][\wÀ-ÖØ-öø-ÿĀ-ſ'’.\-]*", value)
    if not words:
        return False
    for w in words:
        base = _pn_word_base(w)
        if (_pn_token_is_procedural(base) or _pn_is_suffix_token(w)
                or len(base) < 2):
            continue
        return False                     # a distinctive word remains
    return True


# A business name anchored by a PURCHASE cue ("purchased ... from Worthington
# Ford", "Worthington Ford, the dealership"): a non-party dealer has no
# corporate suffix, no caption entry, and no label anchor, so nothing else can
# see it — yet it re-identifies the sale. Reported, never replaced: a
# non-party is scrubbed only via the curated list (--term / the "Other Names"
# spreadsheet column), exactly because inferring parties is the failure mode
# the closed-entity rule exists to prevent.
_PN_REVIEW_DEALER_RES = (
    re.compile(r"(?i:purchased|bought|leased|acquired|sold)[^.\n]{0,60}?"
               r"(?i:from|at|by|to)[ \t]+(?P<n>" + _PN_REVIEW_NAME + r")"),
    re.compile(r"(?P<n>" + _PN_REVIEW_NAME + r")"
               r"(?=[ \t]*,?[ \t]*(?:(?i:a|an|the)[ \t]+)?"
               r"(?i:dealer(?:ship)?|authorized\s+repair)\b)"),
)


def _pn_dealer_review_findings(text, known_fakes=()):
    """[("possible business (purchase)", sample), ...] — capitalized names
    beside a purchase/dealer cue that no term covers. Same suppression rules
    as the person scan: institution stop-words and this run's own fakes are
    never reported."""
    known = {w.lower() for w in known_fakes}
    out, seen = [], set()
    for rx in _PN_REVIEW_DEALER_RES:
        for m in rx.finditer(text):
            name = re.sub(r"\s+", " ", m.group("n")).strip()
            words = [w.strip(".,").lower() for w in name.split()]
            if any(w in _PN_REVIEW_NAME_STOP for w in words):
                continue
            if all(w in known or w in _PN_COMMON_WORDS for w in words):
                continue
            if name.lower() in seen:
                continue
            seen.add(name.lower())
            out.append(("possible business (purchase)", name))
    return out


def _pn_word_is_own_fake(word, known):
    """True when `word` is one of this run's own stand-ins — so a REVIEW scan
    doesn't report already-faked content as an unscrubbed name. Covers the bare
    fake ("Keswick"), and a WELDED fake where a column splice glued the fake to
    its neighbour ("CORPORATIOLORNE10" carries "lorne10"; "POSTBOX4.ORGPANY"
    carries "postbox4"): the fake survives as a substring of the welded token,
    so a token containing a known fake (length >= 5, to avoid a coincidental
    short match) is treated as already-scrubbed too."""
    base = _pn_word_base(word)
    if base in known:
        return True
    return any(k in base for k in known if len(k) >= 5)


def _pn_is_name_word(w):
    """True when `w` could be part of a multi-word party name — a capitalized (or
    all-caps) alphabetic word that is not common / procedural / role / corporate-
    suffix vocab. A soft `no` keep is RELEASED where its word sits next to one of
    these (a possible party such as "Cal Equipment"), so only the standalone word
    stays kept while the phrase is still faked or flagged for review."""
    if not w or not w[:1].isalpha() or not w[:1].isupper():
        return False
    base = _pn_word_base(w)
    if len(base) < 2:
        return False
    return not (base in _PN_COMMON_WORDS or base in _PN_PLEADING_WORDS
                or base in _PN_DECL_DESCRIPTOR or base in _PN_REVIEW_NAME_STOP
                or _pn_is_role_token(w) or _pn_is_suffix_token(w))


def _pn_review_is_neutral(word, known):
    """True when `word` cannot be a fresh unscrubbed name: one of our own fakes
    (bare or welded), a common/procedural/role/connector word, a corporate
    suffix, or a short all-caps acronym prefix ("ASIC", "CBA", "PDT", "ADP")
    that a splice glued onto a name. Only such words remaining means the
    distinctive name was our own fake and the finding is noise."""
    base = _pn_word_base(word)
    return (_pn_word_is_own_fake(word, known)
            or len(base) < 3
            or base in _PN_COMMON_WORDS or base in _PN_PLEADING_WORDS
            or base in _PN_DECL_DESCRIPTOR or base in _PN_REVIEW_NAME_STOP
            or _pn_is_role_token(word) or _pn_is_entity_keep(base)
            or _pn_token_is_procedural(base)
            or (word.isupper() and len(base) <= 4))


def _pn_person_review_findings(text, known_fakes=()):
    """[("possible person name", sample), ...] — capitalised names the scrubber
    never saw. `known_fakes` holds the lower-cased words this run minted, so the
    stand-ins it just wrote in are not reported back as findings."""
    known = {w.lower() for w in known_fakes}
    out, seen = [], set()
    for rx in _PN_REVIEW_PERSON_RES:
        for m in rx.finditer(text):
            # A trailing quote/apostrophe is sentence punctuation the name
            # regex swallowed ("Sheila Bird'…"), and it forks the dedup: the
            # same person then rows once bare and once quoted.
            name = re.sub(r"\s+", " ", m.group("n")).strip().rstrip("'’\"“”")
            if _pn_is_never_fake(name):
                continue      # court-form boilerplate ("Default Only") — leave it
            words = [w.strip(".,").lower() for w in name.split()]
            if any(w in _PN_REVIEW_NAME_STOP for w in words):
                continue
            # A phrase polluted by our OWN fakes is not a fresh survivor: when a
            # fake stands beside a non-name word (a splice's "ASIC Pruett
            # Keswick", a department fragment "Nolan Relations", "Operations
            # Calder"), the strict all-fake test failed on that one word and the
            # fake got re-reported. Suppress unless a real 2-word name still
            # stands after the fakes and neutral words are removed.
            if any(_pn_word_is_own_fake(w, known) for w in name.split()):
                core = [w for w in name.split()
                        if not _pn_review_is_neutral(w, known)]
                if len(core) < 2:
                    continue
            elif all(w in known for w in words):
                continue      # a fake we minted, not a survivor
            if _pn_is_procedural_phrase(name):
                continue      # a document type, not a person ("Reply Brief")
            if name.lower() in seen:
                continue
            seen.add(name.lower())
            out.append(("possible person name", name))
    return out


# ── Common-word gazetteer ────────────────────────────────────────────────────
# Frequency-derived list of capitalised words that are ordinary English or
# legal boilerplate, NOT identity — the distinctiveness gate for the unknown-
# name scanner below. A deny-list grown one incident at a time ("M.D.", "ISO",
# "RJN") can't keep up; a gazetteer states the rule once: a word here is never,
# by itself, evidence of an unscrubbed name. Drawn from high-frequency English
# function/verb vocabulary plus the procedural vocabulary of motion practice.
_PN_COMMON_WORDS = frozenset("""
the a an and or nor but if then else when where while for to of in on at by
with from into upon under over between among through during before after
above below out off again further once here there all any both each few more
most other some such no not only own same so than too very can will just may
shall must might would could should did does do done was were been being is
are am has have had having he she it they them his her its their this that
these those which who whom whose what
oppose opposes opposed opposition motion motions moving moved move compel
compels compelling compelled strike stricken dismiss dismissal demurrer
demurrers reply replies brief briefs briefing hearing hearings trial trials
action actions complaint complaints answer answers petition petitions appeal
appeals arbitration arbitrate arbitrator arbitrators agreement agreements
contract contracts exhibit exhibits declaration declarations judgment
judgments order orders ruling rulings notice notices request requests
requests demand demands discovery deposition depositions interrogatory
interrogatories sanction sanctions damages relief injunction stay stays
statement statements facts fact introduction conclusion argument arguments
summary points authorities authority memorandum support further supplemental
seek seeks seeking sought contend contends contended contention contentions
object objects objected objection objections evidentiary
amended first second third fourth fifth sixth seventh eighth ninth tenth
eleventh twelfth thirteenth fourteenth fifteenth sixteenth seventeenth
eighteenth nineteenth twentieth cause causes proposed joint separate general
special service served serving filed filing files hereby herein hereto
whereas therefore pursuant regarding concerning including without within
entitled entitle avoid avoids avoided delegation clause clauses reserve
reserves reserved rights right cite cites cited grant grants granted denies
denied deny reject rejects rejected waive waives waived enforce enforces
enforced enforceable unconscionable binding severable applies apply applied
governs govern governed controls control controlled fails fail failed lacks
lack lacked standing merits issue issues matter matters cause causes claim
claims defense defenses
warranty warranties warrant legal name names customer customers relation
relations practice practices standard standards consumer consumers repair
repairs vehicle vehicles motor motors strategic
allege alleges alleged alleging allegation allegations liability liabilities
based assign assigns assigned assigning purpose purposes case cases supporting
supported supports our supreme unlimited limited jurisdiction jurisdictions
jurisdictional delaware venue venued
january february march april may june july august september october november
december monday tuesday wednesday thursday friday saturday sunday
establish establishes establishing established show shows showing shown
admit admits admitted complain complains complained request requests requested
approve approves approved articulate articulates articulated experience
experiences experienced resign resigns resigned motivate motivates motivated
prevent prevents prevented grow grows grew dispute disputes disputed undisputed
accommodate accommodates accommodated cannot direct indirect legitimate
intolerable rational irrational reasonable unreasonable financially voluntarily
involuntarily wrongful wrongfully intentional intentionally negligent emotional
physical mental against although age disability discrimination distress element
elements employee employees employer employers evidence failure infliction
information intent pretext reason reasons resource resources retaliation
retaliate retaliated termination terminate terminated violation violate
violated whistleblower human board directors director operation operations
decision decisions maker makers appendix accommodation retaliatory harass
harasses harassed harassment protected adverse discriminate discriminated
discriminatory hostile pretextual constructive resignation whistleblowing
deposition depo depos volume vol testified testimony declares declared
established alleged fabricated financial articulate
made make makes making day days irate feha dec decs grieve grievance
undisputed disputed genuine material fact facts issue issues each every
none no yes true false
fax telephone number numbers reservation date cure confer floor suite
owner builder builders membership arrangement unincorporated
interindemnity helpdesk
hospital hospitals medical center centers surgery surgical physician
physicians clinic clinics healthcare""".split())


def _pn_unknown_name_findings(text, neutral_words):
    """[("unscrubbed name?", sample), ...] — the HIGH-RECALL tier of a two-tier
    detection design. The term/detector tier auto-replaces only what it can
    match with high precision; this tier's one job is to surface anything
    name-shaped it cannot confidently clear, so absence from the term list
    raises a flag instead of granting a pass (the failure mode behind every
    "name we didn't know to look for" leak: Travelers, ToFF, CULTUREEDIT).

    A candidate is a capitalised run right after a party-role anchor
    ("Defendant Travelers moved…", "Attorneys for Sunrise Motors"). It is
    reported unless every word is neutral: a fake this run minted, a common/
    boilerplate word (gazetteer above), a connector, a corporate suffix, a role
    word, or a protected public entity/locality. Reported, never replaced —
    that is the reviewer's call."""
    anchor = re.compile(
        r"(?:Plaintiffs?|Defendants?|Cross-(?:Complainants?|Defendants?)|"
        r"Petitioners?|Respondents?|Appellants?|Movants?|Declarant|"
        r"Attorneys?\s+for|Counsel\s+for)"
        r"[,:]?\s+"
        r"(?P<n>[A-Z][\w&.'’-]*"
        r"(?:[ \t]+(?:(?:of|the|and|de|la|&)[ \t]+)?[A-Z][\w&.'’-]*){0,4})")
    neutral = {w.lower() for w in neutral_words}
    out, seen = [], set()
    for m in anchor.finditer(text):
        name = re.sub(r"\s+", " ", m.group("n")).strip(" .,;:'’\"“”")
        if not name or name.lower() in seen:
            continue
        if _pn_is_never_fake(name):        # court-form boilerplate — leave it
            continue
        if _pn_is_public_entity(name) or _pn_is_protected_locality(name):
            continue
        def _neutral(w):
            base = _pn_word_base(w)
            # A token that CONTAINS a known fake is a welded stand-in ("HENDRY2
            # CORPORATIOLORNE10", "POSTBOX4.ORGPANY") — already scrubbed, so
            # neutral. `neutral` here is the set of minted fake words. A welded
            # run of role/document words ("DefendantAttorneys") is boilerplate
            # that lost its space, not a name — the segmentation check reads it.
            if _pn_word_is_own_fake(w, neutral):
                return True
            return (len(base) < 3
                    or base in _PN_COMMON_WORDS or _pn_is_entity_keep(base)
                    or _pn_is_role_token(w)
                    or base in _PN_REVIEW_NAME_STOP
                    or _pn_token_is_procedural(base))
        words = name.split()
        # "Attorneys for Defendant HEALTH NET" captures "Defendant HEALTH NET"
        # — the second role word is part of the RUN, not the anchor. Reported
        # with the role word attached, the row (a) dodged every dedup against
        # the same name found elsewhere and (b) made a plain `yes` decision
        # wrong: the fix faked the word "Defendant" document-wide, which is
        # what the operator's `[Defendant]` bracket keeps were working around.
        # Trim the leading role words so the row is the name alone.
        while words and _pn_is_role_token(words[0]):
            words.pop(0)
        if not words:
            continue
        name = " ".join(words)
        if not name or name.lower() in seen:
            continue
        flags = [not _neutral(w) for w in words]
        if not any(flags):
            continue
        # A title-case HEADING reads "Defendants Are Entitled To An Order…":
        # function-word-led, with the odd distinctive word in the minority. A
        # party name leads with its distinctive token ("Travelers", "Sunrise
        # Motors Group"). Report only a name-shaped run — first word
        # distinctive, or distinctive words at least half of it.
        if not flags[0] and sum(flags) * 2 < len(flags):
            continue
        seen.add(name.lower())
        out.append(("unscrubbed name?", name))
    return out


def _pn_review_findings(text, known_fakes=()):
    """[(class, sample), ...] — identifier shapes in `text` that a human must
    review before the export is shared. De-duplicated; whitelisted citation
    hosts and the tool's own neutral fake domains are not reported."""
    text = _NFKC(text)
    out, seen = [], set()
    for cls, rx in _PN_REVIEW_RES.items():
        for m in rx.finditer(text):
            sample = re.sub(r"\s+", " ", m.group(0)).strip()
            if cls == "url/domain":
                if _pn_url_whitelisted(sample) or _pn_url_fragmentary(sample):
                    continue
                host = _pn_url_host(sample)
                # Our own fake domain — the base pool OR a numbered mint of it
                # (`postbox2.org`) that known_fakes now carries — is not a leak.
                if host in _PN_EMAIL_DOMAINS or host in known_fakes:
                    continue
                # A WELDED own fake: a column splice glued the fake host to its
                # neighbour ("www.wadsworth.com.mallory", "drvance.com" where
                # "Vance" is this run's minted surname). The host is already
                # scrubbed content, not a survivor — same substring rule as
                # _pn_word_is_own_fake, length-gated so a short coincidental
                # match can't hide a real domain.
                if any(k in host for k in known_fakes if len(k) >= 5):
                    continue
                # The host half of an E-MAIL, not a website: the url regex
                # matches "acme.com" inside "jane@acme.com", and OCR that spaced
                # the address out ("jane @ acme.com") leaves that fragment
                # standing. An e-mail is always faked, so its host is never the
                # operator's decision — reporting it is pure worksheet noise.
                if re.search(r"@[ \t]*$", text[:m.start()]):
                    continue
            key = (cls, sample.lower())
            if key not in seen:
                seen.add(key)
                out.append((cls, sample))
    out += _pn_person_review_findings(text, known_fakes)
    out += _pn_dealer_review_findings(text, known_fakes)
    out += _pn_declarant_ref_findings(text, known_fakes)
    return out


def _pn_alnum_core(real):
    """`real` reduced to lower-case alphanumerics after its trailing corporate
    suffix/connector tokens are dropped, so `Schilleci & Tortorici, P.C.` and the
    spliced `SCHILLECITORTORICI...` share the key `schillecitortorici`."""
    toks = real.split()
    while toks and _pn_is_entity_keep(_pn_word_base(toks[-1])):
        toks.pop()
    return re.sub(r"[^a-z0-9]", "", _pn_ascii_fold(" ".join(toks)).lower())


def _pn_fake_core_display(fake):
    """The fake with its trailing corporate suffix/connector tokens dropped, so
    a welded-span replacement inserts the bare stand-in ("Brightpath") where a
    real name's core was welded to a neighbour, not the full "Brightpath, LLC"
    that would drag a stray suffix into the middle of the run."""
    toks = str(fake).split()
    while toks and _pn_is_entity_keep(_pn_word_base(toks[-1])):
        toks.pop()
    return re.sub(r"[,\s]+$", "", " ".join(toks)) or str(fake)


def _pn_address_adjacency(records):
    """[warning, ...] for any two `address` records on the same street whose
    numbers are adjacent or overlapping. This is the failure that changed the
    record: `414 S. Maple` and `416 S. Maple` were faked to two unrelated
    streets, severing an ADU from the residence it sits behind."""
    by_street = {}
    for r in records:
        if r["category"] != "address":
            continue
        street, nums = _pn_addr_street_key(r["real"])
        if street and nums:
            by_street.setdefault(street, []).append(
                (min(nums), max(nums), r["real"], str(r["fake"])))
    warns = []
    for entries in by_street.values():
        entries.sort()
        for i in range(len(entries)):
            lo1, hi1, a, fa = entries[i]
            for lo2, hi2, b, fb in entries[i + 1:]:
                if lo2 <= hi1 + 2 and lo1 <= hi2 + 2:  # within 2, or overlapping
                    # Two SPELLINGS of one parcel ("Dr." vs "Drive", with or
                    # without the city tail) that resolved to the SAME fake
                    # street are the injective key working, not a conflict —
                    # nine such warnings buried the one that matters. DIFFERENT
                    # house numbers are different parcels and always warn.
                    if ((lo1, hi1) == (lo2, hi2)
                            and _pn_addr_street_key(fa) == _pn_addr_street_key(fb)):
                        continue
                    warns.append(f"'{a}' and '{b}' are adjacent/overlapping on "
                                 f"one street but were faked separately")
    return warns


def _pn_build_pattern(term, *, whole_word):
    """Literal-match regex body (NFKC-normalized, escaped), with optional
    word-boundary guards. Runs of whitespace in the term match ANY whitespace
    run in the text, so a party name broken across a line wrap ("Toyota of\\n
    Downtown") or double-spaced after a period is still caught."""
    body = r"\s+".join(re.escape(p) for p in _NFKC(term).split())
    if whole_word:
        body = rf"(?<!\w)(?:{body})(?!\w)"
    return body


class _PnTerm:
    __slots__ = ("category", "real", "fake", "pattern", "flags", "priority",
                 "source", "whole_word", "count", "loaded", "derived")

    def __init__(self, category, real, fake, *, whole_word, case_sensitive,
                 priority, source, derived=False):
        self.category = category
        self.real = real
        self.fake = fake
        self.whole_word = whole_word
        self.pattern = _pn_build_pattern(real, whole_word=whole_word)
        self.flags = 0 if case_sensitive else re.IGNORECASE
        self.priority = priority
        self.source = source
        self.count = 0
        self.loaded = False   # set on a term pre-bound from a reused key
        # A SYNTHETIC spelling this tool invented to widen matching — a
        # near-miss variant ("Esstrada"), a line-wrap spelling ("Smith- Jones").
        # It is a real value only if a document actually contained it, so it
        # earns a reversal-key row by MATCHING and never merely by existing.
        self.derived = derived


# ── Spreadsheet key (the E-Court order-template export) ──────────────────────
# Columns match the E-Court "Order Template Input" export exactly (see the
# lacourt-ecourt-suite order-template module). "Other Names" is the column that
# module appends specifically for this pseudonym pass — every attorney, law
# firm, and non-party found on the Parties page, "; "-separated — so it MUST be
# treated as a name source. (The "Case Number" header ends with a non-breaking
# space in the real export; _pn_norm_header folds it to a normal space.)
_PN_NAME_HEADERS = {
    "title plaintiff", "other plaintiffs", "title defendant", "other defendants",
    "crosscomplainants", "cross-complainants", "crossdefendants",
    "cross-defendants", "movant", "other names",
}
_PN_CASENO_HEADERS = {"case number"}
_PN_SKIP_PARTY_RE = re.compile(
    r"^(does?\b|roes?\b|all\s|et\s+al\.?$|and$|,+$)", re.IGNORECASE)


def _pn_norm_header(h):
    return h.replace("\xa0", " ").strip().lower() if isinstance(h, str) else ""


# ── Parsing one name cell ───────────────────────────────────────────────────
# "Other Names" is "; "-separated and parses cleanly. "Other Defendants" is not:
# it is the caption's party list copied out as prose —
#   (Dream Team), Glenwood Group, Mortgage 2000, Inc. (Mortgage 2000), South Bay
#   Equity, Lisa Grayson, Andrew Dinsky, WILLIAM KNOX, doing business as GLENWOOD
#   GROUP, NERSES BRONSOZIAN, erroneously sued as NICK BRONSOZIAN, and Denise
#   Padilla, as Trustee of the Sunshine Trust (collectively "Defendants")
# Splitting only on newline/semicolon/pipe turns that into ONE term, which
# matches nothing, so every defendant but the first-named one goes unscrubbed.
# Splitting naively on every comma is just as wrong: it severs "Mortgage 2000"
# from "Inc.", strands "doing business as GLENWOOD GROUP", and turns "as Trustee
# of the Sunshine Trust" into a party. So the comma is a boundary EXCEPT before
# a corporate suffix, a dba/aka marker, or a capacity phrase.
_PN_CELL_COLLECTIVE_RE = re.compile(
    r"\(\s*(?:collectively|together|jointly|hereinafter|referred\s+to)[^)]*\)",
    re.IGNORECASE)
_PN_CELL_PAREN_RE = re.compile(r"\(([^)]*)\)")
_PN_CELL_AND_RE = re.compile(r"(\s+(?:and|&)\s+)")
_PN_CELL_LEAD_AND_RE = re.compile(r"^\s*(?:and|&)\s+", re.IGNORECASE)


def _pn_ends_with_suffix(s):
    toks = s.split()
    return bool(toks) and re.sub(r"[.\s]", "", _pn_word_base(toks[-1])) in _PN_ENTITY_SUFFIXES_NORM


def _pn_cell_continues(piece):
    """True when a comma-separated piece is really the tail of the piece before
    it: a corporate suffix ("Inc."), a dba/aka marker, or a capacity phrase."""
    if not piece:
        return False
    first = _pn_word_base(piece.split()[0])
    if re.sub(r"[.\s]", "", first) in _PN_ENTITY_SUFFIXES_NORM:
        return True
    return bool(_PN_DBA_SPLIT_RE.match(piece) or _PN_AKA_SPLIT_RE.match(piece)
                or _PN_CAPACITY_RE.match(piece))


def _pn_split_cell_commas(text):
    pieces = []
    for raw in text.split(","):
        p = raw.strip()
        if not p:
            continue
        if pieces and _pn_cell_continues(p):
            pieces[-1] = pieces[-1] + ", " + p
        else:
            pieces.append(p)
    return pieces


def _pn_split_cell_and(piece):
    """Split "Mortgage 2000, Inc. and South Bay Equity" into two names, but keep
    "Smith & Jones LLP" whole: only a left side that ENDS in a corporate suffix
    can be a complete name on its own. The original separator is preserved so a
    rejoined value still matches the document text."""
    parts = _PN_CELL_AND_RE.split(piece)
    if len(parts) < 3:
        return [piece]
    out = [parts[0]]
    for sep, nxt in zip(parts[1::2], parts[2::2]):
        if _pn_ends_with_suffix(out[-1]):
            out.append(nxt)
        else:
            out[-1] += sep + nxt
    return out


def _pn_split_cell(value):
    """[(name, is_short_form), ...] for one spreadsheet cell.

    A parenthetical is the court's own short form for the party just named
    ("Mortgage 2000, Inc. (Mortgage 2000)"). It is returned separately and
    flagged, because it must become a whole-phrase term and must NOT be expanded
    into bare word tokens — "Dream Team" as a person would register "Team"."""
    if value is None:
        return []
    out = []
    for chunk in re.split(r"[\n;|]+", str(value)):
        chunk = _PN_CELL_COLLECTIVE_RE.sub(" , ", chunk)
        shorts = []

        def _grab(m):
            inner = m.group(1).strip().strip('"“”\' ')
            if len(inner.split()) >= 2 and not _pn_is_party_role(inner):
                shorts.append(inner)
            return " , "   # the parenthetical also ends the name before it

        chunk = _PN_CELL_PAREN_RE.sub(_grab, chunk)
        out.extend((s, True) for s in shorts)
        for piece in _pn_split_cell_commas(chunk):
            piece = _PN_CELL_LEAD_AND_RE.sub("", piece)
            for name in _pn_split_cell_and(piece):
                out.append((name, False))

    clean = []
    for name, is_short in out:
        p = re.sub(r"\s+", " ", name.strip().strip('"').strip())
        if len(p) < 2 or _PN_SKIP_PARTY_RE.match(p) or not re.search(r"[A-Za-z0-9]", p):
            continue
        if _pn_is_party_role(p):  # a bare role label ("Plaintiff") is not a name
            continue
        clean.append((p, is_short))
    return clean


def _pn_terms_from_xlsx(path, extra_name_headers, log):
    """Pull party names, case numbers and attorney names from the workbook.
    Returns (names, casenos) as ordered, de-duplicated lists."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("reading a --key spreadsheet needs openpyxl "
                           "(pip install openpyxl)")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    name_headers = {_pn_norm_header(h)
                    for h in _PN_NAME_HEADERS | set(extra_name_headers or [])}
    names, casenos, seen_n, seen_c = [], [], set(), set()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        cols = {i: _pn_norm_header(h) for i, h in enumerate(header)}
        for row in rows:
            for i, cell in enumerate(row):
                h = cols.get(i, "")
                if h in name_headers or "attorney" in h:
                    for v, is_short in _pn_split_cell(cell):
                        if v.lower() not in seen_n:
                            seen_n.add(v.lower()); names.append((v, is_short))
                elif h in _PN_CASENO_HEADERS or "case number" in h:
                    for v, _s in _pn_split_cell(cell):
                        if v.lower() not in seen_c:
                            seen_c.add(v.lower()); casenos.append(v)
    wb.close()
    log.info(f"  Pseudonym key: {len(names)} name(s), {len(casenos)} case number(s)")
    return names, casenos


# ── Reusing a previously written key ─────────────────────────────────────────
# The key spreadsheet this tool writes IS a complete real->fake map of a run.
# Feeding it back in (drop the forgotten PDF and pseudonym_key.xlsx in one
# folder) makes a solo re-run reproduce the batch fakes EXACTLY — a party is
# "Ernest N Ramirez" -> "Radley N Vance" here just as in the batch — while
# anything genuinely new in this file gets a fresh fake that can't collide with
# one already in the key. The written-back key carries every loaded row forward
# plus the new ones, so it never shrinks.
_PN_KEY_HEADERS = ("Category", "Real Value", "Replacement", "Status", "Source",
                   "Occurrences")
# The regex-detector categories: in the batch these were computed live at
# priority 3, so a loaded literal must sit ABOVE that to reproduce the exact
# stored fake instead of letting the detector recompute a different one.
_PN_KEY_DETECTOR_CATS = frozenset({"email", "phone", "address", "url", "ssn"})

# Term sources whose binding is AUTHORITATIVE rather than inferred: the operator
# stated this value names a party, so its row is written to the key even when it
# matched nothing in this batch. That pins the fake for the run that finally
# meets the party (the "add the document I forgot" path) instead of leaving it
# to be re-derived — or, worse, never registered at all. Every other source is a
# guess the documents themselves produced (a declarant read off a signature, a
# prefix, a court-staff name), and a guess that matched nothing stays out of the
# reversal key exactly as before.
_PN_KEY_UNMATCHED_SOURCES = frozenset({"spreadsheet", "--term"})

# Status for a key row that is FORWARD-ONLY: a synthetic spelling this tool
# invented to widen matching (a wrap-split hyphen, an OCR near-miss), pointing
# at another row's fake. Real->fake is right and must stay pinned so a re-run
# reproduces the delivered exports; fake->real is ambiguous, and `DeAnonymize`
# retires a fake claimed by two Real Values rather than guess, which left the
# pseudonym standing in the tentative. The marker tells the macro which of the
# two rows owns the reversal so exactly one does. See write_key.
_PN_KEY_ALT_STATUS = "alt spelling"
_PN_KEY_TOKEN_CATS = frozenset({"person-token", "entity-token", "short-name",
                                "address_fragment"})

# Categories whose ALPHANUMERIC core is a genuine welded PARTY NAME worth
# recovering from a column-spliced caption (the reduced-substring passes below).
# A structured identifier — a domain/email/address/phone — must NOT take part:
# its core is a substring of the very party it belongs to, so "raytheon.com"
# (core "raytheoncom") nests inside "RAYTHEON COMPANY" (reduced "raytheoncompany")
# and, being LONGER than the party core "raytheon", wins the longest-first pass
# and splices a domain fake into the party name ("POSTBOX.ORGPANY").
_PN_WELD_CORE_CATS = frozenset({
    "person", "entity", "person-token", "entity-token", "short-name",
    "display-name",
})

# How long an alphanumeric core must be before the reduced-substring passes will
# hunt for it. The bar is high because a reduced core carries no word boundary,
# so a short one coincides inside ordinary words ("law" in "lawsuit").
_PN_WELD_CORE_MIN = 8
# …which is longer than most FIRST and LAST names — "michael", "carroll",
# "amezcua", "maria", "juan" are all 4-7 — so the gate skipped precisely the
# values the pass exists to recover. A spliced caption welds the party's bare
# name tokens to their neighbours ("MICHAELCARROLL", "AMEZCUApain") and the
# whole-word patterns cannot reach them either, so the defendant's name rode
# into the export while `surviving_reals` reported the file clean.
#
# A SHORT core is therefore allowed, but only where a coincidence is
# implausible and the value is not a guess:
#   * a PERSON name (an entity's words are generic — "law", "firm", "brothers"
#     — and nest inside real words; a person token does not),
#   * named by the operator's OWN party list (`_trusted_party_tokens`, which
#     already drops connectors, role words and corporate suffixes, so "De" of
#     "Cruz De Amezcua" never qualifies),
#   * not an ordinary English word, and shaped like a name, and
#   * actually WELDED to a neighbour in the source (`_pn_span_is_welded`) —
#     a clean standalone occurrence is the boundary-anchored pass's business,
#     and if it left one it did so deliberately (a keep, a citation).
_PN_WELD_SHORT_CORE_MIN = 4
_PN_WELD_SHORT_CORE_CATS = frozenset({"person", "person-token"})


def _pn_span_is_welded(src, start, end):
    """True when `src[start:end]` runs straight into an alphanumeric neighbour.

    That is the whole reason the reduced pass exists: a welded span has no word
    boundary, so the ordinary term patterns cannot match it. A span with clean
    boundaries on both sides was reachable and was left standing on purpose."""
    before = src[start - 1] if start > 0 else ""
    after = src[end] if end < len(src) else ""
    return bool((before and before.isalnum()) or (after and after.isalnum()))

# A KEEP loses to an exact party match: a kept word that falls inside one of
# these FULL-name term matches is released so the real party is still faked
# ("Cal" kept alone, "CAL EQUIPMENT FE RANCH, LLC" faked whole). Deliberately
# NOT the bare *-token / short-name / display-name categories — those are how a
# standalone kept word would get faked, and the keep must win over them.
_PN_PARTY_OVERRIDE_CATS = frozenset({"person", "entity", "case_number"})


def _pn_key_looks_like_ours(path):
    """True when `path` is a key THIS tool wrote (header row = _PN_KEY_HEADERS),
    as opposed to an E-Court "Order Template Input" name list."""
    try:
        import openpyxl
    except ImportError:
        return False
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        header = next(ws.iter_rows(values_only=True), ())
        wb.close()
    except Exception:
        return False
    return tuple((h or "").strip() if isinstance(h, str) else h
                 for h in (header or ())[:6]) == _PN_KEY_HEADERS


def _pn_load_key(path, registry, log):
    """Load a key THIS tool wrote as authoritative real->fake bindings, seeding
    `registry` so new values can't collide and re-harvested names recompose
    identically. Returns `(terms, key_decisions)` — the list of pre-bound
    _PnTerm objects, plus operator KEEP/KEEP-PART decisions typed into the
    Replacement column (see below).

    Each row becomes a literal term carrying its stored fake, so the value
    reproduces verbatim however it was originally composed (a full e-mail, a
    house number + street). The registry's used-pool, name-token memo and
    domain memo are seeded from the rows so a genuinely new value drawn later
    avoids every fake already handed out and reuses a party's established
    token/domain.

    The Replacement column also accepts the same operator instructions the LEAKS
    worksheet's Fix? column does, so a mistake baked into the key can be fixed in
    place: `no` (leave this Real Value verbatim — do not fake it) and a
    `[bracketed]` keep-spec (keep the bracketed part verbatim, auto-fake the
    rest). Such a row builds no faking term; the decision is returned in
    `key_decisions` (value_lower -> dict shaped like a leak decision) for the
    caller to apply and persist."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [(_pn_norm_header(h)) for h in next(rows, ())]
    idx = {name: header.index(name) for name in
           ("category", "real value", "replacement", "status", "source",
            "occurrences")
           if name in header}
    terms, seen, key_decisions = [], set(), {}
    for row in rows:
        def cell(name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None
        cat = (cell("category") or "").strip()
        real = cell("real value")
        fake = cell("replacement")
        if not cat or real in (None, "") or fake in (None, ""):
            continue
        real, fake = str(real), str(fake)
        if _pn_is_never_fake(real):
            continue                      # court-form boilerplate — drop any
                                          # stale binding so it is never re-faked
        if (cat, real.lower()) in seen:
            continue
        seen.add((cat, real.lower()))

        # Operator control words typed into the Replacement column (same
        # vocabulary the LEAKS Fix? column accepts). Build no faking term; hand
        # the decision back for the caller to apply (keep-protection / fragment
        # faking) and persist to the worksheet + master log.
        ctrl = fake.strip()
        if ctrl.lower() in ("no", "n"):
            key_decisions[real.lower()] = {
                "value": real, "type": "KEEP", "fix": "no", "replacement": None,
                "fake_values": None, "fixcell": None, "notes": "pseudonym key"}
            continue
        if "[" in ctrl:
            frags = _pn_bracket_keep(real, ctrl)
            if frags == []:               # whole value bracketed -> keep all
                key_decisions[real.lower()] = {
                    "value": real, "type": "KEEP", "fix": "no",
                    "replacement": None, "fake_values": None,
                    "fixcell": ctrl, "notes": "pseudonym key"}
                continue
            if frags:                     # keep bracketed part(s), fake the rest
                key_decisions[real.lower()] = {
                    "value": real, "type": "KEEP-PART", "fix": "yes",
                    "replacement": None, "fake_values": frags,
                    "fixcell": ctrl, "notes": "pseudonym key"}
                continue
            # frags is None: bracketed text isn't a substring of the value — fall
            # through and treat the cell as an ordinary explicit replacement.
        # Self-heal a key written before the self-map guard existed: a row whose
        # replacement equals its real value ("M & M" -> "M & M") never scrubs and
        # loops --fix-leaks. Re-mint a distinct fake; the next write_key persists
        # the correction. Seed the registry (below) from the CORRECTED fake.
        if _pn_norm_map(fake) == _pn_norm_map(real):
            remint = _pn_fake_initials_name(real, registry)
            if _pn_norm_map(remint) != _pn_norm_map(real):
                log.info(f"  Pseudonymize: repaired self-mapped key row {real!r} "
                         f"-> {remint!r}")
                fake = remint
        source = (cell("source") or "spreadsheet").strip() or "spreadsheet"
        try:
            occ = int(cell("occurrences") or 0)
        except (TypeError, ValueError):
            occ = 0

        # Seed the registry so re-derivations and new values stay consistent.
        for w in [fake] + fake.split():
            registry._used.add(w.lower())
        if cat in ("person-token", "entity-token"):
            registry._memo.setdefault(("name_or_entity", real.lower()), fake)
        if cat in ("email", "url"):
            rh = real.split("@", 1)[-1] if cat == "email" else _pn_url_host(real)
            fh = fake.split("@", 1)[-1] if cat == "email" else _pn_url_host(fake)
            rh, fh = rh.lower().lstrip("www."), fh.lower()
            if rh and fh and "." in rh:
                registry._memo.setdefault(("domain", rh), fh)
                registry._domain_reals.setdefault(rh, fh)
                registry._used.add(fh.lower())
        # A case number and a street are re-derived through their OWN registry
        # slots, which the row above does not fill — so a value written a second
        # way in a LATER-ADDED document ("1888 Century Park E." beside the
        # first run's "…Park East", a case number picked up from its label)
        # composed a brand-new fake and the same parcel/case shipped under two
        # stand-ins. Seed those slots too, keyed exactly as the fakers key them
        # (the street on its number-stripped identity), so every spelling folds
        # onto the fake this key already carries.
        elif cat == "case_number":
            registry._memo.setdefault(("caseno", real.lower()), fake)
        elif cat == "address":
            core = _pn_addr_street_key(real)[0]
            fake_core = _pn_addr_parts(fake)[0]
            if core and fake_core:
                registry._memo.setdefault(("street", core.lower()), fake_core)
                registry._used.add(fake_core.lower())

        # A person-token row DERIVED from the judge's (or a staff member's)
        # full name exists for the reversal macro only. Loading it as a match
        # term would fake the judge's BARE surname everywhere on a key-reuse
        # run — which the first run deliberately does not do (bare surname only
        # behind a judicial title). The registry memo above still carries the
        # binding, so re-derived fakes stay consistent; just no term.
        if cat == "person-token" and source in ("judge", "court-staff"):
            continue

        if cat in _PN_KEY_DETECTOR_CATS:
            priority, whole, csens = 4, True, False
        elif cat in _PN_KEY_TOKEN_CATS:
            priority = 1
            whole = True
            # An all-caps short acronym ("SLP") stays case-sensitive so a
            # lowercase word is left alone; other short forms match either case.
            csens = (cat == "short-name" and real.isalpha()
                     and real == real.upper() and len(real) <= 5)
        elif cat in ("person", "entity"):
            priority, whole, csens = 2, False, False
        else:                       # case_number, identifiers, display-name
            priority, whole, csens = 2, True, False
        t = _PnTerm(cat, real, fake, whole_word=whole, case_sensitive=csens,
                    priority=priority, source=source)
        t.count = occ
        t.loaded = True
        # Round-trip the `alt spelling` marker. Without it the re-run's
        # write_key sees two ordinary rows on one fake and re-picks the owner by
        # hit count — which the wrap-split spelling usually wins, since it is
        # the form a line break actually leaves in the text. Ownership would
        # then flip to the synthetic spelling on the first re-run, undoing the
        # whole point of the marker.
        t.derived = (str(cell("status") or "").strip().lower()
                     == _PN_KEY_ALT_STATUS)
        terms.append(t)
    wb.close()
    terms.sort(key=lambda t: (-t.priority, -len(t.real)))
    log.info(f"  Pseudonym key REUSED: loaded {len(terms)} binding(s) from "
             f"{path.name} — fakes will match the original run")
    if key_decisions:
        kept = sum(1 for d in key_decisions.values() if d["fix"] == "no")
        part = len(key_decisions) - kept
        log.info(f"  Pseudonym key: {kept} value(s) marked 'no' (keep verbatim)"
                 + (f" and {part} bracketed keep-spec(s)" if part else "")
                 + " in the Replacement column will be honored.")
    return terms, key_decisions


class _PnTokenOrderRecorder(_PnFakeRegistry):
    """A throwaway registry that ENUMERATES the bare name/entity tokens a term
    build will mint, and the pool each is drawn from — the fakes it hands out are
    discarded. `_pn_build_terms` runs one first, then pre-binds the real registry
    SHORTEST-FIRST, so a welded token ("ADLERMICHAEL") always sees its base
    ("ADLER") already bound and folds onto it whatever order the document
    presented them in. The set of tokens minted does not depend on order, so this
    dry run reliably names them even though its own fold order is arbitrary."""

    def __init__(self):
        super().__init__()
        self.pool_of = {}    # real_lower -> (pool, seed_tag) of its first mint

    def token(self, real, words, seed_tag):
        if (self._memo_tag(seed_tag), real.lower()) not in self._memo:
            self.pool_of.setdefault(real.lower(), (words, seed_tag))
        return super().token(real, words, seed_tag)


def _pn_build_terms(names, casenos, extra_terms, registry=None, _prewarm=True):
    """Turn raw strings into _PnTerm objects with stable fake replacements.
    A shared `registry` keeps every fake unique across the whole case; pass the
    SAME registry to the Pseudonymizer so declarant names added later can't
    reuse a fake already assigned here.

    `_prewarm` makes OCR/typo folding independent of the order names appear in.
    An edit-distance fold is symmetric — whichever near-variant is minted first
    becomes the canonical one and the rest fold onto it — but a WELD is strictly
    longer than its base and can only fold once the base is bound. So a scratch
    pass enumerates the bare tokens and they are pre-bound shortest-first, which
    guarantees every base precedes any token that contains or extends it."""
    if registry is None:
        registry = _PnFakeRegistry()
    if _prewarm:
        rec = _PnTokenOrderRecorder()
        _pn_build_terms(names, casenos, extra_terms, rec, _prewarm=False)
        for tok in sorted(rec.pool_of, key=lambda s: (len(s), s)):
            pool, tag = rec.pool_of[tok]
            registry.token(tok, pool, tag)      # bases before welds/extensions
    terms = []
    for raw in names:
        raw, is_short = raw if isinstance(raw, tuple) else (raw, False)
        if _pn_is_never_fake(raw):
            continue                      # court-form boilerplate — never faked
        if is_short:
            _pn_append_shortname_term(terms, raw, "spreadsheet", registry)
        else:
            _pn_append_name_terms(terms, raw, "spreadsheet", registry)
    for raw in casenos:
        if _pn_is_never_fake(raw):        # a form number is not a case number
            continue
        terms.append(_PnTerm("case_number", raw, _pn_fake_caseno(raw, registry),
                             whole_word=True, case_sensitive=False, priority=2,
                             source="spreadsheet"))
    for raw in extra_terms or []:
        if _pn_is_never_fake(raw):
            continue
        if re.search(r"\d", raw) and not re.search(r"[A-Za-z]{2}", raw):
            terms.append(_PnTerm("case_number", raw, _pn_fake_caseno(raw, registry),
                                 whole_word=False, case_sensitive=False,
                                 priority=2, source="--term"))
        else:
            _pn_append_name_terms(terms, raw, "--term", registry)
    # De-duplicate on (real, category); keep the highest-priority instance.
    dedup = {}
    for t in terms:
        k = (t.real.lower(), t.category)
        if k not in dedup or t.priority > dedup[k].priority:
            dedup[k] = t
    return sorted(dedup.values(), key=lambda t: (-t.priority, -len(t.real)))


def _pn_find_folder_key(folder, log):
    """Locate a key spreadsheet sitting in the CASE FOLDER itself, so a re-run
    resolves its own inputs with no Downloads copy and no typed --key. Order of
    preference:
      1. this tool's own `pseudonym_key.xlsx` — REUSE its bindings so the re-run
         reproduces the original fakes; then
      2. the E-Court `Order*.xlsx` input template (e.g. "Order_Template_Input") —
         mint FRESH fakes from it, which is how a run STARTS OVER after the
         pseudonym key was deleted.
    Returns a Path or None. Excel lock files (~$...) are ignored."""
    own = folder / "pseudonym_key.xlsx"
    if own.is_file():
        log.info("  Pseudonymize: reusing this folder's pseudonym_key.xlsx")
        return own
    orders = [p for p in folder.glob("*.xlsx")
              if not p.name.startswith("~$")
              and p.name.lower().startswith("order")]
    if orders:
        newest = max(orders, key=lambda p: p.stat().st_mtime)
        log.info(f"  Pseudonymize: starting over from this folder's E-Court "
                 f"template: {newest.name}")
        return newest
    return None


def _pn_find_party_template(folder, log):
    """The E-Court `Order*.xlsx` party template for this case, or None.

    Distinct from `_pn_find_folder_key`, which PREFERS this tool's own
    pseudonym_key.xlsx: this looks past the key for the original party LIST, so
    a key-reuse run can still see a party the key never bound. Case folder
    first (where the workflow puts it), then Downloads (where the E-Court
    export lands) — except for an ALL-WORD folder, which has no template of its
    own, so the Downloads guess would supplement this case's key with a
    STRANGER'S parties (see `_is_word_only_folder`)."""
    orders = [p for p in folder.glob("*.xlsx")
              if not p.name.startswith("~$")
              and p.name.lower().startswith("order")]
    if orders:
        return max(orders, key=lambda p: p.stat().st_mtime)
    if _is_word_only_folder(folder):
        return None
    for base in (Path(os.environ.get("USERPROFILE") or Path.home()) / "Downloads",
                 Path.home() / "Downloads"):
        if not base.is_dir():
            continue
        orders = [p for p in base.glob("*.xlsx")
                  if not p.name.startswith("~$")
                  and p.name.lower().startswith("order")]
        if orders:
            return max(orders, key=lambda p: p.stat().st_mtime)
        break
    return None


def _pn_supplement_key_terms(terms, folder, name_column, registry, log):
    """Terms for a listed party the REUSED key never bound, drawn through the
    same `registry` so every existing fake is untouched.

    `write_key` deliberately omits a term that matched nothing (the key is the
    reversal artifact: a Real Value that was never faked must never be in it),
    and a key-reuse run never re-reads the party template — so a party who is
    named ONLY in a document that was missing from the first batch had no
    binding and no term, and shipped in the clear when that document was added.
    That is the ordinary "re-run with the document I forgot" path, and it is
    the one case where the key alone is not the whole truth about the case.

    Re-derivation is safe because the registry is memo-seeded from the key:
    a token the key already bound resolves to its stored fake, so a supplement
    can only ADD a fake, never move one. Terms already loaded are dropped by
    (category, real), so nothing is registered twice."""
    template = _pn_find_party_template(folder, log)
    if template is None:
        return []
    try:
        names, casenos = _pn_terms_from_xlsx(template, name_column, log)
    except RuntimeError as e:
        log.warning(f"  Pseudonymize: could not re-read the party template "
                    f"{template.name}: {e}. A party named only in a newly "
                    f"added document may not be scrubbed.")
        return []
    have = {(t.category, str(t.real).lower()) for t in terms}
    fresh = [t for t in _pn_build_terms(names, casenos, [], registry)
             if (t.category, str(t.real).lower()) not in have]
    if fresh:
        added = sorted({t.real for t in fresh if t.category in
                        ("person", "entity", "case_number")})
        log.info(f"  Pseudonym key: {len(fresh)} binding(s) added from "
                 f"{template.name} for listed value(s) the key had not bound"
                 + (f" ({', '.join(added[:6])})" if added else "") + ".")
    return fresh


def _pn_find_downloads_key(log):
    """Locate the key spreadsheet automatically: the most recently modified
    .xlsx in the user's Downloads folder. That's where the E-Court order-
    template export lands, so a plain --pseudonymize run picks it up with no
    path to type. Returns a Path or None (with a logged reason)."""
    downloads = None
    for cand in (Path(os.environ.get("USERPROFILE") or Path.home()) / "Downloads",
                 Path.home() / "Downloads"):
        if cand.is_dir():
            downloads = cand
            break
    if downloads is None:
        log.warning("  Pseudonymize: no Downloads folder found; "
                    "pass --key to name the spreadsheet explicitly")
        return None
    # Ignore Excel lock files (~$...) and any key we previously wrote.
    sheets = [p for p in downloads.glob("*.xlsx")
              if not p.name.startswith("~$") and p.name != "pseudonym_key.xlsx"]
    if not sheets:
        log.warning(f"  Pseudonymize: no .xlsx in {downloads}; "
                    "pass --key to name the spreadsheet explicitly")
        return None
    # Prefer the E-Court export ("Order_Template_Input*.xlsx" — the same
    # "Order*.xlsx" name the Word mail-merge macro looks for); fall back to the
    # newest .xlsx of any name if none is present.
    order = [p for p in sheets if p.name.lower().startswith("order")]
    newest = max(order or sheets, key=lambda p: p.stat().st_mtime)
    if order:
        log.info(f"  Pseudonymize: using E-Court export: {newest.name}")
    else:
        log.info(f"  Pseudonymize: no Order*.xlsx found; using most recent "
                 f"Downloads spreadsheet: {newest.name}")
    return newest


# ── Declarant names ("DECLARATION OF X" / "I, X, declare") ───────────────────
# A declaration's own title and signature name the declarant, who is often NOT
# in the spreadsheet key (a client, a witness, an attorney the E-Court export
# didn't capture). These patterns pull that name out so it can be scrubbed like
# any other party name. A stop-list rejects the many "DECLARATION OF <process>"
# phrases (service, mailing, due diligence, custodian of records, counsel, …)
# that aren't personal names.
_PN_DECL_STOPWORDS = {
    "service", "mailing", "posting", "publication", "due", "diligence",
    "non-service", "custodian", "records", "record", "merits", "probable",
    "readiness", "completion", "compliance", "good", "non-monetary", "monetary",
    "status", "counsel", "moving", "responding", "demurring", "party", "parties",
    "no", "non", "related", "non-receipt", "receipt", "prejudice", "support",
    "defendant", "defendants", "plaintiff", "plaintiffs", "movant", "the",
    "undersigned", "reasonable", "attempted", "electronic", "personal",
}
# Includes accented Latin letters (see _PN_LAT) so "Declaration of Teresa C.
# Alarcón" captures the whole surname instead of truncating to "Alarc" at the
# "ó" — the truncated stem broke key round-trip and produced the welded fake
# "Isleyón". The leading letter may itself be accented ("Óscar", "Ángela").
_PN_DECL_NAME_WORD = rf"[{_PN_LAT_UPPER}][{_PN_LAT}.'’-]*"
_PN_DECL_NAME = (
    rf"{_PN_DECL_NAME_WORD}(?:\s+{_PN_DECL_NAME_WORD}){{0,3}}"
    r"(?:,\s*(?i:esq|jr|sr|ii|iii|iv|m\.?d|ph\.?d|c\.?p\.?a|r\.?n|d\.?d\.?s)\.?)?"
)
_PN_DECL_TITLE_RE = re.compile(
    r"(?i:(?:supplemental|amended|further|second|third|fourth|reply|corrected|"
    r"joint)\s+)?(?i:declaration\s+of)\s+(" + _PN_DECL_NAME + r")")
_PN_DECL_SELF_RE = re.compile(
    r"\bI,\s+(" + _PN_DECL_NAME + r")\s*,\s*(?i:declare|hereby|do\s+hereby|"
    r"state|being\s+duly\s+sworn)")

# A declaration is cited by its declarant's name — "Smith Decl.", "Alarcón
# Declaration, ¶ 4", "Yu Dec." — which is the single highest-value place a
# witness or party name leaks (see the Alarcón / Yu misses). This review scan
# ALWAYS surfaces the name in front of a Declaration reference so a human checks
# it, unless it is a fake this run minted or procedural boilerplate ("Reply
# Declaration" names a document). Up to three capitalised words are captured so
# "Gregory Yu Decl." keeps both name tokens.
#
# The separator before the reference word is [ \t]* (not +), so a dropped space
# still resolves: OCR that welds "Smith Decl." into "SmithDecl." leaves the name
# rewritable, and the greedy name group backtracks to the reference word.
#
# The reference word is captured LOOSELY — a D-word that may carry OCR stray
# marks — and validated letters-only against _PN_DECL_REF_WORDS, so "Dec.laration"
# / "Decl.aration" still read as "Declaration". It must start with a capital D
# and end at a non-letter (?![A-Za-z]), so it can't bite into a name that merely
# begins that way ("Declan"); a bare "Dec"/"Decl"/"Declaration" that letters out
# to something else ("December", "Decker", "Donald") is dropped by the validator.
_PN_DECL_REF_WORDS = frozenset({"dec", "decl", "declaration"})
_PN_DECL_REF_RE = re.compile(
    r"(?<![A-Za-z])(?P<n>" + _PN_DECL_NAME_WORD
    + r"(?:[ \t]+" + _PN_DECL_NAME_WORD + r"){0,2}?)"
    + r"[ \t]*(?P<ref>D[A-Za-z.'’\-]*)(?![A-Za-z])")


# Descriptor words that precede "Declaration" without naming the declarant —
# "Supporting Declaration", "Expert Witness Declaration", "Custodian of Records
# Declaration". A name made ENTIRELY of these (plus boilerplate) is not a person
# and is neither flagged nor pseudonymized.
_PN_DECL_DESCRIPTOR = frozenset({
    "expert", "witness", "custodian", "records", "supporting", "responding",
    "opposing", "rebuttal", "percipient", "sworn", "verified", "authenticating",
    "foundational", "initial", "final", "corrected", "additional", "further",
    "supplemental", "amended", "joint", "declarant", "moving", "responsive",
})


def _pn_declarant_ref_names(text):
    """Clean declarant names cited before a "Declaration"/"Decl."/"Dec."
    reference in `text`. Leading signal/role/boilerplate words are trimmed
    ("See Smith Decl." -> "Smith"), and a value that is entirely procedural
    ("Reply Declaration") or descriptive ("Supporting Declaration") is dropped.
    The Declaration word itself is NEVER part of a returned name."""
    out, seen = [], set()
    nfkc = _NFKC(text)
    for m in _PN_DECL_REF_RE.finditer(nfkc):
        # Validate the reference word letters-only ("Dec.laration" -> declaration).
        refletters = "".join(c for c in m.group("ref") if c.isalpha()).lower()
        if refletters not in _PN_DECL_REF_WORDS:
            continue
        # A bare "Dec." followed by a number is a date ("Dec. 5, 2024"), not a
        # declaration short cite.
        if refletters == "dec" and re.match(r"[ \t]*\d", nfkc[m.end():]):
            continue
        toks = m.group("n").split()
        # Drop leading words that aren't the declarant ("See"/"The"/"Plaintiff").
        while toks and (_pn_word_base(toks[0]) in _PN_COMMON_WORDS
                        or _pn_word_base(toks[0]) in _PN_PLEADING_WORDS
                        or _pn_word_base(toks[0]) in SIGNAL_PREFIXES
                        or _pn_word_base(toks[0]) in _PN_DECL_DESCRIPTOR
                        or _pn_is_role_token(toks[0])):
            toks.pop(0)
        name = " ".join(toks).strip(" .,;:")
        low = name.lower()
        if not name or low in seen:
            continue
        if _pn_is_procedural_phrase(name):
            continue
        if all(_pn_word_base(w) in _PN_DECL_DESCRIPTOR for w in toks):
            continue
        seen.add(low)
        out.append(name)
    return out


def _pn_declarant_ref_findings(text, known_fakes=()):
    """[("unscrubbed declarant name?", "Alarcón"), ...] — a backstop review for
    a declarant name that reached the OUTPUT un-pseudonymized (e.g. on a spliced
    page the registrar couldn't scrub). A name already replaced with one of this
    run's fakes is not reported."""
    known = {w.lower() for w in known_fakes}
    out = []
    for name in _pn_declarant_ref_names(text):
        toks = name.split()
        if all(_pn_word_is_own_fake(w, known) for w in toks):
            continue                       # already scrubbed to our fake
        # A fake dragged along by a non-name prefix ("ASIC Pruett Keswick",
        # "CBA. Radley", "Keswick's January Radley") is already scrubbed too:
        # once the fakes are removed, what remains is not a declarant name.
        if any(_pn_word_is_own_fake(w, known) for w in toks):
            rest = [w for w in toks if not _pn_word_is_own_fake(w, known)]
            if not _pn_is_personlike_declarant(" ".join(rest)):
                continue
        out.append(("unscrubbed declarant name?", name))
    return out


def _pn_is_personlike_declarant(name):
    core = re.sub(r",.*$", "", name).strip()
    words = core.split()
    if not words or words[0].lower().strip(".") in _PN_DECL_STOPWORDS:
        return False
    # Two-plus name words, or a single surname with a professional suffix.
    return len(words) >= 2 or bool(
        re.search(r",\s*(?i:esq|jr|sr|m\.?d|ph\.?d|c\.?p\.?a)", name))


# Words that end a declarant's name and begin the description of what the
# declaration is about. Without this, "DECLARATION OF TANOOA SPARKS REGARDING
# COMPLIANCE" registers the person "TANOOA SPARKS REGARDING COMPLIANCE" and the
# bare tokens "REGARDING" and "COMPLIANCE" — which then replace those two words
# wherever they appear in the brief.
_PN_DECL_TRAIL_STOP = frozenset({
    "as", "in", "re", "regarding", "concerning", "for", "on", "of", "to",
    "and", "with", "support", "opposition", "response", "reply", "compliance",
    "successor", "interest", "pursuant", "filed", "submitted",
    # "Declaration of X ISO Motion" / "... RJN" — the title, not the name.
    "iso", "rjn", "motion", "notice",
    # Caption field labels: a declarant name that wrapped a line must not run on
    # into "License #:" / "State Bar No." / "Reservation ID:" on the next line.
    "license", "licence", "bar", "state", "sbn", "no", "reservation", "res",
    "id", "case", "dept", "department", "telephone", "tel", "fax", "facsimile",
    "email", "e-mail", "attorney", "attorneys",
})


def _pn_trim_declarant(name):
    """Cut a captured declarant name at the first word that starts describing
    the declaration rather than naming the declarant."""
    out = []
    for w in name.split():
        if w.strip('.,:;').lower() in _PN_DECL_TRAIL_STOP:
            break
        out.append(w)
    return " ".join(out).strip().rstrip(",")


def _pn_declarant_names(text):
    """Raw declarant names found in `text` via the title/self-ID patterns,
    de-duplicated and filtered to person-like values."""
    names, seen = [], set()
    for rx in (_PN_DECL_TITLE_RE, _PN_DECL_SELF_RE):
        for m in rx.finditer(text):
            # Collapse any line wrap in the captured name to single spaces.
            nm = re.sub(r"\s+", " ", m.group(1)).strip().rstrip(".").strip()
            nm = _pn_trim_declarant(nm)
            if _pn_is_personlike_declarant(nm) and nm.lower() not in seen:
                seen.add(nm.lower())
                names.append(nm)
    return names


# ── Label-anchored names (exhibits, contracts, signature blocks) ─────────────
# Pleading paper yields names through DECLARATION OF / I, X, declare / Law
# Offices of. EXHIBITS — letterheads, contracts, proofs of service — carry real
# names with none of those anchors, so every one of them used to survive. These
# labels state that a name follows, which is a reliable, low-risk anchor.
_PN_LABEL_NAME = r"[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z0-9.'’-]+){0,3}"
# The gap between a label and its value. A contract's two-column layout extracts
# as "Property owner:\nRoxanne and Thomas Purscelley", so a horizontal-only gap
# matched nothing and the exhibit's owner names were never registered. At most
# ONE newline, or the label would reach across a blank line into unrelated prose.
_PN_LABEL_GAP = r"[ \t]*\n?[ \t]*"
_PN_LABEL_RES = (
    re.compile(r"(?i:property\s+owners?)[ \t]*:?" + _PN_LABEL_GAP +
               r"(?P<n>" + _PN_LABEL_NAME + r"(?:[ \t]+(?:and|&)[ \t]+"
               + _PN_LABEL_NAME + r")?)"),
    # NB: the plural `s` lives INSIDE the case-insensitive group. Left outside it
    # matched "Contractors:" but not the "CONTRACTORS:" an exhibit actually prints.
    re.compile(r"(?i:(?:contractor|owner|client|tenant|landlord|buyer|seller)s?)"
               r"[ \t]*:" + _PN_LABEL_GAP + r"(?P<n>" + _PN_LABEL_NAME + r")"),
    re.compile(r"/s/[ \t]*(?P<n>" + _PN_LABEL_NAME + r")"),
    re.compile(r"(?i:attn|attention)[ \t]*:?" + _PN_LABEL_GAP +
               r"(?P<n>" + _PN_LABEL_NAME + r")"),
    re.compile(r"(?i:dear)[ \t]+(?i:mr|ms|mrs|dr|messrs)\.?[ \t]+"
               r"(?P<n>" + _PN_LABEL_NAME + r")"),
    # An attorney's name immediately BEFORE their bar-number label — the one
    # place a letterhead prints counsel with no other anchor ("Michael D.
    # Mortenson, State Bar No. 230831"). A run faked the bar number while the
    # name beside it shipped in cleartext; the trailing label makes this a
    # zero-false-positive anchor.
    re.compile(r"(?P<n>" + _PN_LABEL_NAME + r")\s*,?\s*(?:Esq\.?\s*,?\s*)?"
               r"\(?(?i:State\s+Bar\s+No|SBN|S\.\s*B\.?\s*(?:N|#))"),
)


# A "Name  (312) 555-1212" contact line: a capitalized personal name followed
# ON THE SAME LINE by a phone number. Contract letterheads and service lists
# print a roster this way ("CONTRACTORS:\nJose Gomez (323)…\nJuan Olivas (562)…"),
# and the label rule above caught only the FIRST entry, so `Juan Olivas` and
# `Viviana Gomez` leaked. The trailing phone number makes this a low-false-
# positive anchor — prose and address lines don't carry one.
_PN_ROSTER_LINE_RE = re.compile(
    r"(?m)^[ \t]*(?P<n>[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z0-9.'’-]+){1,2})"
    r"[ \t]+\(?\d{3}\)?[ \t.\-]?[ \t]*\d{3}[ \t.\-]\d{4}\b")


def _pn_label_names(text):
    """Personal names read out of exhibit/contract/signature labels
    ("Property owner:", "/s/", "Attn:", "Dear Mr. X", "Contractor:") and out of
    roster lines that pair a name with a phone number. A single "Roxanne and
    Thomas Purscelley" yields both, sharing the trailing surname."""
    out, seen = [], set()
    for rx in (*_PN_LABEL_RES, _PN_ROSTER_LINE_RE):
        for m in rx.finditer(text):
            raw = re.sub(r"\s+", " ", m.group("n")).strip()
            pieces = [p.strip() for p in re.split(r"\s+(?:and|&)\s+", raw) if p.strip()]
            # "Roxanne and Thomas Purscelley" -> give the bare first name the
            # shared surname so it registers as a person, not a lone first name.
            if pieces and len(pieces[-1].split()) >= 2:
                surname = pieces[-1].split()[-1]
                pieces = [p if len(p.split()) >= 2 else f"{p} {surname}"
                          for p in pieces]
            for piece in pieces:
                piece = _pn_trim_declarant(piece).strip()
                words = piece.split()
                if len(words) < 2 or _pn_is_party_role(piece):
                    continue
                # Never harvest a protected locality as a person — a roster line
                # "Los Angeles (213) 555-1212" must not turn the county into a
                # fake name and scrub it out of "County of Los Angeles".
                if _pn_is_protected_locality(piece):
                    continue
                if any(_pn_word_base(w) in _PN_NON_NAME_WORDS or _pn_is_role_token(w)
                       for w in words):
                    continue
                if not all(w[:1].isupper() for w in words if w[:1].isalpha()):
                    continue
                if piece.lower() not in seen:
                    seen.add(piece.lower())
                    out.append(piece)
    return out


# ── Court personnel: presiding judge, court staff, department number ─────────
# A judicial title — the cue that lets a BARE surname be faked ("Judge Whitaker"
# -> "Judge <fake>"). "hon/honorable/justice/commissioner/referee" are never
# verbs so they match case-insensitively; "Judge" must be capitalized so a
# lowercase verb ("the jury must judge Smith's credibility") is not a title. An
# optional middle initial ("Judge D. Whitaker") is allowed between title and
# surname.
_PN_JUDICIAL_TITLE = (
    r"(?:(?i:(?:the\s+)?hon(?:orable)?|justice|commissioner|referee)\.?"
    r"|(?:The\s+)?(?:Judge|JUDGE))")
_PN_TITLE_LEAD = _PN_JUDICIAL_TITLE + r"[ \t]+(?:[A-Z]\.[ \t]+)?"

# Court staff, anchored on a role label ("Judicial Assistant: Jane Doe",
# "Courtroom Clerk: ...", "Court Reporter: ...").
_PN_COURT_STAFF_RE = re.compile(
    r"(?i:judicial\s+assistant|courtroom\s+assistant|court(?:room)?\s+clerk|"
    r"deputy\s+clerk|clerk\s+of\s+(?:the\s+)?court|court\s+reporter|bailiff|"
    r"court\s+attendant|research\s+attorney|law\s+clerk)"
    r"[ \t]*:?" + _PN_LABEL_GAP + r"(?P<n>" + _PN_LABEL_NAME + r")")

# The same staff roles with the NAME FIRST — how a clerk actually signs a
# court-generated notice ("By: N. Lachikian, Deputy Clerk"). The label-first
# rule above cannot see this shape, so the clerk's name was never registered
# as court personnel; it rode into the review sheet as a "possible person
# name" on every notice of appeal instead of being scrubbed.
_PN_COURT_STAFF_NAME_FIRST_RE = re.compile(
    r"(?P<n>" + _PN_LABEL_NAME + r")[ \t]*,[ \t]*"
    r"(?i:judicial\s+assistant|courtroom\s+assistant|court(?:room)?\s+clerk|"
    r"deputy\s+clerk|court\s+reporter|bailiff|court\s+attendant|"
    r"research\s+attorney|law\s+clerk)(?![\w])")

# A department / courtroom number ("Department 515", "Dept. 515", "Dept 72",
# "Department No. 515"). Only the LABELED number is faked, keeping the label
# word — a bare "515" elsewhere (a page, a dollar amount) is left alone.
_PN_DEPT_RE = re.compile(
    r"\b(?:dept|department)\.?[ \t]*(?:no\.?|number|#)?[ \t]*:?[ \t]*"
    r"(\d{1,3}[A-Z]?)(?!\w)", re.IGNORECASE)

# The judge is DISCOVERED from the document via a title ("Hon. Dana Whitaker",
# "Judge Whitaker") — no name is hard-coded. A title + 2-3 words is a full name
# (faked wherever it appears together); a title + one word is a bare surname
# (faked only behind a title). Up to three name words are captured, then trimmed
# of trailing non-name words.
_PN_JUDGE_CAPTURE_RE = re.compile(
    _PN_JUDICIAL_TITLE + r"[ \t]+(?:[A-Z]\.[ \t]+)?"
    r"(?P<n>[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z.'’-]+){0,2})")
# Words that follow a title but do not name the judge ("Presiding", "of the
# Superior Court", "pro tem"), so the capture is CUT at the first one. Includes
# the ruling verbs that follow a judge's name in an ALL-CAPS heading ("JUDGE
# SMITH ORDERED THE PARTIES…") — without them the verb was swallowed into the
# "full name" ("SMITH ORDERED") and that phrase was then faked in ordinary
# prose ("Smith ordered lunch").
_PN_JUDGE_STOP = frozenset({
    "presiding", "court", "department", "dept", "division", "superior",
    "honorable", "pro", "tem", "sitting", "assigned", "designate", "designated",
    "judge", "justice", "of", "the", "for", "all", "purposes",
    "ordered", "orders", "overruled", "overrules", "sustained", "sustains",
    "ruled", "rules", "held", "holds", "found", "finds", "granted", "grants",
    "denied", "denies", "issued", "issues", "entered", "enters", "signed",
    "stated", "states", "noted", "notes", "hereby", "presided", "presides",
    "ret", "retired", "wrote", "writes", "declared", "declares", "concluded",
    "concludes", "determined", "determines", "adopted", "adopts", "vacated",
    "continued", "set", "scheduled",
})


# A display name in front of an e-mail address: "Tommy Purscelley
# <tpurscelley@x.com>". The address is scrubbed by the e-mail detector, but the
# name beside it used to survive, re-identifying the party the address hid.
_PN_EMAIL_DISPLAY_RE = re.compile(
    r"(?P<name>[A-Z][A-Za-z.'’-]+(?:[ \t]+[A-Z][A-Za-z.'’-]+){0,3})[ \t]*"
    r"<[ \t]*[^<>@\s]+@[^<>\s]+[ \t]*>")


def _pn_reflow(original, fake):
    """`fake`, laid back into the line structure `original` occupied.

    A party name is routinely broken across a line ("DREAM TEAM REAL ESTATE\\n
    CONSULTANTS, INC."). Matching across the break is only safe if the
    replacement puts the break back, or the .txt loses a numbered line and every
    pinpoint cite after it shifts. Entity and person fakes are built word by
    word, so they carry the same token count as the real name and the ORIGINAL
    whitespace runs can simply be re-used between the fake's words. When the
    counts differ (a phone number, an e-mail), the newlines are appended after
    the fake instead — the line count survives even though the wrap moves."""
    if "\n" not in original:
        return fake
    seps = re.findall(r"\s+", original)
    otoks, ftoks = original.split(), fake.split()
    if ftoks and len(otoks) == len(ftoks) == len(seps) + 1:
        return "".join(t + s for t, s in zip(ftoks, seps)) + ftoks[-1]
    return fake + "\n" * original.count("\n")


# ── "dba" pairs found in the document text ──────────────────────────────────
# The E-Court export's "Title Defendant" is frequently just the legal name
# ("Dream Team Real Estate Consultants Inc."), while the pleadings call the
# party by its fictitious name ("Equity Union") on nearly every page. Neither
# the spreadsheet nor a declarant scan yields that name, so it is read straight
# off the text, in either word order:
#     <legal name> dba <fictitious name>
#     <fictitious name>, a dba of <legal name>
# Both sides are then registered as terms, which also picks up a legal name the
# export missed entirely ("Mortgage 2000, Inc.").
_PN_DBA_CONNECT = r"of|the|and|de|la|&"
_PN_DBA_WORD = r"[A-Z0-9][\w&.'’-]*"
# The LEGAL name may wrap: a long caption name is routinely split across lines.
_PN_DBA_PHRASE = (rf"{_PN_DBA_WORD}"
                  rf"(?:[\s,]+(?:{_PN_DBA_CONNECT}|{_PN_DBA_WORD})){{0,6}}")
# The FICTITIOUS name may not. Two-column pages extract with the opposing
# signature block interleaved between the lines of the caption, so a phrase that
# jumps a line break after "dba" swallows an attorney's name and address
# ("William Knox, dba\nPeter Pruett, Esq. Glenwood Group 633 ...").
_PN_DBA_PHRASE_1L = (rf"{_PN_DBA_WORD}"
                     rf"(?:[ \t,]+(?:{_PN_DBA_CONNECT}|{_PN_DBA_WORD})){{0,5}}")
_PN_DBA_TEXT_RE = re.compile(
    rf"(?P<head>{_PN_DBA_PHRASE})[\s,]*\(?{_PN_DBA_MARK}\)?[ \t,]*"
    rf"(?P<tail>{_PN_DBA_PHRASE_1L})")
_PN_DBA_OF_RE = re.compile(
    rf"(?P<tail>{_PN_DBA_PHRASE_1L})\s*,\s*an?\s+{_PN_DBA_MARK}\s+of[ \t]+"
    rf"(?P<head>{_PN_DBA_PHRASE_1L})")
# A phrase that is only these words names nothing.
_PN_DBA_PHRASE_STOP = {"the", "a", "an", "this", "that", "its", "his", "her",
                       "their", "and", "or", "of", "is", "was", "court", "no"}
# Never inside a business name; their presence means the phrase ran off the end
# of the name and into a signature block.
_PN_DBA_PHRASE_REJECT = re.compile(
    r"(?<![A-Za-z])(?i:esq|jr|sr|ph\.?d|m\.?d|attorneys?|telephone|facsimile)"
    r"(?![A-Za-z])")
# A 3+ digit number that is FOLLOWED by more words is a street number, not part
# of the name ("Glenwood Group 633 W. Fifth St."). A trailing one is part of it
# ("Mortgage 2000"), as is one before a corporate suffix ("Mortgage 2000, Inc.").
_PN_DBA_ADDRESS_RE = re.compile(r"[\s,]+\d{3,}\s+\S")


def _pn_trim_address(s):
    m = _PN_DBA_ADDRESS_RE.search(s)
    if not m:
        return s
    rest = s[m.start():].split()
    if len(rest) > 1 and _pn_is_entity_keep(_pn_word_base(rest[1])):
        return s
    return s[:m.start()]


def _pn_clean_phrase(s):
    """Collapse wrapped whitespace and shave what a greedy phrase match drags
    along: punctuation, dangling connector words, and the leading role label a
    pleading puts in front of every party ("Defendants Dream Team Real Estate
    Consultants Inc.")."""
    s = re.sub(r"\s+", " ", s).strip().strip(" ,;:.\"'()")
    while True:
        m = re.match(r"(\S+)\s+(?=\S)", s)
        if not m or not _pn_is_role_token(m.group(1)):
            break
        s = s[m.end():]
    while True:
        m = re.search(rf"[\s,]+(?:{_PN_DBA_CONNECT})$", s, re.IGNORECASE)
        if not m:
            return s.strip(" ,;")
        s = s[:m.start()]


def _pn_dba_pairs(text):
    """[(legal_name, fictitious_name), ...] read off `text`, de-duplicated."""
    pairs, seen = [], set()
    for rx in (_PN_DBA_TEXT_RE, _PN_DBA_OF_RE):
        for m in rx.finditer(text):
            head = _pn_trim_address(_pn_clean_phrase(m.group("head")))
            tail = _pn_trim_address(_pn_clean_phrase(m.group("tail")))
            # Both sides must be at least two words. A one-word name read out of
            # prose ("dba Equity") would become a substring term that fires
            # inside every longer name and every ordinary use of the word; a
            # single-word dba has to be supplied with --term instead.
            if len(head.split()) < 2 or len(tail.split()) < 2:
                continue
            if head.lower() in _PN_DBA_PHRASE_STOP or tail.lower() in _PN_DBA_PHRASE_STOP:
                continue
            if _PN_DBA_PHRASE_REJECT.search(head) or _PN_DBA_PHRASE_REJECT.search(tail):
                continue
            if _pn_is_party_role(head) or _pn_is_party_role(tail):
                continue
            key = (head.lower(), tail.lower())
            if key not in seen:
                seen.add(key)
                pairs.append((head, tail))
    return pairs


# ── Law firm names ("LAW OFFICES OF <name>") ────────────────────────────────
# The E-Court export's attorney columns routinely omit the firm, so the firm's
# name — which contains a real person's name — survives on every signature block
# and proof of service. The label is a reliable anchor; the name after it is
# captured up to the first comma, so a trailing "P.C." / "LLP" is left off and
# the name takes the person path (registering the bare surname too).
_PN_FIRM_RE = re.compile(
    r"(?i:law\s+(?:offices?|firm|group)\s+of)[ \t]+"
    r"(?P<name>[A-Z][\w.'’-]*(?:[ \t]+[A-Z][\w.'’-]*){1,3})")
# A firm named by its professional-corporation suffix rather than by a "Law
# Offices of" label: "Schilleci & Tortorici, P.C.". Only the partner whose name
# also appeared as an attorney of record was ever scrubbed, so the letterhead
# came out half-real — "SCHILLECI & HALLORAN, P.C." — which fingerprints the firm
# more precisely than leaving it alone would have. The suffix is captured with
# the name so `_pn_looks_like_entity` routes it down the entity path.
_PN_FIRM_WORD = r"[A-Z][\w.'’-]*"
_PN_FIRM_SUFFIX_RE = re.compile(
    r"\b(?P<name>" + _PN_FIRM_WORD +
    r"(?:[ \t]+(?:&|and)?[ \t]*" + _PN_FIRM_WORD + r"){0,3}"
    r"[ \t]*,[ \t]*(?i:A[ \t]*P[ \t]*L[ \t]*C|P\.?\s?C\.?|L\.?L\.?P\.?|"
    r"L\.?L\.?C\.?|P\.?L\.?L\.?C\.?))(?![\w])")
_PN_FIRM_REJECT = frozenset({"the", "los", "california", "record", "counsel"})


def _pn_firm_names(text):
    """Entity names read out of "Law Offices of X" labels and out of a trailing
    professional-corporation suffix ("Schilleci & Tortorici, P.C.")."""
    out, seen = [], set()
    for m in _PN_FIRM_RE.finditer(text):
        name = _pn_clean_phrase(m.group("name"))
        words = name.split()
        if len(words) < 2:
            continue
        if any(_pn_word_base(w) in _PN_FIRM_REJECT or _pn_is_role_token(w)
               or _pn_word_base(w) in _PN_NON_NAME_WORDS for w in words):
            continue
        if name.lower() not in seen:
            seen.add(name.lower())
            out.append(name)
    for m in _PN_FIRM_SUFFIX_RE.finditer(text):
        name = re.sub(r"\s+", " ", m.group("name")).strip()
        head = name.split(",")[0]
        words = [w for w in head.split() if w not in ("&", "and")]
        if not words:
            continue
        if any(_pn_word_base(w) in _PN_FIRM_REJECT or _pn_is_role_token(w)
               or _pn_word_base(w) in _PN_NON_NAME_WORDS for w in words):
            continue
        if name.lower() not in seen:
            seen.add(name.lower())
            out.append(name)
    return out


class Pseudonymizer:

    """Applies stable pseudonymization to plain text (the .txt export only).

    Terms (party/case/attorney names from a spreadsheet key and/or --term)
    plus regex PII detectors are matched against each page's text; the highest-
    priority, longest non-overlapping matches win, and each is swapped for its
    deterministic fake. Occurrence counts accumulate across the whole folder so
    one key file can be written at the end.
    """

    def __init__(self, terms, detectors, registry=None):
        self.terms = terms
        self.detectors = detectors
        # Shared with _pn_build_terms so declarant fakes minted later stay unique
        # against the terms already assigned (no two reals share one fake).
        self.registry = registry if registry is not None else _PnFakeRegistry()
        self.texts_applied = 0   # how many text bodies were run through apply()
        self.written = []        # .txt paths written with pseudonymization applied
        self.leaked = set()      # real_lower values that survived in some export
        self.leaked_by_file = {} # written txt path -> {real_lower} that survived
                                 # IN THAT FILE, so the gate quarantines only the
                                 # files carrying a leak, not the whole batch
        self.review = []         # (class, sample) open-world findings across files
        self.leak_report = []    # {file, type, value, where} rows for the
                                 # human-triage worksheet (page:line located)
        self.suppressed = set()  # value_lowers the reviewer marked "no fix" in
                                 # the worksheet — never quarantine on these
        # KEEP protection comes in two tiers (from the LEAKS worksheet, the
        # pseudonym key OR the cross-folder master KEEP sheet):
        self.keep_soft = set()   # a `no` value — keep the EXACT word only where
                                 # it stands alone; released inside a multi-word
                                 # party name run (a possible party like "Cal
                                 # Equipment"), so the phrase is still faked.
        # A bracketed keep-spec part from a decision THIS folder made. Such a
        # keep beats even a full party match, because the operator's bracket
        # already says how to split the name: keep this fragment, fake the
        # rest — and the rest IS separately registered as a term, so the party
        # is still scrubbed ("Alder Law, P.C." -> "Aldous Law, P.C."). An
        # INHERITED keep-spec contributes no such fragment terms, so it must
        # keep losing to the party match or the whole name would ride through.
        self.keep_strict_local = set()
        self.keep_strict = set() # a bracketed keep-spec part — "this fragment is
                                 # never a name": kept verbatim even next to
                                 # names ("[Plaintiff]" stays in "Plaintiff X").
                                 # Both still yield to a full party-name match.
        self.kept_hits = set()   # kept strings (lowercased) that ACTUALLY matched
                                 # text this run — so the master KEEP log can bump
                                 # only the decisions a run really used
        self._own_fakes = set()  # detector fakes we minted, so a re-scrub of
                                 # already-fake text never re-fakes them
        # (category, real_lower) -> record dict {category, real, fake, source,
        # count, pattern, flags}
        # Compiled-regex cache keyed by (pattern, flags). Python's own re cache
        # holds only 512 entries, so a case with >512 terms recompiled every
        # term's pattern on every page (regex COMPILATION was ~75% of the
        # pseudonymize+scan runtime). This cache is unbounded, so each pattern
        # compiles once for the whole run.
        self._rx_cache = {}
        self.records = {}
        for t in terms:
            # Never install a self-identical mapping (fake == real): it is a
            # no-op scrub that ships the real value and loops --fix-leaks.
            fake = _pn_guard_distinct_fake(t.real, t.fake, self.registry)
            if fake is None:
                continue
            t.fake = fake
            self.records[(t.category, t.real.lower())] = {
                "category": t.category, "real": t.real, "fake": fake,
                "source": t.source, "count": t.count, "pattern": t.pattern,
                "flags": t.flags, "loaded": getattr(t, "loaded", False),
                "derived": getattr(t, "derived", False)}
        # Real values pre-bound from a reused key: authoritative, so the
        # citation-only prune must never drop one (a party that happens to
        # appear only in a citation in THIS file is still a known party).
        self._loaded_reals = {t.real.lower() for t in terms
                              if getattr(t, "loaded", False)}

    def has_spreadsheet_terms(self):
        return any(t.source == "spreadsheet" for t in self.terms)

    def spreadsheet_hits(self):
        """Replacements made from PRIMARY spreadsheet-key values — full party
        names, entities, case numbers — excluding bare surname/given-name
        tokens. Zero means the key didn't correspond to the documents: a bare
        token can match unrelated prose incidentally, so counting only primary
        values keeps that from masking a wrong/stale sheet."""
        return sum(r["count"] for r in self.records.values()
                   if r["source"] == "spreadsheet" and r["category"] != "person-token")

    def a_case_number_fake(self):
        """A fake case number (any), for use as a neutral filename prefix."""
        for r in self.records.values():
            if r["category"] == "case_number":
                return re.sub(r"[^A-Za-z0-9]", "", r["fake"])
        return None

    def _add_terms(self, new_terms):
        """Add freshly minted terms, skipping any real value already tracked
        and any value pruned as citation-only (a per-file re-harvest must not
        resurrect it). Returns True when anything was added."""
        added = False
        for t in new_terms:
            k = (t.category, t.real.lower())
            if k in self.records:
                continue
            if t.real.lower() in getattr(self, "_pruned_reals", ()):
                continue
            # Never install a self-identical mapping (fake == real): a no-op
            # scrub that ships the real value and loops --fix-leaks.
            fake = _pn_guard_distinct_fake(t.real, t.fake, self.registry)
            if fake is None:
                continue
            t.fake = fake
            self.terms.append(t)
            self.records[k] = {
                "category": t.category, "real": t.real, "fake": fake,
                "source": t.source, "count": 0, "pattern": t.pattern,
                "flags": t.flags, "loaded": getattr(t, "loaded", False),
                "derived": getattr(t, "derived", False)}
            added = True
        if added:
            # Keep terms ordered biggest/most-specific first for apply()'s
            # overlap resolution.
            self.terms.sort(key=lambda t: (-t.priority, -len(t.real)))
            self._trusted_tok_cache = None   # recompute over the new term set
        return added

    def register_short_names(self, text):
        """Register a parenthetical short form the document itself defines:
        `Defendant Glenwood Group ("Glenwood")` or the initialism
        `Tom of Finland Foundation ("ToFF")`.

        A short form qualifies two ways: (a) every one of its words appears in
        the parent name (`Glenwood` of `Glenwood Group`), or (b) it is a single
        token that spells the parent's initials in order (`ToFF`). Both keep
        this from becoming a licence to register arbitrary words — a bare
        capitalised word that is neither a parent word nor an initialism (e.g.
        `("Team")` from `Dream Team Real Estate`) is rejected. The definition may
        carry a short descriptor between the name and the parenthetical
        (`…, INC., A PUBLIC BENEFIT CORPORATION ("TOFF" OR …)`) and more than one
        quoted alternative; every alternative is considered. The fake is built
        from the parent's own fake so the forms stay one party. Idempotent; call
        before apply()."""
        # Optional descriptor run ("…, A PUBLIC BENEFIT CORPORATION") then a
        # parenthetical whose whole body is captured so each quoted alternative
        # can be pulled out.
        paren = (r"[\s,;]*(?:[A-Za-z0-9][A-Za-z0-9.,&\-' ]{0,70}?)?[\s,;]*"
                 r"\(\s*(?P<body>[^()]{0,160})\)")
        for t in list(self.terms):
            if t.category not in ("entity", "person"):
                continue
            real_words = t.real.split()
            fake_words = str(t.fake).split()
            parent_words = {_pn_word_base(w) for w in real_words}
            try:
                rx = self._compiled(t.pattern + paren, t.flags)
            except re.error:
                continue
            for m in rx.finditer(text):
                for short in _pn_paren_short_forms(m.group("body")):
                    if _pn_is_party_role(short):
                        continue
                    words = short.split()
                    # A single-word short form that is a COMMON word ("Eighth" of
                    # "Eighth & Broadway Investments") must never become a term:
                    # case-insensitive, it faked every "eighth cause of action".
                    # Businesses use common words far more than people, so a bare
                    # common word is dropped outright rather than registered.
                    if len(words) == 1 and _pn_is_generic_token(_pn_word_base(short)):
                        continue
                    is_parent_word = bool(words) and all(
                        _pn_word_base(w) in parent_words for w in words)
                    if is_parent_word:
                        fake = (_pn_fake_entity(short, self.registry)
                                if t.category == "entity"
                                else _pn_fake_person(short, self.registry)[0])
                    elif len(words) == 1:
                        fake = _pn_initialism_fake(short, real_words, fake_words)
                        if fake is None:
                            continue
                    else:
                        continue
                    # A single-word BUSINESS short form that is a WORD of the
                    # parent name requires capitalization (case-sensitive):
                    # businesses borrow ordinary words ("Broadway", "Summit"), so
                    # match only the capitalised form, never a lower-case prose
                    # occurrence. An INITIALISM ("ToFF", written TOFF/Toff) stays
                    # case-insensitive, and a person's bare surname stays
                    # case-insensitive (OCR can lower-case it).
                    require_cap = (t.category == "entity" and len(words) == 1
                                   and is_parent_word)
                    self._add_terms([_PnTerm("short-name", short, fake,
                                             whole_word=True,
                                             case_sensitive=require_cap,
                                             priority=1, source="document")])

    def prune_citation_only_terms(self, text):
        """Drop every DOCUMENT-harvested name term whose occurrences in `text`
        all lie inside protected citation spans — the corpus-wide equivalent of
        "don't harvest a party out of the authorities table".

        A lemon-law brief cites "FCA US, LLC", "Mercedes-Benz USA, LLC", "BMW
        of N. Am., LLC" dozens of times; the suffix-anchored harvester swept
        those (and welded fragments like "N. Am., LLC") in as entities. Each
        was then correctly PRESERVED inside its citations by span protection —
        and the leak scan counted every preserved occurrence as a leak,
        quarantining ten clean exports. A name that never appears outside a
        citation is an authority, not a party: it needs no fake, no key row,
        and no leak status. Pruned values are remembered so a later per-file
        re-harvest cannot resurrect them. Call after ALL register_* passes,
        with the FULL corpus text (a name citation-only in one filing may be a
        real party in another)."""
        text = _NFKC(text)
        spans = self._protected_citation_spans(text)
        if not spans:
            return []
        self._pruned_reals = getattr(self, "_pruned_reals", set())
        doomed = []
        loaded = getattr(self, "_loaded_reals", ())
        for t in list(self.terms):
            if (t.source != "document"
                    or t.real.lower() in loaded
                    or t.category not in ("person", "entity", "person-token",
                                          "entity-token", "short-name")):
                continue
            try:
                ms = list(self._compiled(t.pattern, t.flags).finditer(text))
            except re.error:
                continue
            if ms and all(any(m.start() < e and s < m.end() for s, e in spans)
                          for m in ms):
                doomed.append(t)
        for t in doomed:
            self.terms.remove(t)
            self.records.pop((t.category, t.real.lower()), None)
            self._pruned_reals.add(t.real.lower())
        if doomed:
            self._trusted_tok_cache = None
        return [t.real for t in doomed]

    def register_entity_acronyms(self, text):
        """Register the bare initialism of every tracked multi-word entity that
        the document ACTUALLY USES — briefs adopt a firm's acronym ("SLP" for
        Strategic Legal Practices) without ever writing the defining
        parenthetical that register_short_names needs, and the bare initialism
        then re-identifies the firm all over the prose.

        Occurrence-gated (the acronym must appear standalone in `text`) and
        case-sensitive (a 2-5 letter uppercase token; prose words never
        match), guarded against document abbreviations (ISO/RJN), common
        words, and role words. The fake is the initials of the entity's OWN
        fake, so the long and short forms stay recognisably one firm.
        Idempotent; call before apply()."""
        def initials(name):
            words = _pn_acronym_trim(name.replace(",", " ").split())
            core = [w for w in words
                    if _pn_word_base(w) not in _PN_ACRONYM_CONNECTIVES
                    and len(_pn_word_base(w)) >= 2]
            return "".join(w[0].upper() for w in core) if len(core) >= 2 else ""
        new = []
        for t in list(self.terms):
            if t.category != "entity":
                continue
            acr = initials(t.real)
            if not 2 <= len(acr) <= 5:
                continue
            low = acr.lower()
            if (low in _PN_DOC_ABBREV or low in _PN_COMMON_WORDS
                    or low in _PN_PARTY_ROLE_WORDS
                    or ("short-name", low) in self.records):
                continue
            if not re.search(rf"(?<!\w){re.escape(acr)}(?!\w)", text):
                continue
            fake = initials(str(t.fake))
            if not fake or fake == acr:
                fake = self.registry.alnum(acr, "acronym")
            new.append(_PnTerm("short-name", acr, fake, whole_word=True,
                               case_sensitive=True, priority=1,
                               source="document"))
        self._add_terms(new)

    def register_firm_names(self, text):
        """Register the name inside a "Law Offices of X" label. The attorney
        columns of the E-Court export frequently omit the firm, leaving a real
        name on every signature block and proof of service. Idempotent; call
        before apply()."""
        for raw in _pn_firm_names(text):
            new = []
            _pn_append_name_terms(new, raw, "document", self.registry)
            self._add_terms(new)

    def register_label_names(self, text):
        """Register names read out of exhibit/contract/signature labels
        ("Property owner:", "/s/", "Attn:", "Dear Mr. X", "Contractor:") — the
        one place real names appear off the spreadsheet AND off the pleading
        anchors. Idempotent; call before apply()."""
        for raw in _pn_label_names(text):
            new = []
            _pn_append_name_terms(new, raw, "document", self.registry)
            self._add_terms(new)

    def register_dba_names(self, text):
        """Read "<entity> dba <name>" / "<name>, a dba of <entity>" out of `text`
        and register BOTH sides as terms. The spreadsheet's Title Defendant is
        often just the legal name, so without this the fictitious name the
        pleadings actually use is never scrubbed. Shared words keep one fake
        apiece (see `_pn_append_name_terms`). Idempotent; call before apply()."""
        for head, tail in _pn_dba_pairs(text):
            new = []
            _pn_append_name_terms(new, f"{head} dba {tail}", "document",
                                  self.registry)
            self._add_terms(new)

    def register_localities(self, text):
        """No-op under the current policy. City, state, and ZIP are kept
        verbatim — only the house number and street name inside an address are
        faked — so no standalone locality term is registered and a bare
        "Montebello" or "90640" three lines below an address is left as written.
        Retained for call-site stability (the pipeline still calls it)."""
        return

    def register_addresses(self, text):
        """For every full street address in `text`, register its house-number +
        street-name FRAGMENT (suffix, suite and locality stripped) as a term
        mapped to the SAME fake street.

        A corrupted-scan text layer can wrap an address so its street-type
        suffix lands on the next line ("...business address 1055 E Colorado" |
        "Blvd. ..."); the address detector needs the suffix on the line, so the
        bare "1055 E Colorado" is not matched and the real street leaks. Learning
        the fragment from a clean occurrence of the same address (e.g. the
        letterhead) scrubs the suffix-less one to the same fake. Whole-word and
        low priority, so the full-address detector wins wherever the address is
        intact. Idempotent; call before apply()."""
        _DIRS = {"n", "s", "e", "w", "ne", "nw", "se", "sw"}
        new = []
        for m in _PN_ADDR_RE.finditer(_NFKC(text)):
            street, _suite, _cz = _pn_addr_parts(m.group(0))
            frag = _pn_addr_strip_suffix(street)
            toks = frag.split()
            # house number + a distinctive name word (never a bare "1 A St").
            names = [t for t in toks[1:] if t.strip(".,").lower() not in _DIRS]
            if (len(toks) < 2 or not re.match(r"^\d", frag)
                    or not any(len(t.strip(".,")) >= 4 for t in names)):
                continue
            if ("address_fragment", frag.lower()) in self.records:
                continue
            fake = _pn_addr_strip_suffix(self._fake_street_core(m.group(0)))
            if fake:
                new.append(_PnTerm("address_fragment", frag, fake,
                                   whole_word=True, case_sensitive=False,
                                   priority=1, source="document"))
        self._add_terms(new)

    def register_identifiers(self, text):
        """Detect label-anchored identifiers in `text` (bar number, contractor
        licence, court reservation ID, attorney file number) and register each
        as a whole-word term with a digit-for-digit fake.

        Each of these resolves to a real name in one public lookup, so they are
        replaced, not merely reported. Registering the VALUE (rather than
        rewriting only the labelled occurrence) also scrubs the bare repeats —
        a caption's "Res. I.D." echoed on an exhibit stamp — and lets
        `surviving_reals` report one that gets through. Idempotent; call before
        apply()."""
        new = []
        for cls, val in _pn_identifier_values(text):
            cat = cls.replace(" ", "_")
            if (cat, val.lower()) in self.records:
                continue
            # SSN shares the detector's faker (valid shape, digit-seeded) so a
            # labelled run-together SSN and a dashed one elsewhere map to the
            # same fake; an alphanumeric id (confirmation code) is faked char-wise
            # so its distinctive letters change too; the rest keep the digit faker.
            if cls == "ssn":
                fake = self._fake_ssn(val)
            elif cls == "production number":
                fake = _pn_fake_production(val, self.registry)
            elif cls in _PN_ALNUM_IDS:
                fake = self.registry.alnum(val, cat)
            else:
                fake = self.registry.digits(val, cat)
            new.append(_PnTerm(cat, val, fake,
                               whole_word=True, case_sensitive=False,
                               priority=2, source="document"))
        # A tracked party name INSIDE an alphanumeric token that carries a
        # digit run ("FMC_RAMIREZ_ERNEST_000007", "Ramirez000013"): the word-
        # boundary lookarounds can't see the name (`_` and digits are \w), so
        # the enclosing token is faked whole, char-wise. Requires a digit run
        # of 3+ (a document stamp, never prose) and a distinctive 4+ char
        # party word, so ordinary text can't qualify.
        party_bases = {b for (cat, _rl), rec in self.records.items()
                       if cat in ("person", "entity", "person-token",
                                  "entity-token", "short-name")
                       for b in (_pn_word_base(w) for w in rec["real"].split())
                       if len(b) >= 4 and not _pn_is_entity_keep(b)}
        if party_bases:
            # An interior apostrophe (straight or curly) is part of the token,
            # not a boundary: "FORD_O'WBYSOZ_91084" must be read whole so the
            # party base "ford" inside it is seen and the stamp faked entire —
            # otherwise the apostrophe split the run and "FORD" leaked.
            for m in re.finditer(r"(?<![A-Za-z0-9_'’])(?=[A-Za-z0-9_'’]*\d{3,})"
                                 r"([A-Za-z][A-Za-z0-9_'’]{5,})(?![A-Za-z0-9_])",
                                 _NFKC(text)):
                tok = m.group(1)
                low = tok.lower()
                if ("production number", low) in self.records:
                    continue
                if any(b in low for b in party_bases):
                    new.append(_PnTerm("production number", tok,
                                       _pn_fake_production(tok, self.registry),
                                       whole_word=True, case_sensitive=False,
                                       priority=2, source="document"))
        # VINs: a 17-char VIN-alphabet run, faked char-wise into the VIN alphabet
        # so the result stays a valid-shaped but fictitious VIN. The UNCUED
        # 17-char branch additionally requires a VALID ISO check digit —
        # federal docket numbers ("221CV01063RGKAFMX") share the alphabet and
        # length, and faking one rewrites a citation. A real VIN passes; a
        # docket essentially never does; a VIN behind a printed "VIN" label
        # stays loose (the label is the guard there).
        for m in _PN_VIN_RE.finditer(_NFKC(text)):
            vin = m.group(1) or m.group(2)
            if not vin or ("vin", vin.lower()) in self.records:
                continue
            if (m.group(2) and (len(vin) != 17
                                or vin[8].upper() != _pn_vin_check_digit(vin))):
                continue
            new.append(_PnTerm("vin", vin, _pn_fake_vin(vin, self.registry),
                               whole_word=True, case_sensitive=False,
                               priority=2, source="document"))
        self._add_terms(new)

    def register_declarant_names(self, text):
        """Detect declarant names in `text` ("DECLARATION OF X", "I, X,
        declare") and add them as person terms (full name + bare tokens) so a
        declaration's own title and signature are scrubbed even when the
        declarant isn't in the spreadsheet key. Idempotent; call before apply()."""
        for raw in _pn_declarant_names(text):
            if _pn_is_party_role(raw):
                continue
            fake_full, bare = _pn_fake_person(raw, self.registry)
            new_terms = [_PnTerm("person", raw, fake_full, whole_word=False,
                                 case_sensitive=False, priority=2,
                                 source="declarant")]
            new_terms += [_PnTerm("person-token", rt, ft, whole_word=True,
                                  case_sensitive=False, priority=1,
                                  source="declarant")
                          for rt, ft, _s in bare
                          if not _pn_is_role_token(rt)
                          and not _pn_is_generic_token(_pn_word_base(rt))]
            self._add_terms(new_terms)

    def register_declarant_refs(self, text):
        """Auto-pseudonymize the name a brief uses to cite a declaration —
        "Smith Decl., ¶ 3", "Alarcón Declaration", "Yu Dec." — even when that
        declaration itself isn't in this batch, so the reference is SCRUBBED,
        not merely flagged. Only the name is registered; the word "Declaration"
        and its variants ("Decl.", "Dec.") are never faked. A descriptive or
        procedural lead ("Supporting Declaration", "Reply Decl.") names no
        person and is skipped. Idempotent; call before apply()."""
        new = []
        for raw in _pn_declarant_ref_names(text):
            if _pn_is_party_role(raw) or ("person", raw.lower()) in self.records:
                continue
            toks = raw.split()
            # The surname must look like a name, not ordinary vocabulary — auto-
            # scrubbing is higher stakes than flagging, so a bare descriptive or
            # common word never triggers it. Two-letter surnames (Yu, Ng, Le) are
            # allowed, except those that spell a short English word ("As", "Of").
            surname = _pn_word_base(toks[-1])
            if (len(surname) < 2 or surname in _PN_COMMON_WORDS
                    or surname in _PN_PLEADING_WORDS or surname in _PN_DECL_DESCRIPTOR
                    or surname in _PN_REVIEW_NAME_STOP
                    or (len(surname) <= 2 and surname in _PN_SHORT_TOKEN_STOP)
                    or not _pn_is_name_token(toks[-1])):
                continue
            fake_full, bare = _pn_fake_person(raw, self.registry)
            new.append(_PnTerm("person", raw, fake_full, whole_word=False,
                               case_sensitive=False, priority=2,
                               source="declarant"))
            new += [_PnTerm("person-token", rt, ft, whole_word=True,
                            case_sensitive=False, priority=1, source="declarant")
                    for rt, ft, _s in bare
                    if not _pn_is_role_token(rt)
                    and not _pn_is_generic_token(_pn_word_base(rt))
                    # Hard guard: the Declaration word never becomes a term.
                    and _pn_word_base(rt) not in {"declaration", "decl", "dec"}]
        self._add_terms(new)

    def register_court_names(self, text):
        """Precautionary scrub of the venue's court personnel and courtroom, all
        DISCOVERED from the document (no name is hard-coded):

          * the presiding JUDGE, found via a judicial title. A full name behind
            the title ("Hon. Dana Whitaker") is faked wherever it appears
            TOGETHER; the bare surname is faked ONLY behind a title ("Judge
            Whitaker" -> "Judge <fake>"), never on its own. The surname's fake is
            the same in both places.
          * COURT STAFF, by role label ("Judicial Assistant: Jane Doe") — the
            full name only;
          * the DEPARTMENT NUMBER, faked digit-for-digit within its label and
            consistently across the run ("Department 515" -> "Department 372"),
            so a bare number is untouched.

        Idempotent; call before apply()."""
        text = _NFKC(text)
        new = []

        def _name_word_ok(w):
            """True when `w` can be part of a court person's name: a name-shaped
            token that is neither a stop/role word nor ordinary vocabulary. The
            gazetteer check matters in ALL-CAPS headings, where casing can't
            separate a name from the sentence continuing after it."""
            base = _pn_word_base(w)
            return (_pn_is_name_token(w)
                    and base not in _PN_JUDGE_STOP
                    and base not in _PN_NON_NAME_WORDS
                    and base not in _PN_REVIEW_NAME_STOP
                    and not _pn_is_role_token(w)
                    and not _pn_is_generic_token(base))

        def _clean_name(name):
            """The capture reduced to the leading run of plausible NAME words —
            CUT at the first stop/vocabulary word, so a heading's sentence can't
            ride in ("JUDGE SMITH ORDERED THE PARTIES…" -> ["SMITH"]). A trailing
            possessive is stripped from each word ("Whitaker's" -> "Whitaker"),
            so the possessive form shares the base surname's fake."""
            raw = re.sub(r"\s+", " ", name).strip().strip(".,;:").split()
            words = []
            for w in raw:
                w = re.sub(r"['’][sS]$", "", w)
                if not w or not _name_word_ok(w):
                    break
                words.append(w)
            if words and _pn_is_protected_locality(" ".join(words)):
                return []
            return words

        # Discover judge name(s) behind a judicial title. A full name (2-3 words)
        # is faked wherever it appears together; every surname seen behind a title
        # is faked behind a title. Both use the same registry token, so the
        # surname's fake matches across the full name and the titled forms.
        judge_surnames = {}
        for m in _PN_JUDGE_CAPTURE_RE.finditer(text):
            words = _clean_name(m.group("n"))
            if not words:
                continue
            surname = words[-1]
            judge_surnames.setdefault(
                surname.lower(), _pn_fake_name_token(surname, self.registry))
            if len(words) >= 2:
                full = " ".join(words)
                if ("person", full.lower()) not in self.records:
                    fake_full, _b = _pn_fake_person(full, self.registry)
                    new.append(_PnTerm("person", full, fake_full,
                                       whole_word=False, case_sensitive=False,
                                       priority=3, source="judge"))
        # Fake "<title> <surname>" wherever it appears, keeping the title and
        # faking only the surname; a bare surname elsewhere is left alone. The
        # trailing lookahead is (?!\w) — NOT (?!['’]) — so a possessive is left
        # in place and keeps its base surname's fake ("Judge Whitaker's" ->
        # "Judge Ashford's", the same "Ashford" as the full name).
        for surname_l, fake in judge_surnames.items():
            rx = re.compile(r"(" + _PN_TITLE_LEAD + r")(" + re.escape(surname_l)
                            + r")(?!\w)", re.IGNORECASE)
            for m in rx.finditer(text):
                whole = re.sub(r"\s+", " ", m.group(0))
                if ("court-title", whole.lower()) in self.records:
                    continue
                lead = re.sub(r"\s+", " ", m.group(1))
                new.append(_PnTerm("court-title", whole, lead + fake,
                                   whole_word=True, case_sensitive=False,
                                   priority=3, source="judge"))

        def _clean_tail(name):
            """The TRAILING run of plausible name words — for the name-first
            staff shape, where the capture reaches back into lead-in words
            ("By N. Lachikian") that _clean_name's cut-at-first-bad-word rule
            would reject whole."""
            raw = re.sub(r"\s+", " ", name).strip().strip(".,;:").split()
            words = []
            for w in reversed(raw):
                w = re.sub(r"['’][sS]$", "", w)
                # A bare initial is part of the name even when its letter
                # spells a word ("A." reduced to base "a" failed the
                # vocabulary gate, so "A. Latchinian" lost its initial and
                # the whole signature was skipped).
                if not w or not (re.fullmatch(r"[A-Z]\.?", w)
                                 or _name_word_ok(w)):
                    break
                words.append(w)
            words.reverse()
            if words and _pn_is_protected_locality(" ".join(words)):
                return []
            return words

        # Court staff — full name only, no free-standing tokens. Two shapes:
        # role label first ("Deputy Clerk: Jane Doe"), and name first, the way
        # a clerk signs a generated notice ("By: N. Lachikian, Deputy Clerk").
        seen_staff = set()
        for rx, cleaner in ((_PN_COURT_STAFF_RE, _clean_name),
                            (_PN_COURT_STAFF_NAME_FIRST_RE, _clean_tail)):
            for m in rx.finditer(text):
                words = cleaner(_pn_trim_declarant(m.group("n")))
                raw = " ".join(words)
                if len(words) < 2 or raw.lower() in seen_staff:
                    continue
                seen_staff.add(raw.lower())
                if ("person", raw.lower()) in self.records:
                    continue
                fake_full, _bare = _pn_fake_person(raw, self.registry)
                new.append(_PnTerm("person", raw, fake_full, whole_word=False,
                                   case_sensitive=False, priority=2,
                                   source="court-staff"))

        # Department number: fake the LABELED digits only, keeping the label.
        for m in _PN_DEPT_RE.finditer(text):
            num, whole = m.group(1), m.group(0)
            if ("department", whole.lower()) in self.records:
                continue
            fake = whole[: m.start(1) - m.start()] + self.registry.digits(
                num, "department")
            new.append(_PnTerm("department", whole, fake, whole_word=True,
                               case_sensitive=False, priority=2, source="court"))
        self._add_terms(new)

    def _detector_record(self, cat, real, faker):
        rk = (cat, real.lower())
        rec = self.records.get(rk)
        if rec is None:
            # E-mails, addresses and URLs are faked through the registry so the
            # map stays injective (one parcel/domain -> one fake) and an e-mail
            # reuses a party's name-token fake (zcoderre@… -> z<fake-Coderre>@…).
            if cat == "email":
                fake = self._fake_email(real)
            elif cat == "address":
                fake = self._fake_street(real)
            elif cat == "url":
                fake = self._fake_url(real)
            elif cat == "phone":
                fake = self._fake_phone(real)
            elif cat == "ssn":
                fake = self._fake_ssn(real)
            else:
                fake = faker(real)
            rec = {"category": cat, "real": real, "fake": fake,
                   "source": "regex", "count": 0,
                   "pattern": re.escape(_NFKC(real)), "flags": 0}
            self.records[rk] = rec
            # Store trimmed of trailing punctuation so a re-detected match that
            # picks up a following period (a fake suffix like "Ave." meeting a
            # sentence stop) still resolves to the same own-fake.
            self._own_fakes.add(fake.lower().rstrip(" .,;:"))
        return rec

    def _detector_cands(self, text, offset=0):
        out = []
        for cat in self.detectors:
            regex, faker = _PN_DETECTORS[cat]
            for m in regex.finditer(text):
                match = m.group(0)
                # Never rewrite a whitelisted citation URL — the authorities
                # appendix emits those deliberately — nor a fake we ourselves
                # minted (so re-scrubbing already-fake text is a fixed point).
                # A line-wrap's lead fragment ("http://www") is skipped too:
                # it names no host, and a record minted for it re-matched the
                # head of every intact kept URL as a phantom folder-wide leak.
                if cat == "url" and (_pn_url_whitelisted(match)
                                     or _pn_url_fragmentary(match)):
                    continue
                # A bare 10-digit run that is really an OCR'd date ("06/06/2025"
                # -> "0610612025") must not be faked as a phone; the date stays.
                if cat == "phone" and _pn_phone_is_ocr_date(match):
                    continue
                if match.lower().rstrip(" .,;:") in self._own_fakes:
                    continue
                rec = self._detector_record(cat, match, faker)
                out.append((3, m.start() + offset, m.end() + offset, rec))
        return out

    def _fake_street_core(self, real):
        """The faked "<number> <name> <suffix>" street alone (no suite, no
        City/ST/ZIP), keyed on the STREET IDENTITY so every spelling of one
        parcel shares one fake."""
        suffix = _pn_addr_suffix_of(real)
        core, _nums = _pn_addr_street_key(real)

        def make(rng):
            return (f"{rng.randrange(100, 9999)} "
                    f"{rng.choice(_PN_STREET_NAMES)} {suffix}")
        return self.registry.unique(core, "street", make)

    def _fake_street(self, real):
        """Injective address fake keyed on the STREET IDENTITY (number-stripped
        street name), so every spelling of one parcel — with or without the
        city tail, "Ave." vs "Avenue", "414" vs "414-416" — shares one fake.

        Only the house NUMBER and street NAME are replaced. The street-type
        suffix, the unit/suite, and the ENTIRE City/State/ZIP tail are kept
        verbatim: the locality identifies no one and is kept by house standard,
        and keeping it verbatim removes the tail-rebuild that had corrupted a
        spelled-out state ("California" -> "ia")."""
        _street, suite, cityzip = _pn_addr_parts(real)
        return self._fake_street_core(real) + suite + cityzip

    def _fake_city(self, city):
        # A protected locality ("Los Angeles") is kept even inside an address
        # tail — only the street, number and ZIP around it are faked.
        if _pn_is_protected_locality(city):
            return city
        return _pn_case_like(city, self.registry.token(city.lower(),
                                                       _PN_CITY_NAMES, "city"))

    def _fake_zip(self, zip_digits):
        return self.registry.digits(zip_digits, "zip")

    def _fake_cityzip(self, cityzip):
        """Rebuild a "…, City, ST 90640" tail: fake city, real state, fake ZIP.
        Separators are preserved verbatim."""
        if not cityzip:
            return ""
        m = _PN_CITYZIP_PARSE_RE.match(cityzip)
        if not m:
            return cityzip
        city = m.group("city").strip(" ,")
        return _pn_fake_cityzip(cityzip, self._fake_city(city),
                                self._fake_zip(m.group("zip")))

    def _fake_url(self, real):
        """Fake a URL, mapping its host to the SAME fake domain `_fake_email`
        assigns that domain, so the firm site and the firm e-mail agree."""
        m = re.match(r"(?P<scheme>https?://)?(?P<www>www\.)?(?P<host>[^/\s]+)",
                     real.strip(), re.IGNORECASE)
        host = m.group("host") if m else real
        return f"{m.group('scheme') or ''}{m.group('www') or ''}{_pn_fake_domain(host, self.registry)}"

    def _fake_phone(self, real):
        """A valid-NANP fake, format preserved, drawn through the registry so two
        different real numbers can never collide onto one fake (the bijection the
        pseudonym key round-trip depends on). Seeded on the digits alone, so a
        number written two ways maps to one fake."""
        digits = re.sub(r"\D", "", real)
        fake = self.registry.unique(
            digits, "phone", lambda rng: _pn_phone_digits(rng, len(digits)))
        return _pn_reapply_digits(fake, real)

    def _fake_ssn(self, real):
        """A valid-shaped SSN fake, format preserved, drawn through the registry
        for injectivity and seeded on the digits so every separator spelling of
        one SSN resolves to the same fake."""
        digits = re.sub(r"\D", "", real)
        fake9 = self.registry.unique(digits, "ssn", _pn_ssn_digits)
        return _pn_reapply_digits(fake9, real)

    def _compiled(self, pattern, flags):
        """A cached compiled regex for (pattern, flags), so a term's pattern is
        compiled once for the whole run instead of on every page (Python's own
        re cache evicts past 512 entries, which a large case blows through)."""
        key = (pattern, flags)
        rx = self._rx_cache.get(key)
        if rx is None:
            rx = re.compile(pattern, flags)
            self._rx_cache[key] = rx
        return rx

    def _term_cands(self, text):
        out = []
        for t in self.terms:
            for m in self._compiled(t.pattern, t.flags).finditer(text):
                if m.start() != m.end():
                    out.append((t.priority, m.start(), m.end(),
                                self.records[(t.category, t.real.lower())]))
        return out

    def _display_name_cands(self, text, offset=0):
        """Candidates for the NAME in a "Name <addr@domain>" display pair. The
        address is scrubbed by the e-mail detector; this scrubs the name beside
        it too, reusing the party's fake when the name is known and minting a
        stable person fake otherwise."""
        out = []
        for m in _PN_EMAIL_DISPLAY_RE.finditer(text):
            name, start = m.group("name"), m.start("name")
            # Trim a leading non-name lead-in the greedy capture dragged in
            # ("Contact Tommy Purscelley <…>" -> "Tommy Purscelley"), so prose is
            # not rewritten along with the name.
            lead = re.match(r"(?:(?:Contact|Email|From|To|Cc|By|See|Call|Re|And|"
                            r"With)\b[ \t]+)+", name)
            if lead:
                start += lead.end()
                name = name[lead.end():]
            words = name.split()
            # Skip a pure role/label or a name with no plausible name word (so a
            # mailer like "Notifications <no-reply@…>" is left alone).
            if _pn_is_party_role(name) or not any(_pn_is_name_token(w) for w in words):
                continue
            # A name we ourselves already faked (a prior run's fake now sitting
            # beside its already-faked e-mail) must not be re-faked — otherwise
            # each re-run mints a fresh generation off the last, growing the key
            # and the reversal chain by one every time. Same guard as
            # _detector_cands; without it the empty-detector fix-leaks path,
            # which exists precisely to stop re-faking, was bypassed here.
            if name.lower().rstrip(" .,;:") in self._own_fakes:
                continue
            rk = ("display-name", name.lower())
            rec = self.records.get(rk)
            if rec is None:
                rec = {"category": "display-name", "real": name,
                       "fake": _pn_fake_person(name, self.registry)[0],
                       "source": "regex", "count": 0,
                       "pattern": re.escape(_NFKC(name)), "flags": 0}
                self.records[rk] = rec
                # Register the mint as an own-fake (as _detector_record does), so
                # it is recognised and skipped on the next pass and re-run.
                self._own_fakes.add(rec["fake"].lower().rstrip(" .,;:"))
            out.append((2, start + offset, m.end("name") + offset, rec))
        return out

    def _display_name_repeat_cands(self, text, offset=0):
        """Candidates for a known display-name STANDING ALONE, away from the
        "Name <addr@domain>" pair that first revealed it.

        The pair is where the name is recognised; it is rarely where the name is
        printed most. A firm's paralegal appears once beside her address and
        then on every page of the letterhead and every signature block.
        `_display_name_cands` mints the record into `self.records` — so
        `surviving_reals` reports it and `write_key` writes a row for it — but
        never into `self.terms`, and `_term_cands` iterates `self.terms`. So
        every standalone occurrence rode through untouched while the key
        promised a reversal the exports did not honour: the macro would map the
        fake back onto pages where the real name was still printed.

        Word-boundary matched and case-insensitive (`_substitute` case-matches
        the fake), so an ALL-CAPS letterhead is covered and a name can never
        match inside a longer word."""
        out = []
        for (cat, _rl), rec in self.records.items():
            if cat != "display-name":
                continue
            real = _NFKC(str(rec["real"]))
            if not real:
                continue
            pat = r"(?<!\w)" + re.escape(real) + r"(?!\w)"
            for m in self._compiled(pat, re.IGNORECASE).finditer(text):
                if m.start() != m.end():
                    out.append((2, m.start() + offset, m.end() + offset, rec))
        return out

    def _trusted_party_tokens(self):
        """Distinctive lower-cased word bases of the parties named in the
        spreadsheet key (and any --term), used to decide the caption exemption.
        Corporate suffixes, connectors and role words are excluded so only a
        genuinely identifying token ("orellana", "acquisition") counts."""
        cached = getattr(self, "_trusted_tok_cache", None)
        if cached is not None:
            return cached
        toks = set()
        for t in self.terms:
            if t.source not in ("spreadsheet", "--term"):
                continue
            if t.category not in ("person", "entity", "person-token",
                                  "entity-token"):
                continue
            for w in t.real.split():
                base = _pn_word_base(w)
                if (len(base) >= 3 and not _pn_is_entity_keep(base)
                        and not _pn_is_role_token(w)):
                    toks.add(base)
        self._trusted_tok_cache = toks
        return toks

    def _side_is_trusted(self, side):
        trusted = self._trusted_party_tokens()
        return any(_pn_word_base(w) in trusted for w in side.split())

    def _protected_citation_spans(self, text):
        """Spans of published-authority citations whose party names must NOT be
        rewritten. Renaming a cited decision is a worse failure than leaving a
        party name in (the method's cardinal invariant), so any candidate
        overlapping one of these spans is dropped in `_substitute`.

        Caption exemption: a "X v. Y" case span is NOT protected — it is the
        ruling's own caption and must still be replaced — when BOTH sides name a
        trusted party from the key. One trusted side is not enough (Sanchez v.
        Valencia Holding Co. has a trusted-looking plaintiff yet is authority)."""
        spans = []
        try:
            cites = find_all_citations(text)
        except Exception:
            return spans
        for c in cites:
            if (c.get("kind") == "case" and c.get("plaintiff") is not None
                    and self._side_is_trusted(c["plaintiff"])
                    and self._side_is_trusted(c.get("defendant", ""))):
                continue  # this document's own caption — replace it
            spans.append(tuple(c["span"]))
        # A "…, No. <docket>, <year> WL <n>" run names an authority even when
        # the citation parser can't read it (no "v." — a case cited by its
        # defendant name alone; a court paren the OCR dropped the year from).
        # Protect the run plus the name immediately before it. Protection-only:
        # the worst case is a nearby value left unreplaced (a warning), never
        # a renamed authority.
        for m in _PN_WL_RUN_RE.finditer(text):
            spans.append(m.span())
        # "In re <Name> ... Litig." and "<Name> Warranty Cases" (and their bare
        # short forms) name authorities the strict parser can't read — a party
        # token welded into one otherwise gets rewritten inside the cite.
        # Protection-only, same as the WL run above.
        for rx in (_PN_INRE_LITIG_RE, _PN_CASES_RUN_RE):
            for m in rx.finditer(text):
                spans.append(m.span())
        return spans

    def _keep_spans(self, text):
        """Spans of operator-KEPT values in `text` — a value marked `no` (keep
        the whole value) or the bracketed part of a keep-spec. A candidate
        overlapping one of these is dropped in `_substitute`, so the kept text is
        left verbatim even when a bare-token term or a PII detector would fake it.

        Two rules make this precise:
          * EXACT WORD only — the keep matches `value` on WORD BOUNDARIES, so a
            `no` on "Cal" keeps the standalone word but never the "Cal" inside
            "California" or "Calvin".
          * a FULL party-name match WINS — a kept word that falls inside a full
            person/entity/case_number term (a real party this case fakes) is
            RELEASED so the party is still faked. So "Cal" is kept on its own but
            "CAL EQUIPMENT FE RANCH, LLC" is faked whole; "Doe" is kept alone but
            "John Doe" is faked. Bare tokens and detectors do NOT release a keep."""
        if not self.keep_soft and not self.keep_strict:
            return []
        # Spans of full party-name matches, which override EITHER kind of keep.
        party = []
        for t in self.terms:
            if t.category not in _PN_PARTY_OVERRIDE_CATS:
                continue
            for m in self._compiled(t.pattern, t.flags).finditer(text):
                if m.start() != m.end():
                    party.append((m.start(), m.end()))

        def in_party(s, e):
            return any(s < pe and ps < e for ps, pe in party)

        spans = []

        def collect(values, soft, party_wins=True):
            for v in sorted(values, key=len, reverse=True):
                if not v:
                    continue
                # The name-run release only makes sense for a single-word keep
                # (a `no` on one word); a multi-word phrase or a structured value
                # like an e-mail is kept on the full-party rule alone.
                run_check = soft and bool(re.fullmatch(r"[A-Za-z][A-Za-z.'’\-]*", v))
                hit = False
                for m in re.finditer(r"(?<!\w)" + re.escape(v) + r"(?!\w)",
                                     text, re.IGNORECASE):
                    s, e = m.span()
                    if s == e or (party_wins and in_party(s, e)):
                        continue      # a full party match here — party fakes it
                    # A soft `no` keep is also released where the word sits in a
                    # multi-word capitalized name run (a possible party), so only
                    # the standalone word is kept. A strict bracket keep is not.
                    if run_check and self._in_name_run(text, s, e):
                        continue
                    spans.append((s, e))
                    hit = True
                if hit:
                    self.kept_hits.add(v.lower())

        local_strict = self.keep_strict_local & self.keep_strict
        collect(local_strict, soft=False, party_wins=False)
        collect(self.keep_strict - local_strict, soft=False)
        collect(self.keep_soft, soft=True)
        return spans

    def _in_name_run(self, text, s, e):
        """True when the word at [s,e] is immediately adjacent to another
        name-shaped word (part of a multi-word proper-noun phrase like "Cal
        Equipment"), so a soft `no` keep there is released. Adjacency is a single
        run of spaces/tabs — a comma, period or line break ends the phrase."""
        r = re.match(r"[ \t]+([A-Za-z][A-Za-z.'’&\-]*)", text[e:])
        if r and _pn_is_name_word(r.group(1)):
            return True
        l = re.search(r"([A-Za-z][A-Za-z.'’&\-]*)[ \t]+$", text[:s])
        return bool(l and _pn_is_name_word(l.group(1)))

    def _mask_protected_citations(self, text):
        """`text` with every protected citation span blanked to spaces, for the
        LEAK scans: a party name correctly preserved inside a cited authority
        ("Silvio v. Ford Motor Co.") is the protection working, not a leak —
        counting it quarantined ten clean exports over text that was required
        to stay."""
        spans = self._protected_citation_spans(text)
        if not spans:
            return text
        chars = list(text)
        for s, e in spans:
            for i in range(s, min(e, len(chars))):
                if not chars[i].isspace():
                    chars[i] = " "
        return "".join(chars)

    def _substitute(self, text, cands, reflow=False, count=True, protected=None):
        # Biggest / most-specific first; skip anything overlapping a chosen span.
        cands.sort(key=lambda c: (-c[0], -(c[2] - c[1])))
        chosen, occ = [], []
        for _prio, s, e, rec in cands:
            if any(s < oe and os < e for os, oe in occ):
                continue
            # Never rewrite inside a protected citation span (P0-A): a candidate
            # overlapping a cited decision's name is dropped so the authority
            # survives byte-for-byte.
            if protected and any(s < pe and ps < e for ps, pe in protected):
                continue
            occ.append((s, e))
            chosen.append((s, e, rec))
        for s, e, rec in sorted(chosen, key=lambda c: -c[0]):  # right-to-left
            # Match the casing of the text being replaced: a caption written in
            # ALL CAPS gets an ALL-CAPS fake, prose gets the canonical form. The
            # stored fake carries whatever casing the spreadsheet cell had, so
            # without this an all-caps "Title Defendant" shouts in every
            # sentence. An e-mail fake is left alone — its casing is meaningless
            # and upper-casing a domain looks wrong.
            fake = rec["fake"]
            if rec["category"] not in ("email", "url"):
                fake = _pn_case_like(text[s:e], fake)
            if reflow:
                fake = _pn_reflow(text[s:e], fake)
            tail = text[e:]
            # A fake ending in a period ("Ave.", "Inc.") meeting a following
            # sentence period would double it ("Ave.. Email"); drop the extra.
            if fake.endswith(".") and tail[:1] == ".":
                tail = tail[1:]
            text = text[:s] + fake + tail
            if count:
                rec["count"] += 1
        return text

    def apply(self, text, count=True):
        """Return `text` with every match replaced by its stable fake.
        Single pass over the ORIGINAL offsets so a fake can never be re-matched
        by a later term. `count=False` runs the substitution without recording
        occurrences — used for the leak re-scan, which must not inflate the
        numbers written into the key spreadsheet."""
        if count:
            self.texts_applied += 1
        text = _NFKC(text)
        # The repeat pass runs AFTER _display_name_cands so a name first met in
        # this very text is already a record and its other occurrences are
        # covered on the same pass, not only from the next page onward.
        cands = (self._detector_cands(text) + self._term_cands(text)
                 + self._display_name_cands(text)
                 + self._display_name_repeat_cands(text))
        return self._substitute(
            text, cands, count=count,
            protected=self._protected_citation_spans(text) + self._keep_spans(text))

    def apply_lines(self, bodies):
        """Pseudonymize a pleading page's line BODIES as one text, returning one
        body per input line.

        Matching line by line cannot see a party name that the printed page
        broke across two numbered lines — which is how a long caption name
        ("DREAM TEAM REAL ESTATE / CONSULTANTS, INC.") almost always appears, so
        that name silently survived. Terms are matched against the joined text;
        the PII detectors still run per line, so a printed line number can never
        be read as the leading digits of a street address. `_pn_reflow` puts the
        line breaks back, so the line count — and every pinpoint cite keyed to
        it — is unchanged."""
        self.texts_applied += 1
        bodies = [_NFKC(b) for b in bodies]
        joined = "\n".join(bodies)
        cands = (self._term_cands(joined) + self._display_name_cands(joined)
                 + self._display_name_repeat_cands(joined))
        off = 0
        for b in bodies:
            cands += self._detector_cands(b, off)
            off += len(b) + 1
        out = self._substitute(
            joined, cands, reflow=True,
            protected=self._protected_citation_spans(joined)
            + self._keep_spans(joined)).split("\n")
        # _pn_reflow preserves every newline, so this holds; fall back rather
        # than ever hand back the wrong number of lines.
        return out if len(out) == len(bodies) else [self.apply(b) for b in bodies]

    def _fake_email(self, real):
        """Fake an e-mail address, reusing the fakes already assigned to any
        party name-tokens that appear in its local part. So if "Coderre" ->
        "Forsythe" and the address is zcoderre@gmail.com, the fake local part
        keeps the same stand-in: zforsythe@…. The domain is always neutralised.

        EVERY dot-separated piece of the local part is rewritten, not only the
        pieces that hold a known name. Substituting the name alone left the rest
        of the handle standing — `roxane.enterprise1@gmail.com` came out as
        `tolliver.enterprise1@letterbox.co`, and `enterprise1` is a real,
        searchable handle. A piece with no known name in it is swapped for a
        stable word drawn from the registry (unique, so it can never collide
        with a party's fake)."""
        local, sep, domain = real.partition("@")
        if not sep:
            return _pn_fake_email(real)
        tokens = self.registry.tokens_for("nametok")
        # Only distinctive tokens (>=4 chars) so short/common fragments don't
        # rewrite unrelated letters; longest-first for greedy, non-overlapping
        # substitution.
        items = sorted(((r, f) for r, f in tokens.items() if len(r) >= 4),
                       key=lambda rf: -len(rf[0]))
        by_real = {r: f for r, f in items}
        pat = (re.compile("|".join(re.escape(r) for r, _ in items), re.IGNORECASE)
               if items else None)

        # Split on the separators the local part is allowed to use, keeping them.
        pieces = re.split(r"([._%+\-])", local)
        out = []
        for i, piece in enumerate(pieces):
            if i % 2 or not piece:      # a separator, or an empty run
                out.append(piece)
                continue
            hit = False
            if pat is not None:
                def repl(mm):
                    nonlocal hit
                    hit = True
                    return by_real[mm.group(0).lower()].lower()
                piece = pat.sub(repl, piece)
            if hit:
                # A known name carried the piece. Strip any residual digits
                # ("roxane2" -> "tolliver"): a numeric suffix is part of the
                # real handle, not of the name.
                piece = re.sub(r"\d+", "", piece)
            else:
                # No known name: replace the whole piece with a stable, unique
                # stand-in word rather than letting the real handle through.
                piece = self.registry.token(piece.lower(), _PN_NAME_WORDS,
                                            "emaillocal").lower()
            out.append(piece)
        local = "".join(out).strip("._%+-") or _pn_rng(
            "email", real.lower()).choice(_PN_NAME_WORDS).lower()
        return f"{local}@{_pn_fake_domain(domain, self.registry)}"

    def surviving_reals(self, text):
        """Real values that still appear in `text` — a leak the caller should
        surface. Every tracked value is checked, including one that was replaced
        ZERO times: a term whose only occurrences were line-wrapped used to
        match nothing, so it had no replacement count, so the old count>0 guard
        skipped it and the leak was reported as clean."""
        text = self._mask_protected_citations(_NFKC(text))
        # A value the operator marked KEEP is present ON PURPOSE — that is what
        # the decision means — so it is not a leak. Reporting it anyway put a
        # permanent row in LEAKS.xlsx that no answer could ever clear: `no` is
        # what produced it, and the durable decision lives on the master KEEP
        # sheet, so consuming the local worksheet never retired the row either.
        # The review scans already suppress kept values; this tier did not.
        #
        # Scoped to the categories a keep actually survives in. A keep is
        # RELEASED inside a full party match (`_keep_spans`), so a
        # person/entity/case_number real that is still standing was faked
        # nowhere and IS a leak — the safety rule that stops a keep leaving a
        # party in the clear must not be undone here.
        kept = {str(v).lower() for v in (self.keep_soft | self.keep_strict)}
        out = []
        for rec in self.records.values():
            if (kept and rec["category"] not in _PN_PARTY_OVERRIDE_CATS
                    and str(rec["real"]).lower() in kept):
                continue
            try:
                if self._compiled(rec["pattern"], rec["flags"]).search(text):
                    out.append(rec["real"])
            except re.error:
                pass
        return out

    def scrub_emails(self, text):
        """Write-side sweep for a tracked E-MAIL address the pattern pass left
        standing, replaced with the fake already assigned to it.

        An e-mail is ALWAYS faked — that is the policy, so a worksheet row asking
        whether to fake one is never a decision, just a row to clear. The way to
        earn that is to leave none behind: `apply()` resolves overlapping
        candidates and yields to protected spans, so an address can survive a
        second occurrence it could not claim. A literal sweep afterwards catches
        those, using the SAME fake the record already carries — deterministic,
        injective and reversible, since the key row already exists.

        Scoped to `email` records on purpose: an address never appears inside a
        published citation, so this cannot rename an authority (a bare URL can,
        which is why urls are left to the pattern pass and its citation
        protection)."""
        src = _NFKC(text)
        for (cat, _rl), rec in self.records.items():
            if cat != "email":
                continue
            real, fake = str(rec["real"]), str(rec["fake"])
            if not real or real == fake:
                continue
            pat = re.compile(re.escape(real), re.IGNORECASE)
            src, n = pat.subn(fake, src)
            if n:
                rec["count"] += n
        return src

    def _weld_core(self, rec):
        """`(core, short)` — the alphanumeric core the reduced passes may hunt
        for `rec` by, and whether it qualified only under the SHORT-name rule
        (which additionally demands a real weld at the match site). `("", False)`
        when the record is not eligible at all. See `_PN_WELD_CORE_MIN`."""
        if rec["category"] not in _PN_WELD_CORE_CATS:
            return "", False
        core = _pn_alnum_core(rec["real"])
        if len(core) >= _PN_WELD_CORE_MIN:
            return core, False
        if (len(core) >= _PN_WELD_SHORT_CORE_MIN
                and rec["category"] in _PN_WELD_SHORT_CORE_CATS
                and core in self._trusted_party_tokens()
                and core not in _PN_COMMON_WORDS
                and _pn_is_name_token(core.title())):
            return core, True
        return "", False

    def _reduce_with_index(self, src):
        """`(reduced, idx)` — `src` folded to bare lower-case alphanumerics, with
        `idx[i]` the offset in `src` that produced `reduced[i]`, so a hit in the
        reduction can be located back in the original text."""
        reduced, idx = [], []
        for i, ch in enumerate(src):
            c = _pn_ascii_fold(ch).lower()
            if "a" <= c <= "z" or "0" <= c <= "9":
                reduced.append(c)
                idx.append(i)
        return "".join(reduced), idx

    def surviving_reals_reduced(self, text):
        """Tracked reals found in the ALPHANUMERIC-ONLY reduction of `text`.

        On a column-spliced caption the whole-word patterns match nothing —
        `Schilleci & Tortorici` extracts as `SCHILLECITORTORICILAW` with no
        boundaries — yet the reduction `schillecitortoricilaw` still CONTAINS
        `schillecitortorici`. Which cores qualify is `_weld_core`'s call, and it
        must stay the mirror of `scrub_welded`'s: a value this reports but that
        one cannot cure is a leak the substituter was never able to clean."""
        masked = self._mask_protected_citations(_NFKC(text))
        red, idx = self._reduce_with_index(masked)
        out = []
        for rec in self.records.values():
            core, short = self._weld_core(rec)
            if not core:
                continue
            k = red.find(core)
            while k >= 0:
                if not short or _pn_span_is_welded(
                        masked, idx[k], idx[k + len(core) - 1] + 1):
                    out.append(rec["real"])
                    break
                k = red.find(core, k + 1)
        return out

    def scrub_welded(self, text):
        """Write-side mirror of `surviving_reals_reduced`: remove a tracked real
        that survives only in the ALPHANUMERIC reduction of `text`.

        On a column-spliced caption a party name welds to its neighbour
        (`CULTUREEDITservice`, `AZUL@SCHILLECITORTORICILAW`), so the boundary-
        anchored term patterns leave it standing while the reduced substring test
        still finds it — detection then out-runs replacement and the leak gate
        quarantines an export the substituter simply couldn't clean. Here the
        welded span is located in the ORIGINAL text (via a reduced-index map) and
        replaced with the party's fake, so replacement can never lag detection.

        Longest cores first and non-overlapping, so a name nested in a longer
        firm name doesn't double-substitute; the fake is case-matched to the
        span it replaces. Even a long core can coincide inside an unrelated
        word, so callers run this ONLY on a page already flagged spliced — a
        page the tool already reports as corrupted and asks a human to verify."""
        src = _NFKC(text)
        reduced, idx = self._reduce_with_index(src)
        if not reduced:
            return text
        # The same citation protection the ordinary substitution path applies
        # (P0-A): a cited decision whose party name contains a tracked core
        # ("Sanchez v. Valencia Holding Co.") must survive byte-for-byte even
        # on a splice-flagged page — renaming an authority is a worse failure
        # than leaving a welded name in.
        protected = self._protected_citation_spans(src)
        cands = []
        for rec in self.records.values():
            core, short = self._weld_core(rec)
            if core:
                cands.append((core, _pn_fake_core_display(rec["fake"]), rec,
                              short))
        cands.sort(key=lambda c: -len(c[0]))
        taken = [False] * len(reduced)
        repls = []
        for core, fake, rec, short in cands:
            start = 0
            while True:
                k = reduced.find(core, start)
                if k < 0:
                    break
                end = k + len(core)
                o_s, o_e = idx[k], idx[end - 1] + 1
                if any(o_s < pe and ps < o_e for ps, pe in protected):
                    start = k + 1
                    continue
                # A short name only earns a cure where it is genuinely WELDED —
                # standing alone it was the boundary-anchored pass's to make,
                # and that pass yields to keeps and citations this one cannot
                # see.
                if short and not _pn_span_is_welded(src, o_s, o_e):
                    start = k + 1
                    continue
                if not any(taken[k:end]):
                    for z in range(k, end):
                        taken[z] = True
                    repls.append((o_s, o_e, fake, rec))
                start = k + 1
        if not repls:
            return text
        out = src
        for o_s, o_e, fake, rec in sorted(repls, key=lambda r: -r[0]):
            out = out[:o_s] + _pn_case_like(out[o_s:o_e], fake) + out[o_e:]
            # Record the welded replacement so the reversal key carries this
            # mapping: the fake is now in the document, so ReAnonymize must be
            # able to undo it (an unrecorded row is omitted from the key).
            rec["count"] += 1
        return out

    def note_leaks(self, reals):
        """Record that `reals` survived in some export, for the key's Status."""
        for r in reals:
            self.leaked.add(r.lower())

    def has_leaks(self):
        """True when any real value survived in an export (a `leaked` key row).
        What that does to the run is tiered (see `primary_leaks` and the
        `leak_gate` config): the pseudonymization is a precaution against
        casual recognition of a public filing, so only a survivor that defeats
        that purpose outright blocks delivery."""
        return bool(self.leaked)

    # Categories whose surviving real value names the party OUTRIGHT — a full
    # person/entity name or a document-defined short form. One of these in an
    # export makes the case immediately recognizable, which is the one failure
    # the whole precaution exists to prevent. A bare token, an address, or an
    # identifier surviving is a missed spot to review, not a defeat.
    _PRIMARY_LEAK_CATS = ("person", "entity", "short-name", "display-name")

    def primary_leaks(self):
        """The leaked values that name a party outright (see
        _PRIMARY_LEAK_CATS) — the survivors that justify quarantining the
        exports under the default `leak_gate = primary`."""
        out = set()
        for (cat, rl), rec in self.records.items():
            if cat in self._PRIMARY_LEAK_CATS and rl in self.leaked:
                out.add(rec["real"])
        return out

    def files_to_quarantine(self, gating):
        """The written .txt paths that actually contain one of the `gating`
        values — so only the offending exports are renamed to *.LEAK, not the
        whole batch. A file whose leak was in `gating` is held; every other file
        is delivered. `gating` is the gate's blocking set (strict: all leaks;
        primary: party-naming leaks)."""
        want = {str(v).lower() for v in gating}
        return [txt for txt in self.written
                if self.leaked_by_file.get(txt, set()) & want]

    def known_fake_words(self):
        """Every word this run has minted as a stand-in, lower-cased. Lets the
        open-world review scan tell a fake it just wrote from a real survivor.

        An e-mail/URL fake is stored whole ("local@postbox2.org",
        "www.postbox2.org"), so a plain whitespace split never yields the bare
        registrable host — and the review scan then flagged the tool's OWN
        numbered fake domain (`postbox2.org`) as an unknown identifier. Contribute
        the bare host too (drop a local part, scheme, leading www., and any
        path) so those false positives are suppressed while a real domain in a
        source document is still surfaced."""
        words = set()
        for rec in self.records.values():
            for w in str(rec["fake"]).split():
                w = w.strip(".,").lower()
                if not w:
                    continue
                words.add(w)
                host = re.sub(r"^https?://", "", w.split("@", 1)[-1])
                if host.startswith("www."):
                    host = host[4:]
                host = host.split("/", 1)[0].rstrip(".")
                if "." in host and re.fullmatch(r"[a-z0-9.\-]+", host):
                    words.add(host)
        return words

    def review_scan(self, text):
        """Run the open-world identifier-shape scan, accumulate the findings for
        the folder-level report, and return them."""
        found = _pn_review_findings(text, self.known_fake_words())
        seen = {(c, s.lower()) for c, s in self.review}
        for c, s in found:
            if (c, s.lower()) not in seen:
                seen.add((c, s.lower()))
                self.review.append((c, s))
        return found

    def review_definition_survivors(self, source, output):
        """Backstop for the acronym short-form path: a document-defined
        parenthetical short form (`("TOFF")`) whose token is an initialism of a
        TRACKED party and still appears verbatim in the finished `output` — which
        means `register_short_names` missed that definition shape and the highest-
        value survivor (a one-lookup key inversion) shipped uncaught.

        Scoped to initialisms of a tracked party, so an ordinary defined term
        (`("Agreement")`, `("ISO")`, `("RJN")`) is never surfaced — it is not a
        party's initials. Accumulates into the folder review list and returns the
        new findings."""
        parties = [(rec["real"].split(), str(rec["fake"]).split())
                   for (cat, _rl), rec in self.records.items()
                   if cat in ("entity", "person")]
        if not parties:
            return []
        findings, local = [], set()
        for m in re.finditer(r"\(([^()]{0,120})\)", _NFKC(source)):
            for form in _pn_paren_short_forms(m.group(1)):
                low = form.lower()
                if len(form.split()) != 1 or low in local:
                    continue
                # A 2-char title-case token ("El") is a word, not an acronym;
                # only an all-caps form or a 3+ char initialism is worth a
                # reviewer's attention.
                if len(form) < 3 and not form.isupper():
                    continue
                if not re.search(r"(?<!\w)" + re.escape(form) + r"(?!\w)", output):
                    continue          # already scrubbed / absent — fine
                if any(_pn_initialism_fake(form, rw, fw) for rw, fw in parties):
                    local.add(low)
                    findings.append(("party acronym", form))
        seen = {(c, s.lower()) for c, s in self.review}
        for c, s in findings:
            if (c, s.lower()) not in seen:
                seen.add((c, s.lower()))
                self.review.append((c, s))
        return findings

    def reid_scan(self, text):
        """Adversarial re-identification pass over the FINISHED output: any
        label-anchored identifier or VIN shape still present whose value is not
        a fake this run minted. register_identifiers scrubs these at the source,
        but it reads the source text — an identifier that only materialises in
        the output rendering (a column re-join, a differently-wrapped line)
        would slip past it and still invert the whole pseudonym map in one
        public lookup. Framing the check as its own output-side scan means the
        export is never certified clean while carrying one. Findings accumulate
        into the folder review list, tagged REID so they sort above the ordinary
        review noise."""
        fake_vals = set()
        for rec in self.records.values():
            fake_vals.add(re.sub(r"[^a-z0-9]", "", str(rec["fake"]).lower()))
        findings = []
        for cls, val in _pn_identifier_values(text):
            if cls not in _PN_REID_CLASSES:
                continue
            if re.sub(r"[^a-z0-9]", "", val.lower()) in fake_vals:
                continue          # our own stand-in, riding under its label
            # When in doubt, a short bare number is not a re-identification key.
            # A retail "DEAL# 512" / "Inventory No. 55" resolves to nothing on
            # its own, so reporting it is noise. Bar numbers (a definite State
            # Bar lookup, and 5-7 digits by definition) are the one short-number
            # class that IS a key, so they are exempt; every other class here is
            # alphanumeric or already 6+ digits, so this only ever filters the
            # weakly-distinctive account-id stamps.
            digits = re.sub(r"\D", "", val)
            if (cls != "bar number" and not re.search(r"[A-Za-z]", val)
                    and len(digits) < 6):
                continue
            findings.append((f"REID {cls}", val))
        for m in _PN_VIN_RE.finditer(_NFKC(text)):
            vin = (m.group(1) or m.group(2) or "").upper()
            # The uncued shape must carry a valid check digit — otherwise a
            # federal docket number gets reported as a surviving VIN.
            if (m.group(2) and (len(vin) != 17
                                or vin[8] != _pn_vin_check_digit(vin))):
                continue
            if vin and vin.lower() not in fake_vals:
                findings.append(("REID vin", vin))
        seen = {(c, s.lower()) for c, s in self.review}
        out = []
        for c, s in findings:
            if (c, s.lower()) not in seen:
                seen.add((c, s.lower()))
                self.review.append((c, s))
                out.append((c, s))
        return out

    def apply_to_citation(self, cite, s):
        """Pseudonymize a CITATION-derived string (an appendix display line, a
        search query) only when `cite` is this case's own caption — both
        v.-sides trusted — which surfaces as a slip cite and must scrub. Every
        real authority passes through verbatim: a bare cite key re-parsed by
        apply() often can't be re-detected as a citation (the key format drops
        the reporter context), so the party-token terms would rewrite the
        authority's name — the appendix shipped "McGee v. <fake>, LLC" beside
        a body that correctly kept the real name."""
        if (cite.get("kind") == "case" and cite.get("plaintiff") is not None
                and self._side_is_trusted(cite["plaintiff"])
                and self._side_is_trusted(cite.get("defendant", ""))):
            return self.apply(s)
        return s

    def unknown_name_scan(self, text):
        """High-recall tier over the FINISHED output: role-anchored capitalised
        runs that are neither our own fakes nor common words (see
        _pn_unknown_name_findings). Run on the output, a correctly scrubbed
        party shows up as its fake (neutral, silent); an unlisted survivor —
        the "Travelers" class of leak — gets flagged. Accumulates into the
        folder review list."""
        findings = _pn_unknown_name_findings(text, self.known_fake_words())
        seen = {(c, s.lower()) for c, s in self.review}
        out = []
        for c, s in findings:
            if (c, s.lower()) not in seen:
                seen.add((c, s.lower()))
                self.review.append((c, s))
                out.append((c, s))
        return out

    def _status(self, rec):
        if rec["real"].lower() in self.leaked:
            return "leaked"
        return "replaced" if rec["count"] > 0 else "no match"

    def alias_candidates(self):
        """Groups of registered PERSON names that share a given name but differ
        in surname — e.g. "Roxane Estrada" / "Roxane Guzman", the same party
        written two ways, which were faked with UNRELATED surnames. The tool
        does not merge these automatically: linking by shared given name is
        inference, and inference is how a pseudonymizer conflates two different
        people. It surfaces them instead, so a reviewer can add an explicit
        alias term (`--term`) that binds both to one fake. Returns
        [[real, real, ...], ...], each inner list a possible-alias cluster."""
        # Firm-wrapper lead words: a "Law Office(s) of <person>" string routed
        # into the person pool must not cluster on the shared word "Law" (which
        # paired two unrelated firms as if they were one attorney's aliases).
        # Strip a leading wrapper run so the given name is the real personal one.
        lead = {"law", "office", "offices", "firm", "firms", "of", "the",
                "and", "&"}
        groups = {}
        for (cat, _rl), rec in self.records.items():
            if cat != "person":
                continue
            core = [t for t in rec["real"].split()
                    if len(t.strip(".")) > 1 and re.search(r"[A-Za-z]", t)
                    and not _pn_is_suffix_token(t)]
            while core and core[0].lower().rstrip(".") in lead:
                core.pop(0)
            if len(core) < 2:
                continue
            given, surname = core[0].lower().rstrip("."), core[-1].lower().rstrip(".")
            groups.setdefault(given, {}).setdefault(surname, rec["real"])
        return [sorted(d.values()) for d in groups.values() if len(d) > 1]

    def override_fixes(self, explicit):
        """Force operator-typed replacements (from the leak worksheet's Fix?
        column) to win over any auto-derived fake, under WHATEVER category the
        value was recorded. Marks each overridden row authoritative so write_key
        persists it and a later run reuses the typed fake instead of re-deriving
        one. `explicit` is {real_value: replacement}; a replacement equal to its
        own value is skipped (a self-map never scrubs)."""
        want = {str(k).lower(): str(v) for k, v in (explicit or {}).items()
                if v and str(v).strip().lower() != str(k).strip().lower()}
        if not want:
            return
        for (_cat, rl), rec in self.records.items():
            if rl in want:
                rec["fake"] = want[rl]
                rec["loaded"] = True
                rec["source"] = "leak-fix"

    def write_key(self, path, log):
        """Write the REVERSAL key the Word macro consumes. A row is written when
        it MATCHED (Occurrences > 0), when it came from a REUSED key, or when it
        is an AUTHORITATIVE binding — a party the E-Court template or a --term
        named (`_PN_KEY_UNMATCHED_SOURCES`) — and never when it contains a
        newline (Word's Find cannot match one, so such a row is dead).

        The old rule wrote only what matched, on the reasoning that `ReAnonymize`
        must never replace a Real Value that was never a party. That reasoning
        holds for an INFERRED value — a pre-scan guess, a declarant read off a
        signature block — and those are still dropped when they match nothing.
        It does not hold for a name the operator's own party template lists: that
        value IS a party, its fake is already minted (every term gets one at
        build time, matched or not), and dropping the row threw away the only
        durable record of a binding the case will need the moment the party is
        finally mentioned. Concretely: a party named ONLY in a document missing
        from the first batch had no row, so adding that document later either
        leaked the name or renamed it inconsistently.

        Such a row carries Status "no match", so the sheet still says plainly
        which values were never written into an export.

        xlsx if openpyxl is available, else JSON alongside.
        """
        # Nothing to say — EXCEPT when this run retired every row it loaded (the
        # operator marked them KEEP): the stale key on disk still names real
        # values against fakes no longer applied, so it must be rewritten empty
        # rather than left standing.
        if not self.records and not getattr(self, "_retired_key_rows", None):
            return

        # Values with no fake to reverse, so no key row. Which keeps qualify
        # depends on whether the decision is the operator's OWN:
        #   * a LOCAL keep (this folder's key or LEAKS) retires the binding, so
        #     the value really is left verbatim — drop any row for it, including
        #     a KEEP-PART's flagged phrase ("John Doe's Opposition"), whose place
        #     the surviving fragment ("John Doe") takes; but
        #   * an INHERITED keep (another case's, via the master sheet) never
        #     retires this case's binding — a current-case party is faked anyway
        #     (see _pn_retire_kept_key_terms), so its row MUST stay or the export
        #     carries a fake nothing can reverse.
        decisions = getattr(self, "_keep_decisions", None)
        if decisions is None:
            keep_low = {str(k).lower() for k in (self.keep_soft | self.keep_strict)}
        else:
            local = getattr(self, "_keep_local", ())
            keep_low = set()
            for vl, d in decisions.items():
                if vl not in local:
                    continue
                keep_low |= {p.lower() for p in _pn_decision_keep_parts(d)}
                if d.get("fake_values") is not None:
                    keep_low.add(vl)

        def _reversible(r):
            # A row that matched this run (count>0), one loaded from a reused
            # key (preserve it even if it didn't re-match — in fix-leaks mode the
            # text is already scrubbed, so loaded party names never re-match yet
            # must stay in the key for the reversal macro), or an AUTHORITATIVE
            # binding the party template / --term named, which pins the fake for
            # the run that finally meets that party. A value the operator marked
            # KEEP is left verbatim, so it has no fake to reverse — drop any
            # (e.g. harvested) row for it so the key doesn't show a fake that
            # was never applied.
            return ((r["count"] > 0 or r.get("loaded")
                     or (r.get("source") in _PN_KEY_UNMATCHED_SOURCES
                         and not r.get("derived")))
                    and str(r["real"]).lower() not in keep_low
                    and "\n" not in str(r["real"]) and "\n" not in str(r["fake"]))
        keyrows = [r for r in self.records.values() if _reversible(r)]

        # Emit a row per first/last name so a token-level reversal macro can
        # undo a bare "Kingsley" — not only the full "Ivers Kingsley". Each
        # person/entity fake is composed word-for-word, so its tokens are
        # recovered by zipping the real and fake word lists; only tokens not
        # already keyed (as a person-token / entity-token) are added.
        seen = {(r["category"], str(r["real"]).lower()) for r in keyrows}
        derived = []
        for r in keyrows:
            if r["category"] not in ("person", "entity"):
                continue
            tokcat = f"{r['category']}-token"
            for rtok, ftok in _pn_name_token_rows(r["real"], r["fake"]):
                dk = (tokcat, rtok.lower())
                if dk in seen:
                    continue
                seen.add(dk)
                derived.append({"category": tokcat, "real": rtok, "fake": ftok,
                                "source": r["source"], "count": r["count"]})
        keyrows += derived

        # ONE REVERSIBLE row per fake. The registry is INJECTIVE, so two rows
        # sharing a Replacement are never two different parties — they are
        # always alternate SPELLINGS of one real value, deliberately registered
        # against the canonical value's fake so that every spelling scrubs: a
        # wrap-split hyphen ("Ardeshirpour- Zartoshti" for
        # "Ardeshirpour-Zartoshti"), a `_pn_name_variants` near-miss ("Sarra"
        # for "Sara"). Right for the forward pass, and unusable in reverse:
        # `DeAnonymize.bas` cannot tell a synthetic spelling from the real one,
        # so it treats a fake claimed by two Real Values as ambiguous and
        # retires the mapping outright — fail-safe, but the fake is then left
        # standing in the tentative. Registering those spellings therefore made
        # every hyphenated party and every OCR-matched name un-restorable.
        #
        # The non-canonical rows are marked `alt spelling` in Status rather than
        # dropped. Both halves of the contract need them:
        #   * the macro reads Status and uses an `alt spelling` row only in the
        #     REAL->FAKE direction, where it is unambiguous, so exactly one row
        #     reverses each fake and the ambiguity guard never fires; and
        #   * `_pn_load_key` still pins them, so a re-run that has only the key
        #     (no party template to rebuild the variants from) reproduces the
        #     already-delivered exports BYTE-IDENTICALLY. Dropping the rows
        #     broke that: the re-run scrubbed the wrap-split spelling half by
        #     half ("Sedgwick- Linford") instead of as one token.
        # The canonical row is the one the fake was composed from (not
        # `derived`); when a group is ALL derived — the canonical value never
        # matched and isn't authoritative, so it earned no row — one variant is
        # promoted, or the fake would reverse to nothing.
        alt_rows, owner_count, nalt = set(), {}, 0
        by_fake = {}
        for r in keyrows:
            by_fake.setdefault(str(r["fake"]).lower(), []).append(r)
        for group in by_fake.values():
            if len(group) < 2:
                continue
            canon = [r for r in group if not r.get("derived")] or group
            # Deterministic when a group somehow holds several canonical rows:
            # the most-seen, then the shortest spelling.
            keep = min(canon, key=lambda r: (-r["count"], len(str(r["real"])),
                                             str(r["real"]).lower()))
            # (category, real) is this list's own primary key — the records
            # dict is keyed on it and the derived rows were deduped against the
            # same set.
            def rk(r):
                return (r["category"], str(r["real"]).lower())
            # Status on the OWNER row answers for the whole group: the fake
            # reached the export however it was spelled, and the owner is the
            # row that reverses it. Otherwise it reads "no match" — a binding
            # nothing ever used, which the macro rightly discounts — whenever
            # the export happened to carry only a wrap-split or misspelt form.
            # Occurrences stays per-row (each says how many arrived spelled
            # THAT way) so a re-run can keep loading it without inflating it.
            owner_count[rk(keep)] = sum(r["count"] for r in group)
            for r in group:
                if r is not keep:
                    alt_rows.add(rk(r))
                    nalt += 1
        if nalt:
            log.info(f"  Pseudonymize: {nalt} key row(s) marked "
                     f"'{_PN_KEY_ALT_STATUS}' (a synthetic spelling of another "
                     f"row's value — forward-only, not reversible)")

        # People (and their bare tokens / short forms) sort to the TOP of the
        # key, entities next, every other class after in name order — the reader
        # scans party names first, not an alphabetical run that buries them
        # between "address" and "phone".
        rank = {"person": 0, "person-token": 1, "display-name": 2,
                "short-name": 3, "entity": 4, "entity-token": 5}
        keyrows.sort(key=lambda r: (rank.get(r["category"], 50),
                                    r["category"], str(r["real"]).lower()))
        headers = ["Category", "Real Value", "Replacement", "Status", "Source",
                   "Occurrences"]

        def row_status(r):
            key = (r["category"], str(r["real"]).lower())
            if key in alt_rows:
                return _PN_KEY_ALT_STATUS
            if owner_count.get(key):
                return "leaked" if r["real"].lower() in self.leaked else "replaced"
            return self._status(r)


        # An audit spreadsheet from an earlier version of this tool would sit
        # beside the key looking current; remove it so it can't mislead.
        audit_base = (re.sub(r"[ _-]?key$", "", path.stem, flags=re.IGNORECASE)
                      or path.stem)
        for stale in (path.with_name(f"{audit_base} audit{path.suffix}"),
                      path.with_name(f"{audit_base} audit.json")):
            try:
                if stale.exists():
                    stale.unlink()
            except OSError:
                pass

        try:
            import openpyxl
        except ImportError:
            import json
            kp = path.with_suffix(".json")
            kp.write_text(json.dumps(
                {"mappings": [{"category": r["category"], "real": r["real"],
                               "replacement": r["fake"],
                               "status": row_status(r),
                               "source": r["source"],
                               "occurrences": r["count"]}
                              for r in keyrows]}, indent=2), encoding="utf-8")
            log.info(f"  openpyxl not installed; reversal key written as JSON: "
                     f"{kp.name} ({len(keyrows)} mapping(s))")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pseudonym Key"
        ws.append(headers)
        for r in keyrows:
            ws.append([r["category"], r["real"], r["fake"], row_status(r),
                       r["source"], r["count"]])
        wb.save(path)
        for cluster in self.alias_candidates():
            log.warning("  possible alias (same given name, different surname) "
                        "faked separately — add a --term to link: "
                        + " | ".join(cluster))
        log.info(f"  Reversal key written: {path.name} "
                 f"({len(keyrows)} mapping(s))")


# ── Partial (prefix) entity terms ───────────────────────────────────────────
# A column-splice or a bad wrap can leave only the FRONT of an entity name in
# the text ("DREAM TEAM REAL ESTATE" with the opposing caption welded on where
# "CONSULTANTS, INC." belongs). A prefix term catches that. It is dangerous by
# nature: the front of a caption name is where the generic words live, and a
# careless rule rewrites "real estate broker" or "Los Angeles County" all over
# the document. Three guards keep it honest.
_PN_PREFIX_MIN_WORDS = 3        # never a bigram: "Real Estate" is not a name
_PN_PREFIX_MIN_CORE = 4         # only names with room to spare


def _pn_core_indices(name):
    """Indices of the distinctive (non-connector, non-suffix) words of `name`."""
    toks = name.split()
    return [i for i, t in enumerate(toks)
            if re.search(r"[A-Za-z0-9]", t) and not _pn_is_entity_keep(_pn_word_base(t))]


def _pn_prefix_is_safe(prefix, next_word, corpus):
    """True when every occurrence of `prefix` in `corpus` is either part of the
    full name (the name's next word follows) or written in ALL CAPS (a caption
    fragment). A single lowercase or title-case occurrence standing on its own
    means the prefix is ordinary English and must not become a term.

    This is the guard that separates "Dream Team Real" (every loose occurrence
    is the defendant, in a caption) from "Los Angeles County" (which turns up in
    prose) — empirically, per case, with no word list to maintain."""
    pat = _pn_build_pattern(prefix, whole_word=True)
    nxt = re.compile(r"\s+" + re.escape(next_word) + r"(?!\w)", re.IGNORECASE)
    loose = 0
    for m in re.finditer(pat, corpus, re.IGNORECASE):
        if nxt.match(corpus, m.end()):
            continue            # just the full name; the full term already wins
        if not m.group(0).isupper():
            return False        # ordinary prose — reject the prefix outright
        loose += 1
    return loose > 0            # a prefix that never stands alone adds nothing


def _pn_register_prefix_terms(pseudonymizer, corpus, log):
    """Register the SHORTEST safe prefix of each long entity name. Shortest,
    because a shorter prefix catches more truncations; safe, because
    `_pn_prefix_is_safe` has to clear it against the whole corpus first."""
    added = []
    for t in list(pseudonymizer.terms):
        if t.category != "entity":
            continue
        toks = t.real.split()
        core = _pn_core_indices(t.real)
        if len(core) < _PN_PREFIX_MIN_CORE:
            continue
        for k in range(_PN_PREFIX_MIN_WORDS, len(core)):
            prefix = re.sub(r"[.,;]+$", "", " ".join(toks[:core[k - 1] + 1])).strip()
            next_word = _pn_word_affixes(toks[core[k]])[1]
            if not prefix or not next_word:
                continue
            if not _pn_prefix_is_safe(prefix, next_word, corpus):
                continue
            new = _PnTerm("entity-prefix", prefix,
                          _pn_fake_entity(prefix, pseudonymizer.registry),
                          whole_word=True, case_sensitive=False, priority=0,
                          source="prefix")
            if pseudonymizer._add_terms([new]):
                added.append(prefix)
            break               # shortest safe prefix only
    if added:
        log.info(f"  Pseudonymize: partial-name matching added "
                 f"{len(added)} prefix term(s): {', '.join(added)}")


# ────────────────────────────────────────────────────────────────────────────
# Plain-text export (companion .txt for lightweight upload)
# ────────────────────────────────────────────────────────────────────────────
# Write a UTF-8 text file next to each PDF that has a real text layer, so a
# text-only copy can be uploaded (e.g. to Claude) instead of page images.
# Scanned-only PDFs — no native text on any page — are skipped: a text export
# of them would be empty (or, if OCR ran, lower-fidelity), and it wouldn't save
# any "image space" because their content only exists as images.
#
# The text-vs-scanned decision is made from the NATIVE text (before OCR adds a
# layer). It extracts whenever the document OPENS with real native text — the
# common case for an e-filed brief or declaration — regardless of how many
# scanned exhibit images follow. If the head isn't native (e.g. the file opens
# on a scanned cover), it falls back to a random sample of the rest, so a text
# body that starts a few pages in is still caught. A scanned-only PDF matches
# neither test and is skipped.
_TEXT_PAGE_MIN_WORDS = 8    # a page with this many words counts as "has text"
# A head page needs more than this to count as a native document body: enough
# to clear a native e-filing stamp ("Electronically FILED by ...") that can sit
# on an otherwise-scanned page (~15-25 words). A genuine opening page — a
# caption block, or ~28 pleading line-numbers plus prose — clears it easily.
_HEAD_MIN_WORDS = 40


def _sample_page_indices(n_pages: int):
    """Page indices to inspect for the text-layer decision: the first few
    pages plus a small random sample of the remainder. Seeded by page count
    so the decision is reproducible from run to run."""
    import random
    head = list(range(min(3, n_pages)))
    rest = list(range(len(head), n_pages))
    if rest:
        head += random.Random(n_pages).sample(rest, min(4, len(rest)))
    return sorted(set(head))


def _pdf_has_text_layer(doc, log: logging.Logger) -> bool:
    """Decide from the native text (call BEFORE OCR) whether a PDF is
    text-based (worth exporting) or effectively all scanned images (skip).

    Primary test: the document opens with real native text — any of the first
    few pages carries >= _HEAD_MIN_WORDS words. This extracts an e-filed brief
    or declaration even when it's followed by many scanned exhibit images.

    Fallback: if the head isn't native (a scanned cover, say), sample the rest
    and extract when at least half those pages carry text. Scanned-only PDFs
    pass neither test."""
    n = doc.page_count
    if n == 0:
        return False
    head_words = max(
        (len(doc[i].get_text("text").split()) for i in range(min(3, n))),
        default=0,
    )
    if head_words >= _HEAD_MIN_WORDS:
        log.info(f"  Text-layer check: native document body "
                 f"(opening page has {head_words} words) - exporting text")
        return True
    sample = _sample_page_indices(n)
    text_pages = sum(
        1 for i in sample
        if len(doc[i].get_text("text").split()) >= _TEXT_PAGE_MIN_WORDS
    )
    frac = text_pages / len(sample)
    log.info(f"  Text-layer check: head not native (max {head_words} words); "
             f"{text_pages}/{len(sample)} sampled page(s) have text ({frac:.0%})")
    return frac >= 0.5


# ── Public verification links for the .txt export ───────────────────────────
# The PDF's citation links point at the active provider (Lexis/Westlaw), which
# needs a paid login. The .txt export instead lists each cited authority with a
# link to a FREE, publicly-searchable database, so a reader (e.g. Claude) can
# get a head start on verifying the authority without an account. All URL
# construction lives in _public_authority_url so the sources are easy to swap.
#
# Our California code abbreviations (CCP, CIV, PEN, …) are exactly the lawCode
# values the official California Legislative Information site uses, so a statute
# resolves to a direct section deep link.
_CA_LEGINFO_CODES = set(WL_SEARCH_PREFIX)


def _public_authority_url(cite, pseudonymizer=None):
    """A free, public URL for verifying a citation, or None:
      * California codes  -> official California Legislative Information (leginfo)
      * U.S.C.            -> Cornell Law School's Legal Information Institute
      * cases             -> Google Scholar case-law search
      * rules             -> Google search
    Used ONLY in the plain-text export; the PDF keeps its Lexis/Westlaw links.

    A case-search query is pseudonymized (as PLAIN text, before URL-encoding)
    when a pseudonymizer is given — otherwise the current case's own caption,
    which shows up as a slip cite, would embed the real party/case values in
    the query string where percent-encoding hides them from a later scrub.
    Real cited authorities have different names, so pseudonymization leaves
    them untouched.
    """
    from urllib.parse import quote

    def _pn(s):
        # Only this case's OWN caption (a slip cite with both sides trusted)
        # is scrubbed; a real authority's query must stay verbatim, or the
        # verification link points at a nonexistent case.
        return (pseudonymizer.apply_to_citation(cite, s)
                if pseudonymizer is not None else s)

    key = cite.get("key", "")
    kind = cite.get("kind")
    if kind == "statute":
        m = re.match(r"^(\d+)\s+U\.S\.C\.\s*§\s*(.+)$", key)
        if m:
            return (f"https://www.law.cornell.edu/uscode/text/"
                    f"{m.group(1)}/{_bare_section(m.group(2))}")
        m = re.match(r"^([A-Z]+)\s*§\s*(.+)$", key)
        if m and m.group(1) in _CA_LEGINFO_CODES:
            return ("https://leginfo.legislature.ca.gov/faces/"
                    "codes_displaySection.xhtml?lawCode="
                    f"{m.group(1)}&sectionNum={_bare_section(m.group(2))}.")
        return None
    if kind == "case":
        return "https://scholar.google.com/scholar?q=" + quote(
            _pn(_disambiguated_lexis_term(key)))
    if kind == "rule":
        return "https://www.google.com/search?q=" + quote(_pn(key))
    return None


def _build_authorities_appendix(full_text, pseudonymizer=None):
    """Return an appendix block listing each UNIQUE cited authority (in order
    of first appearance) with a public verification URL, or '' if none. When a
    pseudonymizer is given, the displayed cite text and the search query are
    pseudonymized as plain text (so the current case's own caption is scrubbed
    without the percent-encoding boundary issue that would sneak it through)."""
    seen = {}
    for c in find_all_citations(full_text):
        if c["key"] in seen:
            continue
        url = _public_authority_url(c, pseudonymizer)
        if url:
            disp = (pseudonymizer.apply_to_citation(c, c["key"])
                    if pseudonymizer is not None else c["key"])
            seen[c["key"]] = (disp, url)
    if not seen:
        return ""
    lines = ["====== Authorities cited (public verification links) ======"]
    lines += [f"{disp}  ->  {url}" for disp, url in seen.values()]
    return "\n".join(lines)


def _page_lined_rows(page):
    """For a pleading-paper page, return [(line_num, segments), ...] in reading
    order, so the .txt can mirror the PDF's line numbering. `segments` is
    [(x0, text), ...] — the row split at column gutters. `line_num` is the
    gutter number for the first row under it and None for a continuation row (a
    dense letterhead/caption stacks several rows under one number). Returns None
    for pages without a recognisable line-number column (exhibits, covers,
    signature pages), which the caller renders as plain text.

    Returning the rows (rather than a joined string) lets the caller
    pseudonymize each line's BODY on its own — so the line-number prefix is
    never seen by a detector (e.g. the address regex would otherwise read the
    printed line number as a street number). Returning the SEGMENTS lets it go
    one better and pseudonymize each column on its own."""
    anchors = _detect_line_anchors(page)
    if not anchors:
        return None

    def _strip_markers(rows):
        """Drop legacy right-margin markers from row segments. The flowing-
        text export path strips them via _MARKER_DETECT_RE, but the pleading-
        rows path bypassed that, so a PDF stamped by an older tool version
        exported "[Smith|p3:7¶4]" verbatim — garbage that also carries the
        party short name into a pseudonymized .txt."""
        out = []
        for num, segs in rows:
            cleaned = [(x0, _MARKER_DETECT_RE.sub("", t)) for x0, t in segs]
            out.append((num, [(x0, t) for x0, t in cleaned if t.strip()]
                        or cleaned[:1]))
        return out

    rows = _strip_markers([(a["line_num"], a["segments"])
                           for a in sorted(anchors, key=lambda a: a["y_mid"])])
    # Normalize before detect: when the ordinary extraction shows splice
    # corruption, retry from character geometry — a welded span splits at its
    # column jumps and the party name comes out contiguous, so the ordinary
    # term patterns can replace it. Adopted only when the cure actually clears
    # the splice signature; otherwise the original rows stand and the
    # reduced-span scrub (scrub_welded) remains the net beneath.
    if _page_looks_spliced(rows):
        try:
            cured = _detect_line_anchors(page, desplice=True)
        except Exception:
            cured = None
        if cured:
            crows = _strip_markers([(a["line_num"], a["segments"])
                                    for a in sorted(cured, key=lambda a: a["y_mid"])])
            if not _page_looks_spliced(crows):
                return crows
    return rows


_COLUMN_BAND_TOL = 30.0   # pt; segment left edges this close share a page column
                          # (tuned on one caption layout; override in the config
                          # via `column_band_tol` when a layout splices)


# A caption whose two columns interleaved WITHIN a row segment produces
# tell-tale garbage: long spaceless tokens, a lower->upper transition inside a
# token ("SCHILLECIJoplin"), or an @ glued to an upper-case run. When a page
# looks like this, term matching is unreliable and the page must not be
# certified clean on the whole-word patterns alone.
_PN_SPLICE_LONG_TOKEN = 25
_PN_SPLICE_CASEFLIP_RE = re.compile(r"[a-z][A-Z]")
# The mirror flip: an upper-case RUN welded to a lower-case word ("SERVICEoffices",
# "TRANSMISSION1010"). Two lower-case letters are required so an ordinary
# acronym plural ("ADUs", "PDFs") is not mistaken for a splice.
_PN_SPLICE_UPPERRUN_RE = re.compile(r"[A-Z]{3,}[a-z]{2,}")
# A DANGLING "@" — one at the start of a token, i.e. preceded by whitespace or by
# nothing. A real address never begins with it; a spliced caption does, because
# the local part was carried off into the other column ("JPTAttorney For
# Defendant AZUL CONCRETO, INC.  @SCHILLECITORTORICILAW.COM"). The old rule fired
# on `[A-Z]{8,}@`, which is what a perfectly ordinary all-caps letterhead address
# looks like — every Motion page was reported as corrupt.
_PN_SPLICE_ATRUN_RE = re.compile(r"(?:^|\s)@[A-Za-z]")
# Punctuation used as a rule/leader line: "______", "------", "......".
_PN_SPLICE_RULE_RE = re.compile(r"^(.)\1{4,}$")


def _page_looks_spliced(rows):
    """True when a page's extracted rows show column-splice corruption."""
    if not rows:
        return False
    text = " ".join(t for _num, segs in rows for _x, t in segs)
    if _PN_SPLICE_ATRUN_RE.search(text):
        return True
    for tok in text.split():
        if _PN_SPLICE_RULE_RE.match(tok):
            continue                      # a signature rule, not a long word
        # A long token is only suspicious when it is a long WORD. E-mails, URLs
        # and hyphenated compounds are legitimately long.
        if (len(tok) > _PN_SPLICE_LONG_TOKEN
                and not re.search(r"[-@._/]", tok)):
            return True
        if _PN_SPLICE_UPPERRUN_RE.search(tok):
            return True
        # A surname particle before an all-caps body is a NAME, not a splice:
        # all-caps captions legitimately print "SEAN McDONALD" and "DiGIORNO
        # FOODS" — the case flip is the particle's own convention. Without
        # this every such page drew a "column-spliced, verify by hand" warning
        # and an unnecessary welded-scrub pass.
        if re.match(r"(?:Mc|Mac|Di|De|Da|La|Le|Lo|Fitz|Van|Von|St\.?|"
                    r"O['’]|D['’])[A-Z][A-Za-z'’.,-]*$", tok):
            continue
        # A lower->upper flip inside a token that isn't ordinary CamelCase
        # (needs an upper run of >=3, e.g. "...JOPLIN", not "McKay").
        for m in _PN_SPLICE_CASEFLIP_RE.finditer(tok):
            if re.match(r"[A-Z]{3,}", tok[m.start() + 1:]):
                return True
    return False


def _page_column_bands(rows):
    """Assign each row segment to a page-level column band, keyed on its left
    edge. Returns [band_index_per_segment_per_row], bands ordered left to
    right."""
    xs = sorted({round(x) for _num, segs in rows for x, _t in segs})
    bands = []
    for x in xs:
        if bands and x - bands[-1][-1] <= _COLUMN_BAND_TOL:
            bands[-1].append(x)
        else:
            bands.append([x])
    edges = [b[0] for b in bands]

    def band_of(x):
        for i in range(len(edges) - 1, -1, -1):
            if x >= edges[i] - _COLUMN_BAND_TOL:
                return i
        return 0
    return [[band_of(x) for x, _t in segs] for _num, segs in rows]


def _page_column_streams(rows):
    """[(band, [(row_i, seg_i), ...]), ...] — each page column, top to bottom."""
    from collections import defaultdict
    bands = _page_column_bands(rows)
    streams = defaultdict(list)
    for i, (_num, segs) in enumerate(rows):
        for j in range(len(segs)):
            streams[bands[i][j]].append((i, j))
    return sorted(streams.items())


# ── Caption reconstruction ───────────────────────────────────────────────────
# The caption's party block is where nearly every hard extraction failure has
# lived (wraps, splices, welds, OCR variants) — and it is also the one region
# whose content is KNOWN: parties, roles, DOES boilerplate. So instead of
# repairing its extraction, the block is cut out and RE-RENDERED from the
# tracked parties' fakes. Fail-safe by construction: every word of the block
# must be attributed to a known party, a role label, or caption boilerplate,
# or the page falls back to the ordinary pipeline untouched — reconstruction
# can never silently drop a party it didn't recognise.

# Words that legitimately appear in a party block without naming anyone: the
# capacity/descriptor vocabulary of a California caption.
_PN_CAPTION_VOCAB = frozenset("""
a an and the of on behalf all others similarly situated individual
individuals corporation company entity entities form unknown business
capacity partnership partnerships limited liability california delaware
nevada domestic foreign profit nonprofit public benefit municipal
association organization trust joint venture does doe roes roe inclusive to
through thru et al sued herein as related cross actions action consolidated
successor in interest trustee guardian conservator estate deceased minor by
his her their its dba aka fka
""".split())

_PN_CAPTION_VS_RE = re.compile(r"(?i)^\s*vs?\.?\s*,?\s*$")
_PN_CAPTION_PROLE_RE = re.compile(
    r"(?i)^\s*(?:plaintiffs?|petitioners?|cross-complainants?)\s*[.,;]?\s*$")
_PN_CAPTION_DROLE_RE = re.compile(
    r"(?i)^\s*(?:defendants?|respondents?|cross-defendants?)\s*[.,;]?\s*$")
# A line that is page FURNITURE, never a party-name row: the attorney block's
# contact lines (bar number, address, phone, e-mail), an "Attorneys for" tag,
# the court/county heading, or a label ending in a colon. The walk-back from
# the plaintiff-role row stops before these, so incidental content above the
# caption box can't leak into the party block and spoil attribution.
_PN_CAPTION_STOP_RE = re.compile(
    r"(?i)\d|@|:\s*$|\battorneys?\s+(?:for|at)\b|\bcourt\b|\bcounty of\b|"
    r"\btelephone\b|\btel\b|\bfax\b|\bfacsimile\b|\be-?mail\b|\bsuite\b|"
    r"\bfloor\b|\besq\b|\blaw\s+offices?\b")


def _pn_reconstruct_caption(pseudonymizer, rows):
    """Rebuild a page's caption party block from the tracked parties' fakes.

    Detects the classic layout in the leftmost column band — plaintiff names,
    a role row ("Plaintiff,"), a "vs." row, defendant names, a role row
    ("Defendants.") — beside a second column (case number / title). The name
    rows are replaced by lines rendered from the matched parties' fakes (plus
    any DOES phrase, kept verbatim); role and "vs." rows are kept as printed.
    Row count, line numbers, and every other column are untouched, so pinpoint
    anchoring is unaffected and the right column flows through the ordinary
    scrub.

    Returns the rebuilt rows, or None when anything stops it: no such block,
    no second column beside it, a side that matches no tracked party, or —
    the fail-safe — ANY word in the name rows that is not attributable to a
    tracked party, a role label, or caption boilerplate. On None the caller
    proceeds exactly as before, so reconstruction can only ever replace text
    it fully understands."""
    if not rows:
        return None
    parties = []
    for (cat, _rl), rec in pseudonymizer.records.items():
        if cat not in ("person", "entity"):
            continue
        pb = {_pn_word_base(w) for w in rec["real"].split()
              if len(_pn_word_base(w)) >= 2
              and not _pn_is_entity_keep(_pn_word_base(w))}
        if pb:
            parties.append((rec, pb))
    if not parties:
        return None
    all_party_bases = set().union(*(pb for _r, pb in parties))

    bands = _page_column_bands(rows)
    b0 = []                                   # (row_i, seg_j, text) in band 0
    for i, (_num, segs) in enumerate(rows):
        for j in range(len(segs)):
            if bands[i][j] == 0:
                b0.append((i, j, segs[j][1]))
                break
    texts = [t for _i, _j, t in b0]

    def _find(rx, lo):
        for k in range(lo, len(texts)):
            if rx.match(texts[k]):
                return k
        return -1
    p_role = _find(_PN_CAPTION_PROLE_RE, 0)
    vs_k = _find(_PN_CAPTION_VS_RE, p_role + 1) if p_role >= 0 else -1
    d_role = _find(_PN_CAPTION_DROLE_RE, vs_k + 1) if vs_k >= 0 else -1
    if p_role < 0 or vs_k < 0 or d_role < 0 or d_role - p_role > 40:
        return None

    # Walk back from the plaintiff-role row to the top of the name run,
    # stopping before any furniture line (attorney contacts, court heading).
    start = p_role
    while (start > 0 and texts[start - 1].strip()
           and not _PN_CAPTION_PROLE_RE.match(texts[start - 1])
           and not _PN_CAPTION_VS_RE.match(texts[start - 1])
           and not _PN_CAPTION_STOP_RE.search(texts[start - 1])):
        start -= 1
    if start == p_role:
        return None                           # no plaintiff name rows found

    # Two-column requirement: the block must sit beside another column.
    rng = {i for i, _j, _t in b0[start:d_role + 1]}
    if not any(bands[i][j] >= 1
               for i in rng for j in range(len(rows[i][1]))):
        return None

    p_text = " ".join(texts[start:p_role])
    d_text = " ".join(texts[vs_k + 1:d_role])

    # Fail-safe attribution, scoped to what could actually BE a party: only a
    # DISTINCTIVE CAPITALIZED word not attributable to a tracked party blocks
    # reconstruction (an unlisted cross-defendant, welded junk). Everything
    # else is incidental and must not stop the rebuild: lowercase descriptors
    # ("an individual, on behalf of himself"), common/boilerplate words,
    # street/contact vocabulary, localities, and digit tokens are all
    # irrelevant to who the parties are.
    for w in (p_text + " " + d_text).split():
        base = _pn_word_base(w)
        if (not base or len(base) == 1 or not re.search(r"[a-z]", base)
                or not w[:1].isupper()
                or base in _PN_CAPTION_VOCAB
                or base in _PN_PARTY_ROLE_WORDS
                or base in _PN_COMMON_WORDS
                or base in _PN_REVIEW_NAME_STOP
                or _pn_is_entity_keep(base)
                or base in all_party_bases
                or _pn_is_protected_locality(base)):
            continue
        return None

    def _side_parties(txt):
        ws = {_pn_word_base(w) for w in txt.split()}
        hits = sorted((r for r, pb in parties if pb <= ws),
                      key=lambda r: -len(r["real"]))
        out, covered = [], set()
        for r in hits:
            pb = next(pb for rr, pb in parties if rr is r)
            if not pb <= covered:             # skip a variant already covered
                covered |= pb
                out.append(r)
        return out
    ps, ds = _side_parties(p_text), _side_parties(d_text)
    if not ps or not ds:
        return None

    def _cased(rec, txt):
        # Match how the party's OWN words are printed (captions are usually
        # ALL CAPS while the descriptors around them are not).
        pb = next(pb for rr, pb in parties if rr is rec)
        own = [w for w in txt.split() if _pn_word_base(w) in pb]
        joined = "".join(own)
        alpha = [c for c in joined if c.isalpha()]
        upper = sum(c.isupper() for c in alpha)
        return (rec["fake"].upper()
                if alpha and upper > 0.8 * len(alpha) else rec["fake"])
    m = re.search(r"(?i)\bdoes?\s+\d+\s*(?:through|to|thru|[-–])\s*\d+"
                  r"(?:,?\s*inclusive)?", d_text)
    p_line = "; ".join(_cased(r, p_text) for r in ps) + ","
    d_line = "; ".join(_cased(r, d_text) for r in ds)
    if m:
        d_line += f"; and {m.group(0)}"
    d_line += ","

    # Rebuild: the first name row of each side carries the rendered line, the
    # rest go blank; role and "vs." rows are kept exactly as printed.
    new_text = {}
    for k in range(start, p_role):
        new_text[k] = p_line if k == start else ""
    for k in range(vs_k + 1, d_role):
        new_text[k] = d_line if k == vs_k + 1 else ""
    out_rows = []
    replace = {(b0[k][0], b0[k][1]): new_text[k] for k in new_text}
    for i, (num, segs) in enumerate(rows):
        nsegs = []
        for j, (x, t) in enumerate(segs):
            if (i, j) in replace:
                t = replace[(i, j)]
                if not t:
                    continue                  # blank filler row: drop the seg
            nsegs.append((x, t))
        out_rows.append((num, nsegs))
    for r in ps + ds:
        # Record the replacement so the reversal key carries the mapping.
        r["count"] += 1
    return out_rows


def _pn_apply_page_rows(pseudonymizer, rows):
    """Pseudonymize a pleading page column by column and return one display
    string per row.

    The caption party block, when confidently detected and fully attributed,
    is RE-RENDERED from the tracked parties' fakes first (see
    _pn_reconstruct_caption) — extraction pathologies in that block (wraps,
    splices, OCR variants) then never reach the matching layer at all.

    A party name that the printed page wrapped mid-name is only contiguous
    within its own column; read straight across the page it is interrupted by
    whatever the other column holds on that line. Running each column as its own
    line stream restores the contiguity `apply_lines` needs, then the row is
    re-assembled left to right exactly as it reads on paper."""
    rebuilt = _pn_reconstruct_caption(pseudonymizer, rows)
    if rebuilt is not None:
        rows = rebuilt
    faked = {}
    for _band, items in _page_column_streams(rows):
        bodies = [rows[i][1][j][1] for i, j in items]
        for (i, j), text in zip(items, pseudonymizer.apply_lines(bodies)):
            faked[(i, j)] = text
    return [" ".join(faked[(i, j)] for j in range(len(segs))).strip()
            for i, (_num, segs) in enumerate(rows)]


def _page_detect_text(page):
    """Column-ordered plain text for a page, for NAME DETECTION and leak
    checking only. On a two-column caption the on-paper reading order splices
    the columns together; reading each column top to bottom instead makes a
    wrapped party name contiguous, which is what the dba scanner, the declarant
    scanner, and `surviving_reals` need to see."""
    rows = _page_lined_rows(page)
    if not rows:
        return _MARKER_DETECT_RE.sub("", page.get_text("text"))
    out = []
    for _band, items in _page_column_streams(rows):
        out.append("\n".join(rows[i][1][j][1] for i, j in items))
    return "\n".join(out)


def _pn_scrubbed_stem(stem: str, pseudonymizer, log):
    """`stem` pseudonymized for use as a shareable filename. Separators
    (_ / -) are normalized to spaces so name tokens are word-bounded, the stem
    is run through the pseudonymizer, and if a tracked real value still
    survives the result falls back to a neutral, non-revealing name
    (<fake-case-no> <hash> or document <hash>)."""
    spaced = re.sub(r"[_\-]+", " ", stem).strip()
    faked = pseudonymizer.apply(spaced)
    digest = _pn_hashlib.sha256(stem.encode("utf-8")).hexdigest()[:6]
    if pseudonymizer.surviving_reals(faked):
        prefix = pseudonymizer.a_case_number_fake() or "document"
        safe = f"{prefix} {digest}"
        log.info(f"  Filename still held a real value after scrub; using "
                 f"neutral name {safe}.txt")
    else:
        safe = re.sub(r"\s+", " ", faked).strip() or f"document {digest}"
    return safe


def _pseudonymized_txt_path(out_dir: Path, pdf_path: Path, pseudonymizer, log):
    """Return the .txt output path (inside `out_dir`) with the source PDF's
    stem pseudonymized (see _pn_scrubbed_stem), so a party/attorney name in
    the filename doesn't ride along on the shareable text file.

    COLLISION-SAFE: two different sources can scrub to one name ("Smith
    Decl.pdf" and "Smith_Decl.pdf" both -> "Bennett Decl"), and the second
    export used to silently overwrite the first. A per-run registry on the
    pseudonymizer maps each claimed name to its source stem; a different
    source claiming a taken name gets its (deterministic, per-source) digest
    appended instead."""
    stem = pdf_path.stem
    safe = _pn_scrubbed_stem(stem, pseudonymizer, log)
    names = getattr(pseudonymizer, "_export_names", None)
    if names is None:
        names = pseudonymizer._export_names = {}
    owner = names.setdefault(safe.lower(), stem)
    if owner != stem:
        digest = _pn_hashlib.sha256(stem.encode("utf-8")).hexdigest()[:6]
        safe = f"{safe} {digest}"
        names.setdefault(safe.lower(), stem)
        log.info(f"  Export name collision with '{owner}'; using "
                 f"{safe}.txt for {stem}")
    return out_dir / (safe + ".txt")


# ── Leak-triage worksheet ────────────────────────────────────────────────────
# Every potential leak the tool flags (a real value that survived, plus the
# open-world REVIEW / REID / unknown-name findings) is collected across the
# whole run into one spreadsheet, EACH ROW LOCATED to the printed page and
# gutter line so it can be found in seconds, with a blank "Fix?" column for the
# reviewer to triage. Written next to pdf_linker.log; removed on a clean run.
_PN_PAGE_HEADER_RE = re.compile(
    r"^====== Page (\d+)(?: \(printed p\. (.+?)\))? ======$")
_PN_GUTTER_RE = re.compile(r"^\s*(\d+)\s{2}\S")


def _pn_body_lines(body):
    """Parse an assembled .txt body into [(page_label, line_label, text), ...],
    tracking the current page (from '====== Page N ======' headers, preferring
    the printed page number) and the current gutter line number, so a finding
    can be reported as 'p.8:16'. A continuation row keeps the last gutter
    number; the authorities appendix is labelled 'appendix'."""
    page, gutter = "?", None
    out = []
    for raw in body.split("\n"):
        mh = _PN_PAGE_HEADER_RE.match(raw)
        if mh:
            page, gutter = (mh.group(2) or mh.group(1)), None
            continue
        if raw.startswith("====== Authorities"):
            page, gutter = "appendix", None
            continue
        mg = _PN_GUTTER_RE.match(raw)
        if mg:
            gutter = mg.group(1)
        out.append((page, gutter, raw))
    return out


def _pn_locate(parsed, needle, limit=12):
    """Human-readable 'p.<page>:<line>' locations of `needle` in a parsed body
    (case-insensitive substring), de-duplicated, capped. '(not located)' when a
    welded/reduced finding has no clean line to point at."""
    nl = needle.lower()
    seen, out = set(), []
    for page, gutter, text in parsed:
        if nl in text.lower():
            loc = f"p.{page}:{gutter}" if gutter else f"p.{page}"
            if loc not in seen:
                seen.add(loc)
                out.append(loc)
    if not out:
        return "(not located)"
    return ", ".join(out[:limit]) + (" …" if len(out) > limit else "")


# Worksheet column order. The flagged value leads and its Fix? decision sits
# right beside it — the two columns a reviewer actually works — with the
# locating detail (which file, what kind, where) trailing. Each header is
# paired with the row-dict key it renders, so a reorder is a one-line change
# here and the writer, widths and the Fix? dropdown all follow. The reader is
# header-name driven, so column order never affects round-tripping.
_PN_LEAK_COLUMNS = (
    ("Value", "value", 34),
    ("Fix? (yes/no)", "fixcell", 12),
    ("File", "file", 20),
    ("Type", "type", 22),
    ("Where (page:line)", "where", 30),
    ("Notes", "notes", 24),
)
_PN_LEAK_HEADERS = tuple(h for h, _k, _w in _PN_LEAK_COLUMNS)
_PN_LEAK_ABSENT = "(no longer present)"

# Sheet title of the per-folder genuine-leak triage.
_PN_LEAK_SHEET = "LEAKS"

# The review worksheet's filename (stem). Renamed from the historical
# "pdf_linker_leaks" to a plain "LEAKS"; the old name is still READ so a
# folder triaged under the previous version keeps its decisions.
_PN_LEAK_STEM = "LEAKS"
_PN_LEAK_LEGACY_STEMS = ("pdf_linker_leaks",)


def _pn_leak_xlsx_path(folder):
    """The worksheet path to WRITE: the current name."""
    return folder / f"{_PN_LEAK_STEM}.xlsx"


def _pn_leak_txt_path(folder):
    """The plaintext-fallback checklist path to WRITE."""
    return folder / f"{_PN_LEAK_STEM}.txt"


def _pn_existing_leak_xlsx(folder):
    """The worksheet path to READ: the current name if present, else a legacy
    name a prior run left behind (so its Fix? decisions still round-trip)."""
    current = _pn_leak_xlsx_path(folder)
    if current.is_file():
        return current
    for stem in _PN_LEAK_LEGACY_STEMS:
        legacy = folder / f"{stem}.xlsx"
        if legacy.is_file():
            return legacy
    return current


def _pn_bracket_keep(value, cell):
    """Parse a Fix? cell that BRACKETS the part(s) of `value` to KEEP verbatim,
    faking only the rest — the operator's way of saying "this fragment isn't a
    name, but the remainder is". `[Human Resources] Information` on the leak
    "Raytheon's Human Resources Information" keeps "Human Resources" and auto-
    fakes the rest.

    Returns the list of value fragments that DO need faking (the value with each
    bracketed keep-part cut out), or None when the cell is not a bracket spec or
    a bracketed part is not actually a substring of the value (so the caller
    falls back to treating the whole cell as an explicit replacement). An empty
    list means the operator bracketed the entire value — keep all of it."""
    keeps = [k.strip() for k in re.findall(r"\[([^\[\]]*)\]", cell)]
    keeps = [k for k in keeps if k]
    if not keeps:
        return None                       # no brackets — not this rule
    work = value
    for k in keeps:
        i = work.lower().find(k.lower())
        if i < 0:
            return None                   # bracketed text isn't part of the value
        work = work[:i] + "\x00" + work[i + len(k):]
    frags = [f.strip(" \t,.;:'’\"-–—") for f in work.split("\x00")]
    return [f for f in frags if re.search(r"[A-Za-z0-9]", f)]


def _pn_decision_keep_parts(d):
    """The substrings a KEEP/KEEP-PART decision protects verbatim: the whole
    value for a `no`, or the bracketed part(s) for a keep-spec. Shared by the
    LEAKS worksheet and the pseudonym key so both feed `Pseudonymizer.keep_values`
    identically. Returns [] for any other decision (yes / explicit / blank)."""
    if d.get("fix") == "no":
        return [d["value"]]
    if d.get("fake_values") is not None and d.get("fixcell"):
        keeps = [k.strip() for k in re.findall(r"\[([^\[\]]*)\]", str(d["fixcell"]))]
        return [k for k in keeps if k]
    return []


def _pn_decision_is_keep(d):
    """True when a decision is a KEEP: `no` (leave verbatim) or a bracketed
    keep-spec (fake only the non-bracketed fragment(s)). These are the durable
    decisions that live in the worksheet's KEEP sheet and drive re-runs; every
    other decision (yes / explicit replacement / undecided) belongs to the
    transient genuine-leak triage."""
    return d.get("fix") == "no" or d.get("fake_values") is not None


def _pn_parse_decision_rows(rows):
    """Parse the (values_only) rows of a decision sheet — LEAKS or KEEP — into
    {value_lower: {value, type, fix, replacement, fake_values, fixcell, notes}}.
    `fix` is normalised to 'yes'/'no'/''. A Fix? entry that is NOT a reserved
    control word ('yes'/'no' and the bare y/n shorthands) is one of:
      * BRACKETED text that is part of the value — keep the bracketed part
        verbatim, auto-fake the rest (`fake_values` holds the fragment(s)); or
      * anything else — an explicit typed replacement (`replacement`)."""
    if not rows:
        return {}
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]

    def idx(name):
        return hdr.index(name) if name in hdr else None
    ci = {k: idx(k) for k in ("type", "value", "fix? (yes/no)", "notes",
                             "cases", "origin")}
    out = {}
    for r in rows[1:]:
        def cell(k):
            i = ci.get(k)
            return r[i] if i is not None and i < len(r) else None
        val = cell("value")
        if val in (None, ""):
            continue
        val = str(val)
        raw = str(cell("fix? (yes/no)") or "").strip()
        low = raw.lower()
        replacement, fake_values, fixcell = None, None, None
        if low in ("yes", "y"):
            fix, replacement = "yes", None
        elif low in ("no", "n"):
            fix, replacement = "no", None
        elif low == "":
            fix, replacement = "", None
        else:
            frags = _pn_bracket_keep(val, raw)
            fixcell = raw                 # echo the operator's text back verbatim
            if frags is None:
                fix, replacement = "yes", raw           # explicit replacement
            elif frags:
                fix, fake_values = "yes", frags          # keep bracketed, fake rest
            else:
                fix = "no"                # whole value bracketed -> keep all of it
        out[val.lower()] = {"value": val, "type": str(cell("type") or "").strip(),
                            "fix": fix, "replacement": replacement,
                            "fake_values": fake_values, "fixcell": fixcell,
                            "notes": str(cell("notes") or "").strip(),
                            # Master KEEP only: which case folders this decision
                            # came from, so a run can tell its OWN past decision
                            # from another matter's (see `_pn_decision_is_ours`).
                            "cases": str(cell("cases") or "").strip(),
                            "origin": str(cell("origin") or "").strip()}
    return out


def _pn_decision_is_ours(d, folder_name):
    """True when `d` was made in THIS case folder.

    The local `LEAKS.xlsx` is consumed once its rows are resolved, so a
    decision the operator made here survives only on the cross-folder master
    KEEP sheet — where it would otherwise look exactly like another matter's.
    The master rows carry the case folders they came from, so a re-run can
    still recognise its own past decision and keep applying its full meaning
    (keep the bracket, fake the remainder) instead of demoting it to the
    keep-only treatment an inherited decision gets."""
    name = str(folder_name or "").strip().lower()
    if not name:
        return False
    origin = str(d.get("origin") or "").strip().lower()
    if origin:
        return name == origin
    # A sheet written before the Origin column existed: fall back to the Cases
    # list, which for a decision only ever applied in ONE folder names it.
    cases = [c.strip().lower() for c in str(d.get("cases") or "").split(";")
             if c.strip()]
    return len(cases) == 1 and cases[0] == name


def _pn_keep_values(decisions):
    """Return (strict, soft) sets of exact strings to leave verbatim, gathered
    from every KEEP decision. A `no` value is a SOFT keep — the whole value, kept
    only where the word stands alone. A bracketed keep-spec part is a STRICT keep
    — "this fragment is never a name", kept even next to names. Both are matched
    on WORD BOUNDARIES and both lose to a full party match (see
    Pseudonymizer._keep_spans), so a keep never leaves a real party un-faked."""
    strict, soft = set(), set()
    for d in decisions.values():
        if d.get("fix") == "no":
            if d.get("value"):
                soft.add(d["value"])
        elif d.get("fake_values") is not None and d.get("fixcell"):
            for k in re.findall(r"\[([^\[\]]*)\]", str(d["fixcell"])):
                k = k.strip()
                if k:
                    strict.add(k)
    return strict, soft


def _pn_prior_fake_words(terms):
    """Lower-cased words of every fake a REUSED key already handed out — the
    stand-ins already standing in the exports. Used to recognise a HALF-SCRUBBED
    leak value (see `_pn_strip_prior_fakes`)."""
    out = set()
    for t in terms:
        f = str(t.fake)
        out.add(f.lower())
        out.update(w.lower() for w in _PN_WORD_RE.findall(f))
    return out


def _pn_strip_prior_fakes(value, fakes):
    """Drop the words of `value` that are already this key's OWN fakes, keeping
    the genuinely real remainder — or None when nothing real is left.

    A flagged leak is routinely HALF-SCRUBBED: only the bare surname was bound
    ("Penuela" -> "Sable"), so the export reads "Melissa Sable" and the review
    scan rightly surfaces it — "Melissa" IS a real name still standing. But
    marking it yes used to register the whole phrase as a REAL value, which put
    a fake ("Sable") in the key's Real Value column and faked it a second time
    ("Melissa Sable" -> "Ramsey Ellery"). The reversal macro then walked a
    two-generation chain that never reached "Penuela", and every re-run grew
    another generation.

    Faking only the real remainder keeps the existing stand-in and round-trips:
    "Melissa" gets its own fake, "Sable" stays as it is, and the macro reverses
    the pair back to "Melissa Penuela"."""
    words = str(value).split()
    keep = [w for w in words if _pn_word_affixes(w)[1].lower() not in fakes]
    if len(keep) == len(words):
        return value                  # nothing of ours in it — take it as it is
    rest = " ".join(keep).strip(" \t,.;:'’\"-–—")
    return rest if re.search(r"[A-Za-z0-9]", rest) else None


def _pn_drop_prior_fakes_from_terms(values, prior_fakes, log):
    """Apply `_pn_strip_prior_fakes` across the leak-fix values a run is about to
    turn into terms, logging what changed so the operator can see why the key
    holds "Melissa" rather than the "Melissa Sable" they marked."""
    if not prior_fakes:
        return list(values)
    out, changed, dropped = [], [], []
    for v in values:
        rest = _pn_strip_prior_fakes(v, prior_fakes)
        if rest is None:
            dropped.append(v)
            continue
        if rest != v:
            changed.append(f"{v!r} -> {rest!r}")
        out.append(rest)
    if changed:
        log.info(f"  Pseudonymize: {len(changed)} flagged value(s) were already "
                 f"part-scrubbed — faking only the real remainder so a prior "
                 f"fake never becomes a Real Value: {'; '.join(changed[:6])}")
    if dropped:
        log.info(f"  Pseudonymize: {len(dropped)} flagged value(s) are entirely "
                 f"this run's own fakes — nothing left to scrub: "
                 f"{', '.join(repr(v) for v in dropped[:6])}")
    return out


def _pn_retire_kept_key_terms(terms, decisions, registry, log):
    """A KEEP decision RETIRES the pseudonym key's own binding for that value, so
    the key a re-run writes stays CLEAN — only real values and the fakes actually
    applied, never an operator control word and never a row for a value that is
    no longer faked the way the row says.

    This is needed because a loaded key row is a FULL PARTY match, and the
    full-party override in `_keep_spans` deliberately beats a keep (so a keep can
    never leave a real party in the clear). Against the operator's OWN edit that
    override misfires two ways:

      * `no` on a keyed value — the text is faked anyway, while `write_key` drops
        the row for being a keep, leaving a fake in the exports that NOTHING in
        the key reverses; and
      * a `[bracketed]` keep-spec on a keyed value — the whole value is faked,
        bracketed part and all, and the flagged phrase survives as a key row.

    Retiring the term fixes both: a `no` value is genuinely left verbatim and its
    row disappears, and a keep-spec leaves only its non-bracketed fragment
    (rebuilt as its own term by the caller) — so the key ends up holding
    "John Doe" -> "Yorke Deverell", not "John Doe's Opposition" -> "['s
    Opposition]".

    A `no` also retires the bare person-/entity-token rows DERIVED from its words
    (`write_key` emits one per name token) unless the word still belongs to
    another party this case fakes — otherwise a value marked `no` would simply be
    reassembled word by word by its own leftover token rows.

    A keep-spec's fragment INHERITS the retired fake when the two align word for
    word, by pre-seeding the registry memo that the caller's `_pn_build_terms`
    reads next: exports already circulating as "Yorke Deverell" keep saying it.

    Returns `(surviving_terms, retired_reals)` — the input list is left alone.
    `retired_reals` also tells `write_key` that a retirement happened, so a key
    whose every row was retired is REWRITTEN empty instead of left on disk still
    naming real values against fakes this run never applied."""
    keeps = {vl: d for vl, d in decisions.items() if _pn_decision_is_keep(d)}
    if not keeps:
        return terms, []
    retired, gone_words, kept_terms = [], set(), []
    for t in terms:
        d = keeps.get(str(t.real).lower())
        if d is None:
            kept_terms.append(t)
            continue
        retired.append(t)
        if d.get("fix") == "no":
            gone_words.update(w.lower()
                              for w in _PN_WORD_RE.findall(str(t.real)))
            continue
        for frag in d.get("fake_values") or []:
            for rtok, ftok in _pn_name_token_rows(frag, t.fake):
                registry._memo.setdefault(("name_or_entity", rtok.lower()), ftok)
    if not retired:
        return terms, []
    # A word another party of this case still carries is not orphaned by the
    # retirement — "Doe" stays faked for "Jane Doe" even if "John Doe" is kept.
    live = {w.lower() for t in kept_terms
            if t.category in _PN_PARTY_OVERRIDE_CATS
            for w in _PN_WORD_RE.findall(str(t.real))}
    orphan = gone_words - live
    survivors = [t for t in kept_terms
                 if not (t.category in _PN_KEY_TOKEN_CATS
                         and str(t.real).lower() in orphan)]
    log.info(f"  Pseudonym key: retired {len(retired)} binding(s) marked KEEP "
             f"({', '.join(str(t.real) for t in retired[:6])}) — the re-run key "
             f"carries the corrected value only.")
    return survivors, [str(t.real) for t in retired]


def _pn_read_leak_decisions(folder):
    """{value_lower: decision} read back from this folder's LEAKS worksheet the
    reviewer annotated on a prior run — the transient, per-case genuine-leak
    triage. (The durable `no`/bracket KEEP decisions live in the cross-folder
    master KEEP sheet, read separately by `_pn_read_master_keep`.) Empty when
    there is no readable worksheet.

    `fixcell` is the exact string to write back so the annotation round-trips."""
    xlsx = _pn_existing_leak_xlsx(folder)
    if not xlsx.is_file():
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return {}
    return _pn_parse_decision_rows(rows)


def _leak_sev(t):
    """Row severity for sorting AND for picking a merged row's type: a real
    LEAK outranks a map-inverting REID, which outranks ordinary review."""
    return 0 if t == "LEAK" else 1 if str(t).startswith("REID") else 2


def _pn_merge_where(wheres):
    """Union the per-file location strings (each a ", "-joined list from
    `_pn_locate`) into one deduped list. A real location wins over a sentinel
    ("(not located)"), and the list is capped so a value found on dozens of
    pages doesn't blow out the cell."""
    toks = []
    for w in wheres:
        for t in str(w).split(", "):
            t = t.strip()
            if t and t not in toks:
                toks.append(t)
    real = [t for t in toks if not t.startswith("(")]
    toks = real or toks
    return ", ".join(toks[:12]) + (" …" if len(toks) > 12 else "")


# ── Master workbook (cross-case, accumulates over time) ──────────────────────
# One workbook next to pdf_linker.config holds two sheets:
#   * "Master Leaks" — the tally of genuine leaks seen across matters (opt-in
#      via master_leaks = on); and
#   * "KEEP" — the DURABLE no/bracket decisions, a single persistent sheet
#      across every folder and run. It is the preservation vehicle: a value
#      marked "leave it alone" (or a bracketed keep-spec) is remembered here and
#      re-applied on every future run in any folder, accumulating Times Seen /
#      Cases / dates so the screening can be tuned from real history over time.
_PN_MASTER_HEADERS = ("Value", "Type", "Times Seen", "Cases",
                      "First Seen", "Last Seen")
_PN_MASTER_LEAK_SHEET = "Master Leaks"
_PN_MASTER_KEEP_SHEET = "KEEP"
# The KEEP sheet's Fix? column is named so _pn_parse_decision_rows can read the
# instruction ("no" or a [bracket] spec) straight back out.
_PN_MASTER_KEEP_HEADERS = ("Value", "Fix? (yes/no)", "Type", "Times Seen",
                           "Cases", "First Seen", "Last Seen", "Notes",
                           # The folder that MADE this decision. "Cases" lists
                           # every folder the keep has since protected text in,
                           # which is real history but not authorship — and
                           # only the author fakes a keep-spec's remainder.
                           "Origin")


def _pn_master_path(cfg):
    """The cross-case master workbook path (ALWAYS available — the KEEP sheet is
    the preservation vehicle, not opt-in). Config `master_leaks_path` wins, then
    the PDF_LINKER_MASTER env var (used to relocate or isolate the master), else
    it lives next to pdf_linker.config."""
    cfg = cfg or {}
    override = (cfg.get("master_leaks_path") or "").strip().strip('"')
    if override:
        return Path(override)
    env = (os.environ.get("PDF_LINKER_MASTER") or "").strip()
    if env:
        return Path(env)
    return _config_path().with_name("master_leaks.xlsx")


def _pn_master_leaks_path(cfg):
    """Path to the master workbook for the genuine-leak TALLY, or None when that
    tally is disabled (`master_leaks = off` and no `master_leaks_path`). The KEEP
    sheet in the same file is maintained regardless (see _pn_master_path)."""
    cfg = cfg or {}
    override = (cfg.get("master_leaks_path") or "").strip().strip('"')
    if not _config_bool(cfg, "master_leaks", False) and not override:
        return None
    return _pn_master_path(cfg)


def _pn_master_load(master_path):
    """Load the master workbook for update (all sheets preserved), or a fresh
    empty one. Normal (not read-only) mode so a save keeps every OTHER sheet."""
    import openpyxl
    if master_path.exists():
        try:
            return openpyxl.load_workbook(master_path)
        except Exception:
            pass
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _pn_master_sheet_rows(wb, title):
    """(values_only) rows of the named sheet, or [] when it isn't present."""
    for ws in wb.worksheets:
        if ws.title.strip().lower() == title.strip().lower():
            return list(ws.iter_rows(values_only=True))
    return []


def _pn_master_replace_sheet(wb, title, headers, data_rows, widths):
    """Drop and recreate the named sheet with `headers` + `data_rows`, leaving
    every other sheet untouched. `data_rows` is a list of value lists."""
    for ws in list(wb.worksheets):
        if ws.title.strip().lower() == title.strip().lower():
            wb.remove(ws)
    ws = wb.create_sheet(title)
    ws.append(list(headers))
    for row in data_rows:
        ws.append(row)
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _pn_master_save(wb, master_path, log, what):
    """Atomically save the master workbook (temp + replace). Ensures at least one
    sheet exists. Best-effort: a master open in Excel is warned about, not fatal."""
    if not wb.worksheets:
        wb.create_sheet(_PN_MASTER_LEAK_SHEET)
    tmp = master_path.with_name(master_path.name + ".tmp")
    try:
        wb.save(tmp)
        os.replace(tmp, master_path)
        return True
    except OSError as e:
        log.warning(f"  Master {what}: could not write {master_path} ({e}) — is "
                    f"it open in Excel? This run's values were not recorded.")
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass
        return False


def _pn_update_master_leaks(master_path, values, case_name, today, log):
    """Merge this run's genuine-leak values into the master workbook's
    'Master Leaks' sheet — one row per distinct value — WITHOUT disturbing the
    KEEP sheet. `values` is [(value, type), ...]."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return
    wb = _pn_master_load(master_path)
    rows = {}   # value_lower -> {value, type, times, cases:set, first, last}
    for r in _pn_master_sheet_rows(wb, _PN_MASTER_LEAK_SHEET)[1:]:
        if not r or r[0] in (None, ""):
            continue
        val = str(r[0])
        times = r[2] if len(r) > 2 else 1
        cases_raw = r[3] if len(r) > 3 else ""
        rows[val.lower()] = {
            "value": val,
            "type": (str(r[1]) if len(r) > 1 and r[1] else "LEAK"),
            "times": int(times) if str(times).isdigit() else 1,
            "cases": {c.strip() for c in str(cases_raw or "").split(";")
                      if c.strip() and "cases" not in c},
            "first": (str(r[4]) if len(r) > 4 and r[4] else today),
            "last": (str(r[5]) if len(r) > 5 and r[5] else today)}

    for value, vtype in values:
        vl = value.lower()
        g = rows.get(vl)
        if g is None:
            rows[vl] = {"value": value, "type": vtype, "times": 1,
                        "cases": {case_name}, "first": today, "last": today}
        else:
            g["times"] += 1
            g["cases"].add(case_name)
            g["last"] = today
            if _leak_sev(vtype) < _leak_sev(g["type"]):
                g["type"] = vtype

    data = []
    for g in sorted(rows.values(), key=lambda g: g["value"].lower()):
        cases = sorted(c for c in g["cases"] if c)
        cell = "; ".join(cases) if len(cases) <= 8 else f"{len(cases)} cases"
        data.append([g["value"], g["type"], g["times"], cell, g["first"], g["last"]])
    _pn_master_replace_sheet(wb, _PN_MASTER_LEAK_SHEET, _PN_MASTER_HEADERS, data,
                             (34, 12, 11, 44, 12, 12))
    if _pn_master_save(wb, master_path, log, "leak log"):
        log.info(f"  Master leak log updated: {master_path} "
                 f"({len(rows)} distinct value(s))")


def _pn_read_master_keep(cfg):
    """The durable cross-folder KEEP decisions {value_lower: decision} from the
    master workbook's KEEP sheet — the `no`/bracket instructions to re-apply on
    every run in every folder. Empty when the sheet is absent/unreadable."""
    path = _pn_master_path(cfg)
    if not path.exists():
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        rows = _pn_master_sheet_rows(wb, _PN_MASTER_KEEP_SHEET)
        wb.close()
    except Exception:
        return {}
    return _pn_parse_decision_rows(rows)


def _pn_update_master_keep(cfg, record_map, case_name, today, log):
    """Merge this run's KEEP decisions into the master workbook's KEEP sheet
    (single persistent sheet across folders and runs), WITHOUT disturbing the
    'Master Leaks' sheet. `record_map` is {value_lower: decision} for the keeps
    made or re-affirmed this run; Times Seen / Cases / dates accumulate so the
    screening can be tuned from real history. Returns the KEEP-sheet row count."""
    if not record_map and not _pn_master_path(cfg).exists():
        return 0
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return 0
    master_path = _pn_master_path(cfg)
    wb = _pn_master_load(master_path)

    rows = {}   # value_lower -> {value, instruction, type, times, cases, first, last, notes}
    for r in _pn_master_sheet_rows(wb, _PN_MASTER_KEEP_SHEET)[1:]:
        if not r or r[0] in (None, ""):
            continue
        val = str(r[0])
        times = r[3] if len(r) > 3 else 1
        cases_raw = r[4] if len(r) > 4 else ""
        rows[val.lower()] = {
            "value": val,
            "instruction": (str(r[1]) if len(r) > 1 and r[1] else "no"),
            "type": (str(r[2]) if len(r) > 2 and r[2] else "KEEP"),
            "times": int(times) if str(times).isdigit() else 1,
            "cases": {c.strip() for c in str(cases_raw or "").split(";")
                      if c.strip() and "cases" not in c},
            "first": (str(r[5]) if len(r) > 5 and r[5] else today),
            "last": (str(r[6]) if len(r) > 6 and r[6] else today),
            "notes": (str(r[7]) if len(r) > 7 and r[7] else ""),
            "origin": (str(r[8]) if len(r) > 8 and r[8] else "")}

    for vl, d in record_map.items():
        instruction = d.get("fixcell") or ("no" if d.get("fix") == "no" else "no")
        vtype = d.get("type") or ("KEEP-PART" if d.get("fake_values") else "KEEP")
        g = rows.get(vl)
        if g is None:
            # First folder to record a value is the one that decided it — an
            # inherited decision already has its row, so it can never land here.
            rows[vl] = {"value": d["value"], "instruction": instruction,
                        "type": vtype, "times": 1, "cases": {case_name},
                        "first": today, "last": today,
                        "notes": d.get("notes", ""), "origin": case_name}
        else:
            g["times"] += 1
            g["cases"].add(case_name)
            g["last"] = today
            g["instruction"] = instruction     # newest instruction wins
            g["type"] = vtype

    data = []
    for g in sorted(rows.values(), key=lambda g: g["value"].lower()):
        cases = sorted(c for c in g["cases"] if c)
        cell = "; ".join(cases) if len(cases) <= 8 else f"{len(cases)} cases"
        data.append([g["value"], g["instruction"], g["type"], g["times"],
                     cell, g["first"], g["last"], g["notes"], g["origin"]])
    _pn_master_replace_sheet(wb, _PN_MASTER_KEEP_SHEET, _PN_MASTER_KEEP_HEADERS,
                             data, (34, 16, 11, 11, 40, 12, 12, 26, 22))
    if _pn_master_save(wb, master_path, log, "KEEP log") and record_map:
        log.info(f"  Master KEEP log updated: {master_path} "
                 f"({len(rows)} kept value(s) tracked).")
    return len(rows)


def _pn_write_leak_report(folder, entries, log, decisions=None, cfg=None,
                          bound=()):
    """Write/refresh the leak-triage worksheet 'LEAKS.xlsx'. Each DISTINCT
    flagged value is ONE row with a 'Fix?' column — the files and page:line
    locations it was found in are aggregated into that row, so a name that
    leaked across nine files is decided once, not nine times. Prior yes/no
    decisions are carried forward (and persisted even when the value no longer
    appears, so the decision keeps applying on later runs). Rows needing
    attention — undecided, or marked-yes-but-still-present — sort to the top;
    resolved ones sink to the bottom. A run with no findings AND no prior
    decisions removes the worksheet. Plain-text checklist fallback without
    openpyxl."""
    decisions = decisions or {}
    xlsx = _pn_leak_xlsx_path(folder)
    txt = _pn_leak_txt_path(folder)
    # A worksheet written under the old name in a prior run would otherwise
    # linger beside the new one, splitting the operator's decisions across two
    # files; remove it once its decisions have been carried into `decisions`.
    stale = [folder / f"{stem}.xlsx" for stem in _PN_LEAK_LEGACY_STEMS]
    stale += [folder / f"{stem}.txt" for stem in _PN_LEAK_LEGACY_STEMS]

    # Group findings by value (case-insensitive): one row per name, with its
    # files and locations aggregated, so the operator marks a repeated leak
    # once. Same value seen with different types keeps the most severe.
    grouped, seen = {}, set()
    for e in entries:
        vl = e["value"].lower()
        seen.add(vl)
        g = grouped.get(vl)
        if g is None:
            grouped[vl] = {"value": e["value"], "type": e["type"],
                           "files": [e["file"]], "wheres": [e["where"]]}
            continue
        if _leak_sev(e["type"]) < _leak_sev(g["type"]):
            g["type"] = e["type"]
        if e["file"] not in g["files"]:
            g["files"].append(e["file"])
        g["wheres"].append(e["where"])

    # Accumulate this run's flagged values into the cross-case master log, so a
    # value that keeps leaking across matters is easy to spot over time. The
    # durable KEEP (no/bracket) decisions are NOT part of this — they live in the
    # master KEEP sheet, maintained by _pn_update_master_keep, so a genuine-leak
    # tally never mixes with a "leave this alone" decision.
    master_path = _pn_master_leaks_path(cfg)
    if master_path is not None and grouped:
        leak_only = [(g["value"], g["type"]) for vl, g in grouped.items()
                     if not _pn_decision_is_keep(decisions.get(vl, {}))]
        if leak_only:
            _pn_update_master_leaks(
                master_path, leak_only, folder.name,
                datetime.date.today().isoformat(), log)

    # The per-folder worksheet is the TRANSIENT genuine-leak triage only; the
    # durable no/bracket KEEP decisions are excluded here (they are preserved
    # cross-folder in the master KEEP sheet) so this file can be cleaned up
    # freely without ever dropping a keep.
    rows = []
    for vl, g in grouped.items():
        d = decisions.get(vl, {})
        if _pn_decision_is_keep(d):
            continue
        files = g["files"]
        file_cell = ", ".join(files) if len(files) <= 3 else f"{len(files)} files"
        # `fixcell` is what the Fix? cell shows: an operator-typed replacement is
        # carried back verbatim (NOT collapsed to "yes"), so a re-run re-reads
        # the exact instruction instead of re-deriving an auto fake.
        rows.append({"file": file_cell, "type": g["type"], "value": g["value"],
                     "where": _pn_merge_where(g["wheres"]),
                     "fix": d.get("fix", ""),
                     "fixcell": (d.get("fixcell") or d.get("replacement")
                                 or d.get("fix", "")),
                     "notes": d.get("notes", ""), "present": True})
    # Persist a Fix?=yes/explicit decision whose value didn't recur this run, so
    # the fix keeps applying — but ONLY while nothing else already holds it.
    # KEEP decisions are omitted (the master KEEP sheet holds them), and so is
    # any value the pseudonym KEY has bound: a Fix?=yes mints a fake that the
    # key pins and every later run re-applies, so carrying the row forward
    # preserves nothing and regenerates LEAKS.xlsx on every clean run — a
    # worksheet whose only content is "(no longer present)", which reads as a
    # leak to review when there is nothing to do.
    bound_low = {str(b).strip().lower() for b in bound}
    for vl, d in decisions.items():
        if vl in bound_low:
            continue
        if vl not in seen and d.get("fix") in ("yes", "no") \
                and not _pn_decision_is_keep(d):
            rows.append({"file": "—", "type": d.get("type") or "(decided)",
                         "value": d["value"], "where": _PN_LEAK_ABSENT,
                         "fix": d["fix"],
                         "fixcell": (d.get("fixcell") or d.get("replacement")
                                     or d["fix"]),
                         "notes": d.get("notes", ""), "present": False})

    def _remove(paths):
        for p in paths:
            try:
                if p.exists():
                    p.unlink()
            except OSError:
                pass

    if not rows:
        _remove((xlsx, txt, *stale))
        log.info("  No potential leaks flagged — no review worksheet written.")
        return

    def _attention(r):        # 0 = needs a look, 2 = resolved -> bottom
        if r["fix"] == "":
            return 0
        if r["fix"] == "yes" and r["present"]:
            return 0          # marked to fix but STILL present — the fix missed
        return 2

    rows.sort(key=lambda r: (_attention(r), _leak_sev(r["type"]),
                             str(r["file"]), r["value"].lower()))
    active = sum(1 for r in rows if _attention(r) == 0)

    try:
        import openpyxl
    except ImportError:
        lines = ["Potential leaks — set Fix? to yes (auto fake), no (leave it), "
                 "type the exact replacement, or [bracket] the part to KEEP "
                 "(the rest is faked)", ""]
        for r in rows:
            mark = {"yes": "[x]", "no": "[-]"}.get(r["fix"], "[ ]")
            lines.append(f"{mark} {r['type']}: {r['value']}  "
                         f"({r['file']} — {r['where']})")
        try:
            txt.write_text("\n".join(lines) + "\n", encoding="utf-8")
            _remove(stale)
            log.warning(f"  Wrote leak-review checklist: {txt.name} "
                        f"({active} to triage).")
        except OSError as ex:
            log.warning(f"  Could not write leak-review checklist: {ex}")
        return

    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _PN_LEAK_SHEET
    ws.append(list(_PN_LEAK_HEADERS))
    for r in rows:
        ws.append([r.get(key, "") for _hdr, key, _w in _PN_LEAK_COLUMNS])
    for i, (_hdr, _key, width) in enumerate(_PN_LEAK_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    try:                                    # a yes/no dropdown on the Fix? column
        from openpyxl.worksheet.datavalidation import DataValidation
        fix_col = get_column_letter(
            next(i for i, (h, _k, _w) in enumerate(_PN_LEAK_COLUMNS, start=1)
                 if h.lower().startswith("fix?")))
        dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
        dv.showErrorMessage = False   # offer the dropdown but ALSO allow a typed
        dv.showInputMessage = True    # replacement in the cell, not just yes/no
        dv.promptTitle = "Fix?"
        dv.prompt = ("yes = auto fake · no = leave it · type the exact "
                     "replacement to use · or [bracket] the part to KEEP and "
                     "the rest is faked")
        ws.add_data_validation(dv)
        dv.add(f"{fix_col}2:{fix_col}{len(rows) + 1}")
    except Exception:
        pass
    try:
        wb.save(xlsx)
        _remove((txt, *stale))
        log.warning(f"  Wrote leak-review worksheet: {xlsx.name} — {active} "
                    f"item(s) to triage (Fix?: 'yes' scrubs with an auto fake, "
                    f"'no' leaves it, or type the exact replacement to use).")
    except OSError as ex:
        log.warning(f"  Could not write leak-review worksheet: {ex}")


def _write_text_version(pdf_path: Path, doc, log: logging.Logger,
                        pseudonymizer=None, text_subdir="Text Files",
                        original_subdir=None) -> bool:
    """Write a plain-text companion for the PDF into a `text_subdir` subfolder
    of the case folder (created if needed), with pdf_linker's own invisible
    right-margin markers stripped. Overwrites on re-runs. Returns True if
    written.

    When `pseudonymizer` is given, each page's text is pseudonymized (party
    names, case numbers, PII → stable fakes) before writing — the PDF itself is
    never modified. The finished text is re-scanned and any real value that
    survived the replacement is LOGGED as a leak for review; the file is still
    written (it is not withheld).

    When `original_subdir` is also given (and a pseudonymizer is running), the
    UNSCRUBBED text is additionally written to that sibling folder under the
    PDF's real name — a QA / reference copy. It carries real names by design,
    so it is never tracked for the leak gate and never quarantined; the folder
    name is meant to flag that it must not be shared.
    """
    out_dir = pdf_path.parent / text_subdir
    txt_path = out_dir / (pdf_path.stem + ".txt")
    orig_pages = []   # pre-pseudonymization page text, for citation detection
    detect_pages = []  # column-ordered page text, for name detection / leak check
    page_blocks = []  # (header, rows_or_text) before pseudonymization; rows is
                      # a list of (line_num, segments) for pleading pages, else a str
    for i, page in enumerate(doc):
        raw = page.get_text("text")
        # Drop our invisible citation markers (present when re-processing an
        # already-linked PDF) so the export is clean prose.
        clean = _MARKER_DETECT_RE.sub("", raw)
        # Collapse the runs of blank lines that marker removal and pleading
        # gutters leave behind. This flowing text drives citation detection.
        clean = re.sub(r"\n[ \t]*\n(?:[ \t]*\n)+", "\n\n", clean).strip()
        orig_pages.append(clean)

        # Displayed text: on pleading-paper pages, keep per-line rows so we can
        # number each line (a "p.X:Y" cite lands on the right text) AND scrub
        # each line's body without its number prefix; other pages fall back to
        # flowing text.
        rows = _page_lined_rows(page)
        detect_pages.append(_page_detect_text(page))
        # Header carries the printed (footer) page number when present, so the
        # page half of a pinpoint cite is unambiguous too.
        label = _footer_page_label(page)
        header = (f"====== Page {i + 1}"
                  + (f" (printed p. {label})" if label else "")
                  + " ======")
        page_blocks.append((header, rows if rows is not None else clean))

    # Register this document's declarant(s) and any "dba" pair written into the
    # text, so those names are scrubbed even when absent from the spreadsheet
    # key — must happen BEFORE any apply() below. Read from the COLUMN-ORDERED
    # rendering: on a two-column caption the on-paper reading order interleaves
    # the two, and a name spliced with the opposing column matches nothing.
    detect_full = "\n\f\n".join(detect_pages)
    if pseudonymizer is not None:
        pseudonymizer.register_declarant_names(detect_full)
        pseudonymizer.register_declarant_refs(detect_full)
        pseudonymizer.register_court_names(detect_full)
        pseudonymizer.register_dba_names(detect_full)
        pseudonymizer.register_firm_names(detect_full)
        pseudonymizer.register_label_names(detect_full)
        pseudonymizer.register_short_names(detect_full)
        pseudonymizer.register_entity_acronyms(detect_full)
        pseudonymizer.register_identifiers(detect_full)
        pseudonymizer.register_localities(detect_full)
        pseudonymizer.register_addresses(detect_full)

    def build_body(pz):
        """Assemble the page text — scrubbed when `pz` is given, the raw
        extraction when `pz` is None (the original-text copy)."""
        parts = []
        for header, content in page_blocks:
            if isinstance(content, list):   # pleading page: (line_num, segments)
                nums = [num for num, _ in content]
                if pz is not None:
                    # Column by column, so a name the page wrapped mid-name
                    # stays contiguous; rows re-assembled left to right after.
                    bodies = _pn_apply_page_rows(pz, content)
                else:
                    bodies = [" ".join(t for _x, t in segs)
                              for _num, segs in content]
                # A gutter number owns its first row; continuation rows (a
                # stacked letterhead/caption line) carry line_num=None and print
                # under a blank, aligned gutter so the block reads as lines.
                display = "\n".join(
                    ((f"{num:>2}  " if num is not None else "    ") + b).rstrip()
                    for num, b in zip(nums, bodies))
            else:                           # flowing text
                display = pz.apply(content) if pz is not None else content
            parts.append(f"{header}\n{display}")
        out = "\n\n".join(parts) + "\n"
        # Append a list of every cited authority with a free public
        # verification link — the text copy points somewhere anyone can open.
        appendix = _build_authorities_appendix("\n\f\n".join(orig_pages), pz)
        if appendix:
            out += "\n" + appendix + "\n"
        return out

    body = build_body(pseudonymizer)

    if pseudonymizer is not None:
        # Report (but do NOT withhold) any real value that survived the scrub —
        # the .txt is still written; the log flags it for review. Check the
        # column-ordered rendering as well as the display text: a name that the
        # page wrapped mid-name is only contiguous down its own column, so the
        # display text alone would report a page like that as clean.
        scrubbed_detect = pseudonymizer.apply(detect_full, count=False)

        # An e-mail is always faked, so cure any address the pattern pass left
        # standing rather than asking about it in the worksheet.
        body = pseudonymizer.scrub_emails(body)
        scrubbed_detect = pseudonymizer.scrub_emails(scrubbed_detect)

        # Column-splice check: if a page's extraction is corrupted, the whole-
        # word patterns match nothing, so also scan the alphanumeric reduction of
        # what was WRITTEN — `SCHILLECITORTORICILAW` survives the scrub and still
        # contains `schillecitortorici`. Scan the scrubbed output, not the source
        # (the source contains every real name by definition).
        spliced = [i + 1 for i, (_h, c) in enumerate(page_blocks)
                   if isinstance(c, list) and _page_looks_spliced(c)]
        if spliced:
            # A welded party name defeats the boundary-anchored patterns, so
            # cure the written body with the reduced-span replacement (the
            # write-side mirror of surviving_reals_reduced) BEFORE re-scanning —
            # otherwise detection out-runs replacement and quarantines an export
            # the substituter never had a chance to clean.
            body = pseudonymizer.scrub_welded(body)
            scrubbed_detect = pseudonymizer.scrub_welded(scrubbed_detect)
            log.warning(f"  Pseudonymization REVIEW on {pdf_path.name}: caption "
                        f"on page(s) {spliced} appears column-spliced; term "
                        f"matching is unreliable there — verify by hand.")

        survivors = set(pseudonymizer.surviving_reals(body))
        survivors |= set(pseudonymizer.surviving_reals(scrubbed_detect))
        if spliced:
            survivors |= set(pseudonymizer.surviving_reals_reduced(body))
            survivors |= set(pseudonymizer.surviving_reals_reduced(scrubbed_detect))

        if survivors:
            pseudonymizer.note_leaks(survivors)
            shown = ", ".join(sorted(survivors)[:8])
            log.warning(f"  Pseudonymization LEAK on {pdf_path.name}: real "
                        f"value(s) still present in the .txt ({shown}). Review "
                        f"before sharing; add them with --term and re-run.")

        # Open-world REVIEW scan: identifier shapes (licence/bar/reservation/file
        # numbers, bare URLs) that must never ride along even without a key entry.
        review = pseudonymizer.review_scan(body)
        # Backstop: a party acronym defined in the source that survived the scrub
        # (a definition shape register_short_names didn't anticipate).
        review = list(review) + pseudonymizer.review_definition_survivors(
            detect_full, body)
        # High-recall tier: role-anchored name shapes in the output that are
        # neither our fakes nor common words — the "unknown name" net.
        review = list(review) + pseudonymizer.unknown_name_scan(body)
        # Adversarial re-identification pass: a bar number / VIN / reservation
        # shape in the OUTPUT that isn't one of our own fakes. Sorted first —
        # these invert the map in one lookup, so they outrank ordinary review.
        review = pseudonymizer.reid_scan(body) + review
        if review:
            shown = "; ".join(f"{c}: {s}" for c, s in review[:8])
            log.warning(f"  Pseudonymization REVIEW on {pdf_path.name}: "
                        f"identifier(s) to check before sharing ({shown}).")

        # Collect every finding into the run-wide triage worksheet, each located
        # to its printed page and gutter line so it can be found and judged.
        parsed = _pn_body_lines(body)
        for real in sorted(survivors):
            # A pure pleading phrase ("Opposition", "Plaintiff's Opposition to
            # Mot.") is a document type, never a party — never worth a worksheet
            # row no matter which scan produced it.
            if _pn_is_procedural_phrase(real) or _pn_is_email_value(real):
                continue
            pseudonymizer.leak_report.append(
                {"file": pdf_path.name, "type": "LEAK", "value": real,
                 "where": _pn_locate(parsed, real)})
        for cls, sample in review:
            if _pn_is_procedural_phrase(sample) or _pn_is_email_value(sample):
                continue
            pseudonymizer.leak_report.append(
                {"file": pdf_path.name, "type": cls, "value": sample,
                 "where": _pn_locate(parsed, sample)})

        # Pseudonymize the output filename too — the .txt is the artifact that
        # gets shared, so a party/attorney name must not survive in its name.
        txt_path = _pseudonymized_txt_path(out_dir, pdf_path, pseudonymizer, log)

    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        log.warning(f"  Could not create text subfolder {out_dir}: {e}")
        return False

    # Remove any earlier .txt for this PDF that isn't the file we're about to
    # write: the old case-root location (pre-subfolder / non-pseudonymized runs)
    # and the default-named one in the subfolder — so a stale (possibly
    # real-named) copy can't linger beside the current output. ONLY a file this
    # tool plausibly wrote is removed — every export carries the "====== Page"
    # header, so a same-stem file WITHOUT it is the user's own (notes beside
    # the PDF) and used to be silently destroyed here.
    for stale in {pdf_path.with_suffix(".txt"), out_dir / (pdf_path.stem + ".txt")}:
        if stale != txt_path and stale.exists():
            try:
                head = stale.read_text(encoding="utf-8", errors="ignore")[:4000]
                if "====== Page" not in head:
                    continue                     # not ours — leave it alone
                stale.unlink()
            except OSError:
                pass

    try:
        # newline="\n" explicitly: the default translates to the platform's
        # line ending, so every .txt written on Windows shipped with CRLF and
        # diffed against a Linux-produced copy line for line.
        txt_path.write_text(body, encoding="utf-8", newline="\n")
    except Exception as e:
        log.warning(f"  Could not write text version (non-fatal): {e}")
        return False
    if pseudonymizer is not None:
        # Track it so the folder-level no-match check can report a count, and
        # record which real values leaked IN THIS file so the gate can
        # quarantine only the offending exports.
        pseudonymizer.written.append(txt_path)
        if survivors:
            pseudonymizer.leaked_by_file[txt_path] = {s.lower() for s in survivors}
    log.info(f"  Wrote {'pseudonymized ' if pseudonymizer else ''}text "
             f"version: {text_subdir}/{txt_path.name}")

    # Optional UNSCRUBBED reference copy in a sibling folder, under the real
    # filename. Real names by design → never tracked for the leak gate, never
    # quarantined. Written only when pseudonymization is actually running (with
    # it off, the single export above already IS the original).
    if original_subdir and pseudonymizer is not None:
        orig_dir = pdf_path.parent / original_subdir
        orig_path = orig_dir / (pdf_path.stem + ".txt")
        try:
            orig_dir.mkdir(parents=True, exist_ok=True)
            orig_path.write_text(build_body(None), encoding="utf-8", newline="\n")
            log.info(f"  Wrote original text version: "
                     f"{original_subdir}/{orig_path.name}")
        except OSError as e:
            log.warning(f"  Could not write original text version "
                        f"(non-fatal): {e}")
    return True


# ────────────────────────────────────────────────────────────────────────────
# Main per-PDF processing
# ────────────────────────────────────────────────────────────────────────────
def _pn_learn_from_text(pseudonymizer, text, stem=None):
    """Run every document-learning register_* pass over one document's text (and,
    when given, its filename stem). Shared by the PDF pre-scan and the Word-doc
    conversion so both feed the same folder-wide vocabulary."""
    pseudonymizer.register_declarant_names(text)
    pseudonymizer.register_declarant_refs(text)
    pseudonymizer.register_court_names(text)
    # The FILENAME often names a declarant ("Yu Declaration ISO Opp.pdf") whose
    # name may appear nowhere in the body — scan it too, so both the body and the
    # pseudonymized output filename scrub that name.
    if stem:
        pseudonymizer.register_declarant_refs(re.sub(r"[_\-]+", " ", stem))
    pseudonymizer.register_dba_names(text)
    pseudonymizer.register_firm_names(text)
    pseudonymizer.register_label_names(text)
    pseudonymizer.register_short_names(text)
    pseudonymizer.register_entity_acronyms(text)
    pseudonymizer.register_identifiers(text)
    pseudonymizer.register_localities(text)
    pseudonymizer.register_addresses(text)


def _pn_prescan_folder(pdfs, pseudonymizer, log, extra_texts=()):
    """Run EVERY document-learning register_* pass over the whole folder before
    any .txt is written.

    Names, localities and identifiers are read out of document text, so a value
    learned from one filing is only available to filings processed after it —
    and one filing routinely states a party's full address (teaching the city)
    while another states only the bare city, or spells out a "dba" the others
    just use. Running all passes up front removes that ordering dependency
    completely: the folder is scrubbed only after it has been read in full.

    `extra_texts` is a list of (stem, text) pairs from non-PDF sources (Word docs
    bulk-converted in the same run), so a name that appears only in a Word filing
    is harvested folder-wide too."""
    import fitz
    before = len(pseudonymizer.terms)
    corpus = []
    for pdf in pdfs:
        try:
            with fitz.open(pdf) as doc:
                text = "\n\f\n".join(_page_detect_text(page) for page in doc)
        except Exception as e:  # a corrupt file must not abort the whole run
            log.warning(f"  Pseudonymize: could not pre-scan {pdf.name}: {e}")
            continue
        corpus.append(text)
        _pn_learn_from_text(pseudonymizer, text, pdf.stem)
    for stem, text in extra_texts:
        corpus.append(text)
        _pn_learn_from_text(pseudonymizer, text, stem)
    added = len(pseudonymizer.terms) - before
    if added:
        log.info(f"  Pseudonymize: pre-scan of "
                 f"{len(pdfs) + len(extra_texts)} file(s) added "
                 f"{added} term(s) from names, localities and identifiers")
    full = "\n\f\n".join(corpus)
    pruned = pseudonymizer.prune_citation_only_terms(full)
    if pruned:
        log.info(f"  Pseudonymize: dropped {len(pruned)} harvested name(s) "
                 f"that appear only inside case citations "
                 f"({', '.join(sorted(pruned)[:6])})")
    return full


# ── Word-document bulk conversion ────────────────────────────────────────────
# Some folders hold only Word filings, with no PDFs at all. Convert those to the
# same scrubbed .txt exports the PDFs get (no hyperlinking — Word docs are never
# linked), but ONLY when the folder holds no PDFs. A Word doc beside a pile of
# PDFs is the ordinary workflow and must stay untouched, so the pass is gated on
# the folder having no PDFs (checked by the caller in main()).
# Zipped OOXML Word formats we can read with the stdlib alone (a zip of XML).
# The legacy binary .doc has no such structure and needs an external converter,
# so it is not handled here (a warning names any that are found).
_WORD_DOC_SUFFIXES = (".docx", ".docm")


def _pdfs_in_folder(folder):
    """Source PDFs directly in `folder` (not recursive), excluding this tool's
    own `_linked` / `_temp` output. The single definition of "does this folder
    hold a PDF batch" — the Word gate below asks the same question the main
    run does, so the two can never drift apart."""
    return [p for p in folder.glob("*.pdf")
            if not p.stem.endswith("_linked") and not p.stem.endswith("_temp")]


def _is_word_only_folder(folder):
    """True when this folder is an all-WORD batch: Word documents and no PDFs.

    Such a folder is converted to the same scrubbed .txt exports a PDF batch
    gets — but it is typically a one-off (a draft, a set of filings pulled from
    e-mail) that has no E-Court `Order*.xlsx` of its own. The Downloads
    fallback then hands it whatever case was downloaded LAST, so the run
    pseudonymizes against a stranger's party list: this case's parties go
    unscrubbed while another matter's names are hunted for. Callers use this to
    withhold that guess."""
    return bool(_word_docs_in_folder(folder)) and not _pdfs_in_folder(folder)


def _word_docs_in_folder(folder):
    """Convertible Word documents directly in `folder` (not recursive), sorted
    by name. Word's own '~$'-prefixed owner/lock files are skipped."""
    docs = []
    for p in folder.iterdir():
        if not p.is_file() or p.name.startswith("~$"):
            continue
        if p.suffix.lower() in _WORD_DOC_SUFFIXES:
            docs.append(p)
    return sorted(docs, key=lambda p: p.name.lower())


def _extract_docx_text(path):
    """Extract the body text of a .docx/.docm as plain text, dependency-free
    (a Word document is a zip of XML). Paragraphs and table cells become lines;
    <w:tab/> becomes a tab and <w:br/>/<w:cr/> a newline. Headers, footers and
    footnotes live in separate parts and are not included. Raises on a file that
    isn't a readable Word document so the caller can skip and warn."""
    import zipfile
    from xml.etree import ElementTree as ET

    with zipfile.ZipFile(path) as zf:
        data = zf.read("word/document.xml")
    root = ET.fromstring(data)

    out = []

    def walk(node):
        # Local tag name only — the WordprocessingML namespace is verbose and
        # irrelevant to which element this is.
        tag = node.tag.rsplit("}", 1)[-1]
        if tag == "t":            # a run of literal text
            out.append(node.text or "")
            return
        if tag == "tab":
            out.append("\t")
        elif tag in ("br", "cr"):
            out.append("\n")
        for child in node:
            walk(child)
        if tag == "p":            # end of a paragraph → line break
            out.append("\n")

    walk(root)
    text = "".join(out)
    # Trim trailing whitespace per line and collapse runs of blank lines, the
    # same tidy-up the PDF export does, so the two read alike.
    text = "\n".join(ln.rstrip() for ln in text.split("\n"))
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text


def _word_locate(body, needle, limit=12):
    """Human-readable 'line N' locations of `needle` in a Word export body
    (case-insensitive). Word text has no page/gutter structure, so findings are
    located by line number within the export."""
    nl = needle.lower()
    out = []
    for i, line in enumerate(body.split("\n"), 1):
        if nl in line.lower():
            loc = f"line {i}"
            if loc not in out:
                out.append(loc)
    if not out:
        return "(not located)"
    return ", ".join(out[:limit]) + (" …" if len(out) > limit else "")


def _write_word_text_version(src_path, text, log, pseudonymizer=None,
                             text_subdir="Text Files", original_subdir=None):
    """Write a plain-text export of one Word document into the `text_subdir`
    subfolder, mirroring the PDF text pipeline: when a pseudonymizer is given the
    text is scrubbed, leak-scanned and given a name-scrubbed filename exactly as
    a PDF export is (the Word file itself is never modified); Word docs are never
    hyperlinked. Returns True if written."""
    out_dir = src_path.parent / text_subdir
    txt_path = out_dir / (src_path.stem + ".txt")

    if pseudonymizer is not None:
        # Per-file backstop to the folder pre-scan (mirrors _write_text_version):
        # learn any declarant / dba / firm / locality / identifier stated here.
        _pn_learn_from_text(pseudonymizer, text, src_path.stem)
        body = pseudonymizer.apply(text)

        survivors = set(pseudonymizer.surviving_reals(body))
        if survivors:
            pseudonymizer.note_leaks(survivors)
            shown = ", ".join(sorted(survivors)[:8])
            log.warning(f"  Pseudonymization LEAK on {src_path.name}: real "
                        f"value(s) still present in the .txt ({shown}). Review "
                        f"before sharing; add them with --term and re-run.")

        review = list(pseudonymizer.review_scan(body))
        review += pseudonymizer.review_definition_survivors(text, body)
        review += pseudonymizer.unknown_name_scan(body)
        review = pseudonymizer.reid_scan(body) + review
        if review:
            shown = "; ".join(f"{c}: {s}" for c, s in review[:8])
            log.warning(f"  Pseudonymization REVIEW on {src_path.name}: "
                        f"identifier(s) to check before sharing ({shown}).")

        for real in sorted(survivors):
            if _pn_is_procedural_phrase(real) or _pn_is_email_value(real):
                continue
            pseudonymizer.leak_report.append(
                {"file": src_path.name, "type": "LEAK", "value": real,
                 "where": _word_locate(body, real)})
        for cls, sample in review:
            if _pn_is_procedural_phrase(sample) or _pn_is_email_value(sample):
                continue
            pseudonymizer.leak_report.append(
                {"file": src_path.name, "type": cls, "value": sample,
                 "where": _word_locate(body, sample)})

        # Scrub the output filename too — the .txt is the shared artifact, so a
        # party/attorney name in the Word file's name must not ride along.
        txt_path = _pseudonymized_txt_path(out_dir, src_path, pseudonymizer, log)
    else:
        body = text

    body = body.rstrip("\n") + "\n"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        log.warning(f"  Could not create text subfolder {out_dir}: {e}")
        return False
    try:
        txt_path.write_text(body, encoding="utf-8", newline="\n")
    except Exception as e:
        log.warning(f"  Could not write Word text version (non-fatal): {e}")
        return False

    if pseudonymizer is not None:
        pseudonymizer.written.append(txt_path)
        if survivors:
            pseudonymizer.leaked_by_file[txt_path] = {s.lower() for s in survivors}
    log.info(f"  Wrote {'pseudonymized ' if pseudonymizer else ''}text "
             f"version: {text_subdir}/{txt_path.name}")

    # Optional UNSCRUBBED reference copy, under the real filename — same policy
    # as the PDF path (real names by design, never gated, must not be shared).
    if original_subdir and pseudonymizer is not None:
        orig_dir = src_path.parent / original_subdir
        orig_path = orig_dir / (src_path.stem + ".txt")
        try:
            orig_dir.mkdir(parents=True, exist_ok=True)
            orig_path.write_text(text.rstrip("\n") + "\n",
                                 encoding="utf-8", newline="\n")
            log.info(f"  Wrote original text version: "
                     f"{original_subdir}/{orig_path.name}")
        except OSError as e:
            log.warning(f"  Could not write original Word text version "
                        f"(non-fatal): {e}")
    return True


def _convert_word_docs(word_texts, log, pseudonymizer=None,
                       text_subdir="Text Files", original_subdir=None):
    """Convert the already-extracted Word documents to .txt exports. `word_texts`
    is a list of (path, text) pairs (extraction/gating done by the caller).
    Returns the number of exports written."""
    written = 0
    for src_path, text in word_texts:
        try:
            if _write_word_text_version(src_path, text, log, pseudonymizer,
                                        text_subdir, original_subdir):
                written += 1
        except Exception as e:
            log.error(f"  Unhandled error converting Word doc "
                      f"{src_path.name}: {e}")
            log.error(traceback.format_exc())
    return written


# Metadata stamp written into a linked PDF's `keywords`, so a re-run can tell a
# file it already linked and skip the (expensive) re-detect + re-save pass,
# regenerating only the .txt. `--relink` forces the full pass anyway. Existing
# keywords are preserved; the sentinel is just appended.
_PDF_LINK_STAMP = "PDF-Linker-linked"


def _pdf_is_stamped(doc):
    try:
        return _PDF_LINK_STAMP in ((doc.metadata or {}).get("keywords") or "")
    except Exception:
        return False


# The splice PyMuPDF's insert_link writes into a URI that itself contains
# "/Link": ".../Link/NM(fitz-L0)/Document/..." (see the repair call in
# process_pdf). The NM name is always "fitz-L<n>".
_PN_BAD_NM_RE = re.compile(r"/NM\(fitz-L\d+\)")


def _repair_link_uris(doc):
    """Remove the '/NM(fitz-Ln)' fragment PyMuPDF splices into a link URI that
    contains '/Link' (every Westlaw FullText URL). Works via xref surgery on
    each page's Annots array: freshly inserted links are invisible to
    page.get_links() until the doc is saved, and going back through
    update_link would re-run the same buggy string replace. Best-effort."""
    for page in doc:
        try:
            kind, val = doc.xref_get_key(page.xref, "Annots")
        except Exception:
            continue
        if kind == "xref":               # indirect array: resolve it
            try:
                m = re.match(r"(\d+)", val)
                kind, val = doc.xref_get_key(int(m.group(1)), "")
            except Exception:
                continue
        if kind != "array" or not val:
            continue
        for xs in re.findall(r"(\d+)\s+0\s+R", val):
            xref = int(xs)
            try:
                ktype, uri = doc.xref_get_key(xref, "A/URI")
            except Exception:
                continue
            if ktype != "string" or "/NM(fitz-L" not in (uri or ""):
                continue
            fixed = _PN_BAD_NM_RE.sub("", uri)
            escaped = fixed.replace("\\", r"\\").replace("(", r"\(") \
                           .replace(")", r"\)")
            try:
                doc.xref_set_key(xref, "A/URI", f"({escaped})")
            except Exception:
                pass


def _pdf_stamp_linked(doc):
    try:
        md = dict(doc.metadata or {})
        kw = md.get("keywords") or ""
        if _PDF_LINK_STAMP not in kw:
            md["keywords"] = (kw + "; " if kw else "") + _PDF_LINK_STAMP
            doc.set_metadata(md)
    except Exception:
        pass


def process_pdf(pdf_path: Path, log: logging.Logger,
                provider: str = "lexis", extract_text: bool = True,
                pseudonymizer=None, text_subdir="Text Files",
                original_subdir=None, relink: bool = False) -> bool:
    """Process one PDF. Returns True on success."""
    try:
        import fitz
    except ImportError:
        log.error("PyMuPDF (fitz) not installed - aborting. "
                  "Install with: pip install pymupdf")
        return False

    log.info(f"Processing: {pdf_path.name}")
    out_path = pdf_path  # Output will overwrite the original (after we're done with it)
    temp_path = pdf_path.with_name(pdf_path.stem + "_temp.pdf")
    if temp_path.exists():
        # A leftover temp is debris from a run that crashed between save and
        # replace. Skipping (and reporting success!) made the file permanently
        # unprocessed — every later run hit the same temp and skipped again.
        # Remove it and process normally; the save below rewrites it anyway.
        try:
            temp_path.unlink()
            log.warning(f"  Removed stale temp from an interrupted run: "
                        f"{temp_path.name}")
        except OSError as e:
            log.error(f"  Could not remove stale temp {temp_path.name}: {e}")
            return False

    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        log.error(f"  Could not open PDF: {e}")
        return False

    # Check whether any page lacks text. The OCR gate is intentionally
    # triggered by "any page lacks text" rather than "every page lacks
    # text" — mixed-state documents are common in declarations, where the
    # body and exhibit content pages have text but image-only "EXHIBIT A"
    # cover sheets do not. Without per-page OCR, the cover sheets stay
    # textless, _find_exhibit_cover_pages can't see them, and exhibit
    # linking silently produces nothing. _ocr_pdf itself has its own
    # per-page check (it skips any page that already has text), so this
    # outer gate just needs to admit any document with at least one
    # textless page. Idempotent across re-runs because re-runs find every
    # page already has text and _ocr_pdf is a no-op.
    #
    # First, fix pages whose existing text layer is garbled (broken font
    # encoding, unmapped glyphs): rebuild them from a clean OCR pass so every
    # downstream step - the text-export decision, citation extraction, linking
    # - reads accurate text. Only clearly-garbled pages are touched. Non-fatal.
    ocr_changed = False
    try:
        ocr_changed = bool(_reocr_garbled_pages(doc, log))
    except Exception as e:
        log.warning(f"  Garbled-text re-OCR failed (non-fatal): {e}")

    # Decide whether to export a plain-text companion. This uses the NATIVE
    # text and must run before OCR, which would add a text layer to scanned
    # pages and make every document look text-based.
    want_text_export = extract_text and _pdf_has_text_layer(doc, log)

    any_textless = any(not page.get_text("text").strip() for page in doc)
    if any_textless:
        textless_count = sum(1 for p in doc if not p.get_text("text").strip())
        if textless_count == len(doc):
            log.info("  No text layer detected - running OCR")
        else:
            log.info(f"  {textless_count} page(s) lack text - running OCR "
                     f"on those page(s)")
        ocr_changed = bool(_ocr_pdf(doc, log)) or ocr_changed

    # Export the text version now (after OCR fills any minority scanned pages,
    # before marker injection pollutes the text). Non-fatal and independent of
    # the linking passes below. Skipped for scanned-only PDFs.
    if want_text_export:
        try:
            _write_text_version(pdf_path, doc, log, pseudonymizer, text_subdir,
                                original_subdir)
        except Exception as e:
            log.warning(f"  Text export failed (non-fatal): {e}")

    # Fast path for a re-run: this PDF was already linked (metadata stamp), so
    # the citation links, bookmarks and blue underlines are already in it.
    # Skip the whole detect + re-save pass — the .txt above was already
    # regenerated (the usual reason to re-run) — unless --relink forces it.
    # When OCR actually MODIFIED the doc this run (a page gained its text
    # layer), fall through instead: skipping would discard the OCR and re-pay
    # the 300-dpi render on every subsequent run without ever persisting it.
    if _pdf_is_stamped(doc) and not relink and not ocr_changed:
        log.info(f"  Already linked; skipped the link/save pass, regenerated "
                 f".txt only (use --relink to force): {pdf_path.name}")
        doc.close()
        return True

    # Skip link insertion for declarations and separate statements: they
    # rarely contain citation-worthy material in Zachary's workflow, and
    # avoiding the link-insertion pass eliminates spurious annotations on
    # body text that happens to share substrings with TOA citations from
    # related briefing. The right-margin markers still get injected below —
    # those are independent of the link annotations and are exactly what
    # the Word paste macro relies on for declaration citations.
    skip_links = should_skip_linking(pdf_path.name)
    if skip_links:
        log.info(f"  Skipping link insertion (Decl./Separate Statement): "
                 f"{pdf_path.name}")
        all_cites = []
    else:
        # Build full-document text + per-page text spans for supra resolution
        page_texts = []
        for page in doc:
            page_texts.append(page.get_text("text"))

        full_text = "\n\f\n".join(page_texts)
        all_cites = find_all_citations(full_text)
        log.info(f"  Found {len(all_cites)} citations")
        # Page-start offsets in full_text, so a citation whose match text occurs
        # exactly once can be linked on its OWN page instead of searched across
        # every page. The all-pages scan is O(citations x pages) — ~180k page
        # searches on a 448-page OCR'd exhibit with 401 citations, which is why
        # linking crawled. A once-only match lives on a single page (provably),
        # so mapping its span to a page and searching only there is exact and
        # fast; repeated text keeps the full scan.
        _cite_page_starts, _acc = [], 0
        for _pt in page_texts:
            _cite_page_starts.append(_acc)
            _acc += len(_pt) + len("\n\f\n")

    # For each citation, locate occurrences on each page using PyMuPDF's
    # search_for, and add (1) a clickable link annotation and (2) a blue
    # underline annotation. We keep the original text colour untouched —
    # changing text colour on existing PDFs requires redacting and
    # reinserting glyphs, which shifts layout in unpredictable ways. The
    # blue underline plus clickability is unambiguous as a hyperlink cue
    # and matches how other PDF tools (e.g. Adobe Acrobat's web-link
    # converter) handle the same job.
    #
    # CRITICAL — multi-line citations: if a citation in the source text
    # spans a line break (e.g., "Smith v. Jones (2017)\n13 Cal.App.5th 1152"),
    # `match_text` will contain that newline, and `page.search_for(match_text)`
    # returns nothing because PyMuPDF doesn't match across line breaks.
    # The previous fallback — searching for `match_text.splitlines()[0]` —
    # was unsafe: when the first line was a generic substring like "Cal."
    # or "Smith", `search_for` returned matches all over the document and
    # every "Cal." on every page got linked to the same wrong URL.
    # Instead, we now split the match into both halves and require BOTH
    # to be found nearby on the SAME page. If only one half is on a page,
    # we annotate just that half rather than blasting links across unrelated
    # text. We also reject any single-fragment search whose fragment is
    # too short or too generic to be meaningful (`_is_safe_fragment`).
    link_count = 0
    linked_rects: dict = {}

    # Pre-seed linked_rects with link annotations already on each page so a
    # re-run on an already-linked PDF doesn't stack duplicate URI links and
    # double-paint the blue underline. The same `_already_linked` substantial-
    # overlap check that prevents duplicates within one run also catches
    # rects covered by a prior run.
    for _seed_page in doc:
        for _seed_ln in _seed_page.get_links():
            _seed_rect = _seed_ln.get("from")
            if _seed_rect:
                linked_rects.setdefault(_seed_page.number, []).append(_seed_rect)

    def _already_linked(page_num: int, rect) -> bool:
        for r in linked_rects.get(page_num, []):
            inter = rect & r
            if not inter.is_empty:
                smaller = min(rect.get_area(), r.get_area())
                if smaller > 0 and inter.get_area() / smaller > 0.5:
                    return True
        return False

    def _record_linked(page_num: int, rect) -> None:
        linked_rects.setdefault(page_num, []).append(rect)

    import bisect
    for cite in all_cites:
        url = resolve_url(cite, provider)
        if not url:
            continue
        match_text = cite["match_text"]
        # If this exact text appears once in the whole document, it is on ONE
        # page — link it there and skip the other N-1 pages. Otherwise (repeated
        # text) keep the full scan so every occurrence is still linked.
        if full_text.count(match_text) == 1:
            hp = bisect.bisect_right(_cite_page_starts, cite["span"][0]) - 1
            hp = max(0, min(hp, len(doc) - 1))
            pages_iter = (doc[hp],)
        else:
            pages_iter = doc
        found_anywhere = False
        for page in pages_iter:
            quads = _safe_search_for_citation(page, match_text, cite)
            if not quads:
                continue
            for q in quads:
                rect = q.rect
                # Skip quads in the right-margin page-number column:
                # search_for can return a separate quad for the page-ref
                # numbers ("13, 19") on TOA lines. Those quads have x0 well
                # into the right margin. Suppress them so links only cover
                # the citation text, not the page references.
                if rect.x0 > page.rect.width - 90:
                    continue
                # Skip if already covered by a previously-inserted link
                # (prevents duplicate annotations when multiple citation
                # patterns match the same text, e.g. RPC "1.9" and "1.9(a)").
                if _already_linked(page.number, rect):
                    continue
                _record_linked(page.number, rect)
                # 1. Clickable link
                page.insert_link({
                    "kind": fitz.LINK_URI,
                    "from": rect,
                    "uri": url,
                })
                # 2. Blue underline
                underline_y = rect.y1 - 0.5
                page.draw_line(
                    fitz.Point(rect.x0, underline_y),
                    fitz.Point(rect.x1, underline_y),
                    color=LINK_COLOUR,
                    width=1.0,
                )
                found_anywhere = True
        if found_anywhere:
            link_count += 1

    # After all full citations are linked, do a second pass for short-form
    # case references — bare "X v. Y" mentions (no reporter or date) — and
    # link them to the URL of the matching full citation. This handles the
    # common pattern where a brief introduces "Chillon v. Ford Motor Co.,
    # 2023 WL 3035369..." once and then refers to it as just "Chillon v.
    # Ford" or "Chillon v. Ford Motor Co." in the surrounding discussion.
    # Skipped when link insertion is suppressed (Decl./Separate Statement).
    if not skip_links and all_cites:
        link_count += _link_short_form_cases(doc, all_cites, provider, log)

    log.info(f"  Linked {link_count} citation occurrence(s)")

    # Add internal jumps for any Table of Contents found in the front matter.
    # Skipped for the same filename patterns that skip citation linking
    # (Decl./Decl./Separate Statement/Complaint/FAC/SAC/TAC/Proof of Service):
    # those documents don't carry TOCs in Zachary's workflow, and any
    # TOC-shaped match in their body text would be a false positive.
    # Failure here is non-fatal.
    #
    # toc_page_range is the (start, end_inclusive) page-index range
    # occupied by the TOC itself. The bookmark builder uses it to point
    # the Contents branch header at the TOC page; the section-heading
    # detector uses it to skip TOC pages during its scan (heading text
    # inside the TOC is already represented by the Contents children).
    toc_entries: list = []
    toc_page_range = None
    if not skip_links:
        try:
            _, toc_entries, toc_page_range = _link_toc_entries(doc, log)
        except Exception as e:
            log.warning(f"  TOC linking failed (non-fatal): {e}")

    # Section-heading detection. Runs unconditionally on briefs (the
    # same skip_links gate that excludes declarations and complaints
    # excludes them here too). Always-on rather than a TOC fallback:
    # a brief that HAS a TOC may also have pre-numbered headings
    # ("SUMMARY OF ARGUMENT") that wouldn't appear in the TOC because
    # they precede the formal Roman-numeral structure. Those still
    # get bookmarked. The TOC page range is passed in so the scan
    # skips TOC pages (which would otherwise produce phantom heading
    # bookmarks for every TOC entry). Failure is non-fatal.
    section_entries: list = []
    if not skip_links:
        try:
            section_entries = _detect_section_headings(
                doc, log, toc_page_range=toc_page_range)
        except Exception as e:
            log.warning(f"  Section heading detection failed (non-fatal): {e}")

    # Link body references to exhibit pages — this runs regardless of the
    # filename skip rules, because the document type most likely to need it
    # is a Declaration ("Attached hereto as Exhibit 1 is..."), which IS on
    # the skip list. The pass gates itself on the number of detected
    # exhibit covers (>= 2) inside _link_exhibit_references. Failure here
    # is non-fatal.
    exhibit_cover_map: dict = {}
    try:
        _, exhibit_cover_map = _link_exhibit_references(doc, log)
    except Exception as e:
        log.warning(f"  Exhibit linking failed (non-fatal): {e}")

    # Detect paragraph anchors for the bookmark builder's "Paragraphs" branch.
    # (The tool no longer stamps invisible right-margin markers into the page;
    # this pass only reads structure and never modifies the PDF.) Failure here
    # is non-fatal: the linked PDF is still useful without paragraph bookmarks.
    paragraph_anchors: list = []
    try:
        paragraph_anchors = _detect_paragraph_anchors(doc)
    except Exception as e:
        log.warning(f"  Paragraph-anchor detection failed (non-fatal): {e}")

    # Cause-of-action detection and linking for complaints (and amended
    # complaints — FAC/SAC/TAC). Runs on its own gate, independent of
    # `skip_links` (citation/TOC linking is still suppressed for these
    # files — only the cause-of-action structure is added). Produces:
    #   - bookmark entries for the "Causes of Action" branch
    #   - GOTO links from cover-page listings to body headings, when
    #     the cover lists causes in the same ALL-CAPS form.
    # Failure is non-fatal: bookmarks and citation linking still ship
    # even if cause detection fails.
    cause_entries: list = []
    if is_complaint(pdf_path.name):
        try:
            cause_entries, cover_occs, body_occs = _detect_causes_of_action(
                doc, log)
            if cause_entries:
                log.info(f"  Detected {len(cause_entries)} cause(s) of action")
                _link_cover_to_causes(doc, cover_occs, body_occs, log)
        except Exception as e:
            log.warning(f"  Cause-of-action detection failed (non-fatal): {e}")

    # Detect combined-filing sub-documents by unique footer text. Runs
    # on every file; if there's only one distinct footer (typical for
    # standalone briefs and declarations), the function returns []
    # and no Documents branch appears. Failure is non-fatal.
    document_entries: list = []
    try:
        document_entries = _detect_document_footers(doc, log)
    except Exception as e:
        log.warning(f"  Footer-based document detection failed (non-fatal): {e}")

    # Build the PDF outline (sidebar bookmarks) from the structure the
    # three passes above collected. set_toc replaces atomically, so this
    # is idempotent across re-runs and discards any pre-existing outline
    # the source PDF came with. Failure is non-fatal.
    try:
        _set_bookmarks(doc, toc_entries, exhibit_cover_map,
                       paragraph_anchors, log, cause_entries=cause_entries,
                       document_entries=document_entries,
                       section_entries=section_entries,
                       toc_page_range=toc_page_range)
    except Exception as e:
        log.warning(f"  Bookmark build failed (non-fatal): {e}")

    try:
        # Undo PyMuPDF's annotation-naming splice: insert_link names each link
        # by str-replacing the FIRST "/Link" in the annot source with
        # "/Link/NM(fitz-Ln)" — and a Westlaw URL contains "/Link/", so every
        # westlaw deep link saved as ".../Link/NM(fitz-L0)/Document/..." (a
        # 404). Strip the splice from any URI carrying it; this also heals
        # links a prior run corrupted.
        _repair_link_uris(doc)
        # Stamp it as linked so a later re-run can skip this pass (regenerating
        # only the .txt) unless --relink is given.
        _pdf_stamp_linked(doc)
        # Save to temp file first
        doc.save(str(temp_path), garbage=3, deflate=True)
        log.info(f"  Saved to temp: {temp_path.name}")
    except Exception as e:
        log.error(f"  Could not save linked PDF: {e}")
        return False
    finally:
        doc.close()

    # Replace original with the linked version
    try:
        import shutil
        shutil.move(str(temp_path), str(out_path))
        log.info(f"  Replaced original: {out_path.name}")
    except Exception as e:
        log.error(f"  Could not replace original PDF: {e}")
        return False

    return True


# ────────────────────────────────────────────────────────────────────────────
# Config file (edit settings without touching the code)
# ────────────────────────────────────────────────────────────────────────────
# A plain "key = value" file named pdf_linker.config, kept next to this script.
# It lets settings (currently: pseudonymize on/off) be changed without editing
# the code or the command that launches the tool. A command-line flag always
# wins over the config file.
_CONFIG_TRUE = {"1", "true", "yes", "on", "y", "t"}
_CONFIG_FALSE = {"0", "false", "no", "off", "n", "f"}
_CONFIG_TEMPLATE = (
    "# pdf_linker settings. Edit the values below to change behaviour without\n"
    "# touching the code. Lines starting with # are comments. A command-line\n"
    "# flag (e.g. --no-pseudonymize) overrides the matching setting here.\n"
    "\n"
    "# Pseudonymize the .txt exports? on/off (default: on). When on, party /\n"
    "# attorney names, case numbers, and detected PII in the .txt exports are\n"
    "# swapped for stable fakes using the newest Order*.xlsx in Downloads; the\n"
    "# PDFs themselves are never modified.\n"
    "#\n"
    "# Forgot a PDF from the batch? Drop it and the run's pseudonym_key.xlsx in a\n"
    "# folder and run again (or point --key at the key): the tool recognizes its\n"
    "# own key and reproduces the same fakes as the batch, so the new .txt stays\n"
    "# consistent. New values in that file get fresh fakes and are added to the\n"
    "# key.\n"
    "pseudonymize = on\n"
    "\n"
    "# Match a TRUNCATED party name (the front of it) when a two-column page\n"
    "# splices the caption? on/off (default: off). A prefix term is only\n"
    "# registered when it never appears in ordinary prose in this case, but a\n"
    "# false replacement is silent, so review the key before trusting it.\n"
    "partial_names = off\n"
    "\n"
    "# Subfolder (inside each case folder) that the .txt exports are written to.\n"
    "# Default: Text Files.\n"
    "text_subfolder = Text Files\n"
    "\n"
    "# Word documents (.docx/.docm) in a folder that has NO PDFs are bulk-\n"
    "# converted to scrubbed .txt exports too (never hyperlinked). A Word doc\n"
    "# beside PDFs is left alone, so the usual PDF workflow is untouched. Legacy\n"
    "# binary .doc files can't be read (re-save them as .docx).\n"
    "\n"
    "# Also write the UNSCRUBBED text to a second folder, for QA / your own\n"
    "# reference? on/off (default: off). The pseudonymized folder above stays\n"
    "# the shareable one; this second folder holds the ORIGINAL text under the\n"
    "# real filename and, by design, contains real names — it is never checked\n"
    "# by the leak gate and must NOT be shared. Diff the two folders to confirm\n"
    "# the scrub is clean and nothing load-bearing (a citation, a date) changed.\n"
    "keep_original_text = off\n"
    "# Name of that second folder. The default flags it so it isn't mistaken\n"
    "# for the shareable one.\n"
    "original_text_subfolder = Original Text (real names - do not share)\n"
    "\n"
    "# How close (in points) two caption-column left edges may be and still be\n"
    "# treated as ONE page column. Default: 30. Raise it if a two-column caption\n"
    "# splices its columns together in the .txt (a REVIEW warning names the page).\n"
    "column_band_tol = 30\n"
    "\n"
    "# What a surviving real value does to the run. The pseudonymization is a\n"
    "# PRECAUTION against casual recognition of a public filing, so the gate is\n"
    "# tiered:  primary (default) quarantines the exports only when a FULL\n"
    "# party/entity/attorney name or defined short name survives (that defeats\n"
    "# the whole purpose); lesser survivors (a bare token, an identifier) are\n"
    "# warned about and marked in the key but the exports are delivered.\n"
    "# strict quarantines on ANY surviving value; off never quarantines.\n"
    "leak_gate = primary\n"
    "\n"
    "# Accumulate every flagged leak into ONE master spreadsheet across all runs\n"
    "# and case folders, so you can spot a value that keeps leaking over time.\n"
    "# on/off (default: off). This toggle controls the 'Master Leaks' TALLY sheet\n"
    "# only. The 'KEEP' sheet in the SAME workbook — the durable no/[bracket]\n"
    "# decisions, applied across every folder and run — is always maintained (it\n"
    "# is the preservation vehicle for those decisions), so the workbook may be\n"
    "# created for KEEP even with the tally off.\n"
    "master_leaks = off\n"
    "# Where that master workbook lives (both the KEEP and Master Leaks sheets).\n"
    "# Default: master_leaks.xlsx next to this config file (or the PDF_LINKER_MASTER\n"
    "# env var). Point it anywhere stable (e.g. a OneDrive path) so it persists.\n"
    "# master_leaks_path = C:\\Users\\you\\Documents\\Master Leaks.xlsx\n"
)


def _config_path():
    try:
        return Path(__file__).with_name("pdf_linker.config")
    except NameError:
        return Path("pdf_linker.config")


def _read_config(log=None):
    """Return {key: value} from pdf_linker.config next to the script. If the
    file doesn't exist, best-effort write a commented template so it's easy to
    find and edit, and return {}."""
    path = _config_path()
    if not path.is_file():
        try:
            path.write_text(_CONFIG_TEMPLATE, encoding="utf-8")
            if log:
                log.info(f"Wrote default config file: {path}")
        except OSError:
            pass
        return {}
    cfg = {}
    try:
        for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            cfg[k.strip().lower()] = v.strip()
    except OSError as e:
        if log:
            log.warning(f"Could not read config file {path}: {e}")
    return cfg


def _config_bool(cfg, key, default):
    v = cfg.get(key, "").strip().lower()
    if v in _CONFIG_TRUE:
        return True
    if v in _CONFIG_FALSE:
        return False
    return default


# ── Live progress marker ─────────────────────────────────────────────────────
# A run can take many minutes (an OCR-heavy declaration set is minutes per
# file), so a 0-byte marker file dropped beside pdf_linker.log carries the
# estimated FINISH TIME in its NAME — visible at a glance in the folder without
# opening anything ('ETA ~6.04PM (6 of 13).txt'). It is rewritten after each
# PDF and, when the run finishes cleanly, renamed to a 'DONE <clock>.txt' stamp
# of the actual finish time; an ETA marker still left behind means the run
# didn't complete. A clock time (not a shrinking "~10 min") reads better
# because it only refreshes once per file. The name is colon-free so it is
# valid on Windows.
_ETA_MARKER_PREFIX = "ETA"
_DONE_MARKER_PREFIX = "DONE"


def _fmt_clock(dt):
    """A Windows-legal 12-hour clock string with no colon: 5:55 PM -> '5.55PM'."""
    h = dt.hour % 12 or 12
    return f"{h}.{dt.minute:02d}{'AM' if dt.hour < 12 else 'PM'}"


def _clear_eta_markers(folder):
    # Scoped to EMPTY .txt files matching a run-status marker shape, so the short
    # "ETA "/"DONE " prefix can never delete a real (non-empty) file a user named
    # similarly. Both prefixes are cleared, so a stale DONE stamp from a prior
    # run doesn't sit beside a fresh ETA (and vice versa).
    for prefix in (_ETA_MARKER_PREFIX, _DONE_MARKER_PREFIX):
        for m in folder.glob(prefix + " *.txt"):
            try:
                if m.is_file() and m.stat().st_size == 0:
                    m.unlink()
            except OSError:
                pass


def _write_eta_marker(folder, label):
    """Replace the live ETA marker with a fresh 0-byte file whose NAME is the
    estimate, e.g. 'ETA ~6.04PM (6 of 13).txt'. Best-effort: a marker that
    can't be written (locked folder, etc.) is never fatal."""
    _clear_eta_markers(folder)
    try:
        (folder / f"{_ETA_MARKER_PREFIX} {label}.txt").write_text(
            "", encoding="utf-8")
    except OSError:
        pass


def _write_done_marker(folder):
    """Replace the live ETA marker with a 0-byte 'DONE <clock>.txt' stamp on a
    clean finish, so the folder shows the actual finish time at a glance instead
    of the marker simply vanishing. Best-effort, same as the ETA marker."""
    _clear_eta_markers(folder)
    try:
        (folder / f"{_DONE_MARKER_PREFIX} {_fmt_clock(datetime.datetime.now())}.txt"
         ).write_text("", encoding="utf-8")
    except OSError:
        pass


# Relative processing cost of a page for the ETA. A page that must be OCR'd
# runs seconds; a page with a usable text layer is cheap. Weighting by these
# (instead of file bytes) is what keeps the estimate from lurching when a
# scanned declaration comes up — its cost is priced in from the first estimate
# rather than tanking the bytes/sec rate mid-run. The exact ratio need not be
# precise; the sec-per-unit self-calibrates, the ratio only proportions the
# remaining work.
_WORK_OCR_PAGE = 40.0
_WORK_TEXT_PAGE = 1.0


def _pdf_work_weight(pdf):
    """A cheap per-file cost weight for the ETA: count pages that will need OCR
    (empty or garbled text layer — the same test the re-OCR pass uses) heavily,
    text pages lightly. Metadata only — no rendering, no OCR — so it is fast
    even on a big scanned set. Returns None if the file can't be opened (caller
    falls back to a bytes-based weight)."""
    try:
        import fitz
        ocr = text = 0
        with fitz.open(pdf) as doc:
            for page in doc:
                t = page.get_text("text")
                if not t.strip() or _text_looks_garbled(t):
                    ocr += 1
                else:
                    text += 1
        return (ocr * _WORK_OCR_PAGE + text * _WORK_TEXT_PAGE) or _WORK_TEXT_PAGE
    except Exception:
        return None


# Remembered processing throughput (work-units per second) from the last run,
# stored next to the config. It lets a re-run show a projected finish time
# IMMEDIATELY — before the first file lands — instead of sitting on
# "(estimating...)" while a large case's first file is processed. Refined live
# each file and re-saved at the end, so it self-calibrates to this machine.
_ETA_RATE_FILE = "pdf_linker_eta_rate.txt"


def _eta_rate_path():
    return _config_path().with_name(_ETA_RATE_FILE)


def _load_eta_rate():
    try:
        v = float(_eta_rate_path().read_text(encoding="utf-8").strip())
        return v if v > 0 else None
    except Exception:
        return None


def _save_eta_rate(rate):
    try:
        if rate and rate > 0:
            _eta_rate_path().write_text(f"{rate:.6f}", encoding="utf-8")
    except Exception:
        pass


# Remembered --fix-leaks throughput (input BYTES per second), stored separately
# from the full-run rate because the two measure different work: fix-leaks is a
# pure text scrub (no OCR, no PDF), so its bytes/sec is far higher and would
# skew the OCR-weighted full-run rate if they shared a file. A conservative
# default seeds the FIRST fix-leaks run's estimate; it self-calibrates after.
_FIX_ETA_RATE_FILE = "pdf_linker_fixleaks_rate.txt"
_FIX_ETA_DEFAULT_RATE = 1_000_000.0     # bytes/sec, first-run seed only


def _fix_rate_path():
    return _config_path().with_name(_FIX_ETA_RATE_FILE)


def _load_fix_rate():
    try:
        v = float(_fix_rate_path().read_text(encoding="utf-8").strip())
        return v if v > 0 else None
    except Exception:
        return None


def _save_fix_rate(rate):
    try:
        if rate and rate > 0:
            _fix_rate_path().write_text(f"{rate:.6f}", encoding="utf-8")
    except Exception:
        pass


# ── One-click re-run launcher ────────────────────────────────────────────────
# After a run, drop a double-clickable file in the folder that re-runs the tool
# on THIS same folder — the fast path for applying the Fix? decisions saved in
# LEAKS.xlsx, or just re-processing. It targets its own directory
# (%~dp0 on Windows, $(dirname "$0") on POSIX) so it keeps working if the folder
# is moved, and points --key at the folder's own key so the re-run reproduces
# the same fakes (and applies the worksheet decisions on top).
def _rerun_launcher_spec(exe, script, provider, want_key, windows, frozen=False):
    """(filename, content, make_executable) for the re-run launcher.

    `frozen` True means the tool runs as a packaged executable (PyInstaller /
    py2exe): the exe IS the entry point, so the script path must NOT be passed
    (a frozen exe treats a stray .py argument as an unknown target and can open
    an empty console). Otherwise the exe is a Python interpreter and needs the
    script. A running banner and a trailing status line are always printed, so
    the window is never blank even if the interpreter writes nothing itself."""
    prog = f'"{exe}"' if frozen else f'"{exe}" "{script}"'
    if windows:
        key = ' --key "%~dp0pseudonym_key.xlsx"' if want_key else ""
        # Detached background run: `start` hands the work to a separate process
        # and the launcher returns immediately, so no console sits blocking on a
        # `pause`. There is no blackout — the tool appends progress to
        # pdf_linker.log throughout, and stamps a "DONE <time>.txt" file in the
        # folder on a clean finish. The interpreter is the WINDOWLESS pythonw.exe
        # where available (see _write_rerun_launcher), so nothing lingers on
        # screen; a short timeout leaves the confirmation readable, then the
        # launcher window closes on its own.
        content = (
            "@echo off\r\n"
            "REM PDF-Linker re-run - double-click to process this folder in the\r\n"
            "REM background. Picks up any Fix? decisions saved in LEAKS.xlsx.\r\n"
            "REM Runs detached: the work continues after this window closes.\r\n"
            'REM Progress -> pdf_linker.log; a "DONE <time>.txt" file appears\r\n'
            "REM in this folder when the run finishes.\r\n"
            "echo Re-running PDF-Linker in the background (this window closes; the\r\n"
            "echo work keeps running)...\r\n"
            "echo   Progress:  see pdf_linker.log in this folder\r\n"
            'echo   When done: a "DONE ....txt" file appears here (not yet - it is\r\n'
            "echo              still running when this window closes).\r\n"
            f'start "PDF-Linker re-run" {prog} "%~dp0." --provider {provider}{key}\r\n'
            "timeout /t 3 >nul 2>&1\r\n")
        return "Re-run PDF-Linker.bat", content, False
    key = ' --key "$(dirname "$0")/pseudonym_key.xlsx"' if want_key else ""
    content = (
        "#!/bin/sh\n"
        "# PDF-Linker re-run - double-click to process this folder again.\n"
        "# Picks up any Fix? decisions saved in LEAKS.xlsx.\n"
        'echo "Re-running PDF-Linker on this folder..."\n'
        f'{prog} "$(dirname "$0")" --provider {provider}{key}\n'
        'echo "Finished - exit code $?."\n')
    return "Re-run PDF-Linker.command", content, True


def _write_rerun_launcher(folder, provider, want_key, log):
    """Write/refresh the double-click re-run launcher in `folder`. Best-effort;
    a failure is never fatal to the run."""
    exe = Path(sys.executable)
    frozen = bool(getattr(sys, "frozen", False))
    if os.name == "nt" and not frozen:
        # The re-run .bat launches the tool DETACHED (via `start`), so it wants
        # the WINDOWLESS interpreter — pythonw.exe — or a bare console flashes
        # for the length of the run. Prefer the pythonw beside sys.executable;
        # if it isn't there, keep whatever we have (the detached run still works,
        # it just shows a console). Progress goes to pdf_linker.log regardless.
        low = exe.name.lower()
        if low.startswith("python") and not low.startswith("pythonw"):
            cand = exe.with_name("pythonw" + exe.name[len("python"):])
            if cand.exists():
                exe = cand
    script = Path(__file__).resolve()
    name, content, make_exec = _rerun_launcher_spec(
        str(exe), str(script), provider, want_key, os.name == "nt", frozen=frozen)
    try:
        path = folder / name
        path.write_text(content, encoding="utf-8", newline="")
        if make_exec:
            os.chmod(path, 0o755)
        log.info(f"  Wrote one-click re-run launcher: {name}")
    except OSError as e:
        log.warning(f"  Could not write re-run launcher: {e}")


# ── Protected-vocabulary inventory ───────────────────────────────────────────
# The pseudonymizer PROTECTS a fixed vocabulary: words it never fakes and never
# flags as a name, plus citation SHAPES it never rewrites a party name inside.
# `--dump-terms` prints it from these live constants (the source of truth), so
# the inventory can't drift from a hand-kept doc. The word sets are:
#   _PN_PLEADING_WORDS   document types / procedural vocabulary
#   _PN_DECL_DESCRIPTOR  words that precede "Declaration" without naming anyone
#   _PN_SHORT_TOKEN_STOP two-letter English words never registered as a surname
# and the reference words "declaration/decl/dec" are hard-blocked from faking.
def _dump_protected_terms(out=None):
    """Print the protected vocabulary from the live constants. Returns the text
    (also written to `out`, default stdout) so a test can assert on it."""
    import io
    buf = io.StringIO()

    def block(title, note, words):
        buf.write(f"\n{title} ({len(words)})\n  {note}\n")
        line = "   "
        for w in sorted(words):
            if len(line) + len(w) + 2 > 78:
                buf.write(line.rstrip() + "\n")
                line = "   "
            line += " " + w + ","
        buf.write(line.rstrip().rstrip(",") + "\n")

    buf.write("PDF-Linker protected vocabulary "
              "(never faked, never flagged as a name)\n")
    block("Pleading / document-type words  [_PN_PLEADING_WORDS]",
          "A phrase made only of these (plus roles/connectors) is a document, "
          "not a person.", _PN_PLEADING_WORDS)
    block("Declaration descriptors  [_PN_DECL_DESCRIPTOR]",
          "Precede 'Declaration' without naming the declarant "
          "('Supporting Declaration').", _PN_DECL_DESCRIPTOR)
    block("Two-letter word guard  [_PN_SHORT_TOKEN_STOP]",
          "Never registered as a surname, so prose ('as', 'of') is not "
          "rewritten.", _PN_SHORT_TOKEN_STOP)
    buf.write("\nHard never-fake reference words (3)\n"
              "  the declaration reference word itself is never a term:\n"
              "   dec, decl, declaration\n")
    buf.write("\nCourt-form boilerplate  [_PN_NEVER_FAKE]\n"
              "  a form number / field label is never faked or flagged (matched\n"
              "  on an alphanumeric, case-folded reduction):\n"
              "   CASE NUMBER, CIV-100, Default Only\n")
    buf.write("\nCitation shapes protected from party-name rewriting "
              "(regex families, not word lists):\n"
              "   In re <Name> ... Litig.             [_PN_INRE_LITIG_RE]\n"
              "   <Name> Warranty Cases / Warranty    [_PN_CASES_RUN_RE]\n"
              "   <name> <year> WL <number> runs      [_PN_WL_RUN_RE]\n"
              "   reporter cites with multiple pins   [REPORTER_PART]\n")
    text = buf.getvalue()
    (out or sys.stdout).write(text)
    return text


# ── Leak-fix launcher (companion to the re-run launcher) ─────────────────────
def _fix_launcher_spec(exe, script, windows, frozen=False):
    """(filename, content, make_executable) for the 'Apply Leak Fixes' launcher —
    a double-clickable file that runs --fix-leaks on this folder."""
    prog = f'"{exe}"' if frozen else f'"{exe}" "{script}"'
    if windows:
        content = (
            "@echo off\r\n"
            "REM PDF-Linker - apply the Fix? decisions saved in\r\n"
            "REM LEAKS.xlsx to the .txt/.LEAK exports (no PDFs touched).\r\n"
            "echo Applying leak fixes to this folder...\r\n"
            "echo.\r\n"
            f'{prog} "%~dp0." --fix-leaks --key "%~dp0pseudonym_key.xlsx"\r\n'
            "echo.\r\n"
            "echo Finished - exit code %ERRORLEVEL%. You can close this window.\r\n"
            "pause\r\n")
        return "Apply Leak Fixes.bat", content, False
    content = (
        "#!/bin/sh\n"
        "# PDF-Linker - apply LEAKS.xlsx Fix? decisions to the exports.\n"
        'echo "Applying leak fixes to this folder..."\n'
        f'{prog} "$(dirname "$0")" --fix-leaks '
        '--key "$(dirname "$0")/pseudonym_key.xlsx"\n'
        'echo "Finished - exit code $?."\n')
    return "Apply Leak Fixes.command", content, True


def _pn_fix_launcher_paths(folder):
    """Both possible 'Apply Leak Fixes' launcher paths — a folder synced across
    OSes may carry either the .command or the .bat."""
    return [folder / "Apply Leak Fixes.command", folder / "Apply Leak Fixes.bat"]


def _pn_remove_leak_workflow(folder, log):
    """Delete the leak-triage worksheet (under every name it has used) and the
    Apply-Leak-Fixes launcher. Called once no quarantined *.LEAK export remains:
    the leak workflow is resolved, so its own files must not linger and imply
    there is still something to triage. Best-effort; returns the names removed."""
    victims = [_pn_leak_xlsx_path(folder), _pn_leak_txt_path(folder)]
    victims += [folder / f"{stem}.xlsx" for stem in _PN_LEAK_LEGACY_STEMS]
    victims += [folder / f"{stem}.txt" for stem in _PN_LEAK_LEGACY_STEMS]
    victims += _pn_fix_launcher_paths(folder)
    removed = []
    for p in victims:
        try:
            if p.exists():
                p.unlink()
                removed.append(p.name)
        except OSError:
            pass
    return removed


def _write_fix_launcher(folder, log):
    """Write/refresh the double-click 'Apply Leak Fixes' launcher. Best-effort."""
    exe = Path(sys.executable)
    frozen = bool(getattr(sys, "frozen", False))
    if os.name == "nt" and not frozen and exe.stem.lower().startswith("pythonw"):
        cand = exe.with_name(exe.name.lower().replace("pythonw", "python", 1))
        exe = cand if cand.exists() else Path("python.exe")
    script = Path(__file__).resolve()
    name, content, make_exec = _fix_launcher_spec(
        str(exe), str(script), os.name == "nt", frozen)
    try:
        path = folder / name
        path.write_text(content, encoding="utf-8", newline="")
        if make_exec:
            os.chmod(path, 0o755)
        log.info(f"  Wrote leak-fix launcher: {name}")
    except OSError as e:
        log.warning(f"  Could not write leak-fix launcher: {e}")


def _fix_leaks_mode(folder, args, cfg, log):
    """Apply the leak-worksheet Fix?=yes decisions DIRECTLY to the already-
    written .txt / .LEAK exports — no PDFs reopened — then un-quarantine files
    that no longer carry a party-name leak. The fast follow-up to a full run.

    Reuses the hardened core (key load, registry, apply() with citation
    protection, the per-file leak gate), so a fix is the same deterministic,
    reversible fake the tool would have written and never corrupts a citation.
    Returns a process exit code."""
    text_subdir = cfg.get("text_subfolder", "").strip() or "Text Files"
    text_dir = folder / text_subdir
    if not text_dir.is_dir():
        text_dir = folder            # older single-folder layout

    # log-only: the console log handler (attached in main) already echoes these
    # to the terminal — an extra print doubled every message in the window.
    _warn = log.warning

    key_path = Path(args.key) if args.key else (folder / "pseudonym_key.xlsx")
    if not (key_path.is_file() and _pn_key_looks_like_ours(key_path)):
        _warn("--fix-leaks needs this tool's pseudonym_key.xlsx beside the "
              "exports so the fixes stay consistent and reversible; none found "
              f"at {key_path}.")
        return 1

    registry = _PnFakeRegistry()
    try:
        terms, key_decisions = _pn_load_key(key_path, registry, log)
    except RuntimeError as e:
        _warn(f"--fix-leaks: could not read key {key_path.name}: {e}")
        return 1

    master_keep = _pn_read_master_keep(cfg)
    folder_decisions = _pn_read_leak_decisions(folder)
    decisions = {**master_keep, **folder_decisions}
    # KEEP / KEEP-PART edits from the key's Replacement column ('no' or a
    # [bracketed] keep-spec) are authoritative for this run.
    for vl, d in key_decisions.items():
        decisions[vl] = d
    # A master row this folder authored is still OURS: the local LEAKS.xlsx is
    # consumed once resolved, so a decision made here survives only there.
    local_vls = (set(key_decisions) | set(folder_decisions)
                 | {vl for vl, d in decisions.items()
                    if _pn_decision_is_ours(d, folder.name)})
    # Two kinds of "yes": AUTO (let the tool mint a fake) and EXPLICIT (the
    # operator typed the exact replacement into the Fix? cell). An explicit fix
    # bypasses the faker entirely — the operator's deliberate choice of a
    # specific fake, and the cure for anything the faker would handle poorly.
    # Fakes this key already handed out, so a HALF-SCRUBBED flagged value
    # ("Melissa Sable", where only "Penuela" was bound) fakes just its real
    # remainder instead of re-faking our own stand-in — see _pn_strip_prior_fakes.
    prior_fakes = _pn_prior_fake_words(terms)
    auto_terms, explicit = [], {}
    for vl, d in decisions.items():
        # An INHERITED decision contributes its KEEP and nothing else — the
        # bracket generalises, the remainder is the authoring case's party.
        if d["fix"] != "yes" or vl not in local_vls:
            continue
        # Bracketed keep-spec: auto-fake only the non-bracketed fragment(s); the
        # bracketed part is left verbatim (never registered, so never touched).
        if d.get("fake_values") is not None:
            auto_terms.extend(d["fake_values"])
            continue
        repl = (d.get("replacement") or "").strip()
        if repl and repl.lower() != d["value"].strip().lower():
            explicit[d["value"]] = repl
            if _pn_strip_prior_fakes(d["value"], prior_fakes) != d["value"]:
                log.warning(
                    f"  --fix-leaks: typed replacement for {d['value']!r} — part "
                    f"of that value is already a fake this key minted, so the key "
                    f"will carry it as a Real Value. Type the fix against the "
                    f"REAL name instead, or mark the row yes to auto-fake only "
                    f"the un-scrubbed part.")
        elif repl:
            log.warning(f"  --fix-leaks: typed replacement for {d['value']!r} "
                        f"equals the value itself — ignoring (a self-map never "
                        f"scrubs). Type a DIFFERENT replacement.")
        else:
            auto_terms.append(d["value"])
    auto_terms = _pn_drop_prior_fakes_from_terms(auto_terms, prior_fakes, log)
    suppressed = {vl for vl, d in decisions.items() if d["fix"] == "no"}
    # A LOCAL keep decision (this folder's key or LEAKS — not one merely
    # inherited from the master sheet) retires the key's own binding for that
    # value, so the key this run rewrites carries only corrected values.
    terms, retired_reals = _pn_retire_kept_key_terms(
        terms, {vl: d for vl, d in decisions.items() if vl in local_vls},
        registry, log)
    fix_terms = auto_terms + list(explicit)
    if not fix_terms:
        # Nothing to apply. If no quarantined *.LEAK export remains either, the
        # workflow is already resolved — clean up its worksheet and launcher so
        # a done folder doesn't keep both around.
        if not any(p.name.endswith(".txt.LEAK") for p in text_dir.iterdir()
                   if p.is_file()):
            removed = _pn_remove_leak_workflow(folder, log)
            if removed:
                log.info("  --fix-leaks: nothing left to fix — removed "
                         + ", ".join(removed) + ".")
                return 0
        log.info("--fix-leaks: no Fix?=yes rows in LEAKS.xlsx — "
                 "nothing to apply. Mark the leaks you want scrubbed (yes, or "
                 "type the exact replacement) and double-click again.")
        return 0

    terms += _pn_build_terms([], [], list(args.term or []) + auto_terms, registry)
    # Operator-typed replacements: authoritative, reversible, never re-derived.
    # Appended LAST so they win Pseudonymizer.__init__'s last-write-wins records
    # map over any binding loaded from the key or rebuilt here.
    for real, fake in explicit.items():
        t = _PnTerm("person", real, fake, whole_word=False,
                    case_sensitive=False, priority=3, source="leak-fix")
        t.loaded = True          # persist into the key so it survives re-runs
        terms.append(t)
    # No PII detectors: the exports were already scrubbed on the full run, so a
    # detector meeting a fake here could re-fake it. Only the term-based scrub
    # (party names + the flagged values) is wanted.
    pz = Pseudonymizer(terms, [], registry)
    pz.suppressed = suppressed
    pz.keep_strict, pz.keep_soft = _pn_keep_values(decisions)
    pz.keep_strict_local = _pn_keep_values(
        {vl: d for vl, d in decisions.items() if vl in local_vls})[0]
    pz._keep_decisions = {vl: d for vl, d in decisions.items()
                          if _pn_decision_is_keep(d)}
    pz._keep_local = {vl for vl in pz._keep_decisions if vl in local_vls}
    pz._retired_key_rows = retired_reals
    pz.override_fixes(explicit)   # typed fake beats any other record for a value
    for r in pz.records.values():
        pz._own_fakes.add(str(r["fake"]).lower().rstrip(" .,;:"))
    # Record this run's LOCAL keep decisions into the cross-folder master KEEP
    # sheet so a key/worksheet edit made here persists globally.
    _keep_rec = {vl: d for vl, d in decisions.items()
                 if _pn_decision_is_keep(d) and vl in local_vls}
    if _keep_rec:
        _pn_update_master_keep(cfg, _keep_rec, folder.name,
                               datetime.date.today().isoformat(), log)

    def _is_tool_artifact(p):
        """The tool's OWN .txt files in the folder — the worksheet's text
        companion and the ETA/DONE run markers — are not exports and must not
        be scrubbed or tracked (only reachable in the old single-folder layout,
        where text_dir falls back to the case folder itself)."""
        leak_txts = {f"{_PN_LEAK_STEM}.txt"} | {
            f"{stem}.txt" for stem in _PN_LEAK_LEGACY_STEMS}
        return (p.name in leak_txts
                or ((p.name.startswith(_ETA_MARKER_PREFIX + " ")
                     or p.name.startswith(_DONE_MARKER_PREFIX + " "))
                    and p.stat().st_size == 0))

    files = sorted(p for p in text_dir.iterdir()
                   if p.is_file() and (p.suffix == ".txt"
                                       or p.name.endswith(".txt.LEAK"))
                   and not _is_tool_artifact(p))

    # Projected-finish marker for the apply pass, same "ETA <clock>.txt" file the
    # full run drops so the operator sees how long Apply Leak Fixes will take.
    # The scrub is pure text (no OCR), so cost tracks input bytes; the rate
    # self-calibrates from the last fix run (a conservative default seeds the
    # first). Replaced by a DONE stamp when the pass finishes.
    total_bytes = 0
    for f in files:
        try:
            total_bytes += f.stat().st_size
        except OSError:
            pass
    rate = _load_fix_rate() or _FIX_ETA_DEFAULT_RATE
    finish = datetime.datetime.now() + datetime.timedelta(
        seconds=total_bytes / rate if rate else 0)
    _write_eta_marker(folder, f"~{_fmt_clock(finish)} (applying leak fixes)")
    _fix_start = time.monotonic()

    changed = 0
    for f in files:
        try:
            body = f.read_text(encoding="utf-8")
        except Exception as e:
            log.warning(f"  Could not read {f.name}: {e}")
            continue
        is_leak = f.name.endswith(".txt.LEAK")
        scrubbed = pz.apply(body)
        if is_leak:
            # A quarantined file may have been held for a WELDED leak the
            # whole-word patterns can't see ("FORDMOTORCOMPANYDEFENDANT" on a
            # spliced caption). Cure it the way the full run cures a flagged
            # page — quarantine already marks the file as human-review, the
            # precondition scrub_welded requires.
            scrubbed = pz.scrub_welded(scrubbed)
        # Compare against the NFKC form: apply() normalizes unconditionally, so
        # a file whose only difference is normalization has no actual fix in it
        # and is left untouched (not rewritten, not counted as changed).
        if scrubbed != _NFKC(body):
            try:
                f.write_text(scrubbed, encoding="utf-8", newline="\n")
                changed += 1
            except Exception as e:
                log.warning(f"  Could not write {f.name}: {e}")
        survivors = set(pz.surviving_reals(scrubbed))
        if is_leak:
            # ...and gate the un-quarantine on the reduced scan too, so a
            # welded survivor the cure couldn't fix keeps the file held
            # instead of being delivered as "clean".
            survivors |= set(pz.surviving_reals_reduced(scrubbed))
        pz.written.append(f)
        if survivors:
            pz.leaked_by_file[f] = {s.lower() for s in survivors}
            pz.note_leaks(survivors)
            parsed = _pn_body_lines(scrubbed)
            for real in sorted(survivors):
                if _pn_is_email_value(real):
                    continue          # always faked — never a triage decision
                pz.leak_report.append(
                    {"file": f.name, "type": "LEAK", "value": real,
                     "where": _pn_locate(parsed, real)})

    # Un-quarantine every *.LEAK that no longer carries a party-name leak.
    # Parity with the main gate: a value the reviewer marked "no fix" never
    # blocks delivery here either.
    gating = {v for v in pz.primary_leaks()
              if str(v).lower() not in pz.suppressed}
    offenders = set(pz.files_to_quarantine(gating))
    unq = 0
    for f in files:
        if f.name.endswith(".txt.LEAK") and f not in offenders and f.exists():
            # The delivered file's NAME must be scrubbed too: the full run
            # pseudonymizes export names, but a quarantined file kept the name
            # it was written under — un-quarantining "Yu Decl.txt.LEAK" as
            # "Yu Decl.txt" shipped the declarant's name in the filename.
            base_stem = f.name[: -len(".txt.LEAK")]
            safe = _pn_scrubbed_stem(base_stem, pz, log)
            dest = f.with_name(safe + ".txt")
            if dest.exists():
                # A newer full run already wrote this export; the .LEAK is a
                # stale leftover — keep the newer file, drop the stale one.
                try:
                    f.unlink()
                    log.info(f"  Dropped stale quarantine {f.name}; "
                             f"{dest.name} from a later run supersedes it.")
                except OSError as e:
                    log.warning(f"  Could not remove stale {f.name}: {e}")
                continue
            try:
                f.replace(dest)
                unq += 1
            except OSError as e:
                log.warning(f"  Could not un-quarantine {f.name}: {e}")

    pz.write_key(key_path, log)                    # loaded rows preserved

    # Record this pass's throughput so the next run's ETA is sharper, then
    # replace the ETA marker with a DONE stamp (the actual finish time).
    _elapsed = time.monotonic() - _fix_start
    if _elapsed > 0 and total_bytes > 0:
        _save_fix_rate(total_bytes / _elapsed)
    _write_done_marker(folder)

    still = len(offenders)
    if still:
        _pn_write_leak_report(folder, pz.leak_report, log, decisions=decisions,
                              cfg=cfg,
                              bound=[r["real"] for r in pz.records.values()])
    else:
        # Every LEAK file is fixed: the worksheet and the Apply-Leak-Fixes
        # launcher have done their job — remove them instead of leaving stale
        # files that suggest there is still something to triage.
        removed = _pn_remove_leak_workflow(folder, log)
        if removed:
            log.info("  --fix-leaks: all leaks resolved — removed "
                     + ", ".join(removed) + ".")

    log.info(f"--fix-leaks: applied {len(fix_terms)} fix(es) to {changed} "
             f"file(s); {unq} export(s) un-quarantined (*.LEAK -> .txt)"
             + (f"; {still} file(s) still carry a party-name leak — review "
                f"LEAKS.xlsx." if still else "."))
    return 2 if still else 0


# ────────────────────────────────────────────────────────────────────────────
# Entry point
# ────────────────────────────────────────────────────────────────────────────
def main():
    # Backward-compatible CLI: positional <folder>, plus optional --provider.
    # The mail-merge macro calls `pythonw pdf_linker.py "<folder>"` with no
    # flag, which keeps the historical Westlaw-default behaviour.
    import argparse

    parser = argparse.ArgumentParser(
        description="Add citation hyperlinks to PDFs in a case folder."
    )
    parser.add_argument("folder", nargs="?",
                        help="Folder containing PDFs to link.")
    parser.add_argument(
        "--dump-terms", action="store_true",
        help="Print the protected vocabulary the pseudonymizer never fakes or "
             "flags as a name (and the citation shapes it never rewrites), then "
             "exit. No folder needed.",
    )
    parser.add_argument(
        "--relink", action="store_true",
        help="Re-run the citation-linking pass on PDFs already linked in a prior "
             "run. By default a re-run skips linking an already-linked PDF (only "
             "its .txt is regenerated); use this to pick up linking changes.",
    )
    parser.add_argument(
        "--fix-leaks", action="store_true",
        help="Apply the Fix?=yes decisions from LEAKS.xlsx directly to "
             "the .txt/.LEAK exports (no PDFs reopened) and un-quarantine files "
             "that are now clean. Fast follow-up to a full run; needs the "
             "folder's pseudonym_key.xlsx.",
    )
    parser.add_argument(
        "--provider",
        choices=("westlaw", "lexis"),
        default="lexis",
        help="Which service citations should resolve to (default: lexis). "
             "Selects the search-URL form used for each citation.",
    )
    parser.add_argument(
        "--no-text",
        dest="extract_text",
        action="store_false",
        help="Do not write a plain-text (.txt) companion next to each "
             "text-based PDF. By default a .txt export is written for every "
             "PDF that has a real text layer (scanned-only PDFs are skipped).",
    )
    # Pseudonymization — ON by default, applied to the .txt export ONLY, never
    # the PDF.
    # Pseudonymization on/off. Neither flag given -> fall back to the config
    # file (pdf_linker.config next to this script), else default ON. Either
    # flag overrides the config for this run.
    _ps = parser.add_mutually_exclusive_group()
    _ps.add_argument(
        "--pseudonymize", dest="pseudonymize", action="store_true", default=None,
        help="Force pseudonymization of the .txt exports ON for this run "
             "(overrides the config file).",
    )
    _ps.add_argument(
        "--no-pseudonymize", dest="pseudonymize", action="store_false",
        default=None,
        help="Force pseudonymization of the .txt exports OFF for this run "
             "(overrides the config file). By default the .txt exports swap "
             "party/case/attorney names (from the E-Court key spreadsheet) and "
             "detected PII for stable fakes; the PDFs are never modified.",
    )
    _pnp = parser.add_mutually_exclusive_group()
    _pnp.add_argument(
        "--partial-names", dest="partial_names", action="store_true", default=None,
        help="Also match a party name that appears truncated (e.g. a caption "
             "spliced by a two-column page). A prefix is registered only if it "
             "never occurs in ordinary prose in this case. Off by default.",
    )
    _pnp.add_argument(
        "--no-partial-names", dest="partial_names", action="store_false",
        default=None, help="Force partial (prefix) name matching OFF for this run.",
    )
    parser.add_argument(
        "--key", type=Path, default=None, metavar="XLSX",
        help="Spreadsheet with party/case/attorney columns (the E-Court "
             "order-template export) whose values are pseudonymized in the "
             ".txt exports. Default: the most recent Order*.xlsx in your "
             "Downloads folder.",
    )
    parser.add_argument(
        "--term", action="append", metavar="TEXT",
        help="An extra value to pseudonymize in the .txt exports (repeatable).",
    )
    parser.add_argument(
        "--name-column", action="append", metavar="HEADER",
        help="Additional --key header to treat as names (repeatable; "
             "'attorney*' columns are auto-included).",
    )
    parser.add_argument(
        "--detect", metavar="LIST",
        help="Comma-separated PII detectors to run while pseudonymizing "
             f"(default: {','.join(_PN_DEFAULT_DETECTORS)}; "
             f"available: {','.join(_PN_DETECTORS)}).",
    )
    parser.add_argument(
        "--no-detect", action="store_true",
        help="Disable regex PII detection (pseudonymize spreadsheet/--term "
             "values only).",
    )
    parser.add_argument(
        "--key-out", type=Path, default=None, metavar="XLSX",
        help="Where to write the real->fake key (default: "
             "<folder>/pseudonym_key.xlsx).",
    )
    args = parser.parse_args()

    # Child processes (tesseract per OCR page) must fail into the log, never
    # into a Windows error dialog per page.
    _suppress_child_error_dialogs()

    # `--dump-terms` is an inspection command: print the protected vocabulary
    # and exit, no folder required.
    if args.dump_terms:
        _dump_protected_terms()
        sys.exit(0)

    if not args.folder:
        parser.error("the following arguments are required: folder")
    folder = Path(args.folder)
    if not folder.is_dir():
        print(f"Not a folder: {folder}")
        sys.exit(1)

    log_path = folder / "pdf_linker.log"
    logging.basicConfig(
        filename=str(log_path),
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        filemode="a",
    )
    # Also stream progress to the console WHEN ONE EXISTS. Logging went only to
    # the file, so the re-run launcher's window showed nothing and read as an
    # "empty terminal". The normal mail-merge launch is pythonw.exe, which has no
    # console (sys.stdout is None) — guard on it so that path is unaffected.
    _console = getattr(sys, "stdout", None)
    if _console is not None:
        try:
            _ch = logging.StreamHandler(_console)
            _ch.setLevel(logging.INFO)
            _ch.setFormatter(logging.Formatter("%(message)s"))
            logging.getLogger().addHandler(_ch)
        except Exception:
            pass
    log = logging.getLogger("pdf_linker")
    log.info("=" * 60)
    log.info(f"Run started for folder: {folder} (provider={args.provider})")

    # Resolve pseudonymization on/off: an explicit --pseudonymize /
    # --no-pseudonymize flag wins; otherwise the pdf_linker.config file next to
    # the script; otherwise default ON.
    cfg = _read_config(log)

    # Leak-fix mode: apply the worksheet Fix? decisions to the .txt/.LEAK exports
    # directly and exit — no PDFs reopened, no linking. Fast follow-up path.
    if args.fix_leaks:
        log.info(f"--fix-leaks on folder: {folder}")
        sys.exit(_fix_leaks_mode(folder, args, cfg, log))

    if args.pseudonymize is None:
        args.pseudonymize = _config_bool(cfg, "pseudonymize", True)
        log.info(f"Pseudonymization {'ON' if args.pseudonymize else 'OFF'} "
                 f"(from {_config_path().name})")
    else:
        log.info(f"Pseudonymization {'ON' if args.pseudonymize else 'OFF'} "
                 f"(from command line)")

    # Subfolder (within each case folder) that the .txt exports are written to.
    text_subdir = cfg.get("text_subfolder", "").strip() or "Text Files"

    # Optional second folder holding the UNSCRUBBED text, for QA/reference. Off
    # by default so the tool's default output is only the shareable artifact.
    original_subdir = None
    if args.pseudonymize and _config_bool(cfg, "keep_original_text", False):
        original_subdir = (cfg.get("original_text_subfolder", "").strip()
                           or "Original Text (real names - do not share)")

    # Column-band tolerance for splitting two-column caption pages (points).
    global _COLUMN_BAND_TOL
    try:
        _COLUMN_BAND_TOL = float(cfg.get("column_band_tol", "").strip()
                                 or _COLUMN_BAND_TOL)
    except ValueError:
        pass

    def _warn(msg):
        # Surface a folder-level warning both to the log and the console.
        log.warning(msg)
        print(msg, file=sys.stderr)

    # Build a shared pseudonymizer once (one folder is one case, so its terms
    # apply to every .txt). ON by default; applies to the .txt exports only,
    # the PDFs are never modified.
    pseudonymizer = None
    key_path = None
    if args.pseudonymize:
        if args.no_detect:
            detectors = []
        elif args.detect:
            detectors = [d.strip() for d in args.detect.split(",") if d.strip()]
            bad = [d for d in detectors if d not in _PN_DETECTORS]
            if bad:
                print(f"Unknown detector(s): {', '.join(bad)}. "
                      f"Available: {', '.join(_PN_DETECTORS)}")
                sys.exit(1)
        else:
            detectors = list(_PN_DEFAULT_DETECTORS)
        # Key spreadsheet: explicit --key, else this folder's OWN key
        # (pseudonym_key.xlsx to reuse, or an Order*.xlsx template to start over
        # from), else the most recent Order*.xlsx in Downloads (where the E-Court
        # export lands). Folder-local wins so a re-run resolves its own inputs.
        #
        # The Downloads guess is withheld for an ALL-WORD folder: it has no
        # E-Court template of its own, so "newest .xlsx in Downloads" hands it
        # whatever case was downloaded last and the run scrubs against a
        # stranger's party list — this case's parties left in the clear while
        # another matter's names are hunted for, and its key written full of
        # values that were never here. An explicit --key or a template the
        # operator put IN the folder is unambiguous and still honoured.
        key_path = args.key or _pn_find_folder_key(folder, log)
        if key_path is None and not _is_word_only_folder(folder):
            key_path = _pn_find_downloads_key(log)
        registry = _PnFakeRegistry()
        # Decisions feeding this run come from two places:
        #   * this folder's LEAKS worksheet (transient genuine-leak triage), and
        #   * the cross-folder master KEEP sheet (the durable no/bracket
        #     decisions, remembered across every folder and run).
        # The folder's own decisions layer ON TOP of the global keeps, so a case
        # can locally override a global keep when it must.
        master_keep = _pn_read_master_keep(cfg)
        folder_decisions = _pn_read_leak_decisions(folder)
        leak_decisions = {**master_keep, **folder_decisions}
        if master_keep:
            log.info(f"  Master KEEP: {len(master_keep)} durable no/bracket "
                     f"decision(s) loaded from the cross-folder log.")
        # A bracketed keep-spec auto-fakes only its non-bracketed fragment(s);
        # every other Fix?=yes scrubs the whole flagged value.
        #
        # ONLY for a LOCAL decision (this folder's LEAKS triage). A decision
        # merely INHERITED from the cross-folder master sheet contributes its
        # KEEP and nothing else — that is the whole content of a keep-spec: the
        # bracket says "this fragment is never a name", a lesson that
        # generalises to every case. The REMAINDER does not: "Alder" of
        # "Alder Law, P.C. -> [Law]" is one matter's law firm, and hunting it
        # everywhere put it in every folder's log and (once unmatched
        # authoritative bindings were pinned) every folder's key, as a Real
        # Value belonging to a different case. A firm that genuinely recurs is
        # caught by the case's own party list or pre-scan — cross-case
        # inference is exactly what the closed-entity rule exists to prevent.
        # `_pn_keep_values` reads the decisions directly, so the bracketed part
        # is still protected here; only the faking half is withheld.
        # "Ours" = typed in this folder's LEAKS/key this run, OR a master row
        # this folder authored earlier (LEAKS.xlsx is consumed once resolved,
        # so a decision made here lives on only in the master sheet).
        ours = {vl for vl, d in leak_decisions.items()
                if vl in folder_decisions or _pn_decision_is_ours(d, folder.name)}
        fix_terms = []
        for vl, d in leak_decisions.items():
            if d["fix"] != "yes" or vl not in ours:
                continue
            fv = d.get("fake_values")
            fix_terms.extend(fv if fv is not None else [d["value"]])
        suppressed = {vl for vl, d in leak_decisions.items() if d["fix"] == "no"}
        extra_terms = list(args.term or []) + fix_terms
        if fix_terms:
            log.info(f"  Leak worksheet: scrubbing {len(fix_terms)} value(s) "
                     f"marked Fix?=yes ({', '.join(fix_terms[:6])}).")
        if suppressed:
            log.info(f"  Leak worksheet: {len(suppressed)} value(s) marked "
                     f"Fix?=no will be left as-is and not re-flagged.")

        terms, reused_key, key_decisions, retired_reals = [], False, {}, []
        if key_path and not key_path.is_file():
            # The re-run launcher auto-pins the folder's OWN output key
            # (pseudonym_key.xlsx) so a re-run reproduces the prior fakes. If it
            # was deleted — the natural way to force FRESH pseudonyms — that is
            # not an error: fall back to auto-discovery / a fresh mint so the
            # (detached, windowless) re-run still regenerates the .txt instead of
            # aborting silently before it writes anything. Any OTHER explicitly
            # named key that is missing stays fatal, but is now also logged so a
            # background run leaves a trace in pdf_linker.log.
            if key_path.name == "pseudonym_key.xlsx":
                word_only = _is_word_only_folder(folder)
                log.info(f"  {key_path.name} not found — starting over with fresh "
                         f"fakes. Looking for an input key to draw party names "
                         f"from (this folder's Order*.xlsx"
                         + ("" if word_only else " first, then Downloads") + ").")
                key_path = _pn_find_folder_key(folder, log)
                if key_path is None and not word_only:
                    key_path = _pn_find_downloads_key(log)
            else:
                log.error(f"Key spreadsheet not found: {key_path}")
                print(f"Key spreadsheet not found: {key_path}")
                sys.exit(1)
        if key_path:
            try:
                if _pn_key_looks_like_ours(key_path):
                    # A key this tool wrote: reuse its bindings so a follow-up
                    # single-file run reproduces the original run's fakes.
                    terms, key_decisions = _pn_load_key(key_path, registry, log)
                    # A flagged value the operator marked yes can be HALF-
                    # scrubbed ("Melissa Sable", where only "Penuela" was bound).
                    # Fake only its real remainder, so a stand-in this key
                    # already handed out never lands in the Real Value column and
                    # gets faked a second time.
                    extra_terms = _pn_drop_prior_fakes_from_terms(
                        extra_terms, _pn_prior_fake_words(terms), log)
                    # A LOCAL keep decision — typed into this key's Replacement
                    # column or this folder's LEAKS triage — retires the binding
                    # it corrects, so the key rewritten below holds only real
                    # values and the fakes actually applied. A keep merely
                    # INHERITED from the master sheet must not: another case's
                    # keep can never retire this case's own party binding.
                    # Re-running with a document the first batch missed is the
                    # normal way this tool is used (a draft comes back saying a
                    # filing is absent). The key holds only values that MATCHED,
                    # so a party named nowhere but in that missing document had
                    # no binding — and, since a reuse run never re-reads the
                    # party template, no term either. Supplement from the
                    # template; the memo-seeded registry keeps every fake the
                    # first run handed out exactly where it was.
                    #
                    # BEFORE retirement, so an operator KEEP still wins: a value
                    # marked `no` must not come back to life just because it is
                    # also a row in the party template.
                    terms += _pn_supplement_key_terms(
                        terms, folder, args.name_column, registry, log)
                    terms, retired_reals = _pn_retire_kept_key_terms(
                        terms, {**folder_decisions, **key_decisions},
                        registry, log)
                    terms += _pn_build_terms([], [], extra_terms, registry)
                    reused_key = True
                else:
                    names, casenos = _pn_terms_from_xlsx(
                        key_path, args.name_column, log)
                    terms = _pn_build_terms(names, casenos, extra_terms, registry)
            except RuntimeError as e:
                # e.g. openpyxl missing. Don't abort the whole run (linking and
                # PII-only pseudonymization still work) — warn and continue.
                _warn(f"Pseudonymize: could not read key {key_path.name}: {e}. "
                      f"Party names will NOT be pseudonymized in the .txt exports.")
        else:
            terms = _pn_build_terms([], [], extra_terms, registry)

        # KEEP / KEEP-PART edits typed into the pseudonym key's Replacement
        # column ('no' or a [bracketed] keep-spec): a key edit is authoritative
        # for THIS run, so it overrides an inherited decision for the same value.
        if key_decisions:
            key_frags = []
            for d in key_decisions.values():
                if d.get("fake_values"):
                    key_frags.extend(d["fake_values"])
            if key_frags:
                terms += _pn_build_terms([], [], key_frags, registry)
                log.info(f"  Pseudonym key: auto-faking {len(key_frags)} "
                         f"bracketed keep-spec fragment(s).")
            for vl, d in key_decisions.items():
                leak_decisions[vl] = d
                suppressed.add(vl)

        pseudonymizer = Pseudonymizer(terms, detectors, registry)
        pseudonymizer.suppressed = suppressed
        # KEEP protection: a kept word is left verbatim EXCEPT where it falls
        # inside a full party match (a real party this case fakes), which the
        # substitution releases so the party is still scrubbed. So a keep never
        # leaves a real party in the clear, whether learned here or globally.
        local_vls = set(key_decisions) | set(folder_decisions) | ours
        (pseudonymizer.keep_strict,
         pseudonymizer.keep_soft) = _pn_keep_values(leak_decisions)
        # A bracket THIS folder typed also beats the party override: its
        # non-bracketed remainder is registered as its own term, so the party
        # is scrubbed either way and the operator's split is honoured.
        pseudonymizer.keep_strict_local = _pn_keep_values(
            {vl: d for vl, d in leak_decisions.items() if vl in local_vls})[0]
        pseudonymizer._keep_decisions = {
            vl: d for vl, d in leak_decisions.items() if _pn_decision_is_keep(d)}
        pseudonymizer._keep_local = {vl for vl in pseudonymizer._keep_decisions
                                     if vl in local_vls}
        pseudonymizer._retired_key_rows = retired_reals
        if reused_key:
            # Loaded fakes are already-final strings; don't let a detector try
            # to re-fake one it meets in the new file.
            for r in pseudonymizer.records.values():
                pseudonymizer._own_fakes.add(str(r["fake"]).lower().rstrip(" .,;:"))
        log.info(f"Pseudonymizing .txt exports: {len(terms)} term(s), "
                 f"detectors={detectors or 'none'}")

    # Collect PDFs excluding _linked and _temp files.
    pdfs = _pdfs_in_folder(folder)
    sizes = {p: p.stat().st_size for p in pdfs}
    # Weight each file by processing cost — OCR pages dominate runtime, so this
    # is far truer than bytes (a big native-text brief is cheap; a small scanned
    # exhibit is not). Cheap metadata pass, reused for the ETA below. Process the
    # HEAVIEST first: the big scanned evidence files carry the vast majority of
    # the work, so front-loading them runs the slow OCR while you're watching and
    # finishes on a tail of quick files — instead of grinding through many small
    # files and only reaching a monster at the very end (where a stall would cost
    # a whole night). Ties break on size, so order is deterministic.
    weights = {}
    for p in pdfs:
        w = _pdf_work_weight(p)
        weights[p] = w if w is not None else max(1.0, sizes[p] / 50000.0)
    pdfs = sorted(pdfs, key=lambda p: (-weights[p], -sizes[p], p.name))
    log.info(f"Found {len(pdfs)} PDF(s) to process (heaviest first)")

    # Bulk Word→text conversion. Gated on the folder holding NO PDFs, so the
    # usual PDF workflow (even one with a few Word docs mixed in) is untouched:
    # only an all-Word folder triggers the pass. Word docs are converted to the
    # same scrubbed .txt exports the PDFs get, but never hyperlinked. Extracted
    # UP FRONT so the names in them feed the folder-wide pre-scan below; the .txt
    # is written after the PDFs so leak findings land in the same report and gate.
    word_texts = []   # [(path, extracted_text)] for docs to convert this run
    if args.extract_text:
        word_docs = _word_docs_in_folder(folder)
        legacy_doc = [p for p in folder.glob("*.doc")
                      if not p.name.startswith("~$")]
        if word_docs and not pdfs:
            log.info(f"Found {len(word_docs)} Word doc(s) and no PDFs; "
                     f"bulk-converting to text (not hyperlinked)")
            # Say so plainly when the run has NO curated party list. Silence
            # here reads as "fully scrubbed", when in fact only the detectors
            # and the pre-scan harvest are working — a party they don't infer
            # goes out in the clear. The Downloads guess is deliberately not
            # used for this folder (it belongs to whatever case was downloaded
            # last), so the fix is to name the list explicitly.
            if pseudonymizer is not None and not pseudonymizer.terms:
                _warn("Pseudonymize: this all-Word folder has no party list "
                      "(no pseudonym_key.xlsx and no Order*.xlsx in the "
                      "folder). Names will only be caught by the detectors and "
                      "the document pre-scan — a party neither infers will NOT "
                      "be scrubbed. Drop the case's Order*.xlsx in the folder, "
                      "or pass --key/--term, and re-run. A template sitting in "
                      "Downloads is NOT used here: it belongs to whichever case "
                      "was downloaded last, not to this one.")
            for wd in word_docs:
                try:
                    word_texts.append((wd, _extract_docx_text(wd)))
                except Exception as e:
                    log.warning(f"  Could not read Word doc {wd.name} "
                                f"(skipped): {e}")
        elif word_docs:
            log.info(f"Found {len(word_docs)} Word doc(s) beside {len(pdfs)} "
                     f"PDF(s); leaving them alone (Word conversion runs only "
                     f"when the folder has no PDFs)")
        if legacy_doc and not pdfs:
            log.warning(f"  {len(legacy_doc)} legacy .doc file(s) found "
                        f"(e.g. {legacy_doc[0].name}); the old binary .doc "
                        f"format can't be converted — re-save as .docx.")

    # Estimated-finish marker: worth it only for a multi-file batch (a single
    # file has no throughput to project from until it is already done).
    show_eta = len(pdfs) >= 2
    total_work = sum(weights.values()) or 1.0
    if show_eta:
        _write_eta_marker(folder, "(estimating...)")
        # Seed an immediate projected finish from the last run's throughput, so a
        # large re-run shows a time right away instead of "(estimating...)" until
        # the first file lands. Refined live below.
        _seed_rate = _load_eta_rate()
        if _seed_rate:
            finish = datetime.datetime.now() + datetime.timedelta(
                seconds=total_work / _seed_rate)
            _write_eta_marker(folder, f"~{_fmt_clock(finish)} (estimating)")

    # Write the one-click re-run launcher UP FRONT, so an interrupted or hung run
    # still leaves a clickable re-run instead of nothing (it is refreshed at the
    # end too). --key is pinned only if the folder key already exists; on a first
    # run it does not yet, so the re-run re-discovers the key and the
    # deterministic fakes reproduce the same pseudonyms.
    #
    # A WORD-ONLY folder is a real batch, not a no-op: it produces the same
    # scrubbed .txt exports, the same pseudonym key and the same LEAKS
    # worksheet, so it needs the same two launchers. Gating them on `pdfs`
    # alone left an all-Word folder with leaks to triage and nothing to
    # double-click — `--fix-leaks` reads the .txt exports and never reopens a
    # source document, so it works on Word-derived exports unchanged.
    if pdfs or word_texts:
        _want_key = (pseudonymizer is not None
                     and (folder / "pseudonym_key.xlsx").is_file())
        _write_rerun_launcher(folder, args.provider, _want_key, log)
        # Only if a worksheet from a prior run is already waiting — otherwise
        # an interrupted run would leave a launcher for a triage that does not
        # exist. The end-of-run block settles it either way.
        if _want_key and _pn_leak_xlsx_path(folder).is_file():
            _write_fix_launcher(folder, log)

    if pseudonymizer is not None and (pdfs or word_texts):
        corpus = _pn_prescan_folder(
            pdfs, pseudonymizer, log,
            extra_texts=[(p.stem, t) for p, t in word_texts])
        partial = (args.partial_names if args.partial_names is not None
                   else _config_bool(cfg, "partial_names", False))
        if partial:
            _pn_register_prefix_terms(pseudonymizer, corpus, log)

    success = 0
    failed = 0
    done_work = 0.0
    proc_start = time.monotonic()
    for i, pdf in enumerate(pdfs, 1):
        try:
            if process_pdf(pdf, log, provider=args.provider,
                           extract_text=args.extract_text,
                           pseudonymizer=pseudonymizer,
                           text_subdir=text_subdir,
                           original_subdir=original_subdir,
                           relink=args.relink):
                success += 1
            else:
                failed += 1
        except Exception as e:
            log.error(f"Unhandled error processing {pdf.name}: {e}")
            log.error(traceback.format_exc())
            failed += 1
        # Work-weighted estimate (OCR pages counted heavily): the cost of the
        # slow scanned files is priced into total_work from the start, so the
        # projected finish holds steady instead of swinging when one comes up.
        # Cumulative work-units/sec self-calibrates. The marker names the
        # projected finish CLOCK time.
        if show_eta:
            done_work += weights.get(pdf, 0.0)
            elapsed = time.monotonic() - proc_start
            rate = done_work / elapsed if elapsed > 0 else 0
            if i < len(pdfs) and rate > 0:
                finish = datetime.datetime.now() + datetime.timedelta(
                    seconds=(total_work - done_work) / rate)
                clock = _fmt_clock(finish)
                _write_eta_marker(folder, f"~{clock} ({i} of {len(pdfs)})")
                log.info(f"Progress: {i}/{len(pdfs)} files "
                         f"({done_work * 100 / total_work:.0f}% of work); "
                         f"estimated finish ~{clock}")
    if show_eta:
        # Remember this run's throughput so the NEXT run can show a finish time
        # immediately instead of estimating from scratch.
        _elapsed = time.monotonic() - proc_start
        if _elapsed > 0 and done_work > 0:
            _save_eta_rate(done_work / _elapsed)
        _write_done_marker(folder)   # clean finish stamps 'DONE <finish time>'

    # Bulk Word→text conversion (gated above). Done after the PDFs so every leak
    # finding lands in the same worksheet and the same quarantine gate below.
    if word_texts:
        n = _convert_word_docs(word_texts, log, pseudonymizer,
                               text_subdir, original_subdir)
        log.info(f"Converted {n} Word doc(s) to text")
        # Stamp the folder the same way a PDF batch is stamped, so an all-Word
        # run shows at a glance that it FINISHED (a folder with neither marker
        # is indistinguishable from one the tool never touched). No live ETA:
        # the projection is built from per-page OCR weights, and a Word
        # conversion has no OCR to project from — it is done in seconds.
        _write_done_marker(folder)

    # One key file for the whole folder maps every real value to its fake.
    if pseudonymizer is not None:
        # Human-triage worksheet of every located potential leak — written even
        # if the leak gate quarantines below, so the reviewer always gets it.
        # Prior yes/no decisions are carried through (yes was scrubbed above).
        _pn_write_leak_report(folder, pseudonymizer.leak_report, log,
                              leak_decisions, cfg=cfg,
                              bound=[r["real"] for r in
                                     pseudonymizer.records.values()])
        # Record this run's KEEP decisions into the single cross-folder master
        # KEEP sheet: every LOCAL keep (made in this folder), plus any GLOBAL
        # keep that actually protected text here (a real hit) — so Times Seen /
        # Cases / dates accumulate and the durable decision survives cleanup.
        _record = {}
        for vl, d in getattr(pseudonymizer, "_keep_decisions", {}).items():
            hit = any(p.lower() in pseudonymizer.kept_hits
                      for p in _pn_decision_keep_parts(d))
            if vl in pseudonymizer._keep_local or hit:
                _record[vl] = d
        if _record:
            _pn_update_master_keep(cfg, _record, folder.name,
                                   datetime.date.today().isoformat(), log)
        # Two addresses on one street with adjacent numbers were faked to two
        # unrelated streets — the failure that moved an ADU off its parcel.
        for w in _pn_address_adjacency(pseudonymizer.records.values()):
            _warn(f"Pseudonymize REVIEW: {w}. Re-run with the range as one "
                  f"--term, or verify the fake addresses by hand.")
        if pseudonymizer.review:
            shown = "; ".join(f"{c}: {s}" for c, s in pseudonymizer.review[:12])
            _warn(f"Pseudonymize REVIEW: identifier shape(s) present in the "
                  f".txt exports that no key can catch — verify before sharing "
                  f"({shown}).")

        key_out = args.key_out or (folder / "pseudonym_key.xlsx")
        try:
            pseudonymizer.write_key(key_out, log)
        except Exception as e:
            log.warning(f"Could not write pseudonym key (non-fatal): {e}")

        # Report (do NOT delete): if pseudonymization ran on some .txt but the
        # spreadsheet key produced NO primary matches (full party names /
        # entities / case numbers), those exports may still name real parties —
        # the key is the wrong/stale sheet, or none was found. The .txt files
        # are left in place; warn loudly (console + log) so they're reviewed
        # before sharing.
        if pseudonymizer.texts_applied and pseudonymizer.spreadsheet_hits() == 0:
            if not pseudonymizer.has_spreadsheet_terms():
                if key_path:
                    reason = f"the key {key_path.name} held no party/case values"
                elif _is_word_only_folder(folder):
                    # Downloads is deliberately not consulted here, so naming it
                    # would send the operator to fix the wrong thing.
                    reason = ("this all-Word folder had no party list "
                              "(put the case's Order*.xlsx in the folder)")
                else:
                    reason = "no Order*.xlsx was found in Downloads"
            else:
                reason = (f"none of the key values"
                          f"{f' from {key_path.name}' if key_path else ''} "
                          "matched any document (wrong/stale sheet?)")
            _warn(f"!! Pseudonymize: {len(pseudonymizer.written)} .txt export(s) "
                  f"may name real parties because {reason}. Review before "
                  "sharing; fix the key (correct E-Court Order*.xlsx or --key) "
                  "and re-run.")

    log.info(f"Done: {success} succeeded, {failed} failed")

    # One-click re-run launcher (written before the leak gate can exit, so it's
    # there to apply Fix? decisions after a quarantine). --key points at the
    # folder's own key when it exists, so a re-run reproduces the same fakes.
    if pdfs or word_texts:
        want_key = (pseudonymizer is not None
                    and (folder / "pseudonym_key.xlsx").is_file())
        _write_rerun_launcher(folder, args.provider, want_key, log)
        # Companion launcher: apply the worksheet Fix? decisions to the exports
        # directly (no source document reopened) — the fast path after triaging
        # leaks, and the only one that exists for an all-Word folder. It exists
        # to apply THAT worksheet, so it belongs beside it and nowhere else:
        # written when there is something to triage, removed when there is not.
        # This runs after the report is written, so it sees THIS run's verdict
        # (the up-front copy is written from the previous run's state).
        if want_key and _pn_leak_xlsx_path(folder).is_file():
            _write_fix_launcher(folder, log)
        else:
            for _p in _pn_fix_launcher_paths(folder):
                try:
                    if _p.exists():
                        _p.unlink()
                        log.info(f"  Nothing left to triage — removed "
                                 f"{_p.name}.")
                except OSError:
                    pass

    # Leak gate, TIERED (config `leak_gate`): the pseudonymization is a
    # precaution against casual recognition of a public filing, so only a
    # survivor that defeats that purpose outright — a full party/entity/
    # attorney name or defined short form — quarantines the exports (renamed
    # to *.LEAK, preserved for review, exit non-zero). A lesser survivor (a
    # bare token, an identifier, a welded splinter) is warned about and marked
    # Status=leaked in the key, but the exports are delivered: blocking five
    # good files over a spot a casual reader would never notice is the wrong
    # trade. `strict` restores quarantine-on-anything; `off` never quarantines.
    if pseudonymizer is not None and pseudonymizer.has_leaks():
        gate = (cfg.get("leak_gate", "").strip().lower() or "primary")
        if gate not in ("strict", "primary", "off"):
            gate = "primary"
        gating = (pseudonymizer.leaked if gate == "strict"
                  else pseudonymizer.primary_leaks() if gate == "primary"
                  else set())
        # A value the reviewer marked "no fix" in the worksheet never blocks
        # delivery — that is exactly what the mark means.
        gating = {v for v in gating if str(v).lower() not in pseudonymizer.suppressed}
        if gating:
            # Quarantine ONLY the files that actually carry a gating leak; the
            # rest of the batch is clean and gets delivered.
            offenders = pseudonymizer.files_to_quarantine(gating)
            quarantined = []
            for txt in offenders:
                try:
                    if txt.exists():
                        dest = txt.with_suffix(txt.suffix + ".LEAK")
                        txt.replace(dest)
                        quarantined.append(dest)
                except OSError as e:
                    log.error(f"  Could not quarantine leaked export "
                              f"{txt.name}: {e}")
            delivered = len(pseudonymizer.written) - len(quarantined)
            shown = ", ".join(sorted(gating)[:8])
            _warn(f"!! Pseudonymize FAILED: party name(s) survived in the "
                  f"exports ({shown}) — the case is recognizable on sight. "
                  f"{len(quarantined)} .txt export(s) with a leak quarantined to "
                  f"*.LEAK and NOT delivered ({delivered} clean export(s) "
                  f"delivered). Add the survivor(s) with --term and re-run.")
            sys.exit(2)
        shown = ", ".join(sorted(pseudonymizer.leaked)[:8])
        _warn(f"Pseudonymize: {len(pseudonymizer.leaked)} lesser value(s) "
              f"survived ({shown}) — no party name is recognizable, so the "
              f"exports are delivered. Add any that matter with --term and "
              f"re-run.")


if __name__ == "__main__":
    main()
